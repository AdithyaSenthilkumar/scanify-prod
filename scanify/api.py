import frappe
import json
from frappe import _
from frappe.utils import flt, nowdate, add_months, get_first_day
from scanify.permissions import require_process, require
import requests
import os
import base64
import mimetypes
from frappe import _
from google import genai as genai_sdk
from google.genai import types as genai_types
from PIL import Image
import io
import zipfile
import tempfile
import time
from frappe.utils import flt, cstr
from frappe.utils.background_jobs import enqueue
import re
import math
from difflib import SequenceMatcher


# ─────────────────────────────────────────────────────────────────────────────
# Report scoping — confine non-admins to their mapped division(s) and region(s).
#
# Role behaviour (via scanify.permissions):
#   • Admin                 → get_allowed_region_codes() returns None = NO restriction
#                             (sees every region; an explicitly picked region still filters).
#   • HO / Regional / RF    → confined to the user's allowed_regions (∩ any picked region).
# These helpers are the single place reports apply that rule, mirroring the pattern
# already used by get_scheme_list_portal / get_pending_scheme_emails.
# ─────────────────────────────────────────────────────────────────────────────

def _effective_division(division=None, user=None):
    """The division a report must query, forced within the caller's allowed set.
    Admin: the requested/active division as-is. Non-admin: if the requested division
    isn't one they're mapped to, fall back to their first allowed division so a direct
    API call can never read another division's data."""
    from scanify.permissions import is_portal_admin, get_allowed_divisions
    if not division:
        division = get_user_division()
    if is_portal_admin(user):
        return division
    allowed = get_allowed_divisions(user)
    if allowed and division not in allowed:
        return allowed[0]
    return division


def _region_match_values(codes, include_names=False):
    """The values a region column may hold for the given region codes. Most tables store
    the Region Master id/code (R0467), but Primary Sales Data stores the region NAME
    ("Nagpur") — see _pri_master_name. Passing include_names=True matches BOTH, so the
    same scope works on either storage (codes and names never collide)."""
    vals = list(codes)
    if include_names and codes:
        vals += [n for n in frappe.get_all(
            "Region Master", filters={"name": ["in", list(codes)]}, pluck="region_name") if n]
    return vals


def _scope_region_sql(conditions, params, column, division, requested_region=None,
                      key="rgnscope", user=None, include_names=False):
    """Append the caller's region restriction to a (conditions, params) SQL pair.

    - Admin: filter by `requested_region` when one was picked, else add nothing.
    - Non-admin: restrict `column` to (allowed_regions ∩ requested_region); when that
      intersection is empty (user allowed no region here) append an always-false
      condition so the query naturally returns zero rows — the report's own aggregation
      then produces its normal empty shape, no special-casing needed at each call site.

    `column` is a qualified SQL column (e.g. "ss.region"); `key` must be unique within
    the query's params. Set include_names=True for region columns that store the region
    NAME rather than the code (Primary Sales Data). Assumes `conditions` is AND-combined."""
    from scanify.permissions import get_allowed_region_codes, clamp_region_codes
    allowed = get_allowed_region_codes(user, division)
    if allowed is None:  # admin — unrestricted
        if requested_region:
            conditions.append(f"{column} IN %({key})s")
            params[key] = tuple(_region_match_values([requested_region], include_names))
        return
    codes = clamp_region_codes(requested_region, division=division, user=user)
    if not codes:
        conditions.append("1=0")  # locked out of every region → no rows
        return
    conditions.append(f"{column} IN %({key})s")
    params[key] = tuple(_region_match_values(codes, include_names))


def _scope_region_sql_pos(conditions, values, column, division, requested_region=None,
                          user=None, include_names=False):
    """Positional-parameter (%s) variant of _scope_region_sql, for queries that collect a
    `values` list instead of a params dict. Call it exactly where the original region
    condition was appended so the value ordering still lines up."""
    from scanify.permissions import get_allowed_region_codes, clamp_region_codes
    allowed = get_allowed_region_codes(user, division)
    if allowed is None:  # admin — unrestricted
        if requested_region:
            vals = _region_match_values([requested_region], include_names)
            conditions.append(f"{column} IN ({', '.join(['%s'] * len(vals))})")
            values.extend(vals)
        return
    codes = clamp_region_codes(requested_region, division=division, user=user)
    if not codes:
        conditions.append("1=0")  # locked out of every region → no rows
        return
    vals = _region_match_values(codes, include_names)
    conditions.append(f"{column} IN ({', '.join(['%s'] * len(vals))})")
    values.extend(vals)


def _allowed_region_codes_or_all(division=None, user=None):
    """Region codes the caller may see in a division, or None for admin (= all).
    Used to scope filter-option dropdowns so non-admins can't even pick a foreign region."""
    from scanify.permissions import get_allowed_region_codes
    return get_allowed_region_codes(user, division)


def _clamp_hqs_to_allowed_regions(hq_list, division=None, user=None):
    """Restrict a resolved HQ list to those inside the caller's mapped regions (admin
    gets it back unchanged). Lets entity-scoped reports keep an 'Organization'/'Zone'
    scope from spanning regions the user may not see."""
    from scanify.permissions import get_allowed_region_codes
    allowed = get_allowed_region_codes(user, division)
    if allowed is None:
        return hq_list
    if not allowed or not hq_list:
        return []
    keep = set(frappe.get_all(
        "HQ Master", filters={"name": ["in", list(hq_list)], "region": ["in", allowed]},
        pluck="name"))
    return [h for h in hq_list if h in keep]


def _bulk_job_region_scope(doc):
    """Region scope for a bulk OCR job's stockist matching: the job's explicitly chosen
    region when set, otherwise the creating user's mapped regions (None = unrestricted,
    i.e. admin). Stops a non-admin's ZIP from matching stockists in regions they are not
    mapped to, even when they left the region toggle off. Resolved from doc.owner because
    the extraction runs as a background job."""
    from scanify.permissions import get_allowed_region_codes
    if doc.region:
        return doc.region
    return get_allowed_region_codes(doc.owner, doc.division)


def _stockist_region_allowed(stockist_pk, division=None, user=None):
    """Whether the caller may see a single stockist's data: admin always; otherwise the
    stockist's region must be in the user's allowed_regions. Used by single-stockist
    reports that have no region column to clamp on."""
    from scanify.permissions import get_allowed_region_codes
    allowed = get_allowed_region_codes(user, division)
    if allowed is None:
        return True
    if not allowed:
        return False
    return frappe.db.get_value("Stockist Master", stockist_pk, "region") in allowed

# Retired / superseded Gemini model IDs mapped to their current replacements.
# A value saved in Scanify Settings before a model was shut down would otherwise
# cause "model not found" errors (e.g. gemini-3-pro-preview was shut down
# 2026-03-09, gemini-3-flash-preview is being deprecated). This keeps extraction
# working and transparently upgrades stale selections.
GEMINI_MODEL_ALIASES = {
    "gemini-3-pro-preview": "gemini-3.1-pro-preview",
    "gemini-3-flash-preview": "gemini-3.5-flash",
    "gemini-3-flash-lite-preview": "gemini-3.1-flash-lite",
}


def resolve_gemini_model(model_name):
    """Map a possibly-retired model ID to a currently-supported one."""
    model_name = (model_name or "").strip() or "gemini-2.5-flash"
    return GEMINI_MODEL_ALIASES.get(model_name, model_name)


def get_gemini_settings():
    """
    Fetch Gemini API settings from Scanify Settings DocType
    Returns: tuple (api_key, model_name, is_enabled)
    """
    try:
        # Try to get settings - single doctype
        settings_name = frappe.db.get_value("Scanify Settings", {"company_name": "Stedman Pharmaceuticals"}, "name")
        
        if not settings_name:
            # Fallback: try getting the single record directly
            settings_name = "Scanify Settings"
        
        # Fetch settings data
        settings_data = frappe.db.get_value(
            "Scanify Settings",
            settings_name,
            ["enable_gemini", "gemini_model_name"],
            as_dict=True
        )
        
        if not settings_data:
            frappe.throw(_("Scanify Settings not found. Please create it first."))
        
        # Check if Gemini is enabled
        if not settings_data.get("enable_gemini"):
            frappe.throw(_("Gemini AI extraction is not enabled in Scanify Settings"))
        
        # Get API key
        api_key = frappe.utils.password.get_decrypted_password(
            "Scanify Settings",
            settings_name,
            "gemini_api_key"
        )
        
        if not api_key:
            frappe.throw(_("Gemini API key not configured in Scanify Settings"))
        
        # Get model name with fallback, upgrading any retired model IDs
        model_name = resolve_gemini_model(settings_data.get("gemini_model_name"))

        frappe.logger().info(f"✅ Gemini settings loaded: Model={model_name}")
        
        return api_key, model_name, True
        
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Gemini Settings Error")
        frappe.throw(_("Error fetching Gemini settings: {0}").format(str(e)))

        
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Gemini Settings Error")
        frappe.throw(_("Error fetching Gemini settings: {0}").format(str(e)))

def build_product_catalog_for_prompt(division=None):
    """
    Build comprehensive product catalog with all matching hints for Gemini
    Returns formatted string for prompt inclusion.

    When a division is given the catalog is scoped to that division (plus legacy
    "Both" rows). This matters now that the same Product Code can exist in
    DIFFERENT divisions: Gemini must only ever see this division's codes.
    """
    filters = {"status": "Active"}
    if division:
        filters["division"] = ["in", [division, "Both"]]
    products = frappe.get_all(
        "Product Master",
        filters=filters,
        fields=[
            "name",
            "product_code",
            "product_name",
            "pack",
            "pack_conversion",
            "division",
            "product_group",
            "pts",
            "ptr",
            "mrp"
        ],
        order_by="division, product_group, sequence asc, product_name"
    )
    
    if not products:
        frappe.throw("No active products found in Product Master")
    
    # Group by division for better organization
    catalog_text = "\n=== PRODUCT MASTER CATALOG ===\n"
    catalog_text += f"Total Products: {len(products)}\n\n"
    
    current_division = None
    current_group = None
    
    for p in products:
        # Division header
        if p.get("division") != current_division:
            current_division = p.get("division")
            catalog_text += f"\n--- {current_division} Division ---\n"
        
        # Group header
        if p.get("product_group") != current_group:
            current_group = p.get("product_group")
            catalog_text += f"\n  [{current_group} Group]\n"
        
        # Product entry with all matching hints
        catalog_text += f"  • Code: {p['product_code']}\n"
        catalog_text += f"    Name: {p['product_name']}\n"
        catalog_text += f"    Pack: {p.get('pack', 'N/A')}\n"
        catalog_text += f"    Conversion: {p.get('pack_conversion', 'N/A')}\n"
        catalog_text += f"    PTS: {p.get('pts', 0)}\n\n"
    
    return catalog_text, products

@frappe.whitelist()
@require_process("secondary")
def extract_stockist_statement(doc_name, file_url):
    """
    Extract stockist statement data using Gemini AI.
    Runs the heavy Gemini call in a background thread to avoid nginx 504 timeouts.
    Frontend should poll check_extraction_status() for completion.
    """
    import threading

    try:
        doc = frappe.get_doc("Stockist Statement", doc_name)

        if doc.extracted_data_status == "In Progress":
            return {"success": True, "message": "Extraction already in progress", "async": True}

        if not file_url:
            return {"success": False, "message": "No file uploaded"}

        # Validate file exists before spawning thread
        from frappe.utils.file_manager import get_file_path
        file_path = get_file_path(file_url)
        if not os.path.exists(file_path):
            return {"success": False, "message": f"File not found: {file_url}"}

        # Mark as In Progress immediately
        doc.extracted_data_status = "In Progress"
        doc.extraction_notes = ""
        doc.save()
        frappe.db.commit()

        # Capture context for background thread
        site = frappe.local.site

        def _run_extraction():
            try:
                frappe.init(site=site)
                frappe.connect()
                _do_extract(doc_name, file_url)
            except Exception as thread_err:
                try:
                    frappe.init(site=site)
                    frappe.connect()
                    _doc = frappe.get_doc("Stockist Statement", doc_name)
                    _doc.extracted_data_status = "Failed"
                    _doc.extraction_notes = f"Extraction failed: {str(thread_err)}"
                    _doc.save()
                    frappe.db.commit()
                except Exception:
                    pass
                frappe.log_error(frappe.get_traceback(), "Gemini Extraction Thread Error")
            finally:
                try:
                    frappe.destroy()
                except Exception:
                    pass

        t = threading.Thread(target=_run_extraction, daemon=True, name=f"extract_{doc_name}")
        t.start()

        return {"success": True, "message": "Extraction started", "async": True}

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Gemini Extraction Error")
        return {"success": False, "message": f"Extraction failed: {str(e)}"}


def _build_correction_map(stockist_code):
    """Build {RAW_PRODUCT_NAME: product_code} map from Stockist Product Correction for a stockist.

    `mapped_product_code` is a Link (the Product Master id/PK). Gemini only ever
    sees business Product Codes from the catalog, so translate each PK to its
    editable product_code for the prompt hints and response comparison. Legacy
    rows where the PK equals the code fall back to the stored value."""
    corrections = frappe.get_all(
        "Stockist Product Correction",
        filters={"stockist_code": stockist_code, "status": "Active"},
        fields=["raw_product_name", "mapped_product_code"],
    )
    if not corrections:
        return {}

    pks = {c["mapped_product_code"] for c in corrections if c.get("mapped_product_code")}
    pk_to_code = {}
    if pks:
        for p in frappe.get_all(
            "Product Master",
            filters={"name": ["in", list(pks)]},
            fields=["name", "product_code"],
        ):
            pk_to_code[p["name"]] = p["product_code"] or p["name"]

    return {
        c["raw_product_name"].strip().upper(): pk_to_code.get(c["mapped_product_code"], c["mapped_product_code"])
        for c in corrections
    }


def _build_correction_prompt(stockist_code):
    """Build stockist-specific correction hints for Gemini without overriding its final decision."""
    correction_map = _build_correction_map(stockist_code)
    if not correction_map:
        return ""

    lines = ["\n=== STOCKIST-SPECIFIC PRODUCT CORRECTION HINTS ==="]
    lines.append(
        "These are past QC-approved mappings for this stockist. Use them as strong hints only when the visible"
        " product text and pack still agree with the mapped product."
    )
    for raw_name, product_code in sorted(correction_map.items()):
        lines.append(f"- {raw_name} -> {product_code}")

    return "\n".join(lines) + "\n"


def _get_first_present_value(data, *keys):
    """Return the first non-empty value for the given keys."""
    for key in keys:
        if key in data and data.get(key) not in (None, ""):
            return data.get(key)
    return 0


def _parse_numeric_value(value):
    """Parse extracted numeric text while preserving negatives and parenthesized values."""
    text = cstr(value).strip()
    if not text:
        return 0

    if text.startswith("(") and text.endswith(")"):
        text = f"-{text[1:-1]}"

    text = text.replace(",", "")
    return flt(text)


def _thinking_config(model_name, gemini3_level="low", gemini25_budget=0):
    """Build a ThinkingConfig with the right knob for the model family.

    Gemini 3 models use thinking_level ("low"/"high"); Gemini 2.5 models use
    thinking_budget (token cap, 0 disables). Passing the wrong knob to a family
    errors or is silently ignored, so route by model id. Used by the lighter
    helper calls (sales-total, filename matching); the main item-extraction path
    uses _build_gemini_generation_config below.
    """
    if "gemini-3" in cstr(model_name).strip().lower():
        return genai_types.ThinkingConfig(thinking_level=gemini3_level)
    return genai_types.ThinkingConfig(thinking_budget=gemini25_budget)


def _build_gemini_generation_config(model_name):
    """Tune Gemini thinking per model family for extraction requests."""
    model_key = cstr(model_name).strip().lower()
    thinking_config = None

    if "gemini-3" in model_key:
        # Gemini 3 family uses thinking_level. Keep Flash-Lite (the cost tier)
        # on low reasoning so we don't pay for heavy thinking tokens and lose the
        # cost advantage; Pro/Flash get high reasoning for accuracy.
        level = "low" if "flash-lite" in model_key else "high"
        thinking_config = genai_types.ThinkingConfig(thinking_level=level)
    else:
        # Gemini 2.5 models use thinkingBudget. Bump it modestly above the previous fixed 1024 cap.
        thinking_config = genai_types.ThinkingConfig(thinking_budget=2048)

    return genai_types.GenerateContentConfig(thinking_config=thinking_config)


_TOTAL_ROW_EXACT = frozenset({
    "TOTAL", "TOTAL QUANTITY", "TOTAL QTY", "TOTAL VALUE", "TOTAL AMOUNT",
    "GRAND TOTAL", "SUB TOTAL", "SUBTOTAL", "NET TOTAL",
})


def _is_total_row(raw_name):
    """Return True if the raw product name is a summary/total line that must be excluded."""
    if raw_name in _TOTAL_ROW_EXACT:
        return True
    if raw_name.startswith("TOTAL ") or raw_name.startswith("GRAND TOTAL"):
        return True
    return False


def _normalize_row_type(row_type, raw_product_name):
    """Normalize Gemini row type output and backfill known special rows from the raw label."""
    normalized = cstr(row_type).strip().lower().replace(" ", "_")
    if normalized in {"product", "others", "branch_transfer"}:
        return normalized

    raw_name = cstr(raw_product_name).strip().upper()

    # Summary/total rows must be excluded before any other classification
    if _is_total_row(raw_name):
        return "total_row"

    if raw_name == "OTHERS" or raw_name.startswith("OTHERS "):
        return "others"
    if raw_name == "BRANCH TRANSFER" or raw_name.startswith("BRANCH TRANSFER"):
        return "branch_transfer"

    return "product"


def _build_division_product_map(statement_division, products_list=None):
    """Build {product_code: product id (PK)} for a statement division.

    Keys are the editable business Product Codes Gemini returns; values are the
    Product Master ids that statement item Link fields must store. Scoped to the
    statement's division so a code reused across divisions resolves to THIS
    division's product. Membership checks use the keys (same semantics as the
    old code set)."""
    division_product_map = {}
    if not statement_division:
        return division_product_map

    source_products = products_list
    if source_products is None:
        source_products = frappe.get_all(
            "Product Master",
            filters={"status": "Active"},
            fields=["name", "product_code", "division"],
        )

    for product in source_products:
        if product.get("division") in (statement_division, "Both"):
            # Legacy products (and products_list rows without `name`) have PK == code.
            division_product_map[product["product_code"]] = product.get("name") or product["product_code"]

    return division_product_map


def _build_region_excluded_codes(statement_region):
    """Build the set of product codes that are EXCLUDED for the statement's region.

    Products carry an `excluded_regions` table (child doctype Product Excluded Region,
    one Region Master link per row). A product whose excluded list contains the
    statement's region code must be skipped during extraction for that statement.
    Empty/no region → empty set (fail-open: nothing excluded)."""
    if not statement_region:
        return set()

    rows = frappe.get_all(
        "Product Excluded Region",
        filters={"region": statement_region, "parenttype": "Product Master"},
        fields=["parent"],
    )
    # `parent` is the Product Master id (PK) — NOT the editable product_code
    # (they only coincide for legacy products). Callers must compare against the
    # resolved product id, not against Gemini's business code.
    return {r["parent"] for r in rows if r.get("parent")}


def _build_statement_item_row(item_data, statement_division=None, division_product_map=None,
                              region_excluded_codes=None):
    """Normalize one Gemini row into a Stockist Statement child row plus bookkeeping metadata.

    Gemini returns the editable business Product Code; the stored Link value must
    be the Product Master id (PK), resolved through division_product_map so a
    code reused across divisions maps to THIS division's product."""
    raw_name = cstr(item_data.get("raw_product_name")).strip().upper()
    row_type = _normalize_row_type(item_data.get("row_type"), raw_name)
    mapping_basis = cstr(item_data.get("mapping_basis")).strip().lower()

    # Exclude total/summary rows — they pollute closing stock calculations
    if row_type == "total_row":
        return None, {"unmapped": False, "auto_mapped": False, "special_row": False, "skipped_total_row": True}

    is_unmapped = bool(item_data.get("unmapped", False))
    product_code = cstr(item_data.get("product_code")).strip() or None
    if row_type != "product":
        product_code = None
        is_unmapped = False

    if product_code and statement_division and division_product_map and product_code not in division_product_map:
        return None, {
            "unmapped": False,
            "auto_mapped": False,
            "special_row": False,
            "skipped_division": True,
        }

    # Resolve the business code to the Product Master id (PK). Without a division
    # map (no division on the statement) fall back to the raw code, which stays
    # valid for legacy products where PK == code.
    product_pk = None
    if product_code:
        product_pk = (division_product_map or {}).get(product_code) or product_code

    # Region exclusion: product is flagged as not-sold in this statement's region.
    # The excluded set holds Product Master ids, so compare the resolved id.
    if product_pk and region_excluded_codes and product_pk in region_excluded_codes:
        return None, {
            "unmapped": False,
            "auto_mapped": False,
            "special_row": False,
            "skipped_region": True,
        }

    mapping_status = "matched"
    if row_type == "product":
        if is_unmapped or not product_code:
            mapping_status = "unmapped"
        elif mapping_basis == "stockist_correction_hint":
            mapping_status = "auto_mapped"

    row_confidence = min(max(_parse_numeric_value(item_data.get("confidence")), 0), 100)
    operational_sales_qty = _parse_numeric_value(item_data.get("operational_sales_qty"))
    product_sales_qty = _parse_numeric_value(item_data.get("sales_qty"))

    if row_type != "product":
        if not operational_sales_qty:
            operational_sales_qty = product_sales_qty
        product_sales_qty = 0

    row = {
        "raw_product_name": raw_name,
        "row_type": row_type,
        "mapping_status": mapping_status,
        "opening_qty": _parse_numeric_value(item_data.get("opening_qty")),
        "purchase_qty": _parse_numeric_value(item_data.get("purchase_qty")),
        "sales_qty": product_sales_qty,
        "operational_sales_qty": operational_sales_qty,
        "free_qty": _parse_numeric_value(item_data.get("free_qty")),
        "return_qty": _parse_numeric_value(item_data.get("return_qty")),
        "misc_out_qty": _parse_numeric_value(item_data.get("misc_out_qty")),
        "closing_qty": _parse_numeric_value(item_data.get("closing_qty")),
        "closing_value": _parse_numeric_value(item_data.get("closing_value")),
        "row_confidence": round(row_confidence, 1),
        "math_check": "N/A",
    }
    if product_pk:
        row["product_code"] = product_pk

    return row, {
        "unmapped": mapping_status == "unmapped",
        "auto_mapped": mapping_status == "auto_mapped",
        "special_row": row_type != "product",
        "skipped_division": False,
    }


def _calculate_confidence_score(statement_rows):
    """Average row-level extraction confidence without mixing in QC penalties."""
    confidence_values = [flt(row.get("row_confidence")) for row in statement_rows]
    return round(sum(confidence_values) / len(confidence_values), 1) if confidence_values else 0


def _build_extraction_notes(items_added, confidence_score, unmapped_count=0, auto_mapped_count=0, special_row_count=0,
    skipped_division_count=0, statement_division=None, skipped_region_count=0):
    """Build consistent extraction notes for single and bulk flows."""
    notes_parts = [f"Successfully extracted {items_added} rows using AI with product catalog"]
    notes_parts.append(f"Extraction confidence: {confidence_score}%")
    if unmapped_count:
        notes_parts.append(f"{unmapped_count} product rows need QC mapping")
    if auto_mapped_count:
        notes_parts.append(f"{auto_mapped_count} correction-assisted rows need review")
    if special_row_count:
        notes_parts.append(f"{special_row_count} special rows were preserved outside product mapping")
    if skipped_division_count:
        notes_parts.append(f"{skipped_division_count} rows skipped (not in {statement_division} division)")
    if skipped_region_count:
        notes_parts.append(f"{skipped_region_count} rows skipped (product excluded in this region)")
    return ". ".join(notes_parts)


def _all_quantities_zero(row):
    """Return True when every movement qty is zero — indicates a phantom/empty row to discard."""
    return (
        not row.get("opening_qty")
        and not row.get("purchase_qty")
        and not row.get("sales_qty")
        and not row.get("operational_sales_qty")
        and not row.get("free_qty")
        and not row.get("closing_qty")
    )


def _build_statement_rows(extracted_data, statement_division=None, products_list=None,
                          statement_region=None):
    """Convert Gemini response rows into statement child rows and summary counts."""
    division_product_map = _build_division_product_map(statement_division, products_list)
    region_excluded_codes = _build_region_excluded_codes(statement_region)
    statement_rows = []
    counts = {
        "unmapped_count": 0,
        "auto_mapped_count": 0,
        "special_row_count": 0,
        "skipped_division_count": 0,
        "skipped_region_count": 0,
    }

    for item_data in extracted_data:
        row, metadata = _build_statement_item_row(
            item_data,
            statement_division=statement_division,
            division_product_map=division_product_map,
            region_excluded_codes=region_excluded_codes,
        )
        if metadata.get("skipped_division"):
            counts["skipped_division_count"] += 1
            continue
        if metadata.get("skipped_region"):
            counts["skipped_region_count"] += 1
            continue

        # Drop total/summary rows and all-zero phantom rows
        if metadata.get("skipped_total_row") or row is None:
            continue
        if _all_quantities_zero(row):
            frappe.logger().info(f"All-zero row dropped: {row.get('raw_product_name')}")
            continue

        if metadata.get("unmapped"):
            counts["unmapped_count"] += 1
        if metadata.get("auto_mapped"):
            counts["auto_mapped_count"] += 1
        if metadata.get("special_row"):
            counts["special_row_count"] += 1

        statement_rows.append(row)

    return statement_rows, counts


def _replace_statement_items(doc, statement_rows):
    """Replace a statement's items with normalized extraction rows."""
    doc.items = []
    for row in statement_rows:
        doc.append("items", row)


def _do_extract(doc_name, file_url):
    """
    Actual extraction logic, runs inside a background thread with its own DB connection.
    """
    api_key, model_name, is_enabled = get_gemini_settings()
    genai_client = genai_sdk.Client(api_key=api_key)

    doc = frappe.get_doc("Stockist Statement", doc_name)

    from frappe.utils.file_manager import get_file_path
    file_path = get_file_path(file_url)

    # Build product catalog scoped to the statement's division so Gemini only
    # sees (and returns) this division's product codes.
    product_catalog, products_list = build_product_catalog_for_prompt(doc.division)

    # Extract with enhanced prompt
    extracted_data = call_gemini_extraction_with_catalog(
        file_path,
        doc.stockist_code,
        product_catalog,
        products_list,
        model_name,
        genai_client
    )

    if not extracted_data or len(extracted_data) == 0:
        doc.extracted_data_status = "Failed"
        doc.extraction_notes = "No data extracted - AI returned empty results"
        doc.save()
        frappe.db.commit()
        return

    statement_rows, counts = _build_statement_rows(
        extracted_data,
        statement_division=doc.division,
        products_list=products_list,
        statement_region=doc.region,
    )
    _replace_statement_items(doc, statement_rows)
    items_added = len(statement_rows)

    if items_added == 0:
        doc.extracted_data_status = "Failed"
        doc.extraction_notes = "No valid products found in extracted data"
        doc.save()
        frappe.db.commit()
        return

    doc.confidence_score = _calculate_confidence_score(statement_rows)

    doc.extracted_data_status = "Completed"
    doc.extraction_notes = _build_extraction_notes(
        items_added,
        doc.confidence_score,
        unmapped_count=counts["unmapped_count"],
        auto_mapped_count=counts["auto_mapped_count"],
        special_row_count=counts["special_row_count"],
        skipped_division_count=counts["skipped_division_count"],
        statement_division=doc.division,
        skipped_region_count=counts["skipped_region_count"],
    )
    doc.populate_previous_month_closing()
    doc.calculate_closing_and_totals()
    doc.calculate_qc_confidence()

    # Extract raw printed sales total from the document (non-blocking, isolated)
    try:
        doc.ocr_raw_sales_total = _extract_statement_sales_total(file_path, genai_client, model_name)
    except Exception:
        doc.ocr_raw_sales_total = "not visible"

    doc.save()
    frappe.db.commit()


@frappe.whitelist()
def check_extraction_status(doc_name):
    """Poll endpoint: returns current extraction status for a Stockist Statement."""
    status = frappe.db.get_value(
        "Stockist Statement", doc_name,
        ["extracted_data_status", "extraction_notes"],
        as_dict=True
    )
    if not status:
        return {"success": False, "message": "Document not found"}
    return {
        "success": True,
        "status": status.extracted_data_status,
        "notes": status.extraction_notes or ""
    }


@frappe.whitelist()
@require_process("secondary")
def save_extracted_statement(doc_name, data):
    """
    Save the extracted and QC-verified data, then immediately submit the statement
    (docstatus → 1). This is a one-way finalisation — no edits after this point.
    Called from the portal after extraction + optional QC edits.
    data: JSON array of product rows (same structure as extracted_data in JS)
    """
    try:
        if isinstance(data, str):
            data = json.loads(data)

        doc = frappe.get_doc("Stockist Statement", doc_name)

        if doc.docstatus == 1:
            frappe.throw(_("This statement has already been submitted and cannot be changed."))

        # Rows may carry the business Product Code (display value) or the id;
        # resolve either to the id within this statement's division.
        code_to_pk = _build_division_product_map(doc.division)
        valid_pks = set(code_to_pk.values())

        # Replace items in doc
        doc.items = []
        for row in data:
            product_code = row.get("productcode") or row.get("product_code") or None
            if product_code and product_code not in valid_pks:
                product_code = code_to_pk.get(product_code) or _resolve_product_pk(product_code, doc.division) or product_code
            mapping_status = row.get("mapping_status") or row.get("mappingstatus") or "matched"
            raw_product_name = row.get("raw_product_name") or row.get("rawproductname") or ""
            row_type = _normalize_row_type(row.get("rowtype") or row.get("row_type"), raw_product_name)
            sales_qty = _parse_numeric_value(_get_first_present_value(row, "salesqty", "sales_qty"))
            operational_sales_qty = _parse_numeric_value(_get_first_present_value(row, "operationalsalesqty", "operational_sales_qty"))

            if row_type != "product":
                product_code = None
                mapping_status = "matched"
                if not operational_sales_qty:
                    operational_sales_qty = sales_qty
                sales_qty = 0

            # Unmapped items are kept in the statement but won't contribute to totals
            item_dict = {
                "raw_product_name": raw_product_name,
                "row_type": row_type,
                "mapping_status": mapping_status,
                "opening_qty": _parse_numeric_value(_get_first_present_value(row, "openingqty", "opening_qty")),
                "purchase_qty": _parse_numeric_value(_get_first_present_value(row, "purchaseqty", "purchase_qty")),
                "sales_qty": sales_qty,
                "operational_sales_qty": operational_sales_qty,
                "free_qty": _parse_numeric_value(_get_first_present_value(row, "freeqty", "free_qty")),
                "free_qty_scheme": _parse_numeric_value(_get_first_present_value(row, "freeqtyscheme", "free_qty_scheme")),
                "pts": _parse_numeric_value(_get_first_present_value(row, "pts")),
                "return_qty": _parse_numeric_value(_get_first_present_value(row, "returnqty", "return_qty")),
                "misc_out_qty": _parse_numeric_value(_get_first_present_value(row, "miscoutqty", "misc_out_qty")),
                "closing_qty": _parse_numeric_value(_get_first_present_value(row, "closingqty", "closing_qty")),
                "closing_value": _parse_numeric_value(_get_first_present_value(row, "closingvalue", "closing_value")),
                "row_confidence": _parse_numeric_value(_get_first_present_value(row, "row_confidence", "confidence")),
            }
            if product_code:
                item_dict["product_code"] = product_code
            doc.append("items", item_dict)

        doc.extracted_data_status = "Completed"
        doc.calculate_closing_and_totals()
        # Save first (persist items)
        doc.save(ignore_permissions=True)
        frappe.db.commit()
        # Submit — sets docstatus = 1, locks the document permanently
        doc.submit()
        frappe.db.commit()

        return {
            "success": True,
            "message": f"Statement submitted with {len(doc.items)} items.",
            "doc_name": doc.name
        }

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Save Extracted Statement Error")
        return {"success": False, "message": str(e)}


@frappe.whitelist()
@require_process("secondary")
def save_draft_statement(doc_name, data):
    """
    Persist the current items of a draft Stockist Statement without submitting.
    Called automatically from the portal QC screen on every edit (debounced).
    """
    try:
        if isinstance(data, str):
            data = json.loads(data)

        doc = frappe.get_doc("Stockist Statement", doc_name)

        if doc.docstatus == 1:
            return {"success": False, "message": "Statement already submitted."}

        # Rows may carry the business Product Code (display value) or the id;
        # resolve either to the id within this statement's division.
        code_to_pk = _build_division_product_map(doc.division)
        valid_pks = set(code_to_pk.values())

        doc.items = []
        for row in data:
            product_code = row.get("productcode") or row.get("product_code") or None
            if product_code and product_code not in valid_pks:
                product_code = code_to_pk.get(product_code) or _resolve_product_pk(product_code, doc.division) or product_code
            mapping_status = row.get("mapping_status") or row.get("mappingstatus") or "matched"
            raw_product_name = row.get("raw_product_name") or row.get("rawproductname") or ""
            row_type = _normalize_row_type(row.get("rowtype") or row.get("row_type"), raw_product_name)
            sales_qty = _parse_numeric_value(_get_first_present_value(row, "salesqty", "sales_qty"))
            operational_sales_qty = _parse_numeric_value(_get_first_present_value(row, "operationalsalesqty", "operational_sales_qty"))

            if row_type != "product":
                product_code = None
                mapping_status = "matched"
                if not operational_sales_qty:
                    operational_sales_qty = sales_qty
                sales_qty = 0

            item_dict = {
                "raw_product_name": raw_product_name,
                "row_type": row_type,
                "mapping_status": mapping_status,
                "opening_qty": _parse_numeric_value(_get_first_present_value(row, "openingqty", "opening_qty")),
                "purchase_qty": _parse_numeric_value(_get_first_present_value(row, "purchaseqty", "purchase_qty")),
                "sales_qty": sales_qty,
                "operational_sales_qty": operational_sales_qty,
                "free_qty": _parse_numeric_value(_get_first_present_value(row, "freeqty", "free_qty")),
                "free_qty_scheme": _parse_numeric_value(_get_first_present_value(row, "freeqtyscheme", "free_qty_scheme")),
                "pts": _parse_numeric_value(_get_first_present_value(row, "pts")),
                "return_qty": _parse_numeric_value(_get_first_present_value(row, "returnqty", "return_qty")),
                "misc_out_qty": _parse_numeric_value(_get_first_present_value(row, "miscoutqty", "misc_out_qty")),
                "closing_qty": _parse_numeric_value(_get_first_present_value(row, "closingqty", "closing_qty")),
                "closing_value": _parse_numeric_value(_get_first_present_value(row, "closingvalue", "closing_value")),
                "row_confidence": _parse_numeric_value(_get_first_present_value(row, "row_confidence", "confidence")),
            }
            if product_code:
                item_dict["product_code"] = product_code
            doc.append("items", item_dict)

        doc.calculate_closing_and_totals()
        doc.save(ignore_permissions=True)
        frappe.db.commit()

        return {"success": True, "message": f"Draft saved ({len(doc.items)} items)."}

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Save Draft Statement Error")
        return {"success": False, "message": str(e)}


def get_stockist_code_map(pks):
    """Map Stockist Master id (PK, e.g. S1685) -> editable business code (e.g. S0009).

    Used for DISPLAY only: records link to stockists by the internal id, but screens and
    reports must show the human-facing Stockist Code. Falls back to the id when a stockist
    has no code set so nothing renders blank.
    """
    pks = [p for p in set(pks or []) if p]
    if not pks:
        return {}
    rows = frappe.get_all(
        "Stockist Master",
        filters={"name": ["in", pks]},
        fields=["name", "stockist_code"],
        limit_page_length=0,
    )
    return {r.name: (r.stockist_code or r.name) for r in rows}


def get_product_code_map(pks):
    """Map Product Master id (PK, e.g. PRD-0001) -> editable business code.

    Used for DISPLAY only: statement/scheme items link to products by the internal
    id, but screens, reports, and exports must show the human-facing Product Code.
    Falls back to the id when a product has no code set (or for legacy rows where
    the id IS the code) so nothing renders blank."""
    pks = [p for p in set(pks or []) if p]
    if not pks:
        return {}
    rows = frappe.get_all(
        "Product Master",
        filters={"name": ["in", pks]},
        fields=["name", "product_code"],
        limit_page_length=0,
    )
    return {r.name: (r.product_code or r.name) for r in rows}


def _apply_product_display_codes(rows, *keys):
    """Rewrite Product Master ids in report rows to editable business codes.

    `rows` is a list of dicts fresh from frappe.db.sql(as_dict=True); each key in
    `keys` (default: "product_code") names a column holding a Product Master id.
    Grouping/joining stays on the id inside SQL — this is display-layer only."""
    keys = keys or ("product_code",)
    ids = {r.get(k) for r in rows for k in keys if r.get(k)}
    code_map = get_product_code_map(ids)
    if not code_map:
        return rows
    for r in rows:
        for k in keys:
            if r.get(k):
                r[k] = code_map.get(r[k], r[k])
    return rows


@frappe.whitelist()
def get_statement_for_view(doc_name):
    """
    Return full statement data for the portal view page.
    """
    try:
        doc = frappe.get_doc("Stockist Statement", doc_name)
        # Build items list enriched with product master info
        items = []
        for item in doc.items:
            product = {}
            row_type = item.row_type or _normalize_row_type(None, item.raw_product_name)
            operational_sales_qty = flt(item.operational_sales_qty or (row_type != "product" and item.sales_qty or 0))
            sales_qty = 0 if row_type != "product" else flt(item.sales_qty)
            if item.product_code:
                product = frappe.db.get_value(
                    "Product Master", item.product_code,
                    ["product_code", "product_name", "pts", "ptr", "pack"], as_dict=True
                ) or {}
            items.append({
                # Human-facing business code for display/round-trip; save paths
                # resolve it back to the id within the statement's division.
                "productcode": product.get("product_code") or item.product_code or "",
                "productname": item.product_name or product.get("product_name", ""),
                "rawproductname": item.raw_product_name or "",
                "rowtype": row_type,
                "mappingstatus": item.mapping_status or "matched",
                "pack": item.pack or product.get("pack", ""),
                "pts": flt(item.pts or product.get("pts", 0)),
                "ptr": flt(product.get("ptr", 0)),
                "conversion_factor": flt(item.conversion_factor or 1),
                "openingqty": flt(item.opening_qty),
                "purchaseqty": flt(item.purchase_qty),
                "salesqty": sales_qty,
                "operationalsalesqty": operational_sales_qty,
                "freeqty": flt(item.free_qty),
                "freeqtyscheme": flt(item.free_qty_scheme),
                "returnqty": flt(item.return_qty),
                "miscoutqty": flt(item.misc_out_qty),
                "closingqty": flt(item.closing_qty),
                "closingvalue": flt(item.closing_value),
                "openingvalue": flt(item.opening_value),
                "purchasevalue": flt(item.purchase_value),
                "salesvaluepts": flt(item.sales_value_pts),
                "salesvalueptr": flt(item.sales_value_ptr),
                "schemedeductedqty": flt(item.scheme_deducted_qty_calc),
                "row_confidence": flt(item.row_confidence),
                "math_check": item.math_check or "N/A",
            })

        return {
            "success": True,
            "doc": {
                "name": doc.name,
                "stockist_code": doc.stockist_code,
                # Human-facing code for display; stockist_code stays the internal id so the
                # primary-sales lookup keeps working.
                "stockist_display_code": get_stockist_code_map([doc.stockist_code]).get(doc.stockist_code, doc.stockist_code),
                "stockist_name": doc.stockist_name,
                "statement_month": str(doc.statement_month),
                "hq": doc.hq,
                "team": doc.team,
                "region": doc.region,
                "zone": doc.zone,
                "extracted_data_status": doc.extracted_data_status,
                "uploaded_file": doc.uploaded_file,
                "docstatus": doc.docstatus,
                "qc_confidence": doc.qc_confidence or "",
                "confidence_score": flt(doc.confidence_score),
                "total_opening_value": flt(doc.total_opening_value),
                "total_purchase_value": flt(doc.total_purchase_value),
                "total_operational_sales_qty": flt(doc.total_operational_sales_qty),
                "total_closing_value": flt(doc.total_closing_value),
                "total_sales_value_pts": flt(doc.total_sales_value_pts),
                "total_sales_value_ptr": flt(doc.total_sales_value_ptr),
                "ocr_raw_sales_total": doc.ocr_raw_sales_total or "not visible",
                "ocr_raw_purchase_total": getattr(doc, "ocr_raw_purchase_total", "") or "",
                "ocr_raw_opening_total": getattr(doc, "ocr_raw_opening_total", "") or "",
                "ocr_raw_closing_total": getattr(doc, "ocr_raw_closing_total", "") or "",
                "hq_name": frappe.db.get_value("HQ Master", doc.hq, "hq_name") if doc.hq else "",
                "team_name": frappe.db.get_value("Team Master", doc.team, "team_name") if doc.team else "",
                "division": doc.division or "",
            },
            "items": items
        }
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Get Statement For View Error")
        return {"success": False, "message": str(e)}


@frappe.whitelist()
def get_primary_sales_for_stockist(stockist_code, statement_month):
    """
    Fetch Primary Sales Data rows for a given stockist and month.
    statement_month can be 'YYYY-MM-DD' or 'YYYY-MM'; we extract YYYY-MM.
    """
    if not stockist_code or not statement_month:
        return {"success": False, "message": "Stockist code and month are required."}

    # Normalise to YYYY-MM then derive first/last day of that month
    month_str = str(statement_month)[:7]          # "2026-02"
    from frappe.utils import getdate, get_last_day
    first_day = getdate(month_str + "-01")
    last_day = get_last_day(first_day)

    rows = frappe.get_all(
        "Primary Sales Data",
        filters={
            "stockist_code": stockist_code,
            "invoicedate": ["between", [first_day, last_day]],
            "iscancelled": 0,
        },
        fields=[
            "invoiceno", "invoicedate", "pcode", "product", "pack",
            "batchno", "quantity", "freeqty", "pts", "ptsvalue",
            "ptr", "ptrvalue",
        ],
        order_by="invoicedate asc, pcode asc",
        limit_page_length=0,
    )

    return {"success": True, "rows": rows, "month": month_str}


def _extract_statement_sales_total(file_path, genai_client, model_name):
    """
    Lightweight, isolated Gemini call that extracts only the printed grand-total
    sales quantity/value from the statement document.

    Returns the raw value as a string (e.g. "1250" or "45,230.50") or the
    literal string "not visible" when no printed total is found.
    Does NOT touch the main item-extraction flow or bulk OCR.
    """
    file_ext = os.path.splitext(file_path)[1].lower()
    mime_type, _ = mimetypes.guess_type(file_path)

    prompt = (
        "You are looking at a pharmaceutical stockist statement.\n"
        "Find the PRINTED TOTAL of the Sales column — this is a footer/grand-total row "
        "that summarises the entire statement (e.g. 'Total Sales', 'Grand Total', "
        "'Total Qty' in the sales column, 'TOTAL QTY', etc.).\n\n"
        "Return ONLY a JSON object with ONE key:\n"
        '{"statement_sales_total": <number or "not visible">}\n\n'
        "Rules:\n"
        "- If a printed sales total is clearly visible, return its numeric value "
        "(strip commas/currency symbols, e.g. 1250 or 45230.50).\n"
        '- If no sales total is printed or you cannot locate it, return {"statement_sales_total": "not visible"}.\n'
        "- Return ONLY valid JSON. No markdown, no explanation."
    )

    # Keep this sub-call light: minimal reasoning, model-family aware.
    generation_config = genai_types.GenerateContentConfig(
        thinking_config=_thinking_config(model_name, gemini3_level="low", gemini25_budget=256)
    )

    try:
        if file_ext in [".pdf", ".jpg", ".jpeg", ".png"]:
            with open(file_path, "rb") as f:
                file_data = f.read()
            if file_ext == ".pdf":
                file_part = genai_types.Part.from_bytes(data=file_data, mime_type="application/pdf")
            else:
                file_part = genai_types.Part.from_bytes(data=file_data, mime_type=mime_type or "image/jpeg")
            response = genai_client.models.generate_content(
                model=model_name, contents=[prompt, file_part], config=generation_config
            )
        elif file_ext in [".csv", ".txt"]:
            with open(file_path, "r", encoding="utf-8") as f:
                file_content = f.read()
            response = genai_client.models.generate_content(
                model=model_name,
                contents=f"{prompt}\n\nCONTENT:\n{file_content}",
                config=generation_config
            )
        elif file_ext in [".xls", ".xlsx"]:
            import pandas as pd
            df = pd.read_excel(file_path)
            file_content = df.to_string()
            response = genai_client.models.generate_content(
                model=model_name,
                contents=f"{prompt}\n\nCONTENT:\n{file_content}",
                config=generation_config
            )
        else:
            return "not visible"

        response_text = response.text.strip()
        # Strip markdown fences if present
        if response_text.startswith("```"):
            response_text = response_text.split("```", 1)[1]
        if response_text.startswith("json"):
            response_text = response_text[4:]
        if response_text.endswith("```"):
            response_text = response_text.rsplit("```", 1)[0]
        response_text = response_text.strip()

        data = json.loads(response_text)
        val = data.get("statement_sales_total")
        if val is None or str(val).strip().lower() in ("not visible", ""):
            return "not visible"
        return str(val).strip()

    except Exception as e:
        frappe.logger().warning(f"Statement sales total extraction failed (non-critical): {e}")
        return "not visible"


def call_gemini_extraction_with_catalog(file_path, stockist_code, product_catalog, products_list, model_name=None, genai_client=None):
    """
    Enhanced extraction that sends the full product catalog to Gemini
    Gemini does the matching directly using product codes
    """
    if not genai_client:
        api_key, model_name, is_enabled = get_gemini_settings()
        genai_client = genai_sdk.Client(api_key=api_key)
    
    try:
        # Determine file type
        mime_type, _ = mimetypes.guess_type(file_path)
        file_ext = os.path.splitext(file_path)[1].lower()
        correction_prompt = _build_correction_prompt(stockist_code)
        correction_map = _build_correction_map(stockist_code)
        generation_config = _build_gemini_generation_config(model_name)
        
        # Enhanced prompt with product catalog
        prompt = f"""You are extracting pharmaceutical stockist statement data for STEDMAN PHARMACEUTICALS.

{product_catalog}
    {correction_prompt}

=== EXTRACTION RULES ===

1. COLUMN DETECTION (DO THIS FIRST — CRITICAL):
   - FIRST: Read ALL column headers in the document from left to right.
   - Map each header to one of the standard fields below:
     . Opening Qty: "Op.Qty", "OPSTK", "Opening", "Open.Qty", "QpnStk", "Op.Stk"
     . Purchase Qty: "Purch.Qty", "PURCH", "Receipt", "Pr.Qty", "Pur", "Recv"
     . Sales Qty (PRIMARY): "Sales", "Sale", "Sl", "Sold", "Sales Qty", "S.Qty"
     . ADDITIONAL SALES-QUANTITY columns — extra sales channels that MUST be ADDED into Sales Qty (see Rule 2C):
         · Branch Sales Qty: "Br.S.Qty", "Br Sales", "Branch Sales", "Branch Sales Qty", or a "Br" column that is clearly a sales quantity
         · Hospital Sales Qty: "HOS.SALES", "Hos", "Hosp", "Hospital Sales", "Hospital Sales Qty", "Hos.Qty"
         · Transfer Sales Qty: "Transfer Sales", "Transfer Sales Qty", "Trf Sales"
         · Other Sales Qty: "Other Sales", "Others Sales", "Oth Sales"
     . Free Qty: "Free Qty", "Free", "Scheme Qty", "Fre"
     . Return Qty: "Return", "Ret", "Sales Ret", "SR", "Sal.Ret"
     . Misc Out: "Misc.Out", "M.Out", "Transfer", "Trans", "Adj"
       (NOTE: a plain "Transfer"/"Trans"/"Misc.Out" stock-movement column is NOT a sales channel — it stays as Misc Out and is NOT added to Sales Qty. Only an explicit "Transfer SALES" column counts as an additional sales channel.)
     . Closing Qty: "Closing", "Cls", "Cl.Bal", "ClsStk", "Closing Qty", "Balance", "Bal"
     . Closing Value: "Closing Value", "Closing Val", "Cl.Value", "Closing Amount", "Cls.Val"
   - IMPORTANT: Distinguish QUANTITY columns (integers) from VALUE/AMOUNT columns (decimals with currency).
     Do NOT extract value columns as quantities. Value columns typically have large decimal numbers.
   - IGNORE previous-month or historical columns: PMSalesQty, Prev.Cls, LM Sales, Last Month, PM Cls, etc.
   - IMPORTANT: "AVSL" is NOT an Opening column. It typically refers to a scheme/free goods allocation or a sales benchmark — do NOT map it to opening_qty.
   - VERIFY: For a few sample rows, check that Opening + Purchase - Sales - Free - Return - MiscOut ≈ Closing.
     If this does NOT add up, re-examine your column assignments — you may have the wrong column mapped.

2. PRODUCT MATCHING (CRITICAL — STRICT MATCHING ONLY):
   - Match each product ONLY when you are confident. Use ALL of these together:
     a) Product NAME must closely match a catalog entry
     b) Pack SIZE must be consistent with the catalog
     c) Pack CONVERSION must be consistent
   - Be aware of common OCR misreads: G↔C, O↔0, I↔1↔L, 5↔S, 8↔B, rn↔m, cl↔d
     Example: "GALCIGEN" might be OCR misread of "CALCIGEN" — use pack size to disambiguate.
   - If a product name is ambiguous or could match multiple catalog entries, mark it UNMAPPED rather than guessing.
   - Return ALL products from the statement, including those you CANNOT match.
   - For matched products: set "product_code" to the EXACT catalog code and "unmapped" to false
   - For unmatched products: set "product_code" to null and "unmapped" to true
    - If you rely on a stockist-specific correction hint and it agrees with the visible row text and pack, set
      "mapping_basis" to "stockist_correction_hint".

2A. SPECIAL ROWS (CRITICAL):
    - Preserve operational rows such as "OTHERS" and "BRANCH TRANSFER".
    - Do NOT force these rows into Product Master mapping.
    - Set "row_type" to one of: "product", "others", "branch_transfer".
    - For special rows, set "product_code" to null, "unmapped" to false, and "mapping_basis" to "special_row".
    - For special rows, copy the printed movement quantity into "operational_sales_qty" and set "sales_qty" to 0.
    - For normal product rows, set "operational_sales_qty" to 0.

2B. ROWS TO EXCLUDE — *** CRITICAL — DO NOT INCLUDE THESE IN OUTPUT ***:
    Pharmaceutical stockist statements always print one or more aggregate/footer rows at the bottom (or
    sometimes mid-table) that summarise the entire statement — e.g. a row for total quantity movement and
    another for total closing value. These are NOT products. Including them would double-count closing stock.
    You MUST silently skip (omit from JSON) any row whose label clearly describes an aggregate, total, or
    summary figure. This includes — but is NOT limited to — the following patterns:
      • "TOTAL QUANTITY" / "TOTAL QTY" / "TOTAL STOCK" / "TOTAL STKS"
      • "TOTAL VALUE" / "TOTAL AMOUNT" / "TOTAL VAL" / "TOTAL CLOSING VALUE"
      • "TOTAL" (standalone)
      • "GRAND TOTAL" / "GRAND TOTAL QTY" / "GRAND TOTAL VALUE"
      • "SUB TOTAL" / "SUBTOTAL"
      • Any row whose first column is blank or dashes and whose numeric columns contain obviously
        aggregated figures (e.g. the sum of every other row in that column)
    Rule of thumb: if the row is a footer/summary line rather than an individual stock-keeping unit,
    EXCLUDE it — regardless of exact wording, language, or abbreviation used.

2C. COMBINED SALES QUANTITY (CRITICAL):
    - A statement may split sales across SEVERAL quantity columns. When ANY additional sales-quantity
      column from Rule 1 is present, "sales_qty" MUST be the SUM of the primary Sales column PLUS every
      additional sales-quantity column present in that row:
        sales_qty = Sales + Branch Sales (Br / Br.S.Qty) + Hospital Sales (Hos / HOS.SALES)
                    + Transfer Sales + Other Sales
      Example (HOS.SALES present): Sales = 4 and HOS.SALES = 120  ->  sales_qty = 124.
      Example (Br.S.Qty present):  S.Qty = 20 and Br.S.Qty = 20   ->  sales_qty = 40.
    - Add ONLY quantity columns. NEVER add any VALUE/AMOUNT column (e.g. "Br.S.Val", "S.Val", "P.Val",
      hospital value, "CL.Value"). Value columns are decimals/currency — they are not sales quantities.
      Branch/Hospital quantity and value columns usually sit side by side; pick the QUANTITY one only.
    - Add a column ONLY if it is ACTUALLY present in this statement. If a channel column is absent, add nothing for it.
    - Do NOT add Free/Scheme qty, Return qty, or a plain Misc.Out/Transfer (stock-movement) column into sales_qty.
    - Do NOT double-count: each sales-quantity column is added exactly once.
    - This rule is about COLUMNS inside a normal product row. It is DISTINCT from the special
      "OTHERS"/"BRANCH TRANSFER" ROWS in Rule 2A — those remain separate rows handled via operational_sales_qty.
    - The math self-check (Opening + Purchase - Sales - Free - Return - MiscOut ≈ Closing) uses this COMBINED sales_qty.

3. QUANTITY EXTRACTION (CRITICAL):
   - Extract quantities for all identified columns AND closing value if present.
   - ONLY extract values for columns that ACTUALLY EXIST in the statement. Do NOT calculate or infer missing columns.
   - If a column does not exist in the statement, set its value to 0. Do NOT compute it from other columns.
   - If Closing Qty column exists in the statement, extract the printed value directly.
   - If Closing Qty column does NOT exist in the statement, set closing_qty to 0. Do NOT calculate it as Opening + Purchase - Sales - etc.
   - If Closing Value column exists in the statement, extract the printed value directly.
   - If Closing Value column does NOT exist in the statement, set closing_value to 0. Do NOT calculate it.
    - Preserve negative quantities exactly as printed.
    - If a quantity is shown in parentheses, output it as a negative number.
   - SELF-CHECK each row: For columns that DO exist, verify Opening + Purchase - Sales - Free - Return - MiscOut should ≈ Closing.
     If a row's math is way off, double-check which column values you assigned.

4. EXTRACTION CONFIDENCE (CRITICAL — per row):
   - For EACH row, assign a "confidence" integer from 0 to 100 representing how confident you are that the extracted data is correct.
   - Score 100: You can clearly read ALL fields in the row, the values make sense, and math sanity checks out. USE 100 LIBERALLY — if you can read it, it's 100.
   - Score 70-99: One or more values are ambiguous — a blurry digit, faded text, overlapping columns, or the math doesn't add up (likely an OCR misread).
   - Score 30-69: Several fields are hard to read, the row layout is confusing, or quantities seem very unusual.
   - Score 0-29: Very poor readability, heavy guessing, or the row is heavily damaged/truncated.
   - For unmapped products: still score based on how well you could READ the quantities (mapping is separate).
   - Sanity check: if Opening + Purchase - Sales - Free - Return - MiscOut does NOT equal Closing and the difference is significant, REDUCE confidence below 100 (the OCR likely got a digit wrong).
   - If some columns are genuinely absent from the statement (e.g. no Free or Return column), that is NOT a reason to reduce confidence — only reduce when you are uncertain about what you read.
   - IMPORTANT: Do NOT artificially lower confidence. If you can read a row clearly and the math checks out, it MUST be 100.

5. OUTPUT FORMAT:
   - Return ONLY a valid JSON array — no markdown, no explanation, no extra text.
   - Include ONLY quantities (NO price/value columns except closing_value).
   - Include ALL products from the statement, matched and unmatched.

EXPECTED JSON FORMAT:
[
  {{
    "product_code": "ARC",
    "raw_product_name": "ARCALION 200MG",
        "row_type": "product",
        "mapping_basis": "catalog_exact",
    "unmapped": false,
    "confidence": 98,
    "opening_qty": 88,
    "purchase_qty": 29,
    "sales_qty": 59,
    "operational_sales_qty": 0,
    "free_qty": 0,
    "return_qty": 0,
    "misc_out_qty": 0,
    "closing_qty": 58,
    "closing_value": 500.00
  }},
  {{
    "product_code": null,
        "raw_product_name": "BRANCH TRANSFER",
        "row_type": "branch_transfer",
        "mapping_basis": "special_row",
        "unmapped": false,
        "confidence": 92,
    "opening_qty": 10,
    "purchase_qty": 5,
    "sales_qty": 0,
    "operational_sales_qty": 8,
    "free_qty": 0,
    "return_qty": 0,
    "misc_out_qty": 0,
    "closing_qty": 7,
    "closing_value": 0
  }}
]

IMPORTANT:
- NO values/prices in output (except closing_value)
- NO markdown formatting
- ONLY valid JSON array
- Always include row_type and mapping_basis for every row
- Use EXACT product codes from catalog for matched products
- Set product_code to null for unmatched products
- Always include raw_product_name for ALL products
- Always include confidence (0-100) for ALL products
- Always include operational_sales_qty for every row
"""
        
        frappe.logger().info(f"Using Gemini model: {model_name}")

        # Retry logic for rate limiting (429) and unavailable (503)
        max_retries = 3
        base_delay = 2
        response = None
        current_model = model_name
        # Fallback model when primary is overloaded (503). Use a stable, widely
        # available GA model that differs from the primary so the retry actually
        # lands on different capacity.
        fallback_model = "gemini-2.5-flash" if model_name != "gemini-2.5-flash" else "gemini-3.5-flash"
        attempt = 0
        used_fallback = False

        while attempt < max_retries:
            try:
                if file_ext in [".pdf", ".jpg", ".jpeg", ".png"]:
                    with open(file_path, "rb") as f:
                        file_data = f.read()

                    if file_ext == ".pdf":
                        file_part = genai_types.Part.from_bytes(
                            data=file_data, mime_type="application/pdf"
                        )
                    else:
                        file_part = genai_types.Part.from_bytes(
                            data=file_data, mime_type=mime_type or "image/jpeg"
                        )

                    response = genai_client.models.generate_content(
                        model=current_model,
                        contents=[prompt, file_part],
                        config=generation_config
                    )

                elif file_ext in [".csv", ".txt"]:
                    with open(file_path, "r", encoding="utf-8") as f:
                        file_content = f.read()
                    response = genai_client.models.generate_content(
                        model=current_model,
                        contents=f"{prompt}\n\nCONTENT:\n{file_content}",
                        config=generation_config
                    )

                elif file_ext in [".xls", ".xlsx"]:
                    import pandas as pd
                    df = pd.read_excel(file_path)
                    file_content = df.to_string()
                    response = genai_client.models.generate_content(
                        model=current_model,
                        contents=f"{prompt}\n\nCONTENT:\n{file_content}",
                        config=generation_config
                    )

                else:
                    frappe.throw(f"Unsupported file type: {file_ext}")

                break  # Success

            except Exception as e:
                err_str = str(e).lower()
                is_rate_limit = "quota" in err_str or "rate" in err_str or "429" in err_str or "resource_exhausted" in err_str
                is_unavailable = "503" in err_str or "unavailable" in err_str or "overloaded" in err_str or "high demand" in err_str

                if (is_rate_limit or is_unavailable) and attempt < max_retries - 1:
                    delay = base_delay * (2 ** attempt)
                    frappe.logger().warning(
                        f"{'503 Unavailable' if is_unavailable else 'Rate limit'} on {current_model} "
                        f"(attempt {attempt + 1}/{max_retries}), retrying in {delay}s..."
                    )
                    time.sleep(delay)
                    attempt += 1
                elif is_unavailable and not used_fallback:
                    # Primary model exhausted retries — switch to fallback
                    frappe.logger().warning(
                        f"{current_model} unavailable after {max_retries} attempts, "
                        f"falling back to {fallback_model}"
                    )
                    current_model = fallback_model
                    used_fallback = True
                    attempt = 0
                    max_retries = 2
                    time.sleep(base_delay)
                elif is_rate_limit or is_unavailable:
                    frappe.throw(
                        f"Gemini API unavailable after retries (tried {model_name}"
                        f"{' and ' + fallback_model if used_fallback else ''}). "
                        f"Please try again later."
                    )
                else:
                    raise e
        
        # Parse response
        try:
            frappe.logger().info("=== GEMINI RAW RESPONSE START ===")
            frappe.logger().info(response.text)
            frappe.logger().info("=== GEMINI RAW RESPONSE END ===")
        except Exception as log_err:
            frappe.logger().error(f"Failed to log Gemini response: {log_err}")
        
        response_text = response.text.strip()
        
        # Clean response
        if response_text.startswith("```"):
            response_text = response_text.split("```", 1)[1]
        if response_text.startswith("json"):
            response_text = response_text[4:]
        if response_text.endswith("```"):
            response_text = response_text.rsplit("```", 1)[0]
        
        response_text = response_text.strip()
        
        frappe.logger().info("=== Cleaned Response Text ===")
        frappe.logger().info(response_text)
        
        # Robust JSON parsing with repair for common Gemini output issues
        extracted_items = None
        try:
            extracted_items = json.loads(response_text)
        except json.JSONDecodeError:
            # Attempt repairs: trailing commas, truncated JSON
            repaired = response_text
            # Remove trailing commas before ] or }
            repaired = re.sub(r',\s*([\]\}])', r'\1', repaired)
            # If JSON array is truncated (no closing ]), try to close it
            if repaired.lstrip().startswith('[') and not repaired.rstrip().endswith(']'):
                # Find last complete object (ending with })
                last_brace = repaired.rfind('}')
                if last_brace > 0:
                    repaired = repaired[:last_brace + 1] + ']'
            try:
                extracted_items = json.loads(repaired)
                frappe.logger().info("JSON parsed after repair")
            except json.JSONDecodeError as je:
                frappe.logger().error(f"JSON repair also failed: {je}")
                frappe.throw(f"Failed to parse AI response as JSON: {je}")
        
        extracted_items = extracted_items or []
        frappe.logger().info(f"Parsed Items Count: {len(extracted_items)}")
        
        # Validate product codes and tag mapping status
        valid_codes = {p["product_code"] for p in products_list}
        validated_items = []
        
        for item in extracted_items:
            raw_name = (item.get("raw_product_name") or "").strip()
            row_type = _normalize_row_type(item.get("row_type"), raw_name)
            mapping_basis = cstr(item.get("mapping_basis")).strip().lower()
            pc = (item.get("product_code") or "").strip() or None
            is_unmapped = item.get("unmapped", False)

            # Discard total/summary rows before any other processing
            if row_type == "total_row":
                frappe.logger().info(f"Total/summary row excluded: {raw_name}")
                continue

            if row_type != "product":
                item["product_code"] = None
                item["raw_product_name"] = raw_name
                item["row_type"] = row_type
                item["mapping_basis"] = mapping_basis or "special_row"
                item["unmapped"] = False
                item["operational_sales_qty"] = _parse_numeric_value(item.get("operational_sales_qty") or item.get("sales_qty"))
                item["sales_qty"] = 0
                validated_items.append(item)
                frappe.logger().info(f"Special row kept: {raw_name} (type={row_type})")
                continue

            if not mapping_basis and raw_name.strip().upper() in correction_map and correction_map[raw_name.strip().upper()] == pc:
                mapping_basis = "stockist_correction_hint"

            if pc and pc in valid_codes:
                # Gemini matched to a valid catalog product
                item["product_code"] = pc
                item["raw_product_name"] = raw_name
                item["row_type"] = row_type
                item["mapping_basis"] = mapping_basis or "catalog_match"
                item["unmapped"] = False
                item["operational_sales_qty"] = _parse_numeric_value(item.get("operational_sales_qty"))
                validated_items.append(item)
            else:
                # Unmatched — keep the item but mark unmapped
                item["product_code"] = None
                item["raw_product_name"] = raw_name
                item["row_type"] = row_type
                item["mapping_basis"] = mapping_basis or "unmapped"
                item["unmapped"] = True
                item["operational_sales_qty"] = _parse_numeric_value(item.get("operational_sales_qty"))
                validated_items.append(item)
                frappe.logger().info(f"Unmapped product kept: {raw_name} (code={pc})")
        
        frappe.logger().info(f"Final Items: {len(validated_items)} (matched: {sum(1 for i in validated_items if not i.get('unmapped'))}, unmapped: {sum(1 for i in validated_items if i.get('unmapped'))})")
        return validated_items
        
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Gemini API Call Error - Enhanced")
        frappe.throw(f"Extraction failed: {str(e)}")


@frappe.whitelist()
def map_filenames_to_stockists_via_gemini(filenames):
    """
    Use Gemini to map a list of filenames to stockist codes in a single call.
    Much more accurate than pure fuzzy matching.
    filenames: JSON string or list of filename strings
    Returns: dict mapping filename -> stockist_code (or null)
    """
    try:
        if isinstance(filenames, str):
            filenames = json.loads(filenames)

        if not filenames:
            return {"success": True, "mapping": {}}

        # Fetch stockist catalog (name + code only to minimize tokens)
        stockists = frappe.get_all(
            "Stockist Master",
            filters={"status": "Active"},
            fields=["name", "stockist_name"],
            order_by="stockist_name asc"
        )

        if not stockists:
            return {"success": False, "message": "No active stockists found"}

        # Build compact catalog
        catalog_lines = [f"{s['name']}|{s['stockist_name']}" for s in stockists]
        catalog_text = "\n".join(catalog_lines)

        filenames_text = "\n".join([f"{i+1}. {f}" for i, f in enumerate(filenames)])

        prompt = f"""You are matching pharmaceutical stockist statement filenames to stockist records.

STOCKIST CATALOG (format: CODE|NAME):
{catalog_text}

FILENAMES TO MATCH:
{filenames_text}

TASK: Match each filename to the most likely stockist code from the catalog above.
- The filename may contain part of the stockist name (possibly misspelled or abbreviated)
- Ignore date parts, month names, years, numbers, extensions
- Look for the distinctive name portion

Return a JSON object where keys are the exact filenames and values are the matched stockist CODE (or null if no good match):
{{
  "filename1.pdf": "CODE1",
  "filename2.jpg": null,
  ...
}}

IMPORTANT:
- Return ONLY valid JSON, no explanation
- Use exact filenames as keys (including extension)
- Use exact stockist CODEs from the catalog as values
- Return null if confidence is low
"""

        api_key, model_name, _ = get_gemini_settings()
        client = genai_sdk.Client(api_key=api_key)

        response = client.models.generate_content(
            model=model_name,
            contents=prompt,
            config=genai_types.GenerateContentConfig(
                thinking_config=_thinking_config(model_name)
            )
        )

        resp_text = response.text.strip()
        if resp_text.startswith("```"):
            resp_text = resp_text.split("```", 1)[1]
        if resp_text.lower().startswith("json"):
            resp_text = resp_text[4:]
        if resp_text.endswith("```"):
            resp_text = resp_text.rsplit("```", 1)[0]
        resp_text = resp_text.strip()

        mapping = json.loads(resp_text)

        # Validate that returned codes actually exist AND that the matched stockist
        # name is plausibly related to the filename (drops confident hallucinations).
        valid_codes = {s["name"] for s in stockists}
        name_by_code = {s["name"]: s["stockist_name"] for s in stockists}
        validated = {}
        for fname, code in mapping.items():
            if code and code in valid_codes \
                    and _stockist_name_plausible_for_filename(fname, name_by_code.get(code, "")):
                validated[fname] = code
            else:
                validated[fname] = None

        return {"success": True, "mapping": validated}

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Filename→Stockist Gemini Mapping Error")
        return {"success": False, "message": str(e)}


@frappe.whitelist()
@require_process("secondary")
def start_bulk_ocr_job(docname):
    """
    Portal entry point: validates the doc, then runs extraction in a background thread.
    Uses threading instead of RQ enqueue to avoid os.fork() deadlocks in WSL/gunicorn environments.
    The thread gets its own Frappe DB connection via frappe.init()/connect() so it is fully
    independent; progress is written to DB and polled by the frontend every 5 seconds.
    """
    import threading

    try:
        doc = frappe.get_doc("Bulk Statement Upload", docname)

        if doc.status in ("In Progress",):
            return {"success": False, "message": f"Job is already {doc.status}. Cannot restart."}

        if not doc.zip_file:
            return {"success": False, "message": "No ZIP file attached to this job."}

        # Capture all context-specific values before the request context is torn down
        site = frappe.local.site
        zip_file_url = doc.zip_file
        month = str(doc.statement_month)

        # Immediately mark as Queued so the UI updates
        doc.status = "Queued"
        doc.save(ignore_permissions=True)
        frappe.db.commit()

        def run_in_thread():
            """Thread target: initialises its own Frappe connection, runs extraction, then destroys."""
            try:
                frappe.init(site=site)
                frappe.connect()
                process_bulk_extraction(docname=docname, month=month, zip_file_url=zip_file_url)
            except Exception as thread_err:
                # Best-effort: mark doc as Failed with error info
                try:
                    frappe.init(site=site)
                    frappe.connect()
                    _doc = frappe.get_doc("Bulk Statement Upload", docname)
                    _doc.status = "Failed"
                    _doc.extraction_log = json.dumps([
                        {"file": "—", "status": "Failed", "message": str(thread_err)}
                    ])
                    _doc.save(ignore_permissions=True)
                    frappe.db.commit()
                except Exception:
                    pass
            finally:
                try:
                    frappe.destroy()
                except Exception:
                    pass

        t = threading.Thread(target=run_in_thread, daemon=True, name=f"bulk_ocr_{docname}")
        t.start()

        return {"success": True, "message": "Bulk OCR job started in background thread", "job_id": docname}

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Start Bulk OCR Job Error")
        return {"success": False, "message": str(e)}


@frappe.whitelist()
def get_bulk_job_status(docname):
    """
    Return current status/progress of a bulk job for live polling.
    """
    try:
        doc = frappe.get_doc("Bulk Statement Upload", docname)
        log_data = []
        if doc.extraction_log:
            try:
                log_data = json.loads(doc.extraction_log)
            except Exception:
                log_data = []

        # Enrich log with live QC confidence from actual statements
        statement_names = [r["statement"] for r in log_data if r.get("statement")]
        if statement_names:
            live_data = {
                row.name: row
                for row in frappe.get_all(
                    "Stockist Statement",
                    filters={"name": ["in", statement_names]},
                    fields=["name", "qc_confidence", "confidence_score"],
                )
            }
            for entry in log_data:
                st = entry.get("statement")
                if st and st in live_data:
                    entry["qc_confidence"] = live_data[st].qc_confidence or entry.get("qc_confidence", "")
                    entry["confidence_score"] = flt(live_data[st].confidence_score)

        return {
            "success": True,
            "status": doc.status,
            "progress": flt(doc.progress),
            "total_files": doc.total_files or 0,
            "success_count": doc.success_count or 0,
            "failed_count": doc.failed_count or 0,
            "skipped_count": doc.skipped_count or 0,
            "log": log_data,
        }
    except Exception as e:
        return {"success": False, "message": str(e)}


@frappe.whitelist()
@require_process("secondary")
def upload_bulk_zip():
    """
    Custom ZIP upload endpoint for portal users who lack desk access.
    Frappe's built-in upload_file restricts non-desk users to a MIME allowlist
    that excludes application/zip, causing a 417.  This endpoint explicitly
    validates that the caller has write permission on Bulk Statement Upload and
    then saves the file with ignore_permissions=True so the ZIP is accepted.
    """
    if frappe.session.user == "Guest":
        frappe.throw("Authentication required", frappe.AuthenticationError)

    frappe.has_permission("Bulk Statement Upload", "write", throw=True)

    files = frappe.request.files
    if "file" not in files:
        frappe.throw("No file provided")

    uploaded = files["file"]
    filename = uploaded.filename or "bulk_upload.zip"
    if not filename.lower().endswith(".zip"):
        frappe.throw("Only ZIP files are accepted here")

    content = uploaded.stream.read()

    file_doc = frappe.get_doc({
        "doctype": "File",
        "file_name": filename,
        "is_private": 0,
        "content": content,
        "folder": "Home",
    })
    file_doc.save(ignore_permissions=True)
    frappe.db.commit()

    return {"file_url": file_doc.file_url}


@frappe.whitelist()
def upload_scheme_attachment():
    """
    Upload a supporting document (order copy / proof) for a Scheme Request.
    Frappe's built-in upload_file restricts non-desk users to a MIME allowlist
    that can exclude PDFs, so portal users upload proof documents through here.
    The File is saved as private and returned; create_scheme_request_v2 later
    links it to the created Scheme Request.
    """
    if frappe.session.user == "Guest":
        frappe.throw("Authentication required", frappe.AuthenticationError)

    files = frappe.request.files
    if "file" not in files:
        frappe.throw("No file provided")

    uploaded = files["file"]
    filename = uploaded.filename or "attachment"
    allowed = (".pdf", ".jpg", ".jpeg", ".png")
    if not filename.lower().endswith(allowed):
        frappe.throw("Only PDF, JPG or PNG files are accepted")

    content = uploaded.stream.read()

    file_doc = frappe.get_doc({
        "doctype": "File",
        "file_name": filename,
        "is_private": 1,
        "content": content,
        "folder": "Home",
    })
    file_doc.save(ignore_permissions=True)
    frappe.db.commit()

    return {"file_url": file_doc.file_url, "file_name": file_doc.name}


def _assert_scheme_attachment_access(sr):
    """Portal-level read check for a scheme's documents.

    Deliberately mirrors how get_scheme_list_portal scopes data (division + mapped
    regions), so ANY role that can see the scheme in the portal — Admin, HO, Regional
    User, Regional Future — can also open its attachments. Plus the requestor, who
    uploaded the file, always gets their own regardless of current scope.
    """
    from scanify.permissions import (is_portal_admin, get_allowed_divisions,
                                     get_allowed_region_codes)
    user = frappe.session.user
    if user == "Guest":
        frappe.throw("Authentication required", frappe.AuthenticationError)

    # Administrator / portal Admin: unrestricted.
    if user == "Administrator" or is_portal_admin(user):
        return

    # The requestor uploaded this document — always allow, even if the scheme now sits
    # outside their current division/region mapping.
    if sr.get("requested_by") == user:
        return

    # Division: check every division the user may use, NOT just the one the switcher is
    # currently on. The download is a separate GET and must not depend on that state.
    division = sr.get("division")
    if division and division not in set(get_allowed_divisions(user) or []):
        frappe.throw("Not permitted to access this scheme's documents", frappe.PermissionError)

    # Region: same rule as the scheme list — None means unrestricted; otherwise the
    # scheme's region must be in the user's mapped regions.
    allowed = get_allowed_region_codes(user, division)
    if allowed is not None and sr.get("region") not in set(allowed):
        frappe.throw("Not permitted to access this scheme's documents", frappe.PermissionError)


@frappe.whitelist(methods=["GET"])
def download_scheme_attachment(scheme_request, idx):
    """Stream a Scheme Request proof attachment, authorised by the PORTAL's own rules.

    Proof files are uploaded BEFORE the scheme exists and are only linked to it
    afterwards. Whenever that link is missing (legacy rows, an interrupted submit),
    Frappe's File.has_permission falls through to its final `return False`, so
    /private/files/<name> 403s for everyone except the uploader and Administrator —
    including a portal Admin holding System Manager. That is why "View/Download" failed
    even though the page itself rendered. Serving through here removes the dependency on
    File.attached_to entirely, and we re-link the File on the way out so Desk and any
    existing direct links start working too."""
    from frappe.utils import cint, get_site_path
    idx = cint(idx)
    if idx not in (1, 2, 3, 4):
        frappe.throw("Invalid attachment number")

    sr = frappe.db.get_value("Scheme Request", scheme_request,
        ["name", "division", "region", "requested_by", "proof_attachment_1",
         "proof_attachment_2", "proof_attachment_3", "proof_attachment_4"], as_dict=True)
    if not sr:
        frappe.throw("Scheme request not found", frappe.DoesNotExistError)

    _assert_scheme_attachment_access(sr)

    url = sr.get("proof_attachment_%d" % idx)
    if not url:
        frappe.throw("No document attached at position %d" % idx)

    filename = os.path.basename(str(url).split("?")[0])
    if not filename or filename in (".", ".."):
        frappe.throw("Invalid attachment")
    sub = "private" if str(url).startswith("/private/") else "public"
    path = get_site_path(sub, "files", filename)
    if not os.path.exists(path):
        frappe.throw("The attached file is missing on the server")

    # Best-effort re-link so Desk / direct links heal too. Never blocks the download.
    try:
        # "is / not set" -> IFNULL(field,'')='' so it matches BOTH NULL and ''.
        # (An IN [None, ''] filter would never match NULL rows.)
        orphan = frappe.db.get_value(
            "File", {"file_url": url, "attached_to_name": ["is", "not set"]}, "name")
        if orphan:
            frappe.db.set_value("File", orphan, {
                "attached_to_doctype": "Scheme Request",
                "attached_to_name": sr.name,
                "attached_to_field": "proof_attachment_%d" % idx,
            }, update_modified=False)
            frappe.db.commit()
    except Exception:
        frappe.clear_last_message()

    with open(path, "rb") as f:
        content = f.read()

    ext = os.path.splitext(filename)[1].lower()
    frappe.local.response.filename = filename
    frappe.local.response.filecontent = content
    frappe.local.response.type = "download"
    # Render images/PDFs in the tab, matching the old direct-link behaviour.
    frappe.local.response.display_content_as = (
        "inline" if ext in (".pdf", ".jpg", ".jpeg", ".png", ".gif") else "attachment")


@frappe.whitelist()
def get_bulk_jobs_list(division=None):
    """
    Return list of bulk OCR jobs for the portal list page.
    """
    try:
        user_division = division
        if not user_division:
            user_division = frappe.db.get_value("User", frappe.session.user, "division") or "Prima"

        filters = {"docstatus": ["in", [0, 1]]}
        if user_division:
            filters["division"] = user_division

        jobs = frappe.get_all(
            "Bulk Statement Upload",
            filters=filters,
            fields=["name", "statement_month", "status", "progress", "total_files",
                    "success_count", "failed_count", "skipped_count", "creation", "modified"],
            order_by="creation desc",
            limit_page_length=100,
        )

        return {"success": True, "jobs": jobs}

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Get Bulk Jobs List Error")
        return {"success": False, "message": str(e)}


@frappe.whitelist()
@require_process("secondary")
def create_bulk_ocr_job(statement_month, zip_file_url, division=None, job_name=None, region=None):
    """
    Create a new Bulk Statement Upload document from portal.
    If `region` is provided, stockist name matching for this job is restricted
    to stockists in that region only (avoids cross-region name collisions, e.g.
    "Vijay Pharma" vs "Vijaya Pharma" in a different region).
    Returns the created doc name.
    """
    try:
        if not division:
            division = frappe.db.get_value("User", frappe.session.user, "division") or "Prima"

        # A non-admin may only restrict to a region they're mapped to. When they pick
        # none, extraction still confines matching to their regions (see
        # _bulk_job_region_scope), so a ZIP can never create statements out of scope.
        _allowed = _allowed_region_codes_or_all(division)
        if region and _allowed is not None and region not in set(_allowed):
            return {"success": False,
                    "message": "You are not permitted to upload for that region."}

        doc = frappe.get_doc({
            "doctype": "Bulk Statement Upload",
            "statement_month": statement_month,
            "zip_file": zip_file_url,
            "division": division,
            "region": region or None,
            "job_name": job_name or "",
            "status": "Pending",
        })
        doc.insert(ignore_permissions=True)
        frappe.db.commit()
        return {"success": True, "docname": doc.name}
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Create Bulk OCR Job Error")
        return {"success": False, "message": str(e)}


def _bulk_job_statement_names(doc):
    """
    Return the list of Stockist Statement names a bulk job created.
    The only link from a job to its statements is the `extraction_log` JSON —
    each successful entry records its `statement` name. There is no Link/parent
    field on Stockist Statement pointing back to the job, which is exactly why
    deleting the job leaves the statements intact.
    """
    names = []
    if doc.extraction_log:
        try:
            for entry in json.loads(doc.extraction_log):
                if isinstance(entry, dict) and entry.get("statement"):
                    names.append(entry["statement"])
        except Exception:
            pass
    return names


def _count_existing_statements(names):
    """How many of the given statement names still exist in the DB."""
    if not names:
        return 0
    return frappe.db.count("Stockist Statement", {"name": ["in", names]})


@frappe.whitelist()
def get_bulk_job_delete_info(docname):
    """
    Info for the delete-confirmation dialog of a bulk OCR job.

    Returns how many statements this job created (per its log) and how many of
    those still exist. Deleting the job does NOT delete any statement — they are
    independent documents — so this lets the UI warn the user with real numbers.
    """
    try:
        if not frappe.db.exists("Bulk Statement Upload", docname):
            return {"success": False, "message": f"Job {docname} does not exist."}
        doc = frappe.get_doc("Bulk Statement Upload", docname)
        names = _bulk_job_statement_names(doc)
        return {
            "success": True,
            "job_name": doc.job_name or doc.name,
            "existing_statements": _count_existing_statements(names),
        }
    except Exception as e:
        return {"success": False, "message": str(e)}


@frappe.whitelist()
@require_process("secondary")
def delete_bulk_job(docname):
    """
    Delete ONLY the Bulk Statement Upload job entry.

    The Stockist Statements the job created are deliberately left untouched: they
    are standalone documents and the job merely references them in its
    extraction_log. This lets a failed/garbage job be cleared and re-uploaded
    without losing any real statement data. Returns the number of statements that
    remain after the job is removed.

    Permissioning matches the portal's other delete flow (delete_stockist_statement):
    the @frappe.whitelist decorator already blocks Guests, and the delete itself runs
    with ignore_permissions because portal roles (e.g. UATadmin/Sales Manager) hold a
    read/write-only DocPerm on this submittable doctype. The action is non-destructive
    to statement data and is recorded to the audit trail below.
    """
    try:
        if not frappe.db.exists("Bulk Statement Upload", docname):
            return {"success": False, "message": f"Job {docname} does not exist."}

        doc = frappe.get_doc("Bulk Statement Upload", docname)
        kept = _count_existing_statements(_bulk_job_statement_names(doc))

        # Audit trail — persists after the job document is deleted.
        frappe.get_doc({
            "doctype": "Comment",
            "comment_type": "Info",
            "reference_doctype": "Bulk Statement Upload",
            "reference_name": docname,
            "content": (
                f"<b>Bulk Job Entry Deleted</b><br>By: {frappe.session.user}<br>"
                f"{kept} extracted statement(s) preserved — only the job record was removed."
            ),
            "comment_email": frappe.session.user,
        }).insert(ignore_permissions=True)

        # Submittable doctype — a submitted job must be cancelled before deletion.
        if doc.docstatus == 1:
            doc.flags.ignore_permissions = True
            doc.cancel()

        frappe.delete_doc("Bulk Statement Upload", docname, ignore_permissions=True, force=True)
        frappe.db.commit()
        return {"success": True, "kept_statements": kept}
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Delete Bulk Job Error")
        return {"success": False, "message": str(e)}


# ========== REST OF THE API FILE (UNCHANGED) ==========
@frappe.whitelist()
@require_process("secondary")
def bulk_extract_statements_async(docname):
    """
    Enqueue bulk extraction as background job
    """
    doc = frappe.get_doc("Bulk Statement Upload", docname)
    
    # Enqueue background job
    job = enqueue(
        method="scanify.api.process_bulk_extraction",
        queue="long",
        timeout=3600,  # 1 hour
        job_name=f"bulk_extract_{docname}",
        docname=docname,
        month=doc.statement_month,
        zip_file_url=doc.zip_file
    )
    
    # Update status
    doc.status = "Queued"
    doc.job_id = job.id
    doc.save(ignore_permissions=True)
    frappe.db.commit()
    
    return {
        "success": True,
        "message": "Bulk extraction job queued successfully",
        "job_id": job.id
    }

def process_bulk_extraction(docname, month, zip_file_url):
    """
    Background job to process bulk extraction
    """
    try:
        doc = frappe.get_doc("Bulk Statement Upload", docname)
        doc.status = "In Progress"
        doc.progress = 0
        doc.save(ignore_permissions=True)
        frappe.db.commit()
        
        # Get ZIP file
        from frappe.utils.file_manager import get_file_path
        file_path = get_file_path(zip_file_url)
        
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {zip_file_url}")
        
        # Extract ZIP
        with tempfile.TemporaryDirectory() as temp_dir:
            with zipfile.ZipFile(file_path, "r") as zip_ref:
                zip_ref.extractall(temp_dir)
            
            # Get all files
            all_files = []
            for root, dirs, files in os.walk(temp_dir):
                for file in files:
                    if file.startswith(".") or file.startswith("__MACOSX") or file == "Thumbs.db":
                        continue
                    
                    file_full_path = os.path.join(root, file)
                    file_ext = os.path.splitext(file)[1].lower()
                    
                    supported_extensions = [".pdf", ".jpg", ".jpeg", ".png", ".csv", ".txt", ".xls", ".xlsx"]
                    if file_ext in supported_extensions:
                        all_files.append((file, file_full_path, file_ext))
            
            doc.total_files = len(all_files)
            doc.save(ignore_permissions=True)
            frappe.db.commit()
            
            # Process each file
            results = []
            success_count = 0
            failed_count = 0
            skipped_count = 0
            
            # Build product catalog once (reuse for all files), scoped to the
            # upload's division so codes reused across divisions can't cross-match.
            product_catalog, products_list = build_product_catalog_for_prompt(doc.division)

            # Initialize Gemini client once for the entire batch
            bulk_api_key, model_name, _ = get_gemini_settings()
            bulk_genai_client = genai_sdk.Client(api_key=bulk_api_key)

            # --- STEP 1: Batch filename -> stockist mapping via Gemini (single call) ---
            all_filenames = [f for f, _, _ in all_files]
            job_region_scope = _bulk_job_region_scope(doc)
            gemini_mapping = {}
            try:
                filters = {"status": "Active"}
                if doc.division:
                    filters["division"] = doc.division
                # Region scoping: the chosen region when set, else the job owner's mapped
                # regions. Prevents cross-region name collisions AND stops a non-admin
                # uploading statements for regions they aren't mapped to.
                if job_region_scope:
                    filters["region"] = (["in", list(job_region_scope)]
                                         if isinstance(job_region_scope, (list, tuple, set))
                                         else job_region_scope)

                stockists_cat = frappe.get_all(
                    "Stockist Master",
                    filters=filters,
                    fields=["name", "stockist_name"],
                    order_by="stockist_name asc"
                )
                catalog_lines = [f"{s['name']}|{s['stockist_name']}" for s in stockists_cat]
                filenames_text = "\n".join([f"{i+1}. {f}" for i, f in enumerate(all_filenames)])
                map_prompt = (
                    "Match pharmaceutical statement filenames to stockist codes.\n\n"
                    "STOCKIST CATALOG (CODE|NAME):\n" + "\n".join(catalog_lines) + "\n\n"
                    "FILENAMES:\n" + filenames_text + "\n\n"
                    "Return JSON object: filename -> CODE (or null). Use exact filenames as keys.\n"
                    "Return ONLY valid JSON."
                )
                map_resp = bulk_genai_client.models.generate_content(
                    model=model_name,
                    contents=map_prompt,
                    config=genai_types.GenerateContentConfig(
                        thinking_config=_thinking_config(model_name)
                    )
                )
                resp_text = map_resp.text.strip()
                if resp_text.startswith("```"):
                    resp_text = resp_text.split("```", 1)[1]
                if resp_text.lower().startswith("json"):
                    resp_text = resp_text[4:]
                if resp_text.endswith("```"):
                    resp_text = resp_text.rsplit("```", 1)[0]
                raw_map = json.loads(resp_text.strip())
                valid_codes = {s["name"] for s in stockists_cat}
                name_by_code = {s["name"]: s["stockist_name"] for s in stockists_cat}
                for fname, code in raw_map.items():
                    # Reject codes Gemini hallucinated whose name shares nothing with the
                    # filename — they fall through to fuzzy, then to "unmatched" (a surfaced
                    # failure the user can reassign) rather than a silent wrong-stockist write.
                    if code and code in valid_codes \
                            and _stockist_name_plausible_for_filename(fname, name_by_code.get(code, "")):
                        gemini_mapping[fname] = code
                    else:
                        gemini_mapping[fname] = None
                frappe.logger().info(f"Gemini batch mapping completed: {len(gemini_mapping)} entries")
            except Exception as map_err:
                frappe.logger().warning(f"Gemini batch mapping failed, using fuzzy fallback: {map_err}")

            for idx, (file, file_full_path, file_ext) in enumerate(all_files, 1):
                try:
                    # Identify stockist - Gemini mapping first, fuzzy fallback
                    stockist_code = gemini_mapping.get(file) if gemini_mapping else None
                    if not stockist_code:
                        stockist_code = identify_stockist_from_filename(
                            file, division=doc.division, region=job_region_scope
                        )
                    
                    if not stockist_code:
                        results.append({
                            "file": file,
                            "status": "Failed",
                            "message": "Could not identify stockist from filename"
                        })
                        failed_count += 1
                        # Update progress + counters so UI reflects this failure
                        doc.progress = (idx / len(all_files)) * 100
                        doc.success_count = success_count
                        doc.failed_count = failed_count
                        doc.skipped_count = skipped_count
                        doc.extraction_log = json.dumps(results)
                        doc.save(ignore_permissions=True)
                        frappe.db.commit()
                        continue
                    
                    # Check if already exists
                    existing = frappe.db.exists("Stockist Statement", {
                        "stockist_code": stockist_code,
                        "statement_month": month
                    })
                    
                    # Resolve stockist name for display
                    stockist_name, stockist_status = frappe.db.get_value(
                        "Stockist Master", stockist_code, ["stockist_name", "status"]
                    ) or (stockist_code, None)
                    stockist_name = stockist_name or stockist_code

                    # Guard: reject inactive stockists
                    if stockist_status and stockist_status != "Active":
                        results.append({
                            "file": file,
                            "status": "Failed",
                            "message": f"Stockist {stockist_name} ({stockist_code}) is inactive. Statement cannot be created for an inactive stockist.",
                            "stockist": stockist_name
                        })
                        failed_count += 1
                        doc.progress = (idx / len(all_files)) * 100
                        doc.success_count = success_count
                        doc.failed_count = failed_count
                        doc.skipped_count = skipped_count
                        doc.extraction_log = json.dumps(results)
                        doc.save(ignore_permissions=True)
                        frappe.db.commit()
                        continue

                    if existing:
                        results.append({
                            "file": file,
                            "status": "Failed",
                            "message": f"Statement already exists for this stockist in this month: {existing}. Duplicate rejected.",
                            "stockist": stockist_name
                        })
                        failed_count += 1
                        # Update progress + counters so UI reflects this failure
                        doc.progress = (idx / len(all_files)) * 100
                        doc.success_count = success_count
                        doc.failed_count = failed_count
                        doc.skipped_count = skipped_count
                        doc.extraction_log = json.dumps(results)
                        doc.save(ignore_permissions=True)
                        frappe.db.commit()
                        continue
                    
                    # Create statement
                    statement_name = f"TEMP-{frappe.generate_hash(length=8)}"
                    
                    # Save file
                    from frappe.utils.file_manager import save_file_on_filesystem
                    file_doc = save_file_to_public(file, file_full_path, "Stockist Statement", statement_name)
                    
                    # Create statement doc
                    statement = frappe.get_doc({
                        "doctype": "Stockist Statement",
                        "stockist_code": stockist_code,
                        "statement_month": month,
                        "uploaded_file": file_doc.file_url,
                        "extracted_data_status": "Pending"
                    })
                    statement.insert(ignore_permissions=True)
                    
                    # Update file attachment
                    file_doc.attached_to_name = statement.name
                    file_doc.save(ignore_permissions=True)
                    
                    # Extract data using enhanced method (reuse already-configured client)
                    extracted_data = call_gemini_extraction_with_catalog(
                        file_full_path,
                        stockist_code,
                        product_catalog,
                        products_list,
                        model_name,
                        bulk_genai_client
                    )
                    
                    if extracted_data and len(extracted_data) > 0:
                        statement_rows, counts = _build_statement_rows(
                            extracted_data,
                            statement_division=doc.division,
                            products_list=products_list,
                            statement_region=statement.region,
                        )
                        _replace_statement_items(statement, statement_rows)
                        statement.confidence_score = _calculate_confidence_score(statement_rows)

                        statement.extracted_data_status = "Completed"
                        statement.extraction_notes = _build_extraction_notes(
                            len(statement_rows),
                            statement.confidence_score,
                            unmapped_count=counts["unmapped_count"],
                            auto_mapped_count=counts["auto_mapped_count"],
                            special_row_count=counts["special_row_count"],
                            skipped_division_count=counts["skipped_division_count"],
                            statement_division=doc.division,
                            skipped_region_count=counts["skipped_region_count"],
                        )
                    else:
                        statement.extracted_data_status = "Failed"
                        statement.extraction_notes = "No data extracted from file"
                    
                    statement.populate_previous_month_closing()
                    statement.calculate_closing_and_totals()
                    statement.calculate_qc_confidence()
                    statement.save(ignore_permissions=True)
                    
                    results.append({
                        "file": file,
                        "status": "Success",
                        "statement": statement.name,
                        "stockist": stockist_name,
                        "items_extracted": len(extracted_data) if extracted_data else 0,
                        "qc_confidence": statement.qc_confidence or "All Matched",
                    })
                    success_count += 1
                    
                except Exception as e:
                    error_msg = str(e)
                    frappe.log_error(
                        f"Error processing {file}: {error_msg}\n{frappe.get_traceback()}",
                        "Bulk Extract File Error"
                    )
                    # stockist_code / stockist_name may not be set if error happened before identification
                    _sc_display = "Unknown"
                    try:
                        _sc_display = stockist_name  # defined after stockist resolution
                    except NameError:
                        try:
                            _sc_display = stockist_code  # fall back to code
                        except NameError:
                            pass
                    results.append({
                        "file": file,
                        "status": "Failed",
                        "message": error_msg,
                        "stockist": _sc_display
                    })
                    failed_count += 1
                
                # Update progress + partial log so UI shows in-flight results
                doc.progress = (idx / len(all_files)) * 100
                doc.success_count = success_count
                doc.failed_count = failed_count
                doc.skipped_count = skipped_count
                doc.extraction_log = json.dumps(results)
                doc.save(ignore_permissions=True)
                frappe.db.commit()
        
        # Final update
        doc.status = "Completed" if failed_count == 0 else "Partially Completed"
        doc.progress = 100
        doc.success_count = success_count
        doc.failed_count = failed_count
        doc.skipped_count = skipped_count
        doc.extraction_log = json.dumps(results, indent=2)
        doc.save(ignore_permissions=True)
        frappe.db.commit()
        
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Bulk Extraction Background Job Error")
        
        try:
            doc = frappe.get_doc("Bulk Statement Upload", docname)
            doc.status = "Failed"
            doc.extraction_log = json.dumps([{"file": "—", "status": "Failed", "message": f"Job failed: {str(e)}"}])
            doc.save(ignore_permissions=True)
            frappe.db.commit()
        except Exception:
            pass

@frappe.whitelist()
@require_process("secondary")
def bulk_extract_statements(month, zip_file_url):
    """
    Bulk extract stock statements from ZIP file
    Creates draft statements for review
    """
    try:
        # Resolve the active division so stockist identification is scoped to it
        _bulk_division = get_user_division()

        # Get ZIP file
        from frappe.utils.file_manager import get_file_path
        file_path = get_file_path(zip_file_url)

        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {zip_file_url}")

        # Create temporary directory
        with tempfile.TemporaryDirectory() as temp_dir:
            # Extract ZIP
            with zipfile.ZipFile(file_path, 'r') as zip_ref:
                zip_ref.extractall(temp_dir)
            
            # Process each file
            results = []
            
            for root, dirs, files in os.walk(temp_dir):
                for file in files:
                    # Skip hidden/system files
                    if file.startswith('.') or file.startswith('__MACOSX') or file == 'Thumbs.db':
                        continue
                    
                    file_full_path = os.path.join(root, file)
                    file_ext = os.path.splitext(file)[1].lower()
                    
                    # Skip unsupported files
                    supported_extensions = ['.pdf', '.jpg', '.jpeg', '.png', '.csv', '.txt', '.xls', '.xlsx']
                    if file_ext not in supported_extensions:
                        results.append({
                            "file": file,
                            "status": "Skipped",
                            "message": f"Unsupported file type: {file_ext}"
                        })
                        continue
                    
                    # Try to identify stockist (scoped to the active division so a
                    # filename only matches stockists in that division)
                    stockist_code = identify_stockist_from_filename(file, division=_bulk_division)

                    if not stockist_code:
                        results.append({
                            "file": file,
                            "status": "Failed",
                            "message": "Could not identify stockist from filename"
                        })
                        continue
                    
                    try:
                        # Check if statement already exists
                        existing = frappe.db.exists("Stockist Statement", {
                            "stockist_code": stockist_code,
                            "statement_month": month
                        })
                        
                        if existing:
                            results.append({
                                "file": file,
                                "status": "Failed",
                                "message": f"Statement already exists for this stockist in this month: {existing}. Duplicate rejected.",
                                "stockist": stockist_code
                            })
                            continue
                        
                        # Create statement document (WITHOUT saving yet - to get name)
                        statement_name = f"TEMP-{frappe.generate_hash(length=8)}"
                        
                        # Copy file to public folder first
                        file_doc = save_file_to_public(file, file_full_path, "Stockist Statement", statement_name)
                        
                        # Now create the actual statement
                        statement = frappe.get_doc({
                            "doctype": "Stockist Statement",
                            "stockist_code": stockist_code,
                            "statement_month": month,
                            "uploaded_file": file_doc.file_url,
                            "extracted_data_status": "Pending"
                        })
                        
                        statement.insert(ignore_permissions=True)
                        
                        # Update file attachment to correct docname
                        file_doc.attached_to_name = statement.name
                        file_doc.save(ignore_permissions=True)
                        
                        # Extract data using the active Gemini extraction path
                        sync_api_key, sync_model_name, _ = get_gemini_settings()
                        sync_genai_client = genai_sdk.Client(api_key=sync_api_key)
                        product_catalog, products_list = build_product_catalog_for_prompt(statement.division)
                        extracted_data = call_gemini_extraction_with_catalog(
                            file_full_path,
                            stockist_code,
                            product_catalog,
                            products_list,
                            sync_model_name,
                            sync_genai_client,
                        )
                        
                        if extracted_data and len(extracted_data) > 0:
                            statement_rows, counts = _build_statement_rows(
                                extracted_data,
                                statement_division=statement.division,
                                products_list=products_list,
                                statement_region=statement.region,
                            )
                            _replace_statement_items(statement, statement_rows)
                            statement.confidence_score = _calculate_confidence_score(statement_rows)
                            statement.extracted_data_status = "Completed"
                            statement.extraction_notes = _build_extraction_notes(
                                len(statement_rows),
                                statement.confidence_score,
                                unmapped_count=counts["unmapped_count"],
                                auto_mapped_count=counts["auto_mapped_count"],
                                special_row_count=counts["special_row_count"],
                                skipped_division_count=counts["skipped_division_count"],
                                statement_division=statement.division,
                                skipped_region_count=counts["skipped_region_count"],
                            )
                        else:
                            statement.extracted_data_status = "Failed"
                            statement.extraction_notes = "No data extracted from file"
                        
                        statement.populate_previous_month_closing()
                        statement.calculate_closing_and_totals()
                        statement.save(ignore_permissions=True)
                        
                        results.append({
                            "file": file,
                            "status": "Success",
                            "statement": statement.name,
                            "stockist": stockist_code,
                            "items_extracted": len(extracted_data) if extracted_data else 0
                        })
                        
                    except Exception as e:
                        error_msg = str(e)
                        frappe.log_error(
                            f"Error processing {file}: {error_msg}\n{frappe.get_traceback()}",
                            "Bulk Extract File Error"
                        )
                        results.append({
                            "file": file,
                            "status": "Failed",
                            "message": error_msg,
                            "stockist": stockist_code
                        })
            
            frappe.db.commit()
            
            success_count = len([r for r in results if r["status"] == "Success"])
            failed_count = len([r for r in results if r["status"] == "Failed"])
            
            return {
                "success": True,
                "total_files": len(results),
                "success_count": success_count,
                "failed_count": failed_count,
                "results": results
            }
            
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Bulk Extraction Error")
        return {
            "success": False,
            "message": str(e)
        }


@frappe.whitelist()
def get_unmatched_filenames_suggestion(zip_file_url):
    """
    Analyze ZIP file and suggest stockist matches for manual correction
    """
    try:
        file_path = frappe.get_site_path('public', zip_file_url.replace('/files/', ''))
        
        if not os.path.exists(file_path):
            return {"success": False, "message": "ZIP file not found"}
        
        suggestions = []
        _division = get_user_division()

        with tempfile.TemporaryDirectory() as temp_dir:
            with zipfile.ZipFile(file_path, 'r') as zip_ref:
                zip_ref.extractall(temp_dir)

            for root, dirs, files in os.walk(temp_dir):
                for file in files:
                    if file.startswith('.') or file.startswith('__'):
                        continue

                    stockist_code = identify_stockist_from_filename(file, division=_division)

                    # Get top 5 candidates (scoped to the active division)
                    name_clean = os.path.splitext(file)[0].upper()
                    _cand_filters = {"status": "Active"}
                    if _division:
                        _cand_filters["division"] = ["in", [_division, "Both"]]
                    stockists = frappe.get_all("Stockist Master",
                        fields=["stockist_code", "stockist_name"],
                        filters=_cand_filters)
                    
                    from difflib import SequenceMatcher
                    candidates = sorted(
                        [(s, SequenceMatcher(None, name_clean, s['stockist_name'].upper()).ratio()) 
                         for s in stockists],
                        key=lambda x: x[1],
                        reverse=True
                    )[:5]
                    
                    suggestions.append({
                        "filename": file,
                        "matched_stockist": stockist_code,
                        "top_candidates": [
                            {
                                "code": s['stockist_code'],
                                "name": s['stockist_name'],
                                "score": round(score * 100, 1)
                            }
                            for s, score in candidates
                        ]
                    })
        
        return {
            "success": True,
            "suggestions": suggestions
        }
    except Exception as e:
        return {"success": False, "message": str(e)}

# Generic distributor/legal-form tokens that carry no identifying signal — shared by
# the fuzzy matcher and the filename↔stockist plausibility guard.
_STOCKIST_STOP_WORDS = frozenset({
    'LLP', 'PVT', 'LTD', 'LIMITED', 'CO', 'COMPANY', 'AND', 'THE', 'A', 'AN',
    'PHARMA', 'PHARMACEUTICAL', 'PHARMACEUTICALS', 'DIST', 'DISTRIBUTOR',
    'DISTRIBUTORS', 'TRADERS', 'ENTERPRISES', 'AGENCY', 'AGENCIES',
})


def _stockist_name_plausible_for_filename(filename, stockist_name):
    """Guardrail against confident-but-wrong filename→stockist matches.

    Both Gemini and fuzzy matching can return a high-confidence stockist whose name
    shares nothing with the filename (e.g. 'Muthu Pharma ...' assigned to 'Ator
    Health'), silently creating the statement under the wrong stockist. A match is
    only accepted when the filename shares a meaningful name token (substring or a
    typo-level near match) with the stockist. Fail-open when there's nothing concrete
    to compare so legitimate sparse filenames aren't rejected.
    """
    if not filename or not stockist_name:
        return False

    name = os.path.splitext(str(filename))[0].upper().replace('-', ' ').replace('_', ' ')
    name = re.sub(
        r'\b(JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC|STATEMENT|STOCK|SALES|REPORT)\b',
        ' ', name, flags=re.IGNORECASE,
    )
    name = re.sub(r'\d+', ' ', name)

    fwords = {w for w in name.split() if len(w) > 2 and w not in _STOCKIST_STOP_WORDS}
    swords = {w for w in str(stockist_name).upper().split()
              if len(w) > 2 and w not in _STOCKIST_STOP_WORDS}
    if not fwords or not swords:
        return True  # nothing meaningful to compare — don't block

    for sw in swords:
        for fw in fwords:
            if sw == fw or sw in fw or fw in sw:
                return True
            if len(sw) >= 4 and len(fw) >= 4 and SequenceMatcher(None, sw, fw).ratio() >= 0.8:
                return True
    return False


def identify_stockist_from_filename(filename, division=None, region=None):
    """
    Identify stockist code from filename using robust fuzzy matching.
    When `division` and/or `region` are provided, the candidate stockist pool is
    narrowed accordingly so a filename only matches stockists in that scope.
    `region` accepts a single region code or a list of codes (e.g. a user's mapped
    regions) —
    this prevents cross-region name collisions (e.g. "Vijay Pharma" being matched
    to "Vijaya Pharma" from a different region).
    """
    import re
    from difflib import SequenceMatcher
    
    # Remove extension and normalize
    name = os.path.splitext(filename)[0]
    name_clean = name.upper().replace('-', ' ').replace('_', ' ').strip()
    
    # Remove common date patterns and keywords
    date_patterns = [
        r'\b\d{1,2}[-/]\d{1,2}[-/]\d{2,4}\b',
        r'\b\d{4}[-/]\d{1,2}[-/]\d{1,2}\b',
        r'\b(JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)\b',
        r'\b(JANUARY|FEBRUARY|MARCH|APRIL|MAY|JUNE|JULY|AUGUST|SEPTEMBER|OCTOBER|NOVEMBER|DECEMBER)\b',
        r'\b20\d{2}\b',  # Years 2000-2099
        r'\bSTATEMENT\b',
        r'\bSTOCK\b',
        r'\bSALES\b',
        r'\bREPORT\b'
    ]
    
    for pattern in date_patterns:
        name_clean = re.sub(pattern, '', name_clean, flags=re.IGNORECASE)
    
    # Clean up extra spaces
    name_clean = ' '.join(name_clean.split()).strip()
    
    if not name_clean or len(name_clean) < 3:
        frappe.log_error(f"Filename too short after cleaning: {filename}", "Stockist ID Failed")
        return None
    
    # Get all active stockists (optionally scoped to a division and/or region)
    stockist_filters = {"status": "Active"}
    if division:
        stockist_filters["division"] = division
    if region:
        # `region` may be a single code or a list of codes (a user's mapped regions).
        stockist_filters["region"] = (["in", list(region)]
                                      if isinstance(region, (list, tuple, set)) else region)
    stockists = frappe.get_all("Stockist Master",
        fields=["name", "stockist_code", "stockist_name", "city"],
        filters=stockist_filters)

    if not stockists:
        frappe.log_error("No active stockists found", "Stockist ID Failed")
        return None

    # Strategy 1: Exact stockist code match.
    # Match on the editable Stockist Code but RETURN the master id (PK) — every
    # downstream step (statement.stockist_code, report joins) links by the id.
    for s in stockists:
        if s['stockist_code'] and s['stockist_code'].upper() in name_clean:
            return s['name']
    
    # Strategy 2: Fuzzy match on stockist name
    best_match = None
    best_score = 0
    
    # Stop words to ignore (shared with the plausibility guard)
    stop_words = _STOCKIST_STOP_WORDS
    
    for s in stockists:
        stockist_name_clean = s['stockist_name'].upper().strip()
        
        # Calculate direct similarity
        similarity = SequenceMatcher(None, name_clean, stockist_name_clean).ratio()
        
        # Word-based matching
        stockist_words = set(stockist_name_clean.split())
        filename_words = set(name_clean.split())
        
        # Filter out stop words and short words
        stockist_words_filtered = {w for w in stockist_words 
                                   if len(w) > 2 and w not in stop_words}
        filename_words_filtered = {w for w in filename_words 
                                  if len(w) > 2 and w not in stop_words}
        
        if not stockist_words_filtered:
            # If no significant words, use all words
            stockist_words_filtered = {w for w in stockist_words if len(w) > 1}
        
        # Calculate word overlap
        common_words = stockist_words_filtered.intersection(filename_words_filtered)
        word_overlap_score = (len(common_words) / len(stockist_words_filtered) 
                             if stockist_words_filtered else 0)
        
        # Check for partial matches (important for names like "Dhanvantri" vs "Dhanvantari")
        partial_match_score = 0
        for s_word in stockist_words_filtered:
            for f_word in filename_words_filtered:
                if len(s_word) >= 4 and len(f_word) >= 4:
                    # Check if one is substring of other
                    if s_word in f_word or f_word in s_word:
                        partial_match_score += 0.3
                    # Check character-level similarity for typos
                    elif SequenceMatcher(None, s_word, f_word).ratio() > 0.8:
                        partial_match_score += 0.2
        
        partial_match_score = min(partial_match_score, 0.5)  # Cap at 0.5
        
        # Weighted combined score
        combined_score = (similarity * 0.4) + (word_overlap_score * 0.4) + (partial_match_score * 0.2)
        
        # Bonus for matching 2+ significant words
        if len(common_words) >= 2:
            combined_score += 0.15
        elif len(common_words) == 1 and len(stockist_words_filtered) == 1:
            # Single unique word match (e.g., "Jyoti")
            combined_score += 0.2
        
        # Bonus for exact word match
        if stockist_words_filtered == filename_words_filtered:
            combined_score += 0.2
        
        if combined_score > best_score:
            best_score = combined_score
            best_match = s
    
    # Strategy 3: City-based matching with additional context
    if best_score < 0.5:
        for s in stockists:
            if s.get('city') and s['city']:
                city_clean = s['city'].upper().strip()
                if len(city_clean) > 3 and city_clean in name_clean:
                    stockist_words = {w.upper() for w in s['stockist_name'].split() 
                                    if len(w) > 3 and w.upper() not in stop_words}
                    filename_words = set(name_clean.split())
                    
                    if stockist_words.intersection(filename_words):
                        city_match_score = 0.55
                        if city_match_score > best_score:
                            best_score = city_match_score
                            best_match = s
    
    # Accept match if confidence is above threshold
    CONFIDENCE_THRESHOLD = 0.40  # Lowered slightly for flexibility
    
    if best_match and best_score >= CONFIDENCE_THRESHOLD \
            and _stockist_name_plausible_for_filename(filename, best_match['stockist_name']):
        frappe.logger().info(
            f"✓ Matched: {filename} -> {best_match['stockist_name']} "
            f"({best_match['stockist_code']}) [Score: {best_score:.2f}]"
        )
        return best_match['name']
    
    # Log failure with top 3 candidates for debugging
    if stockists:
        top_candidates = sorted(
            [(s, SequenceMatcher(None, name_clean, s['stockist_name'].upper()).ratio()) 
             for s in stockists],
            key=lambda x: x[1],
            reverse=True
        )[:3]
        
        candidates_info = "\n".join([
            f"  - {s['stockist_name']} ({s['stockist_code']}): {score:.2f}"
            for s, score in top_candidates
        ])
        
        frappe.log_error(
            f"Could not identify stockist from: {filename}\n"
            f"Clean name: '{name_clean}'\n"
            f"Best match: {best_match['stockist_name'] if best_match else 'None'}\n"
            f"Best score: {best_score:.2f}\n"
            f"Top candidates:\n{candidates_info}",
            "Stockist Identification Failed"
        )
    
    return None

def save_file_to_public(filename, file_path, doctype, docname):
    """Save file to public folder and create File document"""
    try:
        with open(file_path, 'rb') as f:
            content = f.read()
        
        # Ensure docname is valid
        if not docname or not isinstance(docname, (str, int)):
            # Create a temporary name if docname is invalid
            docname = frappe.generate_hash(length=10)
        
        file_doc = frappe.get_doc({
            "doctype": "File",
            "file_name": filename,
            "attached_to_doctype": doctype,
            "attached_to_name": str(docname),  # Ensure it's a string
            "content": content,
            "is_private": 0
        })
        file_doc.save(ignore_permissions=True)
        
        return file_doc
    except Exception as e:
        frappe.log_error(
            f"Error saving file {filename}: {str(e)}\n{frappe.get_traceback()}",
            "File Save Error"
        )
        raise

@frappe.whitelist()
def fetch_previous_month_closing(stockist_code, current_month):
    """Fetch previous month's closing balance to set as opening balance"""
    try:
        from dateutil.relativedelta import relativedelta

        if not stockist_code or not current_month:
            return []

        current_date = frappe.utils.getdate(current_month)
        previous_month = current_date - relativedelta(months=1)
        previous_month_first = get_first_day(previous_month)

        # Find previous month's statement
        prev_statement = frappe.db.get_value("Stockist Statement", {
            "stockist_code": stockist_code,
            "statement_month": previous_month_first,
            "docstatus": 1
        }, "name")

        if not prev_statement:
            frappe.msgprint("No previous month statement found", indicator='orange')
            return []

        # Get items from previous statement
        prev_items = frappe.get_all("Stockist Statement Item",
            filters={"parent": prev_statement},
            fields=["product_code", "product_name", "pack", "closing_qty", "pts", "closing_value"])

        # Items link by the Product Master id; the statement view matches these
        # rows against the display business codes, so remap before returning.
        _apply_product_display_codes(prev_items)

        return prev_items or []

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Fetch Previous Month Error")
        return []

@frappe.whitelist()
def reroute_scheme_request(doc_name, comments):
    try:
        from scanify.permissions import require_manager
        require_manager()
        doc = frappe.get_doc("Scheme Request", doc_name)
        # Access already authorized by require_manager() (portal Admin/HO). Bypass Frappe
        # doctype perms so the save/submit works regardless of the approver's Frappe
        # role/user_type (a Website-User Admin/HO has no Sales Manager submit perm).
        doc.flags.ignore_permissions = True

        if doc.approval_status == "Rerouted":
            frappe.throw("Scheme request already rerouted")

        doc.approval_status = "Rerouted"
        doc.append("approval_log", {
            "approver": frappe.session.user,
            "approval_level": "Manager",
            "action": "Rerouted",
            "action_date": nowdate(),
            "comments": comments or "Rerouted for revision"
        })

        doc.save()
        frappe.db.commit()

        # Auto-notification email removed (2026-07-23): rerouting no longer emails
        # anyone. Approved-scheme mails are sent manually from the "Send Emails"
        # page (/portal/scheme-email).

        return True
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Reroute Scheme Error")
        frappe.throw(str(e))

@frappe.whitelist()
def approve_scheme_request(doc_name, comments):
    """Approve a scheme request"""
    try:
        from scanify.permissions import require_manager
        require_manager()
        doc = frappe.get_doc("Scheme Request", doc_name)
        # Access already authorized by require_manager() (portal Admin/HO). Bypass Frappe
        # doctype perms so the save/submit works regardless of the approver's Frappe
        # role/user_type (a Website-User Admin/HO has no Sales Manager submit perm).
        doc.flags.ignore_permissions = True

        if doc.approval_status == "Approved":
            frappe.throw("Scheme request already approved")

        doc.approval_status = "Approved"
        doc.append("approval_log", {
            "approver": frappe.session.user,
            "approval_level": "Manager",
            "action": "Approved",
            "action_date": nowdate(),
            "comments": comments or "Approved"
        })

        doc.save()
        # Submit the doc (docstatus=1) so it becomes available for scheme deduction
        doc.submit()
        frappe.db.commit()

        # Auto-notification email removed (2026-07-23): approval no longer emails
        # automatically. HO triggers the consolidated approved-scheme mail manually
        # from the "Send Emails" page (/portal/scheme-email).

        return True
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Approve Scheme Error")
        frappe.throw(str(e))

@frappe.whitelist()
def reject_scheme_request(doc_name, comments):
    """Reject a scheme request"""
    try:
        from scanify.permissions import require_manager
        require_manager()
        doc = frappe.get_doc("Scheme Request", doc_name)
        # Access already authorized by require_manager() (portal Admin/HO). Bypass Frappe
        # doctype perms so the save works regardless of the approver's Frappe
        # role/user_type (a Website-User Admin/HO has no Sales Manager write perm).
        doc.flags.ignore_permissions = True

        if doc.approval_status == "Rejected":
            frappe.throw("Scheme request already rejected")

        doc.approval_status = "Rejected"
        doc.append("approval_log", {
            "approver": frappe.session.user,
            "approval_level": "Manager",
            "action": "Rejected",
            "action_date": nowdate(),
            "comments": comments or "Rejected"
        })

        doc.save()
        frappe.db.commit()

        # Auto-notification email removed (2026-07-23): rejection no longer emails
        # the requestor.

        return True
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Reject Scheme Error")
        frappe.throw(str(e))

@frappe.whitelist()
def reopen_scheme_request(doc_name, comments=None):
    """Reopen a Rejected or Rerouted scheme request back to Pending so it can be
    revised and re-submitted for approval. Only these two states are reopenable —
    both are drafts (docstatus 0), so this is a plain status flip with no cancel.
    Approved schemes are submitted (and may have deductions) and are intentionally
    NOT reopenable here. Manager-only, mirroring approve/reject/reroute."""
    try:
        from scanify.permissions import require_manager
        require_manager()
        doc = frappe.get_doc("Scheme Request", doc_name)
        # Access already authorized by require_manager() (portal Admin/HO). Bypass Frappe
        # doctype perms so the save works regardless of the approver's role/user_type.
        doc.flags.ignore_permissions = True

        if doc.approval_status not in ("Rejected", "Rerouted"):
            frappe.throw("Only Rejected or Rerouted scheme requests can be reopened")
        if doc.docstatus == 1:
            frappe.throw("A submitted scheme request cannot be reopened")

        previous = doc.approval_status
        doc.approval_status = "Pending"
        # Fresh decision cycle → allow the new outcome to be emailed again.
        doc.email_sent = 0
        doc.email_sent_on = None
        doc.email_sent_to = None
        doc.append("approval_log", {
            "approver": frappe.session.user,
            "approval_level": "Manager",
            "action": "Reopened",
            "action_date": nowdate(),
            "comments": comments or f"Reopened from {previous} for revision"
        })

        doc.save()
        frappe.db.commit()

        return True
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Reopen Scheme Error")
        frappe.throw(str(e))

# NOTE (2026-07-23): send_scheme_notification() was removed. Approve / reject /
# reroute no longer send any automatic email to the requestor. The only scheme
# email path is the manual, HO-triggered consolidated mail on the "Send Emails"
# page (/portal/scheme-email → send_scheme_emails()), which is unaffected.


# ───────────────────────────────────────────────────────────────
# Approved-Scheme Email — HO manually triggers consolidated mails
# grouped by "To" recipient (CFA), with the RSM note block.
# ───────────────────────────────────────────────────────────────

_SCHEME_EMAIL_RSM_NOTES = [
    "Ensure single supply to the approved counter.",
    "Ensure monthly repeat orders.",
    "Submit proof of supply for any offer above 20%.",
]

# Built-in defaults for the approved-scheme (CFA) mail. Each is used verbatim when
# the matching Scanify Settings field is blank — so a fresh migrate / clear-cache /
# restart reproduces the exact email that was hardcoded before, with zero change.
_SCHEME_EMAIL_DEFAULTS = {
    "subject_template": "Scheme order: {division}/{region} Region/{month}",
    "greeting": "Dear CFA,",
    "intro": "Approved for billing:",
    "rsm_heading": "Dear RSM,",
    "rsm_notes": "\n".join(_SCHEME_EMAIL_RSM_NOTES),
    "signature": "Regards,\nMarketing Admin Team,\nMobile No. 98406 14334",
}


def _scheme_email_cfg():
    """Effective approved-mail config: each Scanify Settings value if set, else the
    built-in default. Never raises — a missing/failed lookup falls back to defaults."""
    fields = {
        "subject_template": "scheme_email_subject_template",
        "greeting": "scheme_email_greeting",
        "intro": "scheme_email_intro",
        "rsm_heading": "scheme_email_rsm_heading",
        "rsm_notes": "scheme_email_rsm_notes",
        "signature": "scheme_email_signature",
    }
    cfg = {}
    for key, fieldname in fields.items():
        val = None
        try:
            val = frappe.db.get_single_value("Scanify Settings", fieldname)
        except Exception:
            val = None
        cfg[key] = val if (val is not None and str(val).strip() != "") else _SCHEME_EMAIL_DEFAULTS[key]
    return cfg


def _signature_html(sig_text):
    """Turn a plain-text signature (one line per row) into <br>-joined HTML, escaping
    each line. Blank falls back to the default signature."""
    text = sig_text if (sig_text and str(sig_text).strip()) else _SCHEME_EMAIL_DEFAULTS["signature"]
    lines = [frappe.utils.escape_html(ln.strip()) for ln in str(text).splitlines() if ln.strip()]
    return "<br>".join(lines)


def _split_emails(raw):
    """Parse a comma/semicolon/newline separated string into a list of valid emails."""
    if not raw:
        return []
    out = []
    for p in re.split(r"[,;\n]+", str(raw)):
        e = frappe.utils.validate_email_address((p or "").strip(), throw=False)
        if e:
            out.append(e)
    return out


def _dedupe_emails(seq):
    """De-duplicate emails case-insensitively, preserving first-seen order."""
    seen, out = set(), []
    for e in seq:
        k = (e or "").strip().lower()
        if k and k not in seen:
            seen.add(k)
            out.append(e.strip())
    return out


def _resolve_scheme_recipients(team, requested_by):
    """To/CC for one scheme = Team Master (to_email/cc_emails) combined with the
    requestor's User profile (scheme_to_email/scheme_cc_emails), de-duplicated."""
    to, cc = [], []
    if team:
        t = frappe.db.get_value("Team Master", team, ["to_email", "cc_emails"], as_dict=True)
        if t:
            to += _split_emails(t.get("to_email"))
            cc += _split_emails(t.get("cc_emails"))
    if requested_by:
        u = frappe.db.get_value("User", requested_by, ["scheme_to_email", "scheme_cc_emails"], as_dict=True)
        if u:
            to += _split_emails(u.get("scheme_to_email"))
            cc += _split_emails(u.get("scheme_cc_emails"))
    return _dedupe_emails(to), _dedupe_emails(cc)


def _scheme_email_month_bounds(month):
    """month='YYYY-MM' (or any date) -> (first_day, last_day, label like 'Jun 26')."""
    from frappe.utils import getdate, get_first_day, get_last_day
    if month and len(str(month)) == 7:
        d = getdate(str(month) + "-01")
    elif month:
        d = getdate(month)
    else:
        d = getdate(nowdate())
    return get_first_day(d), get_last_day(d), d.strftime("%b %y")


@frappe.whitelist()
def get_pending_scheme_emails(division=None, month=None, region=None, mail_type="Approved"):
    """Scheme requests in a month (+ optional region) not yet emailed, for the given
    mail_type: 'Approved' → CFA "approved for billing" mail; 'Rejected'/'Rerouted' →
    a notice back to the requestor. The email_sent flag is shared across types (a
    scheme is only ever in one of these states; reopening resets the flag)."""
    try:
        if mail_type not in ("Approved", "Rejected", "Rerouted"):
            mail_type = "Approved"
        if not division:
            division = get_user_division()
        first_day, last_day, _ = _scheme_email_month_bounds(month)

        conds = ["sr.approval_status = %(status)s", "COALESCE(sr.email_sent,0) = 0",
                 "sr.division = %(division)s",
                 "sr.application_date BETWEEN %(from_date)s AND %(to_date)s"]
        params = {"division": division, "from_date": first_day, "to_date": last_day,
                  "status": mail_type}

        # Region scoping: non-admins (incl. HO) only email schemes in their mapped regions.
        from scanify.permissions import get_allowed_region_codes, clamp_region_codes
        _allowed = get_allowed_region_codes(division=division)
        if _allowed is not None:
            _codes = clamp_region_codes(region, division=division)
            if not _codes:
                conds.append("1=0")
            else:
                conds.append("sr.region IN %(allowed_regions)s")
                params["allowed_regions"] = tuple(_codes)
        elif region:
            conds.append("sr.region = %(region)s")
            params["region"] = region

        rows = frappe.db.sql(f"""
            SELECT sr.name, sr.application_date, sr.division, sr.region, sr.team, sr.hq,
                   sr.doctor_name, sr.stockist_name, sr.requested_by, sr.total_scheme_value,
                   rm.region_name, tm.team_name, hm.hq_name,
                   (SELECT COUNT(*) FROM `tabScheme Request Item` sri WHERE sri.parent = sr.name) AS item_count
            FROM `tabScheme Request` sr
            LEFT JOIN `tabRegion Master` rm ON rm.name = sr.region
            LEFT JOIN `tabTeam Master` tm ON tm.name = sr.team
            LEFT JOIN `tabHQ Master` hm ON hm.name = sr.hq
            WHERE {' AND '.join(conds)}
            ORDER BY rm.region_name, tm.team_name, sr.name
        """, params, as_dict=True)

        for r in rows:
            if mail_type == "Approved":
                to, cc = _resolve_scheme_recipients(r.team, r.requested_by)
            else:
                # Reject/reroute notice goes to the requestor, not the CFA.
                to, cc = _scheme_notice_recipient(r.requested_by), []
            r["to_emails"] = to
            r["cc_emails"] = cc
            r["has_recipient"] = bool(to)
        return {"success": True, "schemes": rows, "count": len(rows), "mail_type": mail_type}
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Get Pending Scheme Emails Error")
        return {"success": False, "message": str(e)}


def _render_scheme_email_html(rows):
    """Render the 'Approved for billing' body: CFA table + RSM notes + signature.
    Styles are inlined for email-client compatibility."""
    esc = frappe.utils.escape_html
    th = "padding:4px 6px;border:1px solid #333;background:#f0f0f0;font-size:11px;text-align:center;"
    td = "padding:4px 6px;border:1px solid #333;font-size:11px;"
    tdc = td + "text-align:center;"
    headers = ["S.No", "Sch. No", "Region", "Team", "HQ", "Doctor Name", "Hospital/Chemist",
               "Stockist", "Product Code", "Pack size", "Order Qty (Strips/Units)", "Free Qty",
               "Spl. Rate in PTS"]
    head = "".join(f'<th style="{th}">{h}</th>' for h in headers)

    body = []
    for i, r in enumerate(rows, start=1):
        free = r["free_qty"] or ""
        spl = f'{flt(r["special_rate"]):g}' if flt(r["special_rate"]) > 0 else ""
        body.append(
            "<tr>"
            f'<td style="{tdc}">{i}</td>'
            f'<td style="{td}">{esc(str(r["sch_no"]))}</td>'
            f'<td style="{td}">{esc(str(r["region"]))}</td>'
            f'<td style="{td}">{esc(str(r["team"]))}</td>'
            f'<td style="{td}">{esc(str(r["hq"]))}</td>'
            f'<td style="{td}">{esc(str(r["doctor"]))}</td>'
            f'<td style="{td}">{esc(str(r["hospital"]))}</td>'
            f'<td style="{td}">{esc(str(r["stockist"]))}</td>'
            f'<td style="{td}">{esc(str(r["product_code"]))}</td>'
            f'<td style="{td}">{esc(str(r["pack"]))}</td>'
            f'<td style="{tdc}">{r["order_qty"]}</td>'
            f'<td style="{tdc}">{free}</td>'
            f'<td style="{tdc}">{esc(str(spl))}</td>'
            "</tr>"
        )
    # Configurable prose (greeting / intro / RSM heading / notes / signature); each
    # falls back to the built-in default when its Scanify Settings field is blank.
    cfg = _scheme_email_cfg()
    note_lines = [ln.strip() for ln in str(cfg["rsm_notes"]).splitlines() if ln.strip()]
    notes = "".join(f"<li>{esc(n)}</li>" for n in note_lines)
    return (
        '<div style="font-family:Arial,sans-serif;font-size:13px;color:#000;">'
        f"<p>{esc(cfg['greeting'])}</p>"
        f"<p><b>{esc(cfg['intro'])}</b></p>"
        '<table style="border-collapse:collapse;border:1px solid #333;">'
        f"<thead><tr>{head}</tr></thead><tbody>{''.join(body)}</tbody></table>"
        f"<br><p>{esc(cfg['rsm_heading'])}</p>"
        f"<ul>{notes}</ul>"
        f"<br><p>{_signature_html(cfg['signature'])}</p>"
        "</div>"
    )


def _build_scheme_email_groups(scheme_names, month=None):
    """Group approved schemes by their combined To-set and build recipients, subject
    and rendered body per group. Returns (groups, unroutable_scheme_names)."""
    if isinstance(scheme_names, str):
        scheme_names = json.loads(scheme_names)

    _, _, month_label = _scheme_email_month_bounds(month)
    groups = {}
    unroutable = []

    for name in (scheme_names or []):
        sr = frappe.db.get_value("Scheme Request", name,
            ["name", "division", "region", "team", "hq", "doctor_name", "hospital_address",
             "stockist_name", "requested_by", "application_date", "approval_status", "email_sent"],
            as_dict=True)
        if not sr:
            continue
        to, cc = _resolve_scheme_recipients(sr.team, sr.requested_by)
        if not to:
            unroutable.append(name)
            continue
        key = tuple(sorted({e.lower() for e in to}))
        g = groups.get(key)
        if not g:
            g = {"to": to, "cc": [], "schemes": []}
            groups[key] = g
        g["cc"] += cc
        g["schemes"].append(sr)

    out = []
    for g in groups.values():
        to_lower = {e.lower() for e in g["to"]}
        g["cc"] = [e for e in _dedupe_emails(g["cc"]) if e.lower() not in to_lower]

        rows, region_names, divisions = [], [], []
        for sr in sorted(g["schemes"], key=lambda s: s["name"]):
            rname = frappe.db.get_value("Region Master", sr.region, "region_name") or sr.region or ""
            tname = frappe.db.get_value("Team Master", sr.team, "team_name") or sr.team or ""
            hname = frappe.db.get_value("HQ Master", sr.hq, "hq_name") or sr.hq or ""
            if rname and rname not in region_names:
                region_names.append(rname)
            if sr.division and sr.division not in divisions:
                divisions.append(sr.division)
            items = frappe.db.sql(
                "SELECT product_code, pack, quantity, free_quantity, special_rate "
                "FROM `tabScheme Request Item` WHERE parent=%s ORDER BY idx", sr.name, as_dict=True)
            for it in items:
                rows.append({
                    "sch_no": sr.name, "region": rname, "team": tname, "hq": hname,
                    "doctor": sr.doctor_name or "", "hospital": sr.hospital_address or "",
                    "stockist": sr.stockist_name or "", "product_code": it.product_code or "",
                    # Order Qty is rounded UP to a whole strip/unit for the mail —
                    # CFA billing must never see fractional order quantities.
                    "pack": it.pack or "", "order_qty": int(math.ceil(flt(it.quantity or 0))),
                    "free_qty": it.free_quantity or 0, "special_rate": it.special_rate or 0,
                })
        division_label = "/".join(divisions) if divisions else (get_user_division() or "")
        region_label = "/".join(region_names) if region_names else ""
        g["rows"] = rows
        subj_tmpl = _scheme_email_cfg()["subject_template"]
        g["subject"] = (subj_tmpl.replace("{division}", division_label)
                                 .replace("{region}", region_label)
                                 .replace("{month}", month_label))
        g["html"] = _render_scheme_email_html(rows)
        out.append(g)
    return out, unroutable


# ───────────────────────────────────────────────────────────────
# Reject / Reroute notices — manually sent to the requestor from the
# same "Send Emails" page, grouped by requestor, with the action reason.
# ───────────────────────────────────────────────────────────────

_SCHEME_NOTICE_CONFIG = {
    "Rejected": {
        "intro": "The following scheme request(s) you raised have been <b>rejected</b>.",
        "closing": "No further action is required unless you wish to raise a fresh request.",
        "accent": "#dc3545",
    },
    "Rerouted": {
        "intro": "The following scheme request(s) you raised have been <b>sent back for revision</b>.",
        "closing": "Please revise the request(s) and re-submit them for approval.",
        "accent": "#fd7e14",
    },
}


def _scheme_notice_recipient(requested_by):
    """To-list for a reject/reroute notice = the requestor's login email."""
    if not requested_by:
        return []
    email = frappe.db.get_value("User", requested_by, "email") or requested_by
    return _split_emails(email)


def _latest_log_comment(scheme_name, action):
    """Most recent approval-log comment for a given action on a scheme (the reason)."""
    row = frappe.db.sql(
        "SELECT comments FROM `tabScheme Approval Log` "
        "WHERE parent=%s AND action=%s ORDER BY idx DESC LIMIT 1",
        (scheme_name, action), as_dict=True)
    return (row[0].comments if row else "") or ""


def _render_scheme_notice_html(rows, mail_type):
    """Render the requestor-facing reject/reroute notice: intro + table + closing +
    signature. Styles inlined for email-client compatibility."""
    cfg = _SCHEME_NOTICE_CONFIG[mail_type]
    esc = frappe.utils.escape_html
    th = "padding:4px 8px;border:1px solid #333;background:#f0f0f0;font-size:12px;text-align:center;"
    td = "padding:4px 8px;border:1px solid #333;font-size:12px;"
    tdc = td + "text-align:center;"
    headers = ["S.No", "Sch. No", "Date", "Doctor", "Stockist", "Items", "Reason"]
    head = "".join(f'<th style="{th}">{h}</th>' for h in headers)
    body = []
    for i, r in enumerate(rows, start=1):
        body.append(
            "<tr>"
            f'<td style="{tdc}">{i}</td>'
            f'<td style="{td}">{esc(str(r["sch_no"]))}</td>'
            f'<td style="{tdc}">{esc(str(r["date"]))}</td>'
            f'<td style="{td}">{esc(str(r["doctor"]))}</td>'
            f'<td style="{td}">{esc(str(r["stockist"]))}</td>'
            f'<td style="{tdc}">{r["item_count"]}</td>'
            f'<td style="{td}">{esc(str(r["reason"]))}</td>'
            "</tr>"
        )
    return (
        '<div style="font-family:Arial,sans-serif;font-size:13px;color:#000;">'
        "<p>Dear Sir/Madam,</p>"
        f'<p style="border-left:4px solid {cfg["accent"]};padding-left:10px;">{cfg["intro"]}</p>'
        '<table style="border-collapse:collapse;border:1px solid #333;">'
        f"<thead><tr>{head}</tr></thead><tbody>{''.join(body)}</tbody></table>"
        f"<br><p>{cfg['closing']}</p>"
        f"<br><p>{_signature_html(_scheme_email_cfg()['signature'])}</p>"
        "</div>"
    )


def _build_scheme_notice_groups(scheme_names, mail_type, month=None):
    """Group Rejected/Rerouted schemes by requestor and build the notice email per
    group. Returns (groups, unroutable_scheme_names). Same group shape as the approved
    builder so preview/send handle both uniformly."""
    if isinstance(scheme_names, str):
        scheme_names = json.loads(scheme_names)

    _, _, month_label = _scheme_email_month_bounds(month)
    groups = {}
    unroutable = []

    for name in (scheme_names or []):
        sr = frappe.db.get_value("Scheme Request", name,
            ["name", "division", "region", "team", "hq", "doctor_name", "stockist_name",
             "requested_by", "application_date", "approval_status", "email_sent"],
            as_dict=True)
        if not sr:
            continue
        to = _scheme_notice_recipient(sr.requested_by)
        if not to:
            unroutable.append(name)
            continue
        key = tuple(sorted({e.lower() for e in to}))
        g = groups.get(key)
        if not g:
            g = {"to": to, "cc": [], "schemes": []}
            groups[key] = g
        g["schemes"].append(sr)

    out = []
    for g in groups.values():
        rows, divisions = [], []
        for sr in sorted(g["schemes"], key=lambda s: s["name"]):
            if sr.division and sr.division not in divisions:
                divisions.append(sr.division)
            item_count = frappe.db.count("Scheme Request Item", {"parent": sr.name})
            rows.append({
                "sch_no": sr.name,
                "date": str(sr.application_date) if sr.application_date else "",
                "doctor": sr.doctor_name or "",
                "stockist": sr.stockist_name or "",
                "item_count": item_count,
                "reason": _latest_log_comment(sr.name, mail_type),
            })
        division_label = "/".join(divisions) if divisions else (get_user_division() or "")
        g["rows"] = rows
        g["subject"] = f"Scheme request {mail_type.lower()} — {division_label} / {month_label}"
        g["html"] = _render_scheme_notice_html(rows, mail_type)
        out.append(g)
    return out, unroutable


def _build_scheme_mail_groups(scheme_names, mail_type, month=None):
    """Dispatch to the approved (CFA) builder or the reject/reroute notice builder."""
    if mail_type in ("Rejected", "Rerouted"):
        return _build_scheme_notice_groups(scheme_names, mail_type, month)
    return _build_scheme_email_groups(scheme_names, month)


@frappe.whitelist()
def preview_scheme_emails(scheme_names, month=None, mail_type="Approved"):
    """Build the consolidated email(s) for the selected schemes WITHOUT sending."""
    try:
        if mail_type not in ("Approved", "Rejected", "Rerouted"):
            mail_type = "Approved"
        groups, unroutable = _build_scheme_mail_groups(scheme_names, mail_type, month)
        return {
            "success": True,
            "groups": [{
                "to": g["to"], "cc": g["cc"], "subject": g["subject"], "html": g["html"],
                "scheme_count": len(g["schemes"]), "row_count": len(g["rows"]),
                "scheme_names": [s["name"] for s in g["schemes"]],
            } for g in groups],
            "unroutable": unroutable,
        }
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Preview Scheme Emails Error")
        return {"success": False, "message": str(e)}


@frappe.whitelist()
def send_scheme_emails(scheme_names, month=None, mail_type="Approved"):
    """Send the consolidated scheme email(s) and stamp each scheme as emailed.
    mail_type 'Approved' → CFA billing mail; 'Rejected'/'Rerouted' → requestor notice.
    Restricted to Sales/System Manager (HO)."""
    try:
        from frappe.utils import now_datetime, cint
        if mail_type not in ("Approved", "Rejected", "Rerouted"):
            mail_type = "Approved"
        roles = frappe.get_roles(frappe.session.user)
        if "Sales Manager" not in roles and "System Manager" not in roles:
            frappe.throw("Not permitted to send scheme emails", frappe.PermissionError)

        groups, unroutable = _build_scheme_mail_groups(scheme_names, mail_type, month)
        now = now_datetime()
        groups_sent = schemes_sent = 0
        errors = []

        for g in groups:
            # Re-verify each scheme is still in the expected state and not already emailed.
            valid = [s for s in g["schemes"]
                     if s.approval_status == mail_type and not cint(s.email_sent)]
            if not valid:
                continue
            try:
                frappe.sendmail(recipients=g["to"], cc=g["cc"], subject=g["subject"],
                                message=g["html"], now=True)
            except Exception as se:
                errors.append(f"{g['subject']}: {se}")
                frappe.log_error(frappe.get_traceback(), "Send Scheme Email Error")
                continue

            recipients_str = "To: " + ", ".join(g["to"])
            if g["cc"]:
                recipients_str += "; CC: " + ", ".join(g["cc"])
            for s in valid:
                frappe.db.set_value("Scheme Request", s["name"], {
                    "email_sent": 1, "email_sent_on": now, "email_sent_to": recipients_str,
                }, update_modified=False)
                schemes_sent += 1
            groups_sent += 1

        frappe.db.commit()
        return {"success": True, "groups_sent": groups_sent, "schemes_sent": schemes_sent,
                "unroutable": unroutable, "errors": errors}
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Send Scheme Emails Error")
        return {"success": False, "message": str(e)}


def _require_scheme_email_manager():
    roles = frappe.get_roles(frappe.session.user)
    if "Sales Manager" not in roles and "System Manager" not in roles:
        frappe.throw("Not permitted to manage scheme email settings", frappe.PermissionError)


@frappe.whitelist()
def get_scheme_email_config():
    """Effective approved-mail config (setting value or built-in default) for the Send
    Emails page editor, plus the defaults so the UI can show placeholders. HO only."""
    _require_scheme_email_manager()
    return {"success": True, "config": _scheme_email_cfg(), "defaults": dict(_SCHEME_EMAIL_DEFAULTS)}


@frappe.whitelist()
def update_scheme_email_config(subject_template=None, greeting=None, intro=None,
                               rsm_heading=None, rsm_notes=None, signature=None):
    """Save approved-mail prose to Scanify Settings. Storing a blank field reverts it to
    the built-in default (the renderer falls back on blank). HO only."""
    _require_scheme_email_manager()
    try:
        mapping = {
            "scheme_email_subject_template": subject_template,
            "scheme_email_greeting": greeting,
            "scheme_email_intro": intro,
            "scheme_email_rsm_heading": rsm_heading,
            "scheme_email_rsm_notes": rsm_notes,
            "scheme_email_signature": signature,
        }
        for fieldname, val in mapping.items():
            frappe.db.set_single_value("Scanify Settings", fieldname, (val or "").strip())
        frappe.db.commit()
        frappe.clear_cache(doctype="Scanify Settings")
        return {"success": True, "config": _scheme_email_cfg()}
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Update Scheme Email Config Error")
        return {"success": False, "message": str(e)}


@frappe.whitelist()
def search_doctors(search_term):
    """Search doctors by name or place"""
    try:
        search_term = f"%{search_term}%"

        doctors = frappe.db.sql("""
            SELECT 
                name,
                doctor_code,
                doctor_name,
                place,
                specialization,
                hospital_address,
                hq,
                team,
                region
            FROM `tabDoctor Master`
            WHERE status = 'Active'
            AND (
                doctor_name LIKE %(search_term)s
                OR place LIKE %(search_term)s
                OR doctor_code LIKE %(search_term)s
            )
            ORDER BY doctor_name
            LIMIT 20
        """, {"search_term": search_term}, as_dict=True)

        return doctors or []
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Doctor Search Error")
        return []

@frappe.whitelist()
def get_stockists_by_team(team):
    """Get all active stockists for a team"""
    try:
        stockists = frappe.get_all("Stockist Master",
            filters={
                "team": team,
                "status": "Active"
            },
            fields=["name", "stockist_code", "stockist_name", "city", "hq"],
            order_by="stockist_name"
        )
        return stockists or []
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Get Stockists Error")
        return []

@frappe.whitelist()
def search_stockists(search_term, team=None):
    """Search stockists by name or city, optionally filter by team"""
    try:
        search_term = f"%{search_term}%"

        conditions = "status = 'Active'"

        if team:
            conditions += f" AND team = '{team}'"

        stockists = frappe.db.sql(f"""
            SELECT 
                name,
                stockist_code,
                stockist_name,
                city,
                hq,
                team,
                region
            FROM `tabStockist Master`
            WHERE {conditions}
            AND (
                stockist_name LIKE %(search_term)s
                OR city LIKE %(search_term)s
                OR stockist_code LIKE %(search_term)s
            )
            ORDER BY stockist_name
            LIMIT 20
        """, {"search_term": search_term}, as_dict=True)

        return stockists or []
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Stockist Search Error")
        return []

@frappe.whitelist()
def get_dashboard_data():
    """Get dashboard KPI data with null safety"""
    try:
        # Total stockists
        total_stockists = frappe.db.count("Stockist Master", {"status": "Active"}) or 0

        # Total schemes this month
        from frappe.utils import get_first_day, get_last_day, today
        first_day = get_first_day(today())
        last_day = get_last_day(today())

        total_schemes = frappe.db.count("Scheme Request", {
            "application_date": ["between", [first_day, last_day]]
        }) or 0

        pending_schemes = frappe.db.count("Scheme Request", {
            "approval_status": "Pending",
            "application_date": ["between", [first_day, last_day]]
        }) or 0

        approved_schemes = frappe.db.count("Scheme Request", {
            "approval_status": "Approved",
            "application_date": ["between", [first_day, last_day]]
        }) or 0

        # Total scheme value this month
        result = frappe.db.sql("""
            SELECT COALESCE(SUM(total_scheme_value), 0) as total
            FROM `tabScheme Request`
            WHERE application_date BETWEEN %s AND %s
            AND approval_status = 'Approved'
        """, (first_day, last_day), as_dict=True)

        total_scheme_value = flt(result[0].total) if result and result[0] else 0

        # Statements processed this month
        statements_processed = frappe.db.count("Stockist Statement", {
            "statement_month": ["between", [first_day, last_day]],
            "docstatus": 1
        }) or 0

        return {
            "total_stockists": total_stockists,
            "total_schemes": total_schemes,
            "pending_schemes": pending_schemes,
            "approved_schemes": approved_schemes,
            "total_scheme_value": total_scheme_value,
            "statements_processed": statements_processed
        }
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Get Dashboard Data Error")
        return {
            "total_stockists": 0,
            "total_schemes": 0,
            "pending_schemes": 0,
            "approved_schemes": 0,
            "total_scheme_value": 0,
            "statements_processed": 0
        }

@frappe.whitelist()
def upload_company_logo():
    """Upload company logo"""
    return {
        "message": "Upload your logo to /public/files/stedman_logo.png"
    }

@frappe.whitelist()
def get_workspace_settings():
    """Get workspace settings including logo"""
    logo_path = frappe.db.get_single_value("Scanify Settings", "company_logo")
    if not logo_path:
        logo_path = "/files/stedman_logo.png"  # default

    return {
        "logo": logo_path,
        "company_name": "Stedman Pharmaceuticals"
    }

@frappe.whitelist()
def get_product_history_for_scheme(product_code, doctor_code=None, hq=None):
    """Get historical data for a product in scheme context"""
    try:
        from frappe.utils import getdate, add_months
        import json

        # The frontend sends the Product Master id; tolerate an editable business
        # code from older clients by resolving it within the user's division.
        product_code = _resolve_product_pk(product_code, get_user_division()) or product_code

        # Get product details
        product = frappe.get_doc("Product Master", product_code)
        
        # Build filters
        filters = {
            "docstatus": ["<", 2],  # Not cancelled
            "approval_status": ["in", ["Approved", "Pending", "Rerouted"]]
        }
        
        if doctor_code:
            filters["doctor_code"] = doctor_code
        
        if hq:
            filters["hq"] = hq
        
        # Get past scheme requests with this product
        schemes = frappe.db.sql("""
            SELECT 
                sr.name,
                sr.application_date,
                sr.doctor_name,
                sr.doctor_code,
                sr.approval_status,
                sri.quantity,
                sri.product_value
            FROM 
                `tabScheme Request` sr
            INNER JOIN 
                `tabScheme Request Item` sri ON sr.name = sri.parent
            WHERE 
                sri.product_code = %(product_code)s
                AND sr.docstatus < 2
                AND sr.application_date >= %(six_months_ago)s
            ORDER BY 
                sr.application_date DESC
            LIMIT 10
        """, {
            "product_code": product_code,
            "six_months_ago": add_months(getdate(), -6)
        }, as_dict=True)
        
        # Calculate aggregates
        total_quantity = 0
        total_value = 0
        total_schemes = len(schemes)
        last_order_date = None
        
        for scheme in schemes:
            total_quantity += flt(scheme.quantity or 0)
            total_value += flt(scheme.product_value or 0)
            if not last_order_date and scheme.application_date:
                last_order_date = scheme.application_date
        
        # Get monthly trend data (last 6 months)
        chart_data = frappe.db.sql("""
            SELECT 
                DATE_FORMAT(sr.application_date, '%%Y-%%m') as month,
                SUM(sri.quantity) as quantity,
                SUM(sri.product_value) as value
            FROM 
                `tabScheme Request` sr
            INNER JOIN 
                `tabScheme Request Item` sri ON sr.name = sri.parent
            WHERE 
                sri.product_code = %(product_code)s
                AND sr.approval_status = 'Approved'
                AND sr.application_date >= %(six_months_ago)s
            GROUP BY 
                DATE_FORMAT(sr.application_date, '%%Y-%%m')
            ORDER BY 
                month DESC
        """, {
            "product_code": product_code,
            "six_months_ago": add_months(getdate(), -6)
        }, as_dict=True)
        
        return {
            "success": True,
            "product_code": product.product_code,
            "product_name": product.product_name,
            "pack": product.pack,
            "pts": product.pts,
            "total_schemes": total_schemes,
            "total_quantity": total_quantity,
            "total_value": total_value,
            "last_order_date": last_order_date.strftime("%Y-%m-%d") if last_order_date else None,
            "recent_schemes": schemes,
            "chart_data": chart_data
        }
        
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Get Product History Error")
        return {
            "success": False,
            "message": str(e)
        }


















@frappe.whitelist()
def get_incentive_calculation_data(filters):
    """
    Fetch incentive calculation data for Prima and Vektra
    Filter by: Month, HQ, Team, Region, Stockist
    """
    try:
        filters = json.loads(filters) if isinstance(filters, str) else filters
        
        # Get stockist statements
        filters_dict = {
            "docstatus": 1  # Submitted only
        }
        
        # Apply month filter
        if filters.get("month"):
            month = frappe.utils.getdate(f"{filters['month']}-01")
            from_date = get_first_day(month)
            to_date = frappe.utils.get_last_day_of_the_month(month)
            filters_dict["statement_month__gte"] = from_date
            filters_dict["statement_month__lte"] = to_date
        
        # Apply HQ filter if selected
        if filters.get("hq"):
            hq_stockists = frappe.get_all(
                "Stockist Master",
                filters={"hq": filters["hq"]},
                fields=["stockist_code"]
            )
            stockist_codes = [s["stockist_code"] for s in hq_stockists]
            filters_dict["stockist_code"] = ["in", stockist_codes]
        
        # Fetch all statements
        statements = frappe.get_all(
            "Stockist Statement",
            filters=filters_dict,
            fields=["name", "stockist_code", "statement_month"]
        )
        
        # Aggregate data
        incentive_data = {}
        
        for stmt in statements:
            stmt_doc = frappe.get_doc("Stockist Statement", stmt["name"])
            stockist = frappe.get_doc("Stockist Master", stmt["stockist_code"])
            
            # Get items grouped by product type (Prima/Vektra)
            for item in stmt_doc.items:
                product = frappe.get_doc("Product Master", item.product_code)
                product_type = product.product_type  # 'Prima' or 'Vektra'
                
                key = f"{stockist_code}_{product_type}_{item.product_code}"
                
                if key not in incentive_data:
                    incentive_data[key] = {
                        "stockist_code": stmt["stockist_code"],
                        "stockist_name": stockist.stockist_name,
                        "hq": stockist.hq,
                        "product_code": product.product_code or item.product_code,
                        "product_name": product.product_name,
                        "product_type": product_type,
                        "total_qty": 0,
                        "total_value": 0,
                        "incentive_amount": 0
                    }
                
                # Calculate values at PTR (Price to Retailer)
                qty = flt(item.sales_qty) + flt(item.free_qty)
                ptr_value = qty * flt(product.ptr or 0)
                
                incentive_data[key]["total_qty"] += qty
                incentive_data[key]["total_value"] += ptr_value
        
        # Calculate incentive based on target and achievement
        result = []
        for key, data in incentive_data.items():
            # Get target from Incentive Master
            incentive_config = frappe.db.get_value(
                "Incentive Master",
                {
                    "product_type": data["product_type"],
                    "status": "Active"
                },
                ["incentive_rate", "min_target"]
            )
            
            if incentive_config:
                incentive_rate, min_target = incentive_config
                incentive_rate = flt(incentive_rate) / 100  # Convert percentage
                
                if data["total_value"] >= flt(min_target):
                    data["incentive_amount"] = data["total_value"] * incentive_rate
            
            result.append(data)
        
        return {
            "success": True,
            "data": result,
            "total_records": len(result)
        }
    
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Incentive Calculation Error")
        return {
            "success": False,
            "message": str(e)
        }

@frappe.whitelist()
def export_incentive_report(filters, format_type="xlsx"):
    """Export incentive report to Excel or PDF"""
    try:
        filters = json.loads(filters) if isinstance(filters, str) else filters
        data = get_incentive_calculation_data(filters)
        
        if not data["success"]:
            frappe.throw(data["message"])
        
        records = data["data"]
        
        if format_type == "xlsx":
            return export_to_excel_incentive(records, filters)
        else:
            return export_to_pdf_incentive(records, filters)
    
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Export Incentive Error")
        frappe.throw(str(e))

def export_to_excel_incentive(records, filters):
    """Generate Excel file for incentive report"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Incentive Report"
    
    # Add header
    ws['A1'] = "Stedman Pharmaceuticals Pvt Ltd"
    ws['A1'].font = Font(bold=True, size=14)
    
    ws['A2'] = f"Incentive Calculation Report"
    ws['A2'].font = Font(bold=True, size=12)
    
    if filters.get("month"):
        ws['A3'] = f"Month: {filters['month']}"
    
    # Column headers
    headers = [
        "Stockist Code", "Stockist Name", "HQ", "Product Code", "Product Name",
        "Product Type", "Total Qty", "Total Value (₹)", "Incentive (₹)"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add data rows
    row = 7
    total_incentive = 0
    
    for record in records:
        ws.cell(row, 1).value = record.get("stockist_code", "")
        ws.cell(row, 2).value = record.get("stockist_name", "")
        ws.cell(row, 3).value = record.get("hq", "")
        ws.cell(row, 4).value = record.get("product_code", "")
        ws.cell(row, 5).value = record.get("product_name", "")
        ws.cell(row, 6).value = record.get("product_type", "")
        ws.cell(row, 7).value = record.get("total_qty", 0)
        ws.cell(row, 8).value = record.get("total_value", 0)
        ws.cell(row, 9).value = record.get("incentive_amount", 0)
        
        # Format currency columns
        ws.cell(row, 8).number_format = '₹ #,##0.00'
        ws.cell(row, 9).number_format = '₹ #,##0.00'
        
        total_incentive += flt(record.get("incentive_amount", 0))
        row += 1
    
    # Add total row
    ws.cell(row, 8).value = "TOTAL INCENTIVE:"
    ws.cell(row, 8).font = Font(bold=True)
    ws.cell(row, 9).value = total_incentive
    ws.cell(row, 9).font = Font(bold=True)
    ws.cell(row, 9).number_format = '₹ #,##0.00'
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    
    # Save to file
    filename = f"Incentive_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = frappe.get_site_path('public', 'exports', filename)
    
    import os
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    wb.save(filepath)
    
    return {
        "success": True,
        "message": "Report generated successfully",
        "file_url": f"/files/{filename}"
    }
@frappe.whitelist()
def get_doctor_history_for_scheme(doctor_code, hq=None):
    """
    Get historical data for a doctor in scheme context
    Shows all past scheme requests for this doctor
    """
    try:
        from frappe.utils import getdate, add_months
        import json
        
        # Get doctor details
        doctor = frappe.get_doc("Doctor Master", doctor_code)
        
        # Build filters
        filters = {
            "docstatus": ("!=", 2),  # Not cancelled
            "doctor_code": doctor_code
        }
        
        if hq:
            filters["hq"] = hq
        
        # Get past scheme requests for this doctor (last 12 months)
        twelve_months_ago = add_months(getdate(), -12)
        schemes = frappe.db.sql("""
            SELECT 
                sr.name,
                sr.application_date,
                sr.stockist_name,
                sr.hq,
                sr.approval_status,
                sr.total_scheme_value,
                COUNT(sri.name) as product_count
            FROM `tabScheme Request` sr
            LEFT JOIN `tabScheme Request Item` sri ON sr.name = sri.parent
            WHERE sr.doctor_code = %(doctor_code)s
                AND sr.docstatus != 2
                AND sr.application_date >= %(twelve_months_ago)s
            GROUP BY sr.name
            ORDER BY sr.application_date DESC
            LIMIT 20
        """, {
            "doctor_code": doctor_code,
            "twelve_months_ago": twelve_months_ago
        }, as_dict=True)
        _apply_product_display_codes(product_summary)
        
        # Calculate aggregates
        total_schemes = len(schemes)
        total_approved = len([s for s in schemes if s.approval_status == "Approved"])
        total_pending = len([s for s in schemes if s.approval_status == "Pending"])
        total_rejected = len([s for s in schemes if s.approval_status == "Rejected"])
        total_value = sum([flt(s.total_scheme_value or 0) for s in schemes])
        
        last_scheme_date = schemes[0].application_date if schemes else None
        
        # Get product-wise breakdown
        product_summary = frappe.db.sql("""
            SELECT 
                sri.product_code,
                sri.product_name,
                SUM(sri.quantity) as total_quantity,
                SUM(sri.free_quantity) as total_free_quantity,
                SUM(sri.product_value) as total_value,
                COUNT(DISTINCT sr.name) as scheme_count
            FROM `tabScheme Request` sr
            INNER JOIN `tabScheme Request Item` sri ON sr.name = sri.parent
            WHERE sr.doctor_code = %(doctor_code)s
                AND sr.approval_status = 'Approved'
                AND sr.application_date >= %(twelve_months_ago)s
            GROUP BY sri.product_code
            ORDER BY total_value DESC
            LIMIT 10
        """, {
            "doctor_code": doctor_code,
            "twelve_months_ago": twelve_months_ago
        }, as_dict=True)
        
        # Get monthly trend data (last 6 months)
        six_months_ago = add_months(getdate(), -6)
        chart_data = frappe.db.sql("""
            SELECT 
                DATE_FORMAT(sr.application_date, '%%Y-%%m') as month,
                COUNT(sr.name) as scheme_count,
                SUM(sr.total_scheme_value) as total_value
            FROM `tabScheme Request` sr
            WHERE sr.doctor_code = %(doctor_code)s
                AND sr.approval_status = 'Approved'
                AND sr.application_date >= %(six_months_ago)s
            GROUP BY DATE_FORMAT(sr.application_date, '%%Y-%%m')
            ORDER BY month DESC
        """, {
            "doctor_code": doctor_code,
            "six_months_ago": six_months_ago
        }, as_dict=True)
        
        return {
            "success": True,
            "doctor_code": doctor.doctor_code,
            "doctor_name": doctor.doctor_name,
            "place": doctor.place or "N/A",
            "specialization": doctor.specialization or "General",
            "hospital_address": doctor.hospital_address or "N/A",
            "hq": doctor.hq or "N/A",
            "total_schemes": total_schemes,
            "total_approved": total_approved,
            "total_pending": total_pending,
            "total_rejected": total_rejected,
            "total_value": total_value,
            "last_scheme_date": last_scheme_date.strftime("%Y-%m-%d") if last_scheme_date else None,
            "recent_schemes": schemes,
            "product_summary": product_summary,
            "chart_data": chart_data
        }
        
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Get Doctor History Error")
        return {"success": False, "message": str(e)}
@frappe.whitelist()
def set_user_division(division):
    """Set user's selected division in session and User document"""
    
    # Validate division
    if division not in ["Prima", "Vektra"]:
        frappe.throw("Invalid division")
    
    try:
        # Store in session
        frappe.session.user_division = division
        
        # Store in User document for persistence
        user_doc = frappe.get_doc("User", frappe.session.user)
        
        # Check if field exists
        if hasattr(user_doc, "division"):
            user_doc.division = division
            user_doc.save(ignore_permissions=True)
            frappe.db.commit()
        else:
            frappe.log_error("Division field not found in User doctype")
        
        return {
            "success": True, 
            "division": division,
            "message": f"Division switched to {division}"
        }
        
    except Exception as e:
        frappe.log_error(f"Could not save division to user: {str(e)}")
        return {
            "success": False,
            "message": f"Error: {str(e)}"
        }


@frappe.whitelist()
def get_user_division():
    """Get user's current division with proper fallback logic"""
    
    # Priority 1: Check session first (for current page load)
    if hasattr(frappe.session, "user_division") and frappe.session.user_division:
        return frappe.session.user_division
    
    # Priority 2: Check User document (persistent storage)
    try:
        user_doc = frappe.get_doc("User", frappe.session.user)
        if hasattr(user_doc, "division") and user_doc.division:
            # Sync to session
            frappe.session.user_division = user_doc.division
            return user_doc.division
    except:
        pass
    
    # Default: Prima
    default_division = "Prima"
    frappe.session.user_division = default_division
    return default_division


@frappe.whitelist()
def get_user_schemes(filters=None):
    """Get schemes for current user based on division and role"""
    user = frappe.session.user
    division = get_user_division(user)
    
    # Build filters
    scheme_filters = {"division": division}
    
    if filters:
        if isinstance(filters, str):
            filters = json.loads(filters)
        
        if filters.get("status"):
            scheme_filters["approval_status"] = filters["status"]
        if filters.get("from_date"):
            scheme_filters["application_date"] = [">=", filters["from_date"]]
        if filters.get("to_date"):
            if "application_date" in scheme_filters:
                scheme_filters["application_date"] = [
                    "between",
                    [filters["from_date"], filters["to_date"]]
                ]
            else:
                scheme_filters["application_date"] = ["<=", filters["to_date"]]
    
    # Get schemes
    schemes = frappe.get_all(
        "Scheme Request",
        filters=scheme_filters,
        fields=[
            "name", "application_date", "doctor_name", "stockist_name",
            "hq", "total_scheme_value", "approval_status"
        ],
        order_by="application_date desc",
        limit=100
    )
    
    return schemes

@frappe.whitelist()
def get_user_hqs(division=None):
    """Get HQs for current user based on division"""
    if not division:
        division = get_user_division()

    filters = {"status": "Active"}
    if division and division != "Both":
        filters["division"] = ["in", [division, "Both"]]

    # Region scoping: non-admins only see HQs within their mapped regions.
    from scanify.permissions import get_allowed_region_codes
    allowed_regions = get_allowed_region_codes(division=division)
    if allowed_regions is not None:
        if not allowed_regions:
            return []
        filters["region"] = ["in", allowed_regions]

    hqs = frappe.get_all(
        "HQ Master",
        filters=filters,
        fields=["name", "hq_name", "team", "region"],
        order_by="hq_name asc",
    )

    # Fetch display names for team and region links
    team_names = {}
    region_names = {}
    for hq in hqs:
        if hq.team and hq.team not in team_names:
            team_names[hq.team] = frappe.db.get_value("Team Master", hq.team, "team_name") or hq.team
        if hq.region and hq.region not in region_names:
            region_names[hq.region] = frappe.db.get_value("Region Master", hq.region, "region_name") or hq.region
        hq["team_name"] = team_names.get(hq.team, hq.team or "")
        hq["region_name"] = region_names.get(hq.region, hq.region or "")
    
    return hqs

@frappe.whitelist()
def get_active_products(division=None):
    """Get all active products for scheme entry"""
    if not division:
        division = get_user_division()
    
    products = frappe.get_all(
        "Product Master",
        filters={"status": "Active"},
        fields=["name", "product_code", "product_name", "pack", "pts", "division"]
    )
    
    # Filter by division if applicable
    if division and division != "Both":
        products = [p for p in products if p.get("division") in (division, "Both")]
    
    return products

@frappe.whitelist()
def get_doctors_for_hq(hq=None, division=None):
    """Get all active doctors for a given HQ and division (for dropdown)"""
    if not division:
        division = get_user_division()

    conditions = ["status = 'Active'"]
    params = {}
    if hq:
        conditions.append("hq = %(hq)s")
        params["hq"] = hq
    if division and division != "Both":
        conditions.append("(division = %(division)s OR division = 'Both')")
        params["division"] = division

    where_sql = " AND ".join(conditions)
    doctors = frappe.db.sql("""
        SELECT name, doctor_code, doctor_name, place, specialization,
               hospital_address, hq, team, region
        FROM `tabDoctor Master`
        WHERE {where_sql}
        ORDER BY doctor_name
        LIMIT 500
    """.format(where_sql=where_sql), params, as_dict=True)

    return doctors

@frappe.whitelist()
def get_approved_doctors_for_hq(hq=None, division=None):
    """Get doctors who have at least one approved scheme in the given HQ/division"""
    if not division:
        division = get_user_division()

    conditions = ["dm.status = 'Active'"]
    params = {}
    if hq:
        conditions.append("sr.hq = %(hq)s")
        params["hq"] = hq
    if division and division != "Both":
        conditions.append("(sr.division = %(division)s OR sr.division = 'Both')")
        params["division"] = division

    where_sql = " AND ".join(conditions)
    doctors = frappe.db.sql("""
        SELECT DISTINCT dm.name, dm.doctor_code, dm.doctor_name, dm.place,
               dm.specialization, dm.hospital_address, dm.hq, dm.team, dm.region
        FROM `tabDoctor Master` dm
        INNER JOIN `tabScheme Request` sr ON sr.doctor_code = dm.name
        WHERE sr.approval_status = 'Approved'
          AND sr.docstatus != 2
          AND {where_sql}
        ORDER BY dm.doctor_name
        LIMIT 500
    """.format(where_sql=where_sql), params, as_dict=True)

    return doctors

@frappe.whitelist()
def get_approved_products_for_doctor(doctor_code=None, division=None):
    """Get products that appear in approved schemes for a specific doctor"""
    if not division:
        division = get_user_division()
    if not doctor_code:
        return []

    conditions = [
        "sr.approval_status = 'Approved'",
        "sr.docstatus != 2",
        "sr.doctor_code = %(doctor_code)s",
    ]
    params = {"doctor_code": doctor_code}
    if division and division != "Both":
        conditions.append("(sr.division = %(division)s OR sr.division = 'Both')")
        params["division"] = division

    where_sql = " AND ".join(conditions)
    products = frappe.db.sql("""
        SELECT DISTINCT pm.name, pm.product_code, pm.product_name, pm.pack, pm.pts, pm.division, pm.sequence
        FROM `tabProduct Master` pm
        INNER JOIN `tabScheme Request Item` sri ON sri.product_code = pm.name
        INNER JOIN `tabScheme Request` sr ON sr.name = sri.parent
        WHERE pm.status = 'Active'
          AND {where_sql}
        ORDER BY COALESCE(pm.sequence, 9999), pm.product_name
    """.format(where_sql=where_sql), params, as_dict=True)

    return products

@frappe.whitelist()
def get_stockists_by_hq(hq, division=None):
    """Get stockists for a specific HQ"""
    if not division:
        division = get_user_division()

    filters = {"hq": hq, "status": "Active"}
    if division and division != "Both":
        filters["division"] = ["in", [division, "Both"]]
    # Non-admins may only list stockists inside their mapped regions.
    _allowed = _allowed_region_codes_or_all(division)
    if _allowed is not None:
        if not _allowed:
            return []
        filters["region"] = ["in", _allowed]

    stockists = frappe.get_all(
        "Stockist Master",
        filters=filters,
        fields=["name", "stockist_code", "stockist_name"]
    )
    
    return stockists

@frappe.whitelist()
def search_doctors(searchterm=None, search_term=None, division=None, hq=None):
    """Search doctors by name or code, optionally filtered by HQ"""
    term = searchterm or search_term or ""
    if not division:
        division = get_user_division()
    
    division_clause = ""
    hq_clause = ""
    params = {"term": f"%{term}%"}
    if division and division != "Both":
        division_clause = "AND (division = %(division)s OR division = 'Both')"
        params["division"] = division
    if hq:
        hq_clause = "AND hq = %(hq)s"
        params["hq"] = hq
    
    doctors = frappe.db.sql("""
        SELECT name, doctor_code, doctor_name, place, specialization, hospital_address, hq, team, region
        FROM `tabDoctor Master`
        WHERE status = 'Active'
        {division_clause}
        {hq_clause}
        AND (doctor_name LIKE %(term)s OR doctor_code LIKE %(term)s OR place LIKE %(term)s)
        ORDER BY doctor_name
        LIMIT 20
    """.format(division_clause=division_clause, hq_clause=hq_clause), params, as_dict=True)
    
    return doctors

@frappe.whitelist()
def create_scheme_request(data):
    """Create a new scheme request from portal"""
    try:
        if isinstance(data, str):
            data = json.loads(data)
        
        # Create scheme document
        doc = frappe.get_doc({
            "doctype": "Scheme Request",
            "application_date": data.get("application_date"),
            "hq": data.get("hq"),
            "doctor_code": data.get("doctor_code"),
            "stockist_code": data.get("stockist_code"),
            "chemist": data.get("chemist"),
            "scheme_notes": data.get("scheme_notes"),
            "requestedby": frappe.session.user,
            "approval_status": "Pending"
        })
        
        # Add items
        for item in data.get("items", []):
            doc.append("items", {
                "product_code": item.get("product_code"),
                "quantity": item.get("quantity"),
                "free_quantity": item.get("free_quantity"),
                "special_rate": item.get("special_rate")
            })
        
        doc.insert(ignore_permissions=False)
        frappe.db.commit()
        
        return {
            "success": True,
            "name": doc.name,
            "message": "Scheme request created successfully"
        }
        
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Create Scheme Request Error")
        return {
            "success": False,
            "message": str(e)
        }
@frappe.whitelist()
def get_master_data(doctype, division=None, status=None):
    """Get records for a master doctype, all filtered by division where applicable."""
    try:
        filters = {}
        # All masters (Region, Team, Zone, State, HQ, Product, Doctor, Stockist) are division-scoped
        if doctype in ["Region Master", "Team Master", "Zone Master", "State Master", "HQ Master", "Product Master", "Doctor Master", "Stockist Master"]:
            if division:
                filters["division"] = ["in", [division, "Both"]]
        meta = frappe.get_meta(doctype)
        # Child-table and layout/display fieldtypes have no column on the parent
        # table, so they can't be SELECTed. "Table MultiSelect" (e.g. Product
        # Master.excluded_regions) is stored in a child table just like "Table".
        excluded_fieldtypes = [
            "Table", "Table MultiSelect", "HTML", "Button", "Column Break",
            "Section Break", "Tab Break", "Heading", "Image"
        ]

        fields = [
            f.fieldname for f in meta.fields
            if f.fieldtype not in excluded_fieldtypes
        ]
        fields = ["name"] + fields
        data = frappe.get_all(
            doctype,
            filters=filters,
            fields=fields,
            order_by="modified desc"
        )

        # Enrich link codes into human-readable labels for the frontend list view table
        for r in data:
            if "hq" in r and r["hq"]:
                r["hq_label"] = frappe.db.get_value("HQ Master", r["hq"], "hq_name") or r["hq"]
            if "team" in r and r["team"]:
                r["team_label"] = frappe.db.get_value("Team Master", r["team"], "team_name") or r["team"]
            if "region" in r and r["region"]:
                r["region_label"] = frappe.db.get_value("Region Master", r["region"], "region_name") or r["region"]
            if "zone" in r and r["zone"]:
                zone_val = frappe.db.get_value("Zone Master", r["zone"], "zone_name")
                r["zone_label"] = zone_val if zone_val else r["zone"]
            if "state" in r and r["state"]:
                state_val = frappe.db.get_value("State Master", r["state"], "state_name")
                r["state_label"] = state_val if state_val else r["state"]
            # Product Master stores excluded regions as codes (for OCR matching) but the
            # list view should show readable names. Resolve each code → region_name.
            if "excluded_region_codes" in r and r["excluded_region_codes"]:
                codes = [c.strip() for c in str(r["excluded_region_codes"]).split(",") if c.strip()]
                names = [
                    frappe.db.get_value("Region Master", c, "region_name") or c
                    for c in codes
                ]
                r["excluded_region_names"] = ", ".join(names)

        return {"success": True, "data": data}
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Get Master Data Error")
        return {"success": False, "message": str(e)}


@frappe.whitelist()
@require_process("masters")
def save_master_record(doctype, name, data):
    """
    Save a master record.

    Region/Team/Zone/State/HQ Masters now use auto-generated codes (R0001, T0001, Z0001, ST0001, HQ0001).
    Division is injected from session when not provided.
    """
    try:
        data = frappe.parse_json(data) if isinstance(data, str) else data

        current_user_division = get_user_division()

        # Inject division for all division-scoped masters
        if doctype in ['HQ Master', 'Product Master', 'Doctor Master', 'Stockist Master', 'Region Master', 'Team Master', 'Zone Master', 'State Master']:
            if not data.get('division'):
                data['division'] = current_user_division

        # Strip auto-generated code fields — backend sets them via before_save hook
        for _dt, _cf in [
            ('HQ Master', 'hq_code'),
            ('Doctor Master', 'doctor_code'),
            ('Region Master', 'region_code'),
            ('Team Master', 'team_code'),
            ('Zone Master', 'zone_code'),
            ('State Master', 'state_code'),
        ]:
            if doctype == _dt:
                data.pop(_cf, None)

        # Stockist Code is system-generated for new records
        if doctype == 'Stockist Master' and not name:
            data.pop('stockist_code', None)

        # sanctioned_strength is auto-computed — never let frontend override it
        if doctype == 'Team Master':
            data.pop('sanctioned_strength', None)

        # Division uses autoname=field:division_name.
        # Accept legacy payloads that may send `name`.
        if doctype == 'Division':
            if data.get('name') and not data.get('division_name'):
                data['division_name'] = data.pop('name')
            if not name and not data.get('division_name'):
                return {'success': False, 'message': "Division Name is required"}

        # -----------------------------
        # CREATE OR LOAD DOC
        # -----------------------------
        if name:
            doc = frappe.get_doc(doctype, name)
        else:
            doc = frappe.new_doc(doctype)

        # -----------------------------
        # RESOLVE LINK FIELDS
        # If the frontend sends a display label instead of the actual doc name, resolve it.
        # For Region and Team, lookup needs to consider division as well.
        # -----------------------------

        meta = frappe.get_meta(doctype)
        for field, value in list(data.items()):
            df = meta.get_field(field)
            if not df or df.fieldtype != 'Link' or not value:
                continue

            linked_doctype = df.options
            
            # Skip if value is already a valid doc name
            if frappe.db.exists(linked_doctype, value):
                continue

            # Try to resolve by _name label field, considering division for composite keys
            label_field = next(
                (f.fieldname for f in frappe.get_meta(linked_doctype).fields
                 if f.fieldname.endswith('_name')),
                None
            )
            
            if label_field:
                lookup_filters = {label_field: value}

                # For division-scoped doctypes, also filter by division for accurate match
                if linked_doctype in ["Region Master", "Team Master", "Zone Master", "State Master", "HQ Master"]:
                    lookup_division = data.get('division') or current_user_division
                    if lookup_division:
                        lookup_filters["division"] = lookup_division

                match = frappe.db.get_value(linked_doctype, lookup_filters, 'name')
                if match:
                    data[field] = match

        # -----------------------------
        # APPLY DATA & SAVE
        # -----------------------------
        # Access is already authorized by @require_process("masters") (portal Admin only)
        # — the portal is the sole entry point and the real security boundary. Save with
        # ignore_permissions so the operation doesn't depend on the acting user's Frappe
        # role/user_type: a portal Admin who is a Website User (no System Manager) would
        # otherwise fail Frappe's doctype permission check ("no doctype access via role").
        doc.update(data)
        doc.save(ignore_permissions=True)
        frappe.db.commit()

        # After saving an HQ, recalculate the linked team's sanctioned strength
        if doctype == 'HQ Master':
            team_name = doc.get('team') or data.get('team')
            if team_name:
                _recalculate_team_sanctioned_strength(team_name)

        return {
            'success': True,
            'message': 'Record saved successfully',
            'name': doc.name
        }

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), 'Save Master Record Error')
        return {'success': False, 'message': str(e)}


@frappe.whitelist()
@require_process("masters")
def delete_master_record(doctype, name):
    """Delete a master record"""
    try:
        # Before deleting an HQ, capture its team so we can recalculate
        team_name = None
        if doctype == 'HQ Master':
            team_name = frappe.db.get_value('HQ Master', name, 'team')

        # Authorized by @require_process("masters") (portal Admin only); delete with
        # ignore_permissions so it works regardless of the admin's Frappe role/user_type.
        frappe.delete_doc(doctype, name, ignore_permissions=True)
        frappe.db.commit()

        # Recalculate team strength after HQ deletion
        if team_name:
            _recalculate_team_sanctioned_strength(team_name)
        
        return {"success": True, "message": "Record deleted successfully"}
    
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Delete Master Record Error")
        return {"success": False, "message": str(e)}


def _recalculate_team_sanctioned_strength(team_name):
    """Sum per_capita across all HQ Masters linked to this team and update sanctioned_strength."""
    try:
        result = frappe.db.sql(
            "SELECT COALESCE(SUM(per_capita), 0) FROM `tabHQ Master` WHERE team = %s",
            team_name
        )
        total = int(result[0][0]) if result else 0
        frappe.db.set_value('Team Master', team_name, 'sanctioned_strength', total)
        frappe.db.commit()
    except Exception:
        frappe.log_error(frappe.get_traceback(), "Recalculate Team Sanctioned Strength Error")


@frappe.whitelist()
@require_process("masters")
def recalculate_team_sanctioned_strength(team_name):
    """Public API to trigger sanctioned_strength recalculation for a team."""
    try:
        _recalculate_team_sanctioned_strength(team_name)
        strength = frappe.db.get_value('Team Master', team_name, 'sanctioned_strength') or 0
        return {"success": True, "sanctioned_strength": strength}
    except Exception as e:
        return {"success": False, "message": str(e)}


@frappe.whitelist()
def get_hq_list(division=None, search=""):
    """Get HQ list filtered by division for dropdown selection in Stockist/Doctor forms."""
    try:
        filters = {"status": "Active"}
        if division:
            filters["division"] = ["in", [division, "Both"]]
        if search:
            filters["hq_name"] = ["like", f"%{search}%"]

        hqs = frappe.get_all(
            "HQ Master",
            filters=filters,
            fields=["name", "hq_name", "team", "region", "zone", "division"],
            order_by="hq_name asc",
            limit=0
        )

        # Resolve zone code (Z0001) to zone_name for display in text fields
        for hq in hqs:
            zone_code = hq.get("zone")
            if zone_code:
                # Check if zone_code looks like a Zone Master auto-code (starts with Z)
                zone_name = frappe.db.get_value("Zone Master", zone_code, "zone_name")
                if zone_name:
                    hq["zone"] = zone_name  # replace code with display name

        return {"success": True, "data": hqs}
    except Exception as e:
        return {"success": False, "message": str(e)}


@frappe.whitelist()
def get_region_list(division=None, search=""):
    """Get Region list filtered by division for dropdown selection.
    Returns auto-code as value, region_name as label.
    Also resolves zone/state codes to display names for the list view.
    """
    try:
        filters = {"status": "Active"}
        if division:
            filters["division"] = ["in", [division, "Both"]]
        if search:
            filters["region_name"] = ["like", f"%{search}%"]

        # Region scoping: non-admins only get their mapped regions in dropdowns.
        from scanify.permissions import get_allowed_region_codes
        allowed_regions = get_allowed_region_codes(division=division)
        if allowed_regions is not None:
            if not allowed_regions:
                return {"success": True, "data": []}
            filters["name"] = ["in", allowed_regions]

        regions = frappe.get_all(
            "Region Master",
            filters=filters,
            fields=["name", "region_name", "region_code", "zone", "state", "division"],
            order_by="region_name asc",
            limit=0
        )

        # Resolve zone/state codes to display names for the list table
        for region in regions:
            if region.get("zone"):
                region["zone_name"] = frappe.db.get_value("Zone Master", region["zone"], "zone_name") or region["zone"]
            else:
                region["zone_name"] = None
            if region.get("state"):
                region["state_name"] = frappe.db.get_value("State Master", region["state"], "state_name") or region["state"]
            else:
                region["state_name"] = None

        return {"success": True, "data": regions}
    except Exception as e:
        return {"success": False, "message": str(e)}


@frappe.whitelist()
def get_team_list(division=None, search=""):
    """Get Team list filtered by division for dropdown selection in HQ form.
    Returns auto-code name as value, team_name as label.
    Also returns the linked region so the HQ form can auto-fill region and zone.
    """
    try:
        filters = {"status": "Active"}
        if division:
            filters["division"] = ["in", [division, "Both"]]
        if search:
            filters["team_name"] = ["like", f"%{search}%"]

        teams = frappe.get_all(
            "Team Master",
            filters=filters,
            fields=["name", "team_name", "region", "division"],
            order_by="team_name asc",
            limit=0
        )

        # Enrich each team with region_name and zone_name (from Region Master / Zone Master)
        for team in teams:
            if team.get("region"):
                region_data = frappe.db.get_value(
                    "Region Master", team["region"],
                    ["region_name", "zone"], as_dict=True
                )
                team["region_name"] = region_data.get("region_name") if region_data else None
                zone_code = region_data.get("zone") if region_data else None
                # Resolve zone code -> zone_name for display in text fields
                team["zone"] = zone_code  # Z0001 — used by zone_select dropdown
                team["zone_name"] = frappe.db.get_value("Zone Master", zone_code, "zone_name") if zone_code else None
            else:
                team["region_name"] = None
                team["zone"] = None
                team["zone_name"] = None

        return {"success": True, "data": teams}
    except Exception as e:
        return {"success": False, "message": str(e)}


@frappe.whitelist()
def get_team_details(team_name):
    """Get Team details including linked region and zone for HQ form auto-population."""
    try:
        if not team_name:
            return {"success": False, "message": "Team name is required"}
        doc = frappe.get_doc("Team Master", team_name)
        region_name = None
        zone_code = None
        zone_name = None
        if doc.region:
            region_data = frappe.db.get_value(
                "Region Master", doc.region,
                ["region_name", "zone"], as_dict=True
            )
            if region_data:
                region_name = region_data.get("region_name")
                zone_code = region_data.get("zone")
                zone_name = frappe.db.get_value("Zone Master", zone_code, "zone_name") if zone_code else None
        return {
            "success": True,
            "data": {
                "team_name": doc.team_name,
                "team_id": doc.name,
                "region": doc.region,           # auto-code: "R0001"
                "region_name": region_name,     # display: "South Region"
                "zone": zone_code,              # auto-code: "Z0001" (for zone_select dropdown)
                "zone_name": zone_name,         # display text for Data zone fields
                "division": doc.division
            }
        }
    except Exception as e:
        return {"success": False, "message": str(e)}


@frappe.whitelist()
def get_hq_details(hq_name):
    """Get HQ details including linked team and region for auto-population in forms."""
    try:
        if not hq_name:
            return {"success": False, "message": "HQ name is required"}
        doc = frappe.get_doc("HQ Master", hq_name)
        # Resolve display labels
        team_label = frappe.db.get_value("Team Master", doc.team, "team_name") if doc.team else None
        region_label = frappe.db.get_value("Region Master", doc.region, "region_name") if doc.region else None
        # Resolve zone code -> zone_name for display in Data fields
        zone_display = doc.zone
        if doc.zone:
            zone_name_val = frappe.db.get_value("Zone Master", doc.zone, "zone_name")
            if zone_name_val:
                zone_display = zone_name_val
        return {
            "success": True,
            "data": {
                "hq_name": doc.hq_name,
                "hq_id": doc.name,
                "team": doc.team,
                "team_label": team_label or doc.team,
                "region": doc.region,
                "region_label": region_label or doc.region,
                "zone": zone_display,  # human-readable zone_name for display
                "division": doc.division
            }
        }
    except Exception as e:
        return {"success": False, "message": str(e)}

@frappe.whitelist()
@require_process("masters")
def portal_link_search(doctype=None, search=""):
    import json

    # 1️⃣ Try normal frappe args (form encoded)
    doctype = doctype or frappe.form_dict.get("doctype")
    search = search or frappe.form_dict.get("search", "")

    # 2️⃣ Try raw JSON body (portal fetch)
    if not doctype and frappe.request.data:
        try:
            raw = frappe.request.data.decode("utf-8")
            data = json.loads(raw)
            doctype = data.get("doctype")
            search = data.get("search", "")
        except:
            pass

    # 3️⃣ Still missing → error
    if not doctype:
        frappe.throw("doctype is required")

    meta = frappe.get_meta(doctype)

    # Find best display field (e.g. region_name, team_name, hq_name)
    display_field = next(
        (f.fieldname for f in meta.fields if f.fieldname.endswith("_name")),
        "name"
    )

    # For all division-scoped masters, filter by the current user's division.
    filters = {display_field: ["like", f"%{search}%"]}
    if doctype in ("Region Master", "Team Master", "Zone Master", "State Master", "HQ Master"):
        try:
            current_division = get_user_division()
            if current_division:
                filters["division"] = ["in", [current_division, "Both"]]
        except Exception:
            pass

    results = frappe.get_all(
        doctype,
        filters=filters,
        fields=["name", display_field],
        limit=15
    )

    # label = human-readable name (region_name, zone_name, etc.), value = auto-code (R0001, Z0001, etc.)
    return [
        {"label": r.get(display_field) or r.name, "value": r.name}
        for r in results
    ]

@frappe.whitelist()
def search_hq_targets(search="", division=None, limit=15):
    """HQ search helper for Sales Target portal entry."""
    search = (search or "").strip()

    division = division or get_user_division()
    limit = int(limit) if str(limit).isdigit() else 15
    limit = min(max(limit, 1), 50)

    filters = {"status": "Active"}
    if division:
        filters["division"] = ["in", [division, "Both"]]

    hq_rows = frappe.get_all(
        "HQ Master",
        filters=filters,
        or_filters=[
            ["name", "like", f"%{search}%"],
            ["hq_name", "like", f"%{search}%"],
        ],
        fields=["name", "hq_name", "team", "region", "zone"],
        order_by="hq_name asc",
        limit_page_length=limit,
    )

    team_ids = list({row.get("team") for row in hq_rows if row.get("team")})
    region_ids = list({row.get("region") for row in hq_rows if row.get("region")})

    team_map = {}
    if team_ids:
        team_map = {
            d.name: d.team_name
            for d in frappe.get_all("Team Master", filters={"name": ["in", team_ids]}, fields=["name", "team_name"])
        }

    region_map = {}
    if region_ids:
        region_map = {
            d.name: d.region_name
            for d in frappe.get_all("Region Master", filters={"name": ["in", region_ids]}, fields=["name", "region_name"])
        }

    for row in hq_rows:
        row["team_name"] = team_map.get(row.get("team"), row.get("team"))
        row["region_name"] = region_map.get(row.get("region"), row.get("region"))

    return hq_rows


@frappe.whitelist()
def get_hq_yearly_target_details(name):
    """Fetch HQ Yearly Target details for portal editing."""
    try:
        doc = frappe.get_doc("HQ Yearly Target", name)
        items = []
        for item in doc.hq_targets:
            hq_name = frappe.db.get_value("HQ Master", item.hq, "hq_name")
            items.append({
                "hq": item.hq,
                "hq_name": hq_name or item.hq,
                "team": item.team,
                "apr": item.apr, "may": item.may, "jun": item.jun,
                "jul": item.jul, "aug": item.aug, "sep": item.sep,
                "oct": item.oct, "nov": item.nov, "dec": item.dec,
                "jan": item.jan, "feb": item.feb, "mar": item.mar,
            })
        return {
            "success": True,
            "doc": {
                "name": doc.name,
                "financial_year": doc.financial_year,
                "start_date": str(doc.start_date),
                "end_date": str(doc.end_date),
                "status": doc.status,
                "docstatus": doc.docstatus,
                "items": items
            }
        }
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Get HQ Yearly Target Details Error")
        return {"success": False, "message": str(e)}


@frappe.whitelist()
@require_process("sales_targets")
def submit_hq_yearly_target_from_portal(name):
    """Submit / Approve an HQ yearly target from the portal."""
    try:
        if not frappe.has_permission("HQ Yearly Target", "submit"):
            return {"success": False, "message": "Not permitted to approve targets"}
        doc = frappe.get_doc("HQ Yearly Target", name)
        if doc.docstatus == 1:
            return {"success": True, "message": "Already approved"}
        doc.submit()
        frappe.db.commit()
        return {"success": True, "message": "Approved successfully"}
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Approve HQ Yearly Target Error")
        return {"success": False, "message": str(e)}


@frappe.whitelist()
@require_process("sales_targets")
def update_hq_yearly_target_from_portal(name, financial_year, start_date, end_date, hq_targets, status="Draft"):
    """Update existing draft HQ Yearly Target from portal."""
    try:
        if not frappe.has_permission("HQ Yearly Target", "write"):
            return {"success": False, "message": "Not permitted"}

        doc = frappe.get_doc("HQ Yearly Target", name)
        if doc.docstatus == 1:
            return {"success": False, "message": "Cannot edit an approved target"}

        if isinstance(hq_targets, str):
            hq_targets = frappe.parse_json(hq_targets)

        doc.financial_year = financial_year
        doc.start_date = start_date
        doc.end_date = end_date
        doc.status = status or "Draft"

        doc.hq_targets = []
        for raw in (hq_targets or []):
            hq = raw.get("hq")
            if not hq:
                continue
            meta = frappe.db.get_value("HQ Master", hq, ["team", "region"], as_dict=True)
            doc.append("hq_targets", {
                "hq": hq,
                "team": meta.get("team") if meta else None,
                "region": meta.get("region") if meta else None,
                "apr": flt(raw.get("apr")),
                "may": flt(raw.get("may")),
                "jun": flt(raw.get("jun")),
                "jul": flt(raw.get("jul")),
                "aug": flt(raw.get("aug")),
                "sep": flt(raw.get("sep")),
                "oct": flt(raw.get("oct")),
                "nov": flt(raw.get("nov")),
                "dec": flt(raw.get("dec")),
                "jan": flt(raw.get("jan")),
                "feb": flt(raw.get("feb")),
                "mar": flt(raw.get("mar")),
            })

        doc.save(ignore_permissions=False, ignore_mandatory=True)
        frappe.db.commit()
        return {"success": True, "message": "Updated successfully", "name": doc.name}
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Update HQ Yearly Target Error")
        return {"success": False, "message": str(e)}


@frappe.whitelist()
@require_process("sales_targets")
def resolve_hq_target_rows_from_file(division=None):
    """Parse uploaded CSV/Excel bulk target file and resolve HQ names to IDs.
    Required columns: HQ Name, Apr, May, Jun, Jul, Aug, Sep, Oct, Nov, Dec, Jan, Feb, Mar
    Returns resolved rows ready to populate the portal form.
    """
    try:
        import pandas as pd
        import io as _io

        uploaded_file = frappe.request.files.get("file")
        if not uploaded_file:
            return {"success": False, "message": "No file uploaded"}

        if not division:
            division = get_user_division()

        filename = uploaded_file.filename or ""
        ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
        content = uploaded_file.read()

        if ext == "csv":
            df = pd.read_csv(_io.BytesIO(content), dtype=str)
        elif ext in ("xlsx", "xls"):
            df = pd.read_excel(_io.BytesIO(content), dtype=str)
        else:
            try:
                df = pd.read_csv(_io.BytesIO(content), dtype=str)
            except Exception:
                df = pd.read_excel(_io.BytesIO(content), dtype=str)

        df.columns = df.columns.str.strip()

        hq_filters = {"status": "Active"}
        if division:
            hq_filters["division"] = ["in", [division, "Both"]]
        all_hqs = frappe.get_all("HQ Master", filters=hq_filters, fields=["name", "hq_name"])
        hq_by_name = {h.hq_name.strip().lower(): h for h in all_hqs}
        hq_by_code = {h.name.strip().lower(): h for h in all_hqs}

        resolved = []
        errors = []
        months = ["apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec", "jan", "feb", "mar"]

        for idx, row in df.iterrows():
            hq_input = str(row.get("HQ Name", "") or "").strip()
            if not hq_input or hq_input.lower() in ("hq name", "hq"):
                continue  # skip header-like rows

            hq_rec = hq_by_name.get(hq_input.lower()) or hq_by_code.get(hq_input.lower())
            if not hq_rec:
                errors.append(f"Row {idx + 2}: HQ '{hq_input}' not found")
                continue

            result_row = {"hq": hq_rec.name, "hq_name": hq_rec.hq_name}
            for m in months:
                col_label = m.capitalize()
                val = row.get(col_label, 0) or row.get(m.upper(), 0) or row.get(m, 0) or 0
                try:
                    result_row[m] = float(val) if str(val).strip() not in ("nan", "None", "") else 0.0
                except (ValueError, TypeError):
                    result_row[m] = 0.0
            resolved.append(result_row)

        return {"success": True, "rows": resolved, "errors": errors}
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Resolve HQ Target Rows Error")
        return {"success": False, "message": str(e)}


@frappe.whitelist()
@require_process("sales_targets")
def create_hq_yearly_target_from_portal(financial_year, start_date, end_date, status="Draft", hq_targets=None):
    """Create HQ Yearly Target with HQ-wise monthly values from portal screen."""
    try:
        if not frappe.has_permission("HQ Yearly Target", "create"):
            frappe.throw(_("Not permitted to create HQ Yearly Target"))

        if isinstance(hq_targets, str):
            hq_targets = frappe.parse_json(hq_targets)

        if not hq_targets or not isinstance(hq_targets, list):
            frappe.throw(_("At least one HQ target row is required"))

        division = get_user_division()
        unique_hqs = list({(row or {}).get("hq") for row in hq_targets if (row or {}).get("hq")})
        if not unique_hqs:
            frappe.throw(_("At least one valid HQ is required"))

        hq_meta = frappe.get_all(
            "HQ Master",
            filters={"name": ["in", unique_hqs]},
            fields=["name", "team", "region", "division"],
            limit_page_length=0,
        )
        hq_map = {row.name: row for row in hq_meta}

        parsed_rows = []
        seen_hq = set()

        for row in hq_targets:
            row = row or {}
            hq = row.get("hq")
            if not hq:
                continue

            if hq in seen_hq:
                frappe.throw(_("Duplicate HQ found: {0}").format(hq))
            seen_hq.add(hq)

            meta = hq_map.get(hq)
            if not meta:
                frappe.throw(_("Invalid HQ selected: {0}").format(hq))

            if division and meta.get("division") not in [division, "Both"]:
                frappe.throw(_("HQ {0} does not belong to selected division {1}").format(hq, division))

            parsed_rows.append({
                "hq": hq,
                "team": meta.get("team"),
                "apr": flt(row.get("apr")),
                "may": flt(row.get("may")),
                "jun": flt(row.get("jun")),
                "jul": flt(row.get("jul")),
                "aug": flt(row.get("aug")),
                "sep": flt(row.get("sep")),
                "oct": flt(row.get("oct")),
                "nov": flt(row.get("nov")),
                "dec": flt(row.get("dec")),
                "jan": flt(row.get("jan")),
                "feb": flt(row.get("feb")),
                "mar": flt(row.get("mar")),
            })

        if not parsed_rows:
            frappe.throw(_("At least one valid HQ target row is required"))

        series = "HQT-.division.-.YYYY.-"
        doc = frappe.get_doc({
            "doctype": "HQ Yearly Target",
            "naming_series": series,
            "division": division,
            "financial_year": financial_year,
            "start_date": start_date,
            "end_date": end_date,
            "status": status or "Draft",
            "hq_targets": parsed_rows,
        })
        doc.insert(ignore_permissions=False, ignore_mandatory=True)
        frappe.db.commit()

        return {
            "success": True,
            "name": doc.name,
            "total_hqs": doc.total_hqs,
            "total_target_amount": doc.total_target_amount,
        }
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Create HQ Yearly Target From Portal Error")
        return {"success": False, "message": str(e)}


@frappe.whitelist()
@require_process("masters")
def import_master_data(doctype, division):
    """Bulk import master data from Excel or CSV file.

    - Resolves link field values (names → codes) within the correct division.
    - Performs upsert: if a record with the same identifying fields + division
      already exists, it updates instead of creating a duplicate.
    - Code fields (hq_code, zone_code, etc.) are auto-generated — never required
      from the user.  Product Code is the only user-entered code.
    - Returns per-row error details so the portal can display them.
    """
    try:
        import pandas as pd

        file = frappe.request.files.get('file')
        if not file:
            return {"success": False, "message": "No file uploaded"}

        filename = file.filename or ""
        ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

        if ext == "csv":
            df = pd.read_csv(file, dtype=str)
        elif ext in ("xlsx", "xls"):
            df = pd.read_excel(file, dtype=str)
        else:
            try:
                import io
                content = file.read()
                if hasattr(file, 'seek'):
                    file.seek(0)
                df = pd.read_csv(io.BytesIO(content), dtype=str)
            except Exception:
                if hasattr(file, 'seek'):
                    file.seek(0)
                df = pd.read_excel(file, dtype=str)

        df.columns = df.columns.str.strip()

        column_mapping = get_column_mapping(doctype)

        imported = 0
        updated = 0
        failed = 0
        errors = []

        for idx, row in df.iterrows():
            row_num = idx + 2  # Excel row (1-indexed header + 1)
            try:
                data = {}
                for excel_col, field_name in column_mapping.items():
                    if excel_col in row and pd.notna(row[excel_col]):
                        val = str(row[excel_col]).strip()
                        if val:
                            data[field_name] = val

                # ---- Inject division ----
                row_division = data.get("division") or division
                if doctype != "Division":
                    data["division"] = row_division

                # ---- Normalize Select field values (case-insensitive match) ----
                _normalize_select_fields(doctype, data)

                # ---- Resolve link-field labels → doc names ----
                _resolve_import_links(doctype, data, row_division)

                # ---- Strip auto-generated code fields (backend creates them) ----
                # Stockist Code is intentionally NOT stripped: it's user-editable and
                # bulk-updatable. before_save still auto-fills it from the id when blank.
                for _dt, _cf in [
                    ("HQ Master", "hq_code"),
                    ("Doctor Master", "doctor_code"),
                    ("Region Master", "region_code"),
                    ("Team Master", "team_code"),
                    ("Zone Master", "zone_code"),
                    ("State Master", "state_code"),
                ]:
                    if doctype == _dt:
                        data.pop(_cf, None)

                # sanctioned_strength is auto-computed
                if doctype == "Team Master":
                    data.pop("sanctioned_strength", None)

                # ---- Upsert ----
                # If an explicit ID (pk) is supplied, update that exact record and never
                # create a new one. The id itself is never written into the doc — autoname
                # owns it on create, and on update it's only used to locate the record.
                pk = data.pop("name", None)
                if pk:
                    if not frappe.db.exists(doctype, pk):
                        raise frappe.ValidationError(
                            f"ID '{pk}' not found — leave the ID column blank to create a new record"
                        )
                    existing = pk
                else:
                    # No id: fall back to identifying-field + division matching
                    existing = _find_existing_master(doctype, data, row_division)

                if existing:
                    doc = frappe.get_doc(doctype, existing)
                    doc.update(data)
                    doc.save(ignore_permissions=True)
                    updated += 1
                else:
                    doc = frappe.get_doc({"doctype": doctype, **data})
                    doc.insert(ignore_permissions=True)
                    imported += 1

                if (imported + updated) % 50 == 0:
                    frappe.db.commit()

            except Exception as e:
                failed += 1
                err_msg = str(e)
                # Make validation errors more readable
                if "already exists in division" in err_msg:
                    errors.append(f"Row {row_num}: {err_msg}")
                else:
                    # Strip HTML tags from frappe error messages
                    import re as _re
                    clean = _re.sub(r"<[^>]+>", "", err_msg).strip()
                    errors.append(f"Row {row_num}: {clean}")

        frappe.db.commit()

        return {
            "success": True,
            "imported": imported,
            "updated": updated,
            "failed": failed,
            "errors": errors[:50]  # Return first 50 errors for visibility
        }

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Import Master Data Error")
        return {"success": False, "message": str(e)}


def _normalize_select_fields(doctype, data):
    """Case-insensitively match imported Select values to their valid options."""
    meta = frappe.get_meta(doctype)
    for field in meta.fields:
        if field.fieldtype != "Select" or not field.options:
            continue
        val = data.get(field.fieldname)
        if not val:
            continue
        valid_options = [o for o in field.options.split("\n") if o]
        val_lower = val.lower()
        for opt in valid_options:
            if opt.lower() == val_lower:
                data[field.fieldname] = opt  # replace with correctly-cased version
                break


def _resolve_import_links(doctype, data, division):
    """Resolve display-name values in import data to actual doc names.

    E.g. Team='Team Alpha' → Team='T0001' (the actual name of that Team Master
    record in the given division).
    """
    # Mapping: data field → (linked doctype, label field)
    link_fields = {
        "team":   ("Team Master",   "team_name"),
        "region": ("Region Master", "region_name"),
        "zone":   ("Zone Master",   "zone_name"),
        "state":  ("State Master",  "state_name"),
        "hq":     ("HQ Master",     "hq_name"),
    }

    for field, (linked_dt, label_field) in link_fields.items():
        val = data.get(field)
        if not val:
            continue

        # Already a valid doc name?
        if frappe.db.exists(linked_dt, val):
            continue

        # Try to resolve by label within the same division
        filters = {label_field: val}
        linked_meta = frappe.get_meta(linked_dt)
        if linked_meta.has_field("division") and division:
            filters["division"] = division

        match = frappe.db.get_value(linked_dt, filters, "name")
        if match:
            data[field] = match


def _find_existing_master(doctype, data, division):
    """Find an existing master record for upsert based on the identifying
    name field + division (+ any extra scope) combination."""
    upsert_keys = {
        "Zone Master":     "zone_name",
        "State Master":    "state_name",
        "Region Master":   "region_name",
        "Team Master":     "team_name",
        "HQ Master":       "hq_name",
        "Product Master":  "product_code",
        "Stockist Master": "stockist_name",
        "Doctor Master":   "doctor_name",
    }

    # Extra identifying fields beyond the name field. Doctors are de-duplicated at the
    # HQ level, not the Region/Division level: the same doctor name under a different HQ
    # is a distinct doctor and must be created, not merged into the existing record.
    # This mirrors Doctor Master.check_duplicate_in_division (name + division + hq).
    upsert_extra_keys = {
        "Doctor Master": ["hq"],
    }

    name_field = upsert_keys.get(doctype)
    if not name_field or not data.get(name_field):
        return None

    filters = {name_field: data[name_field]}

    for extra in upsert_extra_keys.get(doctype, []):
        # Match blanks against blanks (empty string) so a doctor with no HQ upserts
        # consistently rather than colliding with same-name doctors that do have one.
        filters[extra] = data.get(extra) or ""

    # Division-scoped lookup
    meta = frappe.get_meta(doctype)
    if meta.has_field("division") and division:
        filters["division"] = division

    return frappe.db.get_value(doctype, filters, "name")


def get_column_mapping(doctype):
    """Get Excel column to field mapping for each doctype"""
    mappings = {
        "HQ Master": {
            "HQ Name": "hq_name",
            "Team": "team",
            "Region": "region",
            "Zone": "zone",
            "Per Capita": "per_capita",
            "Division": "division",
            "Status": "status"
        },
        "Stockist Master": {
            "ID": "name",
            "Stockist Code": "stockist_code",
            "Stockist Name": "stockist_name",
            "HQ": "hq",
            "Team": "team",
            "Region": "region",
            "Zone": "zone",
            "Address": "address",
            "Contact Person": "contact_person",
            "Phone": "phone",
            "Email": "email",
            "Status": "status"
        },
        "Product Master": {
            # ID is the internal PK — fill it to UPDATE that exact record (e.g. bulk
            # code corrections); leave blank to upsert by Product Code + Division.
            "ID": "name",
            "Product Code": "product_code",
            "Product Name": "product_name",
            "Sequence": "sequence",
            "Product Group": "product_group",
            "Category": "category",
            "Pack": "pack",
            "Pack Conversion": "pack_conversion",
            "PTS": "pts",
            "PTR": "ptr",
            "MRP": "mrp",
            "GST Rate (%)": "gst_rate",
            "Division": "division",
            "Status": "status"
        },
        "Doctor Master": {
            # Doctor Code is auto-generated — skip from import mapping
            "Doctor Name": "doctor_name",
            "Qualification": "qualification",
            "Doctor Category": "doctor_category",
            "Specialization": "specialization",
            "Phone": "phone",
            "Place": "place",
            "Hospital Address": "hospital_address",
            "House Address": "house_address",
            "Division": "division",
            "HQ": "hq",
            "Team": "team",
            "Region": "region",
            "State": "state",
            "Zone": "zone",
            "Chemist Name": "chemist_name",
            "Status": "status"
        },
        "Team Master": {
            "Team Name": "team_name",
            "Region": "region",
            "Division": "division",
            "Status": "status"
            # sanctioned_strength is auto-computed - not in import
            # team_code is auto-generated - not in import
        },
        "Region Master": {
            "Region Name": "region_name",
            "Division": "division",
            "Zone": "zone",
            "State": "state",
            "Status": "status"
            # region_code is auto-generated - not in import
        },
        "Zone Master": {
            "Zone Name": "zone_name",
            "Division": "division",
            "Status": "status"
            # zone_code is auto-generated - not in import
        },
        "State Master": {
            "State Name": "state_name",
            "Division": "division",
            "Status": "status"
            # state_code is auto-generated - not in import
        }
    }
    return mappings.get(doctype, {})


@frappe.whitelist()
def get_zone_list(division=None, search=""):
    """Get all Zone Master records for dropdown population."""
    try:
        filters = {"status": "Active"}
        if division:
            filters["division"] = ["in", [division, "Both"]]
        if search:
            filters["zone_name"] = ["like", f"%{search}%"]
        zones = frappe.get_all(
            "Zone Master",
            filters=filters,
            fields=["name", "zone_name"],
            order_by="zone_name asc",
            limit=0
        )
        return {"success": True, "data": zones}
    except Exception as e:
        return {"success": False, "message": str(e)}


@frappe.whitelist()
def get_state_list(division=None, search=""):
    """Get all State Master records for dropdown population."""
    try:
        filters = {"status": "Active"}
        if division:
            filters["division"] = ["in", [division, "Both"]]
        if search:
            filters["state_name"] = ["like", f"%{search}%"]
        states = frappe.get_all(
            "State Master",
            filters=filters,
            fields=["name", "state_name"],
            order_by="state_name asc",
            limit=0
        )
        return {"success": True, "data": states}
    except Exception as e:
        return {"success": False, "message": str(e)}


def get_code_field(doctype):
    """Legacy helper — kept for backward compatibility but no longer used by import."""
    code_fields = {
        "HQ Master": "hq_name",
        "Team Master": "team_name",
    }
    return code_fields.get(doctype)

@frappe.whitelist()
def searchstockists(searchterm=None, division=None, limit=20):
    searchterm = (searchterm or "").strip()

    limit = int(limit) if str(limit).isdigit() else 20
    limit = min(max(limit, 1), 50)

    # Basic filters
    filters = {"status": "Active"}
    if division:
        filters["division"] = ["in", [division, "Both"]]

    # Non-admins may only pick stockists inside their mapped regions.
    _allowed = _allowed_region_codes_or_all(division)
    if _allowed is not None:
        if not _allowed:
            return []
        filters["region"] = ["in", _allowed]

    # Match by code or name
    # NOTE: adapt fieldnames if your doctype uses stockist_code/stockist_name exactly
    stockists = frappe.get_all(
        "Stockist Master",
        filters=filters,
        fields=["stockist_code", "stockist_name", "hq", "division"],
        or_filters=[
            ["stockist_code", "like", f"%{searchterm}%"],
            ["stockist_name", "like", f"%{searchterm}%"],
        ],
        limit_page_length=limit,
        order_by="stockist_name asc"
    )

    # Pre-build code → name maps so we send human-readable values to the UI
    hq_codes = {st.get("hq") for st in stockists if st.get("hq")}
    hq_info = {}
    if hq_codes:
        for h in frappe.db.sql(
            f"""SELECT name, hq_name, IFNULL(team, '') AS team,
                       IFNULL(region, '') AS region, IFNULL(zone, '') AS zone
                  FROM `tabHQ Master`
                 WHERE name IN ({", ".join(["%s"] * len(hq_codes))})""",
            tuple(hq_codes), as_dict=True,
        ):
            hq_info[h.name] = h

    team_codes = {h.team for h in hq_info.values() if h.team}
    region_codes = {h.region for h in hq_info.values() if h.region}
    zone_codes = {h.zone for h in hq_info.values() if h.zone}

    team_names = dict(frappe.db.sql(
        f"SELECT name, team_name FROM `tabTeam Master` WHERE name IN ({', '.join(['%s']*len(team_codes))})",
        tuple(team_codes))) if team_codes else {}
    region_names = dict(frappe.db.sql(
        f"SELECT name, region_name FROM `tabRegion Master` WHERE name IN ({', '.join(['%s']*len(region_codes))})",
        tuple(region_codes))) if region_codes else {}
    zone_names = dict(frappe.db.sql(
        f"SELECT name, zone_name FROM `tabZone Master` WHERE name IN ({', '.join(['%s']*len(zone_codes))})",
        tuple(zone_codes))) if zone_codes else {}

    for st in stockists:
        h = hq_info.get(st.get("hq")) if st.get("hq") else None
        team_code = h.team if h else None
        region_code = h.region if h else None
        zone_code = h.zone if h else None
        st["team"] = team_code
        st["region"] = region_code
        st["zone"] = zone_code
        st["hq_name"] = (h.hq_name if h else None) or st.get("hq")
        st["team_name"] = team_names.get(team_code) or team_code
        st["region_name"] = region_names.get(region_code) or region_code
        st["zone_name"] = zone_names.get(zone_code) or zone_code

    return stockists

@frappe.whitelist()
def get_scheme_detail(scheme_name):
    """Get full scheme request details for portal view"""
    try:
        doc = frappe.get_doc("Scheme Request", scheme_name)
        # Items link by the Product Master id; show the business code. Product
        # history lookups resolve either form, so display-only replacement is safe.
        item_code_map = get_product_code_map([i.product_code for i in doc.items])
        items = []
        for item in doc.items:
            items.append({
                # product_code = business/display code; product_pk = Product Master id
                # (the Link value the edit modal needs to round-trip and match).
                "product_code": item_code_map.get(item.product_code, item.product_code),
                "product_pk": item.product_code,
                "product_name": item.product_name,
                "pack": item.pack,
                "quantity": flt(item.quantity),
                "free_quantity": flt(item.free_quantity),
                "product_rate": flt(item.product_rate),
                "special_rate": flt(item.special_rate),
                "scheme_percentage": flt(item.scheme_percentage),
                "product_value": flt(item.product_value),
            })
        logs = []
        for log in (doc.approval_log or []):
            logs.append({
                "approver": log.approver,
                "action": log.action,
                "action_date": str(log.action_date) if log.action_date else "",
                "comments": log.comments,
                "approval_level": log.approval_level,
            })
        return {
            "success": True,
            "name": doc.name,
            "application_date": str(doc.application_date) if doc.application_date else "",
            "doctor_code": doc.doctor_code,
            "doctor_name": doc.doctor_name,
            "doctor_place": doc.doctor_place,
            "specialization": doc.specialization,
            "hospital_address": doc.hospital_address,
            "hq": doc.hq,
            "hq_name": frappe.db.get_value("HQ Master", doc.hq, "hq_name") if doc.hq else "",
            "region": doc.region,
            "team": doc.team,
            "stockist_code": doc.stockist_code,
            # Human-facing code for display; stockist_code stays the internal id so the
            # deduction screen's statement lookup keeps working.
            "stockist_display_code": get_stockist_code_map([doc.stockist_code]).get(doc.stockist_code, doc.stockist_code),
            "stockist_name": doc.stockist_name,
            "approval_status": doc.approval_status,
            "repeated_request": int(doc.repeated_request or 0),
            "total_scheme_value": flt(doc.total_scheme_value),
            "scheme_notes": doc.scheme_notes,
            "division": doc.division,
            "requested_by": doc.requested_by,
            "proof_attachment_1": doc.proof_attachment_1,
            "proof_attachment_2": doc.proof_attachment_2,
            "proof_attachment_3": doc.proof_attachment_3,
            "proof_attachment_4": doc.proof_attachment_4,
            "items": items,
            "approval_log": logs,
            "docstatus": doc.docstatus,
        }
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Get Scheme Detail Error")
        return {"success": False, "message": str(e)}


@frappe.whitelist()
def update_scheme_request_items(scheme_request, items):
    """Replace the product rows of a Pending scheme request from the portal edit
    modal.

    Guardrails:
      - Only the original requester or a manager may edit.
      - Editable only while Pending (docstatus 0). Approved/submitted schemes are
        locked.
      - Repeat schemes stay restricted to the doctor's approved products.
      - Free Qty and Special Price remain mutually exclusive per line.
    The doctype's before_save/validate recomputes scheme %, product value,
    total_scheme_value and re-checks the monthly per-product limit.
    """
    try:
        if isinstance(items, str):
            items = json.loads(items)

        doc = frappe.get_doc("Scheme Request", scheme_request)

        # Permission — requester or manager
        roles = frappe.get_roles(frappe.session.user)
        is_manager = ("System Manager" in roles) or ("Sales Manager" in roles)
        if not is_manager and doc.requested_by != frappe.session.user:
            return {"success": False, "message": "You are not allowed to edit this scheme request."}

        # Only Pending drafts are editable
        if doc.approval_status != "Pending" or doc.docstatus == 1:
            return {"success": False, "message": "Only Pending (unsubmitted) scheme requests can be edited."}

        if not items:
            return {"success": False, "message": "At least one product row is required."}

        # Repeat schemes: only the doctor's approved products are allowed
        if int(doc.repeated_request or 0):
            approved_pks = {p["name"] for p in get_approved_products_for_doctor(doc.doctor_code, doc.division)}
            offenders = [
                (it.get("product_name") or it.get("product_code"))
                for it in items if it.get("product_code") not in approved_pks
            ]
            if offenders:
                return {"success": False,
                        "message": "Repeat schemes may only use approved products. Not approved: " + ", ".join(offenders)}

        # Free Qty vs Special Price exclusivity (doctype re-validates on save)
        conflicts = [
            (it.get("product_name") or it.get("product_code"))
            for it in items
            if flt(it.get("free_quantity")) > 0 and flt(it.get("special_rate")) > 0
        ]
        if conflicts:
            return {"success": False,
                    "message": "Free Quantity and Special Price cannot both be set for: " + ", ".join(conflicts)}

        # Validate each row has a product and a positive order qty
        for it in items:
            if not it.get("product_code"):
                return {"success": False, "message": "Every row must have a product selected."}
            if flt(it.get("quantity")) <= 0:
                return {"success": False, "message": "Order quantity must be greater than 0 for all products."}

        # Replace the child rows; the controller recomputes derived fields on save
        doc.set("items", [])
        for it in items:
            doc.append("items", {
                "product_code": it.get("product_code"),
                "product_name": it.get("product_name"),
                "pack": it.get("pack"),
                "quantity": flt(it.get("quantity")),
                "free_quantity": flt(it.get("free_quantity")),
                "product_rate": flt(it.get("product_rate")),
                "special_rate": flt(it.get("special_rate")),
            })
        doc.save()
        frappe.db.commit()

        return {
            "success": True,
            "name": doc.name,
            "total_scheme_value": flt(doc.total_scheme_value),
            "message": "Scheme products updated successfully",
        }
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Update Scheme Request Items Error")
        return {"success": False, "message": str(e)}


@frappe.whitelist()
def get_scheme_requests_for_deduction(division=None, search=""):
    """Get approved+submitted scheme requests for deduction selection"""
    try:
        if not division:
            division = get_user_division()
        search = (search or "").strip()
        rows = frappe.db.sql("""
            SELECT
                sr.name,
                sr.application_date,
                sr.doctor_name,
                sr.doctor_code,
                sr.stockist_code,
                sr.stockist_name,
                sr.hq,
                sr.total_scheme_value,
                sr.approval_status,
                sr.division
            FROM `tabScheme Request` sr
            WHERE sr.docstatus = 1
              AND sr.division = %(division)s
              AND (
                sr.name LIKE %(search)s
                OR sr.doctor_name LIKE %(search)s
                OR sr.stockist_name LIKE %(search)s
              )
            ORDER BY sr.application_date DESC
            LIMIT 50
        """, {"division": division, "search": f"%{search}%"}, as_dict=True)
        return {"success": True, "data": rows}
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Get Scheme Requests For Deduction Error")
        return {"success": False, "message": str(e)}


@frappe.whitelist()
def get_stockist_statements_for_deduction(stockist_code, division=None):
    """Get stockist statements for a stockist for deduction"""
    try:
        if not division:
            division = get_user_division()
        statements = frappe.db.sql("""
            SELECT
                ss.name,
                DATE_FORMAT(ss.statement_month, '%%b-%%Y') as month_label,
                ss.statement_month,
                sm.stockist_name
            FROM `tabStockist Statement` ss
            LEFT JOIN `tabStockist Master` sm ON ss.stockist_code = sm.name
            WHERE ss.stockist_code = %(stockist_code)s
              AND ss.docstatus != 2
              AND (
                ss.division IS NULL
                OR ss.division = %(division)s
                OR ss.division = 'Both'
              )
            ORDER BY ss.statement_month DESC
            LIMIT 24
        """, {"stockist_code": stockist_code, "division": division}, as_dict=True)
        return {"success": True, "data": statements}
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Get Statements For Deduction Error")
        return {"success": False, "message": str(e)}


@frappe.whitelist()
def fetch_deduction_items_portal(scheme_request, stockist_statement, division=None):
    """Fetch items for deduction from scheme + statement (portal version)"""
    try:
        if not division:
            division = get_user_division()
        scheme = frappe.get_doc("Scheme Request", scheme_request)
        statement = frappe.get_doc("Stockist Statement", stockist_statement)

        if getattr(scheme, "division", None) and division and scheme.division not in [division, "Both"]:
            return {"success": False, "message": f"Scheme {scheme_request} does not belong to division {division}"}
        if getattr(statement, "division", None) and division and statement.division not in [division, "Both"]:
            return {"success": False, "message": f"Statement {stockist_statement} does not belong to division {division}"}

        # Validate stockist match
        if scheme.stockist_code != statement.stockist_code:
            return {
                "success": False,
                "message": f"Stockist mismatch: Scheme ({scheme.stockist_code}) vs Statement ({statement.stockist_code})"
            }

        # Build statement product map — keep the item so we can read sales_qty + current pts
        stmt_map = {item.product_code: item for item in statement.items}

        # For skipped-product messages show the business code, not the id.
        display_code_map = get_product_code_map([si.product_code for si in scheme.items])

        items = []
        skipped = []
        for scheme_item in scheme.items:
            if scheme_item.product_code not in stmt_map:
                skipped.append(display_code_map.get(scheme_item.product_code, scheme_item.product_code))
                continue

            scheme_free_qty = flt(scheme_item.free_quantity)
            special_rate = flt(scheme_item.special_rate)
            # Nothing to do for a product with neither free goods nor a discount
            if scheme_free_qty <= 0 and special_rate <= 0:
                continue

            product = frappe.get_doc("Product Master", scheme_item.product_code)
            stmt_item = stmt_map[scheme_item.product_code]
            # Effective current PTS on the statement line (per-line override or master)
            current_pts = flt(stmt_item.pts) or flt(product.pts)
            items.append({
                # Keep the id in product_code — it round-trips into the deduction
                # Link field. display_product_code is the human-facing code.
                "product_code": scheme_item.product_code,
                "display_product_code": product.product_code or scheme_item.product_code,
                "product_name": product.product_name,
                "pack": product.pack,
                "scheme_free_qty": scheme_free_qty,
                "current_sales_qty": flt(stmt_item.sales_qty),
                "deduct_qty": scheme_free_qty,
                "pts": flt(product.pts),
                "current_pts": current_pts,
                "special_rate": special_rate,
                "deducted_value": scheme_free_qty * flt(product.pts),
            })

        return {
            "success": True,
            "items": items,
            "skipped": skipped,
            "scheme_doctor": scheme.doctor_name,
            "scheme_hq": scheme.hq,
        }
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Fetch Deduction Items Portal Error")
        return {"success": False, "message": str(e)}


@frappe.whitelist()
@require_process("deductions")
def create_scheme_deduction_portal(scheme_request, stockist_statement, items, deduction_date=None, division=None):
    """Create a Scheme Deduction document from portal"""
    try:
        if isinstance(items, str):
            items = json.loads(items)

        scheme = frappe.get_doc("Scheme Request", scheme_request)
        statement = frappe.get_doc("Stockist Statement", stockist_statement)

        # Check if deduction already exists
        existing = frappe.db.exists("Scheme Deduction", {
            "scheme_request": scheme_request,
            "stockist_statement": stockist_statement,
            "docstatus": ["!=", 2]
        })
        if existing:
            return {"success": False, "message": f"Deduction already exists: {existing}"}

        if not division:
            division = getattr(scheme, "division", None) or get_user_division()
        if getattr(scheme, "division", None) and scheme.division not in [division, "Both"]:
            return {"success": False, "message": f"Scheme {scheme_request} does not belong to division {division}"}
        if getattr(statement, "division", None) and statement.division not in [division, "Both"]:
            return {"success": False, "message": f"Statement {stockist_statement} does not belong to division {division}"}

        doc = frappe.new_doc("Scheme Deduction")
        doc.scheme_request = scheme_request
        doc.stockist_statement = stockist_statement
        doc.stockist_code = scheme.stockist_code
        doc.doctor_code = scheme.doctor_code
        doc.scheme_date = scheme.application_date
        doc.division = division
        if deduction_date:
            doc.deduction_date = deduction_date

        total_qty = 0
        total_value = 0
        for item in items:
            deduct_qty = flt(item.get("deduct_qty", 0))
            pts = flt(item.get("pts", 0))
            deducted_value = deduct_qty * pts
            doc.append("items", {
                "product_code": item.get("product_code"),
                "product_name": item.get("product_name"),
                "pack": item.get("pack"),
                "scheme_free_qty": flt(item.get("scheme_free_qty", 0)),
                "current_free_qty": flt(item.get("current_sales_qty") or item.get("current_free_qty") or 0),
                "deduct_qty": deduct_qty,
                "pts": pts,
                "deducted_value": deducted_value,
                # Discount: when > 0, on_submit reprices the statement line PTS to this
                "special_rate": flt(item.get("special_rate", 0)),
            })
            total_qty += deduct_qty
            total_value += deducted_value

        doc.total_deducted_qty = total_qty
        doc.total_deducted_value = total_value
        doc.insert(ignore_permissions=False)
        # Auto-submit the deduction (no draft state needed). on_submit applies the
        # deduction to the statement and recalculates its closing/totals.
        doc.submit()

        # Mark the source scheme as Deducted (final lifecycle status)
        frappe.db.set_value("Scheme Request", scheme_request, "approval_status", "Deducted")
        frappe.db.commit()

        return {"success": True, "name": doc.name, "message": "Scheme deduction created successfully"}
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Create Scheme Deduction Portal Error")
        return {"success": False, "message": str(e)}


@frappe.whitelist()
@require_process("deductions")
def create_bulk_scheme_deductions_portal(deductions, deduction_date, division=None):
    """Create multiple Scheme Deductions at once from the auto-deduction portal page"""
    try:
        if isinstance(deductions, str):
            deductions = json.loads(deductions)

        if not division:
            division = get_user_division()

        results = []
        for entry in deductions:
            scheme_request = entry.get("scheme_request")
            stockist_statement = entry.get("stockist_statement")
            items = entry.get("items", [])

            if not scheme_request or not stockist_statement:
                results.append({"success": False, "message": "Missing scheme or statement"})
                continue

            result = create_scheme_deduction_portal(
                scheme_request=scheme_request,
                stockist_statement=stockist_statement,
                items=json.dumps(items) if not isinstance(items, str) else items,
                deduction_date=deduction_date,
                division=division,
            )
            results.append(result)

        created_count = sum(1 for r in results if r.get("success"))
        return {
            "success": True,
            "results": results,
            "created": created_count,
            "total": len(deductions),
            "message": f"{created_count} of {len(deductions)} deductions created",
        }
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Create Bulk Scheme Deductions Portal Error")
        return {"success": False, "message": str(e)}


@frappe.whitelist()
def get_scheme_deductions_portal(division=None, search=None, status=None, from_date=None, to_date=None):
    """List scheme deductions for the portal, enriched with human-readable
    doctor / stockist / statement-month so the UI shows names not raw codes."""
    try:
        if not division:
            division = get_user_division()

        conditions = ["sd.division = %(division)s"]
        params = {"division": division}

        if search:
            conditions.append(
                "(sd.name LIKE %(s)s OR sr.doctor_name LIKE %(s)s "
                "OR sr.stockist_name LIKE %(s)s OR sd.scheme_request LIKE %(s)s)"
            )
            params["s"] = f"%{search}%"
        if status:
            conditions.append("sd.status = %(status)s")
            params["status"] = status
        if from_date and to_date:
            conditions.append("sd.deduction_date BETWEEN %(from_date)s AND %(to_date)s")
            params["from_date"] = from_date
            params["to_date"] = to_date
        elif from_date:
            conditions.append("sd.deduction_date >= %(from_date)s")
            params["from_date"] = from_date
        elif to_date:
            conditions.append("sd.deduction_date <= %(to_date)s")
            params["to_date"] = to_date

        where_sql = " AND ".join(conditions)
        deductions = frappe.db.sql(f"""
            SELECT
                sd.name, sd.scheme_request, sd.stockist_statement, sd.deduction_date,
                sd.total_deducted_qty, sd.total_deducted_value, sd.status, sd.docstatus, sd.creation,
                sr.doctor_name, sr.stockist_name, sr.hq, sr.application_date,
                DATE_FORMAT(ss.statement_month, '%%b-%%Y') as statement_month_label
            FROM `tabScheme Deduction` sd
            LEFT JOIN `tabScheme Request` sr ON sd.scheme_request = sr.name
            LEFT JOIN `tabStockist Statement` ss ON sd.stockist_statement = ss.name
            WHERE {where_sql}
            ORDER BY sd.creation DESC
            LIMIT 200
        """, params, as_dict=True)

        # Resolve HQ display name (small, cached)
        hq_cache = {}
        for d in deductions:
            hq = d.get("hq")
            if hq and hq not in hq_cache:
                hq_cache[hq] = frappe.db.get_value("HQ Master", hq, "hq_name") or hq
            d["hq_name"] = hq_cache.get(hq, hq or "")

        return {"success": True, "data": deductions}
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Get Scheme Deductions Portal Error")
        return {"success": False, "message": str(e)}


@frappe.whitelist()
def get_scheme_deduction_detail(name):
    """Full detail of a single Scheme Deduction for the list click-through popup."""
    try:
        doc = frappe.get_doc("Scheme Deduction", name)
        sr = frappe.db.get_value(
            "Scheme Request", doc.scheme_request,
            ["doctor_name", "stockist_name", "hq", "application_date"], as_dict=True
        ) or {}
        stmt_month = frappe.db.get_value("Stockist Statement", doc.stockist_statement, "statement_month")
        # Items link by the Product Master id; the popup shows the business code.
        item_code_map = get_product_code_map([i.product_code for i in doc.items])
        items = []
        for it in doc.items:
            items.append({
                "product_code": item_code_map.get(it.product_code, it.product_code),
                "product_name": it.product_name,
                "pack": it.pack,
                "scheme_free_qty": flt(it.scheme_free_qty),
                "current_free_qty": flt(it.current_free_qty),
                "deduct_qty": flt(it.deduct_qty),
                "pts": flt(it.pts),
                "deducted_value": flt(it.deducted_value),
            })
        return {
            "success": True,
            "name": doc.name,
            "scheme_request": doc.scheme_request,
            "stockist_statement": doc.stockist_statement,
            "statement_month_label": stmt_month.strftime("%b-%Y") if stmt_month else "",
            "deduction_date": str(doc.deduction_date) if doc.deduction_date else "",
            "status": doc.status,
            "docstatus": doc.docstatus,
            "total_deducted_qty": flt(doc.total_deducted_qty),
            "total_deducted_value": flt(doc.total_deducted_value),
            "doctor_name": sr.get("doctor_name"),
            "stockist_name": sr.get("stockist_name"),
            "items": items,
        }
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Get Scheme Deduction Detail Error")
        return {"success": False, "message": str(e)}


def _month_bounds(month_str):
    """Given 'YYYY-MM' (or 'YYYY-MM-DD') return (first_day, last_day) date strings."""
    import calendar as _cal
    if not month_str:
        return None, None
    parts = str(month_str).split("-")
    y, m = int(parts[0]), int(parts[1])
    last = _cal.monthrange(y, m)[1]
    return f"{y:04d}-{m:02d}-01", f"{y:04d}-{m:02d}-{last:02d}"


@frappe.whitelist()
def get_bulk_deduction_candidates(division=None, zone=None, region=None, team=None,
                                  hq=None, scheme_month=None, statement_month=None):
    """List Approved (not-yet-Deducted) scheme requests for a scheme month, each
    resolved to its stockist's statement for the target statement month.
    Schemes whose stockist has no statement that month are returned with
    can_deduct=False so the UI can show them disabled."""
    try:
        if not division:
            division = get_user_division()

        conditions = [
            "sr.docstatus = 1",
            "sr.approval_status = 'Approved'",
            "sr.division = %(division)s",
        ]
        params = {"division": division}

        s_from, s_to = _month_bounds(scheme_month)
        if s_from:
            conditions.append("sr.application_date BETWEEN %(s_from)s AND %(s_to)s")
            params["s_from"] = s_from
            params["s_to"] = s_to

        if hq:
            conditions.append("sr.hq = %(hq)s")
            params["hq"] = hq
        if team:
            conditions.append("sr.team = %(team)s")
            params["team"] = team
        if region:
            conditions.append("sr.region = %(region)s")
            params["region"] = region
        elif zone:
            # Resolve all regions in the zone, then filter schemes to those regions
            zone_regions = frappe.get_all("Region Master", filters={"zone": zone}, pluck="name")
            if zone_regions:
                conditions.append("sr.region IN %(zone_regions)s")
                params["zone_regions"] = tuple(zone_regions)
            else:
                return {"success": True, "data": [], "statement_month": statement_month}

        where_sql = " AND ".join(conditions)
        schemes = frappe.db.sql(f"""
            SELECT sr.name, sr.application_date, sr.doctor_name, sr.doctor_code,
                   sr.stockist_code, sr.stockist_name, sr.hq, sr.total_scheme_value
            FROM `tabScheme Request` sr
            WHERE {where_sql}
            ORDER BY sr.application_date DESC
            LIMIT 300
        """, params, as_dict=True)

        st_year = st_month = None
        if statement_month:
            _p = str(statement_month).split("-")
            st_year, st_month = int(_p[0]), int(_p[1])

        hq_cache = {}
        out = []
        for sc in schemes:
            hqv = sc.get("hq")
            if hqv and hqv not in hq_cache:
                hq_cache[hqv] = frappe.db.get_value("HQ Master", hqv, "hq_name") or hqv
            sc["hq_name"] = hq_cache.get(hqv, hqv or "")

            row = {
                "scheme_request": sc["name"],
                "application_date": str(sc["application_date"]) if sc["application_date"] else "",
                "doctor_name": sc["doctor_name"],
                "stockist_code": sc["stockist_code"],
                "stockist_name": sc["stockist_name"],
                "hq_name": sc["hq_name"],
                "stockist_statement": None,
                "items": [],
                "total_free_qty": 0,
                "total_value": 0,
                "can_deduct": False,
                "reason": "",
            }

            # Already-deducted guard (belt & suspenders; Approved filter usually covers it)
            existing = frappe.db.exists("Scheme Deduction", {
                "scheme_request": sc["name"], "docstatus": ["!=", 2]
            })
            if existing:
                row["reason"] = f"Already deducted ({existing})"
                out.append(row)
                continue

            # Resolve the stockist's statement for the target month
            stmt_name = None
            if st_year:
                stmt = frappe.db.sql("""
                    SELECT name FROM `tabStockist Statement`
                    WHERE stockist_code = %(sc)s
                      AND YEAR(statement_month) = %(y)s AND MONTH(statement_month) = %(m)s
                      AND docstatus != 2
                      AND (division IS NULL OR division = %(division)s OR division = 'Both')
                    ORDER BY docstatus DESC LIMIT 1
                """, {"sc": sc["stockist_code"], "y": st_year, "m": st_month, "division": division}, as_dict=True)
                stmt_name = stmt[0]["name"] if stmt else None

            if not stmt_name:
                row["reason"] = "No statement for the selected statement month"
                out.append(row)
                continue

            result = fetch_deduction_items_portal(sc["name"], stmt_name, division)
            if not result.get("success"):
                row["reason"] = result.get("message") or "Could not build deduction items"
                out.append(row)
                continue

            items = result.get("items", [])
            row["stockist_statement"] = stmt_name
            row["items"] = items
            row["total_free_qty"] = sum(flt(i.get("deduct_qty")) for i in items)
            row["total_value"] = sum(flt(i.get("deducted_value")) for i in items)
            row["has_discount"] = any(flt(i.get("special_rate")) > 0 for i in items)
            row["can_deduct"] = len(items) > 0
            if not items:
                row["reason"] = "No scheme products present in the statement"
            out.append(row)

        return {"success": True, "data": out, "statement_month": statement_month}
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Get Bulk Deduction Candidates Error")
        return {"success": False, "message": str(e)}


@frappe.whitelist()
@require_process("scheme_delete")
def delete_and_revert_scheme(scheme_request):
    """Undo any deduction for a scheme (reversing the statement via the deduction
    controller's on_cancel) and then delete the scheme request."""
    try:
        if not scheme_request:
            return {"success": False, "message": "No scheme specified"}

        scheme = frappe.get_doc("Scheme Request", scheme_request)
        division = get_user_division()
        if scheme.division and division and division != "Both" and scheme.division not in (division, "Both"):
            return {"success": False, "message": "This scheme belongs to another division"}

        # 1) Cancel + delete any non-cancelled deductions (on_cancel reverses the
        #    statement and recalculates its closing/totals).
        reverted = []
        deductions = frappe.get_all(
            "Scheme Deduction",
            filters={"scheme_request": scheme_request, "docstatus": ["!=", 2]},
            pluck="name",
        )
        for dname in deductions:
            dd = frappe.get_doc("Scheme Deduction", dname)
            if dd.docstatus == 1:
                dd.flags.ignore_permissions = True
                dd.cancel()
            reverted.append(dname)
            # Remove the cancelled deduction so it no longer links to the scheme
            frappe.delete_doc("Scheme Deduction", dname, force=1, ignore_permissions=True)

        # 2) Cancel (if submitted) and delete the scheme request. Access is already
        #    gated by @require_process("scheme_delete") (portal Admin only); use
        #    ignore_permissions throughout so the operation doesn't depend on the admin's
        #    Frappe role/user_type and can't half-fail after the deductions above were
        #    already reverted/removed (consistent with that Scheme Deduction delete).
        if scheme.docstatus == 1:
            scheme.flags.ignore_permissions = True
            scheme.cancel()
        frappe.delete_doc("Scheme Request", scheme_request, force=1, ignore_permissions=True)
        frappe.db.commit()

        msg = "Scheme deleted"
        if reverted:
            msg += f" and {len(reverted)} deduction(s) reverted"
        return {"success": True, "reverted": reverted, "message": msg}
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Delete And Revert Scheme Error")
        return {"success": False, "message": str(e)}


@frappe.whitelist()
def remap_scheme_stockist(scheme_request, stockist_code):
    """Change the stockist on a PENDING scheme request (original or repeated).
    Only the stockist mapping is updated; items/values are untouched."""
    try:
        if not scheme_request or not stockist_code:
            return {"success": False, "message": "Scheme and stockist are required"}

        scheme = frappe.get_doc("Scheme Request", scheme_request)
        division = get_user_division()
        if scheme.division and division and division != "Both" and scheme.division not in (division, "Both"):
            return {"success": False, "message": "This scheme belongs to another division"}
        if scheme.approval_status != "Pending":
            return {"success": False, "message": "Only pending schemes can be remapped"}

        stockist_name = frappe.db.get_value("Stockist Master", stockist_code, "stockist_name")
        if not stockist_name:
            return {"success": False, "message": "Invalid stockist"}

        frappe.db.set_value("Scheme Request", scheme_request, {
            "stockist_code": stockist_code,
            "stockist_name": stockist_name,
        })
        frappe.db.commit()

        return {
            "success": True,
            "stockist_code": stockist_code,
            "stockist_name": stockist_name,
            "stockist_display_code": get_stockist_code_map([stockist_code]).get(stockist_code, stockist_code),
            "message": "Stockist remapped successfully",
        }
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Remap Scheme Stockist Error")
        return {"success": False, "message": str(e)}


@frappe.whitelist()
def get_stockist_statement_history(stockist_code, division=None):
    """Statement-level secondary sales history for a stockist (recent months)."""
    try:
        if not division:
            division = get_user_division()
        if not stockist_code:
            return {"success": True, "data": []}
        rows = frappe.db.sql("""
            SELECT ss.name,
                   DATE_FORMAT(ss.statement_month, '%%b-%%Y') as month_label,
                   ss.statement_month,
                   ss.total_opening_value, ss.total_purchase_value,
                   ss.total_sales_value_pts, ss.total_sales_value_ptr,
                   ss.total_free_value, ss.total_closing_value,
                   ss.docstatus,
                   COALESCE(SUM(si.sales_qty), 0)       as sales_qty,
                   COALESCE(SUM(si.free_qty), 0)        as free_qty,
                   COALESCE(SUM(si.free_qty_scheme), 0) as free_qty_scheme,
                   COUNT(DISTINCT si.product_code)      as product_count
            FROM `tabStockist Statement` ss
            LEFT JOIN `tabStockist Statement Item` si ON si.parent = ss.name
            WHERE ss.stockist_code = %(sc)s
              AND ss.docstatus != 2
              AND (ss.division IS NULL OR ss.division = %(division)s OR ss.division = 'Both')
            GROUP BY ss.name
            ORDER BY ss.statement_month DESC
            LIMIT 24
        """, {"sc": stockist_code, "division": division}, as_dict=True)
        return {"success": True, "data": rows}
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Get Stockist Statement History Error")
        return {"success": False, "message": str(e)}


@frappe.whitelist()
def get_scheme_history_portal(doctor_code=None, stockist_code=None, hq=None, division=None):
    """Recent scheme requests for the doctor / stockist context (for history tab)."""
    try:
        if not division:
            division = get_user_division()
        conditions = ["sr.docstatus != 2", "sr.division = %(division)s"]
        params = {"division": division}
        if doctor_code:
            conditions.append("sr.doctor_code = %(doctor_code)s")
            params["doctor_code"] = doctor_code
        if stockist_code:
            conditions.append("sr.stockist_code = %(stockist_code)s")
            params["stockist_code"] = stockist_code
        if hq:
            conditions.append("sr.hq = %(hq)s")
            params["hq"] = hq
        where_sql = " AND ".join(conditions)
        rows = frappe.db.sql(f"""
            SELECT sr.name, sr.application_date, sr.doctor_name, sr.stockist_name, sr.hq,
                   sr.total_scheme_value, sr.approval_status,
                   COALESCE(sr.repeated_request, 0) as repeated_request,
                   COUNT(sri.name) as product_count,
                   COALESCE(SUM(sri.free_quantity), 0) as total_free_qty
            FROM `tabScheme Request` sr
            LEFT JOIN `tabScheme Request Item` sri ON sri.parent = sr.name
            WHERE {where_sql}
            GROUP BY sr.name
            ORDER BY sr.application_date DESC
            LIMIT 50
        """, params, as_dict=True)
        hq_cache = {}
        for r in rows:
            r["application_date"] = str(r["application_date"]) if r["application_date"] else ""
            hq = r.get("hq")
            if hq and hq not in hq_cache:
                hq_cache[hq] = frappe.db.get_value("HQ Master", hq, "hq_name") or hq
            r["hq_name"] = hq_cache.get(hq, hq or "")
        return {"success": True, "data": rows}
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Get Scheme History Portal Error")
        return {"success": False, "message": str(e)}


@frappe.whitelist()
def get_doctor_scheme_history(doctor_code=None, hq=None, division=None, period_months=3):
    """Doctor Scheme History – product-line detail of a doctor's scheme requests.

    Mirrors the client 'Doctor Scheme History' sheet, one row per scheme line:
      App Date | P Code | PTS | Qty (bx/units) | Free (bx/units) | Spl Rate | Value
    Value / PTS / Special Rate come straight from the stored Scheme Request Item
    (Value = product_value; free goods are never valued). Ordered newest-first.

    period_months: trailing window in months (default 3). 0 / None / 'all' = no limit.
    """
    try:
        if not division:
            division = get_user_division()
        if not doctor_code:
            return {"success": True, "rows": [], "doctor_name": "", "totals": {}}

        conds = ["sr.docstatus != 2", "sr.division = %(division)s", "sr.doctor_code = %(doctor_code)s"]
        params = {"division": division, "doctor_code": doctor_code}
        if hq:
            conds.append("sr.hq = %(hq)s")
            params["hq"] = hq

        # Trailing month window (calendar months, inclusive of the current month).
        try:
            pm_int = int(period_months)
        except (TypeError, ValueError):
            pm_int = 0
        if pm_int and pm_int > 0:
            from datetime import date as _date
            t = _date.today()
            y, mo = t.year, t.month - (pm_int - 1)
            while mo <= 0:
                mo += 12
                y -= 1
            conds.append("sr.application_date >= %(since)s")
            params["since"] = f"{y:04d}-{mo:02d}-01"

        where = " AND ".join(conds)
        rows = frappe.db.sql(f"""
            SELECT sr.name AS scheme, sr.application_date, sr.doctor_name,
                   COALESCE(pm.product_code, sri.product_code) AS product_code,
                   COALESCE(pm.product_name, sri.product_name, '') AS product_name,
                   COALESCE(NULLIF(sri.pack, ''), pm.pack, '') AS pack,
                   IFNULL(sri.product_rate, 0)   AS pts,
                   IFNULL(sri.quantity, 0)       AS quantity,
                   IFNULL(sri.free_quantity, 0)  AS free_quantity,
                   IFNULL(sri.special_rate, 0)   AS special_rate,
                   IFNULL(sri.product_value, 0)  AS value,
                   sr.approval_status
              FROM `tabScheme Request` sr
        INNER JOIN `tabScheme Request Item` sri ON sri.parent = sr.name
         LEFT JOIN `tabProduct Master` pm ON pm.name = sri.product_code
             WHERE {where}
          ORDER BY sr.application_date DESC, sr.creation DESC, sri.idx ASC
        """, params, as_dict=True)

        doctor_name = rows[0]["doctor_name"] if rows else (
            frappe.db.get_value("Doctor Master", doctor_code, "doctor_name") or "")
        # Qty / Free are entered in strips-units; the sheet shows them box-converted
        # (÷ strips-per-box, the first number of an "NxM" pack) — mirrors the scheme
        # order-value maths (Value = box-qty × rate). Free goods are never valued.
        total_value = 0.0
        total_free = 0.0
        requests = set()
        for r in rows:
            conv = _scheme_pack_conversion(r.get("pack")) or 1
            r["quantity"] = round(flt(r.get("quantity")) / conv, 2)
            r["free_quantity"] = round(flt(r.get("free_quantity")) / conv, 2)
            total_value += flt(r.get("value"))
            total_free += flt(r["free_quantity"])
            requests.add(r.get("scheme"))
            ad = r.get("application_date")
            r["application_date"] = frappe.utils.getdate(ad).strftime("%d/%m/%Y") if ad else ""
            r.pop("doctor_name", None)
            r.pop("pack", None)
        return {"success": True, "doctor_name": doctor_name, "rows": rows,
                "totals": {"value": total_value, "free_qty": round(total_free, 2),
                           "lines": len(rows), "requests": len(requests)}}
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Get Doctor Scheme History Error")
        return {"success": False, "message": str(e)}


@frappe.whitelist()
def get_hq_sales_history_3m(hq=None, division=None, end_month=None, sales_mode="after_deduction",
                            months=3, product_group=None, product_category=None):
    """Sales History – Past N (default 3) Months for an HQ.

    Product x month secondary-sales matrix (box quantities) plus the latest
    month's closing stock, and an 'H.Q Value' header row in Rs. Lakhs. Mirrors
    the client 'Sales History – Past 3 Months' sheet:
      P Code | <m1>/Sales | <m2>/Sales | <m3>/Sales | <m3>/Closing

    Box qty = raw statement qty / conversion_factor (strip -> box), summed per
    HQ. H.Q Value = secondary sales value (after/before deduction toggle, same
    semantics as Report 11). end_month (YYYY-MM) defaults to the latest statement
    month on record for the HQ.
    """
    try:
        if not division:
            division = get_user_division()
        if not hq:
            return {"success": True, "products": [], "months": [], "hq_name": "", "hq_value": {}}

        from datetime import date as _date, datetime as _datetime
        import calendar

        try:
            n = int(months)
        except (TypeError, ValueError):
            n = 3
        n = max(1, min(n, 12))

        # Resolve the end month.
        ey = em = None
        if end_month:
            try:
                parts = str(end_month)[:7].split("-")
                ey, em = int(parts[0]), int(parts[1])
            except Exception:
                ey = em = None
        if not ey:
            latest = frappe.db.sql(
                """SELECT MAX(statement_month) FROM `tabStockist Statement`
                    WHERE division=%s AND hq=%s AND docstatus IN (0, 1)""",
                (division, hq))
            latest = latest[0][0] if latest and latest[0] else None
            if latest:
                ey, em = latest.year, latest.month
            else:
                t = _date.today()
                ey, em = t.year, t.month

        # Build the month sequence (oldest -> newest), n months ending at em/ey.
        months_seq = []
        y, mo = ey, em
        for _ in range(n):
            months_seq.append({"year": y, "month": mo, "key": f"{y}-{mo:02d}",
                               "label": _datetime(y, mo, 1).strftime("%b")})
            mo -= 1
            if mo == 0:
                mo = 12
                y -= 1
        months_seq.reverse()
        month_keys = [ms["key"] for ms in months_seq]
        last_key = month_keys[-1]

        first = months_seq[0]
        last = months_seq[-1]
        from_d = f"{first['year']}-{first['month']:02d}-01"
        to_d = f"{last['year']}-{last['month']:02d}-{calendar.monthrange(last['year'], last['month'])[1]:02d}"

        # Optional product group / category narrowing (intersection).
        _pf = _resolve_product_filter(division, None, product_group, product_category)
        prod_clause = ""
        prod_params = []
        if _pf is not None:
            prod_clause = " AND si.product_code IN (" + ", ".join(["%s"] * len(_pf)) + ")"
            prod_params = _pf

        conv = "IFNULL(NULLIF(si.conversion_factor, 0), 1)"
        pts = "IFNULL(si.pts, 0)"
        if sales_mode == "before_deduction":
            val_expr = f"((si.sales_qty + si.free_qty) / {conv}) * {pts}"
        else:
            val_expr = (f"(COALESCE(si.scheme_deducted_qty_calc, "
                        f"(si.sales_qty + si.free_qty - IFNULL(si.free_qty_scheme, 0))) / {conv}) * {pts}")

        rows = frappe.db.sql(f"""
            SELECT COALESCE(pm.product_code, si.product_code) AS product_code,
                   COALESCE(pm.product_name, si.product_name, si.raw_product_name, '') AS product_name,
                   IFNULL(pm.sequence, 999999) AS seq,
                   YEAR(ss.statement_month) AS y, MONTH(ss.statement_month) AS m,
                   SUM(si.sales_qty / {conv})   AS sales_boxes,
                   SUM(si.closing_qty / {conv}) AS closing_boxes,
                   SUM({val_expr})              AS sales_value,
                   SUM(IFNULL(si.closing_value, 0)) AS closing_value
              FROM `tabStockist Statement` ss
        INNER JOIN `tabStockist Statement Item` si
                ON si.parent = ss.name AND si.parenttype = 'Stockist Statement'
         LEFT JOIN `tabProduct Master` pm ON pm.name = si.product_code
             WHERE ss.division = %s AND ss.hq = %s AND ss.docstatus IN (0, 1)
               AND ss.statement_month BETWEEN %s AND %s{prod_clause}
          GROUP BY product_code, product_name, seq, y, m
        """, [division, hq, from_d, to_d] + prod_params, as_dict=True)

        LAKH = 100000.0
        prod_map = {}
        hq_val = {k: 0.0 for k in month_keys}
        hq_close_val = 0.0
        for r in rows:
            key = f"{int(r.y)}-{int(r.m):02d}"
            p = prod_map.get(r.product_code)
            if p is None:
                p = {"product_code": r.product_code, "product_name": r.product_name or "",
                     "seq": r.seq, "sales": {}, "closing": {}}
                prod_map[r.product_code] = p
            p["sales"][key] = flt(r.sales_boxes)
            p["closing"][key] = flt(r.closing_boxes)
            if key in hq_val:
                hq_val[key] += flt(r.sales_value)
            if key == last_key:
                hq_close_val += flt(r.closing_value)

        products = []
        for code, p in prod_map.items():
            products.append({
                "product_code": code,
                "product_name": p["product_name"],
                "monthly": [round(p["sales"].get(k, 0.0), 2) for k in month_keys],
                "closing": round(p["closing"].get(last_key, 0.0), 2),
            })
        products.sort(key=lambda x: (prod_map[x["product_code"]]["seq"], x["product_code"] or ""))

        hq_name = frappe.db.get_value("HQ Master", hq, "hq_name") or hq
        return {
            "success": True,
            "hq_name": hq_name,
            "months": months_seq,
            "products": products,
            "hq_value": {
                "monthly": [round(hq_val[k] / LAKH, 2) for k in month_keys],
                "closing": round(hq_close_val / LAKH, 2),
            },
            "sales_mode": sales_mode,
            "period_label": first["label"] + "-" + str(first["year"])[2:] + " to "
                            + last["label"] + "-" + str(last["year"])[2:],
        }
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Get HQ Sales History 3M Error")
        return {"success": False, "message": str(e)}


@frappe.whitelist()
def get_schemes_for_stockist(stockist_code, division=None):
    """Approved (not-yet-deducted) scheme requests for a given stockist."""
    try:
        if not division:
            division = get_user_division()
        if not stockist_code:
            return {"success": True, "data": []}
        rows = frappe.db.sql("""
            SELECT sr.name, sr.application_date, sr.doctor_name, sr.doctor_code,
                   sr.total_scheme_value, sr.approval_status
            FROM `tabScheme Request` sr
            WHERE sr.docstatus = 1
              AND sr.approval_status = 'Approved'
              AND sr.stockist_code = %(sc)s
              AND sr.division = %(division)s
            ORDER BY sr.application_date DESC
            LIMIT 100
        """, {"sc": stockist_code, "division": division}, as_dict=True)
        for r in rows:
            r["application_date"] = str(r["application_date"]) if r["application_date"] else ""
        return {"success": True, "data": rows}
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Get Schemes For Stockist Error")
        return {"success": False, "message": str(e)}


@frappe.whitelist()
def get_stockist_list_for_deduction(division=None, search=""):
    """Active stockists for the manual-deduction stockist dropdown (name shown,
    internal id used as value per the stockist-code convention)."""
    try:
        if not division:
            division = get_user_division()
        search = (search or "").strip()
        conditions = ["status = 'Active'"]
        params = {}
        if division and division != "Both":
            conditions.append("(division = %(division)s OR division = 'Both')")
            params["division"] = division
        if search:
            conditions.append("(stockist_name LIKE %(s)s OR stockist_code LIKE %(s)s)")
            params["s"] = f"%{search}%"
        where_sql = " AND ".join(conditions)
        rows = frappe.db.sql(f"""
            SELECT name, stockist_code, stockist_name, hq
            FROM `tabStockist Master`
            WHERE {where_sql}
            ORDER BY stockist_name ASC
            LIMIT 500
        """, params, as_dict=True)
        return {"success": True, "data": rows}
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Get Stockist List For Deduction Error")
        return {"success": False, "message": str(e)}


@frappe.whitelist()
def portal_repeat_scheme_request(source_name):
    """Repeat an approved scheme request from portal"""
    try:
        source_doc = frappe.get_doc("Scheme Request", source_name)
        if source_doc.approval_status != "Approved":
            return {"success": False, "message": "Only approved scheme requests can be repeated"}

        new_doc = frappe.new_doc("Scheme Request")
        new_doc.application_date = nowdate()
        new_doc.requested_by = frappe.session.user
        new_doc.team = source_doc.team
        new_doc.region = source_doc.region
        new_doc.hq = source_doc.hq
        new_doc.stockist_code = source_doc.stockist_code
        new_doc.stockist_name = source_doc.stockist_name
        new_doc.doctor_code = source_doc.doctor_code
        new_doc.doctor_name = source_doc.doctor_name
        new_doc.doctor_place = source_doc.doctor_place
        new_doc.specialization = source_doc.specialization
        new_doc.hospital_address = source_doc.hospital_address
        new_doc.scheme_notes = f"Repeated from {source_doc.name}"
        new_doc.approval_status = "Pending"
        new_doc.repeated_request = 1

        for item in source_doc.items:
            new_doc.append("items", {
                "product_code": item.product_code,
                "product_name": item.product_name,
                "pack": item.pack,
                "quantity": item.quantity,
                "free_quantity": item.free_quantity,
                "product_rate": item.product_rate,
                "special_rate": item.special_rate,
                "product_value": item.product_value,
            })

        new_doc.insert()
        frappe.db.commit()
        return {"success": True, "name": new_doc.name, "message": "New scheme request created successfully"}
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Portal Repeat Scheme Error")
        return {"success": False, "message": str(e)}


@frappe.whitelist()
def get_doctor_monthly_limit_info(doctor_code, application_date=None):
    """Get per-product monthly request count for a doctor"""
    try:
        from frappe.utils import getdate, get_first_day, get_last_day
        if not application_date:
            application_date = nowdate()
        app_date = getdate(application_date)
        first_day = get_first_day(app_date)
        last_day = get_last_day(app_date)
        month_name = app_date.strftime("%B %Y")

        rows = frappe.db.sql("""
            SELECT sri.product_code, COUNT(DISTINCT sr.name) as request_count
            FROM `tabScheme Request` sr
            INNER JOIN `tabScheme Request Item` sri ON sr.name = sri.parent
            WHERE sr.doctor_code = %(doctor_code)s
              AND sr.application_date BETWEEN %(first_day)s AND %(last_day)s
              AND sr.docstatus != 2
            GROUP BY sri.product_code
        """, {"doctor_code": doctor_code, "first_day": first_day, "last_day": last_day}, as_dict=True)

        product_counts = {row.product_code: row.request_count for row in rows}
        total_requests = frappe.db.count("Scheme Request", {
            "doctor_code": doctor_code,
            "application_date": ["between", [first_day, last_day]],
            "docstatus": ["!=", 2]
        })

        return {
            "success": True,
            "month": month_name,
            "total_requests": total_requests,
            "product_counts": product_counts,
            "limit_per_product": 3,
        }
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Doctor Monthly Limit Error")
        return {"success": False, "message": str(e)}


@frappe.whitelist()
def create_scheme_request_v2(data):
    """Create a new scheme request from portal (robust version).

    Serves BOTH scheme-new and scheme-repeat. Access is enforced by portal role here
    (the endpoint had no guard before): repeats are open to every portal role, while
    brand-new requests are limited to Admin / HO / Regional-Future per the process
    matrix — so a Regional User may only create repeats. The scheme is then inserted
    with ignore_permissions=True so it doesn't depend on the acting user's Frappe
    role/user_type: a Regional User who is a Website User (no Sales User role) would
    otherwise fail Frappe's doctype create permission. Business validations on Scheme
    Request (monthly limits, etc.) still run — ignore_permissions skips only perms."""
    if isinstance(data, str):
        data = json.loads(data)
    _is_repeat = str(data.get("repeated_request", 0)).lower() in ("1", "true", "yes")
    require("scheme_repeat" if _is_repeat else "scheme_new")
    try:
        doc = frappe.new_doc("Scheme Request")
        doc.application_date = data.get("application_date") or nowdate()
        doc.requested_by = frappe.session.user
        doc.hq = data.get("hq")
        doc.doctor_code = data.get("doctor_code")
        doc.doctor_name = data.get("doctor_name")
        doc.doctor_place = data.get("doctor_place")
        doc.specialization = data.get("specialization")
        doc.hospital_address = data.get("hospital_address")
        doc.team = data.get("team")
        doc.region = data.get("region")
        doc.stockist_code = data.get("stockist_code")
        doc.stockist_name = data.get("stockist_name")
        doc.scheme_notes = data.get("scheme_notes")
        doc.approval_status = "Pending"
        doc.repeated_request = 1 if str(data.get("repeated_request", 0)).lower() in ("1", "true", "yes") else 0

        # Supporting document attachments (uploaded via upload_scheme_attachment)
        for i in range(1, 5):
            url = data.get("proof_attachment_%d" % i)
            if url:
                setattr(doc, "proof_attachment_%d" % i, url)

        for item in data.get("items", []):
            if not item.get("product_code"):
                continue
            qty = flt(item.get("quantity", 0))
            free_qty = flt(item.get("free_quantity", 0))
            rate = flt(item.get("product_rate", 0))
            special_rate = flt(item.get("special_rate", 0))

            # Calculate scheme %
            scheme_pct = 0
            if special_rate > 0 and rate > 0:
                scheme_pct = ((rate - special_rate) / rate) * 100
            elif free_qty > 0 and qty > 0:
                scheme_pct = (free_qty / qty) * 100

            # Order Value = (order qty in strips/units ÷ strips-per-box) × rate-per-box
            # (PTS or special rate). Mirrors Scheme Request._compute_order_value, which
            # recomputes this authoritatively on save. Free qty is not part of the value.
            pack_match = re.match(r'(\d+)\s*[xX]\s*(\d+)', str(item.get("pack") or "").strip().upper())
            conversion_factor = flt(pack_match.group(1)) if pack_match else 1
            if not conversion_factor:
                conversion_factor = 1
            effective_rate = special_rate if special_rate > 0 else rate
            product_value = (qty / conversion_factor) * effective_rate

            doc.append("items", {
                "product_code": item.get("product_code"),
                "product_name": item.get("product_name"),
                "pack": item.get("pack"),
                "quantity": qty,
                "free_quantity": free_qty,
                "product_rate": rate,
                "special_rate": special_rate,
                "scheme_percentage": round(scheme_pct, 2),
                "product_value": product_value,
            })

        doc.insert(ignore_permissions=True)

        # Re-point uploaded proof files to this Scheme Request so private-file
        # access is granted through document permissions when viewing later.
        for i in range(1, 5):
            url = getattr(doc, "proof_attachment_%d" % i, None)
            if not url:
                continue
            # Claim every still-unattached File row for this URL (uploads can produce more
            # than one). Never re-point a file already attached elsewhere — that would
            # steal it from another document. If this link is missed the attachment is
            # still downloadable via download_scheme_attachment(), which re-links it.
            for file_name in frappe.db.get_all(
                    "File",
                    filters={"file_url": url, "attached_to_name": ["is", "not set"]},
                    pluck="name"):
                frappe.db.set_value("File", file_name, {
                    "attached_to_doctype": "Scheme Request",
                    "attached_to_name": doc.name,
                    "attached_to_field": "proof_attachment_%d" % i,
                }, update_modified=False)

        frappe.db.commit()

        return {"success": True, "name": doc.name, "message": "Scheme request created successfully"}
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Create Scheme Request V2 Error")
        return {"success": False, "message": str(e)}


@frappe.whitelist()
def get_scheme_list_portal(division=None, filters=None):
    """Get schemes for portal list view with division filter"""
    try:
        if not division:
            division = get_user_division()
        if isinstance(filters, str):
            filters = json.loads(filters) if filters else {}
        if not filters:
            filters = {}

        where_clauses = ["sr.division = %(division)s"]
        params = {"division": division}

        if filters.get("status"):
            where_clauses.append("sr.approval_status = %(status)s")
            params["status"] = filters["status"]
        if filters.get("from_date") and filters.get("to_date"):
            where_clauses.append("sr.application_date BETWEEN %(from_date)s AND %(to_date)s")
            params["from_date"] = filters["from_date"]
            params["to_date"] = filters["to_date"]
        elif filters.get("from_date"):
            where_clauses.append("sr.application_date >= %(from_date)s")
            params["from_date"] = filters["from_date"]
        elif filters.get("to_date"):
            where_clauses.append("sr.application_date <= %(to_date)s")
            params["to_date"] = filters["to_date"]
        if filters.get("search"):
            where_clauses.append(
                "(sr.name LIKE %(search)s OR sr.doctor_name LIKE %(search)s "
                "OR sr.stockist_name LIKE %(search)s OR sr.hq LIKE %(search)s)"
            )
            params["search"] = f"%{filters['search']}%"
        if filters.get("request_type") == "Repeated":
            where_clauses.append("COALESCE(sr.repeated_request, 0) = 1")
        elif filters.get("request_type") == "Original":
            where_clauses.append("COALESCE(sr.repeated_request, 0) = 0")

        # Region scoping (security): non-admins only see schemes in their mapped regions.
        from scanify.permissions import get_allowed_region_codes, clamp_region_codes
        _allowed = get_allowed_region_codes(division=division)
        if _allowed is not None:
            _codes = clamp_region_codes(filters.get("region"), division=division)
            if not _codes:
                where_clauses.append("1=0")
            else:
                where_clauses.append("sr.region IN %(allowed_regions)s")
                params["allowed_regions"] = tuple(_codes)
        elif filters.get("region"):
            where_clauses.append("sr.region = %(region)s")
            params["region"] = filters["region"]

        where_sql = " AND ".join(where_clauses)

        rows = frappe.db.sql(f"""
            SELECT
                sr.name,
                sr.application_date,
                sr.doctor_name,
                sr.doctor_code,
                sr.hq,
                hqm.hq_name as hq_name,
                sr.stockist_name,
                sr.stockist_code,
                sr.approval_status,
                COALESCE(sr.repeated_request, 0) as repeated_request,
                sr.total_scheme_value,
                sr.requested_by,
                sr.docstatus,
                COUNT(sri.name) as product_count
            FROM `tabScheme Request` sr
            LEFT JOIN `tabScheme Request Item` sri ON sr.name = sri.parent
            LEFT JOIN `tabHQ Master` hqm ON sr.hq = hqm.name
            WHERE {where_sql}
            GROUP BY sr.name
            ORDER BY sr.application_date DESC, sr.creation DESC
            LIMIT 200
        """, params, as_dict=True)

        # Display the human-facing Stockist Code instead of the internal id.
        code_map = get_stockist_code_map([r["stockist_code"] for r in rows])
        for r in rows:
            r["stockist_code"] = code_map.get(r["stockist_code"], r["stockist_code"])

        return {"success": True, "data": rows, "division": division}
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Get Scheme List Portal Error")
        return {"success": False, "message": str(e)}



def _resolve_stockist_pk(code, division=None):
    """Resolve an editable Stockist Code (business code) to the Stockist Master id (PK).

    Division-scoped so the SAME business code reused across two divisions never
    cross-resolves. Falls back to treating the input as the PK itself, which keeps
    legacy callers working for data where stockist_code == name.
    """
    if not code:
        return None
    filters = {"stockist_code": code}
    if division:
        filters["division"] = ["in", [division, "Both"]]
    name = frappe.db.get_value("Stockist Master", filters, "name")
    if name:
        return name
    # Legacy / already-a-PK fallback (data where stockist_code == name). Still
    # honour the division so the fallback can't cross-resolve either.
    pk_division = frappe.db.get_value("Stockist Master", code, "division")
    if pk_division is not None and (not division or pk_division in (division, "Both")):
        return code
    return None


def _resolve_product_pk(code, division=None):
    """Resolve an editable Product Code (business code) to the Product Master id (PK).

    Division-scoped so the SAME business code reused across two divisions never
    cross-resolves. Falls back to treating the input as the PK itself, which keeps
    legacy callers working for data where product_code == name.
    """
    if not code:
        return None
    filters = {"product_code": code}
    if division:
        filters["division"] = ["in", [division, "Both"]]
    name = frappe.db.get_value("Product Master", filters, "name")
    if name:
        return name
    # Legacy / already-a-PK fallback (data where product_code == name). Still
    # honour the division so the fallback can't cross-resolve either.
    pk_division = frappe.db.get_value("Product Master", code, "division")
    if pk_division is not None and (not division or pk_division in (division, "Both")):
        return code
    return None


@frappe.whitelist()
def get_stockist_details(stockist_code, division=None):

    if not stockist_code:
        frappe.throw(_("Stockist code is required"))

    if not division:
        division = get_user_division()

    # Find Stockist Master by editable Stockist Code, scoped to the active division
    # so a code reused across divisions resolves to THIS division's stockist.
    name = _resolve_stockist_pk(stockist_code, division)
    if not name:
        frappe.throw(_("Stockist not found"))

    st = frappe.get_doc("Stockist Master", name)

    team = region = zone = None
    if st.hq:
        team, region, zone = frappe.db.get_value("HQ Master", st.hq, ["team", "region", "zone"]) or (None, None, None)

    hq_name = frappe.db.get_value("HQ Master", st.hq, "hq_name") if st.hq else None
    team_name = frappe.db.get_value("Team Master", team, "team_name") if team else None
    region_name = frappe.db.get_value("Region Master", region, "region_name") if region else None
    zone_name = frappe.db.get_value("Zone Master", zone, "zone_name") if zone else None

    return {
        "stockist_code": st.stockist_code,
        "stockist_name": st.stockist_name,
        "status": getattr(st, "status", None),
        "division": getattr(st, "division", None),
        "hq": st.hq,
        "team": team,
        "region": region,
        "zone": zone,
        "hq_name": hq_name or st.hq,
        "team_name": team_name or team,
        "region_name": region_name or region,
        "zone_name": zone_name or zone,
    }


# ============================================================================
# INSIGHTS / ANALYTICS API
# ============================================================================

@frappe.whitelist()
def get_insights_filter_options(division=None):
    """Return filter dropdown options for Insights page."""
    if not division:
        division = get_user_division()

    # Confine a non-admin's pickers to their mapped regions (and the teams/HQs inside
    # them). Admin (allowed is None) sees everything.
    allowed = _allowed_region_codes_or_all(division)
    region_filters = {"status": "Active", "division": ["in", [division, "Both"]]}
    team_filters = {"status": "Active", "division": ["in", [division, "Both"]]}
    hq_filters = {"status": "Active", "division": division}
    if allowed is not None:
        codes = allowed or ["__no_region__"]
        region_filters["name"] = ["in", codes]
        team_filters["region"] = ["in", codes]
        hq_filters["region"] = ["in", codes]

    regions = frappe.get_all("Region Master", filters=region_filters, pluck="name", order_by="name")
    teams = frappe.get_all("Team Master", filters=team_filters, pluck="name", order_by="name")
    hqs = frappe.get_all("HQ Master", filters=hq_filters, pluck="name", order_by="name")
    financial_years = frappe.db.sql(
        "SELECT DISTINCT financial_year FROM `tabHQ Yearly Target` WHERE division=%s ORDER BY financial_year DESC",
        (division,), as_list=1
    )

    return {
        "regions": regions,
        "teams": teams,
        "hqs": hqs,
        "financial_years": [f[0] for f in financial_years],
    }


def _build_insights_conditions(division, from_date=None, to_date=None, region=None, team=None, hq=None, date_field="creation"):
    """Helper to build WHERE conditions for insights queries."""
    conditions = ["1=1"]
    values = []

    if division:
        conditions.append("division = %s")
        values.append(division)
    if from_date:
        conditions.append(f"{date_field} >= %s")
        values.append(from_date)
    if to_date:
        conditions.append(f"{date_field} <= %s")
        values.append(to_date)
    _scope_region_sql_pos(conditions, values, "region", division, region)
    if team:
        conditions.append("team = %s")
        values.append(team)
    if hq:
        conditions.append("hq = %s")
        values.append(hq)

    return " AND ".join(conditions), values


@frappe.whitelist()
def get_insights_scheme_data(division=None, from_date=None, to_date=None, region=None, team=None, hq=None):
    """Scheme analytics data — trends, approval funnel, top doctors, top HQs, region breakdown."""
    if not division:
        division = get_user_division()

    cond, vals = _build_insights_conditions(division, from_date, to_date, region, team, hq, "application_date")

    # Monthly scheme trend
    monthly_trend = frappe.db.sql(f"""
        SELECT DATE_FORMAT(application_date, '%%Y-%%m') as month,
               COUNT(*) as count,
               COALESCE(SUM(total_scheme_value), 0) as total_value
        FROM `tabScheme Request`
        WHERE {cond}
        GROUP BY month ORDER BY month
    """, vals, as_dict=1)

    # Approval status breakdown
    approval_breakdown = frappe.db.sql(f"""
        SELECT approval_status as status, COUNT(*) as count
        FROM `tabScheme Request`
        WHERE {cond}
        GROUP BY approval_status
    """, vals, as_dict=1)

    # Top 10 doctors by scheme value
    top_doctors = frappe.db.sql(f"""
        SELECT doctor_name, COALESCE(SUM(total_scheme_value), 0) as total_value, COUNT(*) as count
        FROM `tabScheme Request`
        WHERE {cond} AND doctor_name IS NOT NULL AND doctor_name != ''
        GROUP BY doctor_code ORDER BY total_value DESC LIMIT 10
    """, vals, as_dict=1)

    # Top 10 HQs by scheme value
    top_hqs = frappe.db.sql(f"""
        SELECT hq, COALESCE(SUM(total_scheme_value), 0) as total_value, COUNT(*) as count
        FROM `tabScheme Request`
        WHERE {cond} AND hq IS NOT NULL AND hq != ''
        GROUP BY hq ORDER BY total_value DESC LIMIT 10
    """, vals, as_dict=1)

    # Scheme value by region
    region_breakdown = frappe.db.sql(f"""
        SELECT region, COALESCE(SUM(total_scheme_value), 0) as total_value, COUNT(*) as count
        FROM `tabScheme Request`
        WHERE {cond} AND region IS NOT NULL AND region != ''
        GROUP BY region ORDER BY total_value DESC
    """, vals, as_dict=1)

    # KPI: totals
    kpi = frappe.db.sql(f"""
        SELECT COUNT(*) as total_schemes,
               SUM(CASE WHEN approval_status='Approved' THEN 1 ELSE 0 END) as approved,
               COALESCE(SUM(total_scheme_value), 0) as total_value
        FROM `tabScheme Request`
        WHERE {cond}
    """, vals, as_dict=1)[0]

    return {
        "monthly_trend": monthly_trend,
        "approval_breakdown": approval_breakdown,
        "top_doctors": top_doctors,
        "top_hqs": top_hqs,
        "region_breakdown": region_breakdown,
        "kpi": kpi,
    }


@frappe.whitelist()
def get_insights_statement_data(division=None, from_date=None, to_date=None, region=None, team=None, hq=None):
    """Stock statement analytics — submissions, coverage, value trends."""
    if not division:
        division = get_user_division()

    cond, vals = _build_insights_conditions(division, from_date, to_date, region, team, hq, "statement_month")

    # Monthly statement count
    monthly_statements = frappe.db.sql(f"""
        SELECT DATE_FORMAT(statement_month, '%%Y-%%m') as month, COUNT(*) as count
        FROM `tabStockist Statement`
        WHERE {cond}
        GROUP BY month ORDER BY month
    """, vals, as_dict=1)

    # Closing value trend
    closing_value_trend = frappe.db.sql(f"""
        SELECT DATE_FORMAT(statement_month, '%%Y-%%m') as month,
               COALESCE(SUM(total_closing_value), 0) as total_closing
        FROM `tabStockist Statement`
        WHERE {cond}
        GROUP BY month ORDER BY month
    """, vals, as_dict=1)

    # Coverage by HQ — unique stockists who submitted per HQ
    hq_coverage = frappe.db.sql(f"""
        SELECT hq, COUNT(DISTINCT stockist_code) as stockist_count
        FROM `tabStockist Statement`
        WHERE {cond} AND hq IS NOT NULL AND hq != ''
        GROUP BY hq ORDER BY stockist_count DESC LIMIT 15
    """, vals, as_dict=1)

    # Extraction status
    extraction_status = frappe.db.sql(f"""
        SELECT COALESCE(extracted_data_status, 'Pending') as status, COUNT(*) as count
        FROM `tabStockist Statement`
        WHERE {cond}
        GROUP BY extracted_data_status
    """, vals, as_dict=1)

    # KPI
    kpi = frappe.db.sql(f"""
        SELECT COUNT(*) as total_statements,
               COUNT(DISTINCT stockist_code) as unique_stockists,
               COALESCE(SUM(total_closing_value), 0) as total_closing_value
        FROM `tabStockist Statement`
        WHERE {cond}
    """, vals, as_dict=1)[0]

    return {
        "monthly_statements": monthly_statements,
        "closing_value_trend": closing_value_trend,
        "hq_coverage": hq_coverage,
        "extraction_status": extraction_status,
        "kpi": kpi,
    }


@frappe.whitelist()
def get_insights_deduction_data(division=None, from_date=None, to_date=None, region=None, team=None, hq=None):
    """Scheme deduction analytics — monthly value, top stockists, status breakdown."""
    if not division:
        division = get_user_division()

    # Deduction table doesn't have region/team/hq directly — join via scheme_request
    base_cond = ["1=1"]
    base_vals = []
    if division:
        base_cond.append("sd.division = %s")
        base_vals.append(division)
    if from_date:
        base_cond.append("sd.deduction_date >= %s")
        base_vals.append(from_date)
    if to_date:
        base_cond.append("sd.deduction_date <= %s")
        base_vals.append(to_date)
    _scope_region_sql_pos(base_cond, base_vals, "sr.region", division, region)
    if team:
        base_cond.append("sr.team = %s")
        base_vals.append(team)
    if hq:
        base_cond.append("sr.hq = %s")
        base_vals.append(hq)

    cond_str = " AND ".join(base_cond)

    # Monthly deduction value
    monthly_deductions = frappe.db.sql(f"""
        SELECT DATE_FORMAT(sd.deduction_date, '%%Y-%%m') as month,
               COUNT(*) as count,
               COALESCE(SUM(sd.total_deducted_value), 0) as total_value
        FROM `tabScheme Deduction` sd
        LEFT JOIN `tabScheme Request` sr ON sr.name = sd.scheme_request
        WHERE {cond_str}
        GROUP BY month ORDER BY month
    """, base_vals, as_dict=1)

    # Top stockists by deduction value
    top_stockists = frappe.db.sql(f"""
        SELECT sm.stockist_name, sd.stockist_code,
               COALESCE(SUM(sd.total_deducted_value), 0) as total_value,
               COUNT(*) as count
        FROM `tabScheme Deduction` sd
        LEFT JOIN `tabScheme Request` sr ON sr.name = sd.scheme_request
        LEFT JOIN `tabStockist Master` sm ON sm.name = sd.stockist_code
        WHERE {cond_str} AND sd.stockist_code IS NOT NULL
        GROUP BY sd.stockist_code ORDER BY total_value DESC LIMIT 10
    """, base_vals, as_dict=1)

    # Status breakdown
    status_breakdown = frappe.db.sql(f"""
        SELECT sd.status, COUNT(*) as count
        FROM `tabScheme Deduction` sd
        LEFT JOIN `tabScheme Request` sr ON sr.name = sd.scheme_request
        WHERE {cond_str}
        GROUP BY sd.status
    """, base_vals, as_dict=1)

    # KPI
    kpi = frappe.db.sql(f"""
        SELECT COUNT(*) as total_deductions,
               COALESCE(SUM(sd.total_deducted_value), 0) as total_value
        FROM `tabScheme Deduction` sd
        LEFT JOIN `tabScheme Request` sr ON sr.name = sd.scheme_request
        WHERE {cond_str}
    """, base_vals, as_dict=1)[0]

    return {
        "monthly_deductions": monthly_deductions,
        "top_stockists": top_stockists,
        "status_breakdown": status_breakdown,
        "kpi": kpi,
    }


@frappe.whitelist()
def get_insights_masters_data(division=None):
    """Masters overview — stockist count by HQ, HQ distribution by region."""
    if not division:
        division = get_user_division()

    # Stockist count by HQ (top 15)
    stockists_by_hq = frappe.db.sql("""
        SELECT hq, COUNT(*) as count
        FROM `tabStockist Master`
        WHERE division = %s AND status = 'Active'
        AND hq IS NOT NULL AND hq != ''
        GROUP BY hq ORDER BY count DESC LIMIT 15
    """, (division,), as_dict=1)

    # HQ distribution by region
    hqs_by_region = frappe.db.sql("""
        SELECT region, COUNT(*) as count
        FROM `tabHQ Master`
        WHERE division = %s AND status = 'Active'
        AND region IS NOT NULL AND region != ''
        GROUP BY region ORDER BY count DESC
    """, (division,), as_dict=1)

    # Doctors by HQ (top 15)
    doctors_by_hq = frappe.db.sql("""
        SELECT hq, COUNT(*) as count
        FROM `tabDoctor Master`
        WHERE division = %s AND status = 'Active'
        AND hq IS NOT NULL AND hq != ''
        GROUP BY hq ORDER BY count DESC LIMIT 15
    """, (division,), as_dict=1)

    # KPIs
    kpi = {
        "active_stockists": frappe.db.count("Stockist Master", {"division": division, "status": "Active"}),
        "active_hqs": frappe.db.count("HQ Master", {"division": division, "status": "Active"}),
        "active_doctors": frappe.db.count("Doctor Master", {"division": division, "status": "Active"}),
        "active_products": frappe.db.count("Product Master", {"division": division, "status": "Active"}),
    }

    return {
        "stockists_by_hq": stockists_by_hq,
        "hqs_by_region": hqs_by_region,
        "doctors_by_hq": doctors_by_hq,
        "kpi": kpi,
    }


@frappe.whitelist()
def get_insights_targets_data(division=None, financial_year=None):
    """Sales targets vs actuals — monthly comparison, region-wise breakdown."""
    if not division:
        division = get_user_division()

    if not financial_year:
        fy = frappe.db.sql(
            "SELECT financial_year FROM `tabHQ Yearly Target` WHERE division=%s ORDER BY financial_year DESC LIMIT 1",
            (division,), as_list=1
        )
        financial_year = fy[0][0] if fy else None

    if not financial_year:
        return {"monthly": [], "region_wise": [], "kpi": {"total_target": 0, "total_actual": 0}, "financial_year": None}

    # Monthly targets
    months_map = [
        ("apr", "Apr"), ("may", "May"), ("jun", "Jun"),
        ("jul", "Jul"), ("aug", "Aug"), ("sep", "Sep"),
        ("oct", "Oct"), ("nov", "Nov"), ("dec", "Dec"),
        ("jan", "Jan"), ("feb", "Feb"), ("mar", "Mar"),
    ]

    target_sums = frappe.db.sql("""
        SELECT
            COALESCE(SUM(ti.apr),0) as apr, COALESCE(SUM(ti.may),0) as may, COALESCE(SUM(ti.jun),0) as jun,
            COALESCE(SUM(ti.jul),0) as jul, COALESCE(SUM(ti.aug),0) as aug, COALESCE(SUM(ti.sep),0) as sep,
            COALESCE(SUM(ti.oct),0) as oct, COALESCE(SUM(ti.nov),0) as nov, COALESCE(SUM(ti.dec),0) as dec,
            COALESCE(SUM(ti.jan),0) as jan, COALESCE(SUM(ti.feb),0) as feb, COALESCE(SUM(ti.mar),0) as mar,
            COALESCE(SUM(ti.yearly_total),0) as yearly_total
        FROM `tabHQ Yearly Target` yt
        INNER JOIN `tabHQ Target Item` ti ON ti.parent = yt.name
        WHERE yt.docstatus = 1 AND yt.division = %s AND yt.financial_year = %s
    """, (division, financial_year), as_dict=1)[0]

    # Build monthly actuals from scheme requests (approved value by month)
    # Financial year format is like "2025-26" — parse start year
    try:
        start_year = int(financial_year.split("-")[0])
    except Exception:
        start_year = 2025

    month_to_date = {
        "apr": f"{start_year}-04", "may": f"{start_year}-05", "jun": f"{start_year}-06",
        "jul": f"{start_year}-07", "aug": f"{start_year}-08", "sep": f"{start_year}-09",
        "oct": f"{start_year}-10", "nov": f"{start_year}-11", "dec": f"{start_year}-12",
        "jan": f"{start_year+1}-01", "feb": f"{start_year+1}-02", "mar": f"{start_year+1}-03",
    }

    actual_data = frappe.db.sql("""
        SELECT DATE_FORMAT(application_date, '%%Y-%%m') as month,
               COALESCE(SUM(total_scheme_value), 0) as actual_value
        FROM `tabScheme Request`
        WHERE division = %s AND approval_status = 'Approved'
        AND application_date >= %s AND application_date <= %s
        GROUP BY month
    """, (division, f"{start_year}-04-01", f"{start_year+1}-03-31"), as_dict=1)

    actual_map = {a["month"]: a["actual_value"] for a in actual_data}

    monthly = []
    for key, label in months_map:
        monthly.append({
            "month": label,
            "target": flt(getattr(target_sums, key, 0) if hasattr(target_sums, key) else target_sums.get(key, 0)),
            "actual": flt(actual_map.get(month_to_date.get(key, ""), 0)),
        })

    # Region-wise target vs actual
    region_targets = frappe.db.sql("""
        SELECT yt.region,
               COALESCE(SUM(ti.yearly_total), 0) as target_value
        FROM `tabHQ Yearly Target` yt
        INNER JOIN `tabHQ Target Item` ti ON ti.parent = yt.name
        WHERE yt.docstatus = 1 AND yt.division = %s AND yt.financial_year = %s
        GROUP BY yt.region ORDER BY target_value DESC
    """, (division, financial_year), as_dict=1)

    region_actuals = frappe.db.sql("""
        SELECT region, COALESCE(SUM(total_scheme_value), 0) as actual_value
        FROM `tabScheme Request`
        WHERE division = %s AND approval_status = 'Approved'
        AND application_date >= %s AND application_date <= %s
        AND region IS NOT NULL AND region != ''
        GROUP BY region
    """, (division, f"{start_year}-04-01", f"{start_year+1}-03-31"), as_dict=1)

    actual_region_map = {a["region"]: a["actual_value"] for a in region_actuals}

    region_wise = []
    for rt in region_targets:
        region_wise.append({
            "region": rt["region"],
            "target": flt(rt["target_value"]),
            "actual": flt(actual_region_map.get(rt["region"], 0)),
        })

    total_target = flt(target_sums.get("yearly_total", 0))
    total_actual = sum(flt(actual_map.get(month_to_date.get(k, ""), 0)) for k, _ in months_map)

    return {
        "monthly": monthly,
        "region_wise": region_wise,
        "kpi": {"total_target": total_target, "total_actual": total_actual},
        "financial_year": financial_year,
    }


@frappe.whitelist()
def get_insights_products_data(division=None, from_date=None, to_date=None, region=None, team=None, hq=None):
    """Product movement insights — top products by closing value and by scheme value."""
    if not division:
        division = get_user_division()

    # Top products by closing value from statements
    stmt_cond = ["1=1"]
    stmt_vals = []
    if division:
        stmt_cond.append("ss.division = %s")
        stmt_vals.append(division)
    if from_date:
        stmt_cond.append("ss.statement_month >= %s")
        stmt_vals.append(from_date)
    if to_date:
        stmt_cond.append("ss.statement_month <= %s")
        stmt_vals.append(to_date)
    _scope_region_sql_pos(stmt_cond, stmt_vals, "ss.region", division, region)
    if team:
        stmt_cond.append("ss.team = %s")
        stmt_vals.append(team)
    if hq:
        stmt_cond.append("ss.hq = %s")
        stmt_vals.append(hq)

    stmt_where = " AND ".join(stmt_cond)

    top_products_closing = frappe.db.sql(f"""
        SELECT si.product_name, si.product_code,
               COALESCE(SUM(si.closing_value), 0) as total_closing,
               COALESCE(SUM(si.sales_qty / IFNULL(NULLIF(si.conversion_factor, 0), 1)), 0) as total_sales_qty
        FROM `tabStockist Statement Item` si
        INNER JOIN `tabStockist Statement` ss ON ss.name = si.parent
        WHERE {stmt_where} AND si.product_name IS NOT NULL AND si.product_name != ''
        GROUP BY si.product_code ORDER BY total_closing DESC LIMIT 15
    """, stmt_vals, as_dict=1)

    # Top products by scheme value
    sch_cond = ["1=1"]
    sch_vals = []
    if division:
        sch_cond.append("sr.division = %s")
        sch_vals.append(division)
    if from_date:
        sch_cond.append("sr.application_date >= %s")
        sch_vals.append(from_date)
    if to_date:
        sch_cond.append("sr.application_date <= %s")
        sch_vals.append(to_date)
    _scope_region_sql_pos(sch_cond, sch_vals, "sr.region", division, region)
    if team:
        sch_cond.append("sr.team = %s")
        sch_vals.append(team)
    if hq:
        sch_cond.append("sr.hq = %s")
        sch_vals.append(hq)

    sch_where = " AND ".join(sch_cond)

    top_products_scheme = frappe.db.sql(f"""
        SELECT sri.product_name, sri.product_code,
               COALESCE(SUM(sri.product_value), 0) as total_value,
               COALESCE(SUM(sri.quantity), 0) as total_qty
        FROM `tabScheme Request Item` sri
        INNER JOIN `tabScheme Request` sr ON sr.name = sri.parent
        WHERE {sch_where} AND sri.product_name IS NOT NULL AND sri.product_name != ''
        GROUP BY sri.product_code ORDER BY total_value DESC LIMIT 15
    """, sch_vals, as_dict=1)

    _apply_product_display_codes(top_products_closing)
    _apply_product_display_codes(top_products_scheme)

    return {
        "top_products_closing": top_products_closing,
        "top_products_scheme": top_products_scheme,
    }


# ============================================================
#  PORTAL – USER MANAGEMENT & PROFILE APIs
# ============================================================

@frappe.whitelist()
def get_portal_users():
    """Return list of all non-Guest users for the Users management page (System Manager only)."""


    from scanify.permissions import get_portal_role
    if not (get_portal_role() == "Admin"):
        frappe.throw("Not permitted", frappe.PermissionError)

    users = frappe.get_all(
        "User",
        filters={"name": ["not in", ["Guest", "Administrator"]]},
        fields=["name", "email", "first_name", "middle_name", "last_name", "full_name",
                "mobile_no", "user_image", "enabled", "division",
                "portal_role", "allowed_divisions", "allowed_regions",
                "scheme_to_email", "scheme_cc_emails"],
        order_by="full_name asc",
        limit_page_length=500,
    )

    for u in users:
        u["role"] = u.get("portal_role") or get_portal_role(u["name"])
        u["division"] = u.get("division") or ""

    return users


def _norm_csv(v):
    """Normalise a list / JSON-array-string / comma-or-newline string to a clean list."""
    if not v:
        return []
    if isinstance(v, str):
        v = v.strip()
        if v.startswith("["):
            try:
                v = json.loads(v)
            except Exception:
                v = v.split(",")
        else:
            v = v.replace("\n", ",").split(",")
    return [str(x).strip() for x in v if str(x).strip()]


@frappe.whitelist()
def create_portal_user(email, first_name, last_name=None, middle_name=None,
                       role="Regional User", division=None, mobile_no=None, password=None,
                       scheme_to_email=None, scheme_cc_emails=None,
                       allowed_divisions=None, allowed_regions=None):
    """Create a new portal user. Restricted to portal Admin."""
    from scanify.permissions import is_portal_admin, PORTAL_ROLES
    if not is_portal_admin():
        frappe.throw("Not permitted", frappe.PermissionError)

    if not email or not first_name or not password:
        frappe.throw("Email, First Name and Password are required")
    if role not in PORTAL_ROLES:
        frappe.throw(f"Invalid role: {role}")
    if frappe.db.exists("User", email):
        frappe.throw(f"User {email} already exists")

    divs = _norm_csv(allowed_divisions)
    regions = _norm_csv(allowed_regions)
    if role != "Admin" and (not divs or not regions):
        frappe.throw("Division and Region mapping are required for non-admin roles")
    active_division = division or (divs[0] if divs else "Prima")

    user = frappe.get_doc({
        "doctype": "User",
        "email": email,
        "first_name": first_name,
        "middle_name": middle_name or "",
        "last_name": last_name or "",
        "mobile_no": mobile_no or "",
        "division": active_division,
        "portal_role": role,
        "allowed_divisions": ", ".join(divs),
        "allowed_regions": ", ".join(regions),
        "scheme_to_email": (scheme_to_email or "").strip(),
        "scheme_cc_emails": (scheme_cc_emails or "").strip(),
        "enabled": 1,
        "send_welcome_email": 0,
        "user_type": "System User",
    })
    user.insert(ignore_permissions=True)

    from frappe.utils.password import update_password as _update_password
    _update_password(email, password)

    # Grant the underlying Frappe permissions this portal role needs.
    from scanify.permissions import sync_frappe_roles
    sync_frappe_roles(email, role, prune=False)
    frappe.db.commit()

    return {"success": True, "user": email, "message": f"User {email} created successfully"}


@frappe.whitelist()
def update_portal_user(email, first_name=None, last_name=None, middle_name=None,
                       role=None, division=None, mobile_no=None,
                       scheme_to_email=None, scheme_cc_emails=None,
                       allowed_divisions=None, allowed_regions=None, password=None):
    """Update an existing portal user (role, division/region mapping, scheme To/CC,
    and optional password reset). Restricted to portal Admin."""
    from scanify.permissions import is_portal_admin, PORTAL_ROLES
    if not is_portal_admin():
        frappe.throw("Not permitted", frappe.PermissionError)
    if not email or not frappe.db.exists("User", email):
        frappe.throw(f"User {email} not found")

    user = frappe.get_doc("User", email)
    if first_name is not None:
        user.first_name = first_name
    if middle_name is not None:
        user.middle_name = middle_name or ""
    if last_name is not None:
        user.last_name = last_name or ""
    if mobile_no is not None:
        user.mobile_no = mobile_no or ""
    if scheme_to_email is not None:
        user.scheme_to_email = (scheme_to_email or "").strip()
    if scheme_cc_emails is not None:
        user.scheme_cc_emails = (scheme_cc_emails or "").strip()
    if role is not None:
        if role not in PORTAL_ROLES:
            frappe.throw(f"Invalid role: {role}")
        user.portal_role = role
    if allowed_divisions is not None:
        user.allowed_divisions = ", ".join(_norm_csv(allowed_divisions))
    if allowed_regions is not None:
        user.allowed_regions = ", ".join(_norm_csv(allowed_regions))

    # Non-admins must have division + region mapping; keep the active division valid.
    divs = _norm_csv(user.allowed_divisions)
    eff_role = user.portal_role
    if eff_role and eff_role != "Admin" and (not divs or not _norm_csv(user.allowed_regions)):
        frappe.throw("Division and Region mapping are required for non-admin roles")
    if division is not None:
        user.division = division
    elif divs and user.division not in divs:
        user.division = divs[0]

    user.save(ignore_permissions=True)

    # Optional password reset by the portal admin.
    if password:
        if len(password) < 8:
            frappe.throw("Password must be at least 8 characters")
        from frappe.utils.password import update_password as _update_password
        _update_password(email, password)

    # Keep the underlying Frappe permissions in sync with the portal role.
    if role is not None:
        from scanify.permissions import sync_frappe_roles
        sync_frappe_roles(email, user.portal_role, prune=True)

    frappe.db.commit()
    return {"success": True, "user": email, "message": f"User {email} updated successfully"}


@frappe.whitelist()
def delete_portal_user(email):
    """Delete a portal user; if the user has linked records (e.g. scheme requests),
    disable instead. Restricted to portal Admin."""
    from scanify.permissions import is_portal_admin
    if not is_portal_admin():
        frappe.throw("Not permitted", frappe.PermissionError)
    if email in ("Administrator", frappe.session.user):
        frappe.throw("You cannot delete this user")
    if not frappe.db.exists("User", email):
        frappe.throw(f"User {email} not found")
    try:
        frappe.delete_doc("User", email, ignore_permissions=True)
        frappe.db.commit()
        return {"success": True, "mode": "deleted", "message": f"User {email} deleted"}
    except Exception:
        frappe.db.rollback()
        frappe.db.set_value("User", email, "enabled", 0, update_modified=False)
        frappe.db.commit()
        return {"success": True, "mode": "disabled",
                "message": f"User {email} has linked records — disabled instead of deleted"}


@frappe.whitelist()
def set_portal_user_enabled(email, enabled):
    """Enable/disable a portal user. Restricted to portal Admin."""
    from scanify.permissions import is_portal_admin
    if not is_portal_admin():
        frappe.throw("Not permitted", frappe.PermissionError)
    on = 1 if str(enabled).lower() in ("1", "true", "yes") else 0
    frappe.db.set_value("User", email, "enabled", on, update_modified=False)
    frappe.db.commit()
    return {"success": True, "enabled": on}


_PORTAL_ROLE_LABELS = {
    "Admin": "Administrator",
    "HO": "Head Office (HO)",
    "Regional User": "Regional User",
    "Regional User (Future)": "Regional User (Future)",
}


@frappe.whitelist()
def get_my_profile():
    """Return the current user's profile data, including the portal role (the custom
    role mapping), the divisions they may use, and the regions they are scoped to."""
    from scanify.permissions import (
        get_portal_role, is_portal_admin, get_allowed_divisions,
        get_allowed_region_codes,
    )

    user = frappe.session.user
    if user == "Guest":
        frappe.throw("Please login to continue", frappe.PermissionError)

    doc = frappe.get_doc("User", user)

    portal_role = get_portal_role(user)
    admin = is_portal_admin(user)
    divisions = get_allowed_divisions(user) or []

    # Regions the user is mapped to (read-only). Admins are not region-scoped.
    if admin:
        regions = []
    else:
        codes = get_allowed_region_codes(user) or []
        region_rows = frappe.get_all(
            "Region Master", filters={"name": ["in", codes or [""]]},
            fields=["name", "region_name"], order_by="region_name") if codes else []
        regions = [{"code": r["name"], "name": r["region_name"] or r["name"]} for r in region_rows]

    return {
        "email": doc.email,
        "first_name": doc.first_name or "",
        "middle_name": doc.middle_name or "",
        "last_name": doc.last_name or "",
        "full_name": doc.full_name or "",
        "mobile_no": doc.mobile_no or "",
        "user_image": doc.user_image or "",
        "role": portal_role,
        "role_label": _PORTAL_ROLE_LABELS.get(portal_role, portal_role),
        "is_admin": admin,
        "division": doc.division or (divisions[0] if divisions else ""),
        "divisions": divisions,
        "regions": regions,
    }


@frappe.whitelist()
def change_my_password(current_password, new_password):
    """Let a logged-in user change their own password after verifying the current one."""
    from frappe.utils.password import check_password, update_password as _update_password

    user = frappe.session.user
    if user == "Guest":
        frappe.throw("Please login to continue", frappe.PermissionError)

    current_password = (current_password or "").strip()
    new_password = (new_password or "").strip()

    if not current_password or not new_password:
        return {"success": False, "message": "Both current and new password are required."}
    if len(new_password) < 6:
        return {"success": False, "message": "New password must be at least 6 characters."}
    if new_password == current_password:
        return {"success": False, "message": "New password must be different from the current one."}

    # Verify the current password (raises AuthenticationError if wrong).
    try:
        check_password(user, current_password)
    except frappe.AuthenticationError:
        return {"success": False, "message": "Current password is incorrect."}

    # Enforce the site's password strength policy, if enabled. The core helper
    # returns {} when the policy is off, and otherwise a result whose feedback says
    # whether the score passed. Best-effort: never block on a policy-lookup error.
    try:
        from frappe.core.doctype.user.user import test_password_strength
        result = test_password_strength(new_password)
        feedback = (result or {}).get("feedback") or {}
        if result and not feedback.get("password_policy_validation_passed", True):
            parts = [feedback.get("warning") or "Please choose a stronger password."]
            parts += (feedback.get("suggestions") or [])
            return {"success": False, "message": " ".join(p for p in parts if p).strip()}
    except Exception:
        pass

    _update_password(user, new_password)
    frappe.db.commit()
    return {"success": True, "message": "Password changed successfully."}


@frappe.whitelist()
def update_my_profile(first_name, middle_name=None, last_name=None, mobile_no=None):
    """Update the current user's own profile fields."""
    user = frappe.session.user
    if user == "Guest":
        frappe.throw("Please login to continue", frappe.PermissionError)

    doc = frappe.get_doc("User", user)
    doc.first_name = first_name
    doc.middle_name = middle_name or ""
    doc.last_name = last_name or ""
    doc.mobile_no = mobile_no or ""
    doc.save(ignore_permissions=True)
    frappe.db.commit()
    return {"success": True, "message": "Profile updated successfully"}


@frappe.whitelist()
def update_user_image(file_url):
    """Set the current user's profile image after upload."""
    user = frappe.session.user
    if user == "Guest":
        frappe.throw("Please login to continue", frappe.PermissionError)

    if not file_url:
        frappe.throw("No file URL provided")

    doc = frappe.get_doc("User", user)
    doc.user_image = file_url
    doc.save(ignore_permissions=True)
    frappe.db.commit()
    return {"success": True, "user_image": file_url, "message": "Profile image updated"}


# ───────────────────────────────────────────────────────────────
# Audit Trail / Change History — Portal API
# ───────────────────────────────────────────────────────────────

# Category → internal doctype(s) mapping. Order here drives the portal dropdown.
# Keep this in sync with every track_changes doctype so nothing is silently missed.
_AUDIT_CATEGORY_MAP = {
    "Masters": [
        "HQ Master", "Stockist Master", "Product Master",
        "Doctor Master", "Team Master", "Region Master",
        "Zone Master", "State Master",
    ],
    "Stock Statements": ["Stockist Statement"],
    "Scheme Requests": ["Scheme Request"],
    "Scheme Deductions": ["Scheme Deduction"],
    "Sales Targets": ["HQ Yearly Target"],
    "Bulk Statement Upload": ["Bulk Statement Upload"],
    "Primary Sales": ["Primary Sales Upload"],
    "Secondary Sales": ["Secondary Sales Upload"],
    "Stockist Corrections": ["Stockist Product Correction"],
}

# Masters sub-type label → internal doctype
_MASTER_SUBTYPE_MAP = {
    "HQ": "HQ Master",
    "Stockist": "Stockist Master",
    "Product": "Product Master",
    "Doctor": "Doctor Master",
    "Team": "Team Master",
    "Region": "Region Master",
    "Zone": "Zone Master",
    "State": "State Master",
}

# Reverse: doctype → user-facing label
_DOCTYPE_LABEL_MAP = {
    "HQ Master": "HQ Master",
    "Stockist Master": "Stockist Master",
    "Product Master": "Product Master",
    "Doctor Master": "Doctor Master",
    "Team Master": "Team Master",
    "Region Master": "Region Master",
    "Zone Master": "Zone Master",
    "State Master": "State Master",
    "Stockist Statement": "Stock Statement",
    "Scheme Request": "Scheme Request",
    "Scheme Deduction": "Scheme Deduction",
    "HQ Yearly Target": "Sales Target",
    "Bulk Statement Upload": "Bulk Upload",
    "Primary Sales Upload": "Primary Sales",
    "Secondary Sales Upload": "Secondary Sales",
    "Stockist Product Correction": "Stockist Correction",
}

# Fields to hide from diffs (system / internal)
_SYSTEM_FIELDS = {
    "modified", "modified_by", "creation", "owner", "docstatus",
    "idx", "doctype", "name", "parent", "parenttype", "parentfield",
    "_liked_by", "_comments", "_assign", "_user_tags",
    "_seen", "amended_from",
}

# Every action the unified feed can surface.
_AUDIT_ACTIONS = ["Created", "Updated", "Submitted", "Cancelled", "Deleted"]

# How far back the feed looks when no explicit From date is chosen, and a hard
# per-source row cap so a broad query can never pull the whole history at once.
_AUDIT_DEFAULT_WINDOW_DAYS = 90
_AUDIT_SOURCE_CAP = 3000


def _all_audit_doctypes():
    """Flat list of every doctype the audit feed tracks."""
    out = []
    for dts in _AUDIT_CATEGORY_MAP.values():
        out.extend(dts)
    return out


def _doctype_has_division(dt):
    """True if the doctype carries a `division` field (drives division scoping)."""
    try:
        return bool(frappe.get_meta(dt).has_field("division"))
    except Exception:
        return False


def _in_clause(prefix, values, params):
    """Build a parameterised `(%(p0)s, %(p1)s, ...)` IN clause and load params."""
    keys = []
    for i, v in enumerate(values):
        k = f"{prefix}{i}"
        params[k] = v
        keys.append(f"%({k})s")
    return "(" + ", ".join(keys) + ")" if keys else "(NULL)"


def _humanize(s):
    return (s or "").replace("_", " ").strip().title()


def _fmt_val(v):
    """Render a diff value as a short display string."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    return str(v)


def _audit_user_info(emails):
    """{email: {'name': display, 'role': portal_role}} for a set of change authors."""
    from scanify.permissions import get_portal_role
    info = {}
    emails = {e for e in emails if e}
    if not emails:
        return info
    names = {}
    for r in frappe.get_all("User", filters={"name": ["in", list(emails)]},
                            fields=["name", "full_name"]):
        names[r.name] = (r.full_name or "").strip()
    for e in emails:
        disp = names.get(e) or (e.split("@")[0].title() if "@" in e else e)
        try:
            role = get_portal_role(e) or "—"
        except Exception:
            role = "—"
        info[e] = {"name": disp, "role": role}
    return info


def _division_allowed_names(names_by_dt, division):
    """For each division-scoped doctype, the subset of the given record names that
    belong to the active division (or 'Both'). Doctypes without a division field are
    absent from the result and therefore never filtered out."""
    allowed = {}
    if not division:
        return allowed
    for dt, names in names_by_dt.items():
        names = [n for n in names if n]
        if not names or not _doctype_has_division(dt):
            continue
        params = {"div": division}
        in_names = _in_clause("n", names, params)
        try:
            rows = frappe.db.sql(
                f"SELECT name FROM `tab{dt}` WHERE name IN {in_names} "
                f"AND (division IN (%(div)s, 'Both') OR COALESCE(division, '') = '')",
                params, as_dict=True)
            allowed[dt] = {x.name for x in rows}
        except Exception:
            # If the scope query fails, don't hide rows (fail-open, admin-only page).
            continue
    return allowed


def _deleted_doc_division(data):
    """Pull the `division` value out of a Deleted Document snapshot, if present."""
    try:
        d = json.loads(data) if isinstance(data, str) else (data or {})
        if isinstance(d, dict):
            return d.get("division")
    except Exception:
        pass
    return None


def _get_user_display(email):
    """Return a display-friendly name for an email address."""
    if not email:
        return "System"
    full = frappe.db.get_value("User", email, "full_name")
    if full and full.strip():
        return full.strip()
    return email.split("@")[0].title()


def _parse_version_data(data_str):
    """Parse a Version record's data JSON into its four diff buckets. Frappe stores
    field edits in `changed`, child-row additions/removals in `added`/`removed`, and
    child-row field edits in `row_changed` (previously ignored — a real blind spot)."""
    try:
        data = json.loads(data_str) if isinstance(data_str, str) else (data_str or {})
    except Exception:
        data = {}
    if not isinstance(data, dict):
        data = {}
    return {
        "changed": data.get("changed") or [],
        "added": data.get("added") or [],
        "removed": data.get("removed") or [],
        "row_changed": data.get("row_changed") or [],
    }


def _version_action(parsed):
    """Classify a Version as Submitted / Cancelled (docstatus moved) or Updated."""
    for ch in parsed.get("changed", []):
        if ch and ch[0] == "docstatus":
            try:
                new = int(ch[2])
            except Exception:
                new = None
            if new == 1:
                return "Submitted"
            if new == 2:
                return "Cancelled"
    return "Updated"


def _version_change_count(parsed):
    """Number of user-visible changes in a Version (system fields excluded)."""
    n = 0
    for ch in parsed.get("changed", []):
        if ch and ch[0] not in _SYSTEM_FIELDS:
            n += 1
    n += len(parsed.get("added", []))
    n += len(parsed.get("removed", []))
    for rc in parsed.get("row_changed", []):
        try:
            child = rc[3] or []
        except Exception:
            child = []
        meaningful = len([c for c in child if c and c[0] not in _SYSTEM_FIELDS])
        n += meaningful or 1
    return n


def _child_doctype(parent_dt, table_field):
    """Resolve a child table field's target doctype for nicer child-field labels."""
    try:
        df = frappe.get_meta(parent_dt).get_field(table_field)
        return df.options if df else None
    except Exception:
        return None


def _row_change_entry(row, kind):
    """Diff entry for a whole child row being added or removed."""
    table_field = row[0] if isinstance(row, (list, tuple)) and row else ""
    row_data = row[1] if isinstance(row, (list, tuple)) and len(row) > 1 and isinstance(row[1], dict) else {}
    summary = ", ".join(
        f"{k}: {v}" for k, v in row_data.items()
        if k not in _SYSTEM_FIELDS and not k.startswith("_") and v not in (None, "", 0)
    )
    return {
        "label": f"{_humanize(table_field)} — row {'added' if kind == 'added' else 'removed'}",
        "old_value": (summary[:250] if kind == "removed" else ""),
        "new_value": (summary[:250] if kind == "added" else ""),
        "type": kind,
    }


def _row_changed_entries(parent_dt, rc):
    """Diff entries for field-level edits inside an existing child row."""
    out = []
    try:
        table_field = rc[0]
        child_dt = _child_doctype(parent_dt, table_field)
        for c in (rc[3] or []):
            if not c or c[0] in _SYSTEM_FIELDS:
                continue
            child_label = _field_label(child_dt, c[0]) if child_dt else _humanize(c[0])
            out.append({
                "label": f"{_humanize(table_field)} · {child_label}",
                "old_value": _fmt_val(c[1]),
                "new_value": _fmt_val(c[2]),
                "type": "changed",
            })
    except Exception:
        pass
    return out


def _field_label(doctype, fieldname):
    """Map a field name to its human-readable label using DocType meta."""
    try:
        meta = frappe.get_meta(doctype)
        df = meta.get_field(fieldname)
        if df and df.label:
            return df.label
    except Exception:
        pass
    return fieldname.replace("_", " ").title()


def _collect_version_events(doctypes, record_name, changed_by, dt_from, dt_to, division):
    """Updated / Submitted / Cancelled events from `tabVersion`."""
    params = {"f": dt_from, "t": dt_to}
    conds = [
        f"v.ref_doctype IN {_in_clause('vd', doctypes, params)}",
        "v.creation BETWEEN %(f)s AND %(t)s",
    ]
    if record_name:
        conds.append("v.docname LIKE %(rec)s")
        params["rec"] = f"%{record_name}%"
    if changed_by:
        conds.append("v.owner = %(cb)s")
        params["cb"] = changed_by
    where = " AND ".join(conds)
    rows = frappe.db.sql(
        f"SELECT v.name, v.ref_doctype, v.docname, v.owner, v.creation, v.data "
        f"FROM `tabVersion` v WHERE {where} "
        f"ORDER BY v.creation DESC LIMIT {int(_AUDIT_SOURCE_CAP)}",
        params, as_dict=True) or []

    # Division scope: keep only rows whose parent record is in the active division.
    names_by_dt = {}
    for r in rows:
        names_by_dt.setdefault(r.ref_doctype, set()).add(r.docname)
    allowed = _division_allowed_names(names_by_dt, division)

    events = []
    for r in rows:
        if r.ref_doctype in allowed and r.docname not in allowed[r.ref_doctype]:
            continue
        parsed = _parse_version_data(r.data)
        action = _version_action(parsed)
        count = _version_change_count(parsed)
        # An "Updated" version with no user-visible change is noise — drop it.
        if action == "Updated" and count == 0:
            continue
        events.append({
            "id": "ver:" + r.name,
            "record": r.docname,
            "doctype": r.ref_doctype,
            "category": _DOCTYPE_LABEL_MAP.get(r.ref_doctype, r.ref_doctype),
            "action": action,
            "user": r.owner,
            "timestamp": str(r.creation),
            "change_count": count,
            "has_diff": True,
        })
    return events


def _collect_created_events(doctypes, record_name, changed_by, dt_from, dt_to, division):
    """Creation events pulled from each doctype's own table (Version records none)."""
    events = []
    for dt in doctypes:
        params = {"f": dt_from, "t": dt_to}
        conds = ["creation BETWEEN %(f)s AND %(t)s"]
        if record_name:
            conds.append("name LIKE %(rec)s")
            params["rec"] = f"%{record_name}%"
        if changed_by:
            conds.append("owner = %(cb)s")
            params["cb"] = changed_by
        if division and _doctype_has_division(dt):
            conds.append("(division IN (%(div)s, 'Both') OR COALESCE(division, '') = '')")
            params["div"] = division
        where = " AND ".join(conds)
        try:
            rows = frappe.db.sql(
                f"SELECT name, owner, creation FROM `tab{dt}` WHERE {where} "
                f"ORDER BY creation DESC LIMIT {int(_AUDIT_SOURCE_CAP)}",
                params, as_dict=True) or []
        except Exception:
            continue
        label = _DOCTYPE_LABEL_MAP.get(dt, dt)
        for r in rows:
            events.append({
                "id": f"new:{dt}::{r.name}",
                "record": r.name,
                "doctype": dt,
                "category": label,
                "action": "Created",
                "user": r.owner,
                "timestamp": str(r.creation),
                "change_count": 0,
                "has_diff": True,
            })
    return events


def _collect_deleted_events(doctypes, record_name, changed_by, dt_from, dt_to, division):
    """Deletion events from `tabDeleted Document` (the audited row itself is gone)."""
    params = {"f": dt_from, "t": dt_to}
    conds = [
        f"deleted_doctype IN {_in_clause('dd', doctypes, params)}",
        "creation BETWEEN %(f)s AND %(t)s",
    ]
    if record_name:
        conds.append("deleted_name LIKE %(rec)s")
        params["rec"] = f"%{record_name}%"
    if changed_by:
        conds.append("owner = %(cb)s")
        params["cb"] = changed_by
    where = " AND ".join(conds)
    try:
        rows = frappe.db.sql(
            f"SELECT name, deleted_doctype, deleted_name, owner, creation, data "
            f"FROM `tabDeleted Document` WHERE {where} "
            f"ORDER BY creation DESC LIMIT {int(_AUDIT_SOURCE_CAP)}",
            params, as_dict=True) or []
    except Exception:
        return []
    events = []
    for r in rows:
        dt = r.deleted_doctype
        # Division scope via the stored snapshot; keep rows whose division is unknown.
        if division and _doctype_has_division(dt):
            doc_div = _deleted_doc_division(r.data)
            if doc_div and doc_div not in (division, "Both"):
                continue
        events.append({
            "id": "del:" + r.name,
            "record": r.deleted_name,
            "doctype": dt,
            "category": _DOCTYPE_LABEL_MAP.get(dt, dt),
            "action": "Deleted",
            "user": r.owner,
            "timestamp": str(r.creation),
            "change_count": 0,
            "has_diff": True,
        })
    return events


@frappe.whitelist()
@require_process("audit")
def get_audit_trail_portal(
    category=None, sub_type=None, record_name=None,
    changed_by=None, role=None, action=None,
    from_date=None, to_date=None, page=1, page_size=25
):
    """Unified, paginated activity feed (create / update / submit / cancel / delete)
    across every tracked doctype, scoped to the active division. Admin-only."""
    try:
        page = max(int(page), 1)
        page_size = min(max(int(page_size), 5), 100)

        # Which doctypes?
        if category and category in _AUDIT_CATEGORY_MAP:
            doctypes = list(_AUDIT_CATEGORY_MAP[category])
            if category == "Masters" and sub_type and sub_type in _MASTER_SUBTYPE_MAP:
                doctypes = [_MASTER_SUBTYPE_MAP[sub_type]]
        else:
            doctypes = _all_audit_doctypes()
        if not doctypes:
            return {"success": True, "data": [], "total": 0,
                    "page": page, "page_size": page_size}

        # Date window — default to the recent window when no From date is supplied.
        if not from_date:
            from_date = frappe.utils.add_days(frappe.utils.nowdate(), -_AUDIT_DEFAULT_WINDOW_DAYS)
        dt_from = str(from_date) + " 00:00:00"
        dt_to = (str(to_date) + " 23:59:59") if to_date else frappe.utils.now()

        division = get_user_division()
        want = {action} if (action and action in _AUDIT_ACTIONS) else None
        cb = changed_by or None

        events = []
        if want is None or want & {"Updated", "Submitted", "Cancelled"}:
            events.extend(_collect_version_events(doctypes, record_name, cb, dt_from, dt_to, division))
        if want is None or "Created" in want:
            events.extend(_collect_created_events(doctypes, record_name, cb, dt_from, dt_to, division))
        if want is None or "Deleted" in want:
            events.extend(_collect_deleted_events(doctypes, record_name, cb, dt_from, dt_to, division))

        if want is not None:
            events = [e for e in events if e["action"] in want]

        # Resolve author display name + portal role, then apply the role filter.
        uinfo = _audit_user_info({e["user"] for e in events})
        for e in events:
            u = uinfo.get(e["user"], {})
            e["changed_by"] = u.get("name", e["user"])
            e["role"] = u.get("role", "—")
        if role:
            events = [e for e in events if e["role"] == role]

        events.sort(key=lambda x: x["timestamp"], reverse=True)

        total = len(events)
        start = (page - 1) * page_size
        page_rows = events[start:start + page_size]

        return {
            "success": True,
            "data": page_rows,
            "total": total,
            "page": page,
            "page_size": page_size,
            "window_from": str(from_date),
        }
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Audit Trail Portal Error")
        return {"success": False, "message": str(e)}


@frappe.whitelist()
@require_process("audit")
def get_audit_filter_options():
    """Dropdown data for the audit filters: portal users, roles, and actions."""
    try:
        users = frappe.get_all(
            "User", filters={"enabled": 1, "portal_role": ["is", "set"]},
            fields=["name", "full_name", "portal_role"], order_by="full_name asc")
        out_users = [{
            "email": u.name,
            "name": (u.full_name or u.name.split("@")[0].title()).strip(),
            "role": u.portal_role,
        } for u in users]
        if not any(u["email"] == "Administrator" for u in out_users):
            out_users.insert(0, {"email": "Administrator", "name": "Administrator", "role": "Admin"})
        from scanify.permissions import PORTAL_ROLES
        return {
            "success": True,
            "users": out_users,
            "roles": list(PORTAL_ROLES),
            "actions": list(_AUDIT_ACTIONS),
            "default_window_days": _AUDIT_DEFAULT_WINDOW_DAYS,
        }
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Audit Filter Options Error")
        return {"success": False, "message": str(e)}


_SNAPSHOT_SKIP_FIELDTYPES = {
    "Section Break", "Column Break", "Tab Break", "HTML", "Button",
    "Image", "Fold", "Heading", "Table", "Table MultiSelect",
}


def _detail_version(name):
    """Field & child-row diff for an Updated / Submitted / Cancelled event."""
    ver = frappe.db.get_value(
        "Version", name, ["ref_doctype", "docname", "data", "owner", "creation"],
        as_dict=True)
    if not ver:
        return {"success": False, "message": "Record not found"}
    parsed = _parse_version_data(ver.data)
    ref_dt = ver.ref_doctype
    changes = []
    for c in parsed["changed"]:
        if not c or c[0] in _SYSTEM_FIELDS:
            continue
        changes.append({
            "label": _field_label(ref_dt, c[0]),
            "old_value": _fmt_val(c[1]),
            "new_value": _fmt_val(c[2]),
            "type": "changed",
        })
    for a in parsed["added"]:
        if isinstance(a, list) and len(a) >= 2:
            changes.append(_row_change_entry(a, "added"))
    for r in parsed["removed"]:
        if isinstance(r, list) and len(r) >= 2:
            changes.append(_row_change_entry(r, "removed"))
    for rc in parsed["row_changed"]:
        changes.extend(_row_changed_entries(ref_dt, rc))
    return {
        "success": True,
        "record": ver.docname,
        "category": _DOCTYPE_LABEL_MAP.get(ref_dt, ref_dt),
        "changed_by": _get_user_display(ver.owner),
        "timestamp": str(ver.creation),
        "changes": changes,
    }


def _detail_created(dt, name):
    """Snapshot of the record as it stands now, for a Created event."""
    if dt not in _all_audit_doctypes():
        return {"success": False, "message": "Unknown type"}
    if not frappe.db.exists(dt, name):
        return {"success": True, "record": name,
                "category": _DOCTYPE_LABEL_MAP.get(dt, dt), "changes": [],
                "note": "This record has since been deleted."}
    doc = frappe.get_doc(dt, name)
    changes = []
    for df in frappe.get_meta(dt).fields:
        if df.fieldtype in _SNAPSHOT_SKIP_FIELDTYPES or df.fieldname in _SYSTEM_FIELDS:
            continue
        val = doc.get(df.fieldname)
        if val in (None, "", 0):
            continue
        changes.append({
            "label": df.label or _humanize(df.fieldname),
            "old_value": "",
            "new_value": _fmt_val(val),
            "type": "added",
        })
    return {
        "success": True,
        "record": name,
        "category": _DOCTYPE_LABEL_MAP.get(dt, dt),
        "changed_by": _get_user_display(doc.owner),
        "timestamp": str(doc.creation),
        "changes": changes,
    }


def _detail_deleted(name):
    """Snapshot of the record as it was, reconstructed from Deleted Document."""
    dd = frappe.db.get_value(
        "Deleted Document", name,
        ["deleted_doctype", "deleted_name", "data", "owner", "creation"], as_dict=True)
    if not dd:
        return {"success": False, "message": "Record not found"}
    try:
        snap = json.loads(dd.data) if isinstance(dd.data, str) else (dd.data or {})
    except Exception:
        snap = {}
    dt = dd.deleted_doctype
    has_meta = frappe.db.exists("DocType", dt)
    changes = []
    if isinstance(snap, dict):
        for k, v in snap.items():
            if k in _SYSTEM_FIELDS or k.startswith("_"):
                continue
            if isinstance(v, (list, dict)) or v in (None, "", 0):
                continue
            changes.append({
                "label": _field_label(dt, k) if has_meta else _humanize(k),
                "old_value": _fmt_val(v),
                "new_value": "",
                "type": "removed",
            })
    return {
        "success": True,
        "record": dd.deleted_name,
        "category": _DOCTYPE_LABEL_MAP.get(dt, dt),
        "changed_by": _get_user_display(dd.owner),
        "timestamp": str(dd.creation),
        "changes": changes,
    }


@frappe.whitelist()
@require_process("audit")
def get_audit_trail_detail(event_id=None, version_name=None):
    """Detail view for one feed event. `event_id` is prefixed by source:
    `ver:` (Version diff), `new:` (created snapshot), `del:` (deleted snapshot)."""
    try:
        eid = event_id or version_name
        if not eid:
            return {"success": False, "message": "Missing event id"}
        if eid.startswith("ver:"):
            return _detail_version(eid[4:])
        if eid.startswith("new:"):
            dt, _, name = eid[4:].partition("::")
            return _detail_created(dt, name)
        if eid.startswith("del:"):
            return _detail_deleted(eid[4:])
        # Backward-compatibility: a bare Version name.
        return _detail_version(eid)
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Audit Trail Detail Error")
        return {"success": False, "message": str(e)}


# =============================================================================
# EXPORT MASTERS
# =============================================================================

# Master export configuration — maps frontend keys to doctypes and column definitions
_EXPORT_MASTER_CONFIGS = {
    "hq": {
        "title": "HQ Master",
        "doctype": "HQ Master",
        "columns": ["hq_code", "hq_name", "team", "region", "zone", "per_capita", "division", "status"],
        "headers": ["HQ Code", "HQ Name", "Team", "Region", "Zone", "Per Capita", "Division", "Status"],
        "resolve": {"team": "Team Master:team_name", "region": "Region Master:region_name", "zone": "Zone Master:zone_name"},
    },
    "stockist": {
        "title": "Stockist Master",
        "doctype": "Stockist Master",
        "columns": ["name", "stockist_code", "stockist_name", "hq", "team", "region", "zone", "address", "contact_person", "phone", "email", "division", "status"],
        "headers": ["ID", "Stockist Code", "Stockist Name", "HQ", "Team", "Region", "Zone", "Address", "Contact Person", "Phone", "Email", "Division", "Status"],
        "resolve": {"hq": "HQ Master:hq_name", "team": "Team Master:team_name", "region": "Region Master:region_name", "zone": "Zone Master:zone_name"},
    },
    "product": {
        "title": "Product Master",
        "doctype": "Product Master",
        "columns": ["name", "product_code", "product_name", "product_group", "category", "pack", "pack_conversion", "pts", "ptr", "mrp", "gst_rate", "division", "status"],
        "headers": ["ID", "Product Code", "Product Name", "Product Group", "Category", "Pack", "Pack Conversion", "PTS", "PTR", "MRP", "GST Rate", "Division", "Status"],
        "resolve": {},
    },
    "doctor": {
        "title": "Doctor Master",
        "doctype": "Doctor Master",
        "columns": ["doctor_code", "doctor_name", "qualification", "doctor_category", "specialization", "phone", "place", "hq", "team", "region", "state", "zone", "chemist_name", "division", "status"],
        "headers": ["Doctor Code", "Doctor Name", "Qualification", "Category", "Specialization", "Phone", "Place", "HQ", "Team", "Region", "State", "Zone", "Chemist Name", "Division", "Status"],
        "resolve": {"hq": "HQ Master:hq_name", "team": "Team Master:team_name", "region": "Region Master:region_name", "state": "State Master:state_name", "zone": "Zone Master:zone_name"},
    },
    "team": {
        "title": "Team Master",
        "doctype": "Team Master",
        "columns": ["team_code", "team_name", "region", "sanctioned_strength", "division", "status"],
        "headers": ["Team Code", "Team Name", "Region", "Sanctioned Strength", "Division", "Status"],
        "resolve": {"region": "Region Master:region_name"},
    },
    "region": {
        "title": "Region Master",
        "doctype": "Region Master",
        "columns": ["region_code", "region_name", "zone", "state", "division", "status"],
        "headers": ["Region Code", "Region Name", "Zone", "State", "Division", "Status"],
        "resolve": {"zone": "Zone Master:zone_name", "state": "State Master:state_name"},
    },
    "zone": {
        "title": "Zone Master",
        "doctype": "Zone Master",
        "columns": ["zone_code", "zone_name", "division", "status"],
        "headers": ["Zone Code", "Zone Name", "Division", "Status"],
        "resolve": {},
    },
    "state": {
        "title": "State Master",
        "doctype": "State Master",
        "columns": ["state_code", "state_name", "division", "status"],
        "headers": ["State Code", "State Name", "Division", "Status"],
        "resolve": {},
    },
}


def _fetch_export_data(config, division):
    """Fetch and resolve data for a master export."""
    doctype = config["doctype"]
    filters = {}
    if division:
        filters["division"] = ["in", [division, "Both"]]

    data = frappe.get_all(
        doctype,
        filters=filters,
        # dedupe: a config may list "name" as an exportable ID column
        fields=list(dict.fromkeys(["name"] + config["columns"])),
        order_by="name asc",
        limit_page_length=0,
    )

    # Resolve Link fields to human-readable labels
    resolve_map = config.get("resolve", {})
    # Build caches for each linked doctype to avoid N+1 queries
    resolve_caches = {}
    for field, spec in resolve_map.items():
        linked_dt, label_field = spec.split(":")
        all_vals = frappe.get_all(linked_dt, fields=["name", label_field], limit_page_length=0)
        resolve_caches[field] = {v["name"]: v[label_field] for v in all_vals}

    for row in data:
        for field, spec in resolve_map.items():
            val = row.get(field)
            if val and field in resolve_caches:
                row[field] = resolve_caches[field].get(val, val)

    return data


def _generate_excel(config, data, division):
    """Generate a professional Excel file for a master."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = config["title"]

    # Company header
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(config["headers"]))
    ws["A1"] = "Stedman Pharmaceuticals Pvt Ltd"
    ws["A1"].font = Font(bold=True, size=14, color="1e293b")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(config["headers"]))
    ws["A2"] = config["title"]
    ws["A2"].font = Font(bold=True, size=12, color="4f46e5")
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(config["headers"]))
    ws["A3"] = f"Division: {division or 'All'}  |  Exported: {frappe.utils.now_datetime().strftime('%d %b %Y, %I:%M %p')}"
    ws["A3"].font = Font(size=10, color="64748b")
    ws["A3"].alignment = Alignment(horizontal="center")

    # Column headers at row 5
    header_fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="d1d5db"),
        right=Side(style="thin", color="d1d5db"),
        top=Side(style="thin", color="d1d5db"),
        bottom=Side(style="thin", color="d1d5db"),
    )

    for col_idx, header in enumerate(config["headers"], 1):
        cell = ws.cell(row=5, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows starting at row 6
    alt_fill = PatternFill(start_color="f8fafc", end_color="f8fafc", fill_type="solid")
    data_align = Alignment(vertical="center", wrap_text=True)

    for row_idx, row in enumerate(data):
        excel_row = row_idx + 6
        for col_idx, col_key in enumerate(config["columns"], 1):
            cell = ws.cell(row=excel_row, column=col_idx)
            cell.value = row.get(col_key, "")
            cell.alignment = data_align
            cell.border = thin_border
            if row_idx % 2 == 1:
                cell.fill = alt_fill

    # Auto-fit column widths
    for col_idx, col_key in enumerate(config["columns"], 1):
        max_len = len(config["headers"][col_idx - 1])
        for row in data[:100]:  # Sample first 100 rows
            val = str(row.get(col_key, "") or "")
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

    # Footer
    footer_row = len(data) + 7
    ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=len(config["headers"]))
    ws.cell(row=footer_row, column=1).value = f"Total Records: {len(data)}"
    ws.cell(row=footer_row, column=1).font = Font(bold=True, size=10, color="64748b")

    return wb


def _generate_csv_content(config, data):
    """Generate CSV content for a master."""
    import csv
    import io

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(config["headers"])

    for row in data:
        writer.writerow([row.get(col, "") for col in config["columns"]])

    return output.getvalue()


def _generate_pdf_html(config, data, division):
    """Generate professional HTML for PDF conversion via wkhtmltopdf."""
    export_date = frappe.utils.now_datetime().strftime("%d %b %Y, %I:%M %p")
    num_cols = len(config["headers"])

    # Build table rows
    rows_html = ""
    for idx, row in enumerate(data):
        bg = "#f8fafc" if idx % 2 == 1 else "#ffffff"
        rows_html += f'<tr style="background:{bg};">'
        for col in config["columns"]:
            val = row.get(col, "") or ""
            rows_html += f'<td style="padding:6px 8px;border:1px solid #e2e8f0;font-size:9px;">{val}</td>'
        rows_html += "</tr>"

    # Build header cells
    header_cells = ""
    for h in config["headers"]:
        header_cells += f'<th style="padding:8px;border:1px solid #334155;background:#1e293b;color:#fff;font-size:9px;text-align:center;">{h}</th>'

    html = f"""
    <html>
    <head>
        <meta charset="utf-8">
        <style>
            @page {{
                size: {"landscape" if num_cols > 6 else "portrait"};
                margin: 15mm;
            }}
            body {{
                font-family: 'Inter', 'Helvetica Neue', Arial, sans-serif;
                color: #1e293b;
                margin: 0;
            }}
            .header {{
                text-align: center;
                margin-bottom: 20px;
                border-bottom: 2px solid #4f46e5;
                padding-bottom: 12px;
            }}
            .header h1 {{
                font-size: 18px;
                margin: 0;
                color: #1e293b;
            }}
            .header h2 {{
                font-size: 14px;
                margin: 4px 0;
                color: #4f46e5;
                font-weight: 600;
            }}
            .header p {{
                font-size: 10px;
                color: #64748b;
                margin: 4px 0 0 0;
            }}
            table {{
                width: 100%;
                border-collapse: collapse;
                margin-top: 10px;
            }}
            .footer {{
                text-align: center;
                margin-top: 15px;
                font-size: 9px;
                color: #94a3b8;
            }}
        </style>
    </head>
    <body>
        <div class="header">
            <h1>Stedman Pharmaceuticals Pvt Ltd</h1>
            <h2>{config["title"]}</h2>
            <p>Division: {division or "All"}  &bull;  Exported: {export_date}  &bull;  Total Records: {len(data)}</p>
        </div>
        <table>
            <thead><tr>{header_cells}</tr></thead>
            <tbody>{rows_html}</tbody>
        </table>
        <div class="footer">
            &copy; {frappe.utils.now_datetime().year} Stedman Pharmaceuticals Pvt Ltd &bull; Generated by Scanify
        </div>
    </body>
    </html>
    """
    return html


@frappe.whitelist()
def export_master_data(master_type, format_type="xlsx", division=None):
    """Export a single master to Excel, CSV, or PDF."""
    try:
        if master_type not in _EXPORT_MASTER_CONFIGS:
            return {"success": False, "message": f"Unknown master type: {master_type}"}

        config = _EXPORT_MASTER_CONFIGS[master_type]
        data = _fetch_export_data(config, division)

        timestamp = frappe.utils.now_datetime().strftime("%Y%m%d_%H%M%S")
        safe_title = config["title"].replace(" ", "_")

        files_dir = frappe.get_site_path("public", "files")
        os.makedirs(files_dir, exist_ok=True)

        if format_type == "xlsx":
            wb = _generate_excel(config, data, division)
            filename = f"{safe_title}_{timestamp}.xlsx"
            wb.save(os.path.join(files_dir, filename))

        elif format_type == "csv":
            csv_content = _generate_csv_content(config, data)
            filename = f"{safe_title}_{timestamp}.csv"
            with open(os.path.join(files_dir, filename), "w", encoding="utf-8-sig") as f:
                f.write(csv_content)

        elif format_type == "pdf":
            html = _generate_pdf_html(config, data, division)
            from frappe.utils.pdf import get_pdf
            pdf_content = get_pdf(html)
            filename = f"{safe_title}_{timestamp}.pdf"
            with open(os.path.join(files_dir, filename), "wb") as f:
                f.write(pdf_content)

        else:
            return {"success": False, "message": f"Unsupported format: {format_type}"}

        return {
            "success": True,
            "message": f"{config['title']} exported successfully",
            "file_url": f"/files/{filename}",
        }

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Export Master Data Error")
        return {"success": False, "message": str(e)}


@frappe.whitelist()
def export_all_masters_zip(format_type="xlsx", division=None):
    """Export all masters into individual files and bundle them in a ZIP."""
    try:
        import zipfile
        import tempfile

        timestamp = frappe.utils.now_datetime().strftime("%Y%m%d_%H%M%S")
        zip_filename = f"All_Masters_{timestamp}.zip"
        files_dir = frappe.get_site_path("public", "files")
        os.makedirs(files_dir, exist_ok=True)
        zip_filepath = os.path.join(files_dir, zip_filename)

        with zipfile.ZipFile(zip_filepath, "w", zipfile.ZIP_DEFLATED) as zf:
            for master_key, config in _EXPORT_MASTER_CONFIGS.items():
                data = _fetch_export_data(config, division)
                safe_title = config["title"].replace(" ", "_")

                if format_type == "xlsx":
                    wb = _generate_excel(config, data, division)
                    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                        wb.save(tmp.name)
                        zf.write(tmp.name, f"{safe_title}.xlsx")
                        os.unlink(tmp.name)

                elif format_type == "csv":
                    csv_content = _generate_csv_content(config, data)
                    zf.writestr(f"{safe_title}.csv", csv_content)

                elif format_type == "pdf":
                    html = _generate_pdf_html(config, data, division)
                    from frappe.utils.pdf import get_pdf
                    pdf_content = get_pdf(html)
                    zf.writestr(f"{safe_title}.pdf", pdf_content)

        return {
            "success": True,
            "message": "All masters exported successfully",
            "file_url": f"/files/{zip_filename}",
        }

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Export All Masters ZIP Error")
        return {"success": False, "message": str(e)}


# ===================== DELETE STATEMENT APIs =====================

@frappe.whitelist()
def search_stockist_statements(search_term="", division=None):
    """Search stockist statements by stockist name, code or statement name"""
    if not division:
        division = get_user_division()

    term = f"%{search_term}%"
    division_clause = ""
    params = {"term": term}

    if division and division != "Both":
        division_clause = "AND ss.division IN (%(division)s, 'Both')"
        params["division"] = division

    # stockist_code is the internal id; also resolve the term against the human-facing
    # Stockist Code so the box matches what the user actually sees.
    code_pks = frappe.get_all(
        "Stockist Master", filters={"stockist_code": ["like", term]},
        pluck="name", limit_page_length=0,
    )
    code_clause = ""
    if code_pks:
        ph = ", ".join(f"%(scp{i})s" for i in range(len(code_pks)))
        for i, pk in enumerate(code_pks):
            params[f"scp{i}"] = pk
        code_clause = f" OR ss.stockist_code IN ({ph})"

    results = frappe.db.sql("""
        SELECT ss.name, ss.stockist_code, ss.stockist_name, ss.statement_month,
               ss.hq, ss.region, ss.docstatus, ss.division
        FROM `tabStockist Statement` ss
        WHERE (ss.stockist_name LIKE %(term)s
               OR ss.stockist_code LIKE %(term)s
               OR ss.name LIKE %(term)s
               {code_clause})
        {division_clause}
        ORDER BY ss.modified DESC
        LIMIT 20
    """.format(division_clause=division_clause, code_clause=code_clause), params, as_dict=True)

    # Display the human-facing code instead of the internal id.
    code_map = get_stockist_code_map([r["stockist_code"] for r in results])
    for r in results:
        r["stockist_code"] = code_map.get(r["stockist_code"], r["stockist_code"])

    return results


@frappe.whitelist()
def get_statement_summary(doc_name):
    """Get full metadata summary for a stockist statement"""
    try:
        doc = frappe.get_doc("Stockist Statement", doc_name)
        # Resolve human-readable names for HQ and Region codes
        hq_name = frappe.db.get_value("HQ Master", doc.hq, "hq_name") if doc.hq else doc.hq or ""
        region_name = frappe.db.get_value("Region Master", doc.region, "region_name") if doc.region else doc.region or ""
        return {
            "success": True,
            "name": doc.name,
            "stockist_code": get_stockist_code_map([doc.stockist_code]).get(doc.stockist_code, doc.stockist_code),
            "stockist_name": doc.stockist_name,
            "statement_month": str(doc.statement_month) if doc.statement_month else None,
            "hq": doc.hq,
            "hq_name": hq_name,
            "team": doc.team,
            "region": doc.region,
            "region_name": region_name,
            "zone": doc.zone,
            "division": doc.division,
            "docstatus": doc.docstatus,
            "creation": str(doc.creation) if doc.creation else None,
            "uploaded_file": doc.uploaded_file,
            "total_items": len(doc.items) if doc.items else 0,
            "total_opening_value": flt(doc.total_opening_value),
            "total_purchase_value": flt(doc.total_purchase_value),
            "total_sales_value": flt(doc.total_sales_value_pts),
            "total_closing_value": flt(doc.total_closing_value),
        }
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Get Statement Summary Error")
        return {"success": False, "message": str(e)}


@frappe.whitelist()
@require_process("secondary_admin")
def delete_stockist_statement(doc_name, reason):
    """Delete a stockist statement with a mandatory reason logged to audit trail"""
    try:
        reason = (reason or "").strip()
        if not reason or len(reason) < 5:
            return {"success": False, "message": "A reason of at least 5 characters is required."}

        if not frappe.db.exists("Stockist Statement", doc_name):
            return {"success": False, "message": f"Statement {doc_name} does not exist."}

        doc = frappe.get_doc("Stockist Statement", doc_name)

        # Store summary before deletion for audit trail
        summary = (
            f"Stockist: {doc.stockist_name} ({doc.stockist_code}), "
            f"Month: {doc.statement_month}, HQ: {doc.hq}, "
            f"Items: {len(doc.items) if doc.items else 0}"
        )

        # Add Comment for audit trail (persists after doc deletion)
        frappe.get_doc({
            "doctype": "Comment",
            "comment_type": "Info",
            "reference_doctype": "Stockist Statement",
            "reference_name": doc_name,
            "content": f"<b>Statement Deleted</b><br>Reason: {frappe.utils.escape_html(reason)}<br>{summary}",
            "comment_email": frappe.session.user,
        }).insert(ignore_permissions=True)

        # Cancel first if submitted
        if doc.docstatus == 1:
            doc.flags.ignore_permissions = True
            doc.cancel()

        # Delete the document
        frappe.delete_doc("Stockist Statement", doc_name, ignore_permissions=True, force=True)
        frappe.db.commit()

        return {"success": True, "message": f"Statement {doc_name} deleted. Reason logged to audit trail."}

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Delete Stockist Statement Error")
        return {"success": False, "message": str(e)}


# ===================== SCHEME REQUEST: STOCKISTS BY REGION =====================

@frappe.whitelist()
def get_stockists_by_region(region, division=None):
    """Get all active stockists in a region (across all HQs in that region)"""
    if not division:
        division = get_user_division()

    # Non-admins may only list stockists in regions they're mapped to.
    _allowed = _allowed_region_codes_or_all(division)
    if _allowed is not None and region not in set(_allowed):
        return []

    filters = {"region": region, "status": "Active"}
    if division and division != "Both":
        filters["division"] = ["in", [division, "Both"]]

    stockists = frappe.get_all(
        "Stockist Master",
        filters=filters,
        fields=["name", "stockist_code", "stockist_name", "hq"],
        order_by="stockist_name asc"
    )

    # Fetch HQ names for display
    hq_names = {}
    for s in stockists:
        if s.hq and s.hq not in hq_names:
            hq_names[s.hq] = frappe.db.get_value("HQ Master", s.hq, "hq_name") or s.hq
        s["hq_name"] = hq_names.get(s.hq, s.hq or "")

    return stockists


@frappe.whitelist()
def get_stockists_for_hq(hq, division=None):
    """Get all active stockists belonging to a specific HQ."""
    if not division:
        division = get_user_division()
    if not hq:
        return []

    filters = {"hq": hq, "status": "Active"}
    if division and division != "Both":
        filters["division"] = ["in", [division, "Both"]]
    # Non-admins may only list stockists inside their mapped regions.
    _allowed = _allowed_region_codes_or_all(division)
    if _allowed is not None:
        if not _allowed:
            return []
        filters["region"] = ["in", _allowed]

    stockists = frappe.get_all(
        "Stockist Master",
        filters=filters,
        fields=["name", "stockist_code", "stockist_name", "hq"],
        order_by="stockist_name asc"
    )

    hq_name = frappe.db.get_value("HQ Master", hq, "hq_name") or hq
    for s in stockists:
        s["hq_name"] = hq_name

    return stockists


@frappe.whitelist()
def get_bulk_delete_preview(division=None, region=None, team=None, hq=None,
                             stockist_code=None, from_month=None, to_month=None):
    """Return a preview list of statements matching the bulk-delete filters."""
    if not division:
        division = get_user_division()

    conditions = ["ss.division IN (%(division)s, 'Both')"]
    params = {"division": division}

    if region:
        conditions.append("ss.region = %(region)s")
        params["region"] = region
    if team:
        conditions.append("ss.team = %(team)s")
        params["team"] = team
    if hq:
        conditions.append("ss.hq = %(hq)s")
        params["hq"] = hq
    if stockist_code:
        conditions.append("ss.stockist_code = %(stockist_code)s")
        params["stockist_code"] = stockist_code
    if from_month:
        if len(from_month) == 7:
            from_month = from_month + "-01"
        conditions.append("ss.statement_month >= %(from_month)s")
        params["from_month"] = from_month
    if to_month:
        if len(to_month) == 7:
            import calendar as _cal
            y, m = int(to_month[:4]), int(to_month[5:7])
            last_day = _cal.monthrange(y, m)[1]
            to_month = f"{to_month}-{last_day:02d}"
        conditions.append("ss.statement_month <= %(to_month)s")
        params["to_month"] = to_month

    where = " AND ".join(conditions)
    rows = frappe.db.sql(f"""
        SELECT ss.name, ss.stockist_name, ss.stockist_code,
               ss.statement_month, ss.hq, ss.region, ss.docstatus
          FROM `tabStockist Statement` ss
         WHERE {where}
         ORDER BY ss.statement_month DESC, ss.stockist_name ASC
         LIMIT 500
    """, params, as_dict=True)

    # Enrich with human-readable names
    hq_cache, region_cache = {}, {}
    for r in rows:
        if r.hq and r.hq not in hq_cache:
            hq_cache[r.hq] = frappe.db.get_value("HQ Master", r.hq, "hq_name") or r.hq
        if r.region and r.region not in region_cache:
            region_cache[r.region] = frappe.db.get_value("Region Master", r.region, "region_name") or r.region
        r["hq_name"] = hq_cache.get(r.hq, r.hq or "")
        r["region_name"] = region_cache.get(r.region, r.region or "")
        r["month_display"] = str(r.statement_month)[:7] if r.statement_month else ""

    return {"success": True, "statements": rows, "count": len(rows)}


@frappe.whitelist()
@require_process("secondary_admin")
def bulk_delete_stockist_statements(doc_names, reason, division=None):
    """Delete multiple stockist statements with a mandatory reason logged to audit trail."""
    import json as _json
    if not division:
        division = get_user_division()

    reason = (reason or "").strip()
    if not reason or len(reason) < 5:
        return {"success": False, "message": "A reason of at least 5 characters is required."}

    if isinstance(doc_names, str):
        try:
            doc_names = _json.loads(doc_names)
        except Exception:
            doc_names = [d.strip() for d in doc_names.split(",") if d.strip()]

    if not doc_names:
        return {"success": False, "message": "No statements provided for deletion."}

    deleted, errors = [], []
    for doc_name in doc_names:
        try:
            if not frappe.db.exists("Stockist Statement", doc_name):
                errors.append(f"{doc_name}: not found")
                continue
            doc = frappe.get_doc("Stockist Statement", doc_name)
            summary = (
                f"Stockist: {doc.stockist_name} ({doc.stockist_code}), "
                f"Month: {doc.statement_month}, HQ: {doc.hq}"
            )
            frappe.get_doc({
                "doctype": "Comment",
                "comment_type": "Info",
                "reference_doctype": "Stockist Statement",
                "reference_name": doc_name,
                "content": (
                    f"<b>Statement Deleted (Bulk)</b><br>"
                    f"Reason: {frappe.utils.escape_html(reason)}<br>{summary}"
                ),
                "comment_email": frappe.session.user,
            }).insert(ignore_permissions=True)
            if doc.docstatus == 1:
                doc.flags.ignore_permissions = True
                doc.cancel()
            frappe.delete_doc("Stockist Statement", doc_name, ignore_permissions=True, force=True)
            deleted.append(doc_name)
        except Exception as e:
            errors.append(f"{doc_name}: {str(e)}")

    frappe.db.commit()
    msg = f"{len(deleted)} statement(s) deleted."
    if errors:
        msg += f" {len(errors)} error(s): " + "; ".join(errors[:3])
    return {"success": True, "deleted": deleted, "errors": errors, "message": msg}


# ===================== RELOAD / REFRESH STATEMENT TOTALS =====================

@frappe.whitelist()
def get_statements_for_reload(division=None, region=None, team=None, hq=None,
                              stockist_code=None, from_month=None, to_month=None):
    """Return statements matching the filters, for the bulk 'Reload Statements' screen.

    Reuses the same filter query as the bulk-delete preview (pure filter + list,
    no destructive action), so the two screens stay in sync.
    """
    return get_bulk_delete_preview(
        division=division, region=region, team=team, hq=hq,
        stockist_code=stockist_code, from_month=from_month, to_month=to_month,
    )


# What each reload aspect touches, and its default state. Kept here so the API,
# the docstring and the portal UI stay in agreement.
RELOAD_OPTION_DEFAULTS = {
    "org": True,      # HQ / Team / Region / Zone / Division from Stockist Master
    "totals": True,   # closing qty + opening/purchase/sales/closing values
    "pts": False,     # RE-PRICE from Product Master (opt-in — overwrites captured PTS)
    "qc": False,      # QC review status from item mapping_status
}


def _parse_reload_options(options):
    """Normalise the `options` argument into a dict of bools.

    Accepts a JSON string, a dict, or None. Missing/unknown keys fall back to
    RELOAD_OPTION_DEFAULTS (the safe hierarchy + totals refresh that never
    touches captured prices).
    """
    import json as _json

    if options is None:
        return dict(RELOAD_OPTION_DEFAULTS)
    if isinstance(options, str):
        try:
            options = _json.loads(options)
        except Exception:
            options = {}
    if not isinstance(options, dict):
        return dict(RELOAD_OPTION_DEFAULTS)

    def _b(v):
        if isinstance(v, bool):
            return v
        return str(v).strip().lower() in ("1", "true", "yes", "on")

    return {k: _b(options[k]) if k in options else default
            for k, default in RELOAD_OPTION_DEFAULTS.items()}


def _resync_statement_org(doc):
    """Copy HQ / Team / Region / Zone / Division onto the statement from the live
    Stockist Master — what `fetch_from` does for a fresh draft, but applied to
    already-submitted statements too (fetch_from is skipped once a doc is
    submitted, so these snapshots otherwise stay frozen at creation time).

    Returns True if any field actually changed.
    """
    if not doc.stockist_code:
        return False

    sm = frappe.db.get_value(
        "Stockist Master", doc.stockist_code,
        ["hq", "team", "region", "zone", "division"], as_dict=True,
    )
    if not sm:
        return False

    changed = False
    for field in ("hq", "team", "region", "zone"):
        new_val = sm.get(field)
        if (doc.get(field) or None) != (new_val or None):
            doc.set(field, new_val)
            changed = True

    # Division: the master's validate keeps Stockist Master.division in step with
    # its HQ, so trust it — but only overwrite with a real value (never blank the
    # statement's existing division out).
    new_div = sm.get("division")
    if new_div and new_div != doc.division:
        doc.division = new_div
        changed = True

    return changed


@frappe.whitelist()
@require_process("secondary_admin")
def reload_stockist_statements(doc_names, division=None, options=None):
    """Refresh selected aspects of the given statements from the live masters.

    `options` (JSON string / dict of bools) chooses WHAT to refresh:
        org    - re-sync HQ / Team / Region / Zone / Division from Stockist Master.
                 Fixes reports after a stockist's HQ/hierarchy is changed. (default ON)
        totals - recompute closing qty + opening/purchase/sales/closing values from
                 the current quantities and pack sizes, using the PTS already stored
                 on each line (unless `pts` is also chosen). (default ON)
        pts    - RE-PRICE: overwrite each line's PTS with the current Product Master
                 price and re-value everything off it. (default OFF — opt-in). Reports
                 read the statement's own PTS, so most refreshes must NOT do this;
                 only enable it when you deliberately want to reprice history.
        qc     - recompute the QC review status from item mapping_status. (default OFF)

    Persists via db_update so submitted statements can be refreshed WITHOUT
    re-triggering submit-time side effects (e.g. next-month opening).
    """
    import json as _json
    if not division:
        division = get_user_division()

    if isinstance(doc_names, str):
        try:
            doc_names = _json.loads(doc_names)
        except Exception:
            doc_names = [d.strip() for d in doc_names.split(",") if d.strip()]

    if not doc_names:
        return {"success": False, "message": "No statements provided to reload."}

    opts = _parse_reload_options(options)
    do_pts = opts["pts"]
    # Re-pricing only takes effect through a totals recompute, so imply it.
    do_totals = opts["totals"] or do_pts
    do_org = opts["org"]
    do_qc = opts["qc"]

    if not (do_org or do_totals or do_qc):
        return {"success": False,
                "message": "Nothing selected to reload. Choose at least one option."}

    reloaded, errors = [], []
    for doc_name in doc_names:
        try:
            if not frappe.db.exists("Stockist Statement", doc_name):
                errors.append(f"{doc_name}: not found")
                continue

            doc = frappe.get_doc("Stockist Statement", doc_name)
            touch_parent = False
            touch_items = False

            if do_org and _resync_statement_org(doc):
                touch_parent = True

            if do_pts and not getattr(doc, "skip_conversion", 0):
                # Wipe the cached per-line PTS so the recompute re-reads the live
                # Product Master price. On a normal statement item.pts holds a cached
                # copy of the master PTS, so the calc's "non-zero item.pts wins" rule
                # would otherwise keep the OLD rate. Backfilled statements
                # (skip_conversion) carry a genuine source net-rate — leave those be.
                for item in doc.items:
                    if item.product_code:
                        item.pts = 0

            if do_totals:
                # Preserves each line's stored PTS unless `pts` wiped it above.
                doc.calculate_closing_and_totals()
                touch_items = True
                touch_parent = True

            if do_qc:
                doc.calculate_qc_confidence()
                touch_parent = True

            if touch_items:
                for item in doc.items:
                    item.db_update()
            if touch_parent:
                doc.db_update()

            reloaded.append(doc_name)
        except Exception as e:
            frappe.log_error(frappe.get_traceback(), "Reload Statement Error")
            errors.append(f"{doc_name}: {str(e)}")

    frappe.db.commit()

    applied = [label for label, on in (
        ("hierarchy", do_org), ("totals", do_totals),
        ("re-priced PTS", do_pts), ("QC status", do_qc)) if on]
    msg = f"{len(reloaded)} statement(s) refreshed"
    if applied:
        msg += f" ({', '.join(applied)})"
    msg += "."
    if errors:
        msg += f" {len(errors)} error(s): " + "; ".join(errors[:3])
    return {"success": True, "reloaded": reloaded, "errors": errors,
            "applied": applied, "message": msg}


# ===================== DUPLICATE STATEMENT CHECK =====================

@frappe.whitelist()
def check_statement_exists(stockist_code, statement_month, division=None):
    """Check if a stockist statement already exists for the given stockist + month"""
    if not stockist_code or not statement_month:
        return {"exists": False}

    # Statements link by the master id (PK). Resolve the editable code → id,
    # division-scoped, so the existence check is exact even for custom codes.
    if not division:
        division = get_user_division()
    stockist_pk = _resolve_stockist_pk(stockist_code, division) or stockist_code

    # Normalise month to first-of-month
    if len(statement_month) == 7:
        statement_month = statement_month + "-01"

    existing = frappe.db.exists("Stockist Statement", {
        "stockist_code": stockist_pk,
        "statement_month": statement_month
    })

    if existing:
        return {
            "exists": True,
            "statement_name": existing,
            "message": f"A statement already exists for this stockist in this month: {existing}"
        }

    return {"exists": False}


@frappe.whitelist()
@require_process("secondary")
def create_ocr_statement(stockist_code, statement_month, uploaded_file=None, division=None):
    """Create a Draft Stockist Statement for the single-file OCR flow.

    The portal sends the editable Stockist Code, but the statement's `stockist_code`
    is a Link that must hold the master id (PK) so fetch_from can populate
    name/HQ/team/region and reports can join. Resolve the code → id here,
    division-scoped, so a code that differs from the id (or is reused across
    divisions) links to THIS division's stockist. Mirrors create_manual_statement
    and the bulk flow so every creation path stores the PK consistently — replaces
    the old client-side frappe.client.insert that wrote the raw editable code.
    """
    try:
        if not stockist_code or not statement_month:
            return {"success": False, "message": "Stockist code and statement month are required."}

        if not division:
            division = get_user_division()

        stockist_pk = _resolve_stockist_pk(stockist_code, division)
        if not stockist_pk:
            return {"success": False, "message": f"Stockist '{stockist_code}' not found in this division."}

        if frappe.db.get_value("Stockist Master", stockist_pk, "status") != "Active":
            return {"success": False, "message": f"Stockist '{stockist_code}' is inactive. Statement upload is not allowed."}

        # Non-admins may only enter statements for stockists in their mapped regions.
        if not _stockist_region_allowed(stockist_pk, division):
            return {"success": False,
                    "message": f"Stockist '{stockist_code}' is outside the regions you are mapped to."}

        if len(statement_month) == 7:
            statement_month = statement_month + "-01"

        existing = frappe.db.exists("Stockist Statement", {
            "stockist_code": stockist_pk,
            "statement_month": statement_month,
        })
        if existing:
            return {"success": False, "message": f"A statement already exists: {existing}"}

        doc = frappe.new_doc("Stockist Statement")
        doc.stockist_code = stockist_pk
        doc.statement_month = statement_month
        doc.division = division
        if uploaded_file:
            doc.uploaded_file = uploaded_file
        doc.extracted_data_status = "Pending"
        doc.insert(ignore_permissions=True)

        return {"success": True, "name": doc.name}
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Create OCR Statement Error")
        return {"success": False, "message": str(e)}


# ─────────────────────────────────────────────────────
# Primary Sales Upload, List & Export
# ─────────────────────────────────────────────────────

# Column mapping from Excel headers → doctype fields
_PRIMARY_SALES_COL_MAP = {
    "stockistcode": "stockist_code",
    "product_head": "product_head",
    "stockistname": "stockist_name",
    "citypool": "citypool",
    "team": "team",
    "region": "region",
    "act_region": "act_region",
    "zonee": "zonee",
    "invoiceno": "invoiceno",
    "invoicedate": "invoicedate",
    "pcode": "pcode",
    "product": "product",
    "pack": "pack",
    "batchno": "batchno",
    "quantity": "quantity",
    "freeqty": "freeqty",
    "expqty": "expqty",
    "pts": "pts",
    "ptsvalue": "ptsvalue",
    "ptr": "ptr",
    "ptrvalue": "ptrvalue",
    "mrp": "mrp",
    "mrpvalue": "mrpvalue",
    "nrv": "nrv",
    "nrvvalue": "nrvvalue",
    "dsort": "dsort",
    "direct_party": "direct_party",
    "iscancelled": "iscancelled",
}

_NUMERIC_FIELDS = {
    "quantity", "freeqty", "expqty",
    "pts", "ptsvalue", "ptr", "ptrvalue",
    "mrp", "mrpvalue", "nrv", "nrvvalue", "dsort",
}

_BOOL_FIELDS = {"direct_party", "iscancelled"}


def _parse_bool(val):
    """Convert various truthy representations to 1/0."""
    if val is None:
        return 0
    if isinstance(val, bool):
        return 1 if val else 0
    if isinstance(val, (int, float)):
        return 1 if val else 0
    s = str(val).strip().lower()
    return 1 if s in ("true", "1", "yes", "y") else 0


def _parse_num(val):
    """Safely parse a numeric value."""
    if val is None:
        return 0
    try:
        return flt(val)
    except Exception:
        return 0


@frappe.whitelist()
@require_process("primary_upload")
def process_primary_sales_upload(upload_month, file_url):
    """
    Process an uploaded Excel file of primary sales data.
    Validates columns, maps to masters, and creates Primary Sales Data records.
    """
    import openpyxl
    from datetime import datetime

    user_division = get_user_division() or "Prima"

    # Validate month format (YYYY-MM)
    if not upload_month or not re.match(r"^\d{4}-\d{2}$", str(upload_month)):
        return {"success": False, "error": "Invalid month format. Use YYYY-MM."}

    # Get the file path
    file_path = frappe.get_site_path(file_url.lstrip("/"))
    if not os.path.exists(file_path):
        # Try with 'private' prefix
        file_path = frappe.get_site_path("private", "files", os.path.basename(file_url))
    if not os.path.exists(file_path):
        return {"success": False, "error": "Uploaded file not found on server."}

    # Create upload record
    upload_doc = frappe.get_doc({
        "doctype": "Primary Sales Upload",
        "upload_month": upload_month,
        "division": user_division,
        "uploaded_by": frappe.session.user,
        "upload_date": frappe.utils.now(),
        "file": file_url,
        "status": "Processing",
    })
    upload_doc.insert(ignore_permissions=True)
    frappe.db.commit()

    errors = []
    success_count = 0
    total_rows = 0

    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active

        # Read headers from row 1
        raw_headers = [cell.value for cell in ws[1]]
        headers = []
        for h in raw_headers:
            if h is None:
                headers.append(None)
            else:
                headers.append(str(h).strip().lower())

        # Validate required columns exist. Stockist Code is the match key (compared
        # against the editable Stockist Code in the master); name is an optional fallback.
        required_cols = {"stockistcode", "iscancelled"}
        found_cols = set(h for h in headers if h)
        missing = required_cols - found_cols
        if missing:
            upload_doc.status = "Failed"
            upload_doc.error_log = f"Missing required columns: {', '.join(missing)}"
            upload_doc.save(ignore_permissions=True)
            frappe.db.commit()
            return {"success": False, "error": f"Missing required columns: {', '.join(missing)}"}

        # Build caches for stockist resolution.
        # Primary match key: the editable Stockist Code; fallback: Stockist Name.
        # Both map to the master's internal id (name, e.g. S0001) — that is what
        # Stockist Statement links to and what every report filters on, so resolving
        # to the id keeps primary sales aligned even after a code is edited.
        stockist_code_cache = {}
        stockist_name_cache = {}
        for s in frappe.get_all("Stockist Master",
                                filters={"division": ["in", [user_division, "Both"]]},
                                fields=["name", "stockist_code", "stockist_name"],
                                limit_page_length=0):
            if s.stockist_code:
                stockist_code_cache[str(s.stockist_code).strip().lower()] = s.name
            if s.stockist_name:
                stockist_name_cache[s.stockist_name.strip().lower()] = s.name

        # code → id map. The same Product Code may exist in several of these
        # divisions; prefer the uploader's own division on collision.
        product_cache = {}
        for p in frappe.get_all("Product Master",
                                filters={"division": ["in", [user_division, "Both", "ASPR", "Wellness"]]},
                                fields=["product_code", "name", "division"],
                                limit_page_length=0):
            if p.product_code not in product_cache or p.division == user_division:
                product_cache[p.product_code] = p.name

        # Process rows
        batch_size = 100
        batch = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Skip completely empty rows
            if all(v is None for v in row):
                continue

            total_rows += 1
            row_data = {}

            for col_idx, val in enumerate(row):
                if col_idx >= len(headers) or headers[col_idx] is None:
                    continue
                excel_col = headers[col_idx]
                if excel_col in _PRIMARY_SALES_COL_MAP:
                    field_name = _PRIMARY_SALES_COL_MAP[excel_col]
                    if field_name in _BOOL_FIELDS:
                        row_data[field_name] = _parse_bool(val)
                    elif field_name in _NUMERIC_FIELDS:
                        row_data[field_name] = _parse_num(val)
                    elif field_name == "invoicedate":
                        if isinstance(val, datetime):
                            row_data[field_name] = val.strftime("%Y-%m-%d")
                        elif val:
                            row_data[field_name] = str(val)
                        else:
                            row_data[field_name] = None
                    else:
                        row_data[field_name] = str(val).strip() if val else ""

            is_cancelled = row_data.get("iscancelled", 0)

            # Resolve stockist: match the Excel Stockist Code against the editable code in
            # the master first, then fall back to Stockist Name. Store the master's internal
            # id (name, S####) so statements/reports keep matching regardless of code edits.
            excel_stockist_code = (row_data.get("stockist_code") or "").strip()
            excel_stockist_name = (row_data.get("stockist_name") or "").strip()
            if not excel_stockist_code and not excel_stockist_name:
                errors.append(f"Row {row_idx}: Missing stockist code and name")
                continue

            matched_name = None
            if excel_stockist_code:
                matched_name = stockist_code_cache.get(excel_stockist_code.lower())
            if not matched_name and excel_stockist_name:
                matched_name = stockist_name_cache.get(excel_stockist_name.lower())

            if matched_name:
                row_data["stockist_code"] = matched_name
            else:
                ident = excel_stockist_code or excel_stockist_name
                errors.append(f"Row {row_idx}: Stockist '{ident}' not found in Stockist Master (data saved anyway)")
                # Leave the raw Excel stockist_code on the row so the data isn't lost

            # For non-cancelled rows, validate product code if present
            pcode = row_data.get("pcode", "")
            if not is_cancelled and pcode:
                # Warn if product not found but still store the data
                if pcode not in product_cache:
                    errors.append(f"Row {row_idx}: Product code '{pcode}' not found in Product Master (data saved anyway)")

            # Build the record
            row_data["upload_month"] = upload_month
            row_data["division"] = user_division
            row_data["upload_ref"] = upload_doc.name
            row_data["doctype"] = "Primary Sales Data"

            batch.append(row_data)
            success_count += 1

            # Insert in batches
            if len(batch) >= batch_size:
                for rec in batch:
                    doc = frappe.get_doc(rec)
                    doc.insert(ignore_permissions=True)
                batch = []
                frappe.db.commit()

        # Insert remaining
        if batch:
            for rec in batch:
                doc = frappe.get_doc(rec)
                doc.insert(ignore_permissions=True)
            frappe.db.commit()

        wb.close()

    except Exception as e:
        frappe.db.rollback()
        upload_doc.reload()
        upload_doc.status = "Failed"
        upload_doc.error_log = str(e)
        upload_doc.save(ignore_permissions=True)
        frappe.db.commit()
        frappe.log_error(f"Primary Sales Upload Error: {str(e)}", "Primary Sales Upload")
        return {"success": False, "error": str(e)}

    # Update upload record
    upload_doc.reload()
    upload_doc.total_rows = total_rows
    upload_doc.success_count = success_count
    upload_doc.error_count = len([e for e in errors if "not found" not in e])
    upload_doc.status = "Completed"
    if errors:
        upload_doc.error_log = "\n".join(errors[:200])
    upload_doc.save(ignore_permissions=True)
    frappe.db.commit()

    return {
        "success": True,
        "total_rows": total_rows,
        "success_count": success_count,
        "error_count": len(errors),
        "errors": errors[:20],
        "upload_name": upload_doc.name,
    }


@frappe.whitelist()
def get_primary_sales_data(division, page=1, page_size=50,
                           upload_month=None, zonee=None, region=None,
                           team=None, hq=None, product_head=None, iscancelled=None,
                           stockist_search=None, pcode=None, product_search=None,
                           invoiceno=None):
    """Fetch paginated primary sales data with filters.

    zonee/region/team accept either the master *code* (preferred from the new
    cascading dropdowns) or the *name* (legacy callers). They are resolved to
    the stored display value below.
    hq must be the HQ Master *code* — filtered via Stockist Master.hq.
    """

    page = int(page)
    page_size = min(int(page_size), 200)
    offset = (page - 1) * page_size

    # Resolve code → stored display name for the Data fields in Primary Sales Data.
    def _resolve(doctype, value, name_field):
        if not value:
            return value
        # If a Master record exists with that code, swap to its display name.
        actual = frappe.db.get_value(doctype, value, name_field)
        return actual or value

    zone_filter = _resolve("Zone Master", zonee, "zone_name") if zonee else None
    region_filter = _resolve("Region Master", region, "region_name") if region else None
    team_filter = _resolve("Team Master", team, "team_name") if team else None

    conditions = ["division = %(division)s"]
    params = {"division": division, "page_size": page_size, "offset": offset}

    if upload_month:
        conditions.append("upload_month = %(upload_month)s")
        params["upload_month"] = upload_month

    if zone_filter:
        conditions.append("zonee = %(zonee)s")
        params["zonee"] = zone_filter

    if region_filter:
        conditions.append("region = %(region)s")
        params["region"] = region_filter

    if team_filter:
        conditions.append("team = %(team)s")
        params["team"] = team_filter

    if hq:
        conditions.append("stockist_code IN (SELECT name FROM `tabStockist Master` WHERE hq = %(hq)s)")
        params["hq"] = hq

    if product_head:
        conditions.append("product_head = %(product_head)s")
        params["product_head"] = product_head

    if iscancelled is not None and iscancelled != "":
        conditions.append("iscancelled = %(iscancelled)s")
        params["iscancelled"] = int(iscancelled)

    if stockist_search:
        like = f"%{stockist_search}%"
        params["stockist_search"] = like
        # stockist_code stores the internal id; also resolve the search against the
        # human-facing Stockist Code so users can search by either.
        code_pks = frappe.get_all(
            "Stockist Master", filters={"stockist_code": ["like", like]},
            pluck="name", limit_page_length=0,
        )
        search_conds = ["stockist_name LIKE %(stockist_search)s", "stockist_code LIKE %(stockist_search)s"]
        if code_pks:
            ph = ", ".join(f"%(scp{i})s" for i in range(len(code_pks)))
            for i, pk in enumerate(code_pks):
                params[f"scp{i}"] = pk
            search_conds.append(f"stockist_code IN ({ph})")
        conditions.append("(" + " OR ".join(search_conds) + ")")

    if pcode:
        conditions.append("pcode LIKE %(pcode)s")
        params["pcode"] = f"%{pcode}%"

    if product_search:
        conditions.append("product LIKE %(product_search)s")
        params["product_search"] = f"%{product_search}%"

    if invoiceno:
        conditions.append("invoiceno LIKE %(invoiceno)s")
        params["invoiceno"] = f"%{invoiceno}%"

    where_clause = " AND ".join(conditions)

    total = frappe.db.sql(
        f"SELECT COUNT(*) FROM `tabPrimary Sales Data` WHERE {where_clause}",
        params
    )[0][0]

    rows = frappe.db.sql(
        f"""SELECT
            name,
            stockist_code, stockist_name, product_head, pcode, product, pack,
            invoiceno, invoicedate, batchno,
            quantity, freeqty, expqty,
            pts, ptsvalue, ptr, ptrvalue,
            mrp, mrpvalue, nrv, nrvvalue,
            team, region, zonee, citypool,
            direct_party, iscancelled, upload_month, dsort
        FROM `tabPrimary Sales Data`
        WHERE {where_clause}
        ORDER BY stockist_code, invoicedate, pcode
        LIMIT %(page_size)s OFFSET %(offset)s""",
        params, as_dict=True
    )

    # Display the human-facing Stockist Code instead of the internal id.
    code_map = get_stockist_code_map([r["stockist_code"] for r in rows])
    for r in rows:
        r["stockist_code"] = code_map.get(r["stockist_code"], r["stockist_code"])

    return {
        "rows": rows,
        "total": total,
        "page": page,
        "page_size": page_size,
    }


@frappe.whitelist()
def get_primary_sales_count(month, division):
    """Get count of primary sales records for a given month and division."""
    count = frappe.db.count("Primary Sales Data", filters={
        "upload_month": month,
        "division": division,
    })
    return count


@frappe.whitelist()
def export_primary_sales_data(month, division):
    """Export primary sales data as Excel for the given month and division."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    if not month or not division:
        frappe.throw("Month and division are required")

    rows = frappe.db.sql("""
        SELECT
            stockist_code, product_head, stockist_name, citypool,
            team, region, act_region, zonee,
            invoiceno, invoicedate, pcode, product, pack, batchno,
            quantity, freeqty, expqty,
            pts, ptsvalue, ptr, ptrvalue,
            mrp, mrpvalue, nrv, nrvvalue,
            dsort, direct_party, iscancelled
        FROM `tabPrimary Sales Data`
        WHERE upload_month = %(month)s AND division = %(division)s
        ORDER BY stockist_code, invoicedate, pcode
    """, {"month": month, "division": division}, as_dict=True)

    if not rows:
        frappe.throw(f"No data found for {month} in {division} division")

    # Export the human-facing Stockist Code so a re-upload matches by code.
    code_map = get_stockist_code_map([r["stockist_code"] for r in rows])
    for r in rows:
        r["stockist_code"] = code_map.get(r["stockist_code"], r["stockist_code"])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Primary Sales {month}"

    # Excel column headers (matching original upload format)
    excel_headers = [
        "stockistcode", "product_head", "stockistname", "citypool",
        "team", "region", "act_region", "zonee",
        "invoiceno", "invoicedate", "pcode", "product", "pack", "batchno",
        "quantity", "freeqty", "expqty",
        "pts", "ptsvalue", "ptr", "ptrvalue",
        "mrp", "mrpvalue", "nrv", "nrvvalue",
        "dsort", "direct_party", "iscancelled",
    ]

    # Field mapping for data rows
    field_keys = [
        "stockist_code", "product_head", "stockist_name", "citypool",
        "team", "region", "act_region", "zonee",
        "invoiceno", "invoicedate", "pcode", "product", "pack", "batchno",
        "quantity", "freeqty", "expqty",
        "pts", "ptsvalue", "ptr", "ptrvalue",
        "mrp", "mrpvalue", "nrv", "nrvvalue",
        "dsort", "direct_party", "iscancelled",
    ]

    # Header styling
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Write headers
    for col_idx, header in enumerate(excel_headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Write data
    for row_idx, row in enumerate(rows, 2):
        for col_idx, key in enumerate(field_keys, 1):
            val = row.get(key, "")
            if key in ("direct_party", "iscancelled"):
                val = True if val else False
            elif key == "invoicedate" and val:
                val = str(val)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border

    # Auto-width columns
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 3, 30)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Save to temp file and serve
    fname = f"Primary_Sales_{division}_{month}.xlsx"
    file_path = os.path.join(tempfile.gettempdir(), fname)
    wb.save(file_path)

    with open(file_path, "rb") as f:
        file_content = f.read()

    # Clean up
    try:
        os.remove(file_path)
    except OSError:
        pass

    frappe.local.response.filename = fname
    frappe.local.response.filecontent = file_content
    frappe.local.response.type = "download"


@frappe.whitelist()
def get_secondary_sales_upload_log(name):
    """Return the full import log text + counts for a Secondary Sales Upload record."""
    if not name:
        return {"success": False, "message": "Upload name is required"}
    user_division = get_user_division() or "Prima"
    doc = frappe.db.get_value(
        "Secondary Sales Upload", name,
        ["name", "upload_month", "division", "uploaded_by", "upload_date", "status",
         "total_data_rows", "stockists_in_file", "statements_created", "items_created",
         "skipped_existing", "unmatched_stockists", "inactive_stockists",
         "unmapped_products", "create_errors", "log"],
        as_dict=True,
    )
    if not doc:
        return {"success": False, "message": "Upload record not found"}
    # Division guard — users only see their own division's logs
    if doc.division and doc.division != user_division and "System Manager" not in frappe.get_roles():
        return {"success": False, "message": "Not permitted for this division"}
    doc["upload_date"] = str(doc.get("upload_date") or "")
    return {"success": True, "data": doc}


@frappe.whitelist()
@require_process("primary_view")
def save_primary_sales_record(name=None, data=None):
    """Create or update a Primary Sales Data record."""
    try:
        data = frappe.parse_json(data) if isinstance(data, str) else data
        if not data:
            return {"success": False, "message": "No data provided"}

        current_division = get_user_division() or "Prima"
        data["division"] = current_division

        # Numeric fields
        num_fields = [
            "quantity", "freeqty", "expqty", "pts", "ptsvalue",
            "ptr", "ptrvalue", "mrp", "mrpvalue", "nrv", "nrvvalue", "dsort"
        ]
        for nf in num_fields:
            val = data.get(nf)
            if val is not None and val != "":
                try:
                    data[nf] = float(val)
                except (ValueError, TypeError):
                    data[nf] = 0
            else:
                data[nf] = 0

        # Boolean fields
        for bf in ["iscancelled", "direct_party"]:
            val = data.get(bf)
            data[bf] = 1 if val in (1, "1", True, "true", "True") else 0

        if name:
            doc = frappe.get_doc("Primary Sales Data", name)
            doc.update(data)
        else:
            doc = frappe.new_doc("Primary Sales Data")
            doc.update(data)

        doc.save(ignore_permissions=False)
        frappe.db.commit()

        return {"success": True, "message": "Record saved successfully", "name": doc.name}

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Save Primary Sales Record Error")
        return {"success": False, "message": str(e)}


@frappe.whitelist()
@require_process("primary_view")
def delete_primary_sales_record(name):
    """Delete a Primary Sales Data record."""
    try:
        frappe.delete_doc("Primary Sales Data", name, ignore_permissions=False)
        frappe.db.commit()
        return {"success": True, "message": "Record deleted successfully"}
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Delete Primary Sales Record Error")
        return {"success": False, "message": str(e)}


@frappe.whitelist()
def get_primary_sales_month_stockists(month, division=None):
    """List the stockists that have Primary Sales Data in a given month — used to
    populate the stockist filter in the 'Delete Month' dialog. `value` is the stored
    stockist_code (the master id) so the delete filter matches exactly; display_code
    is the human-facing editable code."""
    if not division:
        division = get_user_division()
    if not month:
        return {"success": False, "stockists": [], "total": 0}
    month = str(month)[:7]
    rows = frappe.db.sql("""
        SELECT psd.stockist_code AS value,
               COALESCE(sm.stockist_name, psd.stockist_name, '') AS stockist_name,
               COALESCE(NULLIF(sm.stockist_code, ''), psd.stockist_code) AS display_code,
               COUNT(*) AS row_count
          FROM `tabPrimary Sales Data` psd
     LEFT JOIN `tabStockist Master` sm ON sm.name = psd.stockist_code
         WHERE psd.division = %(division)s AND psd.upload_month = %(month)s
      GROUP BY psd.stockist_code, stockist_name, display_code
      ORDER BY stockist_name
    """, {"division": division, "month": month}, as_dict=True)
    total = sum(int(r.row_count or 0) for r in rows)
    return {"success": True, "stockists": rows, "total": total}


@frappe.whitelist()
@require_process("primary_upload")
def delete_primary_sales_month(month, reason, stockist_codes=None, division=None):
    """Delete all Primary Sales Data for a division + month, optionally limited to a
    set of stockists. A reason (>= 5 chars) is recorded on the month's upload record(s)
    for audit. Reversible by re-uploading the source Excel, so this is intentionally a
    light-touch bulk delete used to fix back-dated / mis-uploaded primary data."""
    if not division:
        division = get_user_division()
    month = str(month or "")[:7]
    reason = (reason or "").strip()
    if not month:
        return {"success": False, "message": "Month is required."}
    if not reason or len(reason) < 5:
        return {"success": False, "message": "A reason of at least 5 characters is required."}

    # Optional stockist filter — accept a JSON array or comma-separated list of the
    # stored stockist_code (master id) values returned by get_primary_sales_month_stockists.
    codes = None
    if stockist_codes:
        if isinstance(stockist_codes, str):
            try:
                codes = json.loads(stockist_codes)
            except Exception:
                codes = [c.strip() for c in stockist_codes.split(",") if c.strip()]
        else:
            codes = list(stockist_codes)
        codes = [c for c in (codes or []) if c]

    filters = {"division": division, "upload_month": month}
    if codes:
        filters["stockist_code"] = ["in", codes]

    count = frappe.db.count("Primary Sales Data", filters)
    if not count:
        return {"success": False, "message": "No primary sales rows found for the selected month/stockists."}

    # Audit trail on the month's upload record(s) before the data is removed.
    try:
        scope = "ALL stockists" if not codes else f"{len(codes)} selected stockist(s)"
        upload_names = frappe.get_all(
            "Primary Sales Upload",
            filters={"division": division, "upload_month": month},
            pluck="name",
        )
        for up in upload_names:
            frappe.get_doc({
                "doctype": "Comment",
                "comment_type": "Info",
                "reference_doctype": "Primary Sales Upload",
                "reference_name": up,
                "content": (
                    f"<b>Primary Sales Deleted</b><br>"
                    f"Month: {month} | Scope: {scope} | Rows: {count}<br>"
                    f"Reason: {frappe.utils.escape_html(reason)}"
                ),
                "comment_email": frappe.session.user,
            }).insert(ignore_permissions=True)
    except Exception:
        frappe.log_error(frappe.get_traceback(), "Primary Sales Delete Audit Error")

    frappe.db.delete("Primary Sales Data", filters)
    frappe.db.commit()

    frappe.logger().info(
        f"Primary Sales deleted: division={division} month={month} "
        f"rows={count} scope={'all' if not codes else len(codes)} "
        f"by={frappe.session.user} reason={reason!r}"
    )
    return {
        "success": True,
        "deleted": count,
        "message": f"Deleted {count} primary sales row(s) for {month}.",
    }


# ──────────────────────────────────────────────────────────────────────────────
# STOCKIST REPORTS
# ──────────────────────────────────────────────────────────────────────────────

@frappe.whitelist()
def get_stockist_report_filter_options(division=None):
    """Return dropdown options for the Stockist Reports page (active masters only)."""
    if not division:
        division = get_user_division()

    # None for admin (= all regions); a list confines a non-admin's pickers below.
    allowed = _allowed_region_codes_or_all(division)

    regions = frappe.db.sql(
        "SELECT DISTINCT name, region_name, IFNULL(zone, '') AS zone FROM `tabRegion Master` WHERE status='Active' AND division IN (%s, 'Both') ORDER BY name",
        (division,), as_dict=True,
    )
    teams = frappe.db.sql(
        """SELECT DISTINCT t.name, t.team_name,
                  IFNULL(t.region, '') AS region,
                  IFNULL(r.zone, '')   AS zone
             FROM `tabTeam Master` t
             LEFT JOIN `tabRegion Master` r ON r.name = t.region
            WHERE t.status='Active' AND t.division IN (%s, 'Both')
            ORDER BY t.name""",
        (division,), as_dict=True,
    )
    hqs = frappe.db.sql(
        "SELECT DISTINCT name, hq_name, IFNULL(team, '') AS team, IFNULL(region, '') AS region, IFNULL(zone, '') AS zone FROM `tabHQ Master` WHERE division=%s AND status='Active' ORDER BY name",
        (division,), as_dict=True,
    )
    zones = frappe.db.sql(
        "SELECT DISTINCT name, zone_name FROM `tabZone Master` WHERE status='Active' AND division IN (%s, 'Both') ORDER BY name",
        (division,), as_dict=True,
    )
    stockists = frappe.db.sql(
        "SELECT name, stockist_name, IFNULL(region, '') AS region FROM `tabStockist Master` WHERE division=%s AND status='Active' ORDER BY stockist_name",
        (division,), as_dict=True,
    )
    products = frappe.db.sql(
        "SELECT product_code, product_name, pack, COALESCE(category, '') AS category, "
        "COALESCE(product_group, '') AS product_group "
        "FROM `tabProduct Master` WHERE division=%s AND status='Active' "
        "ORDER BY COALESCE(sequence, 9999), product_code",
        (division,), as_dict=True,
    )
    product_groups = frappe.db.sql(
        "SELECT DISTINCT product_group FROM `tabProduct Master` "
        "WHERE division=%s AND status='Active' AND IFNULL(product_group,'')<>'' ORDER BY product_group",
        (division,), as_list=1,
    )
    product_categories = frappe.db.sql(
        "SELECT DISTINCT category FROM `tabProduct Master` "
        "WHERE division=%s AND status='Active' AND IFNULL(category,'')<>'' ORDER BY category",
        (division,), as_list=1,
    )
    months = frappe.db.sql(
        "SELECT DISTINCT upload_month FROM `tabPrimary Sales Data` WHERE division=%s ORDER BY upload_month DESC",
        (division,), as_list=1,
    )
    statement_months = frappe.db.sql(
        "SELECT DISTINCT DATE_FORMAT(statement_month, '%%Y-%%m-%%d') as m FROM `tabStockist Statement` WHERE division=%s AND docstatus=1 ORDER BY statement_month DESC",
        (division,), as_list=1,
    )

    # Confine a non-admin's pickers to their mapped regions (and the teams/HQs/stockists
    # inside them). Admin (allowed is None) sees everything.
    if allowed is not None:
        aset = set(allowed)
        regions = [r for r in regions if r.name in aset]
        teams = [t for t in teams if (t.region or "") in aset]
        hqs = [h for h in hqs if (h.region or "") in aset]
        stockists = [s for s in stockists if (s.region or "") in aset]

    return {
        "regions": [{"code": r.name, "name": r.region_name, "zone": r.zone} for r in regions],
        "teams": [{"code": t.name, "name": t.team_name, "region": t.region or "", "zone": t.zone or ""} for t in teams],
        "hqs": [{"code": h.name, "name": h.hq_name, "team": h.team or "", "region": h.region or "", "zone": h.zone or ""} for h in hqs],
        "zones": [{"code": z.name, "name": z.zone_name} for z in zones],
        "stockists": [{
            "code": s.name, "name": s.stockist_name} for s in stockists],
        "products": [{"code": p.product_code, "name": p.product_name or p.product_code,
                       "pack": p.pack or "", "category": p.category or "",
                       "product_group": p.product_group or ""} for p in products],
        "product_groups": [g[0] for g in product_groups],
        "product_categories": [c[0] for c in product_categories],
        "months": [m[0] for m in months],
        "statement_months": [m[0] for m in statement_months],
    }


@frappe.whitelist()
def get_stockist_primary_sales_report(division=None, sales_type="primary", region=None,
                                       from_date=None, to_date=None, team=None, hq=None,
                                       product_codes=None):
    """Report 1 – Stockist Wise Primary Sales Report.
    sales_type: 'primary' (iscancelled=0) or 'creditnote' (iscancelled=1).
    product_codes: optional list (or JSON-string list / comma-separated string) of
                   product codes to filter by.
    """
    if not division:
        division = get_user_division()
    is_cancelled = 1 if sales_type == "creditnote" else 0

    # Normalise product_codes input
    pcodes = _normalise_code_list(product_codes)

    conditions = ["psd.division = %(division)s", "psd.iscancelled = %(is_cancelled)s"]
    params = {"division": division, "is_cancelled": is_cancelled}

    # Primary Sales Data stores the region NAME, so match code-or-name.
    _scope_region_sql(conditions, params, "psd.region", division, region, include_names=True)
    if team:
        conditions.append("psd.team = %(team)s")
        params["team"] = team
    if hq:
        conditions.append("psd.stockist_code IN (SELECT name FROM `tabStockist Master` WHERE hq = %(hq)s)")
        params["hq"] = hq
    if from_date:
        conditions.append("psd.invoicedate >= %(from_date)s")
        params["from_date"] = from_date
    if to_date:
        conditions.append("psd.invoicedate <= %(to_date)s")
        params["to_date"] = to_date
    if pcodes:
        ph = ", ".join([f"%(_pc{i})s" for i in range(len(pcodes))])
        conditions.append(f"psd.pcode IN ({ph})")
        for i, c in enumerate(pcodes):
            params[f"_pc{i}"] = c

    where = " AND ".join(conditions)

    rows = frappe.db.sql(f"""
        SELECT COALESCE(sm.stockist_code, psd.stockist_code) AS stockist_code, psd.stockist_name,
               COALESCE(hm.hq_name, sm.hq, '') AS hq_name,
               psd.pcode AS product_code,
               psd.product AS product_name, psd.pack,
               SUM(psd.quantity) AS total_qty, SUM(psd.ptsvalue) AS total_value
        FROM `tabPrimary Sales Data` psd
        LEFT JOIN `tabStockist Master` sm ON sm.name = psd.stockist_code AND sm.division = %(division)s
        LEFT JOIN `tabHQ Master` hm ON hm.name = sm.hq AND hm.division = %(division)s
        WHERE {where}
        GROUP BY psd.stockist_code, sm.stockist_code, psd.stockist_name, hm.hq_name, sm.hq,
                 psd.pcode, psd.product, psd.pack
        ORDER BY psd.stockist_name, psd.pcode
    """, params, as_dict=True)

    seq_map = {r[0]: (r[1] if r[1] is not None else 9999) for r in frappe.db.sql(
        "SELECT product_code, COALESCE(sequence, 9999) FROM `tabProduct Master` WHERE division=%s AND status='Active'",
        (division,)
    )}

    # ── Pivot: products as rows, stockists as columns (single qty + value/cell) ──
    stockists = {}
    products = {}
    for r in rows:
        sc = r.stockist_code
        pc = r.product_code
        if sc and sc not in stockists:
            stockists[sc] = {"code": sc, "name": r.stockist_name or sc}
        if pc and pc not in products:
            products[pc] = {"code": pc, "name": r.product_name or "", "pack": r.pack or "",
                            "cells": {}, "total_qty": 0.0, "total_value": 0.0}
        if sc and pc:
            cell = products[pc]["cells"].setdefault(sc, {"qty": 0.0, "value": 0.0})
            cell["qty"] += flt(r.total_qty)
            cell["value"] += flt(r.total_value)
            products[pc]["total_qty"] += flt(r.total_qty)
            products[pc]["total_value"] += flt(r.total_value)

    stockist_list = sorted(stockists.values(), key=lambda s: (s["name"] or "").lower())
    product_list = sorted(products.values(),
                          key=lambda p: (seq_map.get(p["code"], 9999), p["code"] or ""))

    col_totals = {s["code"]: {"qty": 0.0, "value": 0.0} for s in stockist_list}
    grand_qty = 0.0
    grand_value = 0.0
    for p in product_list:
        for sc, vals in p["cells"].items():
            if sc in col_totals:
                col_totals[sc]["qty"] += vals["qty"]
                col_totals[sc]["value"] += vals["value"]
        grand_qty += p["total_qty"]
        grand_value += p["total_value"]

    return {
        "success": True,
        "stockists": stockist_list,
        "products": product_list,
        "col_totals": col_totals,
        "grand": {"qty": grand_qty, "value": grand_value},
    }


def _normalise_code_list(value):
    """Normalise a multi-select input into a list of strings.
    Accepts: list, JSON-string list, comma-separated string, or empty/None.
    """
    if not value:
        return []
    if isinstance(value, (list, tuple)):
        return [str(v).strip() for v in value if str(v).strip()]
    if isinstance(value, str):
        v = value.strip()
        if not v:
            return []
        if v.startswith("["):
            try:
                parsed = json.loads(v)
                if isinstance(parsed, (list, tuple)):
                    return [str(x).strip() for x in parsed if str(x).strip()]
            except Exception:
                pass
        return [p.strip() for p in v.split(",") if p.strip()]
    return []


@frappe.whitelist()
def get_stockist_secondary_sales_report(division=None, region=None,
                                         from_date=None, to_date=None, team=None, hq=None):
    """Report 2 – Stockist Wise Secondary Sales Report.

    Pivoted: products as rows, stockists as columns (each stockist column shows
    Before-Deduction and After-Deduction quantities).
    from_date / to_date are first-of-month dates (e.g. 2026-04-01).
    """
    if not division:
        division = get_user_division()

    conditions = ["ss.division = %(division)s", "ss.docstatus IN (0, 1)"]
    params = {"division": division}

    _scope_region_sql(conditions, params, "ss.region", division, region)
    if team:
        conditions.append("ss.team = %(team)s")
        params["team"] = team
    if hq:
        conditions.append("ss.hq = %(hq)s")
        params["hq"] = hq
    if from_date:
        conditions.append("ss.statement_month >= %(from_date)s")
        params["from_date"] = from_date
    if to_date:
        conditions.append("ss.statement_month <= %(to_date)s")
        params["to_date"] = to_date

    where = " AND ".join(conditions)

    rows = frappe.db.sql(f"""
        SELECT ss.stockist_code, ss.stockist_name,
               ss.hq AS hq_code,
               COALESCE(hm.hq_name, ss.hq, '') AS hq_name,
               COALESCE(tm.team_name, hm.team, ss.team, '') AS team_name,
               si.product_code, si.product_name, si.pack,
               SUM((si.sales_qty + si.free_qty) / IFNULL(NULLIF(si.conversion_factor, 0), 1)) AS qty_before,
               SUM((si.sales_qty + si.free_qty - IFNULL(si.free_qty_scheme, 0)) / IFNULL(NULLIF(si.conversion_factor, 0), 1)) AS qty_after,
               SUM(((si.sales_qty + si.free_qty) / IFNULL(NULLIF(si.conversion_factor, 0), 1)) * IFNULL(si.pts, 0)) AS val_before,
               SUM(((si.sales_qty + si.free_qty - IFNULL(si.free_qty_scheme, 0)) / IFNULL(NULLIF(si.conversion_factor, 0), 1)) * IFNULL(si.pts, 0)) AS val_after
        FROM `tabStockist Statement` ss
        INNER JOIN `tabStockist Statement Item` si
            ON si.parent = ss.name AND si.parenttype = 'Stockist Statement'
        LEFT JOIN `tabHQ Master` hm ON hm.name = ss.hq AND hm.division = %(division)s
        LEFT JOIN `tabTeam Master` tm ON tm.name = COALESCE(hm.team, ss.team)
        WHERE {where}
        GROUP BY ss.stockist_code, ss.stockist_name, ss.hq, ss.team, hm.hq_name, hm.team, tm.team_name,
                 si.product_code, si.product_name, si.pack
        ORDER BY team_name, hq_name, ss.stockist_name, si.product_code
    """, params, as_dict=True)

    # Remap the internal id to the human-facing Stockist Code up front so the pivot
    # columns, cell keys and labels all stay consistent.
    code_map = get_stockist_code_map([r.stockist_code for r in rows])
    for r in rows:
        if r.stockist_code:
            r.stockist_code = code_map.get(r.stockist_code, r.stockist_code)
    # Same for products: rows link by the Product Master id; the pivot keys,
    # seq_map lookups and labels all use the business Product Code.
    _apply_product_display_codes(rows)

    # Stockist order (headquarter-wise: Team → HQ → name) and product order (master sequence)
    stockists = {}
    products = {}
    for r in rows:
        if r.stockist_code and r.stockist_code not in stockists:
            stockists[r.stockist_code] = {
                "code": r.stockist_code,
                "name": r.stockist_name or r.stockist_code,
                "hq_name": r.hq_name or "",
                "team_name": r.team_name or "",
            }
        if r.product_code and r.product_code not in products:
            products[r.product_code] = {
                "code": r.product_code,
                "name": r.product_name or "",
                "pack": r.pack or "",
                "cells": {},
                "total_before": 0.0,
                "total_after": 0.0,
                "total_val_before": 0.0,
                "total_val_after": 0.0,
            }
        if r.stockist_code and r.product_code:
            cell = products[r.product_code]["cells"].setdefault(
                r.stockist_code, {"before": 0.0, "after": 0.0, "val_before": 0.0, "val_after": 0.0}
            )
            cell["before"] += flt(r.qty_before)
            cell["after"] += flt(r.qty_after)
            cell["val_before"] += flt(r.val_before)
            cell["val_after"] += flt(r.val_after)
            products[r.product_code]["total_before"] += flt(r.qty_before)
            products[r.product_code]["total_after"] += flt(r.qty_after)
            products[r.product_code]["total_val_before"] += flt(r.val_before)
            products[r.product_code]["total_val_after"] += flt(r.val_after)

    seq_map = {r[0]: (r[1] if r[1] is not None else 9999) for r in frappe.db.sql(
        "SELECT product_code, COALESCE(sequence, 9999) FROM `tabProduct Master` WHERE division=%s AND status='Active'",
        (division,)
    )}

    stockist_list = sorted(
        stockists.values(),
        key=lambda s: ((s.get("team_name") or "").lower(),
                       (s.get("hq_name") or "").lower(),
                       (s.get("name") or "").lower())
    )
    product_list = sorted(
        products.values(),
        key=lambda p: (seq_map.get(p["code"], 9999), p["code"] or "")
    )

    # Per-stockist column totals + grand totals
    col_totals = {s["code"]: {"before": 0.0, "after": 0.0, "val_before": 0.0, "val_after": 0.0} for s in stockist_list}
    grand_before = 0.0
    grand_after = 0.0
    grand_val_before = 0.0
    grand_val_after = 0.0
    for p in product_list:
        for sc, vals in p["cells"].items():
            if sc in col_totals:
                col_totals[sc]["before"] += vals["before"]
                col_totals[sc]["after"] += vals["after"]
                col_totals[sc]["val_before"] += vals.get("val_before", 0.0)
                col_totals[sc]["val_after"] += vals.get("val_after", 0.0)
        grand_before += p["total_before"]
        grand_after += p["total_after"]
        grand_val_before += p["total_val_before"]
        grand_val_after += p["total_val_after"]

    return {
        "success": True,
        "stockists": stockist_list,
        "products": product_list,
        "col_totals": col_totals,
        "grand": {
            "before": grand_before, "after": grand_after,
            "val_before": grand_val_before, "val_after": grand_val_after,
        },
        "from_date": from_date,
        "to_date": to_date,
    }


@frappe.whitelist()
def get_stockist_moving_trend_report(division=None, sales_type="secondary",
                                      stockist_code=None, product_codes=None):
    """Report 3 – Moving Trend (monthly pivot Apr-Mar) for a single stockist.
    sales_type: 'primary' uses Primary Sales Data; 'secondary' uses Statement Items.
    product_codes: optional multi-select product filter.
    Returns per-product monthly QTY and monthly VALUE, plus aggregated value-in-lakhs row.
    """
    if not division:
        division = get_user_division()
    if not stockist_code:
        return {"success": False, "message": "Stockist code is required"}

    # Non-admins may only view stockists in their mapped regions — a stockist outside
    # scope yields no rows (the query below is keyed on this code).
    if not _stockist_region_allowed(stockist_code, division):
        stockist_code = "__region_denied__"

    month_labels = ["apr", "may", "jun", "jul", "aug", "sep",
                    "oct", "nov", "dec", "jan", "feb", "mar"]

    pcodes = _normalise_code_list(product_codes)

    if sales_type == "primary":
        # Primary Sales Data stores the raw business code in `pcode`, so the UI's
        # business-code filter applies directly.
        pcode_sub = ""
        extra_params = {}
        if pcodes:
            ph = ", ".join([f"%(_pc{i})s" for i in range(len(pcodes))])
            pcode_sub = f" AND pcode IN ({ph})"
            for i, c in enumerate(pcodes):
                extra_params[f"_pc{i}"] = c
        base_params = {"division": division, "stockist": stockist_code, **extra_params}
        rows = frappe.db.sql(f"""
            SELECT pcode AS product_code, product AS product_name, pack,
                   MONTH(invoicedate) AS m, YEAR(invoicedate) AS y,
                   SUM(quantity) AS qty,
                   SUM(ptsvalue) AS value
            FROM `tabPrimary Sales Data`
            WHERE division = %(division)s AND stockist_code = %(stockist)s
                  AND iscancelled = 0{pcode_sub}
            GROUP BY pcode, product, pack, MONTH(invoicedate), YEAR(invoicedate)
            ORDER BY pcode, y, m
        """, base_params, as_dict=True)
    else:
        # Secondary sales = sales + free − scheme free (the canonical "true sales" figure).
        # Statement items link by the Product Master id, so resolve the UI's
        # business codes to ids within the division before filtering.
        pcode_sub = ""
        extra_params = {}
        if pcodes:
            resolved_pks = _resolve_product_filter(division, product_codes=pcodes) or ["__no_match__"]
            ph = ", ".join([f"%(_pc{i})s" for i in range(len(resolved_pks))])
            pcode_sub = f" AND si.product_code IN ({ph})"
            for i, c in enumerate(resolved_pks):
                extra_params[f"_pc{i}"] = c
        base_params = {"division": division, "stockist": stockist_code, **extra_params}
        rows = frappe.db.sql(f"""
            SELECT si.product_code, si.product_name, si.pack,
                   MONTH(ss.statement_month) AS m, YEAR(ss.statement_month) AS y,
                   SUM((si.sales_qty + si.free_qty - IFNULL(si.free_qty_scheme, 0)) / IFNULL(NULLIF(si.conversion_factor, 0), 1)) AS qty,
                   SUM(((si.sales_qty + si.free_qty - IFNULL(si.free_qty_scheme, 0)) / IFNULL(NULLIF(si.conversion_factor, 0), 1)) * IFNULL(si.pts, 0)) AS value
            FROM `tabStockist Statement` ss
            INNER JOIN `tabStockist Statement Item` si
                ON si.parent = ss.name AND si.parenttype = 'Stockist Statement'
            WHERE ss.division = %(division)s AND ss.stockist_code = %(stockist)s
                  AND ss.docstatus IN (0, 1){pcode_sub}
            GROUP BY si.product_code, si.product_name, si.pack,
                     MONTH(ss.statement_month), YEAR(ss.statement_month)
            ORDER BY si.product_code, y, m
        """, base_params, as_dict=True)
        # Show business codes, never internal ids.
        _apply_product_display_codes(rows)

    # Look up stockist metadata (name, HQ name) up front so we always have the
    # HQ even when the result is empty.
    stk_info = frappe.db.sql(
        """SELECT sm.stockist_name, COALESCE(hm.hq_name, sm.hq, '') AS hq_name
             FROM `tabStockist Master` sm
        LEFT JOIN `tabHQ Master` hm ON hm.name = sm.hq AND hm.division = %s
            WHERE sm.name = %s AND sm.division = %s""",
        (division, stockist_code, division), as_dict=True,
    )
    stockist_name = stk_info[0].stockist_name if stk_info else stockist_code
    hq_name = stk_info[0].hq_name if stk_info else ""

    # Derive the financial year range from data
    if not rows:
        return {"success": True, "data": [], "fy_label": "",
                "stockist_name": stockist_name, "hq_name": hq_name,
                "value_in_lakhs": {"months": [0] * 12, "total": 0.0}}

    all_dates = [(r.y, r.m) for r in rows]
    min_y = min(d[0] for d in all_dates)
    max_y = max(d[0] for d in all_dates)

    # Determine FY boundaries (Apr of min year to Mar of max year+1)
    fy_start_year = min_y if min(d[1] for d in all_dates if d[0] == min_y) >= 4 else min_y - 1
    fy_end_year = max_y + 1 if max(d[1] for d in all_dates if d[0] == max_y) >= 4 else max_y
    fy_label = f"Apr {str(fy_start_year)[2:]} to Mar {str(fy_end_year)[2:]}"

    # Map month number to FY column index: Apr(4)→0 … Mar(3)→11
    month_map = {4: 0, 5: 1, 6: 2, 7: 3, 8: 4, 9: 5,
                 10: 6, 11: 7, 12: 8, 1: 9, 2: 10, 3: 11}

    products = {}
    monthly_value = [0.0] * 12
    grand_value = 0.0
    for r in rows:
        key = r.product_code
        if key not in products:
            products[key] = {
                "product_code": r.product_code,
                "product_name": r.product_name,
                "pack": r.pack,
                "months": [0] * 12,
                "total": 0,
            }
        idx = month_map.get(r.m)
        if idx is not None:
            products[key]["months"][idx] += flt(r.qty)
            products[key]["total"] += flt(r.qty)
            monthly_value[idx] += flt(r.value)
            grand_value += flt(r.value)

    seq_map = {r[0]: (r[1] if r[1] is not None else 9999) for r in frappe.db.sql(
        "SELECT product_code, COALESCE(sequence, 9999) FROM `tabProduct Master` WHERE division=%s AND status='Active'",
        (division,)
    )}
    data = sorted(products.values(), key=lambda p: (seq_map.get(p["product_code"], 9999), p["product_code"] or ""))

    # Aggregate value-in-lakhs row (monthly + total)
    value_in_lakhs = {
        "months": [round(v / 100000, 2) for v in monthly_value],
        "total": round(grand_value / 100000, 2),
    }
    return {
        "success": True,
        "data": data,
        "fy_label": fy_label,
        "month_labels": month_labels,
        "stockist_name": stockist_name,
        "hq_name": hq_name,
        "value_in_lakhs": value_in_lakhs,
    }


@frappe.whitelist()
def get_stockist_closing_stock_report(division=None, region=None,
                                       from_date=None, to_date=None, group_by="stockist", team=None, hq=None):
    """Report 4 – Stockist Wise Closing Stock Report (from Draft or submitted Stockist Statements).
    group_by: 'stockist' (default) or 'hq'
    """
    if not division:
        division = get_user_division()

    conditions = ["ss.division = %(division)s", "ss.docstatus IN (0, 1)"]
    params = {"division": division}

    _scope_region_sql(conditions, params, "ss.region", division, region)
    if team:
        conditions.append("ss.team = %(team)s")
        params["team"] = team
    if hq:
        conditions.append("ss.hq = %(hq)s")
        params["hq"] = hq
    if from_date:
        conditions.append("ss.statement_month >= %(from_date)s")
        params["from_date"] = from_date
    if to_date:
        conditions.append("ss.statement_month <= %(to_date)s")
        params["to_date"] = to_date

    where = " AND ".join(conditions)

    if group_by == "hq":
        rows = frappe.db.sql(f"""
            SELECT ss.hq AS hq_code,
                   COALESCE(hm.hq_name, ss.hq, '') AS hq_name,
                   si.product_code, si.product_name, si.pack,
                     SUM(si.opening_qty / IFNULL(NULLIF(si.conversion_factor, 0), 1)) AS opening_qty,
                     SUM(si.purchase_qty / IFNULL(NULLIF(si.conversion_factor, 0), 1)) AS purchase_qty,
                     SUM(si.sales_qty / IFNULL(NULLIF(si.conversion_factor, 0), 1)) AS sales_qty,
                     SUM(si.free_qty / IFNULL(NULLIF(si.conversion_factor, 0), 1)) AS free_qty,
                     SUM(si.free_qty_scheme / IFNULL(NULLIF(si.conversion_factor, 0), 1)) AS scheme_free_qty,
                     SUM(si.closing_qty / IFNULL(NULLIF(si.conversion_factor, 0), 1)) AS closing_qty
            FROM `tabStockist Statement` ss
            INNER JOIN `tabStockist Statement Item` si
                ON si.parent = ss.name AND si.parenttype = 'Stockist Statement'
            LEFT JOIN `tabHQ Master` hm ON hm.name = ss.hq AND hm.division = %(division)s
            WHERE {where}
            GROUP BY ss.hq, hm.hq_name,
                     si.product_code, si.product_name, si.pack
            ORDER BY hm.hq_name, si.product_code
        """, params, as_dict=True)
    else:
        rows = frappe.db.sql(f"""
            SELECT ss.stockist_code, ss.stockist_name,
                   si.product_code, si.product_name, si.pack,
                     SUM(si.opening_qty / IFNULL(NULLIF(si.conversion_factor, 0), 1)) AS opening_qty,
                     SUM(si.purchase_qty / IFNULL(NULLIF(si.conversion_factor, 0), 1)) AS purchase_qty,
                     SUM(si.sales_qty / IFNULL(NULLIF(si.conversion_factor, 0), 1)) AS sales_qty,
                     SUM(si.free_qty / IFNULL(NULLIF(si.conversion_factor, 0), 1)) AS free_qty,
                     SUM(si.free_qty_scheme / IFNULL(NULLIF(si.conversion_factor, 0), 1)) AS scheme_free_qty,
                     SUM(si.closing_qty / IFNULL(NULLIF(si.conversion_factor, 0), 1)) AS closing_qty
            FROM `tabStockist Statement` ss
            INNER JOIN `tabStockist Statement Item` si
                ON si.parent = ss.name AND si.parenttype = 'Stockist Statement'
            WHERE {where}
            GROUP BY ss.stockist_code, ss.stockist_name,
                     si.product_code, si.product_name, si.pack
            ORDER BY ss.stockist_name, si.product_code
        """, params, as_dict=True)

    # Rows link by the Product Master id; seq_map and the UI use business codes.
    _apply_product_display_codes(rows)

    seq_map = {r[0]: (r[1] if r[1] is not None else 9999) for r in frappe.db.sql(
        "SELECT product_code, COALESCE(sequence, 9999) FROM `tabProduct Master` WHERE division=%s AND status='Active'",
        (division,)
    )}
    if group_by == "hq":
        rows.sort(key=lambda r: (r.hq_name or "", seq_map.get(r.product_code, 9999), r.product_code or ""))
    else:
        rows.sort(key=lambda r: (r.stockist_name or "", seq_map.get(r.product_code, 9999), r.product_code or ""))

    return {"success": True, "data": rows, "group_by": group_by}


@frappe.whitelist()
def get_region_product_closing_stock(division=None, region=None, from_date=None, to_date=None, group_by="hq", sales_mode="after_deduction"):
    """Report 9 – Product Closing Stock for Region, pivoted by HQ or Stockist (grouped by Team).
    Returns products as rows, HQs/Stockists as columns grouped by team, with Team Totals and Region Total.
    group_by: 'hq' (default) or 'stockist'
    """
    if not division:
        division = get_user_division()

    # "All Regions (Organization)" view: no single region is selected. Instead of
    # one column per HQ/Stockist within a region, we roll one level up the
    # hierarchy (Division -> Zone -> Region) and show one column per Region grouped
    # by Zone. The response reuses the exact same pivot shape (Zone plays the
    # "team" role, Region plays the "column" role) so the render/PDF/Excel pipeline
    # stays unchanged apart from the group/total labels.
    is_org = (not region) or str(region).upper() in ("__ALL__", "ALL", "ORG", "ORGANIZATION")

    conditions = ["ss.division = %(division)s", "ss.docstatus IN (0, 1)"]
    params = {"division": division}
    # Org (all-regions) view: admins span every region; non-admins are limited to their
    # mapped regions. A specific region is validated against the user's allowed set.
    _scope_region_sql(conditions, params, "ss.region", division, None if is_org else region)

    if from_date:
        conditions.append("ss.statement_month >= %(from_date)s")
        params["from_date"] = from_date
    if to_date:
        conditions.append("ss.statement_month <= %(to_date)s")
        params["to_date"] = to_date

    where = " AND ".join(conditions)

    if is_org:
        # Organization view: one column per Region, grouped by Zone
        rows = frappe.db.sql(f"""
            SELECT
                COALESCE(hm.region, ss.region, '') AS col_code,
                COALESCE(rm.region_name, hm.region, ss.region, '') AS col_name,
                COALESCE(rm.zone, '') AS team_code,
                COALESCE(zm.zone_name, rm.zone, '') AS team_name,
                si.product_code,
                si.product_name,
                si.pack,
                SUM(si.closing_qty / IFNULL(NULLIF(si.conversion_factor, 0), 1)) AS closing_qty,
                SUM(COALESCE(si.closing_value, (si.closing_qty / IFNULL(NULLIF(si.conversion_factor, 0), 1)) * si.pts, 0)) AS closing_value
            FROM `tabStockist Statement` ss
            INNER JOIN `tabStockist Statement Item` si
                ON si.parent = ss.name AND si.parenttype = 'Stockist Statement'
            LEFT JOIN `tabHQ Master` hm ON hm.name = ss.hq AND hm.division = %(division)s
            LEFT JOIN `tabRegion Master` rm ON rm.name = COALESCE(hm.region, ss.region)
            LEFT JOIN `tabZone Master` zm ON zm.name = rm.zone
            WHERE {where}
            GROUP BY COALESCE(hm.region, ss.region, ''), rm.region_name, rm.zone, zm.zone_name,
                     si.product_code, si.product_name, si.pack
            ORDER BY zm.zone_name, rm.region_name, si.product_code
        """, params, as_dict=True)

        # Ordered Region column list grouped by Zone (from masters)
        col_list = frappe.db.sql("""
            SELECT rm.name AS col_code, COALESCE(rm.region_name, rm.name, '') AS col_name,
                   COALESCE(rm.zone, '') AS team_code,
                   COALESCE(zm.zone_name, rm.zone, '') AS team_name
            FROM `tabRegion Master` rm
            LEFT JOIN `tabZone Master` zm ON zm.name = rm.zone
            WHERE rm.division IN (%(division)s, 'Both') AND rm.status = 'Active'
            ORDER BY zm.zone_name, rm.region_name
        """, params, as_dict=True)
    elif group_by == "stockist":
        # Stockist-wise pivot: one column per stockist
        rows = frappe.db.sql(f"""
            SELECT
                ss.stockist_code AS col_code,
                ss.stockist_name AS col_name,
                COALESCE(hm.team, ss.team, '') AS team_code,
                COALESCE(tm.team_name, hm.team, ss.team, '') AS team_name,
                si.product_code,
                si.product_name,
                si.pack,
                SUM(si.closing_qty / IFNULL(NULLIF(si.conversion_factor, 0), 1)) AS closing_qty,
                SUM(COALESCE(si.closing_value, (si.closing_qty / IFNULL(NULLIF(si.conversion_factor, 0), 1)) * si.pts, 0)) AS closing_value
            FROM `tabStockist Statement` ss
            INNER JOIN `tabStockist Statement Item` si
                ON si.parent = ss.name AND si.parenttype = 'Stockist Statement'
            LEFT JOIN `tabHQ Master` hm ON hm.name = ss.hq AND hm.division = %(division)s
            LEFT JOIN `tabTeam Master` tm ON tm.name = COALESCE(hm.team, ss.team)
            WHERE {where}
            GROUP BY ss.stockist_code, ss.stockist_name, hm.hq_name, hm.team, tm.team_name,
                     si.product_code, si.product_name, si.pack
            ORDER BY hm.team, hm.hq_name, ss.stockist_name, si.product_code
        """, params, as_dict=True)

        # Build ordered column list from Stockist Master for this region
        col_list = frappe.db.sql("""
            SELECT sm.name AS col_code, sm.stockist_name AS col_name,
                   COALESCE(hm.team, sm.team, '') AS team_code,
                   COALESCE(tm.team_name, hm.team, sm.team, '') AS team_name
            FROM `tabStockist Master` sm
            LEFT JOIN `tabHQ Master` hm ON hm.name = sm.hq
            LEFT JOIN `tabTeam Master` tm ON tm.name = COALESCE(hm.team, sm.team)
            WHERE sm.division = %(division)s AND sm.region = %(region)s AND sm.status = 'Active'
            ORDER BY hm.team, hm.hq_name, sm.stockist_name
        """, params, as_dict=True)
    else:
        # HQ-wise pivot: one column per HQ
        rows = frappe.db.sql(f"""
            SELECT
                ss.hq AS col_code,
                COALESCE(hm.hq_name, ss.hq, '') AS col_name,
                COALESCE(hm.team, ss.team, '') AS team_code,
                COALESCE(tm.team_name, hm.team, ss.team, '') AS team_name,
                si.product_code,
                si.product_name,
                si.pack,
                SUM(si.closing_qty / IFNULL(NULLIF(si.conversion_factor, 0), 1)) AS closing_qty,
                SUM(COALESCE(si.closing_value, (si.closing_qty / IFNULL(NULLIF(si.conversion_factor, 0), 1)) * si.pts, 0)) AS closing_value
            FROM `tabStockist Statement` ss
            INNER JOIN `tabStockist Statement Item` si
                ON si.parent = ss.name AND si.parenttype = 'Stockist Statement'
            LEFT JOIN `tabHQ Master` hm ON hm.name = ss.hq AND hm.division = %(division)s
            LEFT JOIN `tabTeam Master` tm ON tm.name = COALESCE(hm.team, ss.team)
            WHERE {where}
            GROUP BY ss.hq, hm.hq_name, hm.team, tm.team_name,
                     si.product_code, si.product_name, si.pack
            ORDER BY hm.team, hm.hq_name, si.product_code
        """, params, as_dict=True)

        # Build ordered HQ column list from HQ Master for this region
        col_list = frappe.db.sql("""
            SELECT hm.name AS col_code, COALESCE(hm.hq_name, hm.name, '') AS col_name,
                   COALESCE(hm.team, '') AS team_code,
                   COALESCE(tm.team_name, hm.team, '') AS team_name
            FROM `tabHQ Master` hm
            LEFT JOIN `tabTeam Master` tm ON tm.name = hm.team
            WHERE hm.division = %(division)s AND hm.region = %(region)s AND hm.status = 'Active'
            ORDER BY hm.team, hm.hq_name
        """, params, as_dict=True)

    # Rows link by the Product Master id; pivot keys and labels use business codes.
    _apply_product_display_codes(rows)

    # Header/label hints so the same render pipeline can title the grouping and
    # totals correctly for the per-region ("Team"/"Region") vs organization
    # ("Zone"/"Organization") views.
    group_label = "Zone" if is_org else "Team"
    total_label = "Organization" if is_org else "Region"
    if is_org:
        region_name = "All Regions (Organization)"
    else:
        region_name = frappe.db.get_value("Region Master", region, "region_name") or region

    if not rows:
        return {"success": True, "products": [], "team_order": [], "hq_columns": [],
                "value_in_lakhs": {}, "region": "" if is_org else region,
                "region_name": region_name, "group_by": group_by,
                "is_org": is_org, "group_label": group_label, "total_label": total_label}

    # Fall back to data-derived column list if master has no entries
    if not col_list:
        seen = {}
        for r in rows:
            if r.col_code and r.col_code not in seen:
                seen[r.col_code] = {"col_code": r.col_code, "col_name": r.col_name,
                                    "team_code": r.team_code, "team_name": r.team_name}
        col_list = list(seen.values())

    # Build team order: only include teams/columns that actually have data
    data_col_codes = {r.col_code for r in rows if r.col_code}
    team_map = {}
    for c in col_list:
        tc = c.team_code or "Unknown"
        if tc not in team_map:
            team_map[tc] = {"team_code": tc, "team_name": c.team_name or tc, "hqs": []}
        if c.col_code in data_col_codes:
            team_map[tc]["hqs"].append({"col_code": c.col_code, "col_name": c.col_name})

    # Include any data columns not in master
    configured_cols = {c.col_code for c in col_list}
    for r in rows:
        if r.col_code and r.col_code not in configured_cols:
            tc = r.team_code or "Unknown"
            if tc not in team_map:
                team_map[tc] = {"team_code": tc, "team_name": r.team_name or tc, "hqs": []}
            if not any(h["col_code"] == r.col_code for h in team_map[tc]["hqs"]):
                team_map[tc]["hqs"].append({"col_code": r.col_code, "col_name": r.col_name})

    # Filter teams with at least one column
    team_order = [t for t in team_map.values() if t["hqs"]]
    flat_col_codes = [h["col_code"] for t in team_order for h in t["hqs"]]

    # Build product data matrix
    product_data = {}
    for r in rows:
        pc = r.product_code
        if pc not in product_data:
            product_data[pc] = {
                "product_code": pc,
                "product_name": r.product_name,
                "pack": r.pack,
                "col_qty": {},
                "col_value": {},
            }
        cc = r.col_code
        product_data[pc]["col_qty"][cc] = product_data[pc]["col_qty"].get(cc, 0) + (float(r.closing_qty) if r.closing_qty else 0)
        product_data[pc]["col_value"][cc] = product_data[pc]["col_value"].get(cc, 0) + (float(r.closing_value) if r.closing_value else 0)

    # Sort products by sequence then product_code; exclude rows with empty/None product_code
    _seq_map = {r[0]: (r[1] if r[1] is not None else 9999) for r in frappe.db.sql(
        "SELECT product_code, COALESCE(sequence, 9999) FROM `tabProduct Master` WHERE division=%s AND status='Active'",
        (division,)
    )}
    sorted_products = sorted(
        [p for p in product_data.values() if p["product_code"]],
        key=lambda x: (_seq_map.get(x["product_code"], 9999), x["product_code"])
    )

    # Compute value_in_lakhs (total across all products per column/team/region)
    col_total_value = {}
    for pdata in sorted_products:
        for cc, val in pdata["col_value"].items():
            col_total_value[cc] = col_total_value.get(cc, 0) + val

    team_total_value = {}
    for t in team_order:
        tv = sum(col_total_value.get(h["col_code"], 0) for h in t["hqs"])
        team_total_value[t["team_code"]] = tv

    region_total_value = sum(col_total_value.values())

    # Build product rows
    product_rows = []
    for sno, pdata in enumerate(sorted_products, 1):
        team_qty = {}
        for t in team_order:
            team_qty[t["team_code"]] = sum(pdata["col_qty"].get(h["col_code"], 0) for h in t["hqs"])
        region_qty = sum(pdata["col_qty"].values())
        product_rows.append({
            "sno": sno,
            "product_code": pdata["product_code"],
            "product_name": pdata["product_name"],
            "pack": pdata["pack"],
            "col_qty": pdata["col_qty"],
            "team_qty": team_qty,
            "region_qty": region_qty,
        })

    return {
        "success": True,
        "region": "" if is_org else region,
        "region_name": region_name,
        "is_org": is_org,
        "group_label": group_label,
        "total_label": total_label,
        "from_date": from_date,
        "to_date": to_date,
        "group_by": group_by,
        "team_order": team_order,
        "hq_columns": flat_col_codes,
        "products": product_rows,
        "value_in_lakhs": {
            "col_values": {cc: round(v / 100000, 2) for cc, v in col_total_value.items()},
            "team_values": {tc: round(v / 100000, 2) for tc, v in team_total_value.items()},
            "region_total": round(region_total_value / 100000, 2),
        },
    }


@frappe.whitelist()
def get_hq_wise_stockist_report(division=None, region=None, team=None, hq=None):
    """Report 5 – HQ Wise Stockist Report (active stockists grouped by HQ)."""
    if not division:
        division = get_user_division()

    conditions = ["sm.division = %(division)s", "sm.status = 'Active'"]
    params = {"division": division}

    _scope_region_sql(conditions, params, "COALESCE(hm.region, sm.region)", division, region)
    if team:
        conditions.append("COALESCE(hm.team, sm.team) = %(team)s")
        params["team"] = team
    if hq:
        conditions.append("sm.hq = %(hq)s")
        params["hq"] = hq

    where = " AND ".join(conditions)

    rows = frappe.db.sql(f"""
        SELECT COALESCE(sm.stockist_code, sm.name) AS stockist_code, sm.stockist_name,
               sm.hq, COALESCE(hm.hq_name, sm.hq) AS hq_name,
               COALESCE(hm.team, sm.team, '') AS team,
               COALESCE(tm.team_name, hm.team, sm.team, '') AS team_name,
               COALESCE(hm.region, sm.region) AS region,
               COALESCE(rm.region_name, hm.region, sm.region, '') AS region_name
        FROM `tabStockist Master` sm
        LEFT JOIN `tabHQ Master` hm ON hm.name = sm.hq
        LEFT JOIN `tabTeam Master` tm ON tm.name = COALESCE(hm.team, sm.team)
        LEFT JOIN `tabRegion Master` rm ON rm.name = COALESCE(hm.region, sm.region)
        WHERE {where}
        ORDER BY hm.hq_name, sm.stockist_name
    """, params, as_dict=True)

    # Group by HQ
    grouped = {}
    for r in rows:
        hq_key = r.hq or "Unassigned"
        if hq_key not in grouped:
            grouped[hq_key] = {
                "hq_code": r.hq,
                "hq_name": r.hq_name,
                "team": r.team_name or r.team,
                "region": r.region_name or r.region,
                "stockists": [],
            }
        grouped[hq_key]["stockists"].append({
            "stockist_code": r.stockist_code,
            "stockist_name": r.stockist_name,
        })

    return {"success": True, "data": list(grouped.values())}


@frappe.whitelist()
def get_stockist_address_report(division=None, region=None, criteria="ALL", team=None, hq=None):
    """Report 6 – Stockist Address Report.
    criteria: ALL, HQ WISE, TEAM WISE, CITY WISE, STOCKIST NAME
    """
    if not division:
        division = get_user_division()

    conditions = ["sm.division = %(division)s", "sm.status = 'Active'"]
    params = {"division": division}

    _scope_region_sql(conditions, params, "sm.region", division, region)
    if team:
        conditions.append("sm.team = %(team)s")
        params["team"] = team
    if hq:
        conditions.append("sm.hq = %(hq)s")
        params["hq"] = hq

    where = " AND ".join(conditions)

    order_map = {
        "ALL": "sm.stockist_name",
        "HQ WISE": "sm.hq, sm.stockist_name",
        "TEAM WISE": "sm.team, sm.stockist_name",
        "CITY WISE": "sm.city, sm.stockist_name",
        "STOCKIST NAME": "sm.stockist_name",
    }
    order_by = order_map.get(criteria, "sm.stockist_name")

    rows = frappe.db.sql(f"""
        SELECT COALESCE(sm.stockist_code, sm.name) AS stockist_code, sm.stockist_name,
               sm.address, sm.city, sm.phone,
               sm.hq, COALESCE(hm.hq_name, sm.hq, '') AS hq_name,
               sm.team, COALESCE(tm.team_name, sm.team, '') AS team_name,
               sm.region, COALESCE(rm.region_name, sm.region, '') AS region_name
        FROM `tabStockist Master` sm
        LEFT JOIN `tabHQ Master` hm ON hm.name = sm.hq
        LEFT JOIN `tabTeam Master` tm ON tm.name = sm.team
        LEFT JOIN `tabRegion Master` rm ON rm.name = sm.region
        WHERE {where}
        ORDER BY {order_by}
    """, params, as_dict=True)

    # If grouping is needed, build groups
    group_field = None
    if criteria == "HQ WISE":
        group_field = "hq_name"
    elif criteria == "TEAM WISE":
        group_field = "team_name"
    elif criteria == "CITY WISE":
        group_field = "city"

    if group_field:
        grouped = {}
        for r in rows:
            key = r.get(group_field) or "Unassigned"
            grouped.setdefault(key, []).append(r)
        return {"success": True, "data": rows, "groups": grouped, "group_field": group_field}

    return {"success": True, "data": rows, "groups": None, "group_field": None}


@frappe.whitelist()
def export_stockist_report_excel(report_type, division=None, **kwargs):
    """Generate a styled Excel workbook for any of the 6 stockist report types."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    if not division:
        division = get_user_division()

    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    group_font = Font(bold=True, size=11)
    group_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

    wb = openpyxl.Workbook()
    ws = wb.active

    def write_header_row(ws, row_num, headers):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

    def write_data_row(ws, row_num, values):
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=v)
            cell.border = thin_border

    def write_group_row(ws, row_num, label, num_cols):
        cell = ws.cell(row=row_num, column=1, value=label)
        cell.font = group_font
        cell.fill = group_fill
        for c in range(1, num_cols + 1):
            ws.cell(row=row_num, column=c).fill = group_fill
            ws.cell(row=row_num, column=c).border = thin_border

    def write_title_rows(ws, title, subtitle=""):
        ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=14)
        if subtitle:
            ws.cell(row=2, column=1, value=subtitle).font = Font(size=11, italic=True)
        return 4  # data starts at row 4

    region = kwargs.get("region", "")
    from_date = kwargs.get("from_date", "")
    to_date = kwargs.get("to_date", "")
    sales_type = kwargs.get("sales_type", "primary")
    stockist_code = kwargs.get("stockist_code", "")
    criteria = kwargs.get("criteria", "ALL")
    product_codes = kwargs.get("product_codes", "")
    period_label = f"{from_date} to {to_date}" if from_date and to_date else ""
    region_label = region or "All Regions"

    bold_value_font = Font(bold=True, size=11)
    bold_value_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

    def write_value_row(ws, row_num, values):
        """Bold value row (top-of-section monetary aggregate)."""
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=v)
            cell.font = bold_value_font
            cell.fill = bold_value_fill
            cell.border = thin_border
            if col > 1:
                cell.alignment = Alignment(horizontal="right")

    if report_type == "primary_sales":
        # Pivot: Products as rows, Stockists as columns (Quantity). Trailing Total (Rs. Lakhs).
        ws.title = "Primary Sales"
        result = get_stockist_primary_sales_report(division, sales_type, region, from_date, to_date,
                                                    kwargs.get("team") or None, kwargs.get("hq") or None,
                                                    product_codes or None)
        stockist_list = result.get("stockists", []) or []
        product_list = result.get("products", []) or []
        col_totals = result.get("col_totals", {}) or {}
        grand = result.get("grand", {}) or {}
        type_label = "Credit Note" if sales_type == "creditnote" else "Primary"
        row = write_title_rows(ws, f"Stockist Wise {type_label} Sales Report – {division}",
                               f"Region: {region_label}  |  Period: {period_label}")

        lakh_header_font = Font(bold=True, size=11, color="FFFFFF")
        lakh_header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        lakh_cell_font = Font(bold=True, size=11, color="1D4ED8")
        lakh_cell_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
        right_align = Alignment(horizontal="right")

        headers = ["Product Code", "Pack"] + [s["name"] for s in stockist_list]
        write_header_row(ws, row, headers)
        total_col = len(headers) + 1
        cell = ws.cell(row=row, column=total_col, value="Total (Rs. Lakhs)")
        cell.font = lakh_header_font; cell.fill = lakh_header_fill
        cell.alignment = header_align; cell.border = thin_border
        row += 1

        LAKH = 100000.0

        def _z(v):
            return v if v else None

        for p in product_list:
            vals = [p["code"], p["pack"]]
            for s in stockist_list:
                cv = (p["cells"] or {}).get(s["code"], {})
                vals.append(_z(round(flt(cv.get("qty")), 2)))
            write_data_row(ws, row, vals)
            for c in range(3, len(headers) + 1):
                ws.cell(row=row, column=c).alignment = right_align
            tcell = ws.cell(row=row, column=total_col, value=_z(round(flt(p.get("total_value", 0)) / LAKH, 2)))
            tcell.font = lakh_cell_font; tcell.fill = lakh_cell_fill
            tcell.alignment = right_align; tcell.border = thin_border
            row += 1

        # Grand "Value in Lakhs" row — per stockist + grand total
        glabel = ws.cell(row=row, column=1, value="Value in Lakhs")
        glabel.font = lakh_cell_font; glabel.fill = lakh_cell_fill; glabel.border = thin_border
        ws.cell(row=row, column=2).fill = lakh_cell_fill
        ws.cell(row=row, column=2).border = thin_border
        c_cur = 3
        for s in stockist_list:
            ct = col_totals.get(s["code"], {})
            gc = ws.cell(row=row, column=c_cur, value=_z(round(flt(ct.get("value", 0)) / LAKH, 2)))
            gc.font = lakh_cell_font; gc.fill = lakh_cell_fill
            gc.alignment = right_align; gc.border = thin_border
            c_cur += 1
        gtot = ws.cell(row=row, column=total_col, value=_z(round(flt(grand.get("value", 0)) / LAKH, 2)))
        gtot.font = lakh_cell_font; gtot.fill = lakh_cell_fill
        gtot.alignment = right_align; gtot.border = thin_border
        row += 1

    elif report_type == "secondary_sales":
        # Pivot: Products as rows, Stockists as columns (each with Before / After sub-cols).
        ws.title = "Secondary Sales"
        team_kw = kwargs.get("team", "")
        hq_kw = kwargs.get("hq", "")
        result = get_stockist_secondary_sales_report(division, region, from_date, to_date, team_kw or None, hq_kw or None)
        stockist_list = result.get("stockists", []) or []
        product_list = result.get("products", []) or []
        col_totals = result.get("col_totals", {}) or {}
        grand = result.get("grand", {}) or {}

        row = write_title_rows(ws, f"Stockist Wise Secondary Sales Report – {division}",
                               f"Region: {region_label}  |  Period: {period_label}")

        # Styles for bold+blue "value in lakhs" columns/row
        lakh_header_font  = Font(bold=True, size=11, color="FFFFFF")
        lakh_header_fill  = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        lakh_cell_font    = Font(bold=True, size=11, color="1D4ED8")
        lakh_cell_fill    = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
        lakh_cell_align   = Alignment(horizontal="right")

        def write_lakh_header(ws, r, c, value):
            cell = ws.cell(row=r, column=c, value=value)
            cell.font = lakh_header_font; cell.fill = lakh_header_fill
            cell.alignment = header_align; cell.border = thin_border

        def write_lakh_cell(ws, r, c, value):
            cell = ws.cell(row=r, column=c, value=value)
            cell.font = lakh_cell_font; cell.fill = lakh_cell_fill
            cell.alignment = lakh_cell_align; cell.border = thin_border

        # Two-row header: Row 1 = "Product | Pack | <Stockist1 spans 2> ... | Total (₹ Lakhs) spans 2"
        # Row 2 = "Before | After" sub-columns repeated under each stockist
        hdr_row1 = row
        hdr_row2 = row + 1
        ws.cell(row=hdr_row1, column=1, value="Product").font = header_font
        ws.cell(row=hdr_row1, column=1).fill = header_fill
        ws.cell(row=hdr_row1, column=1).alignment = header_align
        ws.cell(row=hdr_row1, column=1).border = thin_border
        ws.merge_cells(start_row=hdr_row1, start_column=1, end_row=hdr_row2, end_column=1)

        ws.cell(row=hdr_row1, column=2, value="Pack").font = header_font
        ws.cell(row=hdr_row1, column=2).fill = header_fill
        ws.cell(row=hdr_row1, column=2).alignment = header_align
        ws.cell(row=hdr_row1, column=2).border = thin_border
        ws.merge_cells(start_row=hdr_row1, start_column=2, end_row=hdr_row2, end_column=2)

        col_cursor = 3
        for s in stockist_list:
            cell = ws.cell(row=hdr_row1, column=col_cursor, value=s["name"])
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
            ws.merge_cells(start_row=hdr_row1, start_column=col_cursor,
                           end_row=hdr_row1, end_column=col_cursor + 1)
            ws.cell(row=hdr_row2, column=col_cursor, value="Before").font = header_font
            ws.cell(row=hdr_row2, column=col_cursor).fill = header_fill
            ws.cell(row=hdr_row2, column=col_cursor).alignment = header_align
            ws.cell(row=hdr_row2, column=col_cursor).border = thin_border
            ws.cell(row=hdr_row2, column=col_cursor + 1, value="After").font = header_font
            ws.cell(row=hdr_row2, column=col_cursor + 1).fill = header_fill
            ws.cell(row=hdr_row2, column=col_cursor + 1).alignment = header_align
            ws.cell(row=hdr_row2, column=col_cursor + 1).border = thin_border
            col_cursor += 2

        # Total column header — bold blue to mark as "value in lakhs" columns
        write_lakh_header(ws, hdr_row1, col_cursor, "Total (Rs. Lakhs)")
        ws.merge_cells(start_row=hdr_row1, start_column=col_cursor,
                       end_row=hdr_row1, end_column=col_cursor + 1)
        write_lakh_header(ws, hdr_row2, col_cursor,     "Before")
        write_lakh_header(ws, hdr_row2, col_cursor + 1, "After")

        total_col_start = col_cursor  # remember for per-product rows
        row = hdr_row2 + 1

        def _z(v):
            return v if v else None

        LAKH = 100000.0

        for p in product_list:
            vals = [p["code"], p["pack"]]
            for s in stockist_list:
                cell_val = (p["cells"] or {}).get(s["code"], {"before": 0, "after": 0})
                vals.append(_z(round(flt(cell_val.get("before")), 2)))
                vals.append(_z(round(flt(cell_val.get("after")), 2)))
            write_data_row(ws, row, vals)
            # Total columns in lakhs — bold + blue
            vb = round(flt(p.get("total_val_before", 0)) / LAKH, 2)
            va = round(flt(p.get("total_val_after",  0)) / LAKH, 2)
            write_lakh_cell(ws, row, total_col_start,     _z(vb))
            write_lakh_cell(ws, row, total_col_start + 1, _z(va))
            row += 1

        # Grand total row — value in lakhs per stockist + grand total columns
        # First two cells label
        for c, v in enumerate(["Value in Lakhs", ""], 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.font = lakh_cell_font; cell.fill = lakh_cell_fill
            cell.border = thin_border
        c_cur = 3
        for s in stockist_list:
            ct = col_totals.get(s["code"], {})
            vb = round(flt(ct.get("val_before", 0)) / LAKH, 2)
            va = round(flt(ct.get("val_after",  0)) / LAKH, 2)
            write_lakh_cell(ws, row, c_cur,     _z(vb))
            write_lakh_cell(ws, row, c_cur + 1, _z(va))
            c_cur += 2
        gvb = round(flt(grand.get("val_before", 0)) / LAKH, 2)
        gva = round(flt(grand.get("val_after",  0)) / LAKH, 2)
        write_lakh_cell(ws, row, c_cur,     _z(gvb))
        write_lakh_cell(ws, row, c_cur + 1, _z(gva))
        row += 1

    elif report_type == "moving_trend":
        ws.title = "Moving Trend"
        result = get_stockist_moving_trend_report(division, sales_type, stockist_code,
                                                   product_codes or None)
        data = result.get("data", [])
        fy_label = result.get("fy_label", "")
        stockist_name = result.get("stockist_name") or stockist_code
        hq_name = result.get("hq_name", "")
        vil = result.get("value_in_lakhs", {"months": [0] * 12, "total": 0})
        ml = ["Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan", "Feb", "Mar"]
        type_label = "Primary" if sales_type == "primary" else "Secondary"
        subtitle = f"Stockist: {stockist_name}"
        if hq_name:
            subtitle += f"  |  HQ: {hq_name}"
        if fy_label:
            subtitle += f"  |  {fy_label}"
        row = write_title_rows(ws, f"Moving Trend ({type_label} Sales) – {division}", subtitle)
        headers = ["Product Code", "Product Name", "Pack"] + ml + ["Total"]
        write_header_row(ws, row, headers)
        row += 1
        # Bold "Value in Lakhs" row at the top
        vil_months = vil.get("months", [0] * 12)
        vil_vals = ["Value in Lakhs", "", ""] + [v if v else None for v in vil_months] + [
            vil.get("total") or None
        ]
        write_value_row(ws, row, vil_vals)
        row += 1
        for d in data:
            vals = [d["product_code"], d["product_name"], d["pack"]] + d["months"] + [d["total"]]
            write_data_row(ws, row, vals)
            row += 1

    elif report_type == "closing_stock":
        ws.title = "Closing Stock"
        group_by_cs = kwargs.get("group_by", "stockist")
        result = get_stockist_closing_stock_report(division, region, from_date, to_date, group_by_cs)
        data = result.get("data", [])
        row = write_title_rows(ws, f"Stockist Wise Closing Stock Report – {division}",
                               f"Region: {region_label}  |  Period: {period_label}")
        name_label = "HQ Name" if group_by_cs == "hq" else "Stockist Name"
        headers = [name_label, "Product Code", "Product Name", "Pack",
                    "Opening Qty", "Purchase Qty", "Sales Qty", "Free Qty", "Scheme Free Qty", "Closing Qty"]
        write_header_row(ws, row, headers)
        row += 1
        current_key = None

        def _z(v):
            v = flt(v)
            return v if v else None

        for d in data:
            if group_by_cs == "hq":
                key = d.hq_code
                name = d.hq_name or d.hq_code or ""
            else:
                key = d.stockist_code
                name = d.stockist_name or ""
            if key != current_key:
                current_key = key
                write_group_row(ws, row, name, len(headers))
                row += 1
            write_data_row(ws, row, [name, d.product_code, d.product_name, d.pack,
                                     _z(d.opening_qty), _z(d.purchase_qty), _z(d.sales_qty),
                                     _z(d.free_qty), _z(d.scheme_free_qty), _z(d.closing_qty)])
            row += 1

    elif report_type == "hq_wise":
        ws.title = "HQ Wise Stockists"
        result = get_hq_wise_stockist_report(division, region)
        data = result.get("data", [])
        row = write_title_rows(ws, f"HQ Wise Stockist Report – {division}",
                               f"Region: {region_label}")
        headers = ["HQ Name", "Team", "Region", "No. of Stockists", "Stockists"]
        write_header_row(ws, row, headers)
        row += 1
        for hq_group in data:
            stockist_names = ", ".join(s.get("stockist_name", "") for s in hq_group.get("stockists", []))
            count = len(hq_group.get("stockists", []))
            write_data_row(ws, row, [hq_group["hq_name"], hq_group["team"], hq_group["region"],
                                     count if count else None, stockist_names])
            ws.cell(row=row, column=5).alignment = Alignment(wrap_text=True, vertical="top")
            row += 1

    elif report_type == "address":
        ws.title = "Stockist Address"
        result = get_stockist_address_report(division, region, criteria)
        data = result.get("data", [])
        groups = result.get("groups")
        row = write_title_rows(ws, f"Stockist Address Report – {division}",
                               f"Region: {region_label}  |  Criteria: {criteria}")
        headers = ["Stockist Name", "Address", "City", "Phone", "HQ", "Team", "Region"]
        write_header_row(ws, row, headers)
        row += 1

        def _row_vals(d):
            return [d.stockist_name or "", d.address or "", d.city or "", d.phone or "",
                    d.hq_name or "", d.team_name or "", d.region_name or ""]

        if groups:
            for grp_name, grp_rows in groups.items():
                write_group_row(ws, row, str(grp_name), len(headers))
                row += 1
                for d in grp_rows:
                    write_data_row(ws, row, _row_vals(d))
                    row += 1
        else:
            for d in data:
                write_data_row(ws, row, _row_vals(d))
                row += 1

    elif report_type in ("sec_moving_trend", "primary_moving_trend"):
        # Reports 7 (secondary) & 10 (primary) – shared Moving Trend output shape
        entity_type = kwargs.get("entity_type", "Team")
        entity_name = kwargs.get("entity_name", "")
        financial_year = kwargs.get("financial_year", "")
        sales_mode = kwargs.get("sales_mode", "after_deduction")
        if report_type == "primary_moving_trend":
            ws.title = "Primary Sales Moving Trend"
            result = get_primary_sales_moving_trend(division, entity_type, entity_name,
                                                     financial_year, product_codes or None)
            report_title_prefix = "Primary Sales Moving Trend"
            sales_mode_label = ""  # primary has no scheme deduction split
        else:
            ws.title = "Sec Sales Moving Trend"
            result = get_secondary_sales_moving_trend(division, entity_type, entity_name,
                                                      financial_year, sales_mode,
                                                      product_codes or None,
                                                      kwargs.get("product_group") or None,
                                                      kwargs.get("product_category") or None)
            report_title_prefix = "Secondary Sales Moving Trend"
            sales_mode_label = "Before Deduction" if sales_mode == "before_deduction" else "After Deduction"
        ml = ["Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan", "Feb", "Mar"]
        fy_label = result.get("fy_label", "")
        entity_display = result.get("entity_display", entity_name)
        sanctioned_strength = result.get("sanctioned_strength", 0)
        subtitle_parts = [f"{entity_type}: {entity_display}",
                          f"Sanctioned Strength: {sanctioned_strength}",
                          fy_label]
        if sales_mode_label:
            subtitle_parts.append(sales_mode_label)
        subtitle = "  |  ".join([p for p in subtitle_parts if p])
        row = write_title_rows(ws, f"{report_title_prefix} – {division}", subtitle)
        headers = ["Product", "Pack"] + ml + ["Target", "Total", "Average", "Per Capita", "%"]
        write_header_row(ws, row, headers)
        row += 1

        # Style helpers for special rows
        section_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        section_font = Font(bold=True, size=11, color="FFFFFF")
        lakhs_fill = PatternFill(start_color="DBE4F3", end_color="DBE4F3", fill_type="solid")
        lakhs_font = Font(bold=True, size=10)

        for sec in result.get("sections", []):
            # Section header row (merged across all columns)
            for c in range(1, len(headers) + 1):
                cell = ws.cell(row=row, column=c, value=sec.get("label", "") if c == 1 else None)
                cell.fill = section_fill
                cell.font = section_font
                cell.border = thin_border
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
            row += 1

            # Value in Lakhs row comes FIRST in each section
            vl = sec.get("value_in_lakhs", {})
            vl_months = vl.get("months", [0] * 12)

            def _bz(v):
                return v if v else None

            vl_vals = ["Value in Lakhs", ""] + [_bz(m) for m in vl_months] + [
                _bz(vl.get("target", 0)), _bz(vl.get("total", 0)),
                _bz(vl.get("average", 0)), _bz(vl.get("per_capita", 0)), _bz(vl.get("pct", 0))
            ]
            for c, v in enumerate(vl_vals, 1):
                cell = ws.cell(row=row, column=c, value=v)
                cell.fill = lakhs_fill
                cell.font = lakhs_font
                cell.border = thin_border
                if c > 2:
                    cell.alignment = Alignment(horizontal="right")
            row += 1

            # Product rows
            for p in sec.get("products", []):
                months = [_bz(v) for v in (p.get("months") or [0] * 12)]
                vals = [p.get("code", ""), p.get("pack", "")] + months + [
                    _bz(p.get("target", 0)), _bz(p.get("total", 0)),
                    _bz(p.get("average", 0)), _bz(p.get("per_capita", 0)), _bz(p.get("pct", 0))
                ]
                write_data_row(ws, row, vals)
                row += 1

    elif report_type == "region_stockist_trend":
        # Report 8 – Region-wise Stockist Secondary Sales Moving Trend
        financial_year = kwargs.get("financial_year", "")
        sales_mode = kwargs.get("sales_mode", "after_deduction")
        ws.title = "Region Stockist Trend"
        result = get_region_wise_stockist_moving_trend(division, region, financial_year, sales_mode,
                                                       product_codes or None,
                                                       kwargs.get("product_group") or None,
                                                       kwargs.get("product_category") or None)
        ml = result.get("month_labels") or ["Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan", "Feb", "Mar"]
        fy_label = result.get("fy_label", "")
        region_name = result.get("region_name", region)
        sales_mode_label = "Before Deduction" if sales_mode == "before_deduction" else "After Deduction"
        row = write_title_rows(ws, f"Region-wise Stockist Secondary Sales Moving Trend – {division}",
                               f"Region: {region_name}  |  {fy_label}  |  {sales_mode_label}  |  Values in ₹ Lakhs")
        headers = ["Customers", "HQ", "Current Month Opening"] + ml + ["Total Sales Value", "Current Month Closing"]
        write_header_row(ws, row, headers)
        row += 1

        team_fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
        team_font = Font(bold=True, size=11, color="FFFFFF")
        subtotal_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        subtotal_font = Font(bold=True, size=10)
        grand_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        grand_font = Font(bold=True, size=11, color="FFFFFF")

        def _bz(v):
            return v if v else None

        for team in result.get("teams", []):
            # Team header (merged)
            for c in range(1, len(headers) + 1):
                cell = ws.cell(row=row, column=c,
                               value=(team.get("team_name", team.get("team", "")) + " Team") if c == 1 else None)
                cell.fill = team_fill
                cell.font = team_font
                cell.border = thin_border
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
            row += 1

            for s in team.get("stockists", []):
                months = [_bz(v) for v in (s.get("months") or [0] * 12)]
                vals = [s.get("stockist_name", ""), s.get("hq", ""), _bz(s.get("opening", 0))] + \
                       months + [_bz(s.get("total_sales", 0)), _bz(s.get("closing", 0))]
                write_data_row(ws, row, vals)
                row += 1

            # Team subtotal
            tt = team.get("total", {})
            tt_months = [_bz(v) for v in (tt.get("months") or [0] * 12)]
            sub_vals = [f"{team.get('team_name', team.get('team', ''))} Team Total", "", _bz(tt.get("opening", 0))] + \
                       tt_months + [_bz(tt.get("total_sales", 0)), _bz(tt.get("closing", 0))]
            for c, v in enumerate(sub_vals, 1):
                cell = ws.cell(row=row, column=c, value=v)
                cell.fill = subtotal_fill
                cell.font = subtotal_font
                cell.border = thin_border
                if c > 2:
                    cell.alignment = Alignment(horizontal="right")
            row += 1
            row += 1  # spacer

        # Grand total
        grand = result.get("grand_total", {})
        grand_months = [_bz(v) for v in (grand.get("months") or [0] * 12)]
        grand_vals = [f"{region_name} Total", "", _bz(grand.get("opening", 0))] + \
                     grand_months + [_bz(grand.get("total_sales", 0)), _bz(grand.get("closing", 0))]
        for c, v in enumerate(grand_vals, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.fill = grand_fill
            cell.font = grand_font
            cell.border = thin_border
            if c > 2:
                cell.alignment = Alignment(horizontal="right")
        row += 1

    elif report_type == "region_product_stock":
        # Report 9 – Product Closing Stock Region Summary (pivot: products × HQs/Stockists)
        group_by = kwargs.get("group_by", "hq")
        from_date = kwargs.get("from_date", "")
        to_date = kwargs.get("to_date", "")
        ws.title = "Product Closing Stock"
        result = get_region_product_closing_stock(division, region, from_date, to_date, group_by)
        region_name = result.get("region_name", region)
        teams = result.get("team_order", [])
        products = result.get("products", [])
        vil = result.get("value_in_lakhs", {})
        is_org = result.get("is_org", False)
        # In the organization view Zone plays the "team" role and Region the "column"
        # role, so the group/grand-total headers switch labels accordingly.
        group_total_hdr = (result.get("group_label", "Team")) + " Total"
        grand_total_hdr = (result.get("total_label", "Region")) + " Total"
        if is_org:
            gb_label = "Region-wise (grouped by Zone)"
            scope_label = region_name
        else:
            gb_label = "Stockist-wise" if group_by == "stockist" else "HQ-wise"
            scope_label = f"Region: {region_name}"
        period_label = f"{from_date} to {to_date}" if from_date and to_date else "All Dates"

        row = write_title_rows(ws, f"Product Closing Stock – Region Summary ({division})",
                               f"{scope_label}  |  {gb_label}  |  Period: {period_label}")

        # Build flat column list
        flat_cols = []
        for team in teams:
            for hq in team.get("hqs", []):
                flat_cols.append({"col_code": hq["col_code"], "col_name": hq.get("col_name", hq["col_code"]),
                                  "team_code": team["team_code"],
                                  "team_name": team.get("team_name", team["team_code"])})

        # Row 1: Team group headers (merged)
        hdr_row1 = row
        row += 1
        hdr_row2 = row
        row += 1

        # Write Row 1 (team group spans)
        col_cursor = 4  # after S.No, Product, Pack
        for c_idx, lbl in enumerate(["S.No", "Product", "Pack"], 1):
            cell = ws.cell(row=hdr_row1, column=c_idx, value=lbl)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
            ws.merge_cells(start_row=hdr_row1, start_column=c_idx, end_row=hdr_row2, end_column=c_idx)

        team_fill2 = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
        team_font2 = Font(bold=True, size=10, color="FFFFFF")
        for team in teams:
            span = len(team.get("hqs", [])) + 1  # HQ cols + Team Total
            cell = ws.cell(row=hdr_row1, column=col_cursor,
                           value=team.get("team_name", team["team_code"]))
            cell.font = team_font2
            cell.fill = team_fill2
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
            if span > 1:
                ws.merge_cells(start_row=hdr_row1, start_column=col_cursor,
                               end_row=hdr_row1, end_column=col_cursor + span - 1)
            col_cursor += span

        # Grand Total header (Region Total, or Organization Total in the org view)
        rt_col = col_cursor
        cell = ws.cell(row=hdr_row1, column=rt_col, value=grand_total_hdr)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.merge_cells(start_row=hdr_row1, start_column=rt_col, end_row=hdr_row2, end_column=rt_col)

        # Row 2: individual HQ/stockist + Team Total headers
        col_cursor2 = 4
        for team in teams:
            for hq in team.get("hqs", []):
                cell = ws.cell(row=hdr_row2, column=col_cursor2,
                               value=hq.get("col_name", hq["col_code"]))
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
                cell.border = thin_border
                col_cursor2 += 1
            # Group Total column (Team Total, or Zone Total in the org view)
            cell = ws.cell(row=hdr_row2, column=col_cursor2, value=group_total_hdr)
            cell.font = Font(bold=True, size=10, color="FFFFFF")
            cell.fill = PatternFill(start_color="475569", end_color="475569", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
            col_cursor2 += 1

        # Value in Lakhs row
        lakhs_fill2 = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
        lakhs_font2 = Font(bold=True, size=10, color="1E40AF")
        vl_col_vals = vil.get("col_values", {})
        vl_team_vals = vil.get("team_values", {})
        vl_region = vil.get("region_total", 0)

        vil_vals = ["Value In Lakhs", "", ""]
        for team in teams:
            for hq in team.get("hqs", []):
                v = vl_col_vals.get(hq["col_code"])
                vil_vals.append(round(v, 2) if v else "")
            tv = vl_team_vals.get(team["team_code"])
            vil_vals.append(round(tv, 2) if tv else "0.00")
        vil_vals.append(round(vl_region, 2) if vl_region else "0.00")

        for c, v in enumerate(vil_vals, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.fill = lakhs_fill2
            cell.font = lakhs_font2
            cell.border = thin_border
            if c > 3:
                cell.alignment = Alignment(horizontal="right")
                # Plain 2-decimal format, no thousands separators
                cell.number_format = "0.00"
        row += 1

        # Product rows — quantities kept as real numbers but displayed without any
        # thousands separators (client wants no commas), up to 2 decimals.
        for p in products:
            p_vals = [p.get("sno", ""), p.get("product_code", ""), p.get("pack", "")]
            for team in teams:
                for hq in team.get("hqs", []):
                    qty = (p.get("col_qty") or {}).get(hq["col_code"])
                    p_vals.append(qty if qty else "")
                tqty = (p.get("team_qty") or {}).get(team["team_code"])
                p_vals.append(tqty if tqty else "")
            rqty = p.get("region_qty")
            p_vals.append(rqty if rqty else "")
            write_data_row(ws, row, p_vals)
            for c in range(4, len(p_vals) + 1):
                ws.cell(row=row, column=c).number_format = "0.##"
            row += 1

    elif report_type == "sec_vs_closing":
        # Report 11 – Secondary Sales Value vs Closing Stock Value
        from_month = kwargs.get("from_month", "")
        to_month = kwargs.get("to_month", "")
        sales_mode = kwargs.get("sales_mode", "after_deduction")
        region_codes = kwargs.get("region_codes", "")
        ws.title = "Sec Sales vs Closing Value"
        result = get_secondary_vs_closing_value_report(division, from_month or None,
                                                        to_month or None, region_codes or None,
                                                        sales_mode,
                                                        product_codes or None,
                                                        kwargs.get("product_group") or None,
                                                        kwargs.get("product_category") or None)
        months_seq = result.get("months", [])
        regions_out = result.get("regions", [])
        grand = result.get("grand_total", {}).get("monthly", [])

        period_label = ""
        if months_seq:
            period_label = months_seq[0]["label"] + " to " + months_seq[-1]["label"]
        sales_mode_label = "Before Deduction" if sales_mode == "before_deduction" else "After Deduction"
        subtitle_parts = ["Organization Secondary Sales & Closing Stock Value", period_label, sales_mode_label]
        subtitle = "  |  ".join([p for p in subtitle_parts if p])
        row = write_title_rows(ws, f"Secondary Sales vs Closing Value – {division}", subtitle)

        # Two-row header
        hdr_row1 = row
        hdr_row2 = row + 1
        # First col: "HQ" (merged vertically across both header rows — set the
        # anchor cell only; the merged cell below it is read-only in openpyxl)
        cell = ws.cell(row=hdr_row1, column=1, value="HQ")
        cell.font = header_font; cell.fill = header_fill; cell.alignment = header_align; cell.border = thin_border
        ws.merge_cells(start_row=hdr_row1, start_column=1, end_row=hdr_row2, end_column=1)
        # Per-month header pairs
        col_cursor = 2
        for ms in months_seq:
            cell = ws.cell(row=hdr_row1, column=col_cursor, value=ms["label"])
            cell.font = header_font; cell.fill = header_fill; cell.alignment = header_align; cell.border = thin_border
            ws.merge_cells(start_row=hdr_row1, start_column=col_cursor,
                           end_row=hdr_row1, end_column=col_cursor + 1)
            for sub, lbl in enumerate(["Sec Val", "Cls Val"]):
                c = ws.cell(row=hdr_row2, column=col_cursor + sub, value=lbl)
                c.font = header_font; c.fill = header_fill; c.alignment = header_align; c.border = thin_border
            col_cursor += 2

        row = hdr_row2 + 1

        # Region groups
        region_total_font = Font(bold=True, color="C00000", size=11)
        region_total_fill = PatternFill(start_color="FFF2F2", end_color="FFF2F2", fill_type="solid")

        def _z(v):
            return v if v else None

        for reg in regions_out:
            for h in reg.get("hqs", []):
                vals = [h["hq_name"]]
                for m in h["monthly"]:
                    vals.append(_z(m.get("sec_val")))
                    vals.append(_z(m.get("cls_val")))
                write_data_row(ws, row, vals)
                row += 1

            # Region total row (bold red)
            tvals = [f"{reg['region_name']} Region Total"]
            for m in reg["totals"]["monthly"]:
                tvals.append(_z(m.get("sec_val")))
                tvals.append(_z(m.get("cls_val")))
            for c, v in enumerate(tvals, 1):
                cell = ws.cell(row=row, column=c, value=v)
                cell.font = region_total_font
                cell.fill = region_total_fill
                cell.border = thin_border
                if c > 1:
                    cell.alignment = Alignment(horizontal="right")
            row += 1

        # Grand total
        if grand:
            gvals = ["Grand Total"]
            for m in grand:
                gvals.append(_z(m.get("sec_val")))
                gvals.append(_z(m.get("cls_val")))
            grand_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
            grand_font = Font(bold=True, color="FFFFFF", size=11)
            for c, v in enumerate(gvals, 1):
                cell = ws.cell(row=row, column=c, value=v)
                cell.font = grand_font
                cell.fill = grand_fill
                cell.border = thin_border
                if c > 1:
                    cell.alignment = Alignment(horizontal="right")
            row += 1

    elif report_type == "monthly_org":
        # Report 13 – Full Monthly Organizational Report (product-level line items)
        month = kwargs.get("month", "")
        result = get_monthly_organizational_report(
            division, month or None,
            kwargs.get("team") or None, kwargs.get("hq") or None,
            kwargs.get("product_codes") or None, kwargs.get("product_group") or None,
            kwargs.get("product_category") or None)
        data_rows = result.get("rows", [])
        grand = result.get("grand_total", {})
        month_label = result.get("month_label", month)

        row = write_title_rows(
            ws, f"Full Monthly Organizational Report – {division}",
            f"Month: {month_label}  |  Qty in boxes  |  PTS Value = (Sales+Free) × Master PTS  |  "
            f"NRV Value = (Sales+Free−Scheme) × NRV  |  values in Rs. (total row in Rs. Lakhs)")
        headers = ["Stockist Code", "Stockist Name", "City/Pool", "Team", "Region",
                   "Product Code", "Product", "Pack",
                   "Sales (Before Scheme)", "Sales (After Scheme)", "Cl. Stock", "Op. Stock",
                   "PTS", "PTS Value", "NRV", "NRV Value",
                   "Cls Value", "Ops Value",
                   "Product Head", "Product Category"]
        write_header_row(ws, row, headers)
        row += 1

        LAKH = 100000.0

        def _L(v):
            return round(flt(v) / LAKH, 2) if v else None  # lakhs (total row)

        def _R(v):
            return round(flt(v), 2) if v else None          # rupees (data rows)

        def _Q(v):
            return round(flt(v), 2) if v else None            # box quantity

        # Overall total pinned at the TOP — qty in boxes, values in Rs. Lakhs.
        # PTS/NRV are per-line rates (cols 13, 15) → no meaningful total, left blank.
        write_value_row(ws, row, [
            "OVERALL TOTAL (Rs. Lakhs)", "", "", "", "", "", "", "",
            _Q(grand.get("sales_before")), _Q(grand.get("sales_after")),
            _Q(grand.get("clstock")), _Q(grand.get("opstock")),
            None, _L(grand.get("ptsvalue")), None, _L(grand.get("nrvvalue")),
            _L(grand.get("clsvalue")), _L(grand.get("opsvalue")), "", ""])
        row += 1

        right_align = Alignment(horizontal="right")
        for r in data_rows:
            write_data_row(ws, row, [
                r.get("stockist_code"), r.get("stockist_name"), r.get("citypool"),
                r.get("team"), r.get("region"),
                r.get("product_code"), r.get("product_name"), r.get("pack"),
                _Q(r.get("sales_before")), _Q(r.get("sales_after")),
                _Q(r.get("clstock")), _Q(r.get("opstock")),
                _R(r.get("pts")), _R(r.get("ptsvalue")), _R(r.get("nrv")), _R(r.get("nrvvalue")),
                _R(r.get("clsvalue")), _R(r.get("opsvalue")),
                r.get("product_head"), r.get("product_category")])
            for c in range(9, 19):  # numeric columns 9..18 right-aligned
                ws.cell(row=row, column=c).alignment = right_align
            row += 1

    elif report_type == "target_report":
        # Report 14 – Target vs Sales (HQ-wise timeline); month pairs = (Target, Sales)
        from_month = kwargs.get("from_month", "")
        to_month = kwargs.get("to_month", "")
        sales_type = kwargs.get("sales_type", "secondary")
        sales_mode = kwargs.get("sales_mode", "after_deduction")
        region_codes = kwargs.get("region_codes", "")
        ws.title = "Target vs Sales"
        result = get_target_vs_sales_report(division, from_month or None, to_month or None,
                                            region_codes or None, sales_type, sales_mode)
        months_seq = result.get("months", [])
        regions_out = result.get("regions", [])
        grand = result.get("grand_total", {}).get("monthly", [])

        period_label = ""
        if months_seq:
            period_label = months_seq[0]["label"] + " to " + months_seq[-1]["label"]
        type_label = "Primary" if sales_type == "primary" else "Secondary"
        if sales_type != "primary":
            type_label += " (" + ("Before Deduction" if sales_mode == "before_deduction" else "After Deduction") + ")"
        subtitle = "  |  ".join([p for p in ["Target vs Sales", period_label, type_label, "Values in Rs. Lakhs"] if p])
        row = write_title_rows(ws, f"Target vs Sales – {division}", subtitle)

        # Two-row header: HQ + per-month (Target, Sales)
        hdr_row1 = row
        hdr_row2 = row + 1
        cell = ws.cell(row=hdr_row1, column=1, value="HQ")
        cell.font = header_font; cell.fill = header_fill; cell.alignment = header_align; cell.border = thin_border
        ws.merge_cells(start_row=hdr_row1, start_column=1, end_row=hdr_row2, end_column=1)
        col_cursor = 2
        for ms in months_seq:
            cell = ws.cell(row=hdr_row1, column=col_cursor, value=ms["label"])
            cell.font = header_font; cell.fill = header_fill; cell.alignment = header_align; cell.border = thin_border
            ws.merge_cells(start_row=hdr_row1, start_column=col_cursor,
                           end_row=hdr_row1, end_column=col_cursor + 1)
            for sub, lbl in enumerate(["Target", "Sales"]):
                c = ws.cell(row=hdr_row2, column=col_cursor + sub, value=lbl)
                c.font = header_font; c.fill = header_fill; c.alignment = header_align; c.border = thin_border
            col_cursor += 2
        row = hdr_row2 + 1

        region_total_font = Font(bold=True, color="C00000", size=11)
        region_total_fill = PatternFill(start_color="FFF2F2", end_color="FFF2F2", fill_type="solid")

        def _z(v):
            return v if v else None

        for reg in regions_out:
            for h in reg.get("hqs", []):
                vals = [h["hq_name"]]
                for m in h["monthly"]:
                    vals.append(_z(m.get("target")))
                    vals.append(_z(m.get("sales")))
                write_data_row(ws, row, vals)
                row += 1
            tvals = [f"{reg['region_name']} Region Total"]
            for m in reg["totals"]["monthly"]:
                tvals.append(_z(m.get("target")))
                tvals.append(_z(m.get("sales")))
            for c, v in enumerate(tvals, 1):
                cell = ws.cell(row=row, column=c, value=v)
                cell.font = region_total_font
                cell.fill = region_total_fill
                cell.border = thin_border
                if c > 1:
                    cell.alignment = Alignment(horizontal="right")
            row += 1

        if grand:
            gvals = ["Grand Total"]
            for m in grand:
                gvals.append(_z(m.get("target")))
                gvals.append(_z(m.get("sales")))
            grand_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
            grand_font = Font(bold=True, color="FFFFFF", size=11)
            for c, v in enumerate(gvals, 1):
                cell = ws.cell(row=row, column=c, value=v)
                cell.font = grand_font
                cell.fill = grand_fill
                cell.border = thin_border
                if c > 1:
                    cell.alignment = Alignment(horizontal="right")
            row += 1

    elif report_type == "gynae_report":
        # Report 15 – Gynae Report (brands × months, with strength + summary rows)
        ws.title = "Gynae Report"
        result = get_gynae_report(division, kwargs.get("entity_type") or "Organization",
                                  kwargs.get("entity_name") or None,
                                  kwargs.get("financial_year") or None,
                                  kwargs.get("sales_mode") or "after_deduction")
        brands = result.get("brands", [])
        ml = result.get("month_labels") or _MONTH_LABELS
        strength = result.get("strength", 0)
        mode_label = "Before Deduction" if (kwargs.get("sales_mode") == "before_deduction") else "After Deduction"
        row = write_title_rows(
            ws, f"Gynae Report – {division}",
            f"{result.get('entity_display','')}  |  Strength: {strength}  |  {result.get('fy_label','')}  |  {mode_label}")

        headers = ["Gynae Brands"] + list(ml) + ["Total"]
        write_header_row(ws, row, headers)
        row += 1

        right_align = Alignment(horizontal="right")
        for b in brands:
            vals = [f"{b['name']} ({b['pack']})" if b.get("pack") else b["name"]] \
                   + [v if v else None for v in b["months"]] + [b["total"] if b["total"] else None]
            write_data_row(ws, row, vals)
            for c in range(2, len(headers) + 1):
                ws.cell(row=row, column=c).alignment = right_align
            row += 1

        # Summary rows (red, bold) — Values in lakh / PCPM in lakh / Avg per Dr (Rs.)
        red_font = Font(bold=True, color="C00000", size=11)

        def _summary(label, series, as_total=True):
            cells = [label] + [v if v else None for v in series]
            cells += [round(sum(series), 2) if as_total else None]
            for c, v in enumerate(cells, 1):
                cell = ws.cell(row=row_ref[0], column=c, value=v)
                cell.font = red_font
                cell.border = thin_border
                if c > 1:
                    cell.alignment = right_align
            row_ref[0] += 1

        row_ref = [row]
        _summary("Values in lakh", result.get("values_lakh", []))
        _summary("PCPM in lakh", result.get("pcpm_lakh", []), as_total=False)
        _summary("Avg per Dr. (Rs.)", result.get("avg_per_dr", []), as_total=False)
        row = row_ref[0]

    elif report_type == "org_sales":
        # Report 16 – Organisational Sales Report (Region wise): products × regions
        # grouped by zone, with per-zone total + organization total columns.
        ws.title = "Organisational Sales"
        result = get_organizational_sales_report(
            division,
            kwargs.get("sales_type") or "primary",
            kwargs.get("from_date") or None,
            kwargs.get("to_date") or None,
            product_codes or None,
            kwargs.get("sales_mode") or "after_deduction",
        )
        zones = result.get("zones", [])
        prods = result.get("products", [])
        vil = result.get("value_in_lakhs", {}) or {}
        type_label = result.get("type_label", "Primary")
        period = result.get("period_label", "")
        row = write_title_rows(
            ws, f"Organization {type_label} Sales – {division}",
            (f"From {period}" if period else ""))

        # Flat column scaffold: ("region"|"zone"|"org", code, label)
        flat_cols = []
        for z in zones:
            for rg in z.get("regions", []):
                flat_cols.append(("region", rg["code"], rg["name"]))
            flat_cols.append(("zone", z["code"], f"{z['name']} Total"))
        flat_cols.append(("org", "", "Organization Total"))

        headers = ["#", "Code", "Pack"] + [c[2] for c in flat_cols]
        write_header_row(ws, row, headers)
        row += 1

        right_align = Alignment(horizontal="right")
        center_align = Alignment(horizontal="center")

        def _z(v):
            return v if v else None

        def _zr(v):
            # Box quantities are shown rounded to whole numbers (client request).
            return round(flt(v)) if v else None

        vil_reg = vil.get("regions", {}) or {}
        vil_zone = vil.get("zone_totals", {}) or {}

        # Value In Lakhs row (first)
        vrow = ["", "Value In Lakhs", ""]
        for kind, code, _lbl in flat_cols:
            if kind == "region":
                vrow.append(_z(vil_reg.get(code)))
            elif kind == "zone":
                vrow.append(_z(vil_zone.get(code)))
            else:
                vrow.append(_z(vil.get("org_total")))
        write_value_row(ws, row, vrow)
        row += 1

        for i, p in enumerate(prods, 1):
            cells = p.get("cells", {}) or {}
            ztot = p.get("zone_totals", {}) or {}
            vals = [i, p.get("code", ""), p.get("pack", "")]
            for kind, code, _lbl in flat_cols:
                if kind == "region":
                    vals.append(_zr(cells.get(code)))
                elif kind == "zone":
                    vals.append(_zr(ztot.get(code)))
                else:
                    vals.append(_zr(p.get("org_total")))
            write_data_row(ws, row, vals)
            ws.cell(row=row, column=1).alignment = center_align
            ws.cell(row=row, column=2).alignment = center_align
            ws.cell(row=row, column=3).alignment = center_align
            for c in range(4, len(headers) + 1):
                ws.cell(row=row, column=c).alignment = right_align
            row += 1

    else:
        frappe.throw("Invalid report type")

    # Auto-fit column widths
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    # Save to temp file and respond
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    xlsx_data = output.getvalue()

    _report_titles = {
        "primary_sales": "Stockist Wise Primary Sales Report",
        "secondary_sales": "Stockist Wise Secondary Sales Report",
        "moving_trend": "Moving Trend Stockist Wise",
        "closing_stock": "Closing Stock Report",
        "hq_wise": "HQ Wise Stockist Report",
        "address": "Stockist Address Report",
        "sec_moving_trend": "Secondary Sales Moving Trend",
        "region_stockist_trend": "Region wise Stockist Secondary Sales Moving Trend",
        "region_product_stock": "Product Closing Stock Region Summary",
        "primary_moving_trend": "Primary Sales Moving Trend",
        "sec_vs_closing": "Secondary Sales vs Closing Stock Value",
        "monthly_org": "Full Monthly Organizational Report",
        "target_report": "Target vs Sales Report",
        "gynae_report": "Gynae Report",
        "org_sales": "Organisational Sales Report Region wise",
    }
    import re as _re
    from datetime import date as _date
    _title = _report_titles.get(report_type, report_type)
    _today = _date.today().strftime("%Y-%m-%d")
    _safe = _re.sub(r'[^\w\s]', ' ', _title).strip()
    _safe = _re.sub(r'\s+', '_', _safe)
    filename = f"{_safe}_{division}_{_today}.xlsx"
    frappe.local.response.filename = filename
    frappe.local.response.filecontent = xlsx_data
    frappe.local.response.type = "download"
    frappe.local.response.content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


# ═══════════════════════════════════════════════════════════════
# SCHEME REPORTS  –  Portal API Methods
# ═══════════════════════════════════════════════════════════════

def get_scheme_report_filter_options(division=None):
    """Return dropdown options for the Scheme Reports portal page (active masters
    only). Mirrors get_stockist_report_filter_options so the page can offer the
    same Reporting-Criteria + cascading Zone→Region→Team→HQ filters, plus Doctor,
    Products / Product Group / Product Category multi-selects. Codes are carried as
    option values; only names are shown."""
    if not division:
        division = get_user_division()

    # None for admin (= all regions); a list confines a non-admin's pickers below.
    allowed = _allowed_region_codes_or_all(division)

    zones = frappe.db.sql(
        "SELECT name, zone_name FROM `tabZone Master` WHERE status='Active' AND division IN (%s, 'Both') ORDER BY name",
        (division,), as_dict=True,
    )
    regions = frappe.db.sql(
        "SELECT name, region_name, IFNULL(zone,'') AS zone FROM `tabRegion Master` WHERE status='Active' AND division IN (%s, 'Both') ORDER BY name",
        (division,), as_dict=True,
    )
    teams = frappe.db.sql(
        """SELECT t.name, t.team_name, IFNULL(t.region,'') AS region, IFNULL(r.zone,'') AS zone
             FROM `tabTeam Master` t
             LEFT JOIN `tabRegion Master` r ON r.name = t.region
            WHERE t.status='Active' AND t.division IN (%s, 'Both') ORDER BY t.name""",
        (division,), as_dict=True,
    )
    hqs = frappe.db.sql(
        "SELECT name, hq_name, IFNULL(team,'') AS team, IFNULL(region,'') AS region, IFNULL(zone,'') AS zone "
        "FROM `tabHQ Master` WHERE division=%s AND status='Active' ORDER BY name",
        (division,), as_dict=True,
    )
    stockists = frappe.db.sql(
        "SELECT name, stockist_name, IFNULL(region, '') AS region FROM `tabStockist Master` WHERE division=%s AND status='Active' ORDER BY stockist_name",
        (division,), as_dict=True,
    )
    doctors = frappe.db.sql(
        "SELECT name, doctor_code, doctor_name, IFNULL(hq,'') AS hq, IFNULL(team,'') AS team, "
        "IFNULL(region,'') AS region, IFNULL(zone,'') AS zone "
        "FROM `tabDoctor Master` WHERE division=%s AND status='Active' ORDER BY doctor_name",
        (division,), as_dict=True,
    )
    products = frappe.db.sql(
        "SELECT product_code, product_name, category, product_group, pack, sequence "
        "FROM `tabProduct Master` WHERE division=%s AND status='Active' ORDER BY COALESCE(sequence, 9999), product_name",
        (division,), as_dict=True,
    )
    product_groups = sorted(set(p.product_group for p in products if p.product_group))
    product_categories = sorted(set(p.category for p in products if p.category))

    # Confine a non-admin's pickers to their mapped regions (and the teams/HQs/doctors/
    # stockists inside them). Admin (allowed is None) sees everything.
    if allowed is not None:
        aset = set(allowed)
        regions = [r for r in regions if r.name in aset]
        teams = [t for t in teams if (t.region or "") in aset]
        hqs = [h for h in hqs if (h.region or "") in aset]
        doctors = [d for d in doctors if (d.region or "") in aset]
        stockists = [s for s in stockists if (s.region or "") in aset]

    return {
        "zones": [{"code": z.name, "name": z.zone_name or z.name} for z in zones],
        "regions": [{"code": r.name, "name": r.region_name or r.name, "zone": r.zone} for r in regions],
        "teams": [{"code": t.name, "name": t.team_name or t.name, "region": t.region, "zone": t.zone} for t in teams],
        "hqs": [{"code": h.name, "name": h.hq_name or h.name, "team": h.team, "region": h.region, "zone": h.zone} for h in hqs],
        "stockists": [{"code": s.name, "name": s.stockist_name or s.name} for s in stockists],
        "doctors": [{"code": d.name, "name": d.doctor_name or d.name, "hq": d.hq,
                      "team": d.team, "region": d.region, "zone": d.zone} for d in doctors],
        "products": [{"code": p.product_code, "name": p.product_name,
                       "category": p.category or "", "group": p.product_group or "",
                       "pack": p.pack or ""} for p in products],
        "product_groups": product_groups,
        "product_categories": product_categories,
    }


def _scheme_geo_conditions(alias, params, zone=None, region=None, team=None, hq=None, doctor=None, division=None):
    """Build WHERE fragments for the shared cascade filters against a Doctor Master
    alias (dm). Each dimension is optional; the cascade sends only what is selected.
    Returns a list of SQL condition strings and mutates `params` in place. `division`
    enables per-user region scoping: a non-admin is confined to their allowed regions
    here (admin unrestricted); a picked region is validated against that set."""
    conds = []
    if zone:
        conds.append(f"{alias}.zone = %(g_zone)s"); params["g_zone"] = zone
    _scope_region_sql(conds, params, f"{alias}.region", division, region, key="g_region")
    if team:
        conds.append(f"{alias}.team = %(g_team)s"); params["g_team"] = team
    if hq:
        conds.append(f"{alias}.hq = %(g_hq)s"); params["g_hq"] = hq
    if doctor:
        conds.append(f"{alias}.name = %(g_doctor)s"); params["g_doctor"] = doctor
    return conds


@frappe.whitelist()
def get_scheme_activity_trend_report(division=None, from_date=None, to_date=None,
                                      doctor_status="Active", reporting_criteria="Organization",
                                      zone=None, region=None, team=None, hq=None, doctor=None,
                                      product_codes=None, product_group=None, product_category=None):
    """Report 1 – Activity Trend Report.
    Monthly pivot (Apr–Mar) showing product-freeqty pairs per doctor.
    """
    if not division:
        division = get_user_division()
    if not from_date or not to_date:
        return {"success": False, "message": "From Date and To Date are required"}

    conditions = [
        "sr.division = %(division)s",
        "sr.docstatus = 1",
        "sr.approval_status = 'Approved'",
        "sr.application_date BETWEEN %(from_date)s AND %(to_date)s",
    ]
    params = {"division": division, "from_date": from_date, "to_date": to_date}

    # Doctor status filter
    if doctor_status and doctor_status != "All":
        conditions.append("dm.status = %(doctor_status)s")
        params["doctor_status"] = doctor_status

    # Product filter (codes / group / category) — resolves to Product Master ids.
    resolved_pks = _resolve_product_filter(division, product_codes=product_codes,
                                           product_group=product_group, product_category=product_category)
    if resolved_pks is not None:
        placeholders = ", ".join(["%(pc_" + str(i) + ")s" for i in range(len(resolved_pks))])
        conditions.append("sri.product_code IN (" + placeholders + ")")
        for i, pc in enumerate(resolved_pks):
            params["pc_" + str(i)] = pc

    # Shared cascade scope (Zone→Region→Team→HQ→Doctor)
    conditions += _scheme_geo_conditions("dm", params, zone, region, team, hq, doctor, division=division)

    where = " AND ".join(conditions)

    rows = frappe.db.sql(f"""
        SELECT dm.region, hm.hq_name, dm.hq, dm.doctor_name, dm.name AS doctor_code,
               COALESCE(pm.product_code, sri.product_code) AS product_code, sri.free_quantity,
               MONTH(sr.application_date) AS m, YEAR(sr.application_date) AS y
        FROM `tabScheme Request` sr
        INNER JOIN `tabScheme Request Item` sri ON sri.parent = sr.name
        INNER JOIN `tabDoctor Master` dm ON dm.name = sr.doctor_code
        INNER JOIN `tabHQ Master` hm ON hm.name = dm.hq
        LEFT JOIN `tabProduct Master` pm ON pm.name = sri.product_code
        WHERE {where}
        ORDER BY dm.region, hm.hq_name, dm.doctor_name, sr.application_date
    """, params, as_dict=True)

    # Build FY label from from_date
    from datetime import datetime
    fd = datetime.strptime(str(from_date), "%Y-%m-%d")
    td = datetime.strptime(str(to_date), "%Y-%m-%d")
    fy_label = f"From {fd.strftime('%B %Y')} To {td.strftime('%B %Y')}"

    # Month index: Apr=0..Mar=11
    month_idx = {4: 0, 5: 1, 6: 2, 7: 3, 8: 4, 9: 5, 10: 6, 11: 7, 12: 8, 1: 9, 2: 10, 3: 11}

    # Aggregate: key=(region, hq, doctor_name) → months[12] = list of "PROD-QTY" strings
    from collections import defaultdict
    doc_data = defaultdict(lambda: [[] for _ in range(12)])

    for r in rows:
        key = (r.region or "", r.hq or "", r.hq_name or "", r.doctor_name or "", r.doctor_code or "")
        idx = month_idx.get(r.m, 0)
        doc_data[key][idx].append(f"{r.product_code}-{int(r.free_quantity or 0)}")

    data = []
    for (reg, hq, hq_name, doctor_name, doctor_code), months in doc_data.items():
        cells = ["\n".join(sorted(set(m))) if m else "" for m in months]
        data.append({
            "region": reg, "hq": hq, "hq_name": hq_name,
            "doctor_name": doctor_name, "doctor_code": doctor_code,
            "months": cells,
        })

    # Sort by region, hq, doctor
    data.sort(key=lambda x: (x["region"], x["hq"], x["doctor_name"]))

    return {"success": True, "data": data, "fy_label": fy_label, "criteria": reporting_criteria}


@frappe.whitelist()
def get_scheme_activity_track_report(division=None, from_date=None, to_date=None,
                                      doctor_status="Active", reporting_criteria="Organization",
                                      zone=None, region=None, team=None, hq=None, doctor=None,
                                      product_codes=None, product_group=None, product_category=None):
    """Report 2 – Activity Track Report.
    Transaction‐level rows: one per Scheme Request Item with full details incl.
    special price and discount value.
    """
    if not division:
        division = get_user_division()
    if not from_date or not to_date:
        return {"success": False, "message": "From Date and To Date are required"}

    conditions = [
        "sr.division = %(division)s",
        "sr.docstatus = 1",
        "sr.approval_status = 'Approved'",
        "sr.application_date BETWEEN %(from_date)s AND %(to_date)s",
    ]
    params = {"division": division, "from_date": from_date, "to_date": to_date}

    if doctor_status and doctor_status != "All":
        conditions.append("dm.status = %(doctor_status)s")
        params["doctor_status"] = doctor_status

    resolved_pks = _resolve_product_filter(division, product_codes=product_codes,
                                           product_group=product_group, product_category=product_category)
    if resolved_pks is not None:
        placeholders = ", ".join(["%(pc_" + str(i) + ")s" for i in range(len(resolved_pks))])
        conditions.append("sri.product_code IN (" + placeholders + ")")
        for i, pc in enumerate(resolved_pks):
            params["pc_" + str(i)] = pc

    conditions += _scheme_geo_conditions("dm", params, zone, region, team, hq, doctor, division=division)

    where = " AND ".join(conditions)

    rows = frappe.db.sql(f"""
        SELECT sr.application_date AS date, dm.region, dm.hq,
               dm.doctor_name, COALESCE(pm.product_code, sri.product_code) AS product_code,
               COALESCE(pm.product_name, sri.product_name) AS product_name,
               sri.quantity, sri.free_quantity, sri.pack, sri.product_rate AS rate,
               sri.special_rate, sri.product_value AS value,
               sr.stockist_code, COALESCE(sm.stockist_name, '') AS stockist_name,
               dm.team
        FROM `tabScheme Request` sr
        INNER JOIN `tabScheme Request Item` sri ON sri.parent = sr.name
        INNER JOIN `tabDoctor Master` dm ON dm.name = sr.doctor_code
        LEFT JOIN `tabProduct Master` pm ON pm.name = sri.product_code
        LEFT JOIN `tabStockist Master` sm ON sm.name = sr.stockist_code
        WHERE {where}
        ORDER BY dm.team, dm.hq, sr.application_date, dm.doctor_name
    """, params, as_dict=True)

    # Add serial numbers and compute totals (incl. discount value)
    data = []
    total_qty = total_free = total_value = total_discount = 0
    for i, r in enumerate(rows, 1):
        special_rate = flt(r.special_rate)
        # Discount value = (order qty ÷ strips-per-box) × (PTS − special price), only
        # when a special price is set. The ÷conv is essential: order value (product_value)
        # is per-box, so without it the discount is inflated by the pack size for NxM packs.
        _conv = _scheme_pack_conversion(r.pack)
        discount_value = (flt(r.quantity) / _conv) * (flt(r.rate) - special_rate) if special_rate > 0 else 0
        data.append({
            "sno": i,
            "date": str(r.date) if r.date else "",
            "region": r.region or "",
            "hq": r.hq or "",
            "doctor_name": r.doctor_name or "",
            "product_code": r.product_code or "",
            "product_name": r.product_name or "",
            "qty": flt(r.quantity),
            "free_qty": flt(r.free_quantity),
            "rate": flt(r.rate),
            "special_rate": special_rate,
            "discount_value": discount_value,
            "value": flt(r.value),
            "stockist_name": r.stockist_name or "",
            "team": r.team or "",
        })
        total_qty += flt(r.quantity)
        total_free += flt(r.free_quantity)
        total_value += flt(r.value)
        total_discount += discount_value

    return {
        "success": True, "data": data,
        "totals": {"qty": total_qty, "free_qty": total_free,
                   "discount_value": total_discount, "value": total_value},
        "criteria": reporting_criteria,
    }


@frappe.whitelist()
def get_new_approval_doctors_report(division=None, from_date=None, to_date=None,
                                     reporting_criteria="Organization",
                                     zone=None, region=None, team=None, hq=None, doctor=None,
                                     product_codes=None, product_group=None, product_category=None):
    """Report 3 – New Approval Doctors.
    Doctors whose first-ever approved scheme falls within the given date range.
    """
    if not division:
        division = get_user_division()
    if not from_date or not to_date:
        return {"success": False, "message": "From Date and To Date are required"}

    params = {"division": division, "from_date": from_date, "to_date": to_date}

    # Subquery: first ever approved date per doctor in this division
    # Then filter doctors whose first date falls in the range
    hierarchy_conds = ["dm.status = 'Active'"]
    hierarchy_conds += _scheme_geo_conditions("dm", params, zone, region, team, hq, doctor, division=division)

    # Product filter on scheme items (optional) — resolve codes/group/category to ids
    product_join = ""
    product_cond = ""
    resolved_pks = _resolve_product_filter(division, product_codes=product_codes,
                                           product_group=product_group, product_category=product_category)
    if resolved_pks is not None:
        placeholders = ", ".join(["%(pc_" + str(i) + ")s" for i in range(len(resolved_pks))])
        product_join = "INNER JOIN `tabScheme Request Item` sri ON sri.parent = sr.name"
        product_cond = "AND sri.product_code IN (" + placeholders + ")"
        for i, pc in enumerate(resolved_pks):
            params["pc_" + str(i)] = pc

    hierarchy_where = " AND ".join(hierarchy_conds)

    rows = frappe.db.sql(f"""
        SELECT dm.name AS doctor_code, dm.doctor_name, dm.place, dm.hospital_address,
               dm.hq, dm.team, dm.region, dm.zone,
               first_scheme.first_date, first_scheme.approved_by
        FROM `tabDoctor Master` dm
        INNER JOIN (
            SELECT sr.doctor_code,
                   MIN(sr.application_date) AS first_date,
                   SUBSTRING_INDEX(GROUP_CONCAT(sr.modified_by ORDER BY sr.application_date), ',', 1) AS approved_by
            FROM `tabScheme Request` sr
            {product_join}
            WHERE sr.division = %(division)s
              AND sr.docstatus = 1
              AND sr.approval_status = 'Approved'
              {product_cond}
            GROUP BY sr.doctor_code
            HAVING first_date BETWEEN %(from_date)s AND %(to_date)s
        ) first_scheme ON first_scheme.doctor_code = dm.name
        WHERE {hierarchy_where}
        ORDER BY dm.region, dm.hq, dm.doctor_name
    """, params, as_dict=True)

    data = []
    for i, r in enumerate(rows, 1):
        data.append({
            "sno": i,
            "approval_date": str(r.first_date) if r.first_date else "",
            "region": r.region or "",
            "hq": r.hq or "",
            "doctor_name": r.doctor_name or "",
            "hospital": r.hospital_address or "",
            "city": r.place or "",
            "approved_by": r.approved_by or "",
            "team": r.team or "",
            "zone": r.zone or "",
        })

    return {"success": True, "data": data}


@frappe.whitelist()
def get_scheme_periodic_report(division=None, from_date=None, to_date=None,
                                reporting_criteria="HQ", zone=None, region=None,
                                team=None, hq=None, doctor=None,
                                product_codes=None, product_group=None, product_category=None):
    """Report 4 – Periodic Report.
    Scheme summary aggregated by the Reporting Criteria dimension
    (Organization / Zone / Region / Team / HQ / Doctor / Stockist / Month / Value),
    including free qty and discount value.
    """
    if not division:
        division = get_user_division()
    if not from_date or not to_date:
        return {"success": False, "message": "From Date and To Date are required"}

    conditions = [
        "sr.division = %(division)s",
        "sr.docstatus = 1",
        "sr.approval_status = 'Approved'",
        "sr.application_date BETWEEN %(from_date)s AND %(to_date)s",
    ]
    params = {"division": division, "from_date": from_date, "to_date": to_date}

    conditions += _scheme_geo_conditions("dm", params, zone, region, team, hq, doctor, division=division)

    resolved_pks = _resolve_product_filter(division, product_codes=product_codes,
                                           product_group=product_group, product_category=product_category)
    if resolved_pks is not None:
        placeholders = ", ".join(["%(pc_" + str(i) + ")s" for i in range(len(resolved_pks))])
        conditions.append("sri.product_code IN (" + placeholders + ")")
        for i, pc in enumerate(resolved_pks):
            params["pc_" + str(i)] = pc

    where = " AND ".join(conditions)

    # Discount value aggregate: (order qty ÷ strips-per-box) × (PTS − special price)
    # where a special price is set. The ÷conv keeps discount per-box so it reconciles
    # with SUM(product_value) below (order value is already per-box).
    _conv_sql = _sql_pack_conversion("sri.pack")
    disc_expr = ("SUM(CASE WHEN sri.special_rate > 0 "
                 f"THEN (sri.quantity / {_conv_sql}) * (sri.product_rate - sri.special_rate) "
                 "ELSE 0 END) AS discount_value")

    # Reporting Criteria → GROUP BY mapping
    group_map = {
        "Organization": {"select": "'Organization' AS group_key, 'Whole Organization' AS group_label", "group": "1"},
        "Zone":     {"select": "IFNULL(dm.zone,'') AS group_key, IFNULL(dm.zone,'(No Zone)') AS group_label", "group": "dm.zone"},
        "Region":   {"select": "dm.region AS group_key, dm.region AS group_label", "group": "dm.region"},
        "Team":     {"select": "dm.team AS group_key, dm.team AS group_label", "group": "dm.team"},
        "HQ":       {"select": "dm.hq AS group_key, hm.hq_name AS group_label", "group": "dm.hq"},
        "Doctor":   {"select": "dm.name AS group_key, dm.doctor_name AS group_label, dm.hq, dm.region", "group": "dm.name"},
        "Stockist": {"select": "sr.stockist_code AS group_key, COALESCE(sm.stockist_name,'') AS group_label", "group": "sr.stockist_code"},
        "Month":    {"select": "DATE_FORMAT(sr.application_date, '%%Y-%%m') AS group_key, DATE_FORMAT(sr.application_date, '%%b %%Y') AS group_label", "group": "DATE_FORMAT(sr.application_date, '%%Y-%%m')"},
    }

    gb = group_map.get(reporting_criteria, group_map["HQ"])
    extra_select = gb["select"]
    group_clause = gb["group"]

    # Value mode: per-scheme, sorted by value desc
    if reporting_criteria == "Value":
        rows = frappe.db.sql(f"""
            SELECT sr.name AS group_key, CONCAT(sr.name, ' - ', dm.doctor_name) AS group_label,
                   dm.hq, dm.region,
                   SUM(sri.quantity) AS total_qty,
                   SUM(sri.free_quantity) AS free_qty,
                   {disc_expr},
                   SUM(sri.product_value) AS total_value
            FROM `tabScheme Request` sr
            INNER JOIN `tabScheme Request Item` sri ON sri.parent = sr.name
            INNER JOIN `tabDoctor Master` dm ON dm.name = sr.doctor_code
            LEFT JOIN `tabHQ Master` hm ON hm.name = dm.hq
            WHERE {where}
            GROUP BY sr.name, dm.doctor_name, dm.hq, dm.region
            ORDER BY total_value DESC
        """, params, as_dict=True)
    else:
        stockist_join = ""
        if reporting_criteria == "Stockist":
            stockist_join = "LEFT JOIN `tabStockist Master` sm ON sm.name = sr.stockist_code"

        order_clause = "total_value DESC" if reporting_criteria == "Organization" else group_clause
        rows = frappe.db.sql(f"""
            SELECT {extra_select},
                   SUM(sri.quantity) AS total_qty,
                   SUM(sri.free_quantity) AS free_qty,
                   {disc_expr},
                   SUM(sri.product_value) AS total_value
            FROM `tabScheme Request` sr
            INNER JOIN `tabScheme Request Item` sri ON sri.parent = sr.name
            INNER JOIN `tabDoctor Master` dm ON dm.name = sr.doctor_code
            LEFT JOIN `tabHQ Master` hm ON hm.name = dm.hq
            {stockist_join}
            WHERE {where}
            GROUP BY {group_clause}
            ORDER BY {order_clause}
        """, params, as_dict=True)

    data = []
    for r in rows:
        entry = {
            "group_key": r.group_key or "",
            "group_label": r.group_label or "",
            "total_qty": flt(r.total_qty),
            "free_qty": flt(r.free_qty),
            "discount_value": flt(r.discount_value),
            "total_value": flt(r.total_value),
        }
        if hasattr(r, "hq"):
            entry["hq"] = r.hq or ""
        if hasattr(r, "region"):
            entry["region"] = r.region or ""
        data.append(entry)

    return {"success": True, "data": data, "group_by": reporting_criteria}


@frappe.whitelist()
def get_pending_scheme_deduction_report(division=None, month=None,
                                        reporting_criteria="Organization",
                                        zone=None, region=None, team=None, hq=None, doctor=None,
                                        product_codes=None, product_group=None, product_category=None,
                                        scheme_type="All"):
    """Monthly Pending Scheme Deduction Report.

    Lists Approved scheme lines (by scheme application month) that are NOT yet
    deducted — i.e. no non-cancelled Scheme Deduction exists for the scheme — and
    that carry free goods and/or a special price. This is what still needs to be
    deducted from the stockist statements; at month-end it surfaces what will spill
    into next month.

    A scheme is only 'deducted' via a Scheme Deduction document — manually editing a
    statement does NOT clear a scheme from this report.
    """
    if not division:
        division = get_user_division()

    from datetime import date as _date
    if not month:
        month = _date.today().strftime("%Y-%m")
    from_date, to_date = _month_bounds(month)
    if not from_date:
        return {"success": False, "message": "Invalid month (expected YYYY-MM)."}

    conditions = [
        "sr.division = %(division)s",
        "sr.docstatus = 1",
        "sr.approval_status = 'Approved'",
        "sr.application_date BETWEEN %(from_date)s AND %(to_date)s",
        "NOT EXISTS (SELECT 1 FROM `tabScheme Deduction` sd "
        "WHERE sd.scheme_request = sr.name AND sd.docstatus != 2)",
    ]
    params = {"division": division, "from_date": from_date, "to_date": to_date}

    # A pending line must carry free goods and/or a discount
    if scheme_type == "Free":
        conditions.append("sri.free_quantity > 0")
    elif scheme_type == "Discount":
        conditions.append("sri.special_rate > 0")
    else:
        conditions.append("(sri.free_quantity > 0 OR sri.special_rate > 0)")

    conditions += _scheme_geo_conditions("dm", params, zone, region, team, hq, doctor, division=division)

    resolved_pks = _resolve_product_filter(division, product_codes=product_codes,
                                           product_group=product_group, product_category=product_category)
    if resolved_pks is not None:
        placeholders = ", ".join(["%(pc_" + str(i) + ")s" for i in range(len(resolved_pks))])
        conditions.append("sri.product_code IN (" + placeholders + ")")
        for i, pc in enumerate(resolved_pks):
            params["pc_" + str(i)] = pc

    where = " AND ".join(conditions)

    # Reporting Criteria drives the subtotal ordering/grouping
    order_map = {
        "Organization": "dm.region, hm.hq_name",
        "Zone":         "dm.zone, dm.region",
        "Region":       "dm.region, hm.hq_name",
        "Team":         "dm.team, hm.hq_name",
        "HQ":           "hm.hq_name, dm.doctor_name",
        "Doctor":       "dm.doctor_name",
        "Stockist":     "stockist_name, dm.doctor_name",
    }
    order_by = order_map.get(reporting_criteria, "dm.region, hm.hq_name")

    rows = frappe.db.sql(f"""
        SELECT sr.name AS scheme_name, sr.application_date AS date,
               IFNULL(dm.zone,'') AS zone, dm.region, dm.team, dm.hq,
               COALESCE(hm.hq_name, dm.hq) AS hq_name,
               dm.doctor_name, dm.name AS doctor_code,
               sr.stockist_code, COALESCE(sm.stockist_name, sr.stockist_name, '') AS stockist_name,
               COALESCE(pm.product_code, sri.product_code) AS product_code,
               COALESCE(pm.product_name, sri.product_name) AS product_name,
               sri.pack, sri.quantity, sri.free_quantity, sri.product_rate,
               sri.special_rate, sri.product_value, sri.idx
        FROM `tabScheme Request` sr
        INNER JOIN `tabScheme Request Item` sri ON sri.parent = sr.name
        INNER JOIN `tabDoctor Master` dm ON dm.name = sr.doctor_code
        LEFT JOIN `tabHQ Master` hm ON hm.name = dm.hq
        LEFT JOIN `tabProduct Master` pm ON pm.name = sri.product_code
        LEFT JOIN `tabStockist Master` sm ON sm.name = sr.stockist_code
        WHERE {where}
        ORDER BY {order_by}, sr.application_date, sri.idx
    """, params, as_dict=True)

    def group_label_for(r):
        if reporting_criteria == "Zone":
            return r.zone or "(No Zone)"
        if reporting_criteria == "Region":
            return r.region or "(No Region)"
        if reporting_criteria == "Team":
            return r.team or "(No Team)"
        if reporting_criteria == "HQ":
            return r.hq_name or "(No HQ)"
        if reporting_criteria == "Doctor":
            return r.doctor_name or "(No Doctor)"
        if reporting_criteria == "Stockist":
            return r.stockist_name or "(No Stockist)"
        return "Whole Organization"

    data = []
    schemes_seen = set()
    tot_free = tot_disc = tot_value = 0
    for i, r in enumerate(rows, 1):
        special_rate = flt(r.special_rate)
        # (order qty ÷ strips-per-box) × (PTS − special); ÷conv keeps it per-box so it
        # reconciles with the order value (product_value) below.
        _conv = _scheme_pack_conversion(r.pack)
        discount_value = (flt(r.quantity) / _conv) * (flt(r.product_rate) - special_rate) if special_rate > 0 else 0
        schemes_seen.add(r.scheme_name)
        data.append({
            "sno": i,
            "scheme_name": r.scheme_name,
            "group_label": group_label_for(r),
            "date": str(r.date) if r.date else "",
            "region": r.region or "",
            "hq": r.hq or "",
            "hq_name": r.hq_name or "",
            "team": r.team or "",
            "doctor_name": r.doctor_name or "",
            "stockist_name": r.stockist_name or "",
            "product_code": r.product_code or "",
            "product_name": r.product_name or "",
            "pack": r.pack or "",
            "qty": flt(r.quantity),
            "free_qty": flt(r.free_quantity),
            "rate": flt(r.product_rate),
            "special_rate": special_rate,
            "discount_value": discount_value,
            "value": flt(r.product_value),
        })
        tot_free += flt(r.free_quantity)
        tot_disc += discount_value
        tot_value += flt(r.product_value)

    return {
        "success": True,
        "data": data,
        "group_by": reporting_criteria,
        "month": month,
        "totals": {
            "scheme_count": len(schemes_seen),
            "free_qty": tot_free,
            "discount_value": tot_disc,
            "value": tot_value,
        },
    }


@frappe.whitelist()
def export_scheme_report_excel(report_type, division=None, **kwargs):
    """Generate a styled Excel workbook for scheme reports."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    if not division:
        division = get_user_division()

    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    group_font = Font(bold=True, size=11)
    group_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

    wb = openpyxl.Workbook()
    ws = wb.active

    def write_header_row(ws, row_num, headers):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

    def write_data_row(ws, row_num, values):
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=v)
            cell.border = thin_border

    def write_group_row(ws, row_num, label, num_cols):
        cell = ws.cell(row=row_num, column=1, value=label)
        cell.font = group_font
        cell.fill = group_fill
        for c in range(1, num_cols + 1):
            ws.cell(row=row_num, column=c).fill = group_fill
            ws.cell(row=row_num, column=c).border = thin_border

    def write_title_rows(ws, title, subtitle=""):
        ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=14)
        if subtitle:
            ws.cell(row=2, column=1, value=subtitle).font = Font(size=11, italic=True)
        return 4

    from_date = kwargs.get("from_date", "")
    to_date = kwargs.get("to_date", "")
    month = kwargs.get("month", "")
    region_val = kwargs.get("region", "")
    zone_val = kwargs.get("zone", "")
    team_val = kwargs.get("team", "")
    hq_val = kwargs.get("hq", "")
    doctor_val = kwargs.get("doctor", "")
    doctor_status = kwargs.get("doctor_status", "Active")
    product_codes = kwargs.get("product_codes", None)
    product_group = kwargs.get("product_group", None)
    product_category = kwargs.get("product_category", None)
    reporting_criteria = kwargs.get("reporting_criteria", "Organization")
    scheme_type = kwargs.get("scheme_type", "All")
    period_label = f"{from_date} to {to_date}" if from_date and to_date else ""

    ml = ["Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan", "Feb", "Mar"]

    if report_type == "activity_trend":
        ws.title = "Activity Trend"
        result = get_scheme_activity_trend_report(
            division, from_date, to_date, doctor_status, reporting_criteria,
            zone_val, region_val, team_val, hq_val, doctor_val,
            product_codes, product_group, product_category
        )
        data = result.get("data", [])
        fy_label = result.get("fy_label", "")
        row = write_title_rows(ws, f"Activity Trend Report – {division}", f"{fy_label}")
        headers = ["Region", "HQ", "Doctor Name"] + ml
        write_header_row(ws, row, headers)
        row += 1
        current_region = None
        for d in data:
            if d["region"] != current_region:
                current_region = d["region"]
                write_group_row(ws, row, f"{current_region} Region", len(headers))
                row += 1
            vals = [d["region"], d.get("hq_name") or d["hq"], d["doctor_name"]] + d["months"]
            write_data_row(ws, row, vals)
            row += 1

    elif report_type == "activity_track":
        ws.title = "Activity Track"
        result = get_scheme_activity_track_report(
            division, from_date, to_date, doctor_status, reporting_criteria,
            zone_val, region_val, team_val, hq_val, doctor_val,
            product_codes, product_group, product_category
        )
        data = result.get("data", [])
        totals = result.get("totals", {})
        row = write_title_rows(ws, f"Activity Track Report – {division}",
                               f"Period: {period_label}")
        headers = ["S.No", "Date", "Region", "HQ", "Doctor Name", "Product",
                    "Qty", "Free Qty", "Rate", "Special Price", "Discount Value",
                    "Value", "Stockist"]
        write_header_row(ws, row, headers)
        row += 1
        for d in data:
            write_data_row(ws, row, [d["sno"], d["date"], d["region"], d["hq"],
                                     d["doctor_name"], d.get("product_name") or d["product_code"],
                                     d["qty"], d["free_qty"], d["rate"],
                                     d.get("special_rate", 0), d.get("discount_value", 0),
                                     d["value"], d["stockist_name"]])
            row += 1
        # Totals row
        write_group_row(ws, row, "Total", len(headers))
        ws.cell(row=row, column=7, value=totals.get("qty", 0)).font = group_font
        ws.cell(row=row, column=8, value=totals.get("free_qty", 0)).font = group_font
        ws.cell(row=row, column=11, value=totals.get("discount_value", 0)).font = group_font
        ws.cell(row=row, column=12, value=totals.get("value", 0)).font = group_font

    elif report_type == "new_approval_doctors":
        ws.title = "New Approval Doctors"
        result = get_new_approval_doctors_report(
            division, from_date, to_date, reporting_criteria,
            zone_val, region_val, team_val, hq_val, doctor_val,
            product_codes, product_group, product_category
        )
        data = result.get("data", [])
        row = write_title_rows(ws, f"New Approval Doctors – {division}",
                               f"Period: {period_label}")
        headers = ["S.No", "Approval Date", "Region", "HQ", "Doctor Name",
                    "Hospital", "City", "Approved By"]
        write_header_row(ws, row, headers)
        row += 1
        for d in data:
            write_data_row(ws, row, [d["sno"], d["approval_date"], d["region"],
                                     d["hq"], d["doctor_name"], d["hospital"],
                                     d["city"], d["approved_by"]])
            row += 1

    elif report_type == "periodic":
        ws.title = "Periodic Report"
        result = get_scheme_periodic_report(
            division, from_date, to_date, reporting_criteria,
            zone_val, region_val, team_val, hq_val, doctor_val,
            product_codes, product_group, product_category
        )
        data = result.get("data", [])
        gb = result.get("group_by", "HQ")
        row = write_title_rows(ws, f"Periodic Report ({gb} Wise) – {division}",
                               f"Period: {period_label}")
        headers = [gb, "Total Qty", "Free Qty", "Discount Value", "Total Value"]
        write_header_row(ws, row, headers)
        row += 1
        for d in data:
            write_data_row(ws, row, [d["group_label"], d["total_qty"],
                                     d["free_qty"], d.get("discount_value", 0), d["total_value"]])
            row += 1

    elif report_type == "pending_deduction":
        ws.title = "Pending Deduction"
        result = get_pending_scheme_deduction_report(
            division, month, reporting_criteria,
            zone_val, region_val, team_val, hq_val, doctor_val,
            product_codes, product_group, product_category, scheme_type
        )
        data = result.get("data", [])
        totals = result.get("totals", {})
        row = write_title_rows(ws, f"Monthly Pending Scheme Deduction – {division}",
                               f"Month: {month}  |  Criteria: {reporting_criteria}")
        headers = ["S.No", "Date", "Region", "HQ", "Doctor Name", "Stockist",
                    "Product", "Order Qty", "Free Qty", "PTS", "Special Price",
                    "Discount Value", "Scheme Value"]
        write_header_row(ws, row, headers)
        row += 1
        current_group = None
        for d in data:
            if d["group_label"] != current_group:
                current_group = d["group_label"]
                write_group_row(ws, row, current_group, len(headers))
                row += 1
            write_data_row(ws, row, [d["sno"], d["date"], d["region"], d.get("hq_name") or d["hq"],
                                     d["doctor_name"], d["stockist_name"],
                                     d.get("product_name") or d["product_code"],
                                     d["qty"], d["free_qty"], d["rate"],
                                     d["special_rate"], d["discount_value"], d["value"]])
            row += 1
        write_group_row(ws, row, "Grand Total", len(headers))
        ws.cell(row=row, column=9, value=totals.get("free_qty", 0)).font = group_font
        ws.cell(row=row, column=12, value=totals.get("discount_value", 0)).font = group_font
        ws.cell(row=row, column=13, value=totals.get("value", 0)).font = group_font

    else:
        frappe.throw("Invalid report type")

    # Auto-fit column widths
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    xlsx_data = output.getvalue()

    filename = f"Scheme_Report_{report_type}_{division}.xlsx"
    frappe.local.response.filename = filename
    frappe.local.response.filecontent = xlsx_data
    frappe.local.response.type = "download"
    frappe.local.response.content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


# ═══════════════════════════════════════════════════════════════
# RANKING REPORTS  –  Portal API Methods
# ═══════════════════════════════════════════════════════════════

_MONTH_LABELS = ["Apr", "May", "Jun", "Jul", "Aug", "Sep",
                 "Oct", "Nov", "Dec", "Jan", "Feb", "Mar"]
_MONTH_MAP = {4: 0, 5: 1, 6: 2, 7: 3, 8: 4, 9: 5,
              10: 6, 11: 7, 12: 8, 1: 9, 2: 10, 3: 11}


def _current_fy():
    """Return (fy_start_date, fy_end_date, fy_label) for the current financial year."""
    from datetime import date
    today = date.today()
    if today.month >= 4:
        fy_start = date(today.year, 4, 1)
        fy_end = date(today.year + 1, 3, 31)
    else:
        fy_start = date(today.year - 1, 4, 1)
        fy_end = date(today.year, 3, 31)
    fy_label = f"Apr {str(fy_start.year)[2:]} to Mar {str(fy_end.year)[2:]}"
    return str(fy_start), str(fy_end), fy_label


@frappe.whitelist()
def get_ranking_report_filter_options(division=None):
    """Return dropdown options for the Ranking Reports portal page (active masters only)."""
    if not division:
        division = get_user_division()

    # None for admin (= all regions); a list confines a non-admin's pickers below.
    allowed = _allowed_region_codes_or_all(division)

    zones = frappe.db.sql(
        "SELECT name FROM `tabZone Master` WHERE status='Active' AND division IN (%s, 'Both') ORDER BY name",
        (division,), as_list=1)
    regions = frappe.db.sql(
        "SELECT name, region_name, zone FROM `tabRegion Master` WHERE status='Active' AND division IN (%s, 'Both') ORDER BY name",
        (division,), as_dict=True)
    teams = frappe.db.sql(
        "SELECT name, region FROM `tabTeam Master` WHERE status='Active' AND division IN (%s, 'Both') ORDER BY name",
        (division,), as_dict=True)
    hqs = frappe.db.sql(
        "SELECT name, hq_name, team, IFNULL(region, '') AS region FROM `tabHQ Master` WHERE division=%s AND status='Active' ORDER BY name",
        (division,), as_dict=True)
    stockists = frappe.db.sql(
        "SELECT name, stockist_name, hq, IFNULL(region, '') AS region FROM `tabStockist Master` WHERE division=%s AND status='Active' ORDER BY stockist_name",
        (division,), as_dict=True)
    products = frappe.db.sql(
        "SELECT product_code, product_name, category, product_group, pack, sequence "
        "FROM `tabProduct Master` WHERE division=%s AND status='Active' ORDER BY COALESCE(sequence, 9999), product_name",
        (division,), as_dict=True)
    doctors = frappe.db.sql(
        "SELECT name, doctor_code, doctor_name, hq, region "
        "FROM `tabDoctor Master` WHERE division=%s AND status='Active' ORDER BY doctor_name",
        (division,), as_dict=True)

    # Confine a non-admin's pickers to their mapped regions (and the teams/HQs/stockists/
    # doctors inside them). Admin (allowed is None) sees everything.
    if allowed is not None:
        aset = set(allowed)
        regions = [r for r in regions if r.name in aset]
        teams = [t for t in teams if (t.region or "") in aset]
        hqs = [h for h in hqs if (h.region or "") in aset]
        stockists = [s for s in stockists if (s.region or "") in aset]
        doctors = [d for d in doctors if (d.region or "") in aset]

    return {
        "zones": [z[0] for z in zones],
        "regions": [{"name": r.name, "region_name": r.region_name or r.name, "zone": r.zone or ""} for r in regions],
        "teams": [{"name": t.name, "region": t.region or ""} for t in teams],
        "hqs": [{"name": h.name, "hq_name": h.hq_name or "", "team": h.team or ""} for h in hqs],
        "stockists": [{"code": s.name, "name": s.stockist_name, "hq": s.hq or ""} for s in stockists],
        "products": [{"code": p.product_code, "name": p.product_name,
                       "category": p.category or "", "group": p.product_group or "",
                       "pack": p.pack or ""} for p in products],
        "doctors": [{"code": d.name, "doctor_code": d.doctor_code or d.name,
                      "name": d.doctor_name, "hq": d.hq or "",
                      "region": d.region or ""} for d in doctors],
    }


# ─────────────────────────────────────────────────────────────
# Report 1: Moving Trend Report
# ─────────────────────────────────────────────────────────────
@frappe.whitelist()
def get_ranking_moving_trend_report(division=None, sales_type="secondary",
                                     criteria="Region", from_date=None, to_date=None,
                                     region=None, zone=None):
    """Monthly pivot (Apr–Mar) grouped by criteria (Region / HQ / Team / Doctor / Stockist / Product)."""
    if not division:
        division = get_user_division()

    fy_start, fy_end, fy_label = _current_fy()
    if not from_date:
        from_date = fy_start
    if not to_date:
        to_date = fy_end

    if sales_type == "primary":
        rows = _moving_trend_primary(division, criteria, from_date, to_date, region, zone)
    else:
        rows = _moving_trend_secondary(division, criteria, from_date, to_date, region, zone)

    # Build pivot
    pivoted = {}
    for r in rows:
        key = r.criteria_name
        if key not in pivoted:
            pivoted[key] = {"criteria_name": key, "months": [0] * 12, "total": 0}
        idx = _MONTH_MAP.get(r.m)
        if idx is not None:
            pivoted[key]["months"][idx] += flt(r.qty)
            pivoted[key]["total"] += flt(r.qty)

    data = list(pivoted.values())
    for d in data:
        d["average"] = round(d["total"] / 12, 2)
    data.sort(key=lambda x: x["total"], reverse=True)

    # Region / HQ / Team are grouped by the Master id (code); the sheet shows names.
    if criteria == "HQ":
        hq_lut, team_lut = _name_lut("HQ Master", "hq_name"), _name_lut("Team Master", "team_name")
        for d in data:
            v = d["criteria_name"]
            d["criteria_name"] = hq_lut.get(v) or team_lut.get(v) or v
    elif criteria in ("Region", "Team"):
        lut = _name_lut("Region Master", "region_name") if criteria == "Region" \
            else _name_lut("Team Master", "team_name")
        for d in data:
            d["criteria_name"] = lut.get(d["criteria_name"], d["criteria_name"])

    return {"success": True, "data": data, "fy_label": fy_label,
            "month_labels": _MONTH_LABELS, "criteria": criteria}


def _pri_master_name(doctype, value, name_field):
    """Primary Sales Data stores region/zone/team by their DISPLAY NAME, while the
    ranking dropdowns send the Master's code (its PK). Resolve code → stored name so
    primary-sales filters actually match (mirrors the stockist primary report)."""
    if not value:
        return value
    return frappe.db.get_value(doctype, value, name_field) or value


def _name_lut(doctype, name_field):
    """PK → display-name map for a hierarchy master (single query, fail-open).

    Ranking rows store the Master's id (its internal code) for region / hq / team;
    the printed sheets show only the human name. Callers do ``lut.get(value, value)``
    so a value that is already a name (e.g. Primary Sales stores region by name)
    passes straight through unchanged."""
    rows = frappe.get_all(doctype, fields=["name", name_field])
    return {r["name"]: (r.get(name_field) or r["name"]) for r in rows}


def _to_roman(n):
    """Small positive int → Roman numeral (rank labels I, II, III …)."""
    if not n or n < 1:
        return ""
    table = [(1000, "M"), (900, "CM"), (500, "D"), (400, "CD"),
             (100, "C"), (90, "XC"), (50, "L"), (40, "XL"),
             (10, "X"), (9, "IX"), (5, "V"), (4, "IV"), (1, "I")]
    out = []
    for val, sym in table:
        while n >= val:
            out.append(sym)
            n -= val
    return "".join(out)


def _moving_trend_primary(division, criteria, from_date, to_date, region, zone):
    """Query Primary Sales Data for Moving Trend grouped by criteria."""
    criteria_col_map = {
        "Region": "ps.region",
        "HQ": "IFNULL(sm.hq, ps.team)",
        "Team": "ps.team",
        "Stockist": "CONCAT(ps.stockist_code, ' – ', ps.stockist_name)",
        "Product": "CONCAT(ps.pcode, ' – ', ps.product)",
    }
    criteria_col = criteria_col_map.get(criteria, "ps.region")

    conditions = ["ps.division = %(division)s", "ps.iscancelled = 0",
                   "ps.invoicedate >= %(from_date)s", "ps.invoicedate <= %(to_date)s"]
    params = {"division": division, "from_date": from_date, "to_date": to_date}

    join_sm = ""
    if criteria == "HQ":
        join_sm = " LEFT JOIN `tabStockist Master` sm ON sm.name = ps.stockist_code AND sm.status = 'Active'"
    # Primary Sales Data stores the region NAME, so match code-or-name.
    _scope_region_sql(conditions, params, "ps.region", division, region, include_names=True)
    if zone:
        conditions.append("ps.zonee = %(zone)s")
        params["zone"] = _pri_master_name("Zone Master", zone, "zone_name")

    where = " AND ".join(conditions)
    return frappe.db.sql(f"""
        SELECT {criteria_col} AS criteria_name,
               MONTH(ps.invoicedate) AS m,
               SUM(ps.quantity) AS qty
        FROM `tabPrimary Sales Data` ps
        {join_sm}
        WHERE {where}
        GROUP BY criteria_name, MONTH(ps.invoicedate)
    """, params, as_dict=True)


def _moving_trend_secondary(division, criteria, from_date, to_date, region, zone):
    """Query Stockist Statement Items for Moving Trend grouped by criteria."""
    criteria_col_map = {
        "Region": "ss.region",
        "HQ": "IFNULL(sm.hq, ss.team)",
        "Team": "ss.team",
        "Stockist": "CONCAT(ss.stockist_code, ' – ', ss.stockist_name)",
        # si.product_code stores the Product Master id; show the business code.
        "Product": "CONCAT(COALESCE(pm.product_code, si.product_code), ' – ', si.product_name)",
    }
    criteria_col = criteria_col_map.get(criteria, "ss.region")

    conditions = ["ss.division = %(division)s", "ss.docstatus IN (0, 1)",
                   "ss.statement_month >= %(from_date)s", "ss.statement_month <= %(to_date)s"]
    params = {"division": division, "from_date": from_date, "to_date": to_date}

    join_sm = ""
    if criteria == "HQ":
        join_sm = " LEFT JOIN `tabStockist Master` sm ON sm.name = ss.stockist_code AND sm.status = 'Active'"
    if criteria == "Product":
        join_sm += " LEFT JOIN `tabProduct Master` pm ON pm.name = si.product_code"
    _scope_region_sql(conditions, params, "ss.region", division, region)
    if zone:
        conditions.append("ss.zone = %(zone)s")
        params["zone"] = zone

    where = " AND ".join(conditions)
    return frappe.db.sql(f"""
        SELECT {criteria_col} AS criteria_name,
               MONTH(ss.statement_month) AS m,
             SUM(si.sales_qty / IFNULL(NULLIF(si.conversion_factor, 0), 1)) AS qty
        FROM `tabStockist Statement` ss
        INNER JOIN `tabStockist Statement Item` si
            ON si.parent = ss.name AND si.parenttype = 'Stockist Statement'
        {join_sm}
        WHERE {where}
        GROUP BY criteria_name, MONTH(ss.statement_month)
    """, params, as_dict=True)


# ─────────────────────────────────────────────────────────────
# Report 2: Rupee Wise Report
# ─────────────────────────────────────────────────────────────
@frappe.whitelist()
def get_ranking_rupee_wise_report(division=None, sales_type="secondary",
                                   value_condition="gt", sale_value=0,
                                   from_date=None, to_date=None):
    """Transaction-level rows filtered by value > or < threshold."""
    if not division:
        division = get_user_division()
    sale_value = flt(sale_value)
    val_op = ">=" if value_condition == "gt" else "<="

    if sales_type == "primary":
        conditions = ["ps.division = %(division)s", "ps.iscancelled = 0",
                       f"ps.ptsvalue {val_op} %(sale_value)s"]
        params = {"division": division, "sale_value": sale_value}
        # No region filter on this report — still confine non-admins to their regions.
        _scope_region_sql(conditions, params, "ps.region", division, None, include_names=True)
        if from_date:
            conditions.append("ps.invoicedate >= %(from_date)s")
            params["from_date"] = from_date
        if to_date:
            conditions.append("ps.invoicedate <= %(to_date)s")
            params["to_date"] = to_date
        where = " AND ".join(conditions)

        rows = frappe.db.sql(f"""
            SELECT ps.invoicedate AS date, ps.region, ps.team AS hq,
                   ps.stockist_code, ps.stockist_name,
                   ps.pcode AS product_code, ps.product AS product_name,
                   ps.quantity AS qty, ps.pts AS rate, ps.ptsvalue AS value
            FROM `tabPrimary Sales Data` ps
            WHERE {where}
            ORDER BY ps.invoicedate, ps.stockist_code
            LIMIT 5000
        """, params, as_dict=True)
    else:
        conditions = ["ss.division = %(division)s", "ss.docstatus IN (0, 1)",
                       f"si.sales_value_pts {val_op} %(sale_value)s"]
        params = {"division": division, "sale_value": sale_value}
        # No region filter on this report — still confine non-admins to their regions.
        _scope_region_sql(conditions, params, "ss.region", division, None)
        if from_date:
            conditions.append("ss.statement_month >= %(from_date)s")
            params["from_date"] = from_date
        if to_date:
            conditions.append("ss.statement_month <= %(to_date)s")
            params["to_date"] = to_date
        where = " AND ".join(conditions)

        rows = frappe.db.sql(f"""
            SELECT ss.statement_month AS date, ss.region, ss.hq,
                   ss.stockist_code, ss.stockist_name,
                   si.product_code, si.product_name,
                     (si.sales_qty / IFNULL(NULLIF(si.conversion_factor, 0), 1)) AS qty, si.pts AS rate, si.sales_value_pts AS value
            FROM `tabStockist Statement` ss
            INNER JOIN `tabStockist Statement Item` si
                ON si.parent = ss.name AND si.parenttype = 'Stockist Statement'
            WHERE {where}
            ORDER BY ss.statement_month, ss.stockist_code
            LIMIT 5000
        """, params, as_dict=True)
        # Statement items link by the Product Master id; show business codes.
        _apply_product_display_codes(rows)

    # Region / HQ are stored as Master ids (codes); show their names only.
    reg_lut, hq_lut = _name_lut("Region Master", "region_name"), _name_lut("HQ Master", "hq_name")
    for r in rows:
        r["region"] = reg_lut.get(r.get("region"), r.get("region"))
        r["hq"] = hq_lut.get(r.get("hq"), r.get("hq"))

    # Add serial number
    for i, r in enumerate(rows, 1):
        r["sno"] = i
        if r.get("date"):
            r["date"] = str(r["date"])

    return {"success": True, "data": rows}


# ─────────────────────────────────────────────────────────────
# Report 3: Productwise Ranking (Top N)
# ─────────────────────────────────────────────────────────────
@frappe.whitelist()
def get_ranking_productwise_topn(division=None, product_codes=None, top_n=5,
                                  from_date=None, to_date=None, sales_type="secondary"):
    """Rank selected products by total value, return top N with contribution %."""
    if not division:
        division = get_user_division()
    top_n = int(top_n or 5)

    codes = []
    if product_codes:
        codes = json.loads(product_codes) if isinstance(product_codes, str) else product_codes

    if sales_type == "primary":
        conditions = ["ps.division = %(division)s", "ps.iscancelled = 0"]
        params = {"division": division}
        # No region filter on this report — still confine non-admins to their regions.
        _scope_region_sql(conditions, params, "ps.region", division, None, include_names=True)
        if from_date:
            conditions.append("ps.invoicedate >= %(from_date)s")
            params["from_date"] = from_date
        if to_date:
            conditions.append("ps.invoicedate <= %(to_date)s")
            params["to_date"] = to_date
        if codes:
            conditions.append("ps.pcode IN %(codes)s")
            params["codes"] = codes
        where = " AND ".join(conditions)

        rows = frappe.db.sql(f"""
            SELECT ps.pcode AS product_code, ps.product AS product_name,
                   SUM(ps.quantity) AS total_qty, SUM(ps.ptsvalue) AS total_value
            FROM `tabPrimary Sales Data` ps
            WHERE {where}
            GROUP BY ps.pcode, ps.product
            ORDER BY total_value DESC
        """, params, as_dict=True)
    else:
        conditions = ["ss.division = %(division)s", "ss.docstatus IN (0, 1)"]
        params = {"division": division}
        # No region filter on this report — still confine non-admins to their regions.
        _scope_region_sql(conditions, params, "ss.region", division, None)
        if from_date:
            conditions.append("ss.statement_month >= %(from_date)s")
            params["from_date"] = from_date
        if to_date:
            conditions.append("ss.statement_month <= %(to_date)s")
            params["to_date"] = to_date
        if codes:
            # UI sends business codes; the item Link column stores ids.
            conditions.append("si.product_code IN %(codes)s")
            params["codes"] = _resolve_product_filter(division, product_codes=codes) or ["__no_match__"]
        where = " AND ".join(conditions)

        rows = frappe.db.sql(f"""
            SELECT si.product_code, si.product_name,
                     SUM(si.sales_qty / IFNULL(NULLIF(si.conversion_factor, 0), 1)) AS total_qty,
                   SUM(si.sales_value_pts) AS total_value
            FROM `tabStockist Statement` ss
            INNER JOIN `tabStockist Statement Item` si
                ON si.parent = ss.name AND si.parenttype = 'Stockist Statement'
            WHERE {where}
            GROUP BY si.product_code, si.product_name
            ORDER BY total_value DESC
        """, params, as_dict=True)
        _apply_product_display_codes(rows)

    grand_total = sum(flt(r.total_value) for r in rows)
    data = []
    for rank, r in enumerate(rows[:top_n], 1):
        pct = round(flt(r.total_value) / grand_total * 100, 2) if grand_total else 0
        data.append({
            "rank": rank,
            "product_code": r.product_code,
            "product_name": r.product_name,
            "total_qty": flt(r.total_qty),
            "total_value": flt(r.total_value),
            "contribution_pct": pct,
        })

    return {"success": True, "data": data, "grand_total": grand_total}


# ─────────────────────────────────────────────────────────────
# Report 4: Product Wise Ranking Sheet
#   For each product, dense-rank headquarters by sales quantity and keep the
#   top-N ranks. Mirrors the printed Stedman sheet:
#   Product Code | Pack | Rank | Headquarters | Region | Sales.
# ─────────────────────────────────────────────────────────────
@frappe.whitelist()
def get_ranking_productwise_all(division=None, product_code=None, region=None,
                                 sales_type="secondary", from_date=None, to_date=None,
                                 top_n=2):
    """Product Wise Ranking Sheet — per product, dense-rank HQs by sales, keep top-N ranks.

    ``top_n`` counts *distinct* ranks (default 2 → I & II); tied HQs share a rank,
    so a product can have more rows than ``top_n``. ``product_code`` optionally
    narrows to a single product; otherwise every product with sales is listed,
    ordered by the Product Master sequence."""
    if not division:
        division = get_user_division()
    try:
        top_n = int(top_n or 2)
    except (TypeError, ValueError):
        top_n = 2
    if top_n < 1:
        top_n = 2

    if sales_type == "primary":
        conditions = ["ps.division = %(division)s", "ps.iscancelled = 0"]
        params = {"division": division}
        if product_code:
            conditions.append("ps.pcode = %(pcode)s")
            params["pcode"] = product_code
        # Primary Sales Data stores the region NAME, so match code-or-name.
        _scope_region_sql(conditions, params, "ps.region", division, region, include_names=True)
        if from_date:
            conditions.append("ps.invoicedate >= %(from_date)s")
            params["from_date"] = from_date
        if to_date:
            conditions.append("ps.invoicedate <= %(to_date)s")
            params["to_date"] = to_date
        where = " AND ".join(conditions)

        rows = frappe.db.sql(f"""
            SELECT ps.pcode AS product_code,
                   IFNULL(sm.hq, ps.team) AS hq, ps.region AS region,
                   SUM(ps.quantity) AS sales
            FROM `tabPrimary Sales Data` ps
            LEFT JOIN `tabStockist Master` sm ON sm.name = ps.stockist_code AND sm.status = 'Active'
            WHERE {where}
            GROUP BY ps.pcode, hq, ps.region
        """, params, as_dict=True)
    else:
        conditions = ["ss.division = %(division)s", "ss.docstatus IN (0, 1)"]
        params = {"division": division}
        if product_code:
            # UI sends the business code; the item Link column stores the id.
            conditions.append("si.product_code = %(pcode)s")
            params["pcode"] = _resolve_product_pk(product_code, division) or product_code
        _scope_region_sql(conditions, params, "ss.region", division, region)
        if from_date:
            conditions.append("ss.statement_month >= %(from_date)s")
            params["from_date"] = from_date
        if to_date:
            conditions.append("ss.statement_month <= %(to_date)s")
            params["to_date"] = to_date
        where = " AND ".join(conditions)

        rows = frappe.db.sql(f"""
            SELECT si.product_code, ss.hq AS hq, ss.region AS region,
                   SUM(si.sales_qty / IFNULL(NULLIF(si.conversion_factor, 0), 1)) AS sales
            FROM `tabStockist Statement` ss
            INNER JOIN `tabStockist Statement Item` si
                ON si.parent = ss.name AND si.parenttype = 'Stockist Statement'
            WHERE {where}
            GROUP BY si.product_code, ss.hq, ss.region
        """, params, as_dict=True)
        _apply_product_display_codes(rows)   # item id → business product code

    # Resolve HQ / Region ids to their names (only names on the printed sheet).
    hq_lut, reg_lut = _name_lut("HQ Master", "hq_name"), _name_lut("Region Master", "region_name")

    # Product display info (pack + sequence) keyed by business code.
    prod_info = {}
    for p in frappe.db.sql(
            "SELECT product_code, pack, COALESCE(sequence, 9999) AS seq "
            "FROM `tabProduct Master` WHERE division = %s", (division,), as_dict=True):
        prod_info[p.product_code] = {"pack": p.pack or "", "seq": p.seq}

    # Bucket HQ rows per product. Skip non-product statement lines (no code) and
    # zero-sales rows — the printed sheet only ranks headquarters that actually sold.
    buckets = {}
    for r in rows:
        if not r.product_code:
            continue
        sales = int(round(flt(r.sales)))   # rank/display on the same whole number
        if sales <= 0:
            continue
        buckets.setdefault(r.product_code, []).append({
            "hq": hq_lut.get(r.hq, r.hq) or "",
            "region": reg_lut.get(r.region, r.region) or "",
            "sales": sales,
        })

    # Order products by the master sequence, dense-rank HQs, keep the top-N ranks.
    ordered_codes = sorted(buckets.keys(),
                           key=lambda c: (prod_info.get(c, {}).get("seq", 9999), c or ""))

    data = []
    for pc in ordered_codes:
        hq_rows = sorted(buckets[pc], key=lambda x: x["sales"], reverse=True)
        pack = prod_info.get(pc, {}).get("pack", "")
        rank, prev_sales, first = 0, None, True
        for hr in hq_rows:
            # Dense rank on the displayed value so tied HQs share a rank (I, I, II …).
            if prev_sales is None or hr["sales"] != prev_sales:
                rank += 1
                prev_sales = hr["sales"]
            if rank > top_n:
                break
            data.append({
                "product_code": pc if first else "",   # code shown once per product group
                "pack": pack,
                "rank": rank,
                "rank_roman": _to_roman(rank),
                "hq": hr["hq"],
                "region": hr["region"],
                "sales": hr["sales"],
            })
            first = False

    return {"success": True, "data": data, "top_n": top_n}


# ─────────────────────────────────────────────────────────────
# Report 5: Productwise Ranking Advanced
# ─────────────────────────────────────────────────────────────
@frappe.whitelist()
def get_ranking_productwise_advanced(division=None, sales_type="secondary",
                                      qty_filter=0, region=None, hq_wise=0,
                                      product_codes=None,
                                      from_date=None, to_date=None):
    """Advanced product ranking: Primary/Secondary/Closing, qty >= threshold, group by HQ or Stockist."""
    if not division:
        division = get_user_division()
    qty_filter = flt(qty_filter)
    hq_wise = int(hq_wise or 0)

    codes = []
    if product_codes:
        codes = json.loads(product_codes) if isinstance(product_codes, str) else product_codes

    if sales_type == "closing":
        # Closing stock from statement items
        conditions = ["ss.division = %(division)s", "ss.docstatus IN (0, 1)"]
        params = {"division": division}
        _scope_region_sql(conditions, params, "ss.region", division, region)
        if from_date:
            conditions.append("ss.statement_month >= %(from_date)s")
            params["from_date"] = from_date
        if to_date:
            conditions.append("ss.statement_month <= %(to_date)s")
            params["to_date"] = to_date
        if codes:
            # UI sends business codes; the item Link column stores ids.
            conditions.append("si.product_code IN %(codes)s")
            params["codes"] = _resolve_product_filter(division, product_codes=codes) or ["__no_match__"]
        where = " AND ".join(conditions)

        group_col = "ss.hq" if hq_wise else "CONCAT(ss.stockist_code, ' – ', ss.stockist_name)"
        group_label = "HQ" if hq_wise else "Stockist"

        rows = frappe.db.sql(f"""
            SELECT ss.region, {group_col} AS group_key,
                   si.product_code, si.product_name,
                     SUM(si.closing_qty / IFNULL(NULLIF(si.conversion_factor, 0), 1)) AS qty,
                   SUM(si.closing_value) AS value
            FROM `tabStockist Statement` ss
            INNER JOIN `tabStockist Statement Item` si
                ON si.parent = ss.name AND si.parenttype = 'Stockist Statement'
            WHERE {where}
            GROUP BY ss.region, group_key, si.product_code, si.product_name
            HAVING qty >= %(qty_filter)s
            ORDER BY value DESC
            LIMIT 5000
        """, dict(**params, qty_filter=qty_filter), as_dict=True)

    elif sales_type == "primary":
        conditions = ["ps.division = %(division)s", "ps.iscancelled = 0"]
        params = {"division": division}
        # Primary Sales Data stores the region NAME, so match code-or-name.
        _scope_region_sql(conditions, params, "ps.region", division, region, include_names=True)
        if from_date:
            conditions.append("ps.invoicedate >= %(from_date)s")
            params["from_date"] = from_date
        if to_date:
            conditions.append("ps.invoicedate <= %(to_date)s")
            params["to_date"] = to_date
        if codes:
            conditions.append("ps.pcode IN %(codes)s")
            params["codes"] = codes
        where = " AND ".join(conditions)

        if hq_wise:
            group_col = "IFNULL(sm.hq, ps.team)"
            join_sm = " LEFT JOIN `tabStockist Master` sm ON sm.name = ps.stockist_code AND sm.status = 'Active'"
        else:
            group_col = "CONCAT(ps.stockist_code, ' – ', ps.stockist_name)"
            join_sm = ""
        group_label = "HQ" if hq_wise else "Stockist"

        rows = frappe.db.sql(f"""
            SELECT ps.region, {group_col} AS group_key,
                   ps.pcode AS product_code, ps.product AS product_name,
                   SUM(ps.quantity) AS qty, SUM(ps.ptsvalue) AS value
            FROM `tabPrimary Sales Data` ps
            {join_sm}
            WHERE {where}
            GROUP BY ps.region, group_key, ps.pcode, ps.product
            HAVING qty >= %(qty_filter)s
            ORDER BY value DESC
            LIMIT 5000
        """, dict(**params, qty_filter=qty_filter), as_dict=True)

    else:  # secondary
        conditions = ["ss.division = %(division)s", "ss.docstatus IN (0, 1)"]
        params = {"division": division}
        _scope_region_sql(conditions, params, "ss.region", division, region)
        if from_date:
            conditions.append("ss.statement_month >= %(from_date)s")
            params["from_date"] = from_date
        if to_date:
            conditions.append("ss.statement_month <= %(to_date)s")
            params["to_date"] = to_date
        if codes:
            # UI sends business codes; the item Link column stores ids.
            conditions.append("si.product_code IN %(codes)s")
            params["codes"] = _resolve_product_filter(division, product_codes=codes) or ["__no_match__"]
        where = " AND ".join(conditions)

        group_col = "ss.hq" if hq_wise else "CONCAT(ss.stockist_code, ' – ', ss.stockist_name)"
        group_label = "HQ" if hq_wise else "Stockist"

        rows = frappe.db.sql(f"""
            SELECT ss.region, {group_col} AS group_key,
                   si.product_code, si.product_name,
                     SUM(si.sales_qty / IFNULL(NULLIF(si.conversion_factor, 0), 1)) AS qty,
                   SUM(si.sales_value_pts) AS value
            FROM `tabStockist Statement` ss
            INNER JOIN `tabStockist Statement Item` si
                ON si.parent = ss.name AND si.parenttype = 'Stockist Statement'
            WHERE {where}
            GROUP BY ss.region, group_key, si.product_code, si.product_name
            HAVING qty >= %(qty_filter)s
            ORDER BY value DESC
            LIMIT 5000
        """, dict(**params, qty_filter=qty_filter), as_dict=True)

    if sales_type in ("closing", "secondary"):
        # Statement items link by the Product Master id; show business codes.
        _apply_product_display_codes(rows)

    # Region (and the HQ group key, when grouping HQ-wise) are stored as Master
    # ids; show names only. Stockist grouping already carries the readable name.
    reg_lut = _name_lut("Region Master", "region_name")
    hq_lut = _name_lut("HQ Master", "hq_name") if hq_wise else None

    data = []
    for rank, r in enumerate(rows, 1):
        group_key = r.group_key or ""
        data.append({
            "rank": rank,
            "region": reg_lut.get(r.region, r.region) or "",
            "group_key": hq_lut.get(group_key, group_key) if hq_lut else group_key,
            "product_code": r.product_code,
            "product_name": r.product_name,
            "qty": flt(r.qty),
            "value": flt(r.value),
        })

    return {"success": True, "data": data, "group_label": group_label}


# ─────────────────────────────────────────────────────────────
# Report 6: Moving Trend PCPM Tracker
# ─────────────────────────────────────────────────────────────
@frappe.whitelist()
def get_ranking_pcpm_tracker(division=None, sales_type="secondary",
                              region=None, product_codes=None):
    """Monthly pivot (Apr–Mar) per product with PCPM = total / sanctioned_strength / active_months."""
    if not division:
        division = get_user_division()

    fy_start, fy_end, fy_label = _current_fy()

    codes = []
    if product_codes:
        codes = json.loads(product_codes) if isinstance(product_codes, str) else product_codes

    # Get sanctioned strength = sum of per_capita from active HQ Masters in the region
    strength_conditions = ["hm.division = %(division)s", "hm.status = 'Active'"]
    strength_params = {"division": division}
    # Sanctioned strength must span only the regions the caller may see (empty for an
    # admin with no region picked = every region, as before).
    _strength_rgn = []
    _scope_region_sql(_strength_rgn, strength_params, "tm.region", division, region,
                      key="g_strength_rgn")
    if _strength_rgn:
        strength_conditions.append("""hm.team IN (
            SELECT tm.name FROM `tabTeam Master` tm
            WHERE %s AND tm.status = 'Active')""" % " AND ".join(_strength_rgn))
    strength_where = " AND ".join(strength_conditions)

    strength_row = frappe.db.sql(f"""
        SELECT IFNULL(SUM(hm.per_capita), 0) AS total_strength
        FROM `tabHQ Master` hm
        WHERE {strength_where}
    """, strength_params, as_dict=True)
    sanctioned_strength = flt(strength_row[0].total_strength) if strength_row else 0

    # Get sales data
    if sales_type == "primary":
        conditions = ["ps.division = %(division)s", "ps.iscancelled = 0",
                       "ps.invoicedate >= %(from_date)s", "ps.invoicedate <= %(to_date)s"]
        params = {"division": division, "from_date": fy_start, "to_date": fy_end}
        # Primary Sales Data stores the region NAME, so match code-or-name.
        _scope_region_sql(conditions, params, "ps.region", division, region, include_names=True)
        if codes:
            conditions.append("ps.pcode IN %(codes)s")
            params["codes"] = codes
        where = " AND ".join(conditions)

        rows = frappe.db.sql(f"""
            SELECT ps.pcode AS product_code, ps.product AS product_name,
                   MONTH(ps.invoicedate) AS m,
                   SUM(ps.quantity) AS qty
            FROM `tabPrimary Sales Data` ps
            WHERE {where}
            GROUP BY ps.pcode, ps.product, MONTH(ps.invoicedate)
        """, params, as_dict=True)
    else:
        conditions = ["ss.division = %(division)s", "ss.docstatus IN (0, 1)",
                       "ss.statement_month >= %(from_date)s", "ss.statement_month <= %(to_date)s"]
        params = {"division": division, "from_date": fy_start, "to_date": fy_end}
        _scope_region_sql(conditions, params, "ss.region", division, region)
        if codes:
            # UI sends business codes; the item Link column stores ids.
            conditions.append("si.product_code IN %(codes)s")
            params["codes"] = _resolve_product_filter(division, product_codes=codes) or ["__no_match__"]
        where = " AND ".join(conditions)

        rows = frappe.db.sql(f"""
            SELECT si.product_code, si.product_name,
                   MONTH(ss.statement_month) AS m,
                     SUM(si.sales_qty / IFNULL(NULLIF(si.conversion_factor, 0), 1)) AS qty
            FROM `tabStockist Statement` ss
            INNER JOIN `tabStockist Statement Item` si
                ON si.parent = ss.name AND si.parenttype = 'Stockist Statement'
            WHERE {where}
            GROUP BY si.product_code, si.product_name, MONTH(ss.statement_month)
        """, params, as_dict=True)
        _apply_product_display_codes(rows)

    products = {}
    active_months = set()
    for r in rows:
        key = r.product_code
        if key not in products:
            products[key] = {
                "product_code": r.product_code,
                "product_name": r.product_name,
                "months": [0] * 12,
                "total": 0,
            }
        idx = _MONTH_MAP.get(r.m)
        if idx is not None:
            products[key]["months"][idx] += flt(r.qty)
            products[key]["total"] += flt(r.qty)
            active_months.add(idx)

    num_active_months = len(active_months) or 1
    data = []
    for p in products.values():
        p["average"] = round(p["total"] / 12, 2)
        p["pcpm"] = round(p["total"] / sanctioned_strength / num_active_months, 2) if sanctioned_strength else 0
        data.append(p)

    data.sort(key=lambda x: x["total"], reverse=True)

    return {"success": True, "data": data, "fy_label": fy_label,
            "month_labels": _MONTH_LABELS, "sanctioned_strength": sanctioned_strength,
            "active_months": num_active_months}


# ─────────────────────────────────────────────────────────────
# Ranking Reports – Excel Export
# ─────────────────────────────────────────────────────────────
@frappe.whitelist()
def export_ranking_report_excel(report_type, division=None, **kwargs):
    """Server-side Excel export for all 6 ranking report types."""
    if not division:
        division = get_user_division()

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active

    header_fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    group_fill = PatternFill(start_color="e8edf3", end_color="e8edf3", fill_type="solid")
    group_font = Font(bold=True, size=10)
    thin_border = Border(
        left=Side(style="thin", color="d1d5db"),
        right=Side(style="thin", color="d1d5db"),
        top=Side(style="thin", color="d1d5db"),
        bottom=Side(style="thin", color="d1d5db"),
    )

    def write_header_row(ws, row_num, headers):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    def write_data_row(ws, row_num, values):
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=v)
            cell.border = thin_border

    def write_title_rows(ws, title, subtitle=""):
        ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=14)
        if subtitle:
            ws.cell(row=2, column=1, value=subtitle).font = Font(size=11, italic=True)
        return 4

    ml = ["Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan", "Feb", "Mar"]

    if report_type == "moving_trend":
        ws.title = "Moving Trend"
        result = get_ranking_moving_trend_report(
            division, kwargs.get("sales_type", "secondary"),
            kwargs.get("criteria", "Region"),
            kwargs.get("from_date"), kwargs.get("to_date"),
            kwargs.get("region"), kwargs.get("zone"))
        data = result.get("data", [])
        criteria_label = result.get("criteria", "Region")
        row = write_title_rows(ws, f"Moving Trend Report – {division}",
                               result.get("fy_label", ""))
        headers = [criteria_label] + ml + ["Total", "Average"]
        write_header_row(ws, row, headers)
        row += 1
        for d in data:
            vals = [d["criteria_name"]] + d["months"] + [d["total"], d["average"]]
            write_data_row(ws, row, vals)
            row += 1

    elif report_type == "rupee_wise":
        ws.title = "Rupee Wise"
        result = get_ranking_rupee_wise_report(
            division, kwargs.get("sales_type", "secondary"),
            kwargs.get("value_condition", "gt"),
            kwargs.get("sale_value", 0),
            kwargs.get("from_date"), kwargs.get("to_date"))
        data = result.get("data", [])
        from_d = kwargs.get("from_date", "")
        to_d = kwargs.get("to_date", "")
        period_label = f"{from_d} to {to_d}" if from_d and to_d else ""
        row = write_title_rows(ws, f"Rupee Wise Report – {division}", period_label)
        headers = ["S.No", "Date", "Region", "HQ", "Stockist Code", "Stockist Name",
                    "Product Code", "Product", "Qty", "Rate", "Value", "Value ₹ Lakhs"]
        write_header_row(ws, row, headers)
        row += 1
        for d in data:
            write_data_row(ws, row, [d["sno"], d.get("date", ""), d.get("region", ""),
                                      d.get("hq", ""), d.get("stockist_code", ""),
                                      d.get("stockist_name", ""), d.get("product_code", ""),
                                      d.get("product_name", ""), d.get("qty", 0),
                                      d.get("rate", 0), d.get("value", 0),
                                      round(flt(d.get("value", 0)) / 100000, 2)])
            row += 1

    elif report_type == "productwise_topn":
        ws.title = "Product Ranking Top N"
        result = get_ranking_productwise_topn(
            division, kwargs.get("product_codes"),
            kwargs.get("top_n", 5),
            kwargs.get("from_date"), kwargs.get("to_date"),
            kwargs.get("sales_type", "secondary"))
        data = result.get("data", [])
        row = write_title_rows(ws, f"Productwise Ranking (Top N) – {division}", "")
        headers = ["Rank", "Product Code", "Product Name", "Total Qty",
                    "Total Value", "Value ₹ Lakhs", "Contribution %"]
        write_header_row(ws, row, headers)
        row += 1
        for d in data:
            write_data_row(ws, row, [d["rank"], d["product_code"], d["product_name"],
                                      d["total_qty"], d["total_value"],
                                      round(flt(d["total_value"]) / 100000, 2),
                                      d["contribution_pct"]])
            row += 1

    elif report_type == "productwise_all":
        ws.title = "Product Wise Ranking Sheet"
        result = get_ranking_productwise_all(
            division, kwargs.get("product_code"),
            kwargs.get("region"), kwargs.get("sales_type", "secondary"),
            kwargs.get("from_date"), kwargs.get("to_date"),
            kwargs.get("top_n", 2))
        data = result.get("data", [])
        st = kwargs.get("sales_type", "secondary")
        st_label = "Primary Sales" if st == "primary" else "Secondary Sales"
        row = write_title_rows(ws, f"Product Wise Ranking Sheet – {division}", st_label)
        headers = ["Product Code", "Pack", "Rank", "Headquarters", "Region", "Sales"]
        write_header_row(ws, row, headers)
        row += 1
        for d in data:
            write_data_row(ws, row, [d.get("product_code", ""), d.get("pack", ""),
                                      d.get("rank_roman", ""), d.get("hq", ""),
                                      d.get("region", ""), d.get("sales", 0)])
            row += 1

    elif report_type == "productwise_advanced":
        ws.title = "Product Ranking Advanced"
        result = get_ranking_productwise_advanced(
            division, kwargs.get("sales_type", "secondary"),
            kwargs.get("qty_filter", 0), kwargs.get("region"),
            kwargs.get("hq_wise", 0), kwargs.get("product_codes"),
            kwargs.get("from_date"), kwargs.get("to_date"))
        data = result.get("data", [])
        gl = result.get("group_label", "HQ")
        row = write_title_rows(ws, f"Productwise Ranking Advanced – {division}", "")
        headers = ["Rank", "Region", gl, "Product Code", "Product Name", "Qty", "Value", "Value ₹ Lakhs"]
        write_header_row(ws, row, headers)
        row += 1
        for d in data:
            write_data_row(ws, row, [d["rank"], d["region"], d["group_key"],
                                      d["product_code"], d["product_name"],
                                      d["qty"], d["value"],
                                      round(flt(d["value"]) / 100000, 2)])
            row += 1

    elif report_type == "pcpm_tracker":
        ws.title = "PCPM Tracker"
        result = get_ranking_pcpm_tracker(
            division, kwargs.get("sales_type", "secondary"),
            kwargs.get("region"), kwargs.get("product_codes"))
        data = result.get("data", [])
        row = write_title_rows(ws, f"PCPM Tracker – {division}",
                               f"{result.get('fy_label', '')}  |  Sanctioned Strength: {result.get('sanctioned_strength', 0)}")
        headers = ["Product Code", "Product Name"] + ml + ["Total", "Average", "PCPM"]
        write_header_row(ws, row, headers)
        row += 1
        for d in data:
            vals = [d["product_code"], d["product_name"]] + d["months"] + [d["total"], d["average"], d["pcpm"]]
            write_data_row(ws, row, vals)
            row += 1
    else:
        frappe.throw("Invalid report type")

    # Auto-fit column widths
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    xlsx_data = output.getvalue()

    filename = f"Ranking_Report_{report_type}_{division}.xlsx"
    frappe.local.response.filename = filename
    frappe.local.response.filecontent = xlsx_data
    frappe.local.response.type = "download"
    frappe.local.response.content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


# ═══════════════════════════════════════════════════════════════
# Shared helper – resolve a moving-trend scope to its HQ list + strength
# ═══════════════════════════════════════════════════════════════

def _resolve_entity_hqs(division, entity_type, entity_name):
    """Resolve a scope to (hq_list, sanctioned_strength, entity_display).

    entity_type: 'Organization' (whole division; entity_name ignored), 'Zone',
    'Region', 'Team' or 'HQ' — following Division → Zone → Region → Team → HQ.
    sanctioned_strength sums HQ per-capita (Team uses its sanctioned_strength).
    """
    hq_list = []
    entity_display = entity_name
    sanctioned_strength = 0.0

    if entity_type == "Organization":
        hqs = frappe.get_all("HQ Master",
                             filters={"status": "Active", "division": division},
                             fields=["name"])
        hq_list = [h.name for h in hqs]
        total_pc = frappe.db.sql(
            "SELECT COALESCE(SUM(per_capita), 0) FROM `tabHQ Master` "
            "WHERE status='Active' AND division=%s", (division,))
        sanctioned_strength = flt(total_pc[0][0]) if total_pc else 0
        entity_display = "All Regions (Organization)"

    elif entity_type == "HQ":
        hq_list = [entity_name]
        sanctioned_strength = flt(frappe.db.get_value("HQ Master", entity_name, "per_capita") or 0)
        entity_display = frappe.db.get_value("HQ Master", entity_name, "hq_name") or entity_name

    elif entity_type == "Team":
        hqs = frappe.get_all("HQ Master",
                             filters={"team": entity_name, "status": "Active", "division": division},
                             fields=["name", "hq_name"])
        hq_list = [h.name for h in hqs]
        hq_names = [h.hq_name for h in hqs]
        sanctioned_strength = flt(frappe.db.get_value("Team Master", entity_name, "sanctioned_strength") or 0)
        team_name = frappe.db.get_value("Team Master", entity_name, "team_name") or entity_name
        entity_display = f"{team_name} ({', '.join(hq_names)})" if hq_names else team_name

    elif entity_type == "Region":
        teams = frappe.get_all("Team Master",
                               filters={"region": entity_name, "status": "Active",
                                        "division": ["in", [division, "Both"]]},
                               fields=["name"])
        for t in teams:
            sub_hqs = frappe.get_all("HQ Master",
                                     filters={"team": t.name, "status": "Active", "division": division},
                                     fields=["name"])
            hq_list.extend([h.name for h in sub_hqs])
        total_pc = frappe.db.sql(
            "SELECT COALESCE(SUM(per_capita), 0) FROM `tabHQ Master` "
            "WHERE team IN (SELECT name FROM `tabTeam Master` WHERE region=%s AND status='Active') "
            "AND status='Active' AND division=%s", (entity_name, division))
        sanctioned_strength = flt(total_pc[0][0]) if total_pc else 0
        entity_display = frappe.db.get_value("Region Master", entity_name, "region_name") or entity_name

    elif entity_type == "Zone":
        regions = frappe.get_all("Region Master",
                                 filters={"zone": entity_name, "status": "Active",
                                          "division": ["in", [division, "Both"]]},
                                 fields=["name"])
        for reg in regions:
            sub_teams = frappe.get_all("Team Master",
                                       filters={"region": reg.name, "status": "Active",
                                                "division": ["in", [division, "Both"]]},
                                       fields=["name"])
            for t in sub_teams:
                sub_hqs = frappe.get_all("HQ Master",
                                         filters={"team": t.name, "status": "Active", "division": division},
                                         fields=["name"])
                hq_list.extend([h.name for h in sub_hqs])
        total_pc = frappe.db.sql(
            "SELECT COALESCE(SUM(hm.per_capita), 0) FROM `tabHQ Master` hm "
            "INNER JOIN `tabTeam Master` tm ON hm.team = tm.name "
            "INNER JOIN `tabRegion Master` rm ON tm.region = rm.name "
            "WHERE rm.zone=%s AND hm.status='Active' AND hm.division=%s", (entity_name, division))
        sanctioned_strength = flt(total_pc[0][0]) if total_pc else 0
        entity_display = frappe.db.get_value("Zone Master", entity_name, "zone_name") or entity_name

    # Confine non-admins to HQs inside their mapped regions, whatever entity was asked
    # for — an Organization/Zone scope must not span regions the user may not see. The
    # sanctioned strength is recomputed over the visible HQs so totals stay consistent.
    clamped = _clamp_hqs_to_allowed_regions(hq_list, division)
    if len(clamped) != len(hq_list):
        hq_list = clamped
        if hq_list:
            row = frappe.db.sql(
                "SELECT COALESCE(SUM(per_capita), 0) FROM `tabHQ Master` WHERE name IN %s",
                (tuple(hq_list),))
            sanctioned_strength = flt(row[0][0]) if row else 0.0
        else:
            sanctioned_strength = 0.0

    return hq_list, sanctioned_strength, entity_display


# ═══════════════════════════════════════════════════════════════
# SECONDARY SALES MOVING TREND REPORT  –  Portal API
# ═══════════════════════════════════════════════════════════════

@frappe.whitelist()
def get_secondary_sales_moving_trend(division=None, entity_type="Team",
                                     entity_name=None, financial_year=None, sales_mode="after_deduction",
                                     product_codes=None, product_group=None, product_category=None):
    """Product-wise monthly secondary sales pivot grouped by product category.

    Shows MAIN PRODUCTS, HOSPITAL PRODUCTS, NEW PRODUCTS sections with
    Value in Lakhs, Target (from HQ Yearly Target), Sanctioned Strength,
    Per Capita and achievement %.
    """
    if not division:
        division = get_user_division()
    if entity_type != "Organization" and not entity_name:
        return {"success": False, "message": f"{entity_type} is required"}

    from datetime import date
    today = date.today()

    # Determine financial year
    if financial_year:
        try:
            start_year = int(financial_year.split("-")[0])
        except Exception:
            start_year = today.year if today.month >= 4 else today.year - 1
    else:
        start_year = today.year if today.month >= 4 else today.year - 1
        financial_year = f"{start_year}-{str(start_year + 1)[2:]}"

    fy_start = f"{start_year}-04-01"
    fy_end = f"{start_year + 1}-03-31"
    fy_label = f"Apr {str(start_year)[2:]} to Mar {str(start_year + 1)[2:]}"

    # ── Resolve entity → list of HQs ──
    hq_list, sanctioned_strength, entity_display = _resolve_entity_hqs(division, entity_type, entity_name)

    if not hq_list:
        return {"success": True, "sections": [],
                "entity_display": entity_display, "entity_type": entity_type,
                "sanctioned_strength": sanctioned_strength, "fy_label": fy_label}

    # ── Get all products for this division ──
    products = frappe.db.sql(
        "SELECT name, product_code, product_name, pack, category, pts "
        "FROM `tabProduct Master` WHERE division=%s AND status='Active' "
        "ORDER BY category, COALESCE(sequence, 9999), product_code",
        (division,), as_dict=True
    )
    # Optional product / group / category narrowing (intersection).
    # _resolve_product_filter returns Product Master ids, so match on name.
    _pf = _resolve_product_filter(division, product_codes, product_group, product_category)
    if _pf is not None:
        _allowed = set(_pf)
        products = [p for p in products if p.name in _allowed]

    # ── Get secondary sales (Stockist Statement Items) ──
    hq_placeholders = ", ".join(["%s"] * len(hq_list))
    if sales_mode == "before_deduction":
        _qty_expr = "(si.sales_qty + si.free_qty) / IFNULL(NULLIF(si.conversion_factor, 0), 1)"
    else:
        _qty_expr = "(si.sales_qty + si.free_qty - IFNULL(si.free_qty_scheme, 0)) / IFNULL(NULLIF(si.conversion_factor, 0), 1)"
    _val_expr = f"({_qty_expr}) * IFNULL(si.pts, 0)"
    sec_rows = frappe.db.sql(f"""
        SELECT si.product_code, si.pack,
               MONTH(ss.statement_month) AS m,
             SUM({_qty_expr}) AS qty,
             SUM({_val_expr}) AS value
        FROM `tabStockist Statement` ss
        INNER JOIN `tabStockist Statement Item` si
            ON si.parent = ss.name AND si.parenttype = 'Stockist Statement'
        WHERE ss.division = %s AND ss.docstatus IN (0, 1)
              AND ss.hq IN ({hq_placeholders})
              AND ss.statement_month BETWEEN %s AND %s
        GROUP BY si.product_code, si.pack, MONTH(ss.statement_month)
    """, [division] + hq_list + [fy_start, fy_end], as_dict=True)
    # Items link by the Product Master id; the pivot below is keyed by the
    # business code (matching the master list's product_code).
    _apply_product_display_codes(sec_rows)

    # ── Pivot data: product_code → 12-month array ──
    month_map = {4: 0, 5: 1, 6: 2, 7: 3, 8: 4, 9: 5,
                 10: 6, 11: 7, 12: 8, 1: 9, 2: 10, 3: 11}
    product_data = {}
    for r in sec_rows:
        pc = r.product_code
        if pc not in product_data:
            product_data[pc] = {"months_qty": [0] * 12, "months_val": [0.0] * 12}
        idx = month_map.get(r.m)
        if idx is not None:
            product_data[pc]["months_qty"][idx] += flt(r.qty)
            product_data[pc]["months_val"][idx] += flt(r.value)

    # ── Get HQ target value (in Lakhs) for the FY ──
    target_value = 0.0
    target_rows = frappe.db.sql("""
        SELECT COALESCE(SUM(ti.yearly_total), 0) AS total_target
        FROM `tabHQ Yearly Target` yt
        INNER JOIN `tabHQ Target Item` ti ON ti.parent = yt.name
        WHERE yt.docstatus = 1 AND yt.division = %s AND yt.financial_year = %s
              AND ti.hq IN ({hq_ph})
    """.format(hq_ph=hq_placeholders), [division, financial_year] + hq_list, as_dict=True)
    if target_rows:
        target_value = flt(target_rows[0].total_target)

    # ── Build sections by category ──
    category_order = [
        ("Main Products", "MAIN PRODUCTS"),
        ("Hospital Products", "HOSPITAL PRODUCTS"),
        ("New Products", "NEW PRODUCTS"),
    ]

    # Count months with actual data
    active_months = 0
    for i in range(12):
        has_data = any(pd["months_qty"][i] > 0 for pd in product_data.values())
        if has_data:
            active_months += 1
    active_months = max(active_months, 1)

    sections = []
    grand_total_val = 0.0

    # First pass: compute totals per section
    section_totals = {}
    for cat_key, cat_label in category_order:
        cat_products = [p for p in products if (p.category or "").lower() == cat_key.lower()]
        total_val = sum(sum(product_data.get(p.product_code, {"months_val": [0.0] * 12})["months_val"])
                        for p in cat_products)
        section_totals[cat_key] = total_val
        grand_total_val += total_val

    for cat_key, cat_label in category_order:
        cat_products = [p for p in products if (p.category or "").lower() == cat_key.lower()]
        if not cat_products:
            continue

        section_months_val = [0.0] * 12
        product_rows = []

        for p in cat_products:
            pc = p.product_code
            pd = product_data.get(pc, {"months_qty": [0] * 12, "months_val": [0.0] * 12})

            total_qty = sum(pd["months_qty"])
            avg_qty = round(total_qty / active_months) if total_qty else 0
            per_capita = round(avg_qty / sanctioned_strength) if sanctioned_strength else 0

            for i in range(12):
                section_months_val[i] += pd["months_val"][i]

            product_rows.append({
                "code": pc,
                "pack": p.pack or "",
                "target": 0,
                "months": [int(round(x)) for x in pd["months_qty"]],
                "total": int(round(total_qty)),
                "average": avg_qty,
                "per_capita": per_capita,
            })

        # Section value in lakhs
        section_total_val = section_totals[cat_key]
        section_months_lakhs = [round(v / 100000, 2) for v in section_months_val]
        section_total_lakhs = round(section_total_val / 100000, 2)
        section_avg_lakhs = round(section_total_lakhs / active_months, 2)
        section_pc_lakhs = round(section_avg_lakhs / sanctioned_strength, 2) if sanctioned_strength else 0

        # Target shown only for Main Products
        cat_target = round(target_value, 2) if (cat_key == "Main Products" and target_value > 0) else 0.0
        cat_pct = round((section_total_lakhs / cat_target) * 100) if cat_target else 0.0

        sections.append({
            "label": cat_label,
            "category": cat_key,
            "products": product_rows,
            "value_in_lakhs": {
                "target": cat_target,
                "months": section_months_lakhs,
                "total": section_total_lakhs,
                "average": section_avg_lakhs,
                "per_capita": section_pc_lakhs,
                "pct": cat_pct,
            }
        })

    return {
        "success": True,
        "entity_type": entity_type,
        "entity_display": entity_display,
        "sanctioned_strength": sanctioned_strength,
        "fy_label": fy_label,
        "financial_year": financial_year,
        "active_months": active_months,
        "sections": sections,
        "month_labels": _MONTH_LABELS,
        "sales_mode": sales_mode,
    }


# ═══════════════════════════════════════════════════════════════
# REPORT 10 – PRIMARY SALES MOVING TREND
# Same shape as Report 7 (Secondary Sales Moving Trend) but reads from
# `tabPrimary Sales Data`. No before/after-scheme split — primary sales are
# raw billing figures and the scheme deduction concept does not apply.
# ═══════════════════════════════════════════════════════════════

@frappe.whitelist()
def get_primary_sales_moving_trend(division=None, entity_type="Team",
                                   entity_name=None, financial_year=None,
                                   product_codes=None):
    """Report 10 – Primary Sales Moving Trend.

    Mirrors get_secondary_sales_moving_trend's output shape (sections of products
    with monthly quantities and a Value-in-Lakhs row) so the same renderer can be
    reused on the front-end. Reads Primary Sales Data (iscancelled=0).
    """
    if not division:
        division = get_user_division()
    if entity_type != "Organization" and not entity_name:
        return {"success": False, "message": f"{entity_type} is required"}

    from datetime import date
    today = date.today()

    if financial_year:
        try:
            start_year = int(financial_year.split("-")[0])
        except Exception:
            start_year = today.year if today.month >= 4 else today.year - 1
    else:
        start_year = today.year if today.month >= 4 else today.year - 1
        financial_year = f"{start_year}-{str(start_year + 1)[2:]}"

    fy_start = f"{start_year}-04-01"
    fy_end = f"{start_year + 1}-03-31"
    fy_label = f"Apr {str(start_year)[2:]} to Mar {str(start_year + 1)[2:]}"

    # ── Resolve entity → list of HQs (shared resolver) ──
    hq_list, sanctioned_strength, entity_display = _resolve_entity_hqs(division, entity_type, entity_name)

    if not hq_list:
        return {"success": True, "sections": [],
                "entity_display": entity_display, "entity_type": entity_type,
                "sanctioned_strength": sanctioned_strength, "fy_label": fy_label,
                "month_labels": _MONTH_LABELS}

    # ── Products & optional code filter ──
    products = frappe.db.sql(
        "SELECT product_code, product_name, pack, category, pts "
        "FROM `tabProduct Master` WHERE division=%s AND status='Active' "
        "ORDER BY category, COALESCE(sequence, 9999), product_code",
        (division,), as_dict=True
    )
    pcodes = _normalise_code_list(product_codes)
    if pcodes:
        allowed = set(pcodes)
        products = [p for p in products if p.product_code in allowed]

    # ── Primary Sales by product × month, scoped to the HQs of the chosen entity ──
    # Resolve HQs → stockists (Primary Sales has no direct hq column).
    stockist_rows = frappe.db.sql(
        f"""SELECT name FROM `tabStockist Master`
              WHERE division=%s AND status='Active'
                AND hq IN ({", ".join(["%s"] * len(hq_list))})""",
        [division] + hq_list,
    )
    stockist_codes = [r[0] for r in stockist_rows]
    if not stockist_codes:
        return {"success": True, "sections": [],
                "entity_display": entity_display, "entity_type": entity_type,
                "sanctioned_strength": sanctioned_strength, "fy_label": fy_label,
                "month_labels": _MONTH_LABELS}

    sp_ph = ", ".join(["%s"] * len(stockist_codes))
    pcode_clause = ""
    pcode_args = []
    if pcodes:
        pp = ", ".join(["%s"] * len(pcodes))
        pcode_clause = f" AND pcode IN ({pp})"
        pcode_args = pcodes

    pri_rows = frappe.db.sql(f"""
        SELECT pcode AS product_code,
               MONTH(invoicedate) AS m,
               SUM(quantity) AS qty,
               SUM(ptsvalue) AS value
        FROM `tabPrimary Sales Data`
        WHERE division = %s AND iscancelled = 0
          AND stockist_code IN ({sp_ph})
          AND invoicedate BETWEEN %s AND %s{pcode_clause}
        GROUP BY pcode, MONTH(invoicedate)
    """, [division] + stockist_codes + [fy_start, fy_end] + pcode_args, as_dict=True)

    # ── Pivot ──
    month_map = {4: 0, 5: 1, 6: 2, 7: 3, 8: 4, 9: 5,
                 10: 6, 11: 7, 12: 8, 1: 9, 2: 10, 3: 11}
    product_data = {}
    for r in pri_rows:
        pc = r.product_code
        if pc not in product_data:
            product_data[pc] = {"months_qty": [0] * 12, "months_val": [0.0] * 12}
        idx = month_map.get(r.m)
        if idx is not None:
            product_data[pc]["months_qty"][idx] += flt(r.qty)
            product_data[pc]["months_val"][idx] += flt(r.value)

    # ── HQ target (in ₹) for the FY ──
    target_value = 0.0
    target_rows = frappe.db.sql(f"""
        SELECT COALESCE(SUM(ti.yearly_total), 0) AS total_target
        FROM `tabHQ Yearly Target` yt
        INNER JOIN `tabHQ Target Item` ti ON ti.parent = yt.name
        WHERE yt.docstatus = 1 AND yt.division = %s AND yt.financial_year = %s
              AND ti.hq IN ({", ".join(["%s"] * len(hq_list))})
    """, [division, financial_year] + hq_list, as_dict=True)
    if target_rows:
        target_value = flt(target_rows[0].total_target)

    # ── Count active months (months with any sales) ──
    active_months = 0
    for i in range(12):
        if any(pd["months_qty"][i] > 0 for pd in product_data.values()):
            active_months += 1
    active_months = max(active_months, 1)

    # ── Build sections by category ──
    category_order = [
        ("Main Products", "MAIN PRODUCTS"),
        ("Hospital Products", "HOSPITAL PRODUCTS"),
        ("New Products", "NEW PRODUCTS"),
    ]

    section_totals = {}
    grand_total_val = 0.0
    for cat_key, _ in category_order:
        cat_products = [p for p in products if (p.category or "").lower() == cat_key.lower()]
        total_val = sum(sum(product_data.get(p.product_code, {"months_val": [0.0] * 12})["months_val"])
                        for p in cat_products)
        section_totals[cat_key] = total_val
        grand_total_val += total_val

    sections = []
    for cat_key, cat_label in category_order:
        cat_products = [p for p in products if (p.category or "").lower() == cat_key.lower()]
        if not cat_products:
            continue
        section_months_val = [0.0] * 12
        product_rows = []
        for p in cat_products:
            pc = p.product_code
            pd = product_data.get(pc, {"months_qty": [0] * 12, "months_val": [0.0] * 12})
            total_qty = sum(pd["months_qty"])
            avg_qty = round(total_qty / active_months) if total_qty else 0
            per_capita = round(avg_qty / sanctioned_strength) if sanctioned_strength else 0
            for i in range(12):
                section_months_val[i] += pd["months_val"][i]
            product_rows.append({
                "code": pc, "pack": p.pack or "",
                "target": 0,
                "months": [int(round(x)) for x in pd["months_qty"]],
                "total": int(round(total_qty)),
                "average": avg_qty,
                "per_capita": per_capita,
            })

        section_total_val = section_totals[cat_key]
        section_months_lakhs = [round(v / 100000, 2) for v in section_months_val]
        section_total_lakhs = round(section_total_val / 100000, 2)
        section_avg_lakhs = round(section_total_lakhs / active_months, 2)
        section_pc_lakhs = round(section_avg_lakhs / sanctioned_strength, 2) if sanctioned_strength else 0

        # Target shown only for Main Products
        cat_target = round(target_value, 2) if (cat_key == "Main Products" and target_value > 0) else 0.0
        cat_pct = round((section_total_lakhs / cat_target) * 100) if cat_target else 0.0

        sections.append({
            "label": cat_label,
            "category": cat_key,
            "products": product_rows,
            "value_in_lakhs": {
                "target": cat_target,
                "months": section_months_lakhs,
                "total": section_total_lakhs,
                "average": section_avg_lakhs,
                "per_capita": section_pc_lakhs,
                "pct": cat_pct,
            }
        })

    return {
        "success": True,
        "entity_type": entity_type,
        "entity_display": entity_display,
        "sanctioned_strength": sanctioned_strength,
        "fy_label": fy_label,
        "financial_year": financial_year,
        "active_months": active_months,
        "sections": sections,
        "month_labels": _MONTH_LABELS,
    }


# ═══════════════════════════════════════════════════════════════
# REPORT 16 – ORGANISATIONAL SALES REPORT (REGION WISE)
# Product rows × Region columns, grouped by Zone with a per-Zone total column
# and a trailing Organization Total column. The first body row is "Value In
# Lakhs" (monetary total per region / zone / organisation); product cells are
# quantities. Works for both Primary (`tabPrimary Sales Data`) and Secondary
# (`tabStockist Statement`) sales over a date range.
# ═══════════════════════════════════════════════════════════════

@frappe.whitelist()
def get_organizational_sales_report(division=None, sales_type="primary",
                                    from_date=None, to_date=None,
                                    product_codes=None, sales_mode="after_deduction"):
    """Report 16 – Organisational Sales Report (Region wise).

    Rows = products (code + pack); columns = regions grouped by zone, with a
    per-zone total column and a trailing Organization Total column. The first
    body row is "Value In Lakhs". Cell values are quantities.

    sales_type : 'primary' (tabPrimary Sales Data, iscancelled=0)
                 'secondary' (tabStockist Statement, docstatus IN (0,1))
    sales_mode : 'after_deduction' | 'before_deduction' (secondary only)
    """
    if not division:
        division = get_user_division()

    sales_type = (sales_type or "primary").lower()
    is_secondary = (sales_type == "secondary")

    # ── Column scaffold: zones → regions (all active regions in the division) ──
    zone_rows = frappe.db.sql(
        "SELECT name, zone_name FROM `tabZone Master` "
        "WHERE status='Active' AND division IN (%s, 'Both') ORDER BY name",
        (division,), as_dict=True)
    region_rows = frappe.db.sql(
        "SELECT name, region_name, IFNULL(zone,'') AS zone FROM `tabRegion Master` "
        "WHERE status='Active' AND division IN (%s, 'Both') ORDER BY name",
        (division,), as_dict=True)

    # Confine non-admins to their mapped regions — every column/total below derives
    # from region_rows, so filtering here scopes the whole report.
    _allowed = _allowed_region_codes_or_all(division)
    if _allowed is not None:
        _aset = set(_allowed)
        region_rows = [r for r in region_rows if r.name in _aset]

    zone_name_map = {z.name: (z.zone_name or z.name) for z in zone_rows}

    # Region value resolver: sales tables are inconsistent — Stockist Statement
    # stores the region *code* (R0xxx) while Primary Sales Data stores the region
    # *name* ("Chennai"). Map both forms → canonical Region Master code.
    region_lookup = {}
    for r in region_rows:
        region_lookup[r.name] = r.name
        if r.region_name:
            region_lookup[r.region_name] = r.name
            region_lookup[r.region_name.strip().lower()] = r.name

    def _canon_region(val):
        if not val:
            return None
        return region_lookup.get(val) or region_lookup.get(str(val).strip().lower())

    # Group regions under their zone, preserving zone order then region order.
    regions_by_zone = {}
    for r in region_rows:
        regions_by_zone.setdefault(r.zone or "", []).append(
            {"code": r.name, "name": r.region_name or r.name})

    # Ordered zone codes: configured zones first, then any zone code referenced
    # by a region but missing from Zone Master, then the ungrouped bucket last.
    zone_order = [z.name for z in zone_rows if z.name in regions_by_zone]
    for zc in regions_by_zone:
        if zc and zc not in zone_order:
            zone_order.append(zc)

    UNGROUPED = "__none__"
    zones_out = []
    region_to_zone = {}
    for zc in zone_order:
        regs = regions_by_zone.get(zc, [])
        if not regs:
            continue
        for rg in regs:
            region_to_zone[rg["code"]] = zc
        zones_out.append({"code": zc, "name": zone_name_map.get(zc, zc), "regions": regs})
    if "" in regions_by_zone:
        regs = regions_by_zone[""]
        for rg in regs:
            region_to_zone[rg["code"]] = UNGROUPED
        zones_out.append({"code": UNGROUPED, "name": "Other", "regions": regs})

    all_region_codes = [rg["code"] for z in zones_out for rg in z["regions"]]
    valid_regions = set(all_region_codes)

    # ── Products (optionally narrowed by the multi-select) ──
    products = frappe.db.sql(
        "SELECT product_code, product_name, pack FROM `tabProduct Master` "
        "WHERE division=%s AND status='Active' "
        "ORDER BY COALESCE(sequence, 9999), product_code",
        (division,), as_dict=True)
    pcodes = _normalise_code_list(product_codes)
    if pcodes:
        allowed = set(pcodes)
        products = [p for p in products if p.product_code in allowed]

    # ── Sales rows: region × product → qty, value ──
    if is_secondary:
        if sales_mode == "before_deduction":
            _qty_expr = "(si.sales_qty + si.free_qty) / IFNULL(NULLIF(si.conversion_factor, 0), 1)"
        else:
            _qty_expr = "(si.sales_qty + si.free_qty - IFNULL(si.free_qty_scheme, 0)) / IFNULL(NULLIF(si.conversion_factor, 0), 1)"
        _val_expr = f"({_qty_expr}) * IFNULL(si.pts, 0)"
        conds = ["ss.division = %(division)s", "ss.docstatus IN (0, 1)", "IFNULL(ss.region, '') <> ''"]
        params = {"division": division}
        if from_date:
            conds.append("ss.statement_month >= %(from_date)s"); params["from_date"] = from_date
        if to_date:
            conds.append("ss.statement_month <= %(to_date)s"); params["to_date"] = to_date
        if pcodes:
            # UI sends business codes; the item Link column stores ids.
            resolved_pks = _resolve_product_filter(division, product_codes=pcodes) or ["__no_match__"]
            ph = ", ".join([f"%(_pc{i})s" for i in range(len(resolved_pks))])
            conds.append(f"si.product_code IN ({ph})")
            for i, c in enumerate(resolved_pks):
                params[f"_pc{i}"] = c
        where = " AND ".join(conds)
        sales_rows = frappe.db.sql(f"""
            SELECT ss.region AS region, si.product_code AS product_code,
                   SUM({_qty_expr}) AS qty, SUM({_val_expr}) AS value
            FROM `tabStockist Statement` ss
            INNER JOIN `tabStockist Statement Item` si
                ON si.parent = ss.name AND si.parenttype = 'Stockist Statement'
            WHERE {where}
            GROUP BY ss.region, si.product_code
        """, params, as_dict=True)
        # Items link by the Product Master id; the pivot is keyed by business code.
        _apply_product_display_codes(sales_rows)
    else:
        conds = ["psd.division = %(division)s", "psd.iscancelled = 0", "IFNULL(psd.region, '') <> ''"]
        params = {"division": division}
        if from_date:
            conds.append("psd.invoicedate >= %(from_date)s"); params["from_date"] = from_date
        if to_date:
            conds.append("psd.invoicedate <= %(to_date)s"); params["to_date"] = to_date
        if pcodes:
            ph = ", ".join([f"%(_pc{i})s" for i in range(len(pcodes))])
            conds.append(f"psd.pcode IN ({ph})")
            for i, c in enumerate(pcodes):
                params[f"_pc{i}"] = c
        where = " AND ".join(conds)
        sales_rows = frappe.db.sql(f"""
            SELECT psd.region AS region, psd.pcode AS product_code,
                   SUM(psd.quantity) AS qty, SUM(psd.ptsvalue) AS value
            FROM `tabPrimary Sales Data` psd
            WHERE {where}
            GROUP BY psd.region, psd.pcode
        """, params, as_dict=True)

    # ── Pivot: product_code → region_code → {qty, value} ──
    pdata = {}
    for r in sales_rows:
        rc = _canon_region(r.region)
        if rc is None or rc not in valid_regions:
            continue  # region outside this division's active scaffold
        cell = pdata.setdefault(r.product_code, {})
        c = cell.setdefault(rc, {"qty": 0.0, "value": 0.0})
        c["qty"] += flt(r.qty)
        c["value"] += flt(r.value)

    LAKH = 100000.0

    # ── Build product rows (only products that have sales data) ──
    product_rows = []
    vil_regions = {rc: 0.0 for rc in all_region_codes}  # value accumulators (₹)
    for p in products:
        regcells = pdata.get(p.product_code)
        if not regcells:
            continue
        cells = {}
        zone_tot = {}
        org_qty = 0.0
        for rc in all_region_codes:
            c = regcells.get(rc)
            q = flt(c["qty"]) if c else 0.0
            v = flt(c["value"]) if c else 0.0
            if q:
                cells[rc] = round(q, 2)
            org_qty += q
            zc = region_to_zone.get(rc, UNGROUPED)
            zone_tot[zc] = zone_tot.get(zc, 0.0) + q
            vil_regions[rc] += v
        product_rows.append({
            "code": p.product_code,
            "pack": p.pack or "",
            "cells": cells,
            "zone_totals": {z: round(q, 2) for z, q in zone_tot.items() if q},
            "org_total": round(org_qty, 2),
        })

    # ── Value In Lakhs row (per region / zone / organisation) ──
    vil_zone = {}
    vil_org = 0.0
    for rc in all_region_codes:
        zc = region_to_zone.get(rc, UNGROUPED)
        vil_zone[zc] = vil_zone.get(zc, 0.0) + vil_regions[rc]
        vil_org += vil_regions[rc]
    value_in_lakhs = {
        "regions": {rc: round(v / LAKH, 2) for rc, v in vil_regions.items()},
        "zone_totals": {zc: round(v / LAKH, 2) for zc, v in vil_zone.items()},
        "org_total": round(vil_org / LAKH, 2),
    }

    # ── Labels ──
    from frappe.utils import getdate

    def _dmy(d):
        try:
            return getdate(d).strftime("%d/%m/%Y")
        except Exception:
            return str(d or "")

    type_label = "Secondary" if is_secondary else "Primary"
    period_label = f"{_dmy(from_date)} To {_dmy(to_date)}" if (from_date and to_date) else ""

    return {
        "success": True,
        "sales_type": sales_type,
        "type_label": type_label,
        "title": f"Organization {type_label} Sales",
        "from_date": from_date or "",
        "to_date": to_date or "",
        "period_label": period_label,
        "zones": zones_out,
        "region_codes": all_region_codes,
        "value_in_lakhs": value_in_lakhs,
        "products": product_rows,
        "sales_mode": sales_mode if is_secondary else None,
    }


# ═══════════════════════════════════════════════════════════════
# REPORT 15 – GYNAE REPORT
# Secondary-sales moving trend restricted to product_group = 'Gynae', for the
# whole Organization or a Zone/Region/Team/HQ scope. Brand rows show monthly
# quantities; summary rows give Values-in-lakh, PCPM-in-lakh (value/strength)
# and Avg-per-Dr (₹ = value/strength). Strength = sanctioned field-force.
# ═══════════════════════════════════════════════════════════════

@frappe.whitelist()
def get_gynae_report(division=None, entity_type="Organization", entity_name=None,
                     financial_year=None, sales_mode="after_deduction"):
    """Report 15 – Gynae Report (Gynae-group secondary sales moving trend)."""
    if not division:
        division = get_user_division()

    from datetime import date
    today = date.today()
    if financial_year:
        try:
            start_year = int(financial_year.split("-")[0])
        except Exception:
            start_year = today.year if today.month >= 4 else today.year - 1
    else:
        start_year = today.year if today.month >= 4 else today.year - 1
        financial_year = f"{start_year}-{str(start_year + 1)[2:]}"
    fy_start = f"{start_year}-04-01"
    fy_end = f"{start_year + 1}-03-31"
    fy_label = f"Apr {str(start_year)[2:]} to Mar {str(start_year + 1)[2:]}"

    hq_list, sanctioned_strength, entity_display = _resolve_entity_hqs(division, entity_type, entity_name)

    def _empty():
        return {
            "success": True, "entity_type": entity_type, "entity_display": entity_display,
            "strength": flt(sanctioned_strength), "fy_label": fy_label,
            "financial_year": financial_year, "month_labels": _MONTH_LABELS,
            "brands": [], "values_lakh": [0.0] * 12, "pcpm_lakh": [0.0] * 12,
            "avg_per_dr": [0] * 12, "sales_mode": sales_mode,
        }

    if not hq_list:
        return _empty()

    # ── Gynae products only ──
    products = frappe.db.sql(
        "SELECT name, product_code, product_name, pack FROM `tabProduct Master` "
        "WHERE division=%s AND status='Active' AND LOWER(IFNULL(product_group,''))='gynae' "
        "ORDER BY COALESCE(sequence, 9999), product_code",
        (division,), as_dict=True)
    if not products:
        return _empty()

    # Statement items link by the Product Master id.
    pcodes = [p.name for p in products]
    pc_ph = ", ".join(["%s"] * len(pcodes))
    hq_ph = ", ".join(["%s"] * len(hq_list))

    if sales_mode == "before_deduction":
        _qty_expr = "(si.sales_qty + si.free_qty) / IFNULL(NULLIF(si.conversion_factor, 0), 1)"
    else:
        _qty_expr = "(si.sales_qty + si.free_qty - IFNULL(si.free_qty_scheme, 0)) / IFNULL(NULLIF(si.conversion_factor, 0), 1)"
    _val_expr = f"({_qty_expr}) * IFNULL(si.pts, 0)"

    rows = frappe.db.sql(f"""
        SELECT si.product_code, MONTH(ss.statement_month) AS m,
               SUM({_qty_expr}) AS qty, SUM({_val_expr}) AS val
          FROM `tabStockist Statement` ss
    INNER JOIN `tabStockist Statement Item` si
            ON si.parent = ss.name AND si.parenttype = 'Stockist Statement'
         WHERE ss.division = %s AND ss.docstatus IN (0, 1)
           AND ss.hq IN ({hq_ph})
           AND si.product_code IN ({pc_ph})
           AND ss.statement_month BETWEEN %s AND %s
      GROUP BY si.product_code, MONTH(ss.statement_month)
    """, [division] + hq_list + pcodes + [fy_start, fy_end], as_dict=True)
    # Rows come back keyed by id; the brand pivot uses business codes.
    _apply_product_display_codes(rows)

    month_map = {4: 0, 5: 1, 6: 2, 7: 3, 8: 4, 9: 5, 10: 6, 11: 7, 12: 8, 1: 9, 2: 10, 3: 11}
    pdata = {}
    months_val = [0.0] * 12
    for r in rows:
        idx = month_map.get(r.m)
        if idx is None:
            continue
        pdata.setdefault(r.product_code, [0.0] * 12)
        pdata[r.product_code][idx] += flt(r.qty)
        months_val[idx] += flt(r.val)

    brands = []
    for p in products:
        mq = pdata.get(p.product_code, [0.0] * 12)
        brands.append({
            "code": p.product_code,
            "name": p.product_name or p.product_code,
            "pack": p.pack or "",
            "months": [int(round(x)) for x in mq],
            "total": int(round(sum(mq))),
        })

    LAKH = 100000.0
    strength = flt(sanctioned_strength)
    values_lakh = [round(v / LAKH, 2) for v in months_val]
    pcpm_lakh = [round((v / LAKH) / strength, 2) if strength else 0.0 for v in months_val]
    # Avg per Dr. (Rs.) = respective PCPM in lakh × 2000
    avg_per_dr = [int(round(p * 2000)) for p in pcpm_lakh]

    return {
        "success": True,
        "entity_type": entity_type,
        "entity_display": entity_display,
        "strength": strength,
        "fy_label": fy_label,
        "financial_year": financial_year,
        "month_labels": _MONTH_LABELS,
        "brands": brands,
        "values_lakh": values_lakh,
        "pcpm_lakh": pcpm_lakh,
        "avg_per_dr": avg_per_dr,
        "sales_mode": sales_mode,
    }


# ═══════════════════════════════════════════════════════════════
# REGION-WISE STOCKIST SECONDARY SALES MOVING TREND
# Format: Rows = Stockists; Cols = Opening | Apr..Mar | Total Value | Closing
# Values in ₹ Lakhs; grouped by Team within Region
# Includes Draft (docstatus IN (0,1)) statements
# ═══════════════════════════════════════════════════════════════

@frappe.whitelist()
def get_region_wise_stockist_moving_trend(division=None, region=None, financial_year=None, sales_mode="after_deduction",
                                          product_codes=None, product_group=None, product_category=None):
    """
    Region-wise stockist sales moving trend with opening and closing values.
    Matches the Orissa Secondary Sales Excel format:
      Columns: Stockist | HQ | Current Opening | Apr..Mar (12 months) | Total Sales Value | Current Closing
    Values in ₹ Lakhs.  Grouped by Team within the Region.
    Includes both Draft and Submitted statements (docstatus IN (0,1)).
    """
    if not division:
        division = get_user_division()
    if not region:
        return {"success": False, "message": "Region is required"}
    # Non-admins may only run this for a region they are mapped to.
    _allowed = _allowed_region_codes_or_all(division)
    if _allowed is not None and region not in set(_allowed):
        return {"success": False, "message": "You are not permitted to view this region."}

    from datetime import date
    today = date.today()

    # Determine financial year
    if financial_year:
        try:
            start_year = int(financial_year.split("-")[0])
        except Exception:
            start_year = today.year if today.month >= 4 else today.year - 1
    else:
        start_year = today.year if today.month >= 4 else today.year - 1
        financial_year = f"{start_year}-{str(start_year + 1)[2:]}"

    fy_start = f"{start_year}-04-01"
    fy_end   = f"{start_year + 1}-03-31"
    fy_label = f"Apr {str(start_year)[2:]} to Mar {str(start_year + 1)[2:]}"

    # ── Month headers (FY order Apr=0 … Mar=11) ──
    month_map = {4: 0, 5: 1, 6: 2, 7: 3, 8: 4, 9: 5,
                 10: 6, 11: 7, 12: 8, 1: 9, 2: 10, 3: 11}
    month_labels = [
        f"Apr-{str(start_year)[2:]}", f"May-{str(start_year)[2:]}",
        f"Jun-{str(start_year)[2:]}", f"Jul-{str(start_year)[2:]}",
        f"Aug-{str(start_year)[2:]}", f"Sep-{str(start_year)[2:]}",
        f"Oct-{str(start_year)[2:]}", f"Nov-{str(start_year)[2:]}",
        f"Dec-{str(start_year)[2:]}", f"Jan-{str(start_year + 1)[2:]}",
        f"Feb-{str(start_year + 1)[2:]}", f"Mar-{str(start_year + 1)[2:]}",
    ]

    # ── Get all active stockists in this region ──
    stockists = frappe.db.sql("""
        SELECT sm.name AS code, COALESCE(sm.stockist_code, sm.name) AS display_code,
               sm.stockist_name, sm.hq,
               COALESCE(hm.hq_name, sm.hq, '') AS hq_name,
               COALESCE(hm.team, '') AS team,
               COALESCE(tm.team_name, hm.team, '') AS team_name
        FROM `tabStockist Master` sm
        LEFT JOIN `tabHQ Master` hm ON hm.name = sm.hq AND hm.division = %(division)s
        LEFT JOIN `tabTeam Master` tm ON tm.name = hm.team AND tm.division IN (%(division)s, 'Both')
        WHERE sm.division = %(division)s
          AND sm.region = %(region)s
          AND sm.status = 'Active'
        ORDER BY hm.team, hq_name, sm.stockist_name
    """, {"division": division, "region": region}, as_dict=True)

    if not stockists:
        return {"success": True, "data": [], "teams": [], "fy_label": fy_label,
                "month_labels": month_labels, "region": region}

    stockist_codes = [s.code for s in stockists]
    placeholders = ", ".join(["%s"] * len(stockist_codes))

    # Optional product / group / category narrowing (intersection)
    _pf = _resolve_product_filter(division, product_codes, product_group, product_category)
    prod_clause = ""
    prod_codes_p = []
    if _pf is not None:
        prod_clause = " AND si.product_code IN (" + ", ".join(["%s"] * len(_pf)) + ")"
        prod_codes_p = _pf

    # ── Monthly secondary sales value (₹) per stockist ──
    if sales_mode == "before_deduction":
        _sv_expr = "((si.sales_qty + si.free_qty) / IFNULL(NULLIF(si.conversion_factor, 0), 1)) * IFNULL(si.pts, 0)"
    else:
        # After-deduction → prefer the persisted scheme_deducted_qty_calc; fall back to
        # the raw formula for legacy rows where the field is not yet backfilled.
        _sv_expr = (
            "(COALESCE(si.scheme_deducted_qty_calc, "
            "(si.sales_qty + si.free_qty - IFNULL(si.free_qty_scheme, 0))) "
            "/ IFNULL(NULLIF(si.conversion_factor, 0), 1)) * IFNULL(si.pts, 0)"
        )
    sales_rows = frappe.db.sql(f"""
        SELECT ss.stockist_code,
               MONTH(ss.statement_month) AS m,
             SUM({_sv_expr}) AS sales_value
        FROM `tabStockist Statement` ss
        INNER JOIN `tabStockist Statement Item` si
            ON si.parent = ss.name AND si.parenttype = 'Stockist Statement'
        WHERE ss.division = %s
          AND ss.docstatus IN (0, 1)
          AND ss.stockist_code IN ({placeholders})
          AND ss.statement_month BETWEEN %s AND %s{prod_clause}
        GROUP BY ss.stockist_code, MONTH(ss.statement_month)
    """, [division] + stockist_codes + [fy_start, fy_end] + prod_codes_p, as_dict=True)

    # ── Opening stock of the FIRST statement in the FY (total value across all products) ──
    opening_rows = frappe.db.sql(f"""
        SELECT ss.stockist_code,
               SUM(IFNULL(si.opening_value, 0)) AS opening_value
        FROM `tabStockist Statement` ss
        INNER JOIN `tabStockist Statement Item` si
            ON si.parent = ss.name AND si.parenttype = 'Stockist Statement'
        WHERE ss.division = %s
          AND ss.docstatus IN (0, 1)
          AND ss.stockist_code IN ({placeholders})
          AND ss.statement_month = (
              SELECT MIN(ss2.statement_month)
              FROM `tabStockist Statement` ss2
              WHERE ss2.stockist_code = ss.stockist_code
                AND ss2.division = %s
                AND ss2.docstatus IN (0, 1)
                AND ss2.statement_month BETWEEN %s AND %s
          ){prod_clause}
        GROUP BY ss.stockist_code
    """, [division] + stockist_codes + [division, fy_start, fy_end] + prod_codes_p, as_dict=True)

    # ── Closing stock of the LATEST statement in the FY ──
    closing_rows = frappe.db.sql(f"""
        SELECT ss.stockist_code,
               SUM(IFNULL(si.closing_value, 0)) AS closing_value
        FROM `tabStockist Statement` ss
        INNER JOIN `tabStockist Statement Item` si
            ON si.parent = ss.name AND si.parenttype = 'Stockist Statement'
        WHERE ss.division = %s
          AND ss.docstatus IN (0, 1)
          AND ss.stockist_code IN ({placeholders})
          AND ss.statement_month = (
              SELECT MAX(ss2.statement_month)
              FROM `tabStockist Statement` ss2
              WHERE ss2.stockist_code = ss.stockist_code
                AND ss2.division = %s
                AND ss2.docstatus IN (0, 1)
                AND ss2.statement_month BETWEEN %s AND %s
          ){prod_clause}
        GROUP BY ss.stockist_code
    """, [division] + stockist_codes + [division, fy_start, fy_end] + prod_codes_p, as_dict=True)

    # ── Index into dicts ──
    sales_by_stockist = {}
    for r in sales_rows:
        sales_by_stockist.setdefault(r.stockist_code, {})[r.m] = flt(r.sales_value)

    opening_by_stockist = {r.stockist_code: flt(r.opening_value) for r in opening_rows}
    closing_by_stockist = {r.stockist_code: flt(r.closing_value) for r in closing_rows}

    # ── Build team-grouped output ──
    LAKH = 100000.0

    def to_lakhs(v):
        return round(v / LAKH, 2) if v else 0.0

    teams_dict = {}
    for s in stockists:
        team_key = s.team or "Unassigned"
        team_display = s.team_name or team_key
        if team_key not in teams_dict:
            teams_dict[team_key] = {"team": team_key, "team_name": team_display, "stockists": []}

        monthly_values = [to_lakhs(sales_by_stockist.get(s.code, {}).get(m_num, 0))
                          for m_num in [4, 5, 6, 7, 8, 9, 10, 11, 12, 1, 2, 3]]
        total_sales = round(sum(monthly_values), 2)
        opening = to_lakhs(opening_by_stockist.get(s.code, 0))
        closing = to_lakhs(closing_by_stockist.get(s.code, 0))

        teams_dict[team_key]["stockists"].append({
            "stockist_code": s.display_code or s.code,
            "stockist_name": s.stockist_name,
            "hq": s.hq_name or s.hq or "",
            "opening": opening,
            "months": monthly_values,
            "total_sales": total_sales,
            "closing": closing,
        })

    # ── Compute team totals ──
    teams_list = []
    grand_opening = grand_months = None
    grand_total = 0.0
    grand_closing = 0.0

    for team_key, td in teams_dict.items():
        t_opening = round(sum(s["opening"] for s in td["stockists"]), 2)
        t_months = [round(sum(s["months"][i] for s in td["stockists"]), 2) for i in range(12)]
        t_total = round(sum(t_months), 2)
        t_closing = round(sum(s["closing"] for s in td["stockists"]), 2)
        td["total"] = {"opening": t_opening, "months": t_months, "total_sales": t_total, "closing": t_closing}
        teams_list.append(td)

        grand_total += t_total
        grand_closing += t_closing
        if grand_months is None:
            grand_opening = t_opening
            grand_months = list(t_months)
        else:
            grand_opening = round(grand_opening + t_opening, 2)
            grand_months = [round(grand_months[i] + t_months[i], 2) for i in range(12)]

    region_name = frappe.db.get_value("Region Master", region, "region_name") or region

    return {
        "success": True,
        "region": region,
        "region_name": region_name,
        "division": division,
        "fy_label": fy_label,
        "financial_year": financial_year,
        "month_labels": month_labels,
        "teams": teams_list,
        "sales_mode": sales_mode,
        "grand_total": {
            "opening": round(grand_opening or 0, 2),
            "months": grand_months or [0] * 12,
            "total_sales": round(grand_total, 2),
            "closing": round(grand_closing, 2),
        },
    }


# ═══════════════════════════════════════════════════════════════
# REPORT 11 – SECONDARY SALES VALUE vs CLOSING STOCK VALUE
# HQ rows grouped by Region, with one (Sec Val, Cls Val) pair per month.
# All values in ₹ Lakhs.
# ═══════════════════════════════════════════════════════════════

@frappe.whitelist()
def get_secondary_vs_closing_value_report(division=None, from_month=None, to_month=None,
                                          region_codes=None, sales_mode="after_deduction",
                                          product_codes=None, product_group=None, product_category=None):
    """Report 11 – Secondary Sales Value vs Closing Stock Value (HQ-wise).

    from_month / to_month: YYYY-MM strings (first-of-month resolution).
    region_codes: optional multi-select; if blank, includes all active regions for the division.
    sales_mode: 'after_deduction' (default) or 'before_deduction' (same semantics as Reports 7/8).
    product_group / product_category / product_codes: optional product narrowing (intersection).
    """
    if not division:
        division = get_user_division()

    from datetime import date

    # ── Resolve month range (default: current FY Apr-Mar) ──
    today = date.today()
    if not from_month or not to_month:
        start_year = today.year if today.month >= 4 else today.year - 1
        from_date_d = f"{start_year}-04-01"
        to_year = start_year + 1
        to_date_d = f"{to_year}-03-31"
    else:
        try:
            fy, fm = from_month.split("-")
            ty, tm = to_month.split("-")
            from_date_d = f"{fy}-{fm}-01"
            # last day of to_month
            import calendar
            last_d = calendar.monthrange(int(ty), int(tm))[1]
            to_date_d = f"{ty}-{tm}-{last_d:02d}"
        except Exception:
            return {"success": False, "message": "Invalid month range (expected YYYY-MM)."}

    # ── Build month sequence (year, month, label) ──
    from datetime import datetime
    cur = datetime.strptime(from_date_d, "%Y-%m-%d")
    end = datetime.strptime(to_date_d, "%Y-%m-%d")
    months_seq = []
    while cur <= end:
        label = f"{cur.strftime('%b')}-{str(cur.year)[2:]}"
        months_seq.append({"year": cur.year, "month": cur.month, "label": label,
                            "key": f"{cur.year}-{cur.month:02d}"})
        # advance one month
        if cur.month == 12:
            cur = cur.replace(year=cur.year + 1, month=1)
        else:
            cur = cur.replace(month=cur.month + 1)
    if not months_seq:
        return {"success": False, "message": "Empty month range."}

    # ── Resolve regions ──
    region_codes_n = _normalise_code_list(region_codes)
    # Confine non-admins to their mapped regions: intersect an explicit selection, or
    # default to exactly their regions when nothing was picked.
    _allowed = _allowed_region_codes_or_all(division)
    if _allowed is not None:
        _aset = set(_allowed)
        region_codes_n = [c for c in region_codes_n if c in _aset] if region_codes_n else list(_allowed)
        if not region_codes_n:
            return {"success": True, "regions": [], "months": months_seq, "grand_total": {}}
    region_filter_sql = ""
    region_params = []
    if region_codes_n:
        region_filter_sql = "AND name IN (" + ", ".join(["%s"] * len(region_codes_n)) + ")"
        region_params = region_codes_n

    regions = frappe.db.sql(
        f"""SELECT name AS code, region_name FROM `tabRegion Master`
              WHERE status='Active' AND division IN (%s, 'Both') {region_filter_sql}
              ORDER BY region_name""",
        [division] + region_params, as_dict=True,
    )
    if not regions:
        return {"success": True, "regions": [], "months": months_seq, "grand_total": {}}

    region_codes_all = [r.code for r in regions]
    r_ph = ", ".join(["%s"] * len(region_codes_all))

    # ── HQs for those regions ──
    hqs = frappe.db.sql(
        f"""SELECT name AS code, hq_name, region
              FROM `tabHQ Master`
              WHERE status='Active' AND division=%s AND region IN ({r_ph})
              ORDER BY hq_name""",
        [division] + region_codes_all, as_dict=True,
    )
    if not hqs:
        return {"success": True, "regions": [], "months": months_seq, "grand_total": {}}

    hq_codes = [h.code for h in hqs]
    hq_ph = ", ".join(["%s"] * len(hq_codes))

    # Optional product / group / category narrowing (intersection)
    _pf = _resolve_product_filter(division, product_codes, product_group, product_category)
    prod_clause = ""
    prod_codes_p = []
    if _pf is not None:
        prod_clause = " AND si.product_code IN (" + ", ".join(["%s"] * len(_pf)) + ")"
        prod_codes_p = _pf

    # ── Monthly secondary sales value (in Rs) per HQ ──
    if sales_mode == "before_deduction":
        sv_expr = "((si.sales_qty + si.free_qty) / IFNULL(NULLIF(si.conversion_factor, 0), 1)) * IFNULL(si.pts, 0)"
    else:
        sv_expr = (
            "(COALESCE(si.scheme_deducted_qty_calc, "
            "(si.sales_qty + si.free_qty - IFNULL(si.free_qty_scheme, 0))) "
            "/ IFNULL(NULLIF(si.conversion_factor, 0), 1)) * IFNULL(si.pts, 0)"
        )

    sales_rows = frappe.db.sql(f"""
        SELECT ss.hq,
               YEAR(ss.statement_month) AS y,
               MONTH(ss.statement_month) AS m,
               SUM({sv_expr}) AS sales_value
          FROM `tabStockist Statement` ss
    INNER JOIN `tabStockist Statement Item` si
            ON si.parent = ss.name AND si.parenttype = 'Stockist Statement'
         WHERE ss.division = %s
           AND ss.docstatus IN (0, 1)
           AND ss.hq IN ({hq_ph})
           AND ss.statement_month BETWEEN %s AND %s{prod_clause}
      GROUP BY ss.hq, YEAR(ss.statement_month), MONTH(ss.statement_month)
    """, [division] + hq_codes + [from_date_d, to_date_d] + prod_codes_p, as_dict=True)

    # ── Monthly closing stock value per HQ ──
    closing_rows = frappe.db.sql(f"""
        SELECT ss.hq,
               YEAR(ss.statement_month) AS y,
               MONTH(ss.statement_month) AS m,
               SUM(IFNULL(si.closing_value, 0)) AS closing_value
          FROM `tabStockist Statement` ss
    INNER JOIN `tabStockist Statement Item` si
            ON si.parent = ss.name AND si.parenttype = 'Stockist Statement'
         WHERE ss.division = %s
           AND ss.docstatus IN (0, 1)
           AND ss.hq IN ({hq_ph})
           AND ss.statement_month BETWEEN %s AND %s{prod_clause}
      GROUP BY ss.hq, YEAR(ss.statement_month), MONTH(ss.statement_month)
    """, [division] + hq_codes + [from_date_d, to_date_d] + prod_codes_p, as_dict=True)

    LAKH = 100000.0

    def to_lakhs(v):
        return round(flt(v) / LAKH, 2) if v else 0.0

    # Index rows: (hq, year, month) → value
    sales_idx = {(r.hq, r.y, r.m): flt(r.sales_value) for r in sales_rows}
    closing_idx = {(r.hq, r.y, r.m): flt(r.closing_value) for r in closing_rows}

    # ── Build per-region structure ──
    hq_by_region = {}
    for h in hqs:
        hq_by_region.setdefault(h.region, []).append(h)

    regions_out = []
    grand_monthly = [{"sec_val": 0.0, "cls_val": 0.0} for _ in months_seq]

    for r in regions:
        region_hqs = hq_by_region.get(r.code, [])
        region_monthly = [{"sec_val": 0.0, "cls_val": 0.0} for _ in months_seq]
        hqs_out = []
        for h in region_hqs:
            row_monthly = []
            for mi, ms in enumerate(months_seq):
                sv = to_lakhs(sales_idx.get((h.code, ms["year"], ms["month"]), 0))
                cv = to_lakhs(closing_idx.get((h.code, ms["year"], ms["month"]), 0))
                row_monthly.append({"sec_val": sv, "cls_val": cv})
                region_monthly[mi]["sec_val"] = round(region_monthly[mi]["sec_val"] + sv, 2)
                region_monthly[mi]["cls_val"] = round(region_monthly[mi]["cls_val"] + cv, 2)
            hqs_out.append({"hq_code": h.code, "hq_name": h.hq_name or h.code, "monthly": row_monthly})

        for mi in range(len(months_seq)):
            grand_monthly[mi]["sec_val"] = round(grand_monthly[mi]["sec_val"] + region_monthly[mi]["sec_val"], 2)
            grand_monthly[mi]["cls_val"] = round(grand_monthly[mi]["cls_val"] + region_monthly[mi]["cls_val"], 2)

        regions_out.append({
            "region_code": r.code,
            "region_name": r.region_name or r.code,
            "hqs": hqs_out,
            "totals": {"monthly": region_monthly},
        })

    return {
        "success": True,
        "regions": regions_out,
        "months": months_seq,
        "grand_total": {"monthly": grand_monthly},
        "from_month": from_month,
        "to_month": to_month,
        "sales_mode": sales_mode,
    }


# ═══════════════════════════════════════════════════════════════
# REPORT 14 – TARGET vs SALES (HQ-wise timeline)
# Same shape as Report 11 (HQ rows grouped by Region, one pair per month) but
# the pair is (Target, Sales). Target comes from HQ Yearly Target / HQ Target
# Item monthly fields (stored in ₹ Lakhs). Sales can be Secondary (with the
# after/before-deduction toggle) or Primary (Primary Sales Data). All ₹ Lakhs.
# ═══════════════════════════════════════════════════════════════

@frappe.whitelist()
def get_target_vs_sales_report(division=None, from_month=None, to_month=None,
                               region_codes=None, sales_type="secondary",
                               sales_mode="after_deduction"):
    """Report 14 – Target vs Sales (HQ-wise, monthly).

    from_month / to_month: YYYY-MM. region_codes: optional multi-select.
    sales_type: 'secondary' (default) or 'primary'.
    sales_mode: 'after_deduction' | 'before_deduction' (secondary only).
    """
    if not division:
        division = get_user_division()

    from datetime import date, datetime
    import calendar

    # ── Resolve month range (default: current FY Apr-Mar) ──
    today = date.today()
    if not from_month or not to_month:
        start_year = today.year if today.month >= 4 else today.year - 1
        from_date_d = f"{start_year}-04-01"
        to_date_d = f"{start_year + 1}-03-31"
    else:
        try:
            fy, fm = from_month.split("-")
            ty, tm = to_month.split("-")
            from_date_d = f"{fy}-{fm}-01"
            last_d = calendar.monthrange(int(ty), int(tm))[1]
            to_date_d = f"{ty}-{tm}-{last_d:02d}"
        except Exception:
            return {"success": False, "message": "Invalid month range (expected YYYY-MM)."}

    # ── Build month sequence ──
    cur = datetime.strptime(from_date_d, "%Y-%m-%d")
    end = datetime.strptime(to_date_d, "%Y-%m-%d")
    months_seq = []
    while cur <= end:
        months_seq.append({"year": cur.year, "month": cur.month,
                           "label": f"{cur.strftime('%b')}-{str(cur.year)[2:]}",
                           "key": f"{cur.year}-{cur.month:02d}"})
        if cur.month == 12:
            cur = cur.replace(year=cur.year + 1, month=1)
        else:
            cur = cur.replace(month=cur.month + 1)
    if not months_seq:
        return {"success": False, "message": "Empty month range."}

    # ── Resolve regions & HQs (active, like Report 11) ──
    region_codes_n = _normalise_code_list(region_codes)
    # Confine non-admins to their mapped regions: intersect an explicit selection, or
    # default to exactly their regions when nothing was picked.
    _allowed = _allowed_region_codes_or_all(division)
    if _allowed is not None:
        _aset = set(_allowed)
        region_codes_n = [c for c in region_codes_n if c in _aset] if region_codes_n else list(_allowed)
        if not region_codes_n:
            return {"success": True, "regions": [], "months": months_seq, "grand_total": {}}
    region_filter_sql = ""
    region_params = []
    if region_codes_n:
        region_filter_sql = "AND name IN (" + ", ".join(["%s"] * len(region_codes_n)) + ")"
        region_params = region_codes_n

    regions = frappe.db.sql(
        f"""SELECT name AS code, region_name FROM `tabRegion Master`
              WHERE status='Active' AND division IN (%s, 'Both') {region_filter_sql}
              ORDER BY region_name""",
        [division] + region_params, as_dict=True,
    )
    if not regions:
        return {"success": True, "regions": [], "months": months_seq, "grand_total": {}}

    region_codes_all = [r.code for r in regions]
    r_ph = ", ".join(["%s"] * len(region_codes_all))

    hqs = frappe.db.sql(
        f"""SELECT name AS code, hq_name, region
              FROM `tabHQ Master`
              WHERE status='Active' AND division=%s AND region IN ({r_ph})
              ORDER BY hq_name""",
        [division] + region_codes_all, as_dict=True,
    )
    if not hqs:
        return {"success": True, "regions": [], "months": months_seq, "grand_total": {}}

    hq_codes = [h.code for h in hqs]
    hq_ph = ", ".join(["%s"] * len(hq_codes))

    LAKH = 100000.0

    def to_lakhs(v):
        return round(flt(v) / LAKH, 2) if v else 0.0

    # ── Sales value (₹) per HQ per (year, month) ──
    sales_idx = {}
    if sales_type == "primary":
        stk_rows = frappe.db.sql(
            f"""SELECT name, hq FROM `tabStockist Master`
                  WHERE division=%s AND hq IN ({hq_ph})""",
            [division] + hq_codes)
        stk_hq = {r[0]: r[1] for r in stk_rows}
        stk_codes = list(stk_hq.keys())
        if stk_codes:
            sp_ph = ", ".join(["%s"] * len(stk_codes))
            pri_rows = frappe.db.sql(f"""
                SELECT stockist_code,
                       YEAR(invoicedate) AS y, MONTH(invoicedate) AS m,
                       SUM(ptsvalue) AS val
                  FROM `tabPrimary Sales Data`
                 WHERE division = %s AND iscancelled = 0
                   AND stockist_code IN ({sp_ph})
                   AND invoicedate BETWEEN %s AND %s
              GROUP BY stockist_code, YEAR(invoicedate), MONTH(invoicedate)
            """, [division] + stk_codes + [from_date_d, to_date_d], as_dict=True)
            for r in pri_rows:
                hqc = stk_hq.get(r.stockist_code)
                if hqc:
                    sales_idx[(hqc, r.y, r.m)] = sales_idx.get((hqc, r.y, r.m), 0.0) + flt(r.val)
    else:
        if sales_mode == "before_deduction":
            sv_expr = "((si.sales_qty + si.free_qty) / IFNULL(NULLIF(si.conversion_factor, 0), 1)) * IFNULL(si.pts, 0)"
        else:
            sv_expr = (
                "(COALESCE(si.scheme_deducted_qty_calc, "
                "(si.sales_qty + si.free_qty - IFNULL(si.free_qty_scheme, 0))) "
                "/ IFNULL(NULLIF(si.conversion_factor, 0), 1)) * IFNULL(si.pts, 0)"
            )
        sec_rows = frappe.db.sql(f"""
            SELECT ss.hq, YEAR(ss.statement_month) AS y, MONTH(ss.statement_month) AS m,
                   SUM({sv_expr}) AS val
              FROM `tabStockist Statement` ss
        INNER JOIN `tabStockist Statement Item` si
                ON si.parent = ss.name AND si.parenttype = 'Stockist Statement'
             WHERE ss.division = %s AND ss.docstatus IN (0, 1)
               AND ss.hq IN ({hq_ph})
               AND ss.statement_month BETWEEN %s AND %s
          GROUP BY ss.hq, YEAR(ss.statement_month), MONTH(ss.statement_month)
        """, [division] + hq_codes + [from_date_d, to_date_d], as_dict=True)
        for r in sec_rows:
            sales_idx[(r.hq, r.y, r.m)] = flt(r.val)

    # ── Target (₹ Lakhs) per HQ per month, from HQ Yearly Target ──
    month_field = {1: "jan", 2: "feb", 3: "mar", 4: "apr", 5: "may", 6: "jun",
                   7: "jul", 8: "aug", 9: "sep", 10: "oct", 11: "nov", 12: "dec"}

    def _fy_of(year, month):
        sy = year if month >= 4 else year - 1
        return f"{sy}-{str(sy + 1)[2:]}"

    fys = sorted({_fy_of(ms["year"], ms["month"]) for ms in months_seq})
    fy_ph = ", ".join(["%s"] * len(fys))
    target_idx = {}  # (hq, fy) -> {field: lakhs}
    tgt_rows = frappe.db.sql(f"""
        SELECT ti.hq AS hq, yt.financial_year AS fy,
               ti.apr, ti.may, ti.jun, ti.jul, ti.aug, ti.sep,
               ti.oct, ti.nov, ti.dec, ti.jan, ti.feb, ti.mar
          FROM `tabHQ Yearly Target` yt
    INNER JOIN `tabHQ Target Item` ti ON ti.parent = yt.name
         WHERE yt.docstatus = 1 AND yt.division = %s
           AND yt.financial_year IN ({fy_ph})
           AND ti.hq IN ({hq_ph})
    """, [division] + fys + hq_codes, as_dict=True)
    for r in tgt_rows:
        key = (r.hq, r.fy)
        acc = target_idx.setdefault(key, {})
        for f in month_field.values():
            acc[f] = acc.get(f, 0.0) + flt(r.get(f))

    def _target_for(hq_code, ms):
        fy = _fy_of(ms["year"], ms["month"])
        field = month_field[ms["month"]]
        return round(flt(target_idx.get((hq_code, fy), {}).get(field, 0)), 2)

    # ── Build per-region structure ──
    hq_by_region = {}
    for h in hqs:
        hq_by_region.setdefault(h.region, []).append(h)

    region_names = {r.code: (r.region_name or r.code) for r in regions}
    regions_out = []
    grand_monthly = [{"target": 0.0, "sales": 0.0} for _ in months_seq]

    for r in regions:
        region_hqs = hq_by_region.get(r.code, [])
        region_monthly = [{"target": 0.0, "sales": 0.0} for _ in months_seq]
        hqs_out = []
        for h in region_hqs:
            row_monthly = []
            for mi, ms in enumerate(months_seq):
                tv = _target_for(h.code, ms)
                sv = to_lakhs(sales_idx.get((h.code, ms["year"], ms["month"]), 0))
                row_monthly.append({"target": tv, "sales": sv})
                region_monthly[mi]["target"] = round(region_monthly[mi]["target"] + tv, 2)
                region_monthly[mi]["sales"] = round(region_monthly[mi]["sales"] + sv, 2)
            hqs_out.append({"hq_code": h.code, "hq_name": h.hq_name or h.code, "monthly": row_monthly})

        for mi in range(len(months_seq)):
            grand_monthly[mi]["target"] = round(grand_monthly[mi]["target"] + region_monthly[mi]["target"], 2)
            grand_monthly[mi]["sales"] = round(grand_monthly[mi]["sales"] + region_monthly[mi]["sales"], 2)

        regions_out.append({
            "region_code": r.code,
            "region_name": region_names.get(r.code, r.code),
            "hqs": hqs_out,
            "totals": {"monthly": region_monthly},
        })

    return {
        "success": True,
        "regions": regions_out,
        "months": months_seq,
        "grand_total": {"monthly": grand_monthly},
        "from_month": from_month,
        "to_month": to_month,
        "sales_type": sales_type,
        "sales_mode": sales_mode,
    }


# ═══════════════════════════════════════════════════════════════
# Shared helper – resolve product-attribute filters → product codes
# ═══════════════════════════════════════════════════════════════

def _scheme_pack_conversion(pack_str):
    """Strips-per-box for a pack string = first number of an 'NxM' pack, else 1.

    Mirrors SchemeRequest._get_conversion_factor so box-converted scheme quantities
    (Qty/Free) reconcile with the stored order value (Value = box-qty × rate).
    """
    if not pack_str:
        return 1
    pack_str = str(pack_str).strip().upper()
    match = re.match(r'(\d+)\s*[xX]\s*(\d+)', pack_str)
    if match:
        return flt(match.group(1)) or 1
    return 1


def _sql_pack_conversion(pack_col):
    """SQL expression mirroring _scheme_pack_conversion(): strips-per-box = the first
    number of an 'NxM' pack (e.g. 10x15 -> 10), else 1. Used inside aggregate report
    SQL so discount value stays per-box and reconciles with the stored order value.
    Guards against a 0 factor to avoid divide-by-zero."""
    p = f"LOWER(TRIM({pack_col}))"
    n = f"CAST(SUBSTRING_INDEX({p}, 'x', 1) AS DECIMAL(20,6))"
    return (f"(CASE WHEN {p} REGEXP '^[0-9]+[[:space:]]*x[[:space:]]*[0-9]+' "
            f"AND {n} > 0 THEN {n} ELSE 1 END)")


def _resolve_product_filter(division, product_codes=None, product_group=None, product_category=None):
    """Return a list of Product Master codes matching the given filters.

    Returns None when no product filter is active (caller applies no restriction).
    The three dimensions combine as an intersection (AND); each accepts a JSON
    list or CSV string and a blank value means 'no constraint' for that dimension.
    When filters are active but match nothing, returns a sentinel that matches no
    rows (so the report shows an empty set rather than everything).
    """
    pcodes = _normalise_code_list(product_codes)
    pgroups = _normalise_code_list(product_group)
    pcats = _normalise_code_list(product_category)
    if not (pcodes or pgroups or pcats):
        return None

    conds = ["division = %s", "status = 'Active'"]
    params = [division]
    if pcodes:
        conds.append("product_code IN (" + ", ".join(["%s"] * len(pcodes)) + ")")
        params += pcodes
    if pgroups:
        conds.append("product_group IN (" + ", ".join(["%s"] * len(pgroups)) + ")")
        params += pgroups
    if pcats:
        conds.append("category IN (" + ", ".join(["%s"] * len(pcats)) + ")")
        params += pcats
    rows = frappe.db.sql(
        "SELECT name FROM `tabProduct Master` WHERE " + " AND ".join(conds), params)
    return [r[0] for r in rows] or ["__no_match__"]


# ═══════════════════════════════════════════════════════════════
# REPORT 13 – FULL MONTHLY ORGANIZATIONAL REPORT (product-level)
# Flat one-row-per Stockist × Product line for a single month, mirroring the
# Secondary Sales export, with reconciling value columns (₹):
# Opening | Billed Sales | Free Goods | Scheme Deduction | Net Sales | Closing,
# where Net = Billed + Free − Scheme. Billed ties to the export's nrvvalue
# (stored sales_value_pts) and Closing to clsvalue. Filters: team, hq, product
# multi-select, product group, product category. Amounts returned in rupees;
# the frontend/Excel format to ₹ Lakhs. Overall total is surfaced separately so
# the UI can pin it at the top.
# ═══════════════════════════════════════════════════════════════

@frappe.whitelist()
def get_monthly_organizational_report(division=None, month=None, team=None, hq=None,
                                      product_codes=None, product_group=None,
                                      product_category=None, limit=None):
    """Report 13 – Full Monthly Organizational secondary-sales report (item level).

    limit: optional cap on the number of line-item rows returned (used by the
    on-screen preview to avoid hanging the browser on huge result sets). The
    grand total is always computed over the FULL set; Excel/PDF pass no limit.
    """
    if not division:
        division = get_user_division()

    from datetime import date as _date, datetime as _datetime
    import calendar

    if not month:
        month = _date.today().strftime("%Y-%m")
    try:
        y, m = str(month)[:7].split("-")
        y, m = int(y), int(m)
        first_day = f"{y:04d}-{m:02d}-01"
        last_day = f"{y:04d}-{m:02d}-{calendar.monthrange(y, m)[1]:02d}"
        month_label = _datetime.strptime(first_day, "%Y-%m-%d").strftime("%b-%y")
    except Exception:
        return {"success": False, "message": "Invalid month (expected YYYY-MM)."}

    conds = ["ss.division = %s", "ss.docstatus IN (0, 1)",
             "ss.statement_month BETWEEN %s AND %s"]
    params = [division, first_day, last_day]
    # No region filter on this report — still confine non-admins to their regions.
    _scope_region_sql_pos(conds, params, "ss.region", division, None)
    if team:
        conds.append("ss.team = %s")
        params.append(team)
    if hq:
        conds.append("ss.hq = %s")
        params.append(hq)

    prod_filter = _resolve_product_filter(division, product_codes, product_group, product_category)
    if prod_filter is not None:
        conds.append("si.product_code IN (" + ", ".join(["%s"] * len(prod_filter)) + ")")
        params += prod_filter

    where = " AND ".join(conds)
    conv = "IFNULL(NULLIF(si.conversion_factor, 0), 1)"
    pts = "IFNULL(si.pts, 0)"

    # Report 13 mirrors the Primary Sales Data column format. Quantities are box-converted
    # (raw statement qtys are at strip level — divide by the conversion factor).
    #   PTS rate = Product Master PTS (pm.pts) — the fixed catalogue rate, never modified.
    #   NRV rate = statement-editable si.pts — equals PTS, or lower when a scheme discount
    #              was applied at the statement level.
    # Two qty views are exposed: "before scheme" (sales+free) drives PTS Value (× Master PTS)
    # and "after scheme" (sales+free-scheme) drives NRV Value (× NRV = scheme-deducted sales).
    rows = frappe.db.sql(f"""
        SELECT COALESCE(sm.stockist_code, ss.stockist_code) AS stockist_code,
               ss.stockist_name AS stockist_name,
               COALESCE(hm.hq_name, ss.hq, '') AS hq_name,
               COALESCE(tm.team_name, ss.team, '') AS team,
               COALESCE(rm.region_name, ss.region, '') AS region,
               COALESCE(pm.product_code, si.product_code) AS product_code,
               COALESCE(pm.product_name, si.product_name, si.raw_product_name, '') AS product_name,
               COALESCE(pm.product_group, '') AS product_group,
               COALESCE(pm.category, '') AS category,
               COALESCE(NULLIF(si.pack, ''), pm.pack, '') AS pack,
               {conv}                                            AS conv,
               IFNULL(pm.pts, 0)                                 AS master_pts,
               {pts}                                             AS pts_rate,
               IFNULL(si.sales_qty, 0)                           AS sales_qty,
               IFNULL(si.free_qty, 0)                            AS free_qty,
               IFNULL(si.free_qty_scheme, 0)                     AS free_qty_scheme,
               IFNULL(si.opening_qty, 0)                         AS opening_qty,
               IFNULL(si.closing_qty, 0)                         AS closing_qty
          FROM `tabStockist Statement` ss
    INNER JOIN `tabStockist Statement Item` si
            ON si.parent = ss.name AND si.parenttype = 'Stockist Statement'
     LEFT JOIN `tabStockist Master` sm ON sm.name = ss.stockist_code
     LEFT JOIN `tabHQ Master` hm ON hm.name = ss.hq
     LEFT JOIN `tabTeam Master` tm ON tm.name = ss.team
     LEFT JOIN `tabRegion Master` rm ON rm.name = ss.region
     LEFT JOIN `tabProduct Master` pm ON pm.name = si.product_code
         WHERE {where}
      ORDER BY hq_name, stockist_name, product_name
    """, params, as_dict=True)

    grand = {"sales_before": 0.0, "sales_after": 0.0, "clstock": 0.0, "opstock": 0.0,
             "ptsvalue": 0.0, "nrvvalue": 0.0, "clsvalue": 0.0, "opsvalue": 0.0}
    out_rows = []
    for r in rows:
        conv_f = flt(r.conv) or 1
        pts_rate = flt(r.master_pts)   # PTS rate = Product Master PTS (fixed catalogue rate)
        nrv_rate = flt(r.pts_rate)     # NRV rate = statement-editable si.pts (= PTS, or lower if discounted)

        # Two box-converted sales-qty views:
        #   before scheme = sales + free            (gross — drives PTS Value)
        #   after  scheme = sales + free - scheme   (net "true sales" — drives NRV Value)
        sales_before = (flt(r.sales_qty) + flt(r.free_qty)) / conv_f
        sales_after = (flt(r.sales_qty) + flt(r.free_qty) - flt(r.free_qty_scheme)) / conv_f
        opstock = flt(r.opening_qty) / conv_f
        clstock = flt(r.closing_qty) / conv_f

        ptsvalue = sales_before * pts_rate   # PTS Value = before-scheme qty × Master PTS
        nrvvalue = sales_after * nrv_rate    # NRV Value = after-scheme qty × NRV (scheme-deducted sales value)
        opsvalue = opstock * pts_rate        # opening valued at PTS (Master) rate
        clsvalue = clstock * pts_rate        # closing valued at PTS (Master) rate

        out_rows.append({
            "stockist_code": r.stockist_code or "",
            "stockist_name": r.stockist_name or "",
            "citypool": r.hq_name or "",          # client keeps HQ name in the City/Pool column
            "team": r.team or "",
            "region": r.region or "",
            "product_code": r.product_code or "",
            "product_name": r.product_name or "",
            "pack": r.pack or "",
            "sales_before": sales_before, "sales_after": sales_after,
            "clstock": clstock, "opstock": opstock,
            "pts": pts_rate, "ptsvalue": ptsvalue,
            "nrv": nrv_rate, "nrvvalue": nrvvalue,
            "clsvalue": clsvalue, "opsvalue": opsvalue,
            "product_head": r.product_group or "",     # Product Head == Product Group
            "product_category": r.category or "",
        })
        grand["sales_before"] += sales_before
        grand["sales_after"]  += sales_after
        grand["clstock"]  += clstock
        grand["opstock"]  += opstock
        grand["ptsvalue"] += ptsvalue
        grand["nrvvalue"] += nrvvalue
        grand["clsvalue"] += clsvalue
        grand["opsvalue"] += opsvalue

    total_count = len(out_rows)
    truncated = False
    try:
        lim = int(limit) if limit else 0
    except (TypeError, ValueError):
        lim = 0
    if lim and total_count > lim:
        out_rows = out_rows[:lim]
        truncated = True

    return {
        "success": True,
        "division": division,
        "month": str(month)[:7],
        "month_label": month_label,
        "rows": out_rows,
        "grand_total": grand,
        "total_count": total_count,
        "shown_count": len(out_rows),
        "truncated": truncated,
        "limit": lim or None,
    }


@frappe.whitelist()
def render_spreadsheet_preview(file_url):
    """Render an XLS/XLSX/CSV/TXT file as an HTML table for QC viewer preview."""
    import pandas as pd
    import html as html_mod

    if not file_url or not isinstance(file_url, str):
        return {"html": ""}

    # Resolve to absolute file path on disk
    if file_url.startswith("/files/"):
        file_path = frappe.get_site_path("public", file_url.lstrip("/"))
    elif file_url.startswith("/private/files/"):
        file_path = frappe.get_site_path(file_url.lstrip("/"))
    else:
        file_path = frappe.get_site_path("public", file_url.lstrip("/"))

    if not os.path.isfile(file_path):
        return {"html": "<p class='text-muted text-center'>File not found</p>"}

    ext = os.path.splitext(file_path)[1].lower()

    try:
        if ext == ".csv":
            df = pd.read_csv(file_path, dtype=str, na_filter=False)
        elif ext in (".xls", ".xlsx"):
            df = pd.read_excel(file_path, dtype=str, na_filter=False)
        elif ext == ".txt":
            df = pd.read_csv(file_path, dtype=str, sep="\t", na_filter=False)
        else:
            return {"html": "<p class='text-muted text-center'>Unsupported file type</p>"}

        # Limit to first 500 rows for performance
        truncated = len(df) > 500
        df = df.head(500)

        # Build HTML table with escaped content
        rows_html = []
        # Header
        header_cells = "".join(
            f"<th>{html_mod.escape(str(c))}</th>" for c in df.columns
        )
        rows_html.append(f"<thead><tr>{header_cells}</tr></thead>")

        # Body
        rows_html.append("<tbody>")
        for _, row in df.iterrows():
            cells = "".join(
                f"<td>{html_mod.escape(str(v))}</td>" for v in row
            )
            rows_html.append(f"<tr>{cells}</tr>")
        rows_html.append("</tbody>")

        table_html = f"<table>{''.join(rows_html)}</table>"
        if truncated:
            table_html += "<p class='text-muted text-center mt-2'><small>Showing first 500 rows</small></p>"

        return {"html": table_html}

    except Exception as e:
        frappe.log_error(f"Spreadsheet preview error: {e}", "render_spreadsheet_preview")
        return {"html": f"<p class='text-muted text-center'>Could not parse file</p>"}


@frappe.whitelist()
def get_products_for_division(division=None):
    """Return all active products for a division for manual statement entry.

    `name` is the Product Master id (PK) the frontend must submit as the row
    value; `product_code` is the editable business code shown to the user."""
    if not division:
        division = get_user_division()

    products = frappe.get_all(
        "Product Master",
        filters={"status": "Active", "division": division},
        fields=["name", "product_code", "product_name", "pack", "pts", "ptr", "pack_conversion"],
        order_by="sequence asc, product_name asc",
        limit_page_length=0,
    )
    return products


@frappe.whitelist()
def create_manual_statement(stockist_code, statement_month, items, uploaded_file=None, remarks=None, division=None):
    """
    Create a Stockist Statement from manual entry (no AI extraction).
    items: JSON array of product rows with quantities.
    """
    try:
        if isinstance(items, str):
            items = json.loads(items)

        if not stockist_code or not statement_month:
            return {"success": False, "message": "Stockist code and statement month are required."}

        if not division:
            division = get_user_division()

        # Resolve the editable Stockist Code to the master id (PK), scoped to the active
        # division so a code reused across divisions resolves to THIS division's stockist.
        # The statement (and every report) links by the id, not the editable code.
        stockist_pk = _resolve_stockist_pk(stockist_code, division)
        if not stockist_pk:
            return {"success": False, "message": f"Stockist '{stockist_code}' not found."}

        # Normalise month
        if len(statement_month) == 7:
            statement_month = statement_month + "-01"

        # Check stockist is active
        stockist_status = frappe.db.get_value("Stockist Master", stockist_pk, "status")
        if stockist_status != "Active":
            return {"success": False, "message": f"Stockist '{stockist_code}' is inactive. Statement cannot be created for an inactive stockist."}

        # Check for duplicates (statements link by the master id)
        existing = frappe.db.exists("Stockist Statement", {
            "stockist_code": stockist_pk,
            "statement_month": statement_month,
        })
        if existing:
            return {"success": False, "message": f"A statement already exists: {existing}"}

        doc = frappe.new_doc("Stockist Statement")
        doc.stockist_code = stockist_pk
        doc.statement_month = statement_month
        doc.extracted_data_status = "Completed"
        if uploaded_file:
            doc.uploaded_file = uploaded_file
        if remarks:
            doc.remarks = remarks

        for row in items:
            product_code = row.get("productcode") or row.get("product_code")
            if not product_code:
                continue
            # The row value may be the id (new clients) or the editable business
            # code (older clients); resolve within THIS division so a code reused
            # across divisions links to the right product.
            product_pk = _resolve_product_pk(product_code, division)
            if not product_pk:
                return {"success": False, "message": f"Product '{product_code}' not found in division '{division}'."}
            doc.append("items", {
                "product_code": product_pk,
                # Per-line PTS override (manual scheme-discount path). Zero/blank
                # falls back to Master PTS inside calculate_closing_and_totals().
                "pts": flt(row.get("pts") or 0),
                "opening_qty": flt(row.get("openingqty") or 0),
                "purchase_qty": flt(row.get("purchaseqty") or 0),
                "sales_qty": flt(row.get("salesqty") or 0),
                "free_qty": flt(row.get("freeqty") or 0),
                "free_qty_scheme": flt(row.get("freeqtyscheme") or 0),
                "return_qty": flt(row.get("returnqty") or 0),
                "misc_out_qty": flt(row.get("miscoutqty") or 0),
                "closing_qty": flt(row.get("closingqty") or 0),
                "closing_value": flt(row.get("closingvalue") or 0),
            })

        if not doc.items:
            return {"success": False, "message": "No valid product rows provided."}

        doc.calculate_closing_and_totals()
        doc.save(ignore_permissions=True)
        frappe.db.commit()

        return {
            "success": True,
            "message": f"Statement saved as Draft with {len(doc.items)} items.",
            "doc_name": doc.name,
        }

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Create Manual Statement Error")
        return {"success": False, "message": str(e)}


# ═══════════════════════════════════════════════════════════════
# PRODUCT CORRECTION / MAPPING ENDPOINTS
# ═══════════════════════════════════════════════════════════════

@frappe.whitelist()
def save_product_correction(stockist_code, raw_product_name, mapped_product_code, statement_name=None):
    """
    Create or update a Stockist Product Correction record.
    Called from the portal when QC maps an unmapped item.
    """
    try:
        if not stockist_code or not raw_product_name or not mapped_product_code:
            return {"success": False, "message": "stockist_code, raw_product_name, and mapped_product_code are required"}

        raw_name = raw_product_name.strip().upper()

        # The correction stores a Link (Product Master id). Accept either the id
        # or the editable business code, scoped to the stockist's division.
        stockist_division = frappe.db.get_value("Stockist Master", stockist_code, "division")
        product_pk = _resolve_product_pk(mapped_product_code, stockist_division)
        if not product_pk:
            return {"success": False, "message": f"Product '{mapped_product_code}' not found"}
        mapped_product_code = product_pk

        # Check if correction already exists
        existing = frappe.db.get_value(
            "Stockist Product Correction",
            {"stockist_code": stockist_code, "raw_product_name": raw_name},
            "name",
        )

        if existing:
            doc = frappe.get_doc("Stockist Product Correction", existing)
            doc.mapped_product_code = mapped_product_code
            doc.status = "Active"
            doc.save(ignore_permissions=True)
        else:
            doc = frappe.get_doc({
                "doctype": "Stockist Product Correction",
                "stockist_code": stockist_code,
                "raw_product_name": raw_name,
                "mapped_product_code": mapped_product_code,
                "division": frappe.db.get_value("Stockist Master", stockist_code, "division"),
                "created_from_statement": statement_name or "",
                "status": "Active",
            })
            doc.insert(ignore_permissions=True)

        frappe.db.commit()
        return {"success": True, "correction_name": doc.name, "message": "Correction saved"}

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Save Product Correction Error")
        return {"success": False, "message": str(e)}


@frappe.whitelist()
def get_product_search_for_mapping(query, division=None):
    """
    Search Product Master for mapping dropdown in QC screen.
    Returns matching products by code or name.
    """
    try:
        query = (query or "").strip()
        if len(query) < 2:
            return {"success": True, "results": []}

        filters = [
            ["Product Master", "status", "=", "Active"],
        ]
        if division:
            filters.append(["Product Master", "division", "=", division])

        # Search by product_code or product_name (name kept for legacy ids)
        or_filters = [
            ["Product Master", "product_code", "like", f"%{query}%"],
            ["Product Master", "product_name", "like", f"%{query}%"],
            ["Product Master", "name", "like", f"%{query}%"],
        ]

        # `name` is the id the frontend must submit; `product_code` is the
        # editable business code it must display (they differ for new products).
        results = frappe.get_all(
            "Product Master",
            filters=filters,
            or_filters=or_filters,
            fields=["name", "product_code", "product_name", "pack", "pts", "division"],
            order_by="product_code asc",
            limit_page_length=20,
        )

        return {"success": True, "results": results}

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Product Search For Mapping Error")
        return {"success": False, "message": str(e)}


@frappe.whitelist()
def apply_mapping_and_save_correction(doc_name, row_idx, mapped_product_code):
    """
    Map a single unmapped item in a statement to a product,
    save the correction for future auto-mapping, and recalculate.
    row_idx: 0-based index in the items table.
    """
    try:
        row_idx = int(row_idx)
        doc = frappe.get_doc("Stockist Statement", doc_name)

        if doc.docstatus == 1:
            return {"success": False, "message": "Cannot modify a submitted statement"}

        if row_idx < 0 or row_idx >= len(doc.items):
            return {"success": False, "message": f"Invalid row index: {row_idx}"}

        item = doc.items[row_idx]
        raw_name = (item.raw_product_name or "").strip().upper()

        # Accept either the id (new clients) or the editable business code
        # (older clients), scoped to the statement's division.
        product_pk = _resolve_product_pk(mapped_product_code, doc.division)
        if not product_pk:
            return {"success": False, "message": f"Product '{mapped_product_code}' not found"}

        # Update the item
        item.product_code = product_pk
        item.mapping_status = "matched"

        # Keep extraction confidence intact; mapping review is reflected via QC status.
        conf_values = [flt(i.row_confidence) for i in doc.items]
        doc.confidence_score = round(sum(conf_values) / len(conf_values), 1) if conf_values else 0

        # Recalculate
        doc.calculate_closing_and_totals()
        doc.calculate_qc_confidence()
        doc.save(ignore_permissions=True)
        frappe.db.commit()

        # Show the editable business code in the message, never the id.
        display_code = frappe.db.get_value("Product Master", product_pk, "product_code") or product_pk
        return {
            "success": True,
            "message": f"Mapped '{raw_name}' → {display_code}",
            "qc_confidence": doc.qc_confidence,
            "confidence_score": flt(doc.confidence_score),
        }

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Apply Mapping Error")
        return {"success": False, "message": str(e)}


@frappe.whitelist()
def mark_statement_qc_reviewed(doc_name):
    """
    Mark a draft statement as QC-reviewed by the QC team.
    Sets qc_review_status → 'QC Reviewed' without changing docstatus.
    Draft auto-save already persists the latest item edits, so no items data needed here.
    """
    try:
        doc = frappe.get_doc("Stockist Statement", doc_name)

        if doc.docstatus == 1:
            return {"success": False, "message": "Statement is already finalised."}

        doc.qc_confidence = "QC Reviewed"
        doc.save(ignore_permissions=True)
        frappe.db.commit()

        return {"success": True, "qc_confidence": "QC Reviewed"}

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Mark QC Reviewed Error")
        return {"success": False, "message": str(e)}


@frappe.whitelist()
def finalize_statement(doc_name):
    """
    Finalise (submit) a draft statement — sets docstatus → 1 and locks it permanently.
    The latest items are already persisted by the draft auto-save mechanism.
    Requires extraction to be Completed before finalising.
    """
    try:
        doc = frappe.get_doc("Stockist Statement", doc_name)

        if doc.docstatus == 1:
            return {"success": False, "message": "Statement is already finalised."}

        if doc.extracted_data_status != "Completed":
            return {"success": False, "message": "Statement extraction must be completed before finalising."}

        doc.calculate_closing_and_totals()
        doc.save(ignore_permissions=True)
        frappe.db.commit()
        doc.submit()
        frappe.db.commit()

        return {
            "success": True,
            "message": f"Statement {doc_name} has been finalised successfully.",
            "doc_name": doc.name
        }

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Finalise Statement Error")
        return {"success": False, "message": str(e)}


@frappe.whitelist()
def override_qc_confidence(doc_name, new_confidence):
    """
    Manually override QC confidence for a statement.
    Used when a user confirms auto-mapped items are correct.
    Only allows setting to 'All Matched' from 'Verification Needed'.
    """
    try:
        if new_confidence != "All Matched":
            return {"success": False, "message": "Only 'All Matched' override is supported"}

        doc = frappe.get_doc("Stockist Statement", doc_name)

        if doc.docstatus != 0:
            return {"success": False, "message": "Can only override QC on draft statements"}

        if doc.qc_confidence != "Verification Needed":
            return {"success": False, "message": "Override only available for 'Verification Needed' statements"}

        # Mark all auto_mapped items as matched
        for item in doc.items:
            if item.mapping_status == "auto_mapped":
                item.mapping_status = "matched"

        doc.qc_confidence = "All Matched"
        doc.save(ignore_permissions=True)
        frappe.db.commit()

        return {"success": True, "qc_confidence": doc.qc_confidence}

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Override QC Confidence Error")
        return {"success": False, "message": str(e)}


@frappe.whitelist()
def get_bulk_jobs_list_enhanced(division=None, page=1, page_size=20):
    """
    Enhanced bulk jobs list with job_name and QC confidence summary.
    """
    try:
        if not division:
            division = get_user_division()

        page = int(page or 1)
        page_size = int(page_size or 20)
        offset = (page - 1) * page_size

        filters = {}
        if division:
            filters["division"] = division

        jobs = frappe.get_all(
            "Bulk Statement Upload",
            filters=filters,
            fields=[
                "name", "job_name", "statement_month", "division", "status",
                "progress", "total_files", "success_count", "failed_count",
                "skipped_count", "creation", "zip_file",
            ],
            order_by="creation desc",
            start=offset,
            page_length=page_size,
        )

        total_count = frappe.db.count("Bulk Statement Upload", filters=filters)

        # For each job, get QC confidence summary from its statements
        for job in jobs:
            job["statement_month"] = str(job.get("statement_month") or "")
            job["creation"] = str(job.get("creation") or "")
            log_data = frappe.db.get_value("Bulk Statement Upload", job["name"], "extraction_log")
            if log_data:
                try:
                    log_entries = json.loads(log_data)
                    statement_names = [e.get("statement") for e in log_entries if e.get("statement")]
                    if statement_names:
                        qc_counts = frappe.db.sql("""
                            SELECT qc_confidence, COUNT(*) as cnt
                            FROM `tabStockist Statement`
                            WHERE name IN %(names)s
                            GROUP BY qc_confidence
                        """, {"names": statement_names}, as_dict=True)
                        job["qc_summary"] = {r.qc_confidence: r.cnt for r in qc_counts}
                    else:
                        job["qc_summary"] = {}
                except (json.JSONDecodeError, TypeError):
                    job["qc_summary"] = {}
            else:
                job["qc_summary"] = {}

        return {
            "success": True,
            "jobs": jobs,
            "total": total_count,
            "page": page,
            "page_size": page_size,
        }

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Get Bulk Jobs List Enhanced Error")
        return {"success": False, "message": str(e)}


# ============================================================
# AI CHATBOT - Text2SQL Agentic Pipeline
# ============================================================

SCANIFY_SCHEMA_PROMPT = """
You are Scanify AI Assistant — an expert analytical chatbot for Stedman Pharmaceuticals' Stockist & Scheme Management System.
You convert natural language questions into SQL queries against a MariaDB database, then format results clearly.

## COMPANY CONTEXT
- Company: Stedman Pharmaceuticals
- Application: Scanify — manages stockist statements, scheme requests, primary sales, and sales targets
- Divisions: Prima, Vektra (division-controlled, user sees data for their active division)
- Current user division filter: {division}

## ORGANIZATIONAL HIERARCHY (top → bottom)
Division → Zone → Region → Team → HQ → Stockist
- Each level links to its parent. Zone is top geographic grouping, HQ is the lowest field unit.
- Stockists belong to an HQ. Doctors are also mapped to HQ/Team/Region.

## DATABASE SCHEMA

### Master Tables

**`tabDivision`** — Division master
- name (PK), division_name

**`tabZone Master`** — Zone master
- name (PK, format Z####), zone_code, zone_name, division (FK→tabDivision), status

**`tabState Master`** — State master
- name (PK, format ST####), state_code, state_name, division (FK→tabDivision), status

**`tabRegion Master`** — Region master
- name (PK, format R####), region_code, region_name, division (FK→tabDivision), zone (FK→tabZone Master), state (FK→tabState Master), status

**`tabTeam Master`** — Team master
- name (PK, format T####), team_code, team_name, division (FK→tabDivision), region (FK→tabRegion Master), sanctioned_strength, status

**`tabHQ Master`** — Headquarters master (lowest field unit)
- name (PK, format HQ####), hq_code, hq_name, division (FK→tabDivision), region (FK→tabRegion Master), team (FK→tabTeam Master), zone (Data), per_capita, status

**`tabProduct Master`** — Product catalog
- name (PK, format PR####), product_code, product_name, pack, pack_conversion, division (Select: Prima/Vektra/ASPR/Wellness), product_group (Select: Dentist/Derma/Contus/Xptum/Amino/Ortho/Drez/Gynae/Jusdee/Dygerm/Others), category (Main Product/Hospital Product/New Product), mrp, ptr, pts, gst_rate, status

**`tabStockist Master`** — Stockist (distributor) master
- name (PK, format S####), stockist_code, stockist_name, hq (FK→tabHQ Master), division (FK→tabDivision), team (FK→tabTeam Master), region (FK→tabRegion Master), zone, address, city, contact_person, phone, email, status

**`tabDoctor Master`** — Doctor master
- name (PK, format D####), doctor_code, doctor_name, division (FK→tabDivision), qualification, doctor_category, specialization, phone, place, hospital_address, house_address, hq (FK→tabHQ Master), team (FK→tabTeam Master), region (FK→tabRegion Master), state (FK→tabState Master), zone, chemist_name, status

### Transaction Tables

**`tabStockist Statement`** — Monthly stockist stock statements (Submittable: docstatus 0=Draft,1=Submitted,2=Cancelled)
- name (PK), stockist_code (FK→tabStockist Master), stockist_name, hq (FK→tabHQ Master), division (FK→tabDivision), team (FK→tabTeam Master), region (FK→tabRegion Master), zone, statement_month (Date, 1st of month), uploaded_file, extracted_data_status (Pending/In Progress/Completed/Failed), extraction_notes, qc_confidence, total_opening_value, total_purchase_value, total_free_value, total_sales_value_pts, total_sales_value_ptr, total_closing_value, docstatus, owner, creation, modified

**`tabStockist Statement Item`** — Line items of stockist statement (child table, parent=tabStockist Statement)
- name, parent (FK→tabStockist Statement), parenttype='Stockist Statement', idx, product_code (FK→tabProduct Master), product_name, raw_product_name, mapping_status, pack, opening_qty, prev_month_closing, purchase_qty, sales_qty, free_qty, free_qty_scheme, conversion_factor, return_qty, misc_out_qty, closing_qty, pts, sales_value_pts, sales_value_ptr, opening_value, purchase_value, closing_value, scheme_deducted_qty_calc

**`tabScheme Request`** — Scheme/discount requests (Submittable)
- name (PK, format SCH-YYYY-#####), application_date, requested_by (FK→tabUser), division (FK→tabDivision), region (FK→tabRegion Master), team (FK→tabTeam Master), doctor_code (FK→tabDoctor Master), doctor_name, doctor_place, specialization, hospital_address, hq (FK→tabHQ Master), stockist_code (FK→tabStockist Master), stockist_name, total_scheme_value, approval_status (Pending/Approved/Rejected/Rerouted), repeated_request, scheme_notes, docstatus, owner, creation, modified

**`tabScheme Request Item`** — Line items of scheme request (child table, parent=tabScheme Request)
- name, parent (FK→tabScheme Request), idx, product_code (FK→tabProduct Master), product_name, pack, quantity, free_quantity, product_rate, special_rate, scheme_percentage, product_value

**`tabScheme Approval Log`** — Approval workflow log (child table, parent=tabScheme Request)
- name, parent (FK→tabScheme Request), idx, approver (FK→tabUser), approval_level, action (Approved/Rejected/Rerouted), action_date, comments

**`tabScheme Deduction`** — Deductions applied against statements (Submittable)
- name (PK, format SD-####), scheme_request (FK→tabScheme Request), scheme_date, doctor_code (FK→tabDoctor Master), division (FK→tabDivision), stockist_statement (FK→tabStockist Statement), stockist_code (FK→tabStockist Master), deduction_date, status (Draft/Applied/Cancelled), total_deducted_qty, total_deducted_value, notes, docstatus, owner, creation, modified

**`tabScheme Deduction Item`** — Line items of deduction (child table, parent=tabScheme Deduction)
- name, parent (FK→tabScheme Deduction), idx, product_code (FK→tabProduct Master), product_name, pack, scheme_free_qty, current_free_qty, deduct_qty, pts, deducted_value

**`tabPrimary Sales Data`** — Invoice-level primary sales data
- name (PK, hash), upload_month (YYYY-MM), division (FK→tabDivision), upload_ref (FK→tabPrimary Sales Upload), iscancelled, direct_party, stockist_code (Data), stockist_name, citypool, team (Data), region (Data), act_region, zonee, invoiceno, invoicedate, pcode, product (product name), product_head (product group), pack, batchno, quantity, freeqty, expqty, pts, ptsvalue, ptr, ptrvalue, mrp, mrpvalue, nrv, nrvvalue, dsort

**`tabPrimary Sales Upload`** — Upload job tracking
- name (PK, format PRI-YYYY-#####), upload_month (YYYY-MM), division (FK→tabDivision), uploaded_by (FK→tabUser), upload_date, file, total_rows, success_count, error_count, status (Pending/Processing/Completed/Failed), error_log

**`tabBulk Statement Upload`** — Bulk OCR processing jobs (Submittable)
- name (PK, format BULK-month-####), job_name, statement_month, zip_file, division (FK→tabDivision), status (Pending/Queued/In Progress/Completed/Failed/Partially Completed), progress, total_files, success_count, failed_count, skipped_count, job_id, extraction_log, docstatus, owner, creation, modified

**`tabHQ Yearly Target`** — Annual sales targets per HQ
- name (PK), division (FK→tabDivision), region (FK→tabRegion Master), financial_year (e.g. 2025-26), start_date, end_date, status (Draft/Active/Completed/Cancelled), total_target_amount (Lakhs), total_hqs, upload_date, uploaded_by

**`tabHQ Target Item`** — Monthly target breakdowns (child table, parent=tabHQ Yearly Target)
- name, parent (FK→tabHQ Yearly Target), idx, hq (FK→tabHQ Master), hq_name, team (FK→tabTeam Master), apr, may, jun, jul, aug, sep, oct, nov, dec, jan, feb, mar (all Currency in Lakhs), q1_total, q2_total, q3_total, q4_total, yearly_total

**`tabStockist Product Correction`** — OCR product name corrections/mappings
- name (PK, format SPC-####), stockist_code (FK→tabStockist Master), raw_product_name, mapped_product_code (FK→tabProduct Master), division (FK→tabDivision), created_from_statement (FK→tabStockist Statement), status

## IMPORTANT SQL RULES
1. Always use backticks around table names with spaces: `tabStockist Statement`
2. For submitted documents, filter by docstatus=1 unless user asks for drafts
3. Date fields: statement_month stores 1st of month (e.g. '2025-06-01' for June 2025)
4. Primary Sales upload_month is YYYY-MM format string (e.g. '2025-06')
5. Division filter: ALWAYS add division='{division}' to WHERE clause unless query is explicitly cross-division
6. Currency values: pts/ptr/mrp are per-unit prices. Values (_value suffix) are computed totals.
7. Use LEFT JOINs for optional relationships, INNER JOINs for required ones
8. Limit results to 500 rows max unless user asks for all
9. For aggregations, always include meaningful GROUP BY
10. Financial year runs April to March (e.g. 2025-26 = Apr 2025 - Mar 2026)
11. Use COALESCE for nullable numeric fields in aggregations

## RESPONSE FORMAT
You MUST respond with a valid JSON object in one of these formats:

### For data queries:
{{
  "type": "data",
  "sql": "SELECT ... FROM ... WHERE ...",
  "title": "Brief title describing the results",
  "description": "One-line explanation of what the data shows",
  "columns": ["col1", "col2", ...],
  "format": "table"
}}

### For chart-worthy queries (trends, comparisons, distributions):
{{
  "type": "chart",
  "sql": "SELECT ... FROM ... WHERE ...",
  "title": "Chart title",
  "description": "What the chart shows",
  "chart_type": "bar|line|pie|doughnut|horizontalBar",
  "label_column": "column_name_for_labels",
  "value_columns": ["col1", "col2"],
  "columns": ["col1", "col2", ...],
  "format": "chart"
}}

### For count/single-value queries:
{{
  "type": "metric",
  "sql": "SELECT COUNT(*) as count FROM ...",
  "title": "Metric title",
  "description": "What this metric represents",
  "columns": ["count"],
  "format": "metric"
}}

### For conversational/non-data questions:
{{
  "type": "text",
  "message": "Your helpful response here"
}}

### For errors or unclear questions:
{{
  "type": "error",
  "message": "Explain what's wrong or ask for clarification"
}}

## CHART SELECTION GUIDELINES
- Time series / monthly trends → "line"
- Comparisons across categories (regions, teams, products) → "bar" or "horizontalBar"
- Distribution / share / percentage breakdown → "pie" or "doughnut"
- Top N rankings → "horizontalBar"
- If user says "chart" or "graph", pick the best chart_type automatically
- For multi-series charts, use multiple value_columns

## ANALYTICAL CAPABILITIES
You can answer questions like:
- Sales performance by region/team/HQ/stockist
- Product-wise sales trends and comparisons
- Scheme utilization and deduction analysis
- Target vs achievement analysis
- Stock movement patterns (opening, purchase, sales, closing)
- Top/bottom performers at any hierarchy level
- Month-over-month and year-over-year comparisons
- Stockist statement submission tracking
- Product group analysis
- Zone/Region wise breakdowns

Always think step by step about which tables to join and what aggregations are needed.
"""


@frappe.whitelist()
def chatbot_query(message, conversation_history=None):
    """
    Process a natural language query through the Text2SQL pipeline.
    Uses Gemini to convert question to SQL, executes it, and returns formatted results.
    """
    if not message or not message.strip():
        return {"success": False, "error": "Please enter a question."}

    try:
        api_key, model_name, is_enabled = get_gemini_settings()
    except Exception:
        return {"success": False, "error": "Gemini AI is not configured. Please enable it in Scanify Settings."}

    division = get_user_division()

    # Build the system prompt with division context
    system_prompt = SCANIFY_SCHEMA_PROMPT.format(division=division)

    # Build conversation messages
    messages = []
    if conversation_history:
        if isinstance(conversation_history, str):
            try:
                conversation_history = json.loads(conversation_history)
            except (json.JSONDecodeError, TypeError):
                conversation_history = []
        # Include last 10 exchanges for context
        for entry in conversation_history[-10:]:
            role = entry.get("role", "user")
            content = entry.get("content", "")
            if role == "user":
                messages.append({"role": "user", "content": content})
            elif role == "assistant":
                messages.append({"role": "assistant", "content": content})

    messages.append({"role": "user", "content": message})

    try:
        client = genai_sdk.Client(api_key=api_key)

        # Build contents for Gemini
        contents = []
        for msg in messages:
            role = "user" if msg["role"] == "user" else "model"
            contents.append(genai_types.Content(
                role=role,
                parts=[genai_types.Part(text=msg["content"])]
            ))

        response = client.models.generate_content(
            model=model_name,
            contents=contents,
            config=genai_types.GenerateContentConfig(
                system_instruction=system_prompt,
                temperature=0.1,
                max_output_tokens=8192,
            )
        )

        ai_text = response.text.strip()

        # Extract JSON from response (handle markdown code blocks)
        json_str = ai_text
        if "```json" in json_str:
            json_str = json_str.split("```json")[1].split("```")[0].strip()
        elif "```" in json_str:
            json_str = json_str.split("```")[1].split("```")[0].strip()

        try:
            ai_response = json.loads(json_str)
        except json.JSONDecodeError:
            # If AI returned plain text, treat as text response
            return {
                "success": True,
                "type": "text",
                "message": ai_text,
                "raw_response": ai_text
            }

        response_type = ai_response.get("type", "text")

        if response_type == "text":
            return {
                "success": True,
                "type": "text",
                "message": ai_response.get("message", ai_text),
                "raw_response": ai_text
            }

        if response_type == "error":
            return {
                "success": True,
                "type": "error",
                "message": ai_response.get("message", "I couldn't understand that query."),
                "raw_response": ai_text
            }

        sql = ai_response.get("sql", "")
        if not sql:
            return {
                "success": True,
                "type": "text",
                "message": ai_response.get("message", ai_text),
                "raw_response": ai_text
            }

        # Security: validate the SQL is read-only
        sql_upper = sql.strip().upper()
        # Use word-boundary regex to avoid false positives (e.g. LOAD in UPLOAD)
        forbidden_patterns = [
            r'\bINSERT\b', r'\bUPDATE\b', r'\bDELETE\b', r'\bDROP\b',
            r'\bALTER\b', r'\bCREATE\b', r'\bTRUNCATE\b', r'\bGRANT\b',
            r'\bREVOKE\b', r'\bEXEC\b', r'\bEXECUTE\b', r'\bCALL\b',
            r'\bLOAD\b', r'INTO\s+OUTFILE', r'INTO\s+DUMPFILE',
            r'\bINFORMATION_SCHEMA\b', r'\bSLEEP\s*\(', r'\bBENCHMARK\s*\('
        ]
        for pattern in forbidden_patterns:
            if re.search(pattern, sql_upper):
                return {
                    "success": False,
                    "error": "Only read-only SELECT queries are allowed.",
                    "raw_response": ai_text
                }

        if not sql_upper.lstrip().startswith("SELECT") and not sql_upper.lstrip().startswith("WITH"):
            return {
                "success": False,
                "error": "Only SELECT queries are permitted.",
                "raw_response": ai_text
            }

        # Execute the SQL query
        try:
            results = frappe.db.sql(sql, as_dict=True)
        except Exception as db_err:
            # If query fails, ask Gemini to fix it
            error_msg = str(db_err)
            retry_prompt = f"The SQL query failed with error: {error_msg}\n\nOriginal query: {sql}\n\nPlease fix the SQL and respond with the corrected JSON."
            contents.append(genai_types.Content(
                role="model",
                parts=[genai_types.Part(text=ai_text)]
            ))
            contents.append(genai_types.Content(
                role="user",
                parts=[genai_types.Part(text=retry_prompt)]
            ))

            retry_response = client.models.generate_content(
                model=model_name,
                contents=contents,
                config=genai_types.GenerateContentConfig(
                    system_instruction=system_prompt,
                    temperature=0.1,
                    max_output_tokens=8192,
                )
            )

            retry_text = retry_response.text.strip()
            retry_json = retry_text
            if "```json" in retry_json:
                retry_json = retry_json.split("```json")[1].split("```")[0].strip()
            elif "```" in retry_json:
                retry_json = retry_json.split("```")[1].split("```")[0].strip()

            try:
                retry_parsed = json.loads(retry_json)
                retry_sql = retry_parsed.get("sql", "")
                if retry_sql:
                    # Validate retry SQL too
                    retry_upper = retry_sql.strip().upper()
                    for pattern in forbidden_patterns:
                        if re.search(pattern, retry_upper):
                            return {"success": False, "error": "Only read-only queries are allowed."}
                    if not retry_upper.lstrip().startswith("SELECT") and not retry_upper.lstrip().startswith("WITH"):
                        return {"success": False, "error": "Only SELECT queries are permitted."}

                    results = frappe.db.sql(retry_sql, as_dict=True)
                    ai_response = retry_parsed
                    sql = retry_sql
                else:
                    return {
                        "success": False,
                        "error": f"Query failed: {error_msg}",
                        "raw_response": ai_text
                    }
            except Exception:
                return {
                    "success": False,
                    "error": f"Query failed: {error_msg}",
                    "raw_response": ai_text
                }

        # Format results
        rows = []
        for row in results[:500]:
            formatted_row = {}
            for key, val in row.items():
                if val is None:
                    formatted_row[key] = ""
                elif isinstance(val, (int, float)):
                    formatted_row[key] = val
                else:
                    formatted_row[key] = str(val)
            rows.append(formatted_row)

        columns = ai_response.get("columns", list(rows[0].keys()) if rows else [])

        response_data = {
            "success": True,
            "type": response_type,
            "title": ai_response.get("title", "Query Results"),
            "description": ai_response.get("description", ""),
            "columns": columns,
            "data": rows,
            "total_rows": len(rows),
            "sql": sql,
            "raw_response": ai_text
        }

        # Add chart config if chart type
        if response_type == "chart":
            response_data["chart_type"] = ai_response.get("chart_type", "bar")
            response_data["label_column"] = ai_response.get("label_column", columns[0] if columns else "")
            response_data["value_columns"] = ai_response.get("value_columns", columns[1:] if len(columns) > 1 else columns)

        return response_data

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Chatbot Query Error")
        return {
            "success": False,
            "error": f"An error occurred: {str(e)}"
        }


# ═══════════════════════════════════════════════════════════════
# Secondary Sales (Stockist Statement) Backfill — Import & Export
# ───────────────────────────────────────────────────────────────
# Bulk backfill of one month's secondary sales (stockist statements) from an
# Excel export of the legacy/source system. One Stockist Statement is created
# per stockist for the chosen month, with the product rows as items.
#
# Strictness rules (per requirement):
#   • Stockist  → matched STRICTLY by name. The file's `stockistcode` is from the
#     legacy migration and is NOT trusted for mapping. An unmatched name is an
#     error and that row is skipped.
#   • Product   → mapped by `pcode` only (code is authoritative; product names are
#     never used to map). Unmapped codes are kept on the row (mapping_status =
#     'unmapped') but excluded from value totals by the controller.
#   • PTS        → the per-line `pts` from the file is taken as final (a value that
#     differs from the master means a scheme discount was given). Values are then
#     computed by the Stockist Statement controller (qty × pts, pack conversion).
# ═══════════════════════════════════════════════════════════════

# Excel header → Stockist Statement Item field (quantity inputs only)
_SECONDARY_QTY_COL_MAP = {
    "quantity": "sales_qty",
    "clstock": "closing_qty",
    "opstock": "opening_qty",
}


@frappe.whitelist()
def get_secondary_sales_count(month, division=None, zone=None, region=None, team=None, hq=None):
    """Count statements + line items of secondary sales data for a month (with optional org filters)."""
    from frappe.utils import get_first_day, get_last_day

    if not division:
        division = get_user_division() or "Prima"
    if not month:
        return {"statements": 0, "rows": 0}

    month_str = str(month)[:7] + "-01"
    first_day = get_first_day(month_str)
    last_day = get_last_day(first_day)

    conditions = ["ss.division = %(division)s",
                  "ss.statement_month BETWEEN %(first_day)s AND %(last_day)s",
                  "ss.docstatus IN (0, 1)"]
    params = {"division": division, "first_day": first_day, "last_day": last_day}
    if zone:
        conditions.append("ss.zone = %(zone)s"); params["zone"] = zone
    _scope_region_sql(conditions, params, "ss.region", division, region)
    if team:
        conditions.append("ss.team = %(team)s"); params["team"] = team
    if hq:
        conditions.append("ss.hq = %(hq)s"); params["hq"] = hq

    where = " AND ".join(conditions)
    res = frappe.db.sql(f"""
        SELECT COUNT(DISTINCT ss.name) AS statements, COUNT(si.name) AS `rows`
        FROM `tabStockist Statement` ss
        LEFT JOIN `tabStockist Statement Item` si
            ON si.parent = ss.name AND si.parenttype = 'Stockist Statement'
        WHERE {where}
    """, params, as_dict=True)
    row = res[0] if res else {}
    return {"statements": int(row.get("statements") or 0), "rows": int(row.get("rows") or 0)}


@frappe.whitelist()
def process_secondary_sales_upload(upload_month, file_url):
    """Backfill secondary sales (stockist statements) from one month's Excel file.

    See module header above for the strictness rules. Returns a summary dict with
    counts and a capped list of human-readable warnings/errors.
    """
    import openpyxl
    from datetime import datetime
    from frappe.utils import get_first_day

    user_division = get_user_division() or "Prima"

    # Validate month format (YYYY-MM)
    if not upload_month or not re.match(r"^\d{4}-\d{2}$", str(upload_month)):
        return {"success": False, "error": "Invalid month format. Use YYYY-MM."}

    statement_month = get_first_day(str(upload_month) + "-01")

    # Resolve the uploaded file path
    file_path = frappe.get_site_path(file_url.lstrip("/"))
    if not os.path.exists(file_path):
        file_path = frappe.get_site_path("private", "files", os.path.basename(file_url))
    if not os.path.exists(file_path):
        return {"success": False, "error": "Uploaded file not found on server."}

    # Persistent log record (browsable in desk + on the portal import page)
    log_doc = frappe.get_doc({
        "doctype": "Secondary Sales Upload",
        "upload_month": str(upload_month),
        "division": user_division,
        "uploaded_by": frappe.session.user,
        "upload_date": frappe.utils.now(),
        "file": file_url,
        "status": "Processing",
    })
    log_doc.insert(ignore_permissions=True)
    frappe.db.commit()

    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active

        # Headers from row 1 (lower-cased, trimmed)
        raw_headers = [cell.value for cell in ws[1]]
        headers = [str(h).strip().lower() if h is not None else None for h in raw_headers]

        # Require the stockist code column — it's the match key (compared against the
        # editable Stockist Code in the master). Stockist name is an optional fallback.
        if "stockistcode" not in [h for h in headers if h]:
            wb.close()
            log_doc.reload()
            log_doc.status = "Failed"
            log_doc.log = "Missing required column: stockistcode"
            log_doc.save(ignore_permissions=True)
            frappe.db.commit()
            return {"success": False, "error": "Missing required column: stockistcode", "log_name": log_doc.name}

        # ── Caches ──
        # Stockist resolution: primary match by the editable Stockist Code, fallback by
        # name. Both map to the master's internal id (name, S####) — that is what the
        # Stockist Statement Link stores, so resolving to the id keeps the backfill valid
        # even after a code is edited.
        stockist_code_cache = {}   # editable code -> {id, status}
        stockist_name_cache = {}   # name          -> {id, status}
        for s in frappe.get_all("Stockist Master",
                                filters={"division": ["in", [user_division, "Both"]]},
                                fields=["name", "stockist_code", "stockist_name", "status"],
                                limit_page_length=0):
            if s.stockist_code:
                stockist_code_cache[str(s.stockist_code).strip().lower()] = {
                    "id": s.name, "status": s.status}
            if s.stockist_name:
                stockist_name_cache[s.stockist_name.strip().lower()] = {
                    "id": s.name, "status": s.status}

        # Product: code → id (authoritative mapping by code only). Statement items
        # link by the Product Master id, so resolve the business code here; on a
        # code collision across the allowed divisions, prefer the uploader's own.
        product_code_cache = {}
        for p in frappe.get_all("Product Master",
                                filters={"division": ["in", [user_division, "Both", "ASPR", "Wellness"]]},
                                fields=["product_code", "name", "division"],
                                limit_page_length=0):
            if p.product_code and (p.product_code not in product_code_cache or p.division == user_division):
                product_code_cache[p.product_code] = p.name

        # ── Group rows by matched stockist code ──
        groups = {}                 # stockist_code -> list[item_dict]
        unmatched_names = {}        # raw name -> count (strict skip)
        inactive_names = {}         # name -> code (skipped: inactive stockist)
        unmapped_products = 0
        blank_rows = 0
        total_data_rows = 0
        skipped_zero_sales = 0   # rows where the sales (quantity) column is 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None for v in row):
                continue

            rec = {}
            for col_idx, val in enumerate(row):
                if col_idx >= len(headers) or headers[col_idx] is None:
                    continue
                rec[headers[col_idx]] = val

            stk_code = str(rec.get("stockistcode") or "").strip()
            stk_name = str(rec.get("stockistname") or "").strip()
            pcode = str(rec.get("pcode") or "").strip()
            raw_product = str(rec.get("product") or "").strip()

            # Skip rows with no stockist and no product/qty content
            if not stk_code and not stk_name and not pcode and not raw_product:
                blank_rows += 1
                continue

            total_data_rows += 1

            # Identify the stockist by code first (against the editable Stockist Code),
            # then fall back to name.
            ident = stk_code or stk_name
            if not ident:
                unmatched_names["(blank stockist code/name)"] = unmatched_names.get("(blank stockist code/name)", 0) + 1
                continue

            match = None
            if stk_code:
                match = stockist_code_cache.get(stk_code.lower())
            if not match and stk_name:
                match = stockist_name_cache.get(stk_name.lower())
            if not match:
                unmatched_names[ident] = unmatched_names.get(ident, 0) + 1
                continue
            if (match.get("status") or "Active") != "Active":
                inactive_names[ident] = match["id"]
                continue

            # Group by and link via the internal id (S####) — required for the
            # Stockist Statement Link field, and stable across code edits.
            stockist_code = match["id"]

            # The "quantity" column is the final sales figure. Free quantity is NOT
            # counted toward sales. Any row with zero sales is dropped (these are the
            # previous-month closing carry-over rows the source repeats).
            sales_q = _parse_num(rec.get("quantity"))
            if sales_q == 0:
                skipped_zero_sales += 1
                continue

            # Product mapping is by CODE only; store the resolved id (Link value)
            if pcode and pcode in product_code_cache:
                product_code = product_code_cache[pcode]
                mapping_status = "matched"
            else:
                product_code = None
                mapping_status = "unmapped"
                unmapped_products += 1

            # Valuation rate = NRV (the net / final price rate). It equals PTS when no
            # discount was given and is lower when a discount was applied, so it is the
            # true realised rate. Fall back to the file's PTS, then to the master rate.
            rate = _parse_num(rec.get("nrv")) or _parse_num(rec.get("pts"))

            item = {
                "raw_product_name": raw_product or pcode,
                "row_type": "product",
                "mapping_status": mapping_status,
                "pts": rate,                # per-line valuation rate (NRV / final net rate)
                "free_qty": 0,              # free quantity is not counted toward sales
                "free_qty_scheme": 0,
                "purchase_qty": 0,          # not present in secondary statement files
                "return_qty": 0,
                "misc_out_qty": 0,
            }
            for excel_col, field in _SECONDARY_QTY_COL_MAP.items():
                item[field] = _parse_num(rec.get(excel_col))
            if product_code:
                item["product_code"] = product_code

            groups.setdefault(stockist_code, []).append(item)

        wb.close()

    except Exception as e:
        frappe.log_error(f"Secondary Sales Upload (read) Error: {str(e)}", "Secondary Sales Upload")
        log_doc.reload()
        log_doc.status = "Failed"
        log_doc.log = f"Could not read the Excel file: {str(e)}"
        log_doc.save(ignore_permissions=True)
        frappe.db.commit()
        return {"success": False, "error": f"Could not read the Excel file: {str(e)}", "log_name": log_doc.name}

    # ── Create one Stockist Statement per stockist (DRAFT) ──
    statements_created = 0
    items_created = 0
    skipped_existing = []       # list of {stockist, name}
    create_errors = []          # per-stockist failures

    for stockist_code, items in groups.items():
        try:
            existing = frappe.db.exists("Stockist Statement", {
                "stockist_code": stockist_code,
                "statement_month": statement_month,
            })
            if existing:
                skipped_existing.append({
                    "stockist": stockist_code,
                    "name": existing,
                })
                continue

            doc = frappe.new_doc("Stockist Statement")
            doc.stockist_code = stockist_code
            doc.statement_month = statement_month
            doc.extracted_data_status = "Completed"
            doc.skip_conversion = 1   # quantities are already final → value = qty x rate
            doc.extraction_notes = f"Backfilled via secondary sales bulk upload ({upload_month})."
            for it in items:
                doc.append("items", it)

            # Controller computes conversion factor, values and totals on validate.
            doc.insert(ignore_permissions=True)   # saved as Draft (docstatus = 0)
            frappe.db.commit()

            statements_created += 1
            items_created += len(items)

        except Exception as e:
            frappe.db.rollback()
            create_errors.append(f"{stockist_code}: {str(e)}")
            frappe.log_error(frappe.get_traceback(), "Secondary Sales Upload (create) Error")

    # ── Build human-readable messages (capped for the UI) ──
    messages = []
    if unmatched_names:
        sample = list(unmatched_names.keys())[:15]
        messages.append(
            f"{len(unmatched_names)} stockist(s) not found in Stockist Master by code or name "
            f"(rows skipped): " + "; ".join(sample) +
            (" ..." if len(unmatched_names) > 15 else "")
        )
    if inactive_names:
        messages.append(
            f"{len(inactive_names)} stockist(s) are inactive and were skipped: "
            + "; ".join(list(inactive_names.keys())[:15])
        )
    if skipped_existing:
        names = [f"{x['stockist']} ({x['name']})" for x in skipped_existing[:15]]
        messages.append(
            f"{len(skipped_existing)} stockist(s) already have a statement for this month "
            f"and were skipped — delete those first to re-import: " + "; ".join(names) +
            (" ..." if len(skipped_existing) > 15 else "")
        )
    if unmapped_products:
        messages.append(
            f"{unmapped_products} row(s) had a product code not found in Product Master — "
            f"kept on the statement as unmapped (excluded from value totals)."
        )
    if skipped_zero_sales:
        messages.append(
            f"{skipped_zero_sales} row(s) skipped (sales/quantity = 0). The quantity column "
            f"is the final sales figure; free quantity is not counted."
        )
    if create_errors:
        messages.extend(create_errors[:15])

    # ── Build the FULL (uncapped) log text for persistence ──
    log_lines = [
        f"Secondary Sales Import — {upload_month} ({user_division})",
        f"Run by {frappe.session.user} at {frappe.utils.now()}",
        "",
        f"Data rows read         : {total_data_rows}",
        f"Stockists in file      : {len(groups)}",
        f"Statements created     : {statements_created}",
        f"Item rows created      : {items_created}",
        f"Zero-sales rows skipped: {skipped_zero_sales}",
        f"Skipped (existed)      : {len(skipped_existing)}",
        f"Unmatched stockists    : {len(unmatched_names)}",
        f"Inactive skipped       : {len(inactive_names)}",
        f"Unmapped product rows  : {unmapped_products}",
        f"Create errors          : {len(create_errors)}",
    ]
    if unmatched_names:
        log_lines += ["", "── Unmatched stockists (code/name × row-count) ──"]
        log_lines += [f"  • {nm}  (×{cnt})" for nm, cnt in sorted(unmatched_names.items())]
    if inactive_names:
        log_lines += ["", "── Inactive stockists skipped ──"]
        log_lines += [f"  • {nm}  → {code}" for nm, code in sorted(inactive_names.items())]
    if skipped_existing:
        log_lines += ["", "── Skipped: statement already exists ──"]
        log_lines += [f"  • {x['stockist']}  → {x['name']}" for x in skipped_existing]
    if create_errors:
        log_lines += ["", "── Per-stockist create errors ──"]
        log_lines += [f"  • {err}" for err in create_errors]
    full_log = "\n".join(log_lines)

    # ── Persist the log record ──
    try:
        log_doc.reload()
        log_doc.status = "Completed"
        log_doc.total_data_rows = total_data_rows
        log_doc.stockists_in_file = len(groups)
        log_doc.statements_created = statements_created
        log_doc.items_created = items_created
        log_doc.skipped_existing = len(skipped_existing)
        log_doc.unmatched_stockists = len(unmatched_names)
        log_doc.inactive_stockists = len(inactive_names)
        log_doc.unmapped_products = unmapped_products
        log_doc.create_errors = len(create_errors)
        log_doc.skipped_zero_sales = skipped_zero_sales
        log_doc.log = full_log[:140000]
        log_doc.save(ignore_permissions=True)
        frappe.db.commit()
    except Exception:
        frappe.log_error(frappe.get_traceback(), "Secondary Sales Upload (log save) Error")

    return {
        "success": True,
        "log_name": log_doc.name,
        "month": str(upload_month),
        "division": user_division,
        "total_data_rows": total_data_rows,
        "statements_created": statements_created,
        "items_created": items_created,
        "stockists_in_file": len(groups),
        "skipped_existing": len(skipped_existing),
        "unmatched_stockists": len(unmatched_names),
        "inactive_stockists": len(inactive_names),
        "unmapped_products": unmapped_products,
        "skipped_zero_sales": skipped_zero_sales,
        "create_errors": len(create_errors),
        "messages": messages,
    }


@frappe.whitelist()
def export_secondary_sales_data(month, division=None, zone=None, region=None,
                                team=None, hq=None, docstatus=None):
    """Export secondary sales (stockist statement) data as Excel for a month.

    Output columns mirror the import template so the file round-trips. Optional
    org filters (zone/region/team/hq) narrow the export. Value columns are
    recomputed from the stored quantities and per-line PTS.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from frappe.utils import get_first_day, get_last_day

    if not division:
        division = get_user_division() or "Prima"
    if not month:
        frappe.throw("Month is required")

    month_str = str(month)[:7] + "-01"
    first_day = get_first_day(month_str)
    last_day = get_last_day(first_day)

    conditions = ["ss.division = %(division)s",
                  "ss.statement_month BETWEEN %(first_day)s AND %(last_day)s"]
    params = {"division": division, "first_day": first_day, "last_day": last_day}

    # docstatus: default include draft + submitted
    if docstatus in ("0", "1", 0, 1):
        conditions.append("ss.docstatus = %(docstatus)s"); params["docstatus"] = int(docstatus)
    else:
        conditions.append("ss.docstatus IN (0, 1)")

    if zone:
        conditions.append("ss.zone = %(zone)s"); params["zone"] = zone
    _scope_region_sql(conditions, params, "ss.region", division, region)
    if team:
        conditions.append("ss.team = %(team)s"); params["team"] = team
    if hq:
        conditions.append("ss.hq = %(hq)s"); params["hq"] = hq

    where = " AND ".join(conditions)
    rows = frappe.db.sql(f"""
        SELECT
            ss.stockist_code, ss.stockist_name, ss.statement_month,
            ss.hq, ss.team, ss.region, ss.zone,
            si.product_code, si.product_name, si.raw_product_name, si.pack,
            si.opening_qty, si.purchase_qty, si.sales_qty, si.free_qty,
            si.closing_qty, si.conversion_factor, si.pts,
            si.opening_value, si.sales_value_pts, si.sales_value_ptr, si.closing_value
        FROM `tabStockist Statement` ss
        INNER JOIN `tabStockist Statement Item` si
            ON si.parent = ss.name AND si.parenttype = 'Stockist Statement'
        WHERE {where}
        ORDER BY ss.stockist_name, si.product_code
    """, params, as_dict=True)

    if not rows:
        frappe.throw(f"No secondary sales data found for {str(month)[:7]} in {division} division")

    # Resolve org codes → display names (cached)
    team_names, region_names, zone_names, hq_names = {}, {}, {}, {}
    stk_codes = {}   # Stockist Master PK (S####) -> real stockist_code
    prod_meta = {}   # product_code -> {mrp, ptr, product_group, pts, pack, product_name}

    def _stk_code(pk):
        # ss.stockist_code is a Link to Stockist Master, so it holds the PK (S####),
        # not the human stockist code. Resolve to the real code for export.
        if not pk:
            return ""
        if pk not in stk_codes:
            stk_codes[pk] = frappe.db.get_value("Stockist Master", pk, "stockist_code") or pk
        return stk_codes[pk]

    def _hq_name(code):
        if not code:
            return ""
        if code not in hq_names:
            hq_names[code] = frappe.db.get_value("HQ Master", code, "hq_name") or code
        return hq_names[code]

    def _team_name(code):
        if not code:
            return ""
        if code not in team_names:
            team_names[code] = frappe.db.get_value("Team Master", code, "team_name") or code
        return team_names[code]

    def _region_name(code):
        if not code:
            return ""
        if code not in region_names:
            region_names[code] = frappe.db.get_value("Region Master", code, "region_name") or code
        return region_names[code]

    def _zone_name(code):
        if not code:
            return ""
        if code not in zone_names:
            zone_names[code] = frappe.db.get_value("Zone Master", code, "zone_name") or code
        return zone_names[code]

    def _prod(code):
        # si.product_code is a Link (Product Master id); fetch the business
        # product_code too so the export shows the human-facing code.
        if not code:
            return {}
        if code not in prod_meta:
            prod_meta[code] = frappe.db.get_value(
                "Product Master", code,
                ["product_code", "mrp", "ptr", "pts", "pack", "product_name", "product_group"],
                as_dict=True) or {}
        return prod_meta[code]

    # Output headers (match import template order)
    excel_headers = [
        "stockistcode", "stockistname", "citypool", "team", "region", "act_region",
        "zonee", "statementdate", "pcode", "product", "pack",
        "quantity", "clstock", "freeqty", "opstock",
        "pts", "ptsvalue", "ptr", "ptrvalue", "mrp", "mrpvalue",
        "nrv", "nrvvalue", "clsvalue", "opsvalue", "product_head",
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Secondary Sales {str(month)[:7]}"

    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="217346", end_color="217346", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"), bottom=Side(style="thin"))

    for col_idx, header in enumerate(excel_headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    stmt_date = str(last_day)
    for r_idx, row in enumerate(rows, 2):
        prod = _prod(row.get("product_code"))
        conv = flt(row.get("conversion_factor")) or 1
        sales_base = flt(row.get("sales_qty")) / conv
        # net_rate = the stored per-line valuation rate (NRV / final net rate, after any discount).
        # list_rate = the master PTS (standard price to stockist). Equal when no discount.
        net_rate = flt(row.get("pts")) or flt(prod.get("pts") or 0)
        list_rate = flt(prod.get("pts") or 0) or net_rate
        mrp_rate = flt(prod.get("mrp") or 0)
        ptr_rate = flt(prod.get("ptr") or 0)

        out = [
            _stk_code(row.get("stockist_code")),                  # real stockist code, not Master PK
            row.get("stockist_name") or "",
            _hq_name(row.get("hq")),                              # citypool (mapped from stockist HQ)
            _team_name(row.get("team")),
            _region_name(row.get("region")),
            _region_name(row.get("region")),                     # act_region ≈ region
            _zone_name(row.get("zone")),
            stmt_date,
            prod.get("product_code") or row.get("product_code") or "",   # real product code, not Master PK
            row.get("product_name") or row.get("raw_product_name") or "",
            row.get("pack") or prod.get("pack") or "",
            flt(row.get("sales_qty")),
            flt(row.get("closing_qty")),
            flt(row.get("free_qty")),
            flt(row.get("opening_qty")),
            list_rate,                                            # pts  (list / master rate)
            sales_base * list_rate,                               # ptsvalue (list value)
            ptr_rate,
            flt(row.get("sales_value_ptr")),                      # ptrvalue
            mrp_rate,
            sales_base * mrp_rate,                                # mrpvalue
            net_rate,                                             # nrv  (final net rate)
            flt(row.get("sales_value_pts")),                      # nrvvalue (= qty x net rate)
            flt(row.get("closing_value")),                        # clsvalue (net-based)
            flt(row.get("opening_value")),                        # opsvalue (net-based)
            prod.get("product_group") or "",                      # product_head
        ]
        for c_idx, val in enumerate(out, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = thin_border

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 3, 32)

    ws.freeze_panes = "A2"

    fname = f"Secondary_Sales_{division}_{str(month)[:7]}.xlsx"
    out_path = os.path.join(tempfile.gettempdir(), fname)
    wb.save(out_path)
    with open(out_path, "rb") as f:
        file_content = f.read()
    try:
        os.remove(out_path)
    except OSError:
        pass

    frappe.local.response.filename = fname
    frappe.local.response.filecontent = file_content
    frappe.local.response.type = "download"
