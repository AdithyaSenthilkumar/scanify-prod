import frappe
from frappe.model.document import Document
from frappe.utils import flt, getdate, now
from datetime import datetime
import json
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class DoctorSchemeSummaryReport(Document):
    def validate(self):
        """Validate report parameters"""
        if self.from_date and self.to_date:
            if getdate(self.from_date) > getdate(self.to_date):
                frappe.throw("From Date cannot be greater than To Date")
    
    def before_save(self):
        """Set audit fields"""
        if not self.report_date:
            self.report_date = now()
        if not self.generated_by:
            self.generated_by = frappe.session.user
    
    def on_submit(self):
        """Generate on submit"""
        self.generate_report_data()
        self.calculate_summary()
    
    def generate_report_data(self):
        """Generate doctor-wise scheme summary"""
        try:
            filters = self.build_filters()
            
            schemes = frappe.get_all(
                "Scheme Request",
                filters=filters,
                fields=["name", "doctor_code", "stockist_code", "entry_date", "docstatus"]
            )
            
            if not schemes:
                frappe.msgprint("No schemes found for selected filters", indicator='orange')
                self.report_data = json.dumps({}, default=str)
                return {}
            
            # Aggregate data based on group_by
            if self.group_by == "Doctor":
                summary_data = self.aggregate_by_doctor(schemes)
            elif self.group_by == "HQ":
                summary_data = self.aggregate_by_hq(schemes)
            elif self.group_by == "Team":
                summary_data = self.aggregate_by_team(schemes)
            else:  # Region
                summary_data = self.aggregate_by_region(schemes)
            
            self.report_data = json.dumps(summary_data, default=str)
            self.calculate_summary()
            
            return summary_data
            
        except Exception as e:
            frappe.log_error(frappe.get_traceback(), "Doctor Scheme Summary Error")
            frappe.throw(f"Error generating report: {str(e)}")

    def build_filters(self):
        """Build filters for Scheme Request"""
        filters = {}
        
        # Date filters - use entry_date instead of approval_date
        if self.from_date and self.to_date:
            filters["entry_date"] = ["between", [self.from_date, self.to_date]]
        elif self.from_date:
            filters["entry_date"] = [">=", self.from_date]
        elif self.to_date:
            filters["entry_date"] = ["<=", self.to_date]
        
        # Status filter
        if self.scheme_status == "Approved":
            filters["docstatus"] = 1
        elif self.scheme_status == "Pending":
            filters["docstatus"] = 0
        elif self.scheme_status == "Rejected":
            filters["docstatus"] = 2
        # "All" means no docstatus filter
        
        # Include/exclude cancelled
        if not self.include_cancelled:
            if "docstatus" not in filters:
                filters["docstatus"] = ["!=", 2]
        
        # Doctor filter
        if self.doctor:
            filters["doctor_code"] = self.doctor
        
        # Stockist filter
        if self.stockist:
            filters["stockist_code"] = self.stockist
        
        # HQ filter
        if self.hq:
            hq_team = frappe.db.get_value("HQ Master", self.hq, "team")
            if hq_team:
                doctors = frappe.get_all(
                    "Doctor Master",
                    filters={"team": hq_team, "status": "Active"},
                    pluck="doctor_code"
                )
                if doctors:
                    filters["doctor_code"] = ["in", doctors]
        
        # Team filter
        elif self.team:
            doctors = frappe.get_all(
                "Doctor Master",
                filters={"team": self.team, "status": "Active"},
                pluck="doctor_code"
            )
            if doctors:
                filters["doctor_code"] = ["in", doctors]
        
        # Region filter
        elif self.region:
            teams = frappe.get_all(
                "Team Master",
                filters={"region": self.region, "status": "Active"},
                pluck="name"
            )
            doctors = []
            for team in teams:
                team_doctors = frappe.get_all(
                    "Doctor Master",
                    filters={"team": team, "status": "Active"},
                    pluck="doctor_code"
                )
                doctors.extend(team_doctors)
            
            if doctors:
                filters["doctor_code"] = ["in", doctors]
        
        return filters

    
    def aggregate_by_doctor(self, schemes):
        """Aggregate data by doctor"""
        doctor_data = {}
        
        for scheme in schemes:
            scheme_doc = frappe.get_doc("Scheme Request", scheme["name"])
            doctor_code = scheme_doc.doctor_code
            
            # Get doctor details
            if not frappe.db.exists("Doctor Master", doctor_code):
                continue
            
            doctor = frappe.get_doc("Doctor Master", doctor_code)
            
            # Initialize doctor entry
            if doctor_code not in doctor_data:
                doctor_data[doctor_code] = {
                    "doctor_code": doctor_code,
                    "doctor_name": doctor.doctor_name,
                    "team": doctor.team if hasattr(doctor, 'team') else "N/A",
                    "region": doctor.region if hasattr(doctor, 'region') else "N/A",
                    "total_schemes": 0,
                    "approved_schemes": 0,
                    "pending_schemes": 0,
                    "rejected_schemes": 0,
                    "total_approved_qty": 0,
                    "total_reflected_qty": 0,
                    "total_scheme_value": 0,
                    "products": {},
                    "stockists": {}
                }
            
            # Count schemes by status
            doctor_data[doctor_code]["total_schemes"] += 1
            if scheme_doc.docstatus == 1:
                doctor_data[doctor_code]["approved_schemes"] += 1
            elif scheme_doc.docstatus == 0:
                doctor_data[doctor_code]["pending_schemes"] += 1
            elif scheme_doc.docstatus == 2:
                doctor_data[doctor_code]["rejected_schemes"] += 1
            
            # Process items
            for item in scheme_doc.items:
                product_code = item.product_code
                
                # Apply product filter
                if self.product_code and product_code != self.product_code:
                    continue
                
                free_qty = flt(item.free_quantity)
                
                # Get product details
                product = frappe.db.get_value(
                    "Product Master",
                    product_code,
                    ["product_name", "pack", "pts"],
                    as_dict=True
                )
                
                if not product:
                    continue
                
                scheme_value = free_qty * flt(product.pts)
                
                doctor_data[doctor_code]["total_approved_qty"] += free_qty
                doctor_data[doctor_code]["total_scheme_value"] += scheme_value
                
                # Product breakdown
                if self.show_product_breakdown:
                    if product_code not in doctor_data[doctor_code]["products"]:
                        doctor_data[doctor_code]["products"][product_code] = {
                            "product_name": product.product_name,
                            "pack": product.pack,
                            "approved_qty": 0,
                            "reflected_qty": 0,
                            "scheme_count": 0
                        }
                    
                    doctor_data[doctor_code]["products"][product_code]["approved_qty"] += free_qty
                    doctor_data[doctor_code]["products"][product_code]["scheme_count"] += 1
                
                # Get reflected quantity
                reflected_qty = self.get_reflected_qty(
                    scheme_doc.stockist_code,
                    product_code,
                    self.from_date,
                    self.to_date
                )
                
                doctor_data[doctor_code]["total_reflected_qty"] += reflected_qty
                
                if self.show_product_breakdown:
                    doctor_data[doctor_code]["products"][product_code]["reflected_qty"] += reflected_qty
            
            # Stockist breakdown
            if self.show_stockist_breakdown:
                stockist_code = scheme_doc.stockist_code
                if stockist_code not in doctor_data[doctor_code]["stockists"]:
                    stockist_name = frappe.db.get_value("Stockist Master", stockist_code, "stockist_name")
                    doctor_data[doctor_code]["stockists"][stockist_code] = {
                        "stockist_name": stockist_name,
                        "scheme_count": 0
                    }
                
                doctor_data[doctor_code]["stockists"][stockist_code]["scheme_count"] += 1
        
        # Calculate reflection rates
        for doctor_code in doctor_data:
            total_approved = doctor_data[doctor_code]["total_approved_qty"]
            total_reflected = doctor_data[doctor_code]["total_reflected_qty"]
            
            if total_approved > 0:
                doctor_data[doctor_code]["reflection_rate"] = (total_reflected / total_approved) * 100
            else:
                doctor_data[doctor_code]["reflection_rate"] = 0
        
        return doctor_data
    
    def aggregate_by_hq(self, schemes):
        """Aggregate data by HQ"""
        hq_data = {}
        
        for scheme in schemes:
            scheme_doc = frappe.get_doc("Scheme Request", scheme["name"])
            doctor = frappe.get_doc("Doctor Master", scheme_doc.doctor_code)
            
            # Get HQ from stockist
            stockist = frappe.get_doc("Stockist Master", scheme_doc.stockist_code)
            hq = stockist.hq if hasattr(stockist, 'hq') else "Unknown"
            
            if hq not in hq_data:
                hq_data[hq] = {
                    "hq": hq,
                    "total_schemes": 0,
                    "total_doctors": set(),
                    "total_approved_qty": 0,
                    "total_reflected_qty": 0,
                    "total_scheme_value": 0
                }
            
            hq_data[hq]["total_schemes"] += 1
            hq_data[hq]["total_doctors"].add(scheme_doc.doctor_code)
            
            # Process items
            for item in scheme_doc.items:
                if self.product_code and item.product_code != self.product_code:
                    continue
                
                free_qty = flt(item.free_quantity)
                product = frappe.db.get_value("Product Master", item.product_code, ["pts"], as_dict=True)
                
                if product:
                    hq_data[hq]["total_approved_qty"] += free_qty
                    hq_data[hq]["total_scheme_value"] += free_qty * flt(product.pts)
                    
                    reflected_qty = self.get_reflected_qty(
                        scheme_doc.stockist_code,
                        item.product_code,
                        self.from_date,
                        self.to_date
                    )
                    hq_data[hq]["total_reflected_qty"] += reflected_qty
        
        # Convert sets to counts
        for hq in hq_data:
            hq_data[hq]["total_doctors"] = len(hq_data[hq]["total_doctors"])
            
            total_approved = hq_data[hq]["total_approved_qty"]
            total_reflected = hq_data[hq]["total_reflected_qty"]
            
            if total_approved > 0:
                hq_data[hq]["reflection_rate"] = (total_reflected / total_approved) * 100
            else:
                hq_data[hq]["reflection_rate"] = 0
        
        return hq_data
    
    def aggregate_by_team(self, schemes):
        """Aggregate data by team"""
        team_data = {}
        
        for scheme in schemes:
            scheme_doc = frappe.get_doc("Scheme Request", scheme["name"])
            doctor = frappe.get_doc("Doctor Master", scheme_doc.doctor_code)
            
            team = doctor.team if hasattr(doctor, 'team') else "Unknown"
            
            if team not in team_data:
                team_data[team] = {
                    "team": team,
                    "total_schemes": 0,
                    "total_doctors": set(),
                    "total_approved_qty": 0,
                    "total_reflected_qty": 0,
                    "total_scheme_value": 0
                }
            
            team_data[team]["total_schemes"] += 1
            team_data[team]["total_doctors"].add(scheme_doc.doctor_code)
            
            # Process items
            for item in scheme_doc.items:
                if self.product_code and item.product_code != self.product_code:
                    continue
                
                free_qty = flt(item.free_quantity)
                product = frappe.db.get_value("Product Master", item.product_code, ["pts"], as_dict=True)
                
                if product:
                    team_data[team]["total_approved_qty"] += free_qty
                    team_data[team]["total_scheme_value"] += free_qty * flt(product.pts)
                    
                    reflected_qty = self.get_reflected_qty(
                        scheme_doc.stockist_code,
                        item.product_code,
                        self.from_date,
                        self.to_date
                    )
                    team_data[team]["total_reflected_qty"] += reflected_qty
        
        # Convert sets and calculate rates
        for team in team_data:
            team_data[team]["total_doctors"] = len(team_data[team]["total_doctors"])
            
            total_approved = team_data[team]["total_approved_qty"]
            total_reflected = team_data[team]["total_reflected_qty"]
            
            if total_approved > 0:
                team_data[team]["reflection_rate"] = (total_reflected / total_approved) * 100
            else:
                team_data[team]["reflection_rate"] = 0
        
        return team_data
    
    def aggregate_by_region(self, schemes):
        """Aggregate data by region"""
        region_data = {}
        
        for scheme in schemes:
            scheme_doc = frappe.get_doc("Scheme Request", scheme["name"])
            doctor = frappe.get_doc("Doctor Master", scheme_doc.doctor_code)
            
            region = doctor.region if hasattr(doctor, 'region') else "Unknown"
            
            if region not in region_data:
                region_data[region] = {
                    "region": region,
                    "total_schemes": 0,
                    "total_doctors": set(),
                    "total_approved_qty": 0,
                    "total_reflected_qty": 0,
                    "total_scheme_value": 0
                }
            
            region_data[region]["total_schemes"] += 1
            region_data[region]["total_doctors"].add(scheme_doc.doctor_code)
            
            # Process items
            for item in scheme_doc.items:
                if self.product_code and item.product_code != self.product_code:
                    continue
                
                free_qty = flt(item.free_quantity)
                product = frappe.db.get_value("Product Master", item.product_code, ["pts"], as_dict=True)
                
                if product:
                    region_data[region]["total_approved_qty"] += free_qty
                    region_data[region]["total_scheme_value"] += free_qty * flt(product.pts)
                    
                    reflected_qty = self.get_reflected_qty(
                        scheme_doc.stockist_code,
                        item.product_code,
                        self.from_date,
                        self.to_date
                    )
                    region_data[region]["total_reflected_qty"] += reflected_qty
        
        # Convert sets and calculate rates
        for region in region_data:
            region_data[region]["total_doctors"] = len(region_data[region]["total_doctors"])
            
            total_approved = region_data[region]["total_approved_qty"]
            total_reflected = region_data[region]["total_reflected_qty"]
            
            if total_approved > 0:
                region_data[region]["reflection_rate"] = (total_reflected / total_approved) * 100
            else:
                region_data[region]["reflection_rate"] = 0
        
        return region_data
    
    def get_reflected_qty(self, stockist_code, product_code, from_date, to_date):
        """Get reflected quantity in secondary sales"""
        try:
            statements = frappe.get_all(
                "Stockist Statement",
                filters={
                    "stockist_code": stockist_code,
                    "statement_month": ["between", [from_date, to_date]],
                    "docstatus": 1
                },
                pluck="name"
            )
            
            if not statements:
                return 0
            
            total_sales = 0
            for statement in statements:
                sales_qty = frappe.db.get_value(
                    "Stockist Statement Item",
                    {
                        "parent": statement,
                        "product_code": product_code
                    },
                    "sales_qty"
                )
                total_sales += flt(sales_qty or 0)
            
            return total_sales
        
        except Exception:
            return 0
    
    def calculate_summary(self):
        """Calculate summary statistics"""
        if not self.report_data:
            return
        
        try:
            data = json.loads(self.report_data)
            
            total_doctors = 0
            total_schemes = 0
            total_approved_qty = 0
            total_reflected_qty = 0
            total_scheme_value = 0
            
            for key in data:
                entry = data[key]
                
                if self.group_by == "Doctor":
                    total_doctors += 1
                else:
                    total_doctors += entry.get("total_doctors", 0)
                
                total_schemes += entry.get("total_schemes", 0)
                total_approved_qty += flt(entry.get("total_approved_qty", 0))
                total_reflected_qty += flt(entry.get("total_reflected_qty", 0))
                total_scheme_value += flt(entry.get("total_scheme_value", 0))
            
            self.total_doctors = total_doctors
            self.total_schemes = total_schemes
            self.total_approved_qty = total_approved_qty
            self.total_reflected_qty = total_reflected_qty
            self.total_scheme_value = total_scheme_value
            
            if total_approved_qty > 0:
                self.average_reflection_rate = (total_reflected_qty / total_approved_qty) * 100
            else:
                self.average_reflection_rate = 0
        
        except Exception as e:
            frappe.log_error(f"Error calculating summary: {str(e)}", "Summary Error")


@frappe.whitelist()
def generate_report(doc_name):
    """API to generate report"""
    doc = frappe.get_doc("Doctor Scheme Summary Report", doc_name)
    doc.generate_report_data()
    doc.save()
    frappe.db.commit()
    
    return {
        "success": True,
        "message": "Report generated successfully",
        "total_doctors": doc.total_doctors,
        "total_schemes": doc.total_schemes,
        "reflection_rate": doc.average_reflection_rate
    }


@frappe.whitelist()
def export_to_excel(doc_name):
    """Export to Excel"""
    doc = frappe.get_doc("Doctor Scheme Summary Report", doc_name)
    
    if not doc.report_data:
        frappe.throw("No report data available. Please generate the report first.")
    
    data = json.loads(doc.report_data or "{}")
    
    if not data:
        frappe.throw("No data available to export")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{doc.group_by} Summary"
    
    # Styling
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:J1')
    ws['A1'] = "STEDMAN PHARMACEUTICALS PVT LTD"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:J2')
    ws['A2'] = f"Doctor Scheme Summary Report - Group By {doc.group_by}"
    ws['A2'].font = Font(size=12, italic=True)
    ws['A2'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A3:J3')
    ws['A3'] = f"Period: {doc.from_date} to {doc.to_date}"
    ws['A3'].font = Font(size=10)
    ws['A3'].alignment = Alignment(horizontal='center')
    
    # Headers based on group_by
    row = 5
    
    if doc.group_by == "Doctor":
        headers = ["Doctor Code", "Doctor Name", "Team", "Total Schemes", 
                  "Approved Qty", "Reflected Qty", "Reflection %", "Scheme Value"]
    else:
        headers = [doc.group_by, "Total Doctors", "Total Schemes", 
                  "Approved Qty", "Reflected Qty", "Reflection %", "Scheme Value"]
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Data rows
    row = 6
    
    for key in sorted(data.keys()):
        entry = data[key]
        
        if doc.group_by == "Doctor":
            ws.cell(row=row, column=1).value = entry["doctor_code"]
            ws.cell(row=row, column=2).value = entry["doctor_name"]
            ws.cell(row=row, column=3).value = entry["team"]
            ws.cell(row=row, column=4).value = entry["total_schemes"]
            ws.cell(row=row, column=5).value = flt(entry["total_approved_qty"])
            ws.cell(row=row, column=6).value = flt(entry["total_reflected_qty"])
            ws.cell(row=row, column=7).value = flt(entry["reflection_rate"]) / 100
            ws.cell(row=row, column=8).value = flt(entry["total_scheme_value"])
        else:
            ws.cell(row=row, column=1).value = key
            ws.cell(row=row, column=2).value = entry.get("total_doctors", 0)
            ws.cell(row=row, column=3).value = entry["total_schemes"]
            ws.cell(row=row, column=4).value = flt(entry["total_approved_qty"])
            ws.cell(row=row, column=5).value = flt(entry["total_reflected_qty"])
            ws.cell(row=row, column=6).value = flt(entry["reflection_rate"]) / 100
            ws.cell(row=row, column=7).value = flt(entry["total_scheme_value"])
        
        # Format cells
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            if col >= 4:
                cell.alignment = Alignment(horizontal='right')
            if col == 7:
                cell.number_format = '0.00%'
            elif col >= 5:
                cell.number_format = '#,##0.00'
        
        row += 1
    
    # Totals
    summary_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    summary_font = Font(bold=True)
    
    ws.cell(row=row, column=1).value = "TOTAL"
    
    if doc.group_by == "Doctor":
        ws.cell(row=row, column=4).value = doc.total_schemes
        ws.cell(row=row, column=5).value = doc.total_approved_qty
        ws.cell(row=row, column=6).value = doc.total_reflected_qty
        ws.cell(row=row, column=7).value = doc.average_reflection_rate / 100
        ws.cell(row=row, column=8).value = doc.total_scheme_value
    else:
        ws.cell(row=row, column=2).value = doc.total_doctors
        ws.cell(row=row, column=3).value = doc.total_schemes
        ws.cell(row=row, column=4).value = doc.total_approved_qty
        ws.cell(row=row, column=5).value = doc.total_reflected_qty
        ws.cell(row=row, column=6).value = doc.average_reflection_rate / 100
        ws.cell(row=row, column=7).value = doc.total_scheme_value
    
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = summary_fill
        cell.font = summary_font
        cell.border = border
    
    # Adjust widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    # Save
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    file_doc = frappe.get_doc({
        "doctype": "File",
        "file_name": f"Doctor_Scheme_Summary_{doc.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        "attached_to_doctype": "Doctor Scheme Summary Report",
        "attached_to_name": doc_name,
        "content": output.getvalue(),
        "is_private": 0
    })
    file_doc.save(ignore_permissions=True)
    
    return {
        "success": True,
        "file_url": file_doc.file_url
    }


@frappe.whitelist()
def export_to_pdf(doc_name):
    """Export to PDF"""
    from frappe.utils.pdf import get_pdf
    
    doc = frappe.get_doc("Doctor Scheme Summary Report", doc_name)
    
    if not doc.report_data:
        frappe.throw("No report data available. Please generate the report first.")
    
    html = render_report_html(doc)
    pdf_bytes = get_pdf(html)
    
    file_doc = frappe.get_doc({
        "doctype": "File",
        "file_name": f"Doctor_Scheme_Summary_{doc.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
        "attached_to_doctype": "Doctor Scheme Summary Report",
        "attached_to_name": doc_name,
        "content": pdf_bytes,
        "is_private": 0
    })
    file_doc.save(ignore_permissions=True)
    
    return {
        "success": True,
        "file_url": file_doc.file_url
    }


def render_report_html(doc):
    """Render HTML for PDF"""
    data = json.loads(doc.report_data or "{}")
    
    html = f"""
    <html>
    <head>
        <style>
            @page {{ size: A4 landscape; margin: 15mm; }}
            body {{ font-family: Arial, sans-serif; margin: 0; font-size: 9px; }}
            .header {{ text-align: center; margin-bottom: 15px; }}
            .title {{ font-size: 14px; font-weight: bold; }}
            .subtitle {{ font-size: 10px; color: #666; }}
            table {{ width: 100%; border-collapse: collapse; }}
            th {{ background-color: #1F4E78; color: white; padding: 5px; text-align: center; border: 1px solid #000; font-size: 8px; }}
            td {{ padding: 4px; border: 1px solid #ccc; font-size: 8px; }}
            .number {{ text-align: right; }}
            .text-left {{ text-align: left; }}
            .total-row {{ background-color: #D9E1F2; font-weight: bold; }}
        </style>
    </head>
    <body>
        <div class="header">
            <div class="title">STEDMAN PHARMACEUTICALS PVT LTD</div>
            <div class="subtitle">Doctor Scheme Summary Report - Group By {doc.group_by}</div>
            <div class="subtitle">Period: {doc.from_date} to {doc.to_date}</div>
        </div>
        
        <table>
            <thead>
                <tr>
    """
    
    if doc.group_by == "Doctor":
        html += """
                    <th>Doctor</th>
                    <th>Team</th>
                    <th>Schemes</th>
                    <th>Approved Qty</th>
                    <th>Reflected Qty</th>
                    <th>Reflection %</th>
                    <th>Value</th>
        """
    else:
        html += f"""
                    <th>{doc.group_by}</th>
                    <th>Doctors</th>
                    <th>Schemes</th>
                    <th>Approved Qty</th>
                    <th>Reflected Qty</th>
                    <th>Reflection %</th>
                    <th>Value</th>
        """
    
    html += """
                </tr>
            </thead>
            <tbody>
    """
    
    for key in sorted(data.keys()):
        entry = data[key]
        
        if doc.group_by == "Doctor":
            html += f"""
            <tr>
                <td class='text-left'>{entry['doctor_name']}</td>
                <td class='text-left'>{entry['team']}</td>
                <td class='number'>{entry['total_schemes']}</td>
                <td class='number'>{flt(entry['total_approved_qty']):.0f}</td>
                <td class='number'>{flt(entry['total_reflected_qty']):.0f}</td>
                <td class='number'>{flt(entry['reflection_rate']):.2f}%</td>
                <td class='number'>{flt(entry['total_scheme_value']):.2f}</td>
            </tr>
            """
        else:
            html += f"""
            <tr>
                <td class='text-left'>{key}</td>
                <td class='number'>{entry.get('total_doctors', 0)}</td>
                <td class='number'>{entry['total_schemes']}</td>
                <td class='number'>{flt(entry['total_approved_qty']):.0f}</td>
                <td class='number'>{flt(entry['total_reflected_qty']):.0f}</td>
                <td class='number'>{flt(entry['reflection_rate']):.2f}%</td>
                <td class='number'>{flt(entry['total_scheme_value']):.2f}</td>
            </tr>
            """
    
    if doc.group_by == "Doctor":
        html += f"""
            <tr class='total-row'>
                <td colspan='2' class='text-left'>TOTAL</td>
                <td class='number'>{doc.total_schemes}</td>
                <td class='number'>{flt(doc.total_approved_qty):.0f}</td>
                <td class='number'>{flt(doc.total_reflected_qty):.0f}</td>
                <td class='number'>{flt(doc.average_reflection_rate):.2f}%</td>
                <td class='number'>{flt(doc.total_scheme_value):.2f}</td>
            </tr>
        """
    else:
        html += f"""
            <tr class='total-row'>
                <td class='text-left'>TOTAL</td>
                <td class='number'>{doc.total_doctors}</td>
                <td class='number'>{doc.total_schemes}</td>
                <td class='number'>{flt(doc.total_approved_qty):.0f}</td>
                <td class='number'>{flt(doc.total_reflected_qty):.0f}</td>
                <td class='number'>{flt(doc.average_reflection_rate):.2f}%</td>
                <td class='number'>{flt(doc.total_scheme_value):.2f}</td>
            </tr>
        """
    
    html += """
            </tbody>
        </table>
    </body>
    </html>
    """
    
    return html
