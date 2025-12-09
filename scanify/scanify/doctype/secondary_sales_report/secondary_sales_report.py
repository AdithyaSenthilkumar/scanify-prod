import frappe
from frappe.model.document import Document
from frappe.utils import flt, getdate, add_months, get_first_day, get_last_day, now
from datetime import datetime
import json
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class SecondarySalesReport(Document):
    def validate(self):
        """Validate report parameters"""
        if self.from_date and self.to_date:
            if getdate(self.from_date) > getdate(self.to_date):
                frappe.throw("From Date cannot be greater than To Date")
    
    def before_save(self):
        """Set audit fields"""
        if not self.report_date:
            self.report_date = now()
        if not self.generated_by:
            self.generated_by = frappe.session.user
    
    def on_submit(self):
        """Regenerate on submit"""
        self.generate_report_data()
        self.calculate_totals()
    @frappe.whitelist()
    def generate_report_data(self):
        """Generate secondary sales data based on filters"""
        try:
            filters = self.build_filters()
            
            # Fetch raw stockist statement data
            statements = frappe.get_all(
                "Stockist Statement",
                filters=filters,
                fields=[
                    "name", "stockist_code", "statement_month", 
                    "from_date", "to_date", "docstatus"
                ]
            )
            
            if not statements:
                frappe.msgprint("No data found for selected filters", indicator='orange')
                self.report_data = json.dumps({}, default=str)
                return {}
            
            # Aggregate data based on report type
            aggregated_data = self.aggregate_sales_data(statements)
            
            # Apply scheme deduction if enabled
            if self.include_scheme_deduction:
                aggregated_data = self.apply_scheme_deduction(aggregated_data)
            
            # Store as JSON
            self.report_data = json.dumps(aggregated_data, default=str)
            
            # Calculate totals
            self.calculate_totals()
            
            return aggregated_data
            
        except Exception as e:
            frappe.log_error(frappe.get_traceback(), "Report Generation Error")
            frappe.throw(f"Error generating report: {str(e)}")
    
    def build_filters(self):
        """Build Stockist Statement filters based on report parameters"""
        filters = {
            "docstatus": 1,  # Only submitted statements
        }
        
        # Date filters
        if self.from_date and self.to_date:
            filters["statement_month"] = ["between", [self.from_date, self.to_date]]
        elif self.from_date:
            filters["statement_month"] = [">=", self.from_date]
        elif self.to_date:
            filters["statement_month"] = ["<=", self.to_date]
        
        # Hierarchy filters
        if self.stockist:
            filters["stockist_code"] = self.stockist
        elif self.hq:
            # Get all stockists for this HQ
            stockists = frappe.get_all(
                "Stockist Master",
                filters={"hq": self.hq, "status": "Active"},
                fields=["stockist_code"]
            )
            if stockists:
                stockist_codes = [s["stockist_code"] for s in stockists]
                filters["stockist_code"] = ["in", stockist_codes]
        elif self.team:
            # Get all HQs for this team, then all stockists
            hqs = frappe.get_all(
                "HQ Master",
                filters={"team": self.team, "status": "Active"},
                fields=["name"]
            )
            stockist_list = []
            for hq in hqs:
                stockists = frappe.get_all(
                    "Stockist Master",
                    filters={"hq": hq["name"], "status": "Active"},
                    fields=["stockist_code"]
                )
                stockist_list.extend([s["stockist_code"] for s in stockists])
            
            if stockist_list:
                filters["stockist_code"] = ["in", stockist_list]
        elif self.region:
            # Get all teams for this region
            teams = frappe.get_all(
                "Team Master",
                filters={"region": self.region, "status": "Active"},
                fields=["name"]
            )
            stockist_list = []
            for team in teams:
                hqs = frappe.get_all(
                    "HQ Master",
                    filters={"team": team["name"], "status": "Active"},
                    fields=["name"]
                )
                for hq in hqs:
                    stockists = frappe.get_all(
                        "Stockist Master",
                        filters={"hq": hq["name"], "status": "Active"},
                        fields=["stockist_code"]
                    )
                    stockist_list.extend([s["stockist_code"] for s in stockists])
            
            if stockist_list:
                filters["stockist_code"] = ["in", stockist_list]
        
        return filters
    
    def aggregate_sales_data(self, statements):
        """
        Aggregate statement items based on report_type
        Calculate values from Stockist Statement Item fields
        """
        data_map = {}
        
        for statement in statements:
            # Fetch full statement with items
            stmt_doc = frappe.get_doc("Stockist Statement", statement["name"])
            
            # Check if stockist exists
            if not frappe.db.exists("Stockist Master", stmt_doc.stockist_code):
                frappe.log_error(f"Stockist {stmt_doc.stockist_code} not found", "Missing Stockist")
                continue
            
            stockist = frappe.get_doc("Stockist Master", stmt_doc.stockist_code)
            
            # Determine grouping key
            group_key = self.get_group_key(stockist)
            
            if group_key not in data_map:
                data_map[group_key] = {
                    "opening_value": 0,
                    "primary_value": 0,
                    "monthly_sales": {},
                    "closing_value": 0,
                    "stockist_details": []
                }
            
            # Calculate values from items
            month_key = str(stmt_doc.statement_month)
            
            opening_value = 0
            purchase_value = 0  # Primary sales
            sales_value = 0      # Secondary sales
            closing_value = 0
            
            for item in stmt_doc.items:
                # Get product details for pricing
                product_code = item.product_code
                
                # Fetch PTS rate from Product Master
                product = frappe.db.get_value(
                    "Product Master",
                    {"product_code": product_code},
                    ["pts", "ptr", "mrp"],
                    as_dict=True
                )
                
                if not product:
                    frappe.log_error(f"Product {product_code} not found", "Missing Product")
                    continue
                
                pts_rate = flt(product.pts or 0)
                
                # Calculate values
                opening_value += flt(item.opening_qty) * pts_rate
                purchase_value += flt(item.purchase_qty) * pts_rate
                sales_value += flt(item.sales_qty) * pts_rate
                closing_value += flt(item.closing_qty) * pts_rate
            
            # Store monthly sales
            if month_key in data_map[group_key]["monthly_sales"]:
                data_map[group_key]["monthly_sales"][month_key] += sales_value
            else:
                data_map[group_key]["monthly_sales"][month_key] = sales_value
            
            # Accumulate totals (use latest values for opening/closing)
            data_map[group_key]["opening_value"] = opening_value
            data_map[group_key]["primary_value"] += purchase_value
            data_map[group_key]["closing_value"] = closing_value
            
            # Track stockist details for reference
            stockist_detail = {
                "code": stockist.stockist_code,
                "name": stockist.stockist_name,
                "hq": stockist.hq
            }
            
            if stockist_detail not in data_map[group_key]["stockist_details"]:
                data_map[group_key]["stockist_details"].append(stockist_detail)
        
        return data_map
    
    def get_group_key(self, stockist):
        """Get grouping key based on report type"""
        try:
            if self.report_type == "Stockist Wise":
                return f"{stockist.stockist_name} / HQ: {stockist.hq}"
            
            elif self.report_type == "HQ Wise":
                return stockist.hq
            
            elif self.report_type == "Team Wise":
                if frappe.db.exists("HQ Master", stockist.hq):
                    hq = frappe.get_doc("HQ Master", stockist.hq)
                    return hq.team
                else:
                    return "Unknown Team"
            
            elif self.report_type == "Region Wise":
                if frappe.db.exists("HQ Master", stockist.hq):
                    hq = frappe.get_doc("HQ Master", stockist.hq)
                    if frappe.db.exists("Team Master", hq.team):
                        team = frappe.get_doc("Team Master", hq.team)
                        return team.region
                    else:
                        return "Unknown Region"
                else:
                    return "Unknown Region"
            
            return "Unknown"
        
        except Exception as e:
            frappe.log_error(f"Error getting group key: {str(e)}", "Group Key Error")
            return "Unknown"
    
    def apply_scheme_deduction(self, data_map):
        """Apply scheme deduction to secondary sales"""
        if not self.include_scheme_deduction or not self.scheme_deduction_value:
            return data_map
        
        for group_key in data_map:
            monthly_sales = data_map[group_key]["monthly_sales"]
            total_sales = sum(monthly_sales.values())
            
            if total_sales > 0:
                deduction_ratio = flt(self.scheme_deduction_value) / total_sales
                
                for month_key in monthly_sales:
                    original_value = monthly_sales[month_key]
                    deducted_value = original_value - (original_value * deduction_ratio)
                    monthly_sales[month_key] = max(deducted_value, 0)
        
        return data_map
    
    def calculate_totals(self):
        """Calculate summary totals"""
        if not self.report_data:
            return
        
        try:
            data_map = json.loads(self.report_data)
            
            total_opening = 0
            total_primary = 0
            total_secondary = 0
            total_closing = 0
            
            for group_key in data_map:
                group_data = data_map[group_key]
                total_opening += flt(group_data.get("opening_value", 0))
                total_primary += flt(group_data.get("primary_value", 0))
                total_secondary += sum(flt(v) for v in group_data.get("monthly_sales", {}).values())
                total_closing += flt(group_data.get("closing_value", 0))
            
            self.total_opening_value = total_opening
            self.total_primary_value = total_primary
            self.total_secondary_sales = total_secondary
            self.total_closing_value = total_closing
            
            if self.include_scheme_deduction:
                self.total_scheme_deducted = flt(self.scheme_deduction_value or 0)
        
        except Exception as e:
            frappe.log_error(f"Error calculating totals: {str(e)}", "Totals Calculation Error")


@frappe.whitelist()
def generate_report(doc_name):
    """API method to generate report data"""
    doc = frappe.get_doc("Secondary Sales Report", doc_name)
    doc.generate_report_data()
    doc.save()
    frappe.db.commit()
    
    return {
        "success": True,
        "message": "Report generated successfully",
        "total_opening": doc.total_opening_value,
        "total_primary": doc.total_primary_value,
        "total_secondary": doc.total_secondary_sales,
        "total_closing": doc.total_closing_value
    }


@frappe.whitelist()
def export_to_pdf(doc_name):
    """Export report to PDF with formatting"""
    from frappe.utils.pdf import get_pdf
    
    doc = frappe.get_doc("Secondary Sales Report", doc_name)
    
    if not doc.report_data:
        frappe.throw("No report data available. Please generate the report first.")
    
    # HTML template for PDF
    html = render_report_html(doc)
    
    pdf_bytes = get_pdf(html)
    
    # Save as attachment
    file_doc = frappe.get_doc({
        "doctype": "File",
        "file_name": f"Secondary_Sales_{doc.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
        "attached_to_doctype": "Secondary Sales Report",
        "attached_to_name": doc_name,
        "content": pdf_bytes,
        "is_private": 0
    })
    file_doc.save(ignore_permissions=True)
    
    return {
        "success": True,
        "file_url": file_doc.file_url
    }


@frappe.whitelist()
def export_to_excel(doc_name):
    """Export report to Excel with professional formatting"""
    doc = frappe.get_doc("Secondary Sales Report", doc_name)
    
    if not doc.report_data:
        frappe.throw("No report data available. Please generate the report first.")
    
    data_map = json.loads(doc.report_data or "{}")
    
    if not data_map:
        frappe.throw("No data available to export")
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Secondary Sales"
    
    # Header styling
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title section
    ws.merge_cells('A1:O1')
    ws['A1'] = "STEDMAN PHARMACEUTICALS PVT LTD"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('A2:O2')
    ws['A2'] = f"{doc.report_type} - Secondary Sales Report"
    ws['A2'].font = Font(size=12, italic=True)
    ws['A2'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A3:O3')
    ws['A3'] = f"Period: {doc.from_date} to {doc.to_date}"
    ws['A3'].font = Font(size=10)
    ws['A3'].alignment = Alignment(horizontal='center')
    
    # Column headers (month names without year for cleaner look)
    row = 5
    headers = ["Customers", "Opening Value", "Primary Value", 
               "Apr-25", "May-25", "Jun-25", "Jul-25", "Aug-25", "Sep-25", 
               "Oct-25", "Nov-25", "Dec-25", "Jan-26", "Feb-26", "Mar-26",
               "Year", "Closing Value"]
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # Data rows
    row = 6
    for group_key in sorted(data_map.keys()):
        group_data = data_map[group_key]
        
        ws.cell(row=row, column=1).value = group_key
        ws.cell(row=row, column=2).value = flt(group_data.get("opening_value", 0))
        ws.cell(row=row, column=3).value = flt(group_data.get("primary_value", 0))
        
        # Monthly sales
        monthly_sales = group_data.get("monthly_sales", {})
        months = ["04", "05", "06", "07", "08", "09", "10", "11", "12", "01", "02", "03"]
        year_total = 0
        
        for col_idx, month in enumerate(months, 4):
            value = 0
            # Match month in date string (format: YYYY-MM-DD)
            for month_key in monthly_sales.keys():
                if f"-{month}-" in month_key or month_key.split("-")[1] == month:
                    value = flt(monthly_sales[month_key])
                    break
            
            ws.cell(row=row, column=col_idx).value = value
            year_total += value
        
        ws.cell(row=row, column=16).value = year_total
        ws.cell(row=row, column=17).value = flt(group_data.get("closing_value", 0))
        
        # Apply formatting to data row
        for col in range(1, 18):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            if col > 1:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
        
        row += 1
    
    # Totals row
    totals_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    totals_font = Font(bold=True, size=11)
    
    ws.cell(row=row, column=1).value = f"{doc.report_type} Total"
    ws.cell(row=row, column=2).value = flt(doc.total_opening_value)
    ws.cell(row=row, column=3).value = flt(doc.total_primary_value)
    ws.cell(row=row, column=16).value = flt(doc.total_secondary_sales)
    ws.cell(row=row, column=17).value = flt(doc.total_closing_value)
    
    for col in range(1, 18):
        cell = ws.cell(row=row, column=col)
        cell.fill = totals_fill
        cell.font = totals_font
        cell.border = border
        if col > 1:
            cell.number_format = '#,##0.00'
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 45
    for col in range(2, 18):
        ws.column_dimensions[get_column_letter(col)].width = 12
    
    # Freeze panes (header rows)
    ws.freeze_panes = 'A6'
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Save as file
    file_doc = frappe.get_doc({
        "doctype": "File",
        "file_name": f"Secondary_Sales_{doc.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        "attached_to_doctype": "Secondary Sales Report",
        "attached_to_name": doc_name,
        "content": output.getvalue(),
        "is_private": 0
    })
    file_doc.save(ignore_permissions=True)
    
    return {
        "success": True,
        "file_url": file_doc.file_url
    }


def render_report_html(doc):
    """Render HTML for PDF export"""
    data_map = json.loads(doc.report_data or "{}")
    
    html = f"""
    <html>
    <head>
        <style>
            @page {{ size: A4 landscape; margin: 15mm; }}
            body {{ font-family: Arial, sans-serif; margin: 0; font-size: 9px; }}
            .header {{ text-align: center; margin-bottom: 15px; }}
            .title {{ font-size: 14px; font-weight: bold; margin-bottom: 3px; }}
            .subtitle {{ font-size: 10px; color: #666; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 10px; }}
            th {{ background-color: #1F4E78; color: white; padding: 5px; text-align: center; border: 1px solid #000; font-size: 8px; }}
            td {{ padding: 4px; border: 1px solid #ccc; font-size: 8px; }}
            .total-row {{ background-color: #D9E1F2; font-weight: bold; }}
            .number {{ text-align: right; }}
            .text-left {{ text-align: left; }}
        </style>
    </head>
    <body>
        <div class="header">
            <div class="title">STEDMAN PHARMACEUTICALS PVT LTD</div>
            <div class="subtitle">{doc.report_type} - Secondary Sales Report</div>
            <div class="subtitle">Period: {doc.from_date} to {doc.to_date}</div>
        </div>
        
        <table>
            <thead>
                <tr>
                    <th>Customers</th>
                    <th>Opening</th>
                    <th>Primary</th>
                    <th>Apr</th><th>May</th><th>Jun</th><th>Jul</th><th>Aug</th>
                    <th>Sep</th><th>Oct</th><th>Nov</th><th>Dec</th><th>Jan</th>
                    <th>Year</th>
                    <th>Closing</th>
                </tr>
            </thead>
            <tbody>
    """
    
    # Data rows
    for group_key in sorted(data_map.keys()):
        group_data = data_map[group_key]
        monthly_sales = group_data.get("monthly_sales", {})
        
        html += f"<tr><td class='text-left'>{group_key}</td>"
        html += f"<td class='number'>{flt(group_data.get('opening_value', 0)):.2f}</td>"
        html += f"<td class='number'>{flt(group_data.get('primary_value', 0)):.2f}</td>"
        
        # Monthly columns
        year_total = 0
        for month in ["04", "05", "06", "07", "08", "09", "10", "11", "12", "01"]:
            value = 0
            for month_key in monthly_sales:
                if f"-{month}-" in month_key:
                    value = flt(monthly_sales[month_key])
                    year_total += value
                    break
            html += f"<td class='number'>{value:.2f}</td>"
        
        html += f"<td class='number'>{year_total:.2f}</td>"
        html += f"<td class='number'>{flt(group_data.get('closing_value', 0)):.2f}</td></tr>"
    
    # Totals row
    html += f"""
            <tr class='total-row'>
                <td class='text-left'>{doc.report_type} Total</td>
                <td class='number'>{flt(doc.total_opening_value):.2f}</td>
                <td class='number'>{flt(doc.total_primary_value):.2f}</td>
                <td colspan='10' class='number'></td>
                <td class='number'>{flt(doc.total_secondary_sales):.2f}</td>
                <td class='number'>{flt(doc.total_closing_value):.2f}</td>
            </tr>
            </tbody>
        </table>
    </body>
    </html>
    """
    
    return html
