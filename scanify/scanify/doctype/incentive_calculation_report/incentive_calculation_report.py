import frappe
from frappe.model.document import Document
from frappe.utils import flt, getdate, add_months, get_first_day, get_last_day, now
from datetime import datetime
import json
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class IncentiveCalculationReport(Document):
    def validate(self):
        """Validate report parameters"""
        if self.period_type == "Custom Date Range":
            if self.from_date and self.to_date:
                if getdate(self.from_date) > getdate(self.to_date):
                    frappe.throw("From Date cannot be greater than To Date")
        
        # Set dates based on period type
        if self.period_type == "Quarterly" and self.quarter:
            self.set_quarter_dates()
    
    def set_quarter_dates(self):
        """Set from_date and to_date based on quarter selection"""
        current_year = datetime.now().year
        
        quarter_map = {
            "Q1 (Apr-Jun)": ("04-01", "06-30"),
            "Q2 (Jul-Sep)": ("07-01", "09-30"),
            "Q3 (Oct-Dec)": ("10-01", "12-31"),
            "Q4 (Jan-Mar)": ("01-01", "03-31")
        }
        
        if self.quarter in quarter_map:
            start, end = quarter_map[self.quarter]
            
            # Q4 spans two calendar years
            if self.quarter == "Q4 (Jan-Mar)":
                self.from_date = f"{current_year}-{start}"
                self.to_date = f"{current_year}-{end}"
            else:
                self.from_date = f"{current_year}-{start}"
                self.to_date = f"{current_year}-{end}"
    
    def before_save(self):
        """Set audit fields"""
        if not self.report_date:
            self.report_date = now()
        if not self.generated_by:
            self.generated_by = frappe.session.user
    
    def on_submit(self):
        """Generate on submit"""
        self.generate_report_data()
        self.calculate_totals()
    
    def generate_report_data(self):
        """Generate incentive calculation data"""
        try:
            filters = self.build_filters()
            
            # Fetch stockist statements
            statements = frappe.get_all(
                "Stockist Statement",
                filters=filters,
                fields=["name", "stockist_code", "statement_month"]
            )
            
            if not statements:
                frappe.msgprint("No data found for selected filters", indicator='orange')
                self.report_data = json.dumps({}, default=str)
                return {}
            
            # Aggregate data
            incentive_data = self.calculate_incentives(statements)
            
            # Store as JSON
            self.report_data = json.dumps(incentive_data, default=str)
            
            # Calculate totals
            self.calculate_totals()
            
            return incentive_data
            
        except Exception as e:
            frappe.log_error(frappe.get_traceback(), "Incentive Calculation Error")
            frappe.throw(f"Error generating report: {str(e)}")
    
    def build_filters(self):
        """Build filters for Stockist Statement"""
        filters = {"docstatus": 1}
        
        # Date filters
        if self.from_date and self.to_date:
            filters["statement_month"] = ["between", [self.from_date, self.to_date]]
        elif self.from_date:
            filters["statement_month"] = [">=", self.from_date]
        elif self.to_date:
            filters["statement_month"] = ["<=", self.to_date]
        
        # Hierarchy filters
        if self.hq:
            stockists = frappe.get_all(
                "Stockist Master",
                filters={"hq": self.hq, "status": "Active"},
                pluck="stockist_code"
            )
            if stockists:
                filters["stockist_code"] = ["in", stockists]
        elif self.team:
            hqs = frappe.get_all(
                "HQ Master",
                filters={"team": self.team, "status": "Active"},
                pluck="name"
            )
            stockist_list = []
            for hq in hqs:
                stockists = frappe.get_all(
                    "Stockist Master",
                    filters={"hq": hq, "status": "Active"},
                    pluck="stockist_code"
                )
                stockist_list.extend(stockists)
            
            if stockist_list:
                filters["stockist_code"] = ["in", stockist_list]
        elif self.region:
            teams = frappe.get_all(
                "Team Master",
                filters={"region": self.region, "status": "Active"},
                pluck="name"
            )
            stockist_list = []
            for team in teams:
                hqs = frappe.get_all(
                    "HQ Master",
                    filters={"team": team, "status": "Active"},
                    pluck="name"
                )
                for hq in hqs:
                    stockists = frappe.get_all(
                        "Stockist Master",
                        filters={"hq": hq, "status": "Active"},
                        pluck="stockist_code"
                    )
                    stockist_list.extend(stockists)
            
            if stockist_list:
                filters["stockist_code"] = ["in", stockist_list]
        
        return filters
    
    def calculate_incentives(self, statements):
        """
        Calculate incentives based on sales data
        Structure: {hq: {stockist: {product/total: {sales_qty, sales_value, incentive}}}}
        """
        data_map = {}
        
        for statement in statements:
            stmt_doc = frappe.get_doc("Stockist Statement", statement["name"])
            
            if not frappe.db.exists("Stockist Master", stmt_doc.stockist_code):
                continue
            
            stockist = frappe.get_doc("Stockist Master", stmt_doc.stockist_code)
            hq = stockist.hq
            stockist_name = stockist.stockist_name
            
            # Initialize structure
            if hq not in data_map:
                data_map[hq] = {}
            
            if stockist_name not in data_map[hq]:
                data_map[hq][stockist_name] = {
                    "stockist_code": stockist.stockist_code,
                    "total_sales_qty": 0,
                    "total_sales_value": 0,
                    "total_incentive": 0,
                    "products": {}
                }
            
            # Process items
            for item in stmt_doc.items:
                product_code = item.product_code
                
                # Get product details
                product = frappe.db.get_value(
                    "Product Master",
                    product_code,
                    ["product_name", "pack", "pts", "category", "division"],
                    as_dict=True
                )
                
                if not product:
                    continue
                
                # Filter by division
                if self.division and product.division != self.division:
                    continue
                
                # Filter by category
                if self.product_category and self.product_category != "All Products":
                    if product.category != self.product_category:
                        continue
                
                sales_qty = flt(item.sales_qty)
                sales_value = sales_qty * flt(product.pts)
                
                # Product-wise calculation
                if self.calculation_type in ["Product-wise", "Both"]:
                    if product_code not in data_map[hq][stockist_name]["products"]:
                        data_map[hq][stockist_name]["products"][product_code] = {
                            "product_name": product.product_name,
                            "pack": product.pack,
                            "sales_qty": 0,
                            "sales_value": 0,
                            "incentive_qty": 0,
                            "incentive_amount": 0
                        }
                    
                    data_map[hq][stockist_name]["products"][product_code]["sales_qty"] += sales_qty
                    data_map[hq][stockist_name]["products"][product_code]["sales_value"] += sales_value
                    
                    # Calculate product incentive
                    incentive_qty = sales_qty * flt(self.incentive_rate_per_unit)
                    data_map[hq][stockist_name]["products"][product_code]["incentive_qty"] += incentive_qty
                    data_map[hq][stockist_name]["products"][product_code]["incentive_amount"] += sales_value * flt(self.incentive_rate_per_rupee) / 100
                
                # Add to totals
                data_map[hq][stockist_name]["total_sales_qty"] += sales_qty
                data_map[hq][stockist_name]["total_sales_value"] += sales_value
        
        # Calculate rupee-wise incentives and apply threshold
        for hq in data_map:
            for stockist in data_map[hq]:
                stockist_data = data_map[hq][stockist]
                
                # Check minimum threshold
                if stockist_data["total_sales_value"] >= flt(self.minimum_sales_threshold):
                    if self.calculation_type in ["Rupee-wise", "Both"]:
                        # Rupee-wise incentive
                        stockist_data["total_incentive"] = stockist_data["total_sales_value"] * flt(self.incentive_rate_per_rupee) / 100
                    elif self.calculation_type == "Product-wise":
                        # Sum up product incentives
                        stockist_data["total_incentive"] = sum(
                            p["incentive_amount"] for p in stockist_data["products"].values()
                        )
                else:
                    stockist_data["total_incentive"] = 0
        
        return data_map
    
    def calculate_totals(self):
        """Calculate summary totals"""
        if not self.report_data:
            return
        
        try:
            data_map = json.loads(self.report_data)
            
            total_qty = 0
            total_value = 0
            total_incentive = 0
            total_units = 0
            
            for hq in data_map:
                for stockist in data_map[hq]:
                    stockist_data = data_map[hq][stockist]
                    
                    total_qty += flt(stockist_data.get("total_sales_qty", 0))
                    total_value += flt(stockist_data.get("total_sales_value", 0))
                    total_incentive += flt(stockist_data.get("total_incentive", 0))
                    
                    # Sum product incentive units
                    for product in stockist_data.get("products", {}).values():
                        total_units += flt(product.get("incentive_qty", 0))
            
            self.total_sales_qty = total_qty
            self.total_sales_value = total_value
            self.total_incentive_amount = total_incentive
            self.total_incentive_units = total_units
        
        except Exception as e:
            frappe.log_error(f"Error calculating totals: {str(e)}", "Incentive Totals Error")


@frappe.whitelist()
def generate_report(doc_name):
    """API method to generate report"""
    doc = frappe.get_doc("Incentive Calculation Report", doc_name)
    doc.generate_report_data()
    doc.save()
    frappe.db.commit()
    
    return {
        "success": True,
        "message": "Report generated successfully",
        "total_sales_value": doc.total_sales_value,
        "total_incentive": doc.total_incentive_amount
    }


@frappe.whitelist()
def export_to_excel(doc_name):
    """Export incentive report to Excel"""
    doc = frappe.get_doc("Incentive Calculation Report", doc_name)
    
    if not doc.report_data:
        frappe.throw("No report data available. Please generate the report first.")
    
    data_map = json.loads(doc.report_data or "{}")
    
    if not data_map:
        frappe.throw("No data available to export")
    
    # Create workbook with multiple sheets
    wb = openpyxl.Workbook()
    
    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    # Styling
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws_summary.merge_cells('A1:H1')
    ws_summary['A1'] = "STEDMAN PHARMACEUTICALS PVT LTD"
    ws_summary['A1'].font = Font(bold=True, size=14)
    ws_summary['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws_summary.merge_cells('A2:H2')
    ws_summary['A2'] = f"Incentive Calculation Report - {doc.division} Division"
    ws_summary['A2'].font = Font(size=12, italic=True)
    ws_summary['A2'].alignment = Alignment(horizontal='center')
    
    ws_summary.merge_cells('A3:H3')
    ws_summary['A3'] = f"Period: {doc.from_date or 'N/A'} to {doc.to_date or 'N/A'} | {doc.calculation_type}"
    ws_summary['A3'].font = Font(size=10)
    ws_summary['A3'].alignment = Alignment(horizontal='center')
    
    # Headers
    row = 5
    headers = ["HQ", "Stockist", "Stockist Code", "Sales Qty", "Sales Value", 
               "Incentive Rate", "Incentive Amount", "Qualified"]
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws_summary.cell(row=row, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Data rows
    row = 6
    for hq in sorted(data_map.keys()):
        for stockist in sorted(data_map[hq].keys()):
            stockist_data = data_map[hq][stockist]
            
            sales_value = flt(stockist_data.get("total_sales_value", 0))
            incentive = flt(stockist_data.get("total_incentive", 0))
            qualified = "Yes" if sales_value >= flt(doc.minimum_sales_threshold) else "No"
            
            ws_summary.cell(row=row, column=1).value = hq
            ws_summary.cell(row=row, column=2).value = stockist
            ws_summary.cell(row=row, column=3).value = stockist_data.get("stockist_code", "")
            ws_summary.cell(row=row, column=4).value = flt(stockist_data.get("total_sales_qty", 0))
            ws_summary.cell(row=row, column=5).value = sales_value
            ws_summary.cell(row=row, column=6).value = f"{flt(doc.incentive_rate_per_rupee)}%"
            ws_summary.cell(row=row, column=7).value = incentive
            ws_summary.cell(row=row, column=8).value = qualified
            
            # Formatting
            for col in range(1, 9):
                cell = ws_summary.cell(row=row, column=col)
                cell.border = border
                if col in [4, 5, 7]:
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
            
            row += 1
    
    # Totals row
    totals_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    totals_font = Font(bold=True)
    
    ws_summary.cell(row=row, column=1).value = "TOTAL"
    ws_summary.cell(row=row, column=4).value = doc.total_sales_qty
    ws_summary.cell(row=row, column=5).value = doc.total_sales_value
    ws_summary.cell(row=row, column=7).value = doc.total_incentive_amount
    
    for col in range(1, 9):
        cell = ws_summary.cell(row=row, column=col)
        cell.fill = totals_fill
        cell.font = totals_font
        cell.border = border
        if col in [4, 5, 7]:
            cell.number_format = '#,##0.00'
    
    # Adjust column widths
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 35
    ws_summary.column_dimensions['C'].width = 15
    for col in range(4, 9):
        ws_summary.column_dimensions[get_column_letter(col)].width = 15
    
    # Sheet 2: Product-wise (if applicable)
    if doc.calculation_type in ["Product-wise", "Both"]:
        ws_product = wb.create_sheet("Product-wise")
        
        # Title
        ws_product.merge_cells('A1:I1')
        ws_product['A1'] = "Product-wise Incentive Breakdown"
        ws_product['A1'].font = Font(bold=True, size=14)
        ws_product['A1'].alignment = Alignment(horizontal='center')
        
        # Headers
        row = 3
        product_headers = ["HQ", "Stockist", "Product Code", "Product Name", "Pack",
                          "Sales Qty", "Sales Value", "Incentive Units", "Incentive Amount"]
        
        for col_idx, header in enumerate(product_headers, 1):
            cell = ws_product.cell(row=row, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Data
        row = 4
        for hq in sorted(data_map.keys()):
            for stockist in sorted(data_map[hq].keys()):
                stockist_data = data_map[hq][stockist]
                
                for product_code in sorted(stockist_data.get("products", {}).keys()):
                    product_data = stockist_data["products"][product_code]
                    
                    ws_product.cell(row=row, column=1).value = hq
                    ws_product.cell(row=row, column=2).value = stockist
                    ws_product.cell(row=row, column=3).value = product_code
                    ws_product.cell(row=row, column=4).value = product_data.get("product_name", "")
                    ws_product.cell(row=row, column=5).value = product_data.get("pack", "")
                    ws_product.cell(row=row, column=6).value = flt(product_data.get("sales_qty", 0))
                    ws_product.cell(row=row, column=7).value = flt(product_data.get("sales_value", 0))
                    ws_product.cell(row=row, column=8).value = flt(product_data.get("incentive_qty", 0))
                    ws_product.cell(row=row, column=9).value = flt(product_data.get("incentive_amount", 0))
                    
                    # Formatting
                    for col in range(1, 10):
                        cell = ws_product.cell(row=row, column=col)
                        cell.border = border
                        if col >= 6:
                            cell.number_format = '#,##0.00'
                            cell.alignment = Alignment(horizontal='right')
                    
                    row += 1
        
        # Adjust widths
        ws_product.column_dimensions['A'].width = 25
        ws_product.column_dimensions['B'].width = 35
        ws_product.column_dimensions['C'].width = 15
        ws_product.column_dimensions['D'].width = 30
        ws_product.column_dimensions['E'].width = 12
        for col in range(6, 10):
            ws_product.column_dimensions[get_column_letter(col)].width = 15
    
    # Save
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    file_doc = frappe.get_doc({
        "doctype": "File",
        "file_name": f"Incentive_Calculation_{doc.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        "attached_to_doctype": "Incentive Calculation Report",
        "attached_to_name": doc_name,
        "content": output.getvalue(),
        "is_private": 0
    })
    file_doc.save(ignore_permissions=True)
    
    return {
        "success": True,
        "file_url": file_doc.file_url
    }


@frappe.whitelist()
def export_to_pdf(doc_name):
    """Export to PDF"""
    from frappe.utils.pdf import get_pdf
    
    doc = frappe.get_doc("Incentive Calculation Report", doc_name)
    
    if not doc.report_data:
        frappe.throw("No report data available. Please generate the report first.")
    
    html = render_report_html(doc)
    pdf_bytes = get_pdf(html)
    
    file_doc = frappe.get_doc({
        "doctype": "File",
        "file_name": f"Incentive_Calculation_{doc.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
        "attached_to_doctype": "Incentive Calculation Report",
        "attached_to_name": doc_name,
        "content": pdf_bytes,
        "is_private": 0
    })
    file_doc.save(ignore_permissions=True)
    
    return {
        "success": True,
        "file_url": file_doc.file_url
    }


def render_report_html(doc):
    """Render HTML for PDF"""
    data_map = json.loads(doc.report_data or "{}")
    
    html = f"""
    <html>
    <head>
        <style>
            @page {{ size: A4 landscape; margin: 15mm; }}
            body {{ font-family: Arial, sans-serif; margin: 0; font-size: 9px; }}
            .header {{ text-align: center; margin-bottom: 15px; }}
            .title {{ font-size: 14px; font-weight: bold; }}
            .subtitle {{ font-size: 10px; color: #666; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 10px; }}
            th {{ background-color: #1F4E78; color: white; padding: 5px; text-align: center; border: 1px solid #000; font-size: 8px; }}
            td {{ padding: 4px; border: 1px solid #ccc; font-size: 8px; }}
            .number {{ text-align: right; }}
            .text-left {{ text-align: left; }}
            .total-row {{ background-color: #D9E1F2; font-weight: bold; }}
        </style>
    </head>
    <body>
        <div class="header">
            <div class="title">STEDMAN PHARMACEUTICALS PVT LTD</div>
            <div class="subtitle">Incentive Calculation Report - {doc.division} Division</div>
            <div class="subtitle">Period: {doc.from_date or 'N/A'} to {doc.to_date or 'N/A'} | {doc.calculation_type}</div>
        </div>
        
        <table>
            <thead>
                <tr>
                    <th>HQ</th>
                    <th>Stockist</th>
                    <th>Sales Qty</th>
                    <th>Sales Value</th>
                    <th>Rate</th>
                    <th>Incentive</th>
                    <th>Qualified</th>
                </tr>
            </thead>
            <tbody>
    """
    
    for hq in sorted(data_map.keys()):
        for stockist in sorted(data_map[hq].keys()):
            stockist_data = data_map[hq][stockist]
            
            sales_value = flt(stockist_data.get("total_sales_value", 0))
            incentive = flt(stockist_data.get("total_incentive", 0))
            qualified = "Yes" if sales_value >= flt(doc.minimum_sales_threshold) else "No"
            
            html += f"""
            <tr>
                <td class='text-left'>{hq}</td>
                <td class='text-left'>{stockist}</td>
                <td class='number'>{flt(stockist_data.get('total_sales_qty', 0)):.2f}</td>
                <td class='number'>{sales_value:.2f}</td>
                <td class='number'>{flt(doc.incentive_rate_per_rupee)}%</td>
                <td class='number'>{incentive:.2f}</td>
                <td>{qualified}</td>
            </tr>
            """
    
    html += f"""
            <tr class='total-row'>
                <td colspan='2' class='text-left'>TOTAL</td>
                <td class='number'>{flt(doc.total_sales_qty):.2f}</td>
                <td class='number'>{flt(doc.total_sales_value):.2f}</td>
                <td></td>
                <td class='number'>{flt(doc.total_incentive_amount):.2f}</td>
                <td></td>
            </tr>
            </tbody>
        </table>
    </body>
    </html>
    """
    
    return html
