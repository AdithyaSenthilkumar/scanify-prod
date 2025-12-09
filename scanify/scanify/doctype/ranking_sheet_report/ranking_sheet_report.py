import frappe
from frappe.model.document import Document
from frappe.utils import flt, getdate, now
from datetime import datetime
import json
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class RankingSheetReport(Document):
    def validate(self):
        """Validate report parameters"""
        if self.period_type == "Custom Date Range":
            if self.from_date and self.to_date:
                if getdate(self.from_date) > getdate(self.to_date):
                    frappe.throw("From Date cannot be greater than To Date")
        
        if self.period_type == "Quarterly" and self.quarter:
            self.set_quarter_dates()
    
    def set_quarter_dates(self):
        """Set dates based on quarter"""
        current_year = datetime.now().year
        
        quarter_map = {
            "Q1 (Apr-Jun)": ("04-01", "06-30"),
            "Q2 (Jul-Sep)": ("07-01", "09-30"),
            "Q3 (Oct-Dec)": ("10-01", "12-31"),
            "Q4 (Jan-Mar)": ("01-01", "03-31")
        }
        
        if self.quarter in quarter_map:
            start, end = quarter_map[self.quarter]
            
            if self.quarter == "Q4 (Jan-Mar)":
                self.from_date = f"{current_year}-{start}"
                self.to_date = f"{current_year}-{end}"
            else:
                self.from_date = f"{current_year}-{start}"
                self.to_date = f"{current_year}-{end}"
    
    def before_save(self):
        """Set audit fields"""
        if not self.report_date:
            self.report_date = now()
        if not self.generated_by:
            self.generated_by = frappe.session.user
    
    def on_submit(self):
        """Generate on submit"""
        self.generate_report_data()
        self.calculate_summary()
    
    def generate_report_data(self):
        """Generate ranking data"""
        try:
            filters = self.build_filters()
            
            statements = frappe.get_all(
                "Stockist Statement",
                filters=filters,
                fields=["name", "stockist_code", "statement_month"]
            )
            
            if not statements:
                frappe.msgprint("No data found for selected filters", indicator='orange')
                self.report_data = json.dumps({}, default=str)
                return {}
            
            # Calculate rankings based on type
            if self.ranking_type == "Product-wise":
                ranking_data = self.calculate_product_rankings(statements)
            elif self.ranking_type == "Rupee-wise (HQ)":
                ranking_data = self.calculate_hq_rankings(statements)
            elif self.ranking_type == "Stockist-wise":
                ranking_data = self.calculate_stockist_rankings(statements)
            else:  # Combined
                ranking_data = {
                    "product_wise": self.calculate_product_rankings(statements),
                    "hq_wise": self.calculate_hq_rankings(statements)
                }
            
            self.report_data = json.dumps(ranking_data, default=str)
            self.calculate_summary()
            
            return ranking_data
            
        except Exception as e:
            frappe.log_error(frappe.get_traceback(), "Ranking Sheet Error")
            frappe.throw(f"Error generating report: {str(e)}")
    
    def build_filters(self):
        """Build filters"""
        filters = {"docstatus": 1}
        
        # Date filters
        if self.from_date and self.to_date:
            filters["statement_month"] = ["between", [self.from_date, self.to_date]]
        elif self.from_date:
            filters["statement_month"] = [">=", self.from_date]
        elif self.to_date:
            filters["statement_month"] = ["<=", self.to_date]
        
        # Region filter
        if self.region:
            teams = frappe.get_all(
                "Team Master",
                filters={"region": self.region, "status": "Active"},
                pluck="name"
            )
            stockist_list = []
            for team in teams:
                hqs = frappe.get_all(
                    "HQ Master",
                    filters={"team": team, "status": "Active"},
                    pluck="name"
                )
                for hq in hqs:
                    stockists = frappe.get_all(
                        "Stockist Master",
                        filters={"hq": hq, "status": "Active"},
                        pluck="stockist_code"
                    )
                    stockist_list.extend(stockists)
            
            if stockist_list:
                filters["stockist_code"] = ["in", stockist_list]
        
        return filters
    
    def calculate_product_rankings(self, statements):
        """
        Calculate product-wise rankings
        Returns: {product_code: [{rank, hq/stockist, sales_qty, sales_value, growth%}]}
        """
        product_sales = {}
        
        for statement in statements:
            stmt_doc = frappe.get_doc("Stockist Statement", statement["name"])
            
            if not frappe.db.exists("Stockist Master", stmt_doc.stockist_code):
                continue
            
            stockist = frappe.get_doc("Stockist Master", stmt_doc.stockist_code)
            
            for item in stmt_doc.items:
                product_code = item.product_code
                
                # Get product details
                product = frappe.db.get_value(
                    "Product Master",
                    product_code,
                    ["product_name", "pack", "pts", "category", "division"],
                    as_dict=True
                )
                
                if not product:
                    continue
                
                # Filter by division
                if self.division != "Both" and product.division != self.division:
                    continue
                
                # Filter by category
                if self.product_category and self.product_category != "All Products":
                    if product.category != self.product_category:
                        continue
                
                # Filter by specific product
                if self.product_code and product_code != self.product_code:
                    continue
                
                # Initialize product entry
                if product_code not in product_sales:
                    product_sales[product_code] = {
                        "product_name": product.product_name,
                        "pack": product.pack,
                        "division": product.division,
                        "participants": {}
                    }
                
                # Determine participant key (HQ or Stockist)
                participant_key = f"{stockist.hq}"
                participant_name = stockist.hq
                
                if participant_key not in product_sales[product_code]["participants"]:
                    product_sales[product_code]["participants"][participant_key] = {
                        "name": participant_name,
                        "hq": stockist.hq,
                        "primary_qty": 0,
                        "secondary_qty": 0,
                        "primary_value": 0,
                        "secondary_value": 0
                    }
                
                # Aggregate sales
                pts = flt(product.pts)
                
                if self.sales_type in ["Primary Sales", "Both"]:
                    product_sales[product_code]["participants"][participant_key]["primary_qty"] += flt(item.purchase_qty)
                    product_sales[product_code]["participants"][participant_key]["primary_value"] += flt(item.purchase_qty) * pts
                
                if self.sales_type in ["Secondary Sales", "Both"]:
                    product_sales[product_code]["participants"][participant_key]["secondary_qty"] += flt(item.sales_qty)
                    product_sales[product_code]["participants"][participant_key]["secondary_value"] += flt(item.sales_qty) * pts
        
        # Rank participants for each product
        ranked_data = {}
        
        for product_code, product_data in product_sales.items():
            participants = product_data["participants"]
            
            # Sort by sales value (secondary or primary based on sales_type)
            if self.sales_type == "Primary Sales":
                sorted_participants = sorted(
                    participants.items(),
                    key=lambda x: x[1]["primary_value"],
                    reverse=True
                )
            else:
                sorted_participants = sorted(
                    participants.items(),
                    key=lambda x: x[1]["secondary_value"],
                    reverse=True
                )
            
            # Get top N
            top_n = int(self.top_n_records or 5)
            top_performers = []
            
            for rank, (key, data) in enumerate(sorted_participants[:top_n], 1):
                top_performers.append({
                    "rank": rank,
                    "hq": data["hq"],
                    "name": data["name"],
                    "primary_qty": data["primary_qty"],
                    "secondary_qty": data["secondary_qty"],
                    "primary_value": data["primary_value"],
                    "secondary_value": data["secondary_value"],
                    "total_value": data["primary_value"] + data["secondary_value"]
                })
            
            ranked_data[product_code] = {
                "product_name": product_data["product_name"],
                "pack": product_data["pack"],
                "division": product_data["division"],
                "top_performers": top_performers
            }
        
        return ranked_data
    
    def calculate_hq_rankings(self, statements):
        """
        Calculate HQ-wise rankings based on total rupee value
        Returns: [{rank, hq, region, team, primary_value, secondary_value, total_value}]
        """
        hq_sales = {}
        
        for statement in statements:
            stmt_doc = frappe.get_doc("Stockist Statement", statement["name"])
            
            if not frappe.db.exists("Stockist Master", stmt_doc.stockist_code):
                continue
            
            stockist = frappe.get_doc("Stockist Master", stmt_doc.stockist_code)
            hq = stockist.hq
            
            # Get HQ details
            if hq not in hq_sales:
                hq_doc = frappe.get_doc("HQ Master", hq)
                team_doc = frappe.get_doc("Team Master", hq_doc.team)
                
                hq_sales[hq] = {
                    "hq": hq,
                    "team": hq_doc.team,
                    "region": team_doc.region,
                    "primary_value": 0,
                    "secondary_value": 0
                }
            
            # Aggregate sales
            for item in stmt_doc.items:
                product = frappe.db.get_value(
                    "Product Master",
                    item.product_code,
                    ["pts", "division"],
                    as_dict=True
                )
                
                if not product:
                    continue
                
                # Filter by division
                if self.division != "Both" and product.division != self.division:
                    continue
                
                pts = flt(product.pts)
                
                if self.sales_type in ["Primary Sales", "Both"]:
                    hq_sales[hq]["primary_value"] += flt(item.purchase_qty) * pts
                
                if self.sales_type in ["Secondary Sales", "Both"]:
                    hq_sales[hq]["secondary_value"] += flt(item.sales_qty) * pts
        
        # Calculate total and rank
        for hq in hq_sales:
            hq_sales[hq]["total_value"] = hq_sales[hq]["primary_value"] + hq_sales[hq]["secondary_value"]
        
        # Sort by total value
        sorted_hqs = sorted(
            hq_sales.values(),
            key=lambda x: x["total_value"],
            reverse=True
        )
        
        # Add ranks
        top_n = int(self.top_n_records or 10)
        ranked_hqs = []
        
        for rank, hq_data in enumerate(sorted_hqs[:top_n], 1):
            hq_data["rank"] = rank
            ranked_hqs.append(hq_data)
        
        return ranked_hqs
    
    def calculate_stockist_rankings(self, statements):
        """Calculate stockist-wise rankings"""
        stockist_sales = {}
        
        for statement in statements:
            stmt_doc = frappe.get_doc("Stockist Statement", statement["name"])
            
            if not frappe.db.exists("Stockist Master", stmt_doc.stockist_code):
                continue
            
            stockist = frappe.get_doc("Stockist Master", stmt_doc.stockist_code)
            stockist_key = stmt_doc.stockist_code
            
            if stockist_key not in stockist_sales:
                stockist_sales[stockist_key] = {
                    "stockist_code": stockist_key,
                    "stockist_name": stockist.stockist_name,
                    "hq": stockist.hq,
                    "primary_value": 0,
                    "secondary_value": 0
                }
            
            # Aggregate
            for item in stmt_doc.items:
                product = frappe.db.get_value(
                    "Product Master",
                    item.product_code,
                    ["pts", "division"],
                    as_dict=True
                )
                
                if not product:
                    continue
                
                if self.division != "Both" and product.division != self.division:
                    continue
                
                pts = flt(product.pts)
                
                if self.sales_type in ["Primary Sales", "Both"]:
                    stockist_sales[stockist_key]["primary_value"] += flt(item.purchase_qty) * pts
                
                if self.sales_type in ["Secondary Sales", "Both"]:
                    stockist_sales[stockist_key]["secondary_value"] += flt(item.sales_qty) * pts
        
        # Calculate total and rank
        for stockist_key in stockist_sales:
            stockist_sales[stockist_key]["total_value"] = (
                stockist_sales[stockist_key]["primary_value"] + 
                stockist_sales[stockist_key]["secondary_value"]
            )
        
        sorted_stockists = sorted(
            stockist_sales.values(),
            key=lambda x: x["total_value"],
            reverse=True
        )
        
        top_n = int(self.top_n_records or 10)
        ranked_stockists = []
        
        for rank, stockist_data in enumerate(sorted_stockists[:top_n], 1):
            stockist_data["rank"] = rank
            ranked_stockists.append(stockist_data)
        
        return ranked_stockists
    
    def calculate_summary(self):
        """Calculate summary statistics"""
        if not self.report_data:
            return
        
        try:
            data = json.loads(self.report_data)
            
            total_participants = 0
            total_sales = 0
            max_sales = 0
            
            if self.ranking_type == "Product-wise":
                for product_code in data:
                    product_data = data[product_code]
                    top_performers = product_data.get("top_performers", [])
                    total_participants += len(top_performers)
                    
                    for performer in top_performers:
                        sales_value = flt(performer.get("total_value", 0))
                        total_sales += sales_value
                        max_sales = max(max_sales, sales_value)
            
            elif self.ranking_type == "Rupee-wise (HQ)":
                total_participants = len(data)
                for hq_data in data:
                    sales_value = flt(hq_data.get("total_value", 0))
                    total_sales += sales_value
                    max_sales = max(max_sales, sales_value)
            
            elif self.ranking_type == "Stockist-wise":
                total_participants = len(data)
                for stockist_data in data:
                    sales_value = flt(stockist_data.get("total_value", 0))
                    total_sales += sales_value
                    max_sales = max(max_sales, sales_value)
            
            self.total_participants = total_participants
            self.total_sales_value = total_sales
            self.average_sales_value = total_sales / total_participants if total_participants > 0 else 0
            self.highest_sales_value = max_sales
        
        except Exception as e:
            frappe.log_error(f"Error calculating summary: {str(e)}", "Summary Error")


@frappe.whitelist()
def generate_report(doc_name):
    """API to generate report"""
    doc = frappe.get_doc("Ranking Sheet Report", doc_name)
    doc.generate_report_data()
    doc.save()
    frappe.db.commit()
    
    return {
        "success": True,
        "message": "Report generated successfully",
        "total_participants": doc.total_participants,
        "total_sales_value": doc.total_sales_value
    }


@frappe.whitelist()
def export_to_excel(doc_name):
    """Export rankings to Excel"""
    doc = frappe.get_doc("Ranking Sheet Report", doc_name)
    
    if not doc.report_data:
        frappe.throw("No report data available. Please generate the report first.")
    
    data = json.loads(doc.report_data or "{}")
    
    if not data:
        frappe.throw("No data available to export")
    
    wb = openpyxl.Workbook()
    
    # Styling
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    gold_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    silver_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
    bronze_fill = PatternFill(start_color="CD7F32", end_color="CD7F32", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    if doc.ranking_type == "Product-wise":
        ws = wb.active
        ws.title = "Product Rankings"
        
        # Title
        ws.merge_cells('A1:I1')
        ws['A1'] = "STEDMAN PHARMACEUTICALS PVT LTD"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A2:I2')
        ws['A2'] = f"Product-wise Ranking Sheet - {doc.division} Division"
        ws['A2'].font = Font(size=12, italic=True)
        ws['A2'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A3:I3')
        ws['A3'] = f"Period: {doc.from_date or 'N/A'} to {doc.to_date or 'N/A'} | {doc.sales_type}"
        ws['A3'].font = Font(size=10)
        ws['A3'].alignment = Alignment(horizontal='center')
        
        row = 5
        
        for product_code in sorted(data.keys()):
            product_data = data[product_code]
            
            # Product header
            ws.merge_cells(f'A{row}:I{row}')
            ws[f'A{row}'] = f"{product_code} - {product_data['product_name']} ({product_data['pack']})"
            ws[f'A{row}'].font = Font(bold=True, size=11, color="FFFFFF")
            ws[f'A{row}'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            ws[f'A{row}'].alignment = Alignment(horizontal='left')
            row += 1
            
            # Column headers
            headers = ["Rank", "HQ", "Primary Qty", "Secondary Qty", 
                      "Primary Value", "Secondary Value", "Total Value", "Badge"]
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col_idx)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
                cell.border = border
            
            row += 1
            
            # Top performers
            for performer in product_data.get("top_performers", []):
                rank = performer["rank"]
                
                ws.cell(row=row, column=1).value = rank
                ws.cell(row=row, column=2).value = performer["hq"]
                ws.cell(row=row, column=3).value = flt(performer["primary_qty"])
                ws.cell(row=row, column=4).value = flt(performer["secondary_qty"])
                ws.cell(row=row, column=5).value = flt(performer["primary_value"])
                ws.cell(row=row, column=6).value = flt(performer["secondary_value"])
                ws.cell(row=row, column=7).value = flt(performer["total_value"])
                
                # Badge
                badge = ""
                rank_fill = None
                if rank == 1:
                    badge = "🥇 Gold"
                    rank_fill = gold_fill
                elif rank == 2:
                    badge = "🥈 Silver"
                    rank_fill = silver_fill
                elif rank == 3:
                    badge = "🥉 Bronze"
                    rank_fill = bronze_fill
                
                ws.cell(row=row, column=8).value = badge
                
                # Apply formatting
                for col in range(1, 9):
                    cell = ws.cell(row=row, column=col)
                    cell.border = border
                    if rank_fill and col == 1:
                        cell.fill = rank_fill
                    if col >= 3 and col <= 7:
                        cell.number_format = '#,##0.00'
                        cell.alignment = Alignment(horizontal='right')
                
                row += 1
            
            row += 2  # Space between products
        
        # Adjust widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 30
        for col in range(3, 8):
            ws.column_dimensions[get_column_letter(col)].width = 15
        ws.column_dimensions['H'].width = 12
    
    elif doc.ranking_type == "Rupee-wise (HQ)":
        ws = wb.active
        ws.title = "HQ Rankings"
        
        # Title
        ws.merge_cells('A1:H1')
        ws['A1'] = "STEDMAN PHARMACEUTICALS PVT LTD"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A2:H2')
        ws['A2'] = f"HQ-wise Ranking Sheet - {doc.division} Division"
        ws['A2'].font = Font(size=12, italic=True)
        ws['A2'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A3:H3')
        ws['A3'] = f"Period: {doc.from_date or 'N/A'} to {doc.to_date or 'N/A'}"
        ws['A3'].font = Font(size=10)
        ws['A3'].alignment = Alignment(horizontal='center')
        
        # Headers
        row = 5
        headers = ["Rank", "HQ", "Team", "Region", "Primary Value", 
                  "Secondary Value", "Total Value", "Badge"]
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        row = 6
        
        for hq_data in data:
            rank = hq_data["rank"]
            
            ws.cell(row=row, column=1).value = rank
            ws.cell(row=row, column=2).value = hq_data["hq"]
            ws.cell(row=row, column=3).value = hq_data["team"]
            ws.cell(row=row, column=4).value = hq_data["region"]
            ws.cell(row=row, column=5).value = flt(hq_data["primary_value"])
            ws.cell(row=row, column=6).value = flt(hq_data["secondary_value"])
            ws.cell(row=row, column=7).value = flt(hq_data["total_value"])
            
            # Badge
            badge = ""
            rank_fill = None
            if rank == 1:
                badge = "🥇 Gold"
                rank_fill = gold_fill
            elif rank == 2:
                badge = "🥈 Silver"
                rank_fill = silver_fill
            elif rank == 3:
                badge = "🥉 Bronze"
                rank_fill = bronze_fill
            
            ws.cell(row=row, column=8).value = badge
            
            # Formatting
            for col in range(1, 9):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                if rank_fill and col == 1:
                    cell.fill = rank_fill
                if col >= 5 and col <= 7:
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
            
            row += 1
        
        # Adjust widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 25
        for col in range(5, 8):
            ws.column_dimensions[get_column_letter(col)].width = 15
        ws.column_dimensions['H'].width = 12
    
    # Save
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    file_doc = frappe.get_doc({
        "doctype": "File",
        "file_name": f"Ranking_Sheet_{doc.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        "attached_to_doctype": "Ranking Sheet Report",
        "attached_to_name": doc_name,
        "content": output.getvalue(),
        "is_private": 0
    })
    file_doc.save(ignore_permissions=True)
    
    return {
        "success": True,
        "file_url": file_doc.file_url
    }


@frappe.whitelist()
def export_to_pdf(doc_name):
    """Export to PDF"""
    from frappe.utils.pdf import get_pdf
    
    doc = frappe.get_doc("Ranking Sheet Report", doc_name)
    
    if not doc.report_data:
        frappe.throw("No report data available. Please generate the report first.")
    
    html = render_report_html(doc)
    pdf_bytes = get_pdf(html)
    
    file_doc = frappe.get_doc({
        "doctype": "File",
        "file_name": f"Ranking_Sheet_{doc.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
        "attached_to_doctype": "Ranking Sheet Report",
        "attached_to_name": doc_name,
        "content": pdf_bytes,
        "is_private": 0
    })
    file_doc.save(ignore_permissions=True)
    
    return {
        "success": True,
        "file_url": file_doc.file_url
    }


def render_report_html(doc):
    """Render HTML for PDF"""
    data = json.loads(doc.report_data or "{}")
    
    html = f"""
    <html>
    <head>
        <style>
            @page {{ size: A4 landscape; margin: 15mm; }}
            body {{ font-family: Arial, sans-serif; margin: 0; font-size: 9px; }}
            .header {{ text-align: center; margin-bottom: 15px; }}
            .title {{ font-size: 14px; font-weight: bold; }}
            .subtitle {{ font-size: 10px; color: #666; }}
            table {{ width: 100%; border-collapse: collapse; margin-bottom: 20px; }}
            th {{ background-color: #1F4E78; color: white; padding: 5px; text-align: center; border: 1px solid #000; font-size: 8px; }}
            td {{ padding: 4px; border: 1px solid #ccc; font-size: 8px; }}
            .number {{ text-align: right; }}
            .product-header {{ background-color: #4472C4; color: white; font-weight: bold; padding: 5px; }}
            .rank-1 {{ background-color: #FFD700; }}
            .rank-2 {{ background-color: #C0C0C0; }}
            .rank-3 {{ background-color: #CD7F32; }}
        </style>
    </head>
    <body>
        <div class="header">
            <div class="title">STEDMAN PHARMACEUTICALS PVT LTD</div>
            <div class="subtitle">{doc.ranking_type} - {doc.division} Division</div>
            <div class="subtitle">Period: {doc.from_date or 'N/A'} to {doc.to_date or 'N/A'}</div>
        </div>
    """
    
    if doc.ranking_type == "Product-wise":
        for product_code in sorted(data.keys()):
            product_data = data[product_code]
            
            html += f"""
            <div class="product-header">{product_code} - {product_data['product_name']} ({product_data['pack']})</div>
            <table>
                <thead>
                    <tr>
                        <th>Rank</th>
                        <th>HQ</th>
                        <th>Primary Qty</th>
                        <th>Secondary Qty</th>
                        <th>Primary Value</th>
                        <th>Secondary Value</th>
                        <th>Total Value</th>
                    </tr>
                </thead>
                <tbody>
            """
            
            for performer in product_data.get("top_performers", []):
                rank = performer["rank"]
                rank_class = f"rank-{rank}" if rank <= 3 else ""
                
                html += f"""
                <tr class="{rank_class}">
                    <td>{rank}</td>
                    <td>{performer['hq']}</td>
                    <td class="number">{flt(performer['primary_qty']):.0f}</td>
                    <td class="number">{flt(performer['secondary_qty']):.0f}</td>
                    <td class="number">{flt(performer['primary_value']):.2f}</td>
                    <td class="number">{flt(performer['secondary_value']):.2f}</td>
                    <td class="number">{flt(performer['total_value']):.2f}</td>
                </tr>
                """
            
            html += "</tbody></table>"
    
    elif doc.ranking_type == "Rupee-wise (HQ)":
        html += """
        <table>
            <thead>
                <tr>
                    <th>Rank</th>
                    <th>HQ</th>
                    <th>Team</th>
                    <th>Region</th>
                    <th>Primary Value</th>
                    <th>Secondary Value</th>
                    <th>Total Value</th>
                </tr>
            </thead>
            <tbody>
        """
        
        for hq_data in data:
            rank = hq_data["rank"]
            rank_class = f"rank-{rank}" if rank <= 3 else ""
            
            html += f"""
            <tr class="{rank_class}">
                <td>{rank}</td>
                <td>{hq_data['hq']}</td>
                <td>{hq_data['team']}</td>
                <td>{hq_data['region']}</td>
                <td class="number">{flt(hq_data['primary_value']):.2f}</td>
                <td class="number">{flt(hq_data['secondary_value']):.2f}</td>
                <td class="number">{flt(hq_data['total_value']):.2f}</td>
            </tr>
            """
        
        html += "</tbody></table>"
    
    html += "</body></html>"
    
    return html
