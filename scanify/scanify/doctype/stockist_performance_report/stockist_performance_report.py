import frappe
from frappe.model.document import Document
from frappe.utils import flt, getdate, add_months, now
from datetime import datetime
import json
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class StockistPerformanceReport(Document):
    def validate(self):
        """Validate report parameters"""
        if self.from_date and self.to_date:
            if getdate(self.from_date) > getdate(self.to_date):
                frappe.throw("From Date cannot be greater than To Date")
    
    def before_save(self):
        """Set audit fields"""
        if not self.report_date:
            self.report_date = now()
        if not self.generated_by:
            self.generated_by = frappe.session.user
    
    def on_submit(self):
        """Generate on submit"""
        self.generate_report_data()
        self.calculate_summary()
    
    def generate_report_data(self):
        """Generate stockist performance data"""
        try:
            filters = self.build_filters()
            
            statements = frappe.get_all(
                "Stockist Statement",
                filters=filters,
                fields=["name", "stockist_code", "statement_month"]
            )
            
            if not statements:
                frappe.msgprint("No data found for selected filters", indicator='orange')
                self.report_data = json.dumps({}, default=str)
                return {}
            
            # Aggregate stockist performance
            performance_data = self.aggregate_stockist_performance(statements)
            
            # Compare with previous period if enabled
            if self.compare_with_previous_period:
                performance_data = self.add_previous_period_comparison(performance_data)
            
            self.report_data = json.dumps(performance_data, default=str)
            self.calculate_summary()
            
            return performance_data
            
        except Exception as e:
            frappe.log_error(frappe.get_traceback(), "Stockist Performance Report Error")
            frappe.throw(f"Error generating report: {str(e)}")
    
    def build_filters(self):
        """Build filters for Stockist Statement"""
        filters = {"docstatus": 1}
        
        # Date filters
        if self.from_date and self.to_date:
            filters["statement_month"] = ["between", [self.from_date, self.to_date]]
        elif self.from_date:
            filters["statement_month"] = [">=", self.from_date]
        elif self.to_date:
            filters["statement_month"] = ["<=", self.to_date]
        
        # Stockist filter
        if self.stockist:
            filters["stockist_code"] = self.stockist
        
        # HQ filter
        elif self.hq:
            stockists = frappe.get_all(
                "Stockist Master",
                filters={"hq": self.hq, "status": "Active"},
                pluck="stockist_code"
            )
            if stockists:
                filters["stockist_code"] = ["in", stockists]
        
        # Team filter
        elif self.team:
            hqs = frappe.get_all(
                "HQ Master",
                filters={"team": self.team, "status": "Active"},
                pluck="name"
            )
            stockist_list = []
            for hq in hqs:
                stockists = frappe.get_all(
                    "Stockist Master",
                    filters={"hq": hq, "status": "Active"},
                    pluck="stockist_code"
                )
                stockist_list.extend(stockists)
            
            if stockist_list:
                filters["stockist_code"] = ["in", stockist_list]
        
        # Region filter
        elif self.region:
            teams = frappe.get_all(
                "Team Master",
                filters={"region": self.region, "status": "Active"},
                pluck="name"
            )
            stockist_list = []
            for team in teams:
                hqs = frappe.get_all(
                    "HQ Master",
                    filters={"team": team, "status": "Active"},
                    pluck="name"
                )
                for hq in hqs:
                    stockists = frappe.get_all(
                        "Stockist Master",
                        filters={"hq": hq, "status": "Active"},
                        pluck="stockist_code"
                    )
                    stockist_list.extend(stockists)
            
            if stockist_list:
                filters["stockist_code"] = ["in", stockist_list]
        
        return filters
    
    def aggregate_stockist_performance(self, statements):
        """
        Aggregate performance metrics per stockist
        Returns: {stockist_code: {primary_value, secondary_value, closing_value, products, ratios}}
        """
        stockist_data = {}
        
        for statement in statements:
            stmt_doc = frappe.get_doc("Stockist Statement", statement["name"])
            stockist_code = stmt_doc.stockist_code
            
            if not frappe.db.exists("Stockist Master", stockist_code):
                continue
            
            stockist = frappe.get_doc("Stockist Master", stockist_code)
            
            # Initialize stockist entry
            if stockist_code not in stockist_data:
                stockist_data[stockist_code] = {
                    "stockist_code": stockist_code,
                    "stockist_name": stockist.stockist_name,
                    "hq": stockist.hq if hasattr(stockist, 'hq') else "N/A",
                    "primary_qty": 0,
                    "secondary_qty": 0,
                    "primary_value": 0,
                    "secondary_value": 0,
                    "closing_value": 0,
                    "opening_value": 0,
                    "products": {},
                    "statement_count": 0
                }
            
            stockist_data[stockist_code]["statement_count"] += 1
            
            # Process items
            for item in stmt_doc.items:
                product_code = item.product_code
                
                # Get product details
                product = frappe.db.get_value(
                    "Product Master",
                    product_code,
                    ["product_name", "pack", "pts", "division", "category"],
                    as_dict=True
                )
                
                if not product:
                    continue
                
                # Apply division filter
                if self.division and product.division != self.division:
                    continue
                
                # Apply category filter
                if self.product_category and self.product_category != "All Products":
                    if product.category != self.product_category:
                        continue
                
                pts = flt(product.pts)
                purchase_qty = flt(item.purchase_qty)
                sales_qty = flt(item.sales_qty)
                closing_qty = flt(item.closing_qty)
                opening_qty = flt(item.opening_qty)
                
                # Aggregate values
                stockist_data[stockist_code]["primary_qty"] += purchase_qty
                stockist_data[stockist_code]["secondary_qty"] += sales_qty
                stockist_data[stockist_code]["primary_value"] += purchase_qty * pts
                stockist_data[stockist_code]["secondary_value"] += sales_qty * pts
                stockist_data[stockist_code]["closing_value"] += closing_qty * pts
                stockist_data[stockist_code]["opening_value"] += opening_qty * pts
                
                # Product-level tracking
                if self.show_top_products:
                    if product_code not in stockist_data[stockist_code]["products"]:
                        stockist_data[stockist_code]["products"][product_code] = {
                            "product_name": product.product_name,
                            "pack": product.pack,
                            "primary_qty": 0,
                            "secondary_qty": 0,
                            "primary_value": 0,
                            "secondary_value": 0
                        }
                    
                    stockist_data[stockist_code]["products"][product_code]["primary_qty"] += purchase_qty
                    stockist_data[stockist_code]["products"][product_code]["secondary_qty"] += sales_qty
                    stockist_data[stockist_code]["products"][product_code]["primary_value"] += purchase_qty * pts
                    stockist_data[stockist_code]["products"][product_code]["secondary_value"] += sales_qty * pts
        
        # Calculate ratios and filter
        filtered_data = {}
        
        for stockist_code in stockist_data:
            data = stockist_data[stockist_code]
            
            # Apply minimum filters
            if flt(self.minimum_primary_sales) > 0:
                if data["primary_value"] < flt(self.minimum_primary_sales):
                    continue
            
            if flt(self.minimum_secondary_sales) > 0:
                if data["secondary_value"] < flt(self.minimum_secondary_sales):
                    continue
            
            # Calculate ratios
            if self.calculate_ratios:
                # Primary to Secondary Ratio
                if data["secondary_value"] > 0:
                    data["primary_to_secondary_ratio"] = data["primary_value"] / data["secondary_value"]
                else:
                    data["primary_to_secondary_ratio"] = 0
                
                # Stock Turnover Ratio (Sales / Average Stock)
                avg_stock = (data["opening_value"] + data["closing_value"]) / 2
                if avg_stock > 0:
                    data["stock_turnover_ratio"] = data["secondary_value"] / avg_stock
                else:
                    data["stock_turnover_ratio"] = 0
                
                # Days of Inventory (Closing Stock / Average Daily Sales)
                total_days = (getdate(self.to_date) - getdate(self.from_date)).days + 1
                avg_daily_sales = data["secondary_value"] / total_days if total_days > 0 else 0
                if avg_daily_sales > 0:
                    data["days_of_inventory"] = data["closing_value"] / avg_daily_sales
                else:
                    data["days_of_inventory"] = 0
            
            # Get top products
            if self.show_top_products:
                products = data["products"]
                sorted_products = sorted(
                    products.items(),
                    key=lambda x: x[1]["secondary_value"],
                    reverse=True
                )
                
                top_n = int(self.top_n_products or 5)
                data["top_products"] = []
                
                for product_code, product_data in sorted_products[:top_n]:
                    data["top_products"].append({
                        "product_code": product_code,
                        "product_name": product_data["product_name"],
                        "secondary_value": product_data["secondary_value"],
                        "secondary_qty": product_data["secondary_qty"]
                    })
                
                # Remove full product dict to reduce size
                del data["products"]
            
            filtered_data[stockist_code] = data
        
        return filtered_data
    
    def add_previous_period_comparison(self, current_data):
        """Add previous period comparison data"""
        try:
            # Calculate previous period dates
            months_back = int(self.previous_period_months or 1)
            prev_from_date = add_months(getdate(self.from_date), -months_back)
            prev_to_date = add_months(getdate(self.to_date), -months_back)
            
            # Build filters for previous period
            prev_filters = {"docstatus": 1}
            prev_filters["statement_month"] = ["between", [prev_from_date, prev_to_date]]
            
            # Get stockist codes from current data
            stockist_codes = list(current_data.keys())
            if stockist_codes:
                prev_filters["stockist_code"] = ["in", stockist_codes]
            
            prev_statements = frappe.get_all(
                "Stockist Statement",
                filters=prev_filters,
                fields=["name", "stockist_code"]
            )
            
            # Aggregate previous period data
            prev_data = {}
            
            for statement in prev_statements:
                stmt_doc = frappe.get_doc("Stockist Statement", statement["name"])
                stockist_code = stmt_doc.stockist_code
                
                if stockist_code not in prev_data:
                    prev_data[stockist_code] = {
                        "primary_value": 0,
                        "secondary_value": 0,
                        "closing_value": 0
                    }
                
                for item in stmt_doc.items:
                    product = frappe.db.get_value(
                        "Product Master",
                        item.product_code,
                        ["pts", "division", "category"],
                        as_dict=True
                    )
                    
                    if not product:
                        continue
                    
                    # Apply same filters
                    if self.division and product.division != self.division:
                        continue
                    
                    if self.product_category and self.product_category != "All Products":
                        if product.category != self.product_category:
                            continue
                    
                    pts = flt(product.pts)
                    prev_data[stockist_code]["primary_value"] += flt(item.purchase_qty) * pts
                    prev_data[stockist_code]["secondary_value"] += flt(item.sales_qty) * pts
                    prev_data[stockist_code]["closing_value"] += flt(item.closing_qty) * pts
            
            # Add comparison to current data
            for stockist_code in current_data:
                if stockist_code in prev_data:
                    prev = prev_data[stockist_code]
                    current = current_data[stockist_code]
                    
                    # Calculate growth percentages
                    if prev["primary_value"] > 0:
                        current["primary_growth"] = ((current["primary_value"] - prev["primary_value"]) / prev["primary_value"]) * 100
                    else:
                        current["primary_growth"] = 0
                    
                    if prev["secondary_value"] > 0:
                        current["secondary_growth"] = ((current["secondary_value"] - prev["secondary_value"]) / prev["secondary_value"]) * 100
                    else:
                        current["secondary_growth"] = 0
                    
                    current["prev_primary_value"] = prev["primary_value"]
                    current["prev_secondary_value"] = prev["secondary_value"]
                else:
                    current["primary_growth"] = 0
                    current["secondary_growth"] = 0
                    current["prev_primary_value"] = 0
                    current["prev_secondary_value"] = 0
        
        except Exception as e:
            frappe.log_error(f"Error comparing periods: {str(e)}", "Period Comparison Error")
        
        return current_data
    
    def calculate_summary(self):
        """Calculate summary statistics"""
        if not self.report_data:
            return
        
        try:
            data = json.loads(self.report_data)
            
            total_stockists = len(data)
            total_primary = 0
            total_secondary = 0
            total_closing = 0
            total_ratio_ps = 0
            total_ratio_turnover = 0
            count_ratios = 0
            
            for stockist_code in data:
                entry = data[stockist_code]
                
                total_primary += flt(entry.get("primary_value", 0))
                total_secondary += flt(entry.get("secondary_value", 0))
                total_closing += flt(entry.get("closing_value", 0))
                
                if self.calculate_ratios:
                    total_ratio_ps += flt(entry.get("primary_to_secondary_ratio", 0))
                    total_ratio_turnover += flt(entry.get("stock_turnover_ratio", 0))
                    count_ratios += 1
            
            self.total_stockists = total_stockists
            self.total_primary_value = total_primary
            self.total_secondary_value = total_secondary
            self.total_closing_value = total_closing
            
            if count_ratios > 0:
                self.average_primary_to_secondary_ratio = total_ratio_ps / count_ratios
                self.average_stock_turnover_ratio = total_ratio_turnover / count_ratios
            else:
                self.average_primary_to_secondary_ratio = 0
                self.average_stock_turnover_ratio = 0
        
        except Exception as e:
            frappe.log_error(f"Error calculating summary: {str(e)}", "Summary Error")


@frappe.whitelist()
def generate_report(doc_name):
    """API to generate report"""
    doc = frappe.get_doc("Stockist Performance Report", doc_name)
    doc.generate_report_data()
    doc.save()
    frappe.db.commit()
    
    return {
        "success": True,
        "message": "Report generated successfully",
        "total_stockists": doc.total_stockists,
        "total_secondary_value": doc.total_secondary_value
    }


@frappe.whitelist()
def export_to_excel(doc_name):
    """Export to Excel"""
    doc = frappe.get_doc("Stockist Performance Report", doc_name)
    
    if not doc.report_data:
        frappe.throw("No report data available. Please generate the report first.")
    
    data = json.loads(doc.report_data or "{}")
    
    if not data:
        frappe.throw("No data available to export")
    
    wb = openpyxl.Workbook()
    
    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Performance Summary"
    
    # Styling
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws_summary.merge_cells('A1:L1')
    ws_summary['A1'] = "STEDMAN PHARMACEUTICALS PVT LTD"
    ws_summary['A1'].font = Font(bold=True, size=14)
    ws_summary['A1'].alignment = Alignment(horizontal='center')
    
    ws_summary.merge_cells('A2:L2')
    ws_summary['A2'] = "Stockist Performance Report"
    ws_summary['A2'].font = Font(size=12, italic=True)
    ws_summary['A2'].alignment = Alignment(horizontal='center')
    
    ws_summary.merge_cells('A3:L3')
    ws_summary['A3'] = f"Period: {doc.from_date} to {doc.to_date}"
    ws_summary['A3'].font = Font(size=10)
    ws_summary['A3'].alignment = Alignment(horizontal='center')
    
    # Headers
    row = 5
    headers = ["Stockist Code", "Stockist Name", "HQ", "Primary Value", "Secondary Value",
              "Closing Value", "P/S Ratio", "Stock Turnover", "Days of Inv"]
    
    if doc.compare_with_previous_period:
        headers.extend(["Primary Growth %", "Secondary Growth %"])
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws_summary.cell(row=row, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # Data rows
    row = 6
    
    for stockist_code in sorted(data.keys()):
        entry = data[stockist_code]
        
        ws_summary.cell(row=row, column=1).value = entry["stockist_code"]
        ws_summary.cell(row=row, column=2).value = entry["stockist_name"]
        ws_summary.cell(row=row, column=3).value = entry["hq"]
        ws_summary.cell(row=row, column=4).value = flt(entry["primary_value"])
        ws_summary.cell(row=row, column=5).value = flt(entry["secondary_value"])
        ws_summary.cell(row=row, column=6).value = flt(entry["closing_value"])
        ws_summary.cell(row=row, column=7).value = flt(entry.get("primary_to_secondary_ratio", 0))
        ws_summary.cell(row=row, column=8).value = flt(entry.get("stock_turnover_ratio", 0))
        ws_summary.cell(row=row, column=9).value = flt(entry.get("days_of_inventory", 0))
        
        col = 10
        if doc.compare_with_previous_period:
            ws_summary.cell(row=row, column=col).value = flt(entry.get("primary_growth", 0)) / 100
            ws_summary.cell(row=row, column=col + 1).value = flt(entry.get("secondary_growth", 0)) / 100
            col += 2
        
        # Format cells
        for c in range(1, len(headers) + 1):
            cell = ws_summary.cell(row=row, column=c)
            cell.border = border
            if c >= 4:
                cell.alignment = Alignment(horizontal='right')
            if c >= 4 and c <= 6:
                cell.number_format = '#,##0.00'
            elif c >= 7 and c <= 9:
                cell.number_format = '0.00'
            elif c >= 10 and doc.compare_with_previous_period:
                cell.number_format = '0.00%'
        
        row += 1
    
    # Totals
    summary_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    summary_font = Font(bold=True)
    
    ws_summary.cell(row=row, column=1).value = "TOTAL"
    ws_summary.cell(row=row, column=4).value = doc.total_primary_value
    ws_summary.cell(row=row, column=5).value = doc.total_secondary_value
    ws_summary.cell(row=row, column=6).value = doc.total_closing_value
    ws_summary.cell(row=row, column=7).value = doc.average_primary_to_secondary_ratio
    ws_summary.cell(row=row, column=8).value = doc.average_stock_turnover_ratio
    
    for c in range(1, len(headers) + 1):
        cell = ws_summary.cell(row=row, column=c)
        cell.fill = summary_fill
        cell.font = summary_font
        cell.border = border
    
    # Adjust widths
    ws_summary.column_dimensions['A'].width = 15
    ws_summary.column_dimensions['B'].width = 35
    ws_summary.column_dimensions['C'].width = 20
    for col in range(4, len(headers) + 1):
        ws_summary.column_dimensions[get_column_letter(col)].width = 15
    
    # Sheet 2: Top Products (if enabled)
    if doc.show_top_products:
        ws_products = wb.create_sheet("Top Products")
        
        # Title
        ws_products.merge_cells('A1:F1')
        ws_products['A1'] = "Top Performing Products by Stockist"
        ws_products['A1'].font = Font(bold=True, size=14)
        ws_products['A1'].alignment = Alignment(horizontal='center')
        
        row = 3
        
        for stockist_code in sorted(data.keys()):
            entry = data[stockist_code]
            
            if "top_products" in entry and entry["top_products"]:
                # Stockist header
                ws_products.merge_cells(f'A{row}:F{row}')
                ws_products[f'A{row}'] = f"{entry['stockist_name']} ({entry['stockist_code']})"
                ws_products[f'A{row}'].font = Font(bold=True, size=11)
                ws_products[f'A{row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                row += 1
                
                # Product headers
                prod_headers = ["Rank", "Product Code", "Product Name", "Quantity", "Value", "% of Total"]
                for col_idx, header in enumerate(prod_headers, 1):
                    cell = ws_products.cell(row=row, column=col_idx)
                    cell.value = header
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.border = border
                
                row += 1
                
                # Products
                for rank, product in enumerate(entry["top_products"], 1):
                    ws_products.cell(row=row, column=1).value = rank
                    ws_products.cell(row=row, column=2).value = product["product_code"]
                    ws_products.cell(row=row, column=3).value = product["product_name"]
                    ws_products.cell(row=row, column=4).value = flt(product["secondary_qty"])
                    ws_products.cell(row=row, column=5).value = flt(product["secondary_value"])
                    
                    # Calculate percentage
                    if entry["secondary_value"] > 0:
                        pct = (product["secondary_value"] / entry["secondary_value"]) * 100
                        ws_products.cell(row=row, column=6).value = pct / 100
                        ws_products.cell(row=row, column=6).number_format = '0.00%'
                    
                    for col in range(1, 7):
                        ws_products.cell(row=row, column=col).border = border
                    
                    row += 1
                
                row += 2  # Space between stockists
        
        # Adjust widths
        ws_products.column_dimensions['A'].width = 8
        ws_products.column_dimensions['B'].width = 15
        ws_products.column_dimensions['C'].width = 35
        ws_products.column_dimensions['D'].width = 12
        ws_products.column_dimensions['E'].width = 15
        ws_products.column_dimensions['F'].width = 12
    
    # Save
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    file_doc = frappe.get_doc({
        "doctype": "File",
        "file_name": f"Stockist_Performance_{doc.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        "attached_to_doctype": "Stockist Performance Report",
        "attached_to_name": doc_name,
        "content": output.getvalue(),
        "is_private": 0
    })
    file_doc.save(ignore_permissions=True)
    
    return {
        "success": True,
        "file_url": file_doc.file_url
    }


@frappe.whitelist()
def export_to_pdf(doc_name):
    """Export to PDF"""
    from frappe.utils.pdf import get_pdf
    
    doc = frappe.get_doc("Stockist Performance Report", doc_name)
    
    if not doc.report_data:
        frappe.throw("No report data available. Please generate the report first.")
    
    html = render_report_html(doc)
    pdf_bytes = get_pdf(html)
    
    file_doc = frappe.get_doc({
        "doctype": "File",
        "file_name": f"Stockist_Performance_{doc.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
        "attached_to_doctype": "Stockist Performance Report",
        "attached_to_name": doc_name,
        "content": pdf_bytes,
        "is_private": 0
    })
    file_doc.save(ignore_permissions=True)
    
    return {
        "success": True,
        "file_url": file_doc.file_url
    }


def render_report_html(doc):
    """Render HTML for PDF"""
    data = json.loads(doc.report_data or "{}")
    
    html = f"""
    <html>
    <head>
        <style>
            @page {{ size: A4 landscape; margin: 10mm; }}
            body {{ font-family: Arial, sans-serif; margin: 0; font-size: 8px; }}
            .header {{ text-align: center; margin-bottom: 10px; }}
            .title {{ font-size: 14px; font-weight: bold; }}
            .subtitle {{ font-size: 10px; color: #666; }}
            table {{ width: 100%; border-collapse: collapse; }}
            th {{ background-color: #1F4E78; color: white; padding: 4px; text-align: center; border: 1px solid #000; font-size: 7px; }}
            td {{ padding: 3px; border: 1px solid #ccc; font-size: 7px; }}
            .number {{ text-align: right; }}
            .text-left {{ text-align: left; }}
            .total-row {{ background-color: #D9E1F2; font-weight: bold; }}
        </style>
    </head>
    <body>
        <div class="header">
            <div class="title">STEDMAN PHARMACEUTICALS PVT LTD</div>
            <div class="subtitle">Stockist Performance Report</div>
            <div class="subtitle">Period: {doc.from_date} to {doc.to_date}</div>
        </div>
        
        <table>
            <thead>
                <tr>
                    <th>Stockist</th>
                    <th>HQ</th>
                    <th>Primary</th>
                    <th>Secondary</th>
                    <th>Closing</th>
                    <th>P/S Ratio</th>
                    <th>Turnover</th>
    """
    
    if doc.compare_with_previous_period:
        html += "<th>Pri Growth %</th><th>Sec Growth %</th>"
    
    html += """
                </tr>
            </thead>
            <tbody>
    """
    
    for stockist_code in sorted(data.keys()):
        entry = data[stockist_code]
        
        html += f"""
        <tr>
            <td class='text-left'>{entry['stockist_name'][:30]}</td>
            <td class='text-left'>{entry['hq']}</td>
            <td class='number'>{flt(entry['primary_value']):.2f}</td>
            <td class='number'>{flt(entry['secondary_value']):.2f}</td>
            <td class='number'>{flt(entry['closing_value']):.2f}</td>
            <td class='number'>{flt(entry.get('primary_to_secondary_ratio', 0)):.2f}</td>
            <td class='number'>{flt(entry.get('stock_turnover_ratio', 0)):.2f}</td>
        """
        
        if doc.compare_with_previous_period:
            html += f"""
            <td class='number'>{flt(entry.get('primary_growth', 0)):.1f}%</td>
            <td class='number'>{flt(entry.get('secondary_growth', 0)):.1f}%</td>
            """
        
        html += "</tr>"
    
    html += f"""
            <tr class='total-row'>
                <td colspan='2' class='text-left'>TOTAL</td>
                <td class='number'>{flt(doc.total_primary_value):.2f}</td>
                <td class='number'>{flt(doc.total_secondary_value):.2f}</td>
                <td class='number'>{flt(doc.total_closing_value):.2f}</td>
                <td class='number'>{flt(doc.average_primary_to_secondary_ratio):.2f}</td>
                <td class='number'>{flt(doc.average_stock_turnover_ratio):.2f}</td>
    """
    
    if doc.compare_with_previous_period:
        html += "<td colspan='2'></td>"
    
    html += """
            </tr>
            </tbody>
        </table>
    </body>
    </html>
    """
    
    return html
