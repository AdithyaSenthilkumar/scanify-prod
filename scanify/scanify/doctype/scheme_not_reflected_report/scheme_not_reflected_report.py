import frappe
from frappe.model.document import Document
from frappe.utils import flt, getdate, date_diff, now
from datetime import datetime
import json
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class SchemeNotReflectedReport(Document):
    def validate(self):
        """Validate report parameters"""
        if self.from_date and self.to_date:
            if getdate(self.from_date) > getdate(self.to_date):
                frappe.throw("Statement Period From Date cannot be greater than To Date")
        
        if self.scheme_approval_from_date and self.scheme_approval_to_date:
            if getdate(self.scheme_approval_from_date) > getdate(self.scheme_approval_to_date):
                frappe.throw("Scheme Approval From Date cannot be greater than To Date")
    
    def before_save(self):
        """Set audit fields"""
        if not self.report_date:
            self.report_date = now()
        if not self.generated_by:
            self.generated_by = frappe.session.user
    
    def on_submit(self):
        """Generate on submit"""
        self.generate_report_data()
        self.calculate_summary()
    
    def generate_report_data(self):
        """Generate scheme not reflected analysis"""
        try:
            # Get approved schemes in the period
            schemes = self.get_approved_schemes()
            
            if not schemes:
                frappe.msgprint("No approved schemes found for the selected period", indicator='orange')
                self.report_data = json.dumps({}, default=str)
                return {}
            
            # Analyze each scheme
            unreflected_data = self.analyze_scheme_reflection(schemes)
            
            # Store as JSON
            self.report_data = json.dumps(unreflected_data, default=str)
            
            # Calculate summary
            self.calculate_summary()
            
            return unreflected_data
            
        except Exception as e:
            frappe.log_error(frappe.get_traceback(), "Scheme Not Reflected Report Error")
            frappe.throw(f"Error generating report: {str(e)}")
    
    def get_approved_schemes(self):
        """Get all approved schemes in the period"""
        filters = {
            "docstatus": 1,  # Submitted/Approved
            "approval_date": ["between", [self.scheme_approval_from_date, self.scheme_approval_to_date]]
        }
        
        # Apply filters
        if self.stockist:
            filters["stockist_code"] = self.stockist
        
        if self.doctor:
            filters["doctor_code"] = self.doctor
        
        # For HQ filter - need to get doctors from that HQ's team
        if self.hq:
            # Get team for this HQ
            hq_team = frappe.db.get_value("HQ Master", self.hq, "team")
            if hq_team:
                doctors = frappe.get_all(
                    "Doctor Master",
                    filters={"team": hq_team, "status": "Active"},
                    pluck="doctor_code"
                )
                if doctors:
                    filters["doctor_code"] = ["in", doctors]
        
        # For team filter
        elif self.team:
            doctors = frappe.get_all(
                "Doctor Master",
                filters={"team": self.team, "status": "Active"},
                pluck="doctor_code"
            )
            if doctors:
                filters["doctor_code"] = ["in", doctors]
        
        # For region filter
        elif self.region:
            # Get all teams in this region
            teams = frappe.get_all(
                "Team Master",
                filters={"region": self.region, "status": "Active"},
                pluck="name"
            )
            
            doctors = []
            for team in teams:
                team_doctors = frappe.get_all(
                    "Doctor Master",
                    filters={"team": team, "status": "Active"},
                    pluck="doctor_code"
                )
                doctors.extend(team_doctors)
            
            if doctors:
                filters["doctor_code"] = ["in", doctors]
        
        schemes = frappe.get_all(
            "Scheme Request",
            filters=filters,
            fields=["name", "doctor_code", "stockist_code", "approval_date", "entry_date"]
        )
        
        return schemes

    
    def analyze_scheme_reflection(self, schemes):
        """
        Analyze each scheme to check if products are reflected in secondary sales
        Returns: {scheme_name: {doctor, stockist, products: [{product, approved_qty, reflected_qty, unreflected_qty, closing_qty}]}}
        """
        unreflected_data = {}
        
        for scheme in schemes:
            scheme_doc = frappe.get_doc("Scheme Request", scheme["name"])
            
            # Get doctor and stockist details
            doctor = frappe.get_doc("Doctor Master", scheme_doc.doctor_code)
            stockist = frappe.get_doc("Stockist Master", scheme_doc.stockist_code)
            
            # Calculate days since approval
            days_since_approval = date_diff(getdate(self.to_date), getdate(scheme_doc.approval_date))
            
            # Skip if not enough days passed
            if days_since_approval < int(self.minimum_days_unreflected or 0):
                continue
            
            scheme_key = scheme_doc.name  # Use name as key
            
            unreflected_products = []
            
            # Analyze each product in the scheme
            for item in scheme_doc.items:
                product_code = item.product_code
                
                # Apply product filter
                if self.product_code and product_code != self.product_code:
                    continue
                
                approved_free_qty = flt(item.free_quantity)
                
                # Skip if below minimum
                if approved_free_qty < flt(self.minimum_scheme_qty or 0):
                    continue
                
                # Get secondary sales for this product from stockist in the period
                reflected_qty = self.get_secondary_sales_qty(
                    scheme_doc.stockist_code,
                    product_code,
                    self.from_date,
                    self.to_date
                )
                
                # Get closing stock
                closing_qty = self.get_closing_stock_qty(
                    scheme_doc.stockist_code,
                    product_code,
                    self.to_date
                )
                
                unreflected_qty = approved_free_qty - reflected_qty
                
                # Check if unreflected
                if unreflected_qty > 0:
                    # Include based on partial reflection setting
                    if not self.include_partial_reflection and reflected_qty > 0:
                        continue
                    
                    # Get product details
                    product = frappe.db.get_value(
                        "Product Master",
                        product_code,
                        ["product_name", "pack", "pts"],
                        as_dict=True
                    )
                    
                    if not product:
                        continue
                    
                    closing_value = closing_qty * flt(product.pts)
                    
                    unreflected_products.append({
                        "product_code": product_code,
                        "product_name": product.product_name,
                        "pack": product.pack,
                        "approved_free_qty": approved_free_qty,
                        "reflected_qty": reflected_qty,
                        "unreflected_qty": unreflected_qty,
                        "closing_qty": closing_qty,
                        "closing_value": closing_value,
                        "pts": product.pts,
                        "reflection_percentage": (reflected_qty / approved_free_qty * 100) if approved_free_qty > 0 else 0
                    })
            
            # Only add if there are unreflected products
            if unreflected_products:
                unreflected_data[scheme_key] = {
                    "scheme_number": scheme_key,  # Use name field
                    "doctor_code": scheme_doc.doctor_code,
                    "doctor_name": doctor.doctor_name,
                    "team": doctor.team if hasattr(doctor, 'team') else "N/A",
                    "region": doctor.region if hasattr(doctor, 'region') else "N/A",
                    "city_pool": doctor.city_pool if hasattr(doctor, 'city_pool') else "N/A",
                    "stockist_code": scheme_doc.stockist_code,
                    "stockist_name": stockist.stockist_name,
                    "stockist_hq": stockist.hq if hasattr(stockist, 'hq') else "N/A",
                    "approval_date": str(scheme_doc.approval_date),
                    "entry_date": str(scheme_doc.entry_date),
                    "days_since_approval": days_since_approval,
                    "unreflected_products": unreflected_products
                }
        
        return unreflected_data

    
    def get_secondary_sales_qty(self, stockist_code, product_code, from_date, to_date):
        """Get total secondary sales quantity for a product from stockist in period"""
        try:
            statements = frappe.get_all(
                "Stockist Statement",
                filters={
                    "stockist_code": stockist_code,
                    "statement_month": ["between", [from_date, to_date]],
                    "docstatus": 1
                },
                pluck="name"
            )
            
            if not statements:
                return 0
            
            total_sales_qty = 0
            
            for statement in statements:
                sales_qty = frappe.db.get_value(
                    "Stockist Statement Item",
                    {
                        "parent": statement,
                        "product_code": product_code
                    },
                    "sales_qty"
                )
                
                total_sales_qty += flt(sales_qty or 0)
            
            return total_sales_qty
        
        except Exception:
            return 0
    
    def get_closing_stock_qty(self, stockist_code, product_code, as_of_date):
        """Get closing stock quantity as of a date"""
        try:
            # Get the latest statement on or before the date
            statement = frappe.get_value(
                "Stockist Statement",
                {
                    "stockist_code": stockist_code,
                    "statement_month": ["<=", as_of_date],
                    "docstatus": 1
                },
                "name",
                order_by="statement_month desc"
            )
            
            if not statement:
                return 0
            
            closing_qty = frappe.db.get_value(
                "Stockist Statement Item",
                {
                    "parent": statement,
                    "product_code": product_code
                },
                "closing_qty"
            )
            
            return flt(closing_qty or 0)
        
        except Exception:
            return 0
    
    def calculate_summary(self):
        """Calculate summary statistics"""
        if not self.report_data:
            return
        
        try:
            data = json.loads(self.report_data)
            
            total_schemes_analyzed = len(self.get_approved_schemes())
            total_unreflected_schemes = len(data)
            total_unreflected_qty = 0
            total_closing_value = 0
            total_approved_qty = 0
            total_reflected_qty = 0
            
            for scheme_key in data:
                scheme_data = data[scheme_key]
                
                for product in scheme_data.get("unreflected_products", []):
                    total_unreflected_qty += flt(product.get("unreflected_qty", 0))
                    total_closing_value += flt(product.get("closing_value", 0))
                    total_approved_qty += flt(product.get("approved_free_qty", 0))
                    total_reflected_qty += flt(product.get("reflected_qty", 0))
            
            self.total_schemes_analyzed = total_schemes_analyzed
            self.total_unreflected_schemes = total_unreflected_schemes
            self.total_unreflected_qty = total_unreflected_qty
            self.total_closing_stock_value = total_closing_value
            
            if total_approved_qty > 0:
                self.reflection_percentage = (total_reflected_qty / total_approved_qty) * 100
            else:
                self.reflection_percentage = 0
        
        except Exception as e:
            frappe.log_error(f"Error calculating summary: {str(e)}", "Summary Error")


@frappe.whitelist()
def generate_report(doc_name):
    """API to generate report"""
    doc = frappe.get_doc("Scheme Not Reflected Report", doc_name)
    doc.generate_report_data()
    doc.save()
    frappe.db.commit()
    
    return {
        "success": True,
        "message": "Report generated successfully",
        "total_unreflected": doc.total_unreflected_schemes,
        "reflection_percentage": doc.reflection_percentage
    }


@frappe.whitelist()
def export_to_excel(doc_name):
    """Export to Excel with detailed analysis"""
    doc = frappe.get_doc("Scheme Not Reflected Report", doc_name)
    
    if not doc.report_data:
        frappe.throw("No report data available. Please generate the report first.")
    
    data = json.loads(doc.report_data or "{}")
    
    if not data:
        frappe.throw("No unreflected schemes found")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Unreflected Schemes"
    
    # Styling
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    warning_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    caution_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:N1')
    ws['A1'] = "STEDMAN PHARMACEUTICALS PVT LTD"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:N2')
    ws['A2'] = "Scheme Approved Products Not Reflected in Secondary Sales"
    ws['A2'].font = Font(size=12, italic=True)
    ws['A2'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A3:N3')
    ws['A3'] = f"Statement Period: {doc.from_date} to {doc.to_date} | Scheme Approval: {doc.scheme_approval_from_date} to {doc.scheme_approval_to_date}"
    ws['A3'].font = Font(size=10)
    ws['A3'].alignment = Alignment(horizontal='center')
    
    # Summary
    ws.merge_cells('A4:N4')
    ws['A4'] = f"Total Schemes Analyzed: {doc.total_schemes_analyzed} | Unreflected: {doc.total_unreflected_schemes} | Reflection Rate: {doc.reflection_percentage:.2f}%"
    ws['A4'].font = Font(size=10, bold=True)
    ws['A4'].alignment = Alignment(horizontal='center')
    
    # Headers
    row = 6
    headers = ["Scheme No", "Doctor", "Team", "Stockist", "Stockist HQ", "Product Code", 
              "Product Name", "Approved Qty", "Reflected Qty", "Unreflected Qty",
              "Closing Stock", "Closing Value", "Reflection %", "Days Since Approval"]
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # Data rows
    row = 7
    
    for scheme_key in sorted(data.keys()):
        scheme_data = data[scheme_key]
        
        for product in scheme_data.get("unreflected_products", []):
            ws.cell(row=row, column=1).value = scheme_data["scheme_number"]
            ws.cell(row=row, column=2).value = scheme_data["doctor_name"]
            ws.cell(row=row, column=3).value = scheme_data["team"]
            ws.cell(row=row, column=4).value = scheme_data["stockist_name"]
            ws.cell(row=row, column=5).value = scheme_data.get("stockist_hq", "N/A")
            ws.cell(row=row, column=6).value = product["product_code"]
            ws.cell(row=row, column=7).value = product["product_name"]
            ws.cell(row=row, column=8).value = flt(product["approved_free_qty"])
            ws.cell(row=row, column=9).value = flt(product["reflected_qty"])
            ws.cell(row=row, column=10).value = flt(product["unreflected_qty"])
            ws.cell(row=row, column=11).value = flt(product["closing_qty"])
            ws.cell(row=row, column=12).value = flt(product["closing_value"])
            ws.cell(row=row, column=13).value = flt(product["reflection_percentage"]) / 100
            ws.cell(row=row, column=14).value = scheme_data["days_since_approval"]
            
            # Apply conditional formatting
            reflection_pct = flt(product["reflection_percentage"])
            
            for col in range(1, 15):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                
                # Highlight based on reflection percentage
                if reflection_pct == 0 and col == 13:
                    cell.fill = warning_fill  # Red for 0%
                elif reflection_pct < 50 and col == 13:
                    cell.fill = caution_fill  # Yellow for < 50%
                
                # Number formatting
                if col in [8, 9, 10, 11]:
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal='right')
                elif col == 12:
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
                elif col == 13:
                    cell.number_format = '0.00%'
                    cell.alignment = Alignment(horizontal='right')
            
            row += 1
    
    # Summary row
    summary_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    summary_font = Font(bold=True)
    
    ws.cell(row=row, column=1).value = "TOTAL"
    ws.cell(row=row, column=10).value = doc.total_unreflected_qty
    ws.cell(row=row, column=12).value = doc.total_closing_stock_value
    ws.cell(row=row, column=13).value = doc.reflection_percentage / 100
    
    for col in range(1, 15):
        cell = ws.cell(row=row, column=col)
        cell.fill = summary_fill
        cell.font = summary_font
        cell.border = border
        if col == 13:
            cell.number_format = '0.00%'
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 30
    for col in range(8, 15):
        ws.column_dimensions[get_column_letter(col)].width = 13
    
    # Freeze panes
    ws.freeze_panes = 'A7'
    
    # Save
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    file_doc = frappe.get_doc({
        "doctype": "File",
        "file_name": f"Scheme_Not_Reflected_{doc.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        "attached_to_doctype": "Scheme Not Reflected Report",
        "attached_to_name": doc_name,
        "content": output.getvalue(),
        "is_private": 0
    })
    file_doc.save(ignore_permissions=True)
    
    return {
        "success": True,
        "file_url": file_doc.file_url
    }


@frappe.whitelist()
def export_to_pdf(doc_name):
    """Export to PDF"""
    from frappe.utils.pdf import get_pdf
    
    doc = frappe.get_doc("Scheme Not Reflected Report", doc_name)
    
    if not doc.report_data:
        frappe.throw("No report data available. Please generate the report first.")
    
    html = render_report_html(doc)
    pdf_bytes = get_pdf(html)
    
    file_doc = frappe.get_doc({
        "doctype": "File",
        "file_name": f"Scheme_Not_Reflected_{doc.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
        "attached_to_doctype": "Scheme Not Reflected Report",
        "attached_to_name": doc_name,
        "content": pdf_bytes,
        "is_private": 0
    })
    file_doc.save(ignore_permissions=True)
    
    return {
        "success": True,
        "file_url": file_doc.file_url
    }


def render_report_html(doc):
    """Render HTML for PDF"""
    data = json.loads(doc.report_data or "{}")
    
    html = f"""
    <html>
    <head>
        <style>
            @page {{ size: A4 landscape; margin: 10mm; }}
            body {{ font-family: Arial, sans-serif; margin: 0; font-size: 8px; }}
            .header {{ text-align: center; margin-bottom: 10px; }}
            .title {{ font-size: 14px; font-weight: bold; }}
            .subtitle {{ font-size: 10px; color: #666; }}
            .summary {{ text-align: center; font-size: 9px; font-weight: bold; margin: 5px 0; }}
            table {{ width: 100%; border-collapse: collapse; }}
            th {{ background-color: #1F4E78; color: white; padding: 4px; text-align: center; border: 1px solid #000; font-size: 7px; }}
            td {{ padding: 3px; border: 1px solid #ccc; font-size: 7px; }}
            .number {{ text-align: right; }}
            .text-left {{ text-align: left; }}
            .warning {{ background-color: #FFC7CE; }}
            .caution {{ background-color: #FFEB9C; }}
            .total-row {{ background-color: #D9E1F2; font-weight: bold; }}
        </style>
    </head>
    <body>
        <div class="header">
            <div class="title">STEDMAN PHARMACEUTICALS PVT LTD</div>
            <div class="subtitle">Scheme Approved Products Not Reflected in Secondary Sales</div>
            <div class="subtitle">Period: {doc.from_date} to {doc.to_date}</div>
        </div>
        
        <div class="summary">
            Total Schemes: {doc.total_schemes_analyzed} | Unreflected: {doc.total_unreflected_schemes} | 
            Reflection Rate: {doc.reflection_percentage:.2f}%
        </div>
        
        <table>
            <thead>
                <tr>
                    <th>Scheme</th>
                    <th>Doctor</th>
                    <th>Team</th>
                    <th>Stockist</th>
                    <th>Product</th>
                    <th>Approved</th>
                    <th>Reflected</th>
                    <th>Unreflected</th>
                    <th>Closing</th>
                    <th>Value</th>
                    <th>%</th>
                    <th>Days</th>
                </tr>
            </thead>
            <tbody>
    """
    
    for scheme_key in sorted(data.keys()):
        scheme_data = data[scheme_key]
        
        for product in scheme_data.get("unreflected_products", []):
            reflection_pct = flt(product["reflection_percentage"])
            row_class = ""
            if reflection_pct == 0:
                row_class = "warning"
            elif reflection_pct < 50:
                row_class = "caution"
            
            html += f"""
            <tr class="{row_class}">
                <td class='text-left'>{scheme_data['scheme_number']}</td>
                <td class='text-left'>{scheme_data['doctor_name']}</td>
                <td class='text-left'>{scheme_data['team']}</td>
                <td class='text-left'>{scheme_data['stockist_name'][:20]}</td>
                <td class='text-left'>{product['product_code']}</td>
                <td class='number'>{flt(product['approved_free_qty']):.0f}</td>
                <td class='number'>{flt(product['reflected_qty']):.0f}</td>
                <td class='number'>{flt(product['unreflected_qty']):.0f}</td>
                <td class='number'>{flt(product['closing_qty']):.0f}</td>
                <td class='number'>{flt(product['closing_value']):.2f}</td>
                <td class='number'>{reflection_pct:.1f}%</td>
                <td class='number'>{scheme_data['days_since_approval']}</td>
            </tr>
            """
    
    html += f"""
            <tr class='total-row'>
                <td colspan='7' class='text-left'>TOTAL</td>
                <td class='number'>{flt(doc.total_unreflected_qty):.0f}</td>
                <td colspan='1'></td>
                <td class='number'>{flt(doc.total_closing_stock_value):.2f}</td>
                <td class='number'>{flt(doc.reflection_percentage):.2f}%</td>
                <td></td>
            </tr>
            </tbody>
        </table>
    </body>
    </html>
    """
    
    return html
