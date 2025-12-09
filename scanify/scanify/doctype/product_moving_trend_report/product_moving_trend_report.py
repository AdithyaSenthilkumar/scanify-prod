import frappe
from frappe.model.document import Document
from frappe.utils import flt, getdate, add_months, now
from datetime import datetime
import json
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ProductMovingTrendReport(Document):
    def validate(self):
        """Validate report parameters"""
        if self.from_date and self.to_date:
            if getdate(self.from_date) > getdate(self.to_date):
                frappe.throw("From Date cannot be greater than To Date")
    
    def before_save(self):
        """Set audit fields"""
        if not self.report_date:
            self.report_date = now()
        if not self.generated_by:
            self.generated_by = frappe.session.user
    
    def on_submit(self):
        """Regenerate on submit"""
        self.generate_report_data()
        self.calculate_totals()
    
    def generate_report_data(self):
        """Generate product-wise moving trend data"""
        try:
            filters = self.build_filters()
            
            # Fetch stockist statements
            statements = frappe.get_all(
                "Stockist Statement",
                filters=filters,
                fields=["name", "stockist_code", "statement_month", "from_date", "to_date"]
            )
            
            if not statements:
                frappe.msgprint("No data found for selected filters", indicator='orange')
                self.report_data = json.dumps({}, default=str)
                return {}
            
            # Aggregate product-wise data
            product_data = self.aggregate_product_data(statements)
            
            # Store as JSON
            self.report_data = json.dumps(product_data, default=str)
            
            # Calculate totals
            self.calculate_totals()
            
            return product_data
            
        except Exception as e:
            frappe.log_error(frappe.get_traceback(), "Product Moving Trend Report Error")
            frappe.throw(f"Error generating report: {str(e)}")
    
    def build_filters(self):
        """Build filters for Stockist Statement"""
        filters = {"docstatus": 1}
        
        # Date filters
        if self.from_date and self.to_date:
            filters["statement_month"] = ["between", [self.from_date, self.to_date]]
        elif self.from_date:
            filters["statement_month"] = [">=", self.from_date]
        elif self.to_date:
            filters["statement_month"] = ["<=", self.to_date]
        
        # Stockist hierarchy filters
        if self.stockist:
            filters["stockist_code"] = self.stockist
        elif self.hq:
            stockists = frappe.get_all(
                "Stockist Master",
                filters={"hq": self.hq, "status": "Active"},
                pluck="stockist_code"
            )
            if stockists:
                filters["stockist_code"] = ["in", stockists]
        elif self.team:
            hqs = frappe.get_all(
                "HQ Master",
                filters={"team": self.team, "status": "Active"},
                pluck="name"
            )
            stockist_list = []
            for hq in hqs:
                stockists = frappe.get_all(
                    "Stockist Master",
                    filters={"hq": hq, "status": "Active"},
                    pluck="stockist_code"
                )
                stockist_list.extend(stockists)
            
            if stockist_list:
                filters["stockist_code"] = ["in", stockist_list]
        elif self.region:
            teams = frappe.get_all(
                "Team Master",
                filters={"region": self.region, "status": "Active"},
                pluck="name"
            )
            stockist_list = []
            for team in teams:
                hqs = frappe.get_all(
                    "HQ Master",
                    filters={"team": team, "status": "Active"},
                    pluck="name"
                )
                for hq in hqs:
                    stockists = frappe.get_all(
                        "Stockist Master",
                        filters={"hq": hq, "status": "Active"},
                        pluck="stockist_code"
                    )
                    stockist_list.extend(stockists)
            
            if stockist_list:
                filters["stockist_code"] = ["in", stockist_list]
        
        return filters
    
    def aggregate_product_data(self, statements):
        """
        Aggregate product-wise data from stockist statements
        Structure: {product_code: {hq: {stockist: {month: {primary, secondary, closing}}}}}
        """
        product_map = {}
        
        for statement in statements:
            stmt_doc = frappe.get_doc("Stockist Statement", statement["name"])
            
            if not frappe.db.exists("Stockist Master", stmt_doc.stockist_code):
                continue
            
            stockist = frappe.get_doc("Stockist Master", stmt_doc.stockist_code)
            hq = stockist.hq
            month_key = str(stmt_doc.statement_month)
            
            # Process each item in statement
            for item in stmt_doc.items:
                product_code = item.product_code
                
                # Apply product filters
                if self.product_code and product_code != self.product_code:
                    continue
                
                # Get product details
                product = frappe.db.get_value(
                    "Product Master",
                    product_code,
                    ["product_name", "pack", "pts", "category"],
                    as_dict=True
                )
                
                if not product:
                    continue
                
                # Apply category filter
                if self.product_category and self.product_category != "All Products":
                    if product.category != self.product_category:
                        continue
                
                # Initialize nested structure
                if product_code not in product_map:
                    product_map[product_code] = {
                        "product_name": product.product_name,
                        "pack": product.pack,
                        "pts": product.pts,
                        "hqs": {}
                    }
                
                if hq not in product_map[product_code]["hqs"]:
                    product_map[product_code]["hqs"][hq] = {
                        "stockists": {}
                    }
                
                stockist_key = f"{stockist.stockist_name}"
                if stockist_key not in product_map[product_code]["hqs"][hq]["stockists"]:
                    product_map[product_code]["hqs"][hq]["stockists"][stockist_key] = {
                        "months": {},
                        "stockist_code": stockist.stockist_code
                    }
                
                if month_key not in product_map[product_code]["hqs"][hq]["stockists"][stockist_key]["months"]:
                    product_map[product_code]["hqs"][hq]["stockists"][stockist_key]["months"][month_key] = {
                        "primary_qty": 0,
                        "secondary_qty": 0,
                        "closing_qty": 0,
                        "closing_value": 0,
                        "previous_closing_qty": 0
                    }
                
                month_data = product_map[product_code]["hqs"][hq]["stockists"][stockist_key]["months"][month_key]
                
                # Aggregate quantities
                if self.include_primary_sales:
                    month_data["primary_qty"] += flt(item.purchase_qty)
                
                if self.include_secondary_sales:
                    month_data["secondary_qty"] += flt(item.sales_qty)
                
                if self.include_closing_stock:
                    month_data["closing_qty"] += flt(item.closing_qty)
                    month_data["closing_value"] += flt(item.closing_qty) * flt(product.pts)
                
                # Get previous month closing if enabled
                if self.show_previous_month_closing:
                    prev_month = add_months(getdate(month_key), -1)
                    prev_closing = self.get_previous_month_closing(
                        stockist.stockist_code,
                        product_code,
                        prev_month
                    )
                    month_data["previous_closing_qty"] = prev_closing
        
        return product_map
    
    def get_previous_month_closing(self, stockist_code, product_code, prev_month):
        """Get closing stock from previous month"""
        try:
            prev_stmt = frappe.get_value(
                "Stockist Statement",
                {
                    "stockist_code": stockist_code,
                    "statement_month": prev_month,
                    "docstatus": 1
                },
                "name"
            )
            
            if not prev_stmt:
                return 0
            
            closing_qty = frappe.db.get_value(
                "Stockist Statement Item",
                {
                    "parent": prev_stmt,
                    "product_code": product_code
                },
                "closing_qty"
            )
            
            return flt(closing_qty or 0)
        
        except Exception:
            return 0
    
    def calculate_totals(self):
        """Calculate summary totals"""
        if not self.report_data:
            return
        
        try:
            product_map = json.loads(self.report_data)
            
            total_primary = 0
            total_secondary = 0
            total_closing_qty = 0
            total_closing_value = 0
            
            for product_code in product_map:
                product_data = product_map[product_code]
                
                for hq in product_data.get("hqs", {}):
                    for stockist in product_data["hqs"][hq].get("stockists", {}):
                        for month in product_data["hqs"][hq]["stockists"][stockist].get("months", {}):
                            month_data = product_data["hqs"][hq]["stockists"][stockist]["months"][month]
                            
                            total_primary += flt(month_data.get("primary_qty", 0))
                            total_secondary += flt(month_data.get("secondary_qty", 0))
                            total_closing_qty += flt(month_data.get("closing_qty", 0))
                            total_closing_value += flt(month_data.get("closing_value", 0))
            
            self.total_primary_qty = total_primary
            self.total_secondary_qty = total_secondary
            self.total_closing_qty = total_closing_qty
            self.total_closing_value = total_closing_value
        
        except Exception as e:
            frappe.log_error(f"Error calculating totals: {str(e)}", "Totals Error")


@frappe.whitelist()
def generate_report(doc_name):
    """API method to generate report"""
    doc = frappe.get_doc("Product Moving Trend Report", doc_name)
    doc.generate_report_data()
    doc.save()
    frappe.db.commit()
    
    return {
        "success": True,
        "message": "Report generated successfully",
        "total_primary": doc.total_primary_qty,
        "total_secondary": doc.total_secondary_qty,
        "total_closing": doc.total_closing_qty,
        "total_closing_value": doc.total_closing_value
    }


@frappe.whitelist()
def export_to_excel(doc_name):
    """Export product moving trend to Excel"""
    doc = frappe.get_doc("Product Moving Trend Report", doc_name)
    
    if not doc.report_data:
        frappe.throw("No report data available. Please generate the report first.")
    
    product_map = json.loads(doc.report_data or "{}")
    
    if not product_map:
        frappe.throw("No data available to export")
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Product Moving Trend"
    
    # Styling
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    subheader_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    subheader_font = Font(color="FFFFFF", bold=True, size=10)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:M1')
    ws['A1'] = "STEDMAN PHARMACEUTICALS PVT LTD"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('A2:M2')
    ws['A2'] = "Product-Wise Moving Trend Report"
    ws['A2'].font = Font(size=12, italic=True)
    ws['A2'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A3:M3')
    ws['A3'] = f"Period: {doc.from_date} to {doc.to_date}"
    ws['A3'].font = Font(size=10)
    ws['A3'].alignment = Alignment(horizontal='center')
    
    # Column headers
    row = 5
    headers = [
        "Product Code", "Product Name", "Pack", "HQ", "Stockist",
        "Month", "Prev Closing", "Primary Qty", "Secondary Qty",
        "Closing Qty", "PTS Rate", "Closing Value", "Movement %"
    ]
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # Data rows
    row = 6
    
    for product_code in sorted(product_map.keys()):
        product_data = product_map[product_code]
        product_name = product_data.get("product_name", "")
        pack = product_data.get("pack", "")
        pts = flt(product_data.get("pts", 0))
        
        for hq in sorted(product_data.get("hqs", {}).keys()):
            for stockist in sorted(product_data["hqs"][hq].get("stockists", {}).keys()):
                stockist_data = product_data["hqs"][hq]["stockists"][stockist]
                
                for month in sorted(stockist_data.get("months", {}).keys()):
                    month_data = stockist_data["months"][month]
                    
                    prev_closing = flt(month_data.get("previous_closing_qty", 0))
                    primary_qty = flt(month_data.get("primary_qty", 0))
                    secondary_qty = flt(month_data.get("secondary_qty", 0))
                    closing_qty = flt(month_data.get("closing_qty", 0))
                    closing_value = flt(month_data.get("closing_value", 0))
                    
                    # Calculate movement %
                    movement_pct = 0
                    if prev_closing > 0:
                        movement_pct = ((prev_closing + primary_qty - closing_qty) / prev_closing) * 100
                    
                    # Write row
                    ws.cell(row=row, column=1).value = product_code
                    ws.cell(row=row, column=2).value = product_name
                    ws.cell(row=row, column=3).value = pack
                    ws.cell(row=row, column=4).value = hq
                    ws.cell(row=row, column=5).value = stockist
                    ws.cell(row=row, column=6).value = month
                    ws.cell(row=row, column=7).value = prev_closing
                    ws.cell(row=row, column=8).value = primary_qty
                    ws.cell(row=row, column=9).value = secondary_qty
                    ws.cell(row=row, column=10).value = closing_qty
                    ws.cell(row=row, column=11).value = pts
                    ws.cell(row=row, column=12).value = closing_value
                    ws.cell(row=row, column=13).value = movement_pct
                    
                    # Formatting
                    for col in range(1, 14):
                        cell = ws.cell(row=row, column=col)
                        cell.border = border
                        if col >= 7 and col <= 10:
                            cell.number_format = '#,##0'
                            cell.alignment = Alignment(horizontal='right')
                        elif col >= 11 and col <= 12:
                            cell.number_format = '#,##0.00'
                            cell.alignment = Alignment(horizontal='right')
                        elif col == 13:
                            cell.number_format = '0.00%'
                            cell.alignment = Alignment(horizontal='right')
                    
                    row += 1
    
    # Summary row
    summary_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    summary_font = Font(bold=True)
    
    ws.cell(row=row, column=1).value = "TOTAL"
    ws.cell(row=row, column=8).value = doc.total_primary_qty
    ws.cell(row=row, column=9).value = doc.total_secondary_qty
    ws.cell(row=row, column=10).value = doc.total_closing_qty
    ws.cell(row=row, column=12).value = doc.total_closing_value
    
    for col in range(1, 14):
        cell = ws.cell(row=row, column=col)
        cell.fill = summary_fill
        cell.font = summary_font
        cell.border = border
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 12
    for col in range(7, 14):
        ws.column_dimensions[get_column_letter(col)].width = 13
    
    # Freeze panes
    ws.freeze_panes = 'A6'
    
    # Save
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Create file
    file_doc = frappe.get_doc({
        "doctype": "File",
        "file_name": f"Product_Moving_Trend_{doc.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        "attached_to_doctype": "Product Moving Trend Report",
        "attached_to_name": doc_name,
        "content": output.getvalue(),
        "is_private": 0
    })
    file_doc.save(ignore_permissions=True)
    
    return {
        "success": True,
        "file_url": file_doc.file_url
    }


@frappe.whitelist()
def export_to_pdf(doc_name):
    """Export to PDF"""
    from frappe.utils.pdf import get_pdf
    
    doc = frappe.get_doc("Product Moving Trend Report", doc_name)
    
    if not doc.report_data:
        frappe.throw("No report data available. Please generate the report first.")
    
    html = render_report_html(doc)
    pdf_bytes = get_pdf(html)
    
    file_doc = frappe.get_doc({
        "doctype": "File",
        "file_name": f"Product_Moving_Trend_{doc.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
        "attached_to_doctype": "Product Moving Trend Report",
        "attached_to_name": doc_name,
        "content": pdf_bytes,
        "is_private": 0
    })
    file_doc.save(ignore_permissions=True)
    
    return {
        "success": True,
        "file_url": file_doc.file_url
    }


def render_report_html(doc):
    """Render HTML for PDF export"""
    product_map = json.loads(doc.report_data or "{}")
    
    html = f"""
    <html>
    <head>
        <style>
            @page {{ size: A4 landscape; margin: 10mm; }}
            body {{ font-family: Arial, sans-serif; margin: 0; font-size: 8px; }}
            .header {{ text-align: center; margin-bottom: 10px; }}
            .title {{ font-size: 14px; font-weight: bold; }}
            .subtitle {{ font-size: 10px; color: #666; }}
            table {{ width: 100%; border-collapse: collapse; }}
            th {{ background-color: #1F4E78; color: white; padding: 4px; text-align: center; border: 1px solid #000; font-size: 7px; }}
            td {{ padding: 3px; border: 1px solid #ccc; font-size: 7px; }}
            .number {{ text-align: right; }}
            .text-left {{ text-align: left; }}
            .total-row {{ background-color: #D9E1F2; font-weight: bold; }}
        </style>
    </head>
    <body>
        <div class="header">
            <div class="title">STEDMAN PHARMACEUTICALS PVT LTD</div>
            <div class="subtitle">Product-Wise Moving Trend Report</div>
            <div class="subtitle">Period: {doc.from_date} to {doc.to_date}</div>
        </div>
        
        <table>
            <thead>
                <tr>
                    <th>Code</th>
                    <th>Product</th>
                    <th>Pack</th>
                    <th>HQ</th>
                    <th>Stockist</th>
                    <th>Month</th>
                    <th>Prev Cls</th>
                    <th>Primary</th>
                    <th>Secondary</th>
                    <th>Closing</th>
                    <th>Value</th>
                </tr>
            </thead>
            <tbody>
    """
    
    for product_code in sorted(product_map.keys()):
        product_data = product_map[product_code]
        
        for hq in sorted(product_data.get("hqs", {}).keys()):
            for stockist in sorted(product_data["hqs"][hq].get("stockists", {}).keys()):
                stockist_data = product_data["hqs"][hq]["stockists"][stockist]
                
                for month in sorted(stockist_data.get("months", {}).keys()):
                    month_data = stockist_data["months"][month]
                    
                    html += f"""
                    <tr>
                        <td class='text-left'>{product_code}</td>
                        <td class='text-left'>{product_data.get('product_name', '')}</td>
                        <td class='text-left'>{product_data.get('pack', '')}</td>
                        <td class='text-left'>{hq}</td>
                        <td class='text-left'>{stockist}</td>
                        <td>{month}</td>
                        <td class='number'>{flt(month_data.get('previous_closing_qty', 0)):.0f}</td>
                        <td class='number'>{flt(month_data.get('primary_qty', 0)):.0f}</td>
                        <td class='number'>{flt(month_data.get('secondary_qty', 0)):.0f}</td>
                        <td class='number'>{flt(month_data.get('closing_qty', 0)):.0f}</td>
                        <td class='number'>{flt(month_data.get('closing_value', 0)):.2f}</td>
                    </tr>
                    """
    
    html += f"""
            <tr class='total-row'>
                <td colspan='7' class='text-left'>TOTAL</td>
                <td class='number'>{flt(doc.total_primary_qty):.0f}</td>
                <td class='number'>{flt(doc.total_secondary_qty):.0f}</td>
                <td class='number'>{flt(doc.total_closing_qty):.0f}</td>
                <td class='number'>{flt(doc.total_closing_value):.2f}</td>
            </tr>
            </tbody>
        </table>
    </body>
    </html>
    """
    
    return html
